"""One-shot: add KO translations (subtitle_ko/title_ko/visual_direction_ko) to existing weekly plan Excel."""
import os, sys, json, openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from env_loader import load_env
load_env()

import anthropic

SRC = Path('.tmp/weekly_content_plan_20260424.xlsx')

client = anthropic.Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY'))

wb_old = openpyxl.load_workbook(SRC)
ws_old = wb_old.active

plans = []
for row in ws_old.iter_rows(min_row=2, max_row=21, values_only=True):
    if not row[0]:
        continue
    plans.append({
        'num': row[0], 'category': row[1], 'format': row[2],
        'subtitle': row[3] or '', 'title': row[4] or '',
        'topic': row[5] or '', 'topic_ko': row[6] or '',
        'visual_direction': row[7] or '',
        'caption': row[8] or '', 'caption_ko': row[9] or '',
        'hashtags': row[10] or '', 'trend_ref': row[11] or '',
    })

items = [{'num': p['num'], 'subtitle': p['subtitle'], 'title': p['title'],
          'visual_direction': p['visual_direction']} for p in plans]

prompt = (
    "다음 일본어 필드들을 자연스러운 한국어로 번역하세요.\n"
    "- subtitle / title: 짧고 임팩트 있게 (썸네일용)\n"
    "- visual_direction: 촬영 지시가 그대로 전달되도록 구체적으로\n\n"
    + json.dumps(items, ensure_ascii=False, indent=2)
    + "\n\n출력은 JSON 배열만. 다른 텍스트 금지:\n"
    + '[{"num": 1, "subtitle_ko": "...", "title_ko": "...", "visual_direction_ko": "..."}, ...]'
)

resp = client.messages.create(
    model='claude-sonnet-4-6', max_tokens=8000,
    messages=[{'role': 'user', 'content': prompt}]
)
text = resp.content[0].text.strip()
if text.startswith('```'):
    parts = text.split('```')
    text = parts[1] if len(parts) > 1 else text
    if text.startswith('json'):
        text = text[4:]
    text = text.strip('`\n ')

translations = json.loads(text)
trans = {t['num']: t for t in translations}

for p in plans:
    t = trans.get(p['num'], {})
    p['subtitle_ko'] = t.get('subtitle_ko', '')
    p['title_ko'] = t.get('title_ko', '')
    p['visual_direction_ko'] = t.get('visual_direction_ko', '')

print(f'Translated {len(translations)} plans')

COLUMNS = [
    ('#', 5, False), ('카테고리', 14, False), ('포맷', 14, False),
    ('소제목', 20, False), ('소제목(한국어)', 20, False),
    ('대제목', 30, False), ('대제목(한국어)', 30, False),
    ('주제', 28, True), ('주제(한국어)', 28, True),
    ('비주얼 방향', 55, True), ('비주얼 방향(한국어)', 55, True),
    ('캡션', 55, True), ('캡션(한국어)', 55, True),
    ('해시태그', 40, True), ('참고 트렌드', 30, True),
]

NAVY = '1F4E79'
WHITE = 'FFFFFF'
OFF_WHITE = 'F7F7F7'

def border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '주간 기획안'
ws.sheet_view.showGridLines = False

for ci, (hdr, w, _) in enumerate(COLUMNS, 1):
    c = ws.cell(row=1, column=ci, value=hdr)
    c.font = Font(name='Calibri', bold=True, color=WHITE, size=11)
    c.fill = PatternFill(fill_type='solid', fgColor=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border()
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = 'A2'

for ri, p in enumerate(plans):
    row_idx = ri + 2
    row_bg = OFF_WHITE if ri % 2 == 0 else WHITE
    vals = [
        p['num'], p['category'], p['format'],
        p['subtitle'], p['subtitle_ko'],
        p['title'], p['title_ko'],
        p['topic'], p['topic_ko'],
        p['visual_direction'], p['visual_direction_ko'],
        p['caption'], p['caption_ko'],
        p['hashtags'], p['trend_ref'],
    ]
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row_idx, column=ci, value=v)
        c.fill = PatternFill(fill_type='solid', fgColor=row_bg)
        c.alignment = Alignment(wrap_text=COLUMNS[ci - 1][2], vertical='top')
        c.border = border()
        c.font = Font(name='Calibri', size=10)
    ws.row_dimensions[row_idx].height = 80

wb.save(SRC)
print(f'Saved: {SRC}')
print(f"Sample #{plans[0]['num']}: {plans[0]['title']} → {plans[0]['title_ko']}")
