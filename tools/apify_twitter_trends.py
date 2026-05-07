"""
Apify Twitter Trends — JP育児ハッシュタグ・トレンドを自動収集

Apifyの twitter-scraper (検索モード) を使い、日本の育児関連
ハッシュタグとトレンドキーワードを構造化データとして収集する。
既存の twitter_hashtag.py (Firecrawl版) を補完・代替する安定版。

出力:
  - JSON: .tmp/apify_twitter_trends_{date}.json
  - Excel: .tmp/apify_twitter_trends_{date}.xlsx (Teams送信用)

Usage:
    python tools/apify_twitter_trends.py                    # 全キーワード検索
    python tools/apify_twitter_trends.py --dry-run           # クロールのみ
    python tools/apify_twitter_trends.py --json              # JSON出力
    python tools/apify_twitter_trends.py --notify            # Teams通知付き
"""

import argparse
import json
import os
import re
import sys
import logging
from collections import Counter
from pathlib import Path
from datetime import datetime, timezone, timedelta

from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).parent.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
TMP_DIR.mkdir(exist_ok=True)
JST = timezone(timedelta(hours=9))

# ═══════════════════════════════════════════════════════════════════════
# CONFIG: 検索キーワード
# ═══════════════════════════════════════════════════════════════════════

SEARCH_QUERIES = {
    "育児・子育て": [
        "育児 ママ",
        "子育て あるある",
        "離乳食 レシピ",
        "1歳 成長",
        "イヤイヤ期",
    ],
    "ベビー用品": [
        "ストローマグ 赤ちゃん",
        "哺乳瓶 おすすめ",
        "ベビーカー 比較",
        "抱っこ紐 口コミ",
        "PPSU ボトル",
    ],
    "グロミミ関連": [
        "グロミミ",
        "grosmimi",
        "ステンレス マグ 赤ちゃん",
    ],
}

TWITTER_ACTOR = "apidojo/tweet-scraper"
TWEETS_PER_QUERY = 20
TIMEOUT_SECS = 120


# ═══════════════════════════════════════════════════════════════════════
# CRAWL
# ═══════════════════════════════════════════════════════════════════════

def crawl_trends(queries_dict: dict) -> list[dict]:
    """Search Twitter/X for each query via Apify."""
    token = os.environ.get("APIFY_API_TOKEN")
    if not token:
        logger.error("APIFY_API_TOKEN not set")
        return []

    try:
        from apify_client import ApifyClient
    except ImportError:
        logger.error("apify-client not installed. Run: pip install apify-client")
        return []

    client = ApifyClient(token)
    results = []

    for category, queries in queries_dict.items():
        for query in queries:
            try:
                run = client.actor(TWITTER_ACTOR).call(
                    run_input={
                        "searchTerms": [query],
                        "maxItems": TWEETS_PER_QUERY,
                        "proxyConfig": {"useApifyProxy": True},
                    },
                    timeout_secs=TIMEOUT_SECS,
                )
                items = list(client.dataset(run["defaultDatasetId"]).iterate_items())

                tweets = []
                for item in items[:TWEETS_PER_QUERY]:
                    tweets.append({
                        "text": (item.get("text", item.get("full_text", "")) or "")[:280],
                        "likes": item.get("likeCount", item.get("likes", 0)),
                        "retweets": item.get("retweetCount", item.get("retweets", 0)),
                        "views": item.get("viewCount", item.get("views", 0)),
                        "hashtags": item.get("hashtags", []),
                        "author": (item.get("author", {}) or {}).get("name",
                                   (item.get("user", {}) or {}).get("name", "")),
                        "created_at": item.get("createdAt", item.get("created_at", "")),
                    })

                results.append({
                    "category": category,
                    "query": query,
                    "tweet_count": len(tweets),
                    "tweets": tweets,
                    "error": None,
                })

                logger.info(f'  [OK] "{query}": {len(tweets)} tweets')

            except Exception as e:
                results.append({
                    "category": category,
                    "query": query,
                    "tweet_count": 0,
                    "tweets": [],
                    "error": str(e)[:200],
                })
                logger.error(f'  [ERR] "{query}": {e}')

    return results


# ═══════════════════════════════════════════════════════════════════════
# ANALYZE: Extract hashtags and trends
# ═══════════════════════════════════════════════════════════════════════

def extract_hashtag_pattern(text: str) -> list[str]:
    """Extract Japanese/English hashtags from tweet text."""
    return re.findall(r'#[\w\u3000-\u9FFF\uFF00-\uFFEF]+', text)


def analyze_trends(results: list[dict]) -> dict:
    """Analyze collected tweets to find trending hashtags and topics."""
    all_hashtags = Counter()
    category_hashtags = {}
    high_engagement_tweets = []

    for r in results:
        if r.get("error"):
            continue

        cat = r["category"]
        if cat not in category_hashtags:
            category_hashtags[cat] = Counter()

        for tweet in r["tweets"]:
            # From structured hashtags field
            for h in tweet.get("hashtags", []):
                tag = h.lower().strip("#")
                all_hashtags[tag] += 1
                category_hashtags[cat][tag] += 1

            # From text pattern matching
            for h in extract_hashtag_pattern(tweet.get("text", "")):
                tag = h.lower().strip("#")
                all_hashtags[tag] += 1
                category_hashtags[cat][tag] += 1

            # Track high engagement tweets
            engagement = (tweet.get("likes", 0) +
                          tweet.get("retweets", 0) * 2)
            if engagement > 50:
                high_engagement_tweets.append({
                    "text": tweet["text"][:150],
                    "likes": tweet.get("likes", 0),
                    "retweets": tweet.get("retweets", 0),
                    "author": tweet.get("author", ""),
                    "query": r["query"],
                })

    # Sort by frequency
    high_engagement_tweets.sort(
        key=lambda t: t["likes"] + t["retweets"] * 2, reverse=True
    )

    return {
        "top_hashtags": all_hashtags.most_common(30),
        "category_hashtags": {
            cat: counter.most_common(10)
            for cat, counter in category_hashtags.items()
        },
        "high_engagement_tweets": high_engagement_tweets[:10],
        "total_tweets_analyzed": sum(
            r["tweet_count"] for r in results if not r.get("error")
        ),
        "grosmimi_mentions": all_hashtags.get("グロミミ", 0) + all_hashtags.get("grosmimi", 0),
    }


# ═══════════════════════════════════════════════════════════════════════
# OUTPUT
# ═══════════════════════════════════════════════════════════════════════

def save_json(results: list[dict], analysis: dict, date_str: str) -> Path:
    """Save results and analysis to JSON."""
    output = {
        "crawled_at": datetime.now(JST).isoformat(),
        "analysis": {
            "top_hashtags": [{"tag": t, "count": c} for t, c in analysis["top_hashtags"]],
            "category_hashtags": {
                cat: [{"tag": t, "count": c} for t, c in tags]
                for cat, tags in analysis["category_hashtags"].items()
            },
            "high_engagement_tweets": analysis["high_engagement_tweets"],
            "total_tweets": analysis["total_tweets_analyzed"],
            "grosmimi_mentions": analysis["grosmimi_mentions"],
        },
        "raw_results": results,
    }

    path = TMP_DIR / f"apify_twitter_trends_{date_str}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    logger.info(f"  Saved: {path}")
    return path


def save_excel(analysis: dict, date_str: str) -> Path | None:
    """Save analysis to Excel."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("  openpyxl not installed — skipping Excel")
        return None

    wb = openpyxl.Workbook()

    # Sheet 1: Top Hashtags
    ws1 = wb.active
    ws1.title = "人気ハッシュタグ"
    ws1.cell(row=1, column=1, value="ランク")
    ws1.cell(row=1, column=2, value="ハッシュタグ")
    ws1.cell(row=1, column=3, value="出現回数")
    for i, (tag, count) in enumerate(analysis["top_hashtags"], 2):
        ws1.cell(row=i, column=1, value=i - 1)
        ws1.cell(row=i, column=2, value=f"#{tag}")
        ws1.cell(row=i, column=3, value=count)

    # Sheet 2: Category Breakdown
    ws2 = wb.create_sheet("カテゴリ別")
    ws2.cell(row=1, column=1, value="カテゴリ")
    ws2.cell(row=1, column=2, value="ハッシュタグ")
    ws2.cell(row=1, column=3, value="出現回数")
    row = 2
    for cat, tags in analysis["category_hashtags"].items():
        for tag, count in tags:
            ws2.cell(row=row, column=1, value=cat)
            ws2.cell(row=row, column=2, value=f"#{tag}")
            ws2.cell(row=row, column=3, value=count)
            row += 1

    # Sheet 3: High Engagement
    ws3 = wb.create_sheet("高エンゲージメント")
    ws3.cell(row=1, column=1, value="ツイート")
    ws3.cell(row=1, column=2, value="いいね")
    ws3.cell(row=1, column=3, value="RT")
    ws3.cell(row=1, column=4, value="投稿者")
    ws3.cell(row=1, column=5, value="検索クエリ")
    for i, t in enumerate(analysis["high_engagement_tweets"], 2):
        ws3.cell(row=i, column=1, value=t["text"])
        ws3.cell(row=i, column=2, value=t["likes"])
        ws3.cell(row=i, column=3, value=t["retweets"])
        ws3.cell(row=i, column=4, value=t["author"])
        ws3.cell(row=i, column=5, value=t["query"])

    path = TMP_DIR / f"apify_twitter_trends_{date_str}.xlsx"
    wb.save(str(path))
    logger.info(f"  Saved: {path}")
    return path


def notify_teams(analysis: dict):
    """Send trend summary to Teams."""
    webhook = os.environ.get("TEAMS_WEBHOOK_URL")
    if not webhook:
        logger.info("  TEAMS_WEBHOOK_URL not set — skipping")
        return

    import requests

    ts = datetime.now(JST).strftime("%Y-%m-%d %H:%M JST")
    lines = [f"## 📊 Apify JP育児トレンドレポート\n**日時:** {ts}\n"]

    lines.append("### 人気ハッシュタグ TOP 10")
    for tag, count in analysis["top_hashtags"][:10]:
        lines.append(f"- #{tag} ({count}回)")

    gm = analysis.get("grosmimi_mentions", 0)
    lines.append(f"\n**グロミミ言及:** {gm}回")

    try:
        requests.post(webhook, json={"text": "\n".join(lines)}, timeout=10)
        logger.info("  Teams notification sent")
    except Exception as e:
        logger.warning(f"  Teams notification failed: {e}")


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Apify Twitter Trends — JP育児トレンド収集"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="クロールのみ（保存なし）")
    parser.add_argument("--json", action="store_true",
                        help="JSON出力")
    parser.add_argument("--notify", action="store_true",
                        help="Teams通知")
    args = parser.parse_args()

    now = datetime.now(JST)
    date_str = now.strftime("%Y%m%d_%H%M")

    logger.info("=== Apify Twitter Trends ===")
    logger.info(f"  Timestamp: {now.strftime('%Y-%m-%d %H:%M JST')}")
    logger.info(f"  Queries: {sum(len(v) for v in SEARCH_QUERIES.values())}")
    print()

    # Crawl
    logger.info("[1/3] Searching Twitter/X via Apify...")
    results = crawl_trends(SEARCH_QUERIES)
    if not results:
        logger.error("No results. Check APIFY_API_TOKEN.")
        sys.exit(1)

    success = sum(1 for r in results if not r.get("error"))
    logger.info(f"  Completed: {success}/{len(results)} queries")
    print()

    # Analyze
    logger.info("[2/3] Analyzing hashtags & trends...")
    analysis = analyze_trends(results)
    logger.info(f"  Total tweets: {analysis['total_tweets_analyzed']}")
    logger.info(f"  Unique hashtags: {len(analysis['top_hashtags'])}")
    logger.info(f"  Grosmimi mentions: {analysis['grosmimi_mentions']}")
    print()

    # Output
    if not args.dry_run:
        logger.info("[3/3] Saving results...")
        save_json(results, analysis, date_str)
        save_excel(analysis, date_str)
        if args.notify:
            notify_teams(analysis)
    else:
        logger.info("[3/3] DRY RUN — skipping save")

    if args.json:
        output = {
            "top_hashtags": [{"tag": t, "count": c} for t, c in analysis["top_hashtags"][:15]],
            "grosmimi_mentions": analysis["grosmimi_mentions"],
            "total_tweets": analysis["total_tweets_analyzed"],
        }
        print(json.dumps(output, ensure_ascii=False, indent=2))
    else:
        print("\n=== Top 15 Hashtags ===")
        for i, (tag, count) in enumerate(analysis["top_hashtags"][:15], 1):
            print(f"  {i:2d}. #{tag} ({count})")
        print(f"\n  グロミミ言及: {analysis['grosmimi_mentions']}回")


if __name__ == "__main__":
    main()
