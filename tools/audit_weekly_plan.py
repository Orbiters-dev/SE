"""
WAT Tool: Audit a weekly tweet plan Excel against memory rules.

Rules checked (from memory/feedback_twitter_weekly_flow.md):
- Per-day slot count: must be [10, 13, 17, 19, 21]
- Total tweets: 35 (5 slots * 7 days)
- Product mentions in body: max 2 of 35
  (keywords: グロミミ|grosmimi|PPSU|マグ|ストロー|+CUT|素材|製品開発|試作|漏れない|マグメーカー)
- K-parenting in body: max 2 of 35, slot 13 only
  (keywords: 韓国|K育児|K幼児食|韓国式|반찬|チゲ|トッポッキ|チヂミ|キムチ|미운)
- Body must be 60-80 chars
- Polite form (です/ます) in body: should be rare

Usage:
    python tools/audit_weekly_plan.py path/to/tweet_plan_weekly_*.xlsx
    python tools/audit_weekly_plan.py --latest    # download latest from SharePoint
"""

import sys
import re
import argparse
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import load_workbook

EXPECTED_SLOTS = [10, 13, 17, 19, 21]
PRODUCT_QUOTA = 2
KOREAN_QUOTA = 2
KOREAN_ALLOWED_SLOT = 13
BODY_MIN = 60
BODY_MAX = 80

PRODUCT_RE = re.compile(
    r"グロミミ|grosmimi|PPSU|ppsu|マグ|ストロー|製品開発|試作|素材|\+CUT|漏れない|マグメーカー"
)
KOREAN_RE = re.compile(
    r"韓国|K育児|K幼児食|K-育児|韓国式|반찬|チゲ|トッポッキ|チヂミ|キムチ|미운"
)
POLITE_RE = re.compile(r"です(?![ねよ])|ます(?![ねよ])|でした|ました")
HASHTAG_RE = re.compile(r"#\S+")


def split_body_hashtags(text: str) -> tuple[str, list[str]]:
    """Split tweet into body (no hashtags) and hashtag list."""
    tags = HASHTAG_RE.findall(text)
    body = HASHTAG_RE.sub("", text).strip()
    return body, tags


def audit_excel(path: str) -> dict:
    """Audit a weekly tweet plan Excel. Returns result dict + prints report."""
    wb = load_workbook(path, data_only=True)

    issues = {
        "slot_missing": [],
        "product_violations": [],
        "korean_non13": [],
        "korean_13_excess": [],
        "body_length": [],
        "polite_form": [],
    }
    total = 0
    product_hits = []
    k_non13 = []
    k_13 = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        slots_in_sheet = []
        for r in range(4, ws.max_row + 1):
            time_v = ws.cell(row=r, column=1).value
            tweet = ws.cell(row=r, column=3).value
            if not time_v or not tweet:
                continue
            try:
                slot = int(str(time_v).split(":")[0])
            except Exception:
                continue
            slots_in_sheet.append(slot)
            text = str(tweet)
            body, _tags = split_body_hashtags(text)
            total += 1
            entry = (sheet_name, slot, body[:60])

            if PRODUCT_RE.search(body):
                product_hits.append(entry)
            if KOREAN_RE.search(body):
                if slot == KOREAN_ALLOWED_SLOT:
                    k_13.append(entry)
                else:
                    k_non13.append(entry)
            body_len = len(body)
            if body_len < BODY_MIN or body_len > BODY_MAX:
                issues["body_length"].append((*entry, body_len))
            if POLITE_RE.search(body):
                issues["polite_form"].append(entry)

        missing = [s for s in EXPECTED_SLOTS if s not in slots_in_sheet]
        if missing:
            issues["slot_missing"].append((sheet_name, missing))

    if len(product_hits) > PRODUCT_QUOTA:
        issues["product_violations"] = product_hits[PRODUCT_QUOTA:]
    issues["korean_non13"] = k_non13
    if len(k_13) > KOREAN_QUOTA:
        issues["korean_13_excess"] = k_13[KOREAN_QUOTA:]

    print(f"=== AUDIT REPORT: {Path(path).name} ===")
    print(f"Total tweets: {total} (expected 35)")
    print(f"Sheets: {len(wb.sheetnames)}")
    print()

    fail = False
    if issues["slot_missing"]:
        fail = True
        print("[FAIL] Missing slots per day:")
        for sn, ms in issues["slot_missing"]:
            print(f"  {sn}: missing {ms}")

    print(f"Product mentions in body: {len(product_hits)} (quota={PRODUCT_QUOTA})")
    if issues["product_violations"]:
        fail = True
        print(f"  [FAIL] Excess product mentions ({len(issues['product_violations'])}):")
        for sn, sl, t in issues["product_violations"]:
            print(f"    {sn} slot{sl}: {t}")

    print(f"K-parenting mentions: 13slot={len(k_13)} non13={len(k_non13)}")
    if issues["korean_non13"]:
        fail = True
        print(f"  [FAIL] K-parenting in non-13 slots ({len(k_non13)}):")
        for sn, sl, t in k_non13:
            print(f"    {sn} slot{sl}: {t}")
    if issues["korean_13_excess"]:
        fail = True
        print(f"  [FAIL] K-parenting on slot 13 over quota ({len(issues['korean_13_excess'])}):")
        for sn, sl, t in issues["korean_13_excess"]:
            print(f"    {sn} slot{sl}: {t}")

    if issues["body_length"]:
        print(f"[WARN] Body length out of [{BODY_MIN},{BODY_MAX}]: {len(issues['body_length'])}")
        for sn, sl, t, ln in issues["body_length"][:10]:
            print(f"    {sn} slot{sl} (body={ln}): {t}")

    if issues["polite_form"]:
        print(f"[WARN] Polite form in body: {len(issues['polite_form'])}")
        for sn, sl, t in issues["polite_form"][:10]:
            print(f"    {sn} slot{sl}: {t}")

    print()
    print("RESULT:", "FAIL" if fail else "PASS")
    return {"pass": not fail, "issues": issues, "total": total}


def main():
    p = argparse.ArgumentParser()
    p.add_argument("path", nargs="?", help="Path to weekly Excel")
    p.add_argument("--latest", action="store_true", help="Download latest from SharePoint Teams channel")
    args = p.parse_args()

    if args.latest:
        from teams_upload import _get_channel_drive_folder, _graph_headers, GRAPH_BASE
        import requests
        drive_id, folder_id = _get_channel_drive_folder()
        url = f"{GRAPH_BASE}/drives/{drive_id}/items/{folder_id}/children?$top=200"
        r = requests.get(url, headers=_graph_headers(), timeout=30)
        items = [x for x in r.json().get("value", []) if x["name"].startswith("tweet_plan_weekly_")]
        items.sort(key=lambda x: x["lastModifiedDateTime"], reverse=True)
        if not items:
            print("No weekly Excel found")
            sys.exit(1)
        latest = items[0]
        local = PROJECT_ROOT / ".tmp" / latest["name"]
        dl = requests.get(
            f"{GRAPH_BASE}/drives/{drive_id}/items/{latest['id']}/content",
            headers=_graph_headers(), timeout=60, allow_redirects=True,
        )
        local.write_bytes(dl.content)
        print(f"Downloaded: {local}")
        path = str(local)
    else:
        if not args.path:
            p.print_help()
            sys.exit(1)
        path = args.path

    result = audit_excel(path)
    sys.exit(0 if result["pass"] else 1)


if __name__ == "__main__":
    main()
