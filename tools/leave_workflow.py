"""
휴가 승인 워크플로우 (2-Phase)

Phase 1: 받은편지함에서 휴가 신청 메일 감지 → 승인 회신 드래프트 자동 생성
Phase 2: 보낸편지함에서 승인 메일 발송 감지 → NAS 엑셀 자동 업데이트

Usage:
    python tools/leave_workflow.py --run          # Phase 1 + 2 모두 실행
    python tools/leave_workflow.py --draft-only   # Phase 1만 (드래프트 생성)
    python tools/leave_workflow.py --sync-only    # Phase 2만 (엑셀 업데이트)
    python tools/leave_workflow.py --dry-run      # 테스트 (실제 변경 없음)
    python tools/leave_workflow.py --status       # 현재 상태 확인
"""

import argparse
import json
import os
import re
import sys
from datetime import date as date_cls, datetime, timedelta
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from env_loader import load_env
load_env()

# ── Config ────────────────────────────────────────────────────────────────────
OUTLOOK_EMAIL     = os.getenv("OUTLOOK_EMAIL", "dk.shin@orbiters.co.kr")
OUTLOOK_CLIENT_ID = os.getenv("OUTLOOK_CLIENT_ID", "")
OUTLOOK_TENANT_ID = os.getenv("OUTLOOK_TENANT_ID", "common")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")

SENDER_NAME       = "신동균"
SENDER_TITLE      = "경영지원"
CC_RECIPIENTS     = ["wj.choi@orbiters.co.kr", "mj.lee@orbiters.co.kr"]
INTERNAL_DOMAIN   = "orbiters.co.kr"

BASE_DIR          = Path(__file__).parent.parent
PENDING_PATH      = BASE_DIR / ".tmp" / "leave_pending_drafts.json"
PROCESSED_INBOX   = BASE_DIR / ".tmp" / "leave_inbox_processed.json"
PROCESSED_SENT    = BASE_DIR / ".tmp" / "leave_sent_processed.json"
GRAPH_TOKEN_PATH  = BASE_DIR / "credentials" / "leave_tracker_token.json"
EXCEL_PATH        = Path(r"//Orbiters/경영지원/연차관리/연차관리_2026.xlsx")
EMPLOYEE_DATA     = Path(__file__).parent / "employee_data.json"

GRAPH_SCOPES      = ["Mail.Read", "Mail.ReadWrite", "Mail.Send"]
LEAVE_KEYWORDS    = ["연차", "휴가", "반차", "병가", "공가", "예비군"]
DRAFT_SUBJECT_TAG = "[휴가 승인]"

# Confirm 시트 컬럼: 신청일, 이메일 제목, 휴가 시작, 휴가 종료, 사용일수, 유형, 비고, 잔여, 상태
CONFIRM_HEADERS   = ["신청일", "이메일 제목", "휴가 시작", "휴가 종료", "사용일수", "유형", "비고", "잔여", "상태"]
PERSON_HEADERS    = ["신청일", "이메일 제목", "휴가 시작", "휴가 종료", "사용일수", "유형", "비고", "잔여", "상태"]

# Styles
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
ALT_FILL    = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
NO_FILL     = PatternFill(fill_type=None)
HEADER_FONT = Font(name="Calibri", size=10, bold=True)
DATA_FONT   = Font(name="Calibri", size=10)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"),  bottom=Side(style="thin"))
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)

WEEKDAY_KR = ["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"]


# ── Graph API Auth (MSAL) ────────────────────────────────────────────────────
def get_graph_token() -> str:
    try:
        import msal
    except ImportError:
        print("ERROR: msal 미설치 → pip install msal")
        sys.exit(1)

    GRAPH_TOKEN_PATH.parent.mkdir(parents=True, exist_ok=True)
    cache = msal.SerializableTokenCache()
    if GRAPH_TOKEN_PATH.exists():
        cache.deserialize(GRAPH_TOKEN_PATH.read_text(encoding="utf-8"))

    authority = f"https://login.microsoftonline.com/{OUTLOOK_TENANT_ID}"
    app = msal.PublicClientApplication(OUTLOOK_CLIENT_ID, authority=authority, token_cache=cache)

    accounts = app.get_accounts()
    result = None
    if accounts:
        result = app.acquire_token_silent(GRAPH_SCOPES, account=accounts[0])

    if not result:
        flow = app.initiate_device_flow(scopes=GRAPH_SCOPES)
        print("\n[Microsoft 계정 인증 필요 — 최초 1회만]")
        print(f"  1. 브라우저: {flow['verification_uri']}")
        print(f"  2. 코드 입력: {flow['user_code']}")
        print("  3. 회사 계정으로 로그인\n")
        result = app.acquire_token_by_device_flow(flow)

    if "access_token" in result:
        GRAPH_TOKEN_PATH.write_text(cache.serialize(), encoding="utf-8")
        return result["access_token"]

    print(f"ERROR: 인증 실패 — {result.get('error_description', result)}")
    sys.exit(1)


# ── Persistence Helpers ──────────────────────────────────────────────────────
def _load_json(path: Path) -> dict | list:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {}

def _save_json(path: Path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


# ── Korean Holidays ──────────────────────────────────────────────────────────
def get_kr_holidays(year: int) -> set:
    try:
        import holidays as hol_lib
        return set(hol_lib.KR(years=year).keys())
    except ImportError:
        return {
            date_cls(year, 1, 1), date_cls(year, 3, 1), date_cls(year, 5, 5),
            date_cls(year, 6, 6), date_cls(year, 8, 15), date_cls(year, 10, 3),
            date_cls(year, 10, 9), date_cls(year, 12, 25),
        }

def is_working_day(d: date_cls, kr_holidays: set) -> bool:
    return d.weekday() < 5 and d not in kr_holidays


# ── Employee Data ────────────────────────────────────────────────────────────
def load_employees() -> dict:
    if EMPLOYEE_DATA.exists():
        return json.loads(EMPLOYEE_DATA.read_text(encoding="utf-8"))
    return {"hire_dates": {}, "emails": {}}

def find_employee_by_email(email: str) -> str | None:
    emp = load_employees()
    email_lower = email.lower()
    for name, addr in emp.get("emails", {}).items():
        if addr.lower() == email_lower:
            return name
    return None

def get_annual_leave_days(name: str) -> int:
    """입사일 기준 연차 일수 계산 (근로기준법)."""
    emp = load_employees()
    hire_str = emp.get("hire_dates", {}).get(name)
    if not hire_str:
        return 15
    hire = datetime.strptime(hire_str, "%Y-%m-%d").date()
    today = date_cls.today()
    years_worked = (today - hire).days / 365.25
    if years_worked < 1:
        # 1년 미만: 1개월 만근 시 1일 (최대 11일) — 간이 계산
        months = max(0, (today.year - hire.year) * 12 + today.month - hire.month)
        return min(months, 11)
    elif years_worked < 3:
        return 15
    else:
        extra = int((years_worked - 1) // 2)
        return min(15 + extra, 25)


# ── AI Parsing ───────────────────────────────────────────────────────────────
def parse_leave_email(email: dict) -> dict | None:
    """Claude AI로 자유형 휴가 신청 메일 파싱."""
    import anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    body_text = re.sub(r"<[^>]+>", " ", email.get("body", ""))
    body_text = re.sub(r"\s+", " ", body_text).strip()[:3000]
    current_year = datetime.now().year

    prompt = f"""다음 이메일이 연차/휴가/반차 신청인지 분석해주세요.

발신자: {email['sender_name']} ({email['sender_email']})
제목: {email['subject']}
수신일: {email['received_at']}
본문:
{body_text}

연차/휴가 신청이면 아래 JSON만 응답 (다른 텍스트 없이):
{{
  "is_leave_request": true,
  "name": "신청자 이름",
  "leave_type": "연차 또는 반차",
  "start_date": "YYYY-MM-DD",
  "end_date": "YYYY-MM-DD",
  "days": 1.0,
  "note": "오전/오후 반차 등 특이사항 (없으면 빈 문자열)",
  "reason": "사유 (없으면 빈 문자열)"
}}

규칙:
- 반차는 days=0.5, note에 "오전 반차" 또는 "오후 반차" 명시
- 연차 단일일이면 start_date == end_date
- 연도 불명확 시 {current_year}년 기준
- 주말/공휴일은 제외하고 영업일 기준 days 계산
- JSON만 응답, 마크다운 코드블록 없이

아니면:
{{"is_leave_request": false}}"""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=500,
            messages=[{"role": "user", "content": prompt}],
        )
        text = response.content[0].text.strip()
        text = re.sub(r"^```[a-z]*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
        parsed = json.loads(text)
    except Exception as e:
        print(f"  [파싱 오류] {e}")
        return None

    if not parsed.get("is_leave_request"):
        return None

    parsed["sender_email"] = email["sender_email"]
    parsed["sender_name"]  = email["sender_name"]
    parsed["email_subject"] = email["subject"]
    parsed["email_id"]     = email["id"]
    parsed["received_at"]  = email["received_at"]
    return parsed


# ── Excel Helpers ────────────────────────────────────────────────────────────
def get_leave_balance(name: str) -> dict:
    """NAS 엑셀에서 해당 직원의 연차 잔여 현황 조회."""
    annual = get_annual_leave_days(name)
    used = 0.0

    if not EXCEL_PATH.exists():
        return {"annual": annual, "used": 0.0, "remaining": float(annual)}

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Summary 시트에서 조회
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip() == name:
                annual = row[3] if row[3] else annual  # D열: 발생연차
                used = row[16] if row[16] else 0.0     # Q열: 총사용
                break

    # 개인 시트에서도 확인 (더 정확)
    if name in wb.sheetnames:
        ws = wb[name]
        total = 0.0
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[4] is not None:  # E열: 사용일수
                try:
                    total += float(row[4])
                except (ValueError, TypeError):
                    pass
        if total > 0:
            used = total

    wb.close()
    return {"annual": annual, "used": used, "remaining": annual - used}


def update_excel(parsed: dict, dry_run: bool = False) -> bool:
    """NAS 엑셀에 휴가 기록 추가 (Confirm + 개인 시트 + Summary)."""
    if not EXCEL_PATH.exists():
        print(f"  ERROR: 엑셀 파일 없음 — {EXCEL_PATH}")
        return False

    name       = parsed["name"]
    balance    = get_leave_balance(name)
    new_used   = balance["used"] + parsed["days"]
    remaining  = balance["annual"] - new_used

    if dry_run:
        print(f"  [DRY RUN] 엑셀 업데이트 스킵: {name} {parsed['start_date']}~{parsed['end_date']} ({parsed['days']}일)")
        print(f"  잔여: {balance['remaining']} → {remaining}")
        return True

    wb = openpyxl.load_workbook(EXCEL_PATH)

    # ── Confirm 시트에 추가 ──
    if "Confirm" in wb.sheetnames:
        ws_confirm = wb["Confirm"]
        next_row = ws_confirm.max_row + 1
        subject = f"{DRAFT_SUBJECT_TAG} {name} 님 휴가 승인 안내"
        note = parsed.get("note", "")
        values = [
            parsed["received_at"], subject,
            parsed["start_date"], parsed["end_date"],
            parsed["days"], parsed["leave_type"],
            note, remaining, "승인"
        ]
        for col, val in enumerate(values, 1):
            cell = ws_confirm.cell(row=next_row, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CTR if col != 2 else LFT
        print(f"  Confirm 시트: Row {next_row} 추가")

    # ── 개인 시트에 추가 ──
    sheet_name = name
    if sheet_name in wb.sheetnames:
        ws_person = wb[sheet_name]
        next_row = ws_person.max_row + 1
        subject = f"{DRAFT_SUBJECT_TAG} {name} 님 휴가 승인 안내"
        values = [
            parsed["received_at"], subject,
            parsed["start_date"], parsed["end_date"],
            parsed["days"], parsed["leave_type"],
            parsed.get("note", ""), remaining, "승인"
        ]
        for col, val in enumerate(values, 1):
            cell = ws_person.cell(row=next_row, column=col, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CTR if col != 2 else LFT
        print(f"  {sheet_name} 시트: Row {next_row} 추가")
    else:
        print(f"  WARN: '{sheet_name}' 시트 없음 — 개인 시트 스킵")

    # ── Summary 시트 월별 업데이트 ──
    if "Summary" in wb.sheetnames:
        ws_summary = wb["Summary"]
        try:
            month = int(parsed["start_date"][5:7])
        except (ValueError, IndexError):
            month = datetime.now().month

        for row in ws_summary.iter_rows(min_row=2):
            if row[0].value and str(row[0].value).strip() == name:
                # 월별 컬럼: E(1월)~P(12월), Q(총사용), R(잔여)
                month_col = 4 + month  # E=5(1월), F=6(2월), ...
                current_val = row[month_col - 1].value or 0
                row[month_col - 1].value = current_val + parsed["days"]

                # Q열(총사용) R열(잔여) 업데이트
                row[16].value = new_used       # Q열 (col 17, idx 16)
                row[17].value = remaining      # R열 (col 18, idx 17)
                print(f"  Summary: {name} {month}월 +{parsed['days']}일, 잔여 {remaining}")
                break

    wb.save(EXCEL_PATH)
    wb.close()
    print(f"  엑셀 저장 완료: {EXCEL_PATH}")
    return True


# ── Draft Email Builder ──────────────────────────────────────────────────────
def build_approval_html(parsed: dict, balance: dict) -> str:
    """동균이 승인 회신 템플릿 (HTML)."""
    name       = parsed["name"]
    start_date = parsed["start_date"]
    days       = parsed["days"]
    leave_type = parsed["leave_type"]

    # 요일 계산
    try:
        dt = datetime.strptime(start_date, "%Y-%m-%d")
        weekday = WEEKDAY_KR[dt.weekday()]
        date_display = f"{dt.year}년 {dt.month}월 {dt.day}일 ({weekday})"
    except ValueError:
        date_display = start_date

    # 일수 표시
    days_display = f"{days}일" if days == int(days) else f"{days}일"

    # 잔여 계산 (이번 건 반영)
    new_used = balance["used"] + days
    new_remaining = balance["annual"] - new_used

    # 비고 (반차면 오전/오후 표시)
    note = parsed.get("note", "")

    table_style = (
        'style="border-collapse:collapse;font-family:Malgun Gothic,sans-serif;font-size:13px;'
        'margin:12px 0;"'
    )
    cell_style = 'style="border:1px solid #ddd;padding:8px 16px;"'
    header_style = 'style="border:1px solid #ddd;padding:8px 16px;background:#f5f5f5;color:#555;font-weight:normal;"'
    section_style = 'style="border:1px solid #ddd;padding:8px 16px;background:#e8f0fe;font-weight:bold;text-align:center;"'

    html = f"""<div style="font-family:Malgun Gothic,sans-serif;font-size:14px;color:#333;line-height:1.6;">
<p>안녕하세요, {name} 님.<br>
{SENDER_TITLE} 담당자 {SENDER_NAME}입니다.</p>

<p>신청하신 휴가 일정이 아래와 같이 승인되었음을 안내드립니다.</p>

<table {table_style}>
<tr><td {header_style}>시작일</td><td {cell_style}>{date_display}</td></tr>
<tr><td {header_style}>사용 일수</td><td {cell_style}>{days_display}</td></tr>
<tr><td {header_style}>휴가 구분</td><td {cell_style}>{leave_type}{(' — ' + note) if note else ''}</td></tr>
</table>

<p><strong>[ 연차 잔여 현황 ]</strong></p>
<table {table_style}>
<tr>
<td {section_style}>기존 연차</td>
<td {section_style}>사용 연차</td>
<td {section_style}>잔여 연차</td>
</tr>
<tr>
<td {cell_style} style="text-align:center;border:1px solid #ddd;padding:8px 16px;">{int(balance['annual'])}</td>
<td {cell_style} style="text-align:center;border:1px solid #ddd;padding:8px 16px;">{new_used if new_used == int(new_used) else new_used}</td>
<td {cell_style} style="text-align:center;border:1px solid #ddd;padding:8px 16px;font-weight:bold;color:#1565C0;">{new_remaining if new_remaining == int(new_remaining) else new_remaining}</td>
</tr>
</table>

<p>휴가 기간 동안 재충전의 시간 되시길 바랍니다.</p>

<p>감사합니다.<br>
{SENDER_TITLE} {SENDER_NAME} 드림</p>
</div>"""
    return html


def create_approval_draft(parsed: dict, reply_to_id: str | None = None, dry_run: bool = False) -> str | None:
    """Outlook 임시저장에 승인 회신 드래프트 생성. 드래프트 ID 반환."""
    name    = parsed["name"]
    balance = get_leave_balance(name)
    html    = build_approval_html(parsed, balance)
    subject = f"{DRAFT_SUBJECT_TAG} {name} 님 휴가 승인 안내"

    if dry_run:
        preview_path = BASE_DIR / ".tmp" / "leave_draft_preview.html"
        preview_path.parent.mkdir(parents=True, exist_ok=True)
        preview_path.write_text(html, encoding="utf-8")
        print(f"  [DRY RUN] 드래프트 프리뷰: {preview_path}")
        return "dry-run-id"

    token = get_graph_token()

    # 수신자: 신청자 + CC
    to_recipient = parsed["sender_email"]
    payload = {
        "subject": subject,
        "body": {"contentType": "HTML", "content": html},
        "toRecipients": [{"emailAddress": {"address": to_recipient}}],
        "ccRecipients": [{"emailAddress": {"address": cc}} for cc in CC_RECIPIENTS],
    }

    # 원본 메일에 대한 답장으로 생성 (reply)
    if reply_to_id:
        try:
            # createReply로 답장 드래프트 생성 후 본문 교체
            resp = requests.post(
                f"https://graph.microsoft.com/v1.0/me/messages/{reply_to_id}/createReply",
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                json={}, timeout=15,
            )
            if resp.status_code == 201:
                draft = resp.json()
                draft_id = draft["id"]
                # 드래프트 본문/제목/CC 업데이트
                update_payload = {
                    "subject": subject,
                    "body": {"contentType": "HTML", "content": html},
                    "toRecipients": [{"emailAddress": {"address": to_recipient}}],
                    "ccRecipients": [{"emailAddress": {"address": cc}} for cc in CC_RECIPIENTS],
                }
                resp2 = requests.patch(
                    f"https://graph.microsoft.com/v1.0/me/messages/{draft_id}",
                    headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                    json=update_payload, timeout=15,
                )
                if resp2.status_code == 200:
                    print(f"  ✓ 답장 드래프트 생성: {subject}")
                    return draft_id
        except Exception as e:
            print(f"  [답장 생성 실패, 새 메일로 대체] {e}")

    # Fallback: 새 메일 드래프트
    try:
        resp = requests.post(
            "https://graph.microsoft.com/v1.0/me/messages",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=payload, timeout=15,
        )
        if resp.status_code == 201:
            draft_id = resp.json().get("id", "")
            print(f"  ✓ 드래프트 생성: {subject}")
            return draft_id
        print(f"  ERROR: 드래프트 생성 실패 HTTP {resp.status_code}: {resp.text[:200]}")
    except Exception as e:
        print(f"  ERROR: 드래프트 생성 실패 — {e}")
    return None


# ── Phase 1: Inbox → Draft ───────────────────────────────────────────────────
def fetch_inbox_emails(hours: int = 6) -> list:
    """최근 N시간 이내 받은편지함에서 @orbiters.co.kr 발신 메일 조회."""
    token = get_graph_token()
    since = (datetime.utcnow() - timedelta(hours=hours)).strftime("%Y-%m-%dT%H:%M:%SZ")
    url = (
        "https://graph.microsoft.com/v1.0/me/mailFolders/inbox/messages"
        f"?$filter=receivedDateTime ge {since}"
        f"&$select=id,subject,body,sender,receivedDateTime,conversationId"
        "&$orderby=receivedDateTime desc&$top=50"
    )
    emails = []
    try:
        resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
        if resp.status_code != 200:
            print(f"ERROR: 받은편지함 조회 실패 HTTP {resp.status_code}")
            return []
        for msg in resp.json().get("value", []):
            sender_email = msg.get("sender", {}).get("emailAddress", {}).get("address", "")
            if not sender_email.lower().endswith(f"@{INTERNAL_DOMAIN}"):
                continue
            if sender_email.lower() == OUTLOOK_EMAIL.lower():
                continue  # 내가 보낸 건 제외
            subject = msg.get("subject", "") or ""
            body    = msg.get("body", {}).get("content", "") or ""
            if not any(kw in (subject + body) for kw in LEAVE_KEYWORDS):
                continue
            emails.append({
                "id":           msg["id"],
                "subject":      subject,
                "body":         body,
                "sender_name":  msg.get("sender", {}).get("emailAddress", {}).get("name", ""),
                "sender_email": sender_email,
                "received_at":  (msg.get("receivedDateTime", "") or "")[:10],
                "conversation_id": msg.get("conversationId", ""),
            })
    except Exception as e:
        print(f"ERROR: 받은편지함 조회 실패 — {e}")
    return emails


def phase1_create_drafts(dry_run: bool = False, hours: int = 6):
    """Phase 1: 받은편지함 → 휴가 메일 감지 → 승인 드래프트 생성."""
    print(f"\n{'='*50}")
    print(f"Phase 1: 받은편지함 감시 (최근 {hours}시간)")
    print(f"{'='*50}")

    emails = fetch_inbox_emails(hours=hours)
    print(f"  휴가 키워드 메일: {len(emails)}건")

    processed = set(_load_json(PROCESSED_INBOX) if isinstance(_load_json(PROCESSED_INBOX), list) else [])
    pending = _load_json(PENDING_PATH) if isinstance(_load_json(PENDING_PATH), dict) else {}

    new_emails = [e for e in emails if e["id"] not in processed]
    print(f"  미처리 메일: {len(new_emails)}건")

    created = 0
    for email in new_emails:
        print(f"\n  [{email['sender_name']}] {email['subject'][:50]}")

        parsed = parse_leave_email(email)
        processed.add(email["id"])

        if not parsed:
            print(f"    → 휴가 신청 아님, 스킵")
            continue

        print(f"    → {parsed['name']} / {parsed['leave_type']} / {parsed['start_date']}~{parsed['end_date']} ({parsed['days']}일)")

        draft_id = create_approval_draft(parsed, reply_to_id=email["id"], dry_run=dry_run)
        if draft_id:
            # pending에 저장 (Phase 2에서 발송 감지용)
            pending[draft_id] = {
                **parsed,
                "draft_created_at": datetime.now().isoformat(),
                "conversation_id": email.get("conversation_id", ""),
            }
            created += 1

    _save_json(PROCESSED_INBOX, list(processed))
    _save_json(PENDING_PATH, pending)
    print(f"\n  Phase 1 완료: 드래프트 {created}건 생성")
    return created


# ── Phase 2: Sent Items → Excel ──────────────────────────────────────────────
def fetch_sent_emails(hours: int = 24) -> list:
    """보낸편지함에서 승인 메일 감지."""
    token = get_graph_token()
    since = (datetime.utcnow() - timedelta(hours=hours)).strftime("%Y-%m-%dT%H:%M:%SZ")
    url = (
        "https://graph.microsoft.com/v1.0/me/mailFolders/sentitems/messages"
        f"?$filter=receivedDateTime ge {since}"
        f"&$select=id,subject,sentDateTime,conversationId"
        "&$orderby=sentDateTime desc&$top=50"
    )
    emails = []
    try:
        resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
        if resp.status_code != 200:
            print(f"ERROR: 보낸편지함 조회 실패 HTTP {resp.status_code}")
            return []
        for msg in resp.json().get("value", []):
            subject = msg.get("subject", "") or ""
            if DRAFT_SUBJECT_TAG not in subject:
                continue
            emails.append({
                "id":              msg["id"],
                "subject":         subject,
                "sent_at":         (msg.get("sentDateTime", "") or "")[:10],
                "conversation_id": msg.get("conversationId", ""),
            })
    except Exception as e:
        print(f"ERROR: 보낸편지함 조회 실패 — {e}")
    return emails


def phase2_sync_excel(dry_run: bool = False, hours: int = 24):
    """Phase 2: 보낸편지함에서 승인 메일 발송 감지 → NAS 엑셀 업데이트."""
    print(f"\n{'='*50}")
    print(f"Phase 2: 보낸편지함 감시 (최근 {hours}시간)")
    print(f"{'='*50}")

    sent_emails = fetch_sent_emails(hours=hours)
    print(f"  승인 메일 감지: {len(sent_emails)}건")

    processed_sent = set(_load_json(PROCESSED_SENT) if isinstance(_load_json(PROCESSED_SENT), list) else [])
    pending = _load_json(PENDING_PATH) if isinstance(_load_json(PENDING_PATH), dict) else {}

    new_sent = [e for e in sent_emails if e["id"] not in processed_sent]
    print(f"  미처리 발송: {len(new_sent)}건")

    synced = 0
    for sent in new_sent:
        print(f"\n  발송 감지: {sent['subject'][:50]}")
        processed_sent.add(sent["id"])

        # pending 목록에서 매칭 (conversation_id 또는 제목)
        matched_data = None
        matched_key = None

        for draft_id, data in pending.items():
            # conversation_id 매칭
            if sent.get("conversation_id") and data.get("conversation_id"):
                if sent["conversation_id"] == data["conversation_id"]:
                    matched_data = data
                    matched_key = draft_id
                    break
            # 제목 매칭 (fallback)
            name_in_subject = data.get("name", "")
            if name_in_subject and name_in_subject in sent["subject"]:
                matched_data = data
                matched_key = draft_id
                break

        if not matched_data:
            print(f"    → pending 매칭 실패, 제목에서 직접 파싱 시도")
            # 제목에서 이름 추출: "[휴가 승인] 심원기 님 휴가 승인 안내"
            m = re.search(rf"\{DRAFT_SUBJECT_TAG}\s*(.+?)\s*님", sent["subject"])
            if m:
                name = m.group(1).strip()
                print(f"    → 이름 추출: {name} (엑셀 업데이트는 pending 데이터 필요, 스킵)")
            continue

        print(f"    → 매칭: {matched_data['name']} / {matched_data['start_date']}~{matched_data['end_date']}")

        if update_excel(matched_data, dry_run=dry_run):
            synced += 1
            if matched_key and matched_key in pending:
                del pending[matched_key]

    _save_json(PROCESSED_SENT, list(processed_sent))
    _save_json(PENDING_PATH, pending)
    print(f"\n  Phase 2 완료: 엑셀 {synced}건 업데이트")
    return synced


# ── Status ───────────────────────────────────────────────────────────────────
def cmd_status():
    print("\n=== 휴가 워크플로우 상태 ===\n")

    # Pending drafts
    pending = _load_json(PENDING_PATH) if isinstance(_load_json(PENDING_PATH), dict) else {}
    print(f"  대기 중 드래프트: {len(pending)}건")
    for draft_id, data in pending.items():
        print(f"   {data.get('name','')} / {data.get('start_date','')}~{data.get('end_date','')} "
              f"({data.get('days','')}일 {data.get('leave_type','')})")

    # Processed counts
    inbox_processed = _load_json(PROCESSED_INBOX)
    sent_processed = _load_json(PROCESSED_SENT)
    inbox_count = len(inbox_processed) if isinstance(inbox_processed, list) else 0
    sent_count = len(sent_processed) if isinstance(sent_processed, list) else 0
    print(f"\n  처리된 수신 메일: {inbox_count}건")
    print(f"  처리된 발송 메일: {sent_count}건")

    # Excel status
    print(f"\n  엑셀: {EXCEL_PATH}")
    print(f"  존재: {'O' if EXCEL_PATH.exists() else 'X'}")

    # Token status
    print(f"\n  Graph 토큰: {GRAPH_TOKEN_PATH}")
    print(f"  존재: {'O' if GRAPH_TOKEN_PATH.exists() else 'X (최초 인증 필요)'}")


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

    parser = argparse.ArgumentParser(description="휴가 승인 워크플로우 (2-Phase)")
    parser.add_argument("--run",        action="store_true", help="Phase 1 + 2 모두 실행")
    parser.add_argument("--draft-only", action="store_true", help="Phase 1만 (드래프트 생성)")
    parser.add_argument("--sync-only",  action="store_true", help="Phase 2만 (엑셀 업데이트)")
    parser.add_argument("--dry-run",    action="store_true", help="테스트 모드 (실제 변경 없음)")
    parser.add_argument("--status",     action="store_true", help="현재 상태 확인")
    parser.add_argument("--hours",      type=int, default=6, help="감시 범위 (시간, 기본 6)")
    args = parser.parse_args()

    if args.status:
        cmd_status()
    elif args.draft_only:
        phase1_create_drafts(dry_run=args.dry_run, hours=args.hours)
    elif args.sync_only:
        phase2_sync_excel(dry_run=args.dry_run, hours=args.hours * 4)
    elif args.run:
        phase1_create_drafts(dry_run=args.dry_run, hours=args.hours)
        phase2_sync_excel(dry_run=args.dry_run, hours=args.hours * 4)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
