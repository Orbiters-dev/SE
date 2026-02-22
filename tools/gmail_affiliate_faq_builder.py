"""
gmail_affiliate_faq_builder.py

Gmail 받은편지함(affiliates@onzenna.com)의 이메일을 수집하여
인플루언서/어필리에이트 협업 문의 FAQ를 자동 생성합니다.

Output:
  - Google Sheets (Affiliate_FAQ 시트)
  - .tmp/gmail_affiliate_faq/affiliate_faq_YYYY-MM-DD.xlsx (로컬 백업)

Usage:
    python tools/gmail_affiliate_faq_builder.py [--months 3] [--dry-run]

Prerequisites:
    - Google Cloud Console에서 Gmail API 활성화
    - OAuth 2.0 Client ID(Desktop App) 생성 후 JSON 다운로드
      → credentials/gmail_oauth_credentials.json 에 저장
    - .env: ANTHROPIC_API_KEY
    - .env: GOOGLE_SERVICE_ACCOUNT_PATH=credentials/google_service_account.json
    - .env: GMAIL_OAUTH_CREDENTIALS_PATH=credentials/gmail_oauth_credentials.json
    - .env: AFFILIATE_FAQ_SHEET_ID=<Google Sheets ID>
    - 첫 실행 시 브라우저 팝업 → affiliates@onzenna.com 로그인 후 승인
      이후 실행은 credentials/gmail_token.json으로 자동 인증
"""

import os
import sys
import json
import argparse
import time
import re
import base64
import email
import html
from datetime import datetime, timezone, timedelta
from collections import defaultdict

from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment
import anthropic

# Force UTF-8 output on Windows
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

load_dotenv()

ANTHROPIC_KEY       = os.getenv("ANTHROPIC_API_KEY")
SA_PATH             = os.getenv("GOOGLE_SERVICE_ACCOUNT_PATH", "credentials/google_service_account.json")
OAUTH_CREDS_PATH    = os.getenv("GMAIL_OAUTH_CREDENTIALS_PATH", "credentials/gmail_oauth_credentials.json")
TOKEN_PATH          = os.getenv("GMAIL_TOKEN_PATH", "credentials/gmail_token.json")
SHEET_ID            = os.getenv("AFFILIATE_FAQ_SHEET_ID", "")

AI = anthropic.Anthropic(api_key=ANTHROPIC_KEY) if ANTHROPIC_KEY else None

GMAIL_SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]

SHEET_NAME = "Affiliate_FAQ"
SHEET_HEADERS = [
    "faq_id",
    "Category",
    "Partnership Type",
    "Email Count",
    "Question Variations",
    "Answer",
    "Internal Notes",
    "Related Keywords",
    "Update Log",
]

# High Touch: paid collab, contract, custom content guide, budget negotiation
# Low Touch:  gifted/sample only, single link, standard process, no payment

FAQ_CATEGORIES = [
    "How to Join & Get Started",           # 가입/시작 방법
    "Product Selection & What We Send",    # 어떤 제품 받는지
    "Shipping & Delivery of Product",      # 제품 배송
    "Commission Rate & Structure",         # 커미션 비율/구조
    "Payment Schedule & Method",           # 정산 일정/방법
    "Promo Code & Affiliate Link",         # 프로모 코드/링크 발급
    "Content Requirements & Posting",      # 콘텐츠 요건/게시 규칙
    "Eligibility & Follower Requirements", # 자격 요건
    "Exclusivity & Brand Conflicts",       # 타 브랜드 협업 제한
    "Timeline & Next Steps",               # 다음 단계/타임라인
    "Other",
]

# Auto-reply/spam subject patterns to skip
SKIP_SUBJECT_PATTERNS = [
    r"no.?reply",
    r"auto.?reply",
    r"out of office",
    r"unsubscribe",
    r"자동 회신",
    r"부재중",
    r"delivery failure",
    r"mailer.daemon",
]

# ---------------------------------------------------------------------------
# Step 1: Gmail OAuth 인증 + 이메일 페치
# ---------------------------------------------------------------------------

def get_gmail_service():
    """
    OAuth 2.0으로 Gmail API 서비스 객체를 반환합니다.
    최초 실행 시 브라우저 팝업이 열립니다.
    이후 실행은 credentials/gmail_token.json으로 자동 인증합니다.
    """
    try:
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
    except ImportError:
        print("ERROR: Gmail API 패키지가 설치되어 있지 않습니다.")
        print("  pip install google-auth-oauthlib google-auth-httplib2 google-api-python-client")
        sys.exit(1)

    creds = None

    # 기존 토큰 로드
    if os.path.exists(TOKEN_PATH):
        creds = Credentials.from_authorized_user_file(TOKEN_PATH, GMAIL_SCOPES)

    # 토큰이 없거나 만료된 경우
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            print("  Gmail 토큰 갱신 완료.")
        else:
            if not os.path.exists(OAUTH_CREDS_PATH):
                print(f"ERROR: OAuth 자격증명 파일을 찾을 수 없습니다: {OAUTH_CREDS_PATH}")
                print("  Google Cloud Console에서 OAuth 2.0 Client ID(Desktop App)를 생성하고")
                print(f"  JSON을 {OAUTH_CREDS_PATH}에 저장하세요.")
                sys.exit(1)
            flow = InstalledAppFlow.from_client_secrets_file(OAUTH_CREDS_PATH, GMAIL_SCOPES)
            print("  브라우저가 열립니다. affiliates@onzenna.com 계정으로 로그인하고 권한을 승인하세요.")
            creds = flow.run_local_server(port=0)

        # 토큰 저장
        os.makedirs(os.path.dirname(TOKEN_PATH), exist_ok=True)
        with open(TOKEN_PATH, "w") as f:
            f.write(creds.to_json())
        print(f"  Gmail 토큰 저장됨: {TOKEN_PATH}")

    from googleapiclient.discovery import build
    service = build("gmail", "v1", credentials=creds)
    return service


def fetch_emails(service, months: int) -> list[dict]:
    """
    받은편지함에서 최근 N개월간의 이메일을 페이지네이션으로 모두 가져옵니다.
    자동회신/스팸 패턴의 이메일은 제외합니다.
    """
    cutoff = datetime.now(timezone.utc) - timedelta(days=months * 30)
    # Gmail API 날짜 쿼리 (after:YYYY/MM/DD)
    after_str = cutoff.strftime("%Y/%m/%d")
    query = f"in:inbox after:{after_str}"

    print(f"Step 1: Gmail 이메일 페치 중 ({after_str} 이후, 쿼리: \"{query}\")...")

    all_ids = []
    page_token = None
    page = 1

    while True:
        params = {"userId": "me", "q": query, "maxResults": 500}
        if page_token:
            params["pageToken"] = page_token

        result = service.users().messages().list(**params).execute()
        messages = result.get("messages", [])
        all_ids.extend(messages)
        print(f"  Page {page}: {len(messages)}개 메시지 ID 수집 (누계: {len(all_ids)})")

        page_token = result.get("nextPageToken")
        if not page_token:
            break
        page += 1

    print(f"  총 {len(all_ids)}개 메시지 ID 수집 완료. 본문 파싱 시작...\n")

    skip_re = re.compile("|".join(SKIP_SUBJECT_PATTERNS), re.IGNORECASE)
    emails_data = []

    for i, msg_ref in enumerate(all_ids, 1):
        msg_id = msg_ref["id"]
        try:
            msg = service.users().messages().get(
                userId="me", id=msg_id, format="full"
            ).execute()
        except Exception as e:
            print(f"  [{i}/{len(all_ids)}] SKIP: 메시지 가져오기 실패 (id={msg_id}): {e}")
            continue

        parsed = parse_email(msg)
        if not parsed:
            continue

        # 자동회신/스팸 필터
        if skip_re.search(parsed["subject"]):
            continue

        emails_data.append(parsed)

        if i % 50 == 0:
            print(f"  {i}/{len(all_ids)} 파싱 완료...")
        time.sleep(0.05)  # Gmail API 속도 제한 방지

    print(f"\n  최종 수집: {len(emails_data)}개 이메일 (필터 후)\n")
    return emails_data


# ---------------------------------------------------------------------------
# Step 2: 이메일 파싱
# ---------------------------------------------------------------------------

def _extract_text_from_html(html_body: str) -> str:
    """HTML 태그 제거 후 평문 텍스트 반환."""
    from html.parser import HTMLParser

    class MLStripper(HTMLParser):
        def __init__(self):
            super().__init__()
            self.reset()
            self.fed = []

        def handle_data(self, d):
            self.fed.append(d)

        def get_data(self):
            return " ".join(self.fed)

    s = MLStripper()
    s.feed(html_body)
    text = s.get_data()
    # HTML 엔티티 디코딩
    text = html.unescape(text)
    # 과도한 공백 정리
    text = re.sub(r'\s{3,}', '\n\n', text)
    return text.strip()


def _decode_part(part: dict) -> str:
    """메시지 파트에서 텍스트를 디코딩합니다."""
    data = part.get("body", {}).get("data", "")
    if not data:
        return ""
    try:
        decoded = base64.urlsafe_b64decode(data + "==").decode("utf-8", errors="replace")
        return decoded
    except Exception:
        return ""


def _extract_body(payload: dict) -> str:
    """
    이메일 payload에서 본문 텍스트를 추출합니다.
    plain text 우선, 없으면 HTML을 파싱.
    """
    mime_type = payload.get("mimeType", "")
    parts = payload.get("parts", [])

    # 단일 파트
    if mime_type == "text/plain":
        return _decode_part(payload)
    if mime_type == "text/html":
        return _extract_text_from_html(_decode_part(payload))

    # 멀티파트: plain text 우선 탐색
    plain_text = ""
    html_text = ""

    def walk_parts(parts_list):
        nonlocal plain_text, html_text
        for part in parts_list:
            mt = part.get("mimeType", "")
            sub_parts = part.get("parts", [])
            if mt == "text/plain":
                plain_text += _decode_part(part)
            elif mt == "text/html":
                html_text += _extract_text_from_html(_decode_part(part))
            elif sub_parts:
                walk_parts(sub_parts)

    walk_parts(parts)

    return (plain_text or html_text).strip()


def parse_email(msg: dict) -> dict | None:
    """Gmail API 메시지 객체를 파싱하여 딕셔너리로 반환합니다."""
    payload = msg.get("payload", {})
    headers = {h["name"].lower(): h["value"] for h in payload.get("headers", [])}

    subject = headers.get("subject", "(제목 없음)")
    sender  = headers.get("from", "")
    date_str = headers.get("date", "")

    # 본문 추출
    body = _extract_body(payload)
    if not body.strip():
        return None

    # 길이 제한
    body = body[:2500]
    if len(body) == 2500:
        body += "\n...[truncated]"

    # 날짜 파싱 시도
    try:
        from email.utils import parsedate_to_datetime
        date_parsed = parsedate_to_datetime(date_str)
        date_formatted = date_parsed.strftime("%Y-%m-%d")
    except Exception:
        date_formatted = date_str[:10] if date_str else ""

    return {
        "msg_id":  msg.get("id", ""),
        "subject": subject,
        "sender":  sender,
        "date":    date_formatted,
        "body":    body,
    }


# ---------------------------------------------------------------------------
# Step 3: Claude Haiku로 이메일 분류
# ---------------------------------------------------------------------------

CLASSIFY_SYSTEM = """You are an expert classifier for influencer/affiliate program emails.

This brand runs TWO types of partnerships:
- HIGH TOUCH: Paid collaboration. Includes contract signing, custom content guidelines tailored to the creator's past content, paid fee/commission. The brand invests significantly per creator.
- LOW TOUCH: Gifted/sample only (no payment). Creator receives a product sample. Uses a single affiliate link that contains all guidelines. Simple, scalable process.

Read the email carefully to identify:
1. The SPECIFIC topic being asked about
2. Whether it relates to High Touch (paid, contract, custom) or Low Touch (gifted, link-based, standard), or General (applies to both / unclear)

Respond ONLY with valid JSON. No markdown, no explanation."""

CLASSIFY_PROMPT = """\
Subject: {subject}
From: {sender}
Date: {date}

Email body:
{body}

Classify this email on TWO dimensions:

CATEGORY — the specific topic being asked:
- "How to Join & Get Started" — how to sign up, apply, or begin
- "Product Selection & What We Send" — what product they'll receive, options, variants
- "Shipping & Delivery of Product" — shipping logistics for the gifted product
- "Commission Rate & Structure" — commission %, paid rates, budget, earnings structure
- "Payment Schedule & Method" — when/how they get paid
- "Promo Code & Affiliate Link" — their unique tracking link or discount code
- "Content Requirements & Posting" — what to create, posting rules, platform guidelines, deadlines
- "Eligibility & Follower Requirements" — whether they qualify, minimum followers, niche fit
- "Exclusivity & Brand Conflicts" — restrictions on working with competing brands
- "Timeline & Next Steps" — what happens next, turnaround time, process steps
- "Other" — does not fit above

PARTNERSHIP TYPE — which tier does this email relate to:
- "High Touch" — signals: mentions payment/budget/rate, contract, custom guidelines, negotiation, paid partnership
- "Low Touch" — signals: asks about gifted/sample/free product, affiliate link, standard process, no mention of payment
- "General" — unclear, could apply to both, or early-stage inquiry before tier is decided

Return ONLY this JSON (no markdown fences):
{{
  "category": "<exact category name>",
  "partnership_type": "<High Touch | Low Touch | General>",
  "core_question": "<one sentence: the specific question this person is asking>",
  "is_faq_worthy": <true if this is a common/recurring question type, false if purely personal or one-off>
}}"""


def classify_email(idx: int, total: int, email_data: dict) -> dict | None:
    """Claude Haiku로 단일 이메일 분류."""
    if not AI:
        return None

    prompt = CLASSIFY_PROMPT.format(
        subject=email_data["subject"],
        sender=email_data["sender"],
        date=email_data["date"],
        body=email_data["body"],
    )

    try:
        resp = AI.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=300,
            system=CLASSIFY_SYSTEM,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        raw = re.sub(r'^```(?:json)?\s*', '', raw)
        raw = re.sub(r'\s*```$', '', raw)
        result = json.loads(raw)
        return result
    except json.JSONDecodeError as e:
        print(f"    WARN: JSON 파싱 오류 [{idx}/{total}]: {e}")
        return None
    except Exception as e:
        print(f"    WARN: Haiku 분류 실패 [{idx}/{total}]: {e}")
        return None


# ---------------------------------------------------------------------------
# Step 4: Claude Sonnet으로 FAQ 항목 합성
# ---------------------------------------------------------------------------

_LOW_TOUCH_TERMS = """
LOW TOUCH collaboration terms (gifted/sample, no payment):
- Video: 30 seconds total length
- Must include voiceover + subtitles
- Must use royalty-free music
- Must tag: @zezebaebae_official (IG), @zeze_baebae (TikTok), @grosmimi_usa (IG & TikTok)
- Hashtags: #Grosmimi #PPSU #sippycup #ppsusippycup #Onzenna
- Upload within 1 week of receiving the product
- Must provide original content file (no subtitles/music overlaid) for whitelisting
- Whitelisting code provided upon upload
"""

_HIGH_TOUCH_TERMS = """
HIGH TOUCH collaboration terms (paid partnership, from contract template):
- Deliverables: typically 2 Instagram Reels + 2 TikTok videos featuring Grosmimi products
- Payment: [ ] USD via PayPal, NET 30 days after posting — ALWAYS use [ ] as placeholder, never quote a specific amount
- Timeline: post within 14 days of product delivery; payment may be withheld if not on schedule
- Must tag @onzenna and @grosmimi_usa; include #Ad or #Sponsored disclosure
- Provide original files without subtitles/music (for whitelisting use)
- All content subject to brand review and approval before posting
- Content may be used in Meta whitelisting and Spark Ads campaigns
- Payment terms are confidential
"""

SYNTHESIZE_SYSTEM = f"""You are an expert FAQ writer for an influencer/affiliate program (brands: Onzenna, Grosmimi, ZEZEBAEBAE).

CRITICAL RULES:
1. NEVER use the words "High Touch" or "Low Touch" in the answer field — these are internal labels only. Write naturally.
2. Keep answers CONCISE — 2-4 sentences maximum. Cut all filler.
3. For Commission/Payment questions where the specific rate or amount is variable: use [ ] as a placeholder (e.g., "Our paid rate is [ ] USD").
4. When writing for Low Touch (gifted) questions, incorporate the exact Low Touch terms below.
5. When writing for High Touch (paid) questions, reference the High Touch terms below.
6. Warm, professional tone in English.
{_LOW_TOUCH_TERMS}
{_HIGH_TOUCH_TERMS}
Respond ONLY with valid JSON. No markdown."""

SYNTHESIZE_PROMPT = """\
Category: {category}
Partnership Type (INTERNAL — never use this term in the answer): {partnership_type}
Number of real emails in this group: {count}

Sample emails from real inquiries:
{examples}

Write a FAQ entry. question_variations must capture DIFFERENT angles/sub-concerns — not rephrasing the same question.

Return ONLY this JSON (no markdown fences):
{{
  "question_variations": [
    "<Q1: most common angle>",
    "<Q2: different sub-concern>",
    "<Q3: another distinct angle>",
    "<Q4: edge case or nuanced version>",
    "<Q5: optional — only if genuinely distinct>"
  ],
  "answer": "<2-4 sentences, concise. Specific to the partnership type using the terms in your context. Never say High Touch or Low Touch. Use [ ] for variable rates/amounts.>",
  "internal_notes": "<INTERNAL USE: May use High Touch/Low Touch terminology here. What to check before replying, red flags, tier-specific actions. Be specific.>",
  "related_keywords": "<4-6 comma-separated keywords>"
}}"""


def synthesize_faq(category: str, partnership_type: str, examples: list[dict]) -> dict | None:
    """Claude Sonnet으로 카테고리+티어별 FAQ 항목 합성."""
    if not AI:
        return None

    # 최대 8개 사례 사용
    sample = examples[:8]
    example_texts = []
    for i, ex in enumerate(sample, 1):
        core_q = ex.get("core_question", "")
        subject = ex.get("subject", "")
        body_snippet = ex.get("body", "")[:350]
        example_texts.append(
            f"--- Example {i} ---\n"
            f"Subject: {subject}\n"
            f"Core question: {core_q}\n"
            f"Body snippet: {body_snippet}"
        )

    prompt = SYNTHESIZE_PROMPT.format(
        category=category,
        partnership_type=partnership_type,
        count=len(examples),
        examples="\n\n".join(example_texts),
    )

    try:
        resp = AI.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1800,
            system=SYNTHESIZE_SYSTEM,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        raw = re.sub(r'^```(?:json)?\s*', '', raw)
        raw = re.sub(r'\s*```$', '', raw)

        try:
            result = json.loads(raw)
            # question_variations를 bullet string으로 변환
            if isinstance(result.get("question_variations"), list):
                result["question_variations"] = "\n".join(
                    f"• {q}" for q in result["question_variations"] if q
                )
            return result
        except json.JSONDecodeError:
            pass

        # 정규식 fallback: answer와 internal_notes만 추출
        result = {}
        fields = {
            "answer":          r'"answer"\s*:\s*"((?:[^"\\]|\\.)*)"',
            "internal_notes":  r'"internal_notes"\s*:\s*"((?:[^"\\]|\\.)*)"',
            "related_keywords": r'"related_keywords"\s*:\s*"((?:[^"\\]|\\.)*)"',
        }
        for field, pattern in fields.items():
            m = re.search(pattern, raw, re.DOTALL)
            if m:
                result[field] = m.group(1).replace('\\n', '\n').replace('\\"', '"')

        # question_variations fallback: extract array items
        qv_matches = re.findall(r'"((?:[^"\\]|\\.){10,})"', raw)
        if qv_matches and len(qv_matches) >= 2:
            result["question_variations"] = "\n".join(f"• {q}" for q in qv_matches[:5])

        if len(result) >= 2:
            return result

        print(f"    WARN: JSON 파싱 및 정규식 모두 실패 ({category}/{partnership_type})")
        return None

    except Exception as e:
        print(f"    WARN: Sonnet 합성 실패 ({category}/{partnership_type}): {e}")
        return None


# ---------------------------------------------------------------------------
# Step 5: Google Sheets 기록
# ---------------------------------------------------------------------------

def write_to_sheets(records: list[dict], sheet_id: str, sa_path: str) -> bool:
    """Google Sheets에 FAQ 레코드를 작성합니다."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("ERROR: gspread 또는 google-auth가 설치되어 있지 않습니다.")
        print("  pip install gspread")
        return False

    if not os.path.exists(sa_path):
        print(f"ERROR: 서비스 계정 JSON을 찾을 수 없습니다: {sa_path}")
        return False

    if not sheet_id:
        print("ERROR: AFFILIATE_FAQ_SHEET_ID가 .env에 설정되어 있지 않습니다.")
        return False

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    try:
        creds = Credentials.from_service_account_file(sa_path, scopes=scopes)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(sheet_id)
    except Exception as e:
        print(f"ERROR: 스프레드시트를 열 수 없습니다: {e}")
        return False

    try:
        ws = sh.worksheet(SHEET_NAME)
        ws.clear()
        print(f"  기존 시트 초기화: {SHEET_NAME}")
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=SHEET_NAME, rows=300, cols=len(SHEET_HEADERS))
        print(f"  새 시트 생성: {SHEET_NAME}")

    ws.append_row(SHEET_HEADERS)

    for rec in records:
        row = [
            rec.get("faq_id", ""),
            rec.get("category", ""),
            rec.get("partnership_type", ""),
            rec.get("frequency", ""),
            rec.get("question_variations", ""),
            rec.get("answer", ""),
            rec.get("internal_notes", ""),
            rec.get("related_keywords", ""),
            rec.get("update_log", ""),
        ]
        ws.append_row(row)
        time.sleep(0.3)

    try:
        ws.format(f"A1:{chr(64 + len(SHEET_HEADERS))}1", {"textFormat": {"bold": True}})
    except Exception:
        pass

    print(f"  {len(records)}개 FAQ 항목을 Google Sheets에 저장 완료.")
    return True


# ---------------------------------------------------------------------------
# Step 6: Excel 로컬 백업
# ---------------------------------------------------------------------------

def write_excel_backup(records: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Affiliate FAQ"

    ws.append(SHEET_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    col_widths = [10, 32, 14, 12, 50, 65, 40, 28, 22]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wrap = Alignment(wrap_text=True, vertical="top")
    for rec in records:
        row = [
            rec.get("faq_id", ""),
            rec.get("category", ""),
            rec.get("partnership_type", ""),
            rec.get("frequency", ""),
            rec.get("question_variations", ""),
            rec.get("answer", ""),
            rec.get("internal_notes", ""),
            rec.get("related_keywords", ""),
            rec.get("update_log", ""),
        ]
        ws.append(row)
        for col_idx in range(1, len(SHEET_HEADERS) + 1):
            ws.cell(row=ws.max_row, column=col_idx).alignment = wrap
        ws.row_dimensions[ws.max_row].height = 90

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"  로컬 백업 저장: {output_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Gmail 어필리에이트 이메일로 FAQ 자동 생성"
    )
    parser.add_argument(
        "--months", type=int, default=3,
        help="페치할 이메일 기간 (개월, 기본값: 3)"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Google Sheets 저장 생략, Excel만 저장"
    )
    args = parser.parse_args()

    from output_utils import get_output_path
    output_path = get_output_path("influencer", "affiliate_faq")

    # ------------------------------------------------------------------
    # Step 1: Gmail 인증 + 이메일 페치
    # ------------------------------------------------------------------
    service = get_gmail_service()
    all_emails = fetch_emails(service, args.months)

    if not all_emails:
        print("수집된 이메일이 없습니다. 종료합니다.")
        return

    # ------------------------------------------------------------------
    # Step 2-3: 이메일 파싱 및 Haiku 분류
    # ------------------------------------------------------------------
    print(f"Step 2-3: {len(all_emails)}개 이메일 분류 중 (Claude Haiku)...")
    classified = []

    for i, em in enumerate(all_emails, 1):
        print(f"  [{i}/{len(all_emails)}] {em['date']} | {em['subject'][:50]}")
        result = classify_email(i, len(all_emails), em)

        if result:
            result["msg_id"]  = em["msg_id"]
            result["subject"] = em["subject"]
            result["sender"]  = em["sender"]
            result["date"]    = em["date"]
            result["body"]    = em["body"]
            classified.append(result)
            tier = result.get('partnership_type', 'General')
            print(f"    → [{tier}] {result.get('category')} | FAQ: {result.get('is_faq_worthy')}")
        else:
            print(f"    → 분류 실패 (스킵)")

        time.sleep(0.1)

    print(f"\n분류 완료: {len(classified)}개 / 전체 {len(all_emails)}개\n")

    # FAQ worthy 필터
    faq_emails = [e for e in classified if e.get("is_faq_worthy")]
    print(f"FAQ 대상 이메일: {len(faq_emails)}개 (is_faq_worthy=true)\n")

    if not faq_emails:
        print("FAQ 생성 대상 이메일이 없습니다. --months 값을 늘려보세요.")
        return

    # ------------------------------------------------------------------
    # Step 4: (카테고리 + 파트너십 유형)별 그룹핑 + Sonnet FAQ 합성
    # ------------------------------------------------------------------
    print("Step 4: 카테고리 × 파트너십 유형별 FAQ 합성 중 (Claude Sonnet)...")

    groups = defaultdict(list)
    for item in faq_emails:
        cat  = item.get("category", "Other")
        tier = item.get("partnership_type", "General")
        groups[(cat, tier)].append(item)

    print(f"  그룹 분포 (카테고리 × 티어):")
    for (cat, tier), items in sorted(groups.items(), key=lambda x: -len(x[1])):
        print(f"    [{tier}] {cat}: {len(items)}개")

    faqs = []
    today_log = datetime.now().strftime("%Y-%m-%d")

    for idx, ((category, partnership_type), examples) in enumerate(
        sorted(groups.items(), key=lambda x: -len(x[1])), start=1
    ):
        print(f"\n  [{idx}/{len(groups)}] 합성 중: [{partnership_type}] {category} ({len(examples)}개)")
        result = synthesize_faq(category, partnership_type, examples)

        if not result:
            result = {
                "question_variations": f"• Questions about: {category}",
                "answer": "(FAQ generation failed — please write manually)",
                "internal_notes": "Auto-generation failed",
                "related_keywords": category,
            }

        faqs.append({
            "faq_id":            f"AFF-{idx:03d}",
            "category":          category,
            "partnership_type":  partnership_type,
            "frequency":         len(examples),
            "question_variations": result.get("question_variations", ""),
            "answer":            result.get("answer", ""),
            "internal_notes":    result.get("internal_notes", ""),
            "related_keywords":  result.get("related_keywords", ""),
            "update_log":        f"v1 {today_log}: auto-generated from {len(examples)} emails",
        })

    print(f"\n{len(faqs)}개 FAQ 항목 생성 완료.\n")

    # ------------------------------------------------------------------
    # Step 5: Google Sheets 저장
    # ------------------------------------------------------------------
    if args.dry_run:
        print("Step 5: [DRY RUN] Google Sheets 저장 건너뜀.")
    else:
        print("Step 5: Google Sheets에 저장 중...")
        success = write_to_sheets(faqs, SHEET_ID, SA_PATH)
        if success:
            print(f"  Google Sheets URL: https://docs.google.com/spreadsheets/d/{SHEET_ID}")
        else:
            print("  Google Sheets 저장 실패 — 자격증명을 확인하세요. 로컬 Excel은 저장됩니다.")

    # ------------------------------------------------------------------
    # Step 6: Excel 로컬 백업
    # ------------------------------------------------------------------
    print(f"\nStep 6: Excel 로컬 백업 저장 중...")
    write_excel_backup(faqs, output_path)

    print(f"\n완료! {len(faqs)}개 FAQ 항목 생성.")
    if args.dry_run:
        print(f"결과물 확인: {output_path}")
        print("확인 후 --dry-run 없이 재실행하면 Google Sheets에 저장됩니다.")


if __name__ == "__main__":
    main()
