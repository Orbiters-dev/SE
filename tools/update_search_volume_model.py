"""
Update Search Volume tab in Polar_Financial_Model_new.xlsx

Data sources — completely separated, never mixed:
  Section 1: DataForSEO Google Ads API  → Google Search Volume (monthly historical)
  Section 2: DataForSEO Amazon API      → Amazon Search Volume (current snapshot)
  Section 3: Google Trends              → (not touched — manual source)
  Section 4: GSC zezebaebae.com         → NEW section appended at bottom (monthly impressions)

Headers updated to show source. Each section is independent.

Usage:
    python update_search_volume_model.py [--dry-run]
"""

import sys
import os
import argparse
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))

sys.path.insert(0, os.path.dirname(__file__))
from fetch_keyword_volume import (
    fetch_google_volume_historical,
    fetch_amazon_volume,
    fetch_gsc_monthly,
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
BASE            = os.path.dirname(os.path.dirname(__file__))
EXCEL_PATH      = os.path.join(BASE, "Data Storage", "Polar data", "Polar_Financial_Model_new.xlsx")
EXCEL_COPY      = os.path.join(BASE, "Data Storage", "Polar_Financial_Model_new.xlsx")
SHEET_NAME      = "Search Volume"
SERVICE_ACCOUNT = os.path.join(BASE, "credentials", "google_service_account.json")
GSC_SITE        = "sc-domain:zezebaebae.com"
DATE_FROM       = "2024-01-01"
DATE_TO         = "2026-02-17"

KEYWORDS = [
    ("Onzenna",          "onzenna"),
    ("zezebaebae",       "zezebaebae"),
    ("Grosmimi",         "grosmimi"),
    ("Alpremio",         "alpremio"),
    ("Cha&Mom",          "cha&mom"),
    ("Comme Moi",        "comme moi"),
    ("BabyRabbit",       "babyrabbit"),
    ("Naeiae",           "naeiae"),
    ("Bamboobebe",       "bamboobebe"),
    ("Hattung",          "hattung"),
    ("Beemymagic",       "beemymagic"),
    ("Nature Love Mere", "nature love mere"),
    ("PPSU",             "ppsu"),
    ("PPSU Bottle",      "ppsu bottle"),
    ("PPSU Baby Bottle", "ppsu baby bottle"),
    ("Phyto Seline",     "phyto seline"),
]

# Existing section row anchors (1-indexed)
GOOGLE_HEADER_ROW = 1    # Section title row
GOOGLE_COL_ROW    = 2    # Column header row
GOOGLE_START_ROW  = 3    # First keyword row (Onzenna)

AMAZON_HEADER_ROW = 20
AMAZON_COL_ROW    = 21
AMAZON_START_ROW  = 22

TRENDS_HEADER_ROW = 39   # Keep untouched

# GSC section — appended after Google Trends (which ends at row 56)
GSC_HEADER_ROW    = 58
GSC_COL_ROW       = 59
GSC_START_ROW     = 60

# Month columns: Jan 2024 (col 4 = D) → Feb 2026 (col 29 = AC)
MONTHS = []
for y in [2024, 2025, 2026]:
    for m in range(1, 13):
        if y == 2026 and m > 2:
            break
        MONTHS.append((y, m))
MONTH_TO_COL = {ym: 4 + i for i, ym in enumerate(MONTHS)}   # col 4 = Jan 2024

MONTH_LABELS = [f"{['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][m-1]} {y}"
                for y, m in MONTHS]


# ── Style helpers ─────────────────────────────────────────────────────────────
def style_section_header(cell, color_hex):
    cell.font      = Font(bold=True, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor=color_hex)
    cell.alignment = Alignment(wrap_text=True)

def style_col_header(cell):
    cell.font      = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")

def section(title):
    print(f"\n{'='*60}\n{title}\n{'='*60}")


# ── Main ──────────────────────────────────────────────────────────────────────
def run(dry_run=False):
    api_queries = [q for _, q in KEYWORDS]

    # Step 1: DataForSEO Google Ads
    section("Step 1: DataForSEO Google Ads — monthly historical")
    google_data = fetch_google_volume_historical(
        api_queries, location="US", date_from=DATE_FROM, date_to=DATE_TO
    )
    for name, q in KEYWORDS:
        d = google_data.get(q, {})
        print(f"  {name:<20}  avg={str(d.get('avg')):>8}  cpc={str(d.get('cpc')):>6}  "
              f"monthly_pts={len(d.get('monthly', {}))}")

    # Step 2: DataForSEO Amazon
    section("Step 2: DataForSEO Amazon — current snapshot")
    amazon_results = fetch_amazon_volume(api_queries, location="US")
    amazon_data = {r["keyword"]: r["search_volume"] for r in amazon_results}
    for name, q in KEYWORDS:
        print(f"  {name:<20}  current_vol={str(amazon_data.get(q)):>8}")

    # Step 3: GSC
    section("Step 3: GSC zezebaebae.com — monthly impressions")
    gsc_data = fetch_gsc_monthly(GSC_SITE, DATE_FROM, DATE_TO, SERVICE_ACCOUNT)
    for name, q in KEYWORDS:
        if q in gsc_data:
            total  = sum(gsc_data[q].values())
            n_mo   = len(gsc_data[q])
            print(f"  {name:<20}  total={total:>7,}  months={n_mo}")
        else:
            print(f"  {name:<20}  (no GSC data)")

    if dry_run:
        print("\n[DRY RUN] Skipping Excel write.")
        return

    # Step 4: Write Excel
    section("Step 4: Writing to Excel (sources separated)")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    last_col = max(MONTH_TO_COL.values())  # col 29

    # ── Update Section 1 header — DataForSEO Google Ads ──────────────────────
    ws.cell(row=GOOGLE_HEADER_ROW, column=1).value = (
        "GOOGLE SEARCH VOLUME — Source: DataForSEO Google Ads API  |  Monthly Absolute  |  US  |  Jan 2024 – Feb 2026"
        "  |  ✦ 전체 시장 수요: 구글에서 이 키워드를 검색한 총 횟수 (우리 사이트 여부 무관)"
        "  |  ✦ Total Market Demand: All searches on Google for this keyword (regardless of which site they visited)"
    )
    ws.merge_cells(start_row=GOOGLE_HEADER_ROW, start_column=1,
                   end_row=GOOGLE_HEADER_ROW, end_column=last_col)
    style_section_header(ws.cell(row=GOOGLE_HEADER_ROW, column=1), "2E75B6")

    # Write DataForSEO Google Ads data (pure — no GSC mixing)
    for i, (name, q) in enumerate(KEYWORDS):
        row = GOOGLE_START_ROW + i
        g   = google_data.get(q, {})
        if g.get("avg") is not None:
            ws.cell(row=row, column=2).value = g["avg"]
        if g.get("cpc") is not None:
            ws.cell(row=row, column=3).value = round(g["cpc"], 2)
        for ym, col in MONTH_TO_COL.items():
            val = g.get("monthly", {}).get(ym)
            if val is not None:
                ws.cell(row=row, column=col).value = val
        print(f"  [Google/DataForSEO] {name:<20}  avg={g.get('avg')}  monthly_pts={len(g.get('monthly',{}))}")

    # ── Update Section 2 header — DataForSEO Amazon ───────────────────────────
    ws.cell(row=AMAZON_HEADER_ROW, column=1).value = (
        "AMAZON SEARCH VOLUME — Source: DataForSEO Amazon API  |  Current Snapshot (no monthly breakdown)  |  US"
        "  |  ✦ 전체 시장 수요: 아마존에서 이 키워드를 검색한 총 횟수 (우리 제품 여부 무관)"
        "  |  ✦ Total Market Demand: All searches on Amazon for this keyword (regardless of which product they viewed)"
    )
    ws.merge_cells(start_row=AMAZON_HEADER_ROW, start_column=1,
                   end_row=AMAZON_HEADER_ROW, end_column=last_col)
    style_section_header(ws.cell(row=AMAZON_HEADER_ROW, column=1), "C55A11")

    # Write Amazon data (pure)
    for i, (name, q) in enumerate(KEYWORDS):
        row = AMAZON_START_ROW + i
        vol = amazon_data.get(q)
        if vol is not None:
            ws.cell(row=row, column=2).value = vol
            print(f"  [Amazon/DataForSEO] {name:<20}  vol={vol}")

    # ── Add Section 4 — GSC zezebaebae.com (new, appended) ───────────────────
    # Section header
    gsc_header_cell = ws.cell(row=GSC_HEADER_ROW, column=1)
    gsc_header_cell.value = (
        "GOOGLE SEARCH CONSOLE IMPRESSIONS — Source: GSC zezebaebae.com  |  Monthly Absolute  |  Jan 2024 – Feb 2026"
        "  |  ✦ 우리 사이트 노출량: 구글 검색 결과에 '우리 페이지'가 표시된 횟수 (전체 시장 수요 아님)"
        "  |  ✦ OUR SITE Exposure Only: Times OUR pages appeared in Google Search results for this keyword"
        "  |  ※ NOT total market demand — only reflects zezebaebae.com visibility in Google Search"
    )
    ws.merge_cells(start_row=GSC_HEADER_ROW, start_column=1,
                   end_row=GSC_HEADER_ROW, end_column=last_col)
    style_section_header(gsc_header_cell, "375623")

    # Column headers for GSC section
    ws.cell(row=GSC_COL_ROW, column=1).value = "Keyword"
    ws.cell(row=GSC_COL_ROW, column=2).value = "Avg Monthly Impressions"
    ws.cell(row=GSC_COL_ROW, column=3).value = "Total Impressions"
    for label, (ym, col) in zip(MONTH_LABELS, MONTH_TO_COL.items()):
        c = ws.cell(row=GSC_COL_ROW, column=col)
        c.value = label
        style_col_header(c)
    for col in [1, 2, 3]:
        style_col_header(ws.cell(row=GSC_COL_ROW, column=col))

    # GSC keyword data
    for i, (name, q) in enumerate(KEYWORDS):
        row     = GSC_START_ROW + i
        gsc_mo  = gsc_data.get(q, {})
        ws.cell(row=row, column=1).value = name

        if gsc_mo:
            total   = sum(gsc_mo.values())
            avg_imp = round(total / len(gsc_mo))
            ws.cell(row=row, column=2).value = avg_imp
            ws.cell(row=row, column=3).value = total
            for ym, col in MONTH_TO_COL.items():
                val = gsc_mo.get(ym)
                if val is not None:
                    ws.cell(row=row, column=col).value = val
            print(f"  [GSC] {name:<20}  avg={avg_imp:>6}  total={total:>7,}  months={len(gsc_mo)}")
        else:
            print(f"  [GSC] {name:<20}  (no data)")

    # Save
    ts     = datetime.now().strftime("%Y-%m-%d_%H%M")
    out_ts = EXCEL_PATH.replace(".xlsx", f"_updated_{ts}.xlsx")
    wb.save(out_ts)
    print(f"\n✓ Saved (timestamped): {out_ts}")
    wb.save(EXCEL_PATH)
    print(f"✓ Overwritten (Polar data): {EXCEL_PATH}")
    if os.path.exists(EXCEL_COPY):
        wb.save(EXCEL_COPY)
        print(f"✓ Overwritten (Data Storage): {EXCEL_COPY}")
    print("\nDone. 3 sources clearly separated in Search Volume tab.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    run(dry_run=args.dry_run)
