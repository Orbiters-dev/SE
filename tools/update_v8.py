"""
Update V8 Excel with live API data.

Pulls data from:
  - Rakuten RMS API   → RAKUTEN tab (monthly total sales → L=전체매출)
  - Amazon SP-API     → AMAZON tab (monthly total sales → L=전체매출)
  - Meta Marketing API → META tab (campaign-level monthly data, all columns)

V8 column structure:
  RAKUTEN/AMAZON: B=Month, C=Period, D=Budget, E=Ad Spend, F=Impr, G=Clicks,
                  H=CTR, I=CPC, J=Ad Sales, K=ACOS, L=전체매출, M=Notes
  META:           B=Month, C=Campaign, D=Budget, E=Ad Spend, F=Impressions,
                  G=Reach, H=Link Clicks, I=CTR(Link), J=CPC, K=Frequency, L=Notes

Usage:
  python tools/update_v8.py                    # update current month (all channels)
  python tools/update_v8.py 2026-01            # update specific month
  python tools/update_v8.py --all              # update all months (Nov 2025 ~ current)
"""
import os
os.environ["PYTHONIOENCODING"] = "utf-8"

import sys, argparse, calendar
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

from openpyxl import load_workbook

# Import API modules
sys.path.insert(0, str(Path(__file__).parent))
from rakuten_api import weekly_sales as rakuten_weekly_sales
from amazon_sp_api import weekly_sales as amazon_weekly_sales
import meta_api

# Override Meta credentials for Japan account
# Falls back to META_AD_ACCOUNT_ID / META_ACCESS_TOKEN if JP-specific keys not set
meta_api.AD_ACCOUNT_ID = os.getenv("META_JP_AD_ACCOUNT_ID") or os.getenv("META_AD_ACCOUNT_ID")
meta_api.ACCESS_TOKEN = os.getenv("META_JP_ACCESS_TOKEN") or os.getenv("META_ACCESS_TOKEN")
from meta_api import get_campaign_insights

EXCEL_PATH = Path(__file__).parent.parent / "Japan_Marketing Plan_Monthly_V8.xlsx"

# ── KRW → JPY conversion ─────────────────────────────────────────
KRW_TO_JPY = None

def get_krw_to_jpy():
    global KRW_TO_JPY
    if KRW_TO_JPY is not None:
        return KRW_TO_JPY
    import requests
    try:
        resp = requests.get("https://api.exchangerate-api.com/v4/latest/KRW", timeout=10)
        rate = resp.json()["rates"]["JPY"]
        KRW_TO_JPY = rate
        print(f"  [FX] KRW→JPY rate: {rate}")
    except Exception:
        KRW_TO_JPY = 0.107
        print(f"  [FX] Using fallback KRW→JPY rate: {KRW_TO_JPY}")
    return KRW_TO_JPY


# ── V8 (V1 structure) row mapping ────────────────────────────────
# Each month block = 5 rows (4 categories + TOTAL)
# RAKUTEN: ROOM, RPP, RMS, e.t.c, TOTAL
# AMAZON: SP, SB, SD, e.t.c, TOTAL

MONTHS = [
    "2025-11", "2025-12", "2026-01", "2026-02",
    "2026-03", "2026-04", "2026-05", "2026-06",
]

# Row mapping: month → (block_start, total_row)
# Block starts at row 3, 5 rows per month
def _rak_amz_rows(month_key):
    """Get (block_start, total_row) for RAKUTEN/AMAZON tab."""
    idx = MONTHS.index(month_key)
    start = 3 + idx * 5
    total = start + 4
    return start, total

# META has variable row counts in V1. We need to map specifically.
# V1 META: Nov=3-7(T=7), Dec=8-14(T=14), Jan=15-23(T=23)
# V8 added: Feb=25-31(T=31), Mar=32-38(T=38), Apr=39-45(T=45), May=46-52(T=52), Jun=53-59(T=59)
META_BLOCKS = {
    "2025-11": {"start": 3, "total": 7, "campaign_start": 3, "campaign_slots": 4},
    "2025-12": {"start": 8, "total": 14, "campaign_start": 8, "campaign_slots": 5},
    "2026-01": {"start": 15, "total": 23, "campaign_start": 15, "campaign_slots": 7},
    "2026-02": {"start": 25, "total": 31, "campaign_start": 25, "campaign_slots": 4},
    "2026-03": {"start": 32, "total": 38, "campaign_start": 32, "campaign_slots": 4},
    "2026-04": {"start": 39, "total": 45, "campaign_start": 39, "campaign_slots": 4},
    "2026-05": {"start": 46, "total": 52, "campaign_start": 46, "campaign_slots": 4},
    "2026-06": {"start": 53, "total": 59, "campaign_start": 53, "campaign_slots": 4},
}

# Overview row mapping
OV_BLOCKS = {
    "2025-11": {"rak": 4, "amz": 5, "meta": 6, "inf": 7, "total": 10},
    "2025-12": {"rak": 13, "amz": 14, "meta": 15, "inf": 16, "total": 19},
    "2026-01": {"rak": 22, "amz": 23, "meta": 24, "inf": 25, "total": 28},
    "2026-02": {"rak": 31, "amz": 32, "meta": 33, "inf": 34, "total": 37},
    "2026-03": {"rak": 40, "amz": 41, "meta": 42, "inf": 43, "total": 46},
    "2026-04": {"rak": 49, "amz": 50, "meta": 51, "inf": 52, "total": 55},
    "2026-05": {"rak": 58, "amz": 59, "meta": 60, "inf": 61, "total": 64},
    "2026-06": {"rak": 67, "amz": 68, "meta": 69, "inf": 70, "total": 73},
}


def _week_dates(year, month, week_num):
    """Get start/end dates for a week within a month."""
    last_day = calendar.monthrange(year, month)[1]
    starts = [1, 8, 15, 22, 29]
    ends = [7, 14, 21, 28, last_day]
    if week_num < 1 or week_num > 5:
        return None, None
    start = starts[week_num - 1]
    end = min(ends[week_num - 1], last_day)
    if start > last_day:
        return None, None
    return (f"{year}-{month:02d}-{start:02d}",
            f"{year}-{month:02d}-{end:02d}")


# ===================================================================
#  RAKUTEN: Monthly total sales (sum all weeks)
# ===================================================================
def update_rakuten(wb, year, month):
    """Fill RAKUTEN TOTAL row with monthly sales from API."""
    ws = wb["RAKUTEN"]
    month_key = f"{year}-{month:02d}"
    if month_key not in MONTHS:
        print(f"  [SKIP] {month_key} not in range")
        return

    _, total_row = _rak_amz_rows(month_key)
    print(f"\n  RAKUTEN {month_key}: Fetching weekly sales...")

    total_sales = 0
    total_orders = 0

    for w in range(1, 6):
        date_from, date_to = _week_dates(year, month, w)
        if date_from is None:
            continue
        try:
            result = rakuten_weekly_sales(date_from, date_to)
            total_sales += result["total_sales"]
            total_orders += result["order_count"]
            print(f"    W{w}: ¥{result['total_sales']:,} ({result['order_count']} orders)")
        except Exception as e:
            print(f"    W{w}: ERROR - {e}")

    # Write to TOTAL row: L=전체매출 (col 12)
    ws.cell(row=total_row, column=12).value = total_sales
    ws.cell(row=total_row, column=12).number_format = '[$¥-411]#,##0'

    # Also update Overview
    ov_row = OV_BLOCKS.get(month_key, {}).get("rak")
    if ov_row:
        ov = wb["Overview(Monthly)"]
        ov.cell(row=ov_row, column=18).value = total_sales
        ov.cell(row=ov_row, column=18).number_format = '[$¥-411]#,##0'

    print(f"  → TOTAL row {total_row}: 전체매출=¥{total_sales:,} ({total_orders} orders)")
    print(f"  [OK] RAKUTEN {month_key} updated")


# ===================================================================
#  AMAZON: Monthly total sales (sum all weeks)
# ===================================================================
def update_amazon(wb, year, month):
    """Fill AMAZON TOTAL row with monthly sales from API."""
    ws = wb["AMAZON"]
    month_key = f"{year}-{month:02d}"
    if month_key not in MONTHS:
        print(f"  [SKIP] {month_key} not in range")
        return

    _, total_row = _rak_amz_rows(month_key)
    print(f"\n  AMAZON {month_key}: Fetching weekly sales...")

    total_sales = 0
    total_orders = 0

    for w in range(1, 6):
        date_from, date_to = _week_dates(year, month, w)
        if date_from is None:
            continue
        try:
            result = amazon_weekly_sales(date_from, date_to)
            total_sales += result["total_sales"]
            total_orders += result["order_count"]
            print(f"    W{w}: ¥{result['total_sales']:,} ({result['order_count']} orders)")
        except Exception as e:
            print(f"    W{w}: ERROR - {e}")

    # L=전체매출 (col 12)
    ws.cell(row=total_row, column=12).value = total_sales
    ws.cell(row=total_row, column=12).number_format = '[$¥-411]#,##0'

    # Also update Overview
    ov_row = OV_BLOCKS.get(month_key, {}).get("amz")
    if ov_row:
        ov = wb["Overview(Monthly)"]
        ov.cell(row=ov_row, column=18).value = total_sales
        ov.cell(row=ov_row, column=18).number_format = '[$¥-411]#,##0'

    print(f"  → TOTAL row {total_row}: 전체매출=¥{total_sales:,} ({total_orders} orders)")
    print(f"  [OK] AMAZON {month_key} updated")


# ===================================================================
#  META: Campaign-level monthly data
# ===================================================================
def update_meta(wb, year, month):
    """Fill META tab with campaign-level data.

    V8 META column mapping (1-indexed):
      C=3: Campaign        E=5: Ad Spend (JPY)   F=6: Impressions
      G=7: Reach           H=8: Link Clicks       I=9: CTR(Link) %
      J=10: CPC (JPY)      K=11: Frequency        L=12: Notes
      D=4: Budget — manual, not touched
    """
    ws = wb["META"]
    month_key = f"{year}-{month:02d}"
    block = META_BLOCKS.get(month_key)
    if not block:
        print(f"  [SKIP] {month_key} not in range")
        return

    last_day = calendar.monthrange(year, month)[1]
    today = datetime.now()
    end_day = today.day if (year == today.year and month == today.month) else last_day

    date_from = f"{year}-{month:02d}-01"
    date_to = f"{year}-{month:02d}-{end_day:02d}"

    print(f"\n  META {month_key}: {date_from} ~ {date_to}")
    campaigns = get_campaign_insights(date_from, date_to)

    if not campaigns:
        print("    No campaign data found")
        return

    # Japan Meta account bills in JPY — no FX conversion needed
    campaigns.sort(key=lambda c: float(c.get("spend", 0)), reverse=True)

    # Column indices
    COL_CAMPAIGN = 3   # C
    COL_SPEND    = 5   # E
    COL_IMPR     = 6   # F
    COL_REACH    = 7   # G
    COL_CLICKS   = 8   # H
    COL_CTR      = 9   # I
    COL_CPC      = 10  # J
    COL_FREQ     = 11  # K
    COL_NOTES    = 12  # L
    ALL_COLS = [COL_CAMPAIGN, COL_SPEND, COL_IMPR, COL_REACH,
                COL_CLICKS, COL_CTR, COL_CPC, COL_FREQ, COL_NOTES]

    total_spend = 0
    max_slots = block["campaign_slots"]

    for ci, camp in enumerate(campaigns[:max_slots]):
        r = block["campaign_start"] + ci
        name   = camp.get("campaign_name", "Unknown")
        spend  = round(float(camp.get("spend", 0)))        # JPY
        impr   = int(camp.get("impressions", 0))
        reach  = int(camp.get("reach", 0))
        clicks = int(camp.get("clicks", 0))
        ctr    = float(camp.get("ctr", 0))                 # e.g. 1.23 → store as 0.0123
        cpc    = round(float(camp.get("cpc", 0)))          # JPY
        freq   = round(float(camp.get("frequency", 0)), 2)
        total_spend += spend

        ws.cell(row=r, column=COL_CAMPAIGN).value = name

        c = ws.cell(row=r, column=COL_SPEND, value=spend)
        c.number_format = '[$¥-411]#,##0'

        ws.cell(row=r, column=COL_IMPR, value=impr).number_format = '#,##0'
        ws.cell(row=r, column=COL_REACH, value=reach).number_format = '#,##0'
        ws.cell(row=r, column=COL_CLICKS, value=clicks).number_format = '#,##0'

        c = ws.cell(row=r, column=COL_CTR, value=ctr / 100)  # decimal for % format
        c.number_format = '0.00%'

        c = ws.cell(row=r, column=COL_CPC, value=cpc)
        c.number_format = '[$¥-411]#,##0'

        ws.cell(row=r, column=COL_FREQ, value=freq).number_format = '0.00'

        print(f"    [{ci+1}] {name[:35]}: ¥{spend:,} | {impr:,} impr | {clicks:,} clicks | CTR {ctr:.2f}% | CPC ¥{cpc:,}")

    # Clear remaining slots
    for ci in range(len(campaigns[:max_slots]), max_slots):
        r = block["campaign_start"] + ci
        for col in ALL_COLS:
            ws.cell(row=r, column=col).value = None

    # TOTAL row: E=Ad Spend total
    total_row = block["total"]
    c = ws.cell(row=total_row, column=COL_SPEND, value=total_spend)
    c.number_format = '[$¥-411]#,##0'

    # Update Overview
    ov_row = OV_BLOCKS.get(month_key, {}).get("meta")
    if ov_row:
        ov = wb["Overview(Monthly)"]
        ov.cell(row=ov_row, column=12).value = total_spend
        ov.cell(row=ov_row, column=12).number_format = '[$¥-411]#,##0'

    print(f"  → TOTAL: ¥{total_spend:,} ({len(campaigns[:max_slots])} campaigns)")
    print(f"  [OK] META {month_key} updated")


# ===================================================================
#  MAIN
# ===================================================================
def update(year=None, month=None, all_months=False):
    if year is None or month is None:
        now = datetime.now()
        year, month = now.year, now.month

    print("=" * 60)
    print(f"V8 Dashboard Update — {year}-{month:02d}")
    if all_months:
        print("  Mode: ALL MONTHS (Nov 2025 ~ current)")
    print("=" * 60)

    if not EXCEL_PATH.exists():
        print(f"[ERROR] V8 file not found: {EXCEL_PATH}")
        print("  Run '.tmp/build_v8.py' first.")
        return

    wb = load_workbook(str(EXCEL_PATH))

    # Determine months to update
    if all_months:
        months_to_update = []
        for key in MONTHS:
            y, m = key.split("-")
            y, m = int(y), int(m)
            if (y < year) or (y == year and m <= month):
                months_to_update.append((y, m))
    else:
        months_to_update = [(year, month)]

    for upd_year, upd_month in months_to_update:
        print(f"\n{'─'*40}")
        print(f"RAKUTEN — {upd_year}-{upd_month:02d}")
        print(f"{'─'*40}")
        try:
            update_rakuten(wb, upd_year, upd_month)
        except Exception as e:
            print(f"  [ERROR] Rakuten failed: {e}")

        print(f"\n{'─'*40}")
        print(f"AMAZON — {upd_year}-{upd_month:02d}")
        print(f"{'─'*40}")
        try:
            update_amazon(wb, upd_year, upd_month)
        except Exception as e:
            print(f"  [ERROR] Amazon failed: {e}")

        print(f"\n{'─'*40}")
        print(f"META — {upd_year}-{upd_month:02d}")
        print(f"{'─'*40}")
        try:
            update_meta(wb, upd_year, upd_month)
        except Exception as e:
            print(f"  [ERROR] Meta failed: {e}")

    # Save
    wb.save(str(EXCEL_PATH))
    print(f"\n{'='*60}")
    print(f"[OK] Saved: {EXCEL_PATH.name}")
    print(f"{'='*60}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Update V8 dashboard with API data")
    parser.add_argument("month", nargs="?", help="YYYY-MM (default: current month)")
    parser.add_argument("--all", "-a", action="store_true",
                        help="Update all months (Nov 2025 ~ current)")
    args = parser.parse_args()

    year, month = None, None
    if args.month:
        parts = args.month.split("-")
        year, month = int(parts[0]), int(parts[1])

    update(year, month, all_months=args.all)
