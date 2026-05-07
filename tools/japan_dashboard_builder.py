"""
Japan Marketing Plan Dashboard Builder
======================================
Input:  Japan_Marketing Plan_Monthly_V8.xlsx (or latest version)
Output: .tmp/Japan_Dashboard_<date>.xlsx

Tabs:
  1. Overview_Monthly   - Channel performance summary by month
  2. Ad_Performance     - Rakuten / Amazon / Meta detail
  3. CM_Analysis        - Contribution margin by product × channel
"""

import sys
import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

# ── paths ──────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE   = os.path.join(BASE_DIR, "Japan_Marketing Plan_Monthly_V8.xlsx")
TMP_DIR  = os.path.join(BASE_DIR, ".tmp")
os.makedirs(TMP_DIR, exist_ok=True)
OUTFILE  = os.path.join(TMP_DIR, f"Japan_Dashboard_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

# ── style helpers ──────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

COLORS = {
    "header_dark":  "1F2D3D",   # dark navy
    "header_mid":   "2E4057",   # mid navy
    "rakuten":      "BF0000",   # rakuten red
    "amazon":       "FF9900",   # amazon orange
    "meta":         "1877F2",   # meta blue
    "influencer":   "7B2D8B",   # purple
    "total":        "2E4057",   # navy
    "section":      "E8EDF2",   # light blue-gray
    "alt_row":      "F5F7FA",   # very light
    "white":        "FFFFFF",
    "green":        "27AE60",
    "red_light":    "FDEDEC",
    "cm_header":    "1A535C",   # teal
    "ppsu":         "E8F4FD",
    "stainless":    "FEF9E7",
}

CHANNEL_COLORS = {
    "Rakuten":     COLORS["rakuten"],
    "Amazon":      COLORS["amazon"],
    "Meta":        COLORS["meta"],
    "Influencer":  COLORS["influencer"],
    "LINE":        "00B900",
    "TOTAL":       COLORS["total"],
}

def hdr(ws, cell, value, bg=None, fg="FFFFFF", bold=True, size=10, align="center", wrap=False):
    c = ws[cell]
    c.value = value
    if bg:
        c.fill = _fill(bg)
    c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = _border()

def val(ws, cell, value, fmt=None, bold=False, bg=None, fg="000000", align="right"):
    c = ws[cell]
    c.value = value
    if fmt:
        c.number_format = fmt
    c.font = Font(bold=bold, color=fg, size=10, name="Calibri")
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _border()
    if bg:
        c.fill = _fill(bg)

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def freeze(ws, cell):
    ws.freeze_panes = cell

# ── data reading ───────────────────────────────────────────────────────────

def read_source():
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    return wb

def parse_overview(wb):
    """
    Returns list of dicts:
      month, channel, budget, target_roas, target_sales, cm_pct,
      ad_spend, exec_pct, ad_sales, direct_roas, total_sales,
      cm, cm_ads, roas_ex_inf, blended_roas
    """
    ws = wb["Overview(Monthly)"]
    rows = []
    current_month = None

    CHANNEL_ORDER = ["Rakuten", "Meta (WL → Rakuten)", "Influencer (→ Rakuten)",
                     "Amazon", "LINE", "TOTAL"]

    for row in ws.iter_rows(min_row=4, values_only=True):
        # col B=month, C=channel, D=budget, E=target_roas, F=target_sales,
        # G=cm_pct, H=ad_spend, I=exec_pct, J=ad_sales, K=direct_roas,
        # L=total_sales, M=cm, N=cm_ads, O=roas_ex_inf, P=blended_roas
        month_raw = row[1]
        channel   = row[2]

        if month_raw:
            current_month = str(month_raw).replace("\n", " ").strip()

        if not channel:
            continue

        channel_clean = str(channel).strip()
        if channel_clean not in CHANNEL_ORDER:
            continue

        rows.append({
            "month":        current_month,
            "channel":      channel_clean,
            "budget":       row[3],
            "target_roas":  row[4],
            "target_sales": row[5],
            "cm_pct":       row[6],
            "ad_spend":     row[7],
            "exec_pct":     row[8],
            "ad_sales":     row[9],
            "direct_roas":  row[10],
            "total_sales":  row[11],
            "cm":           row[12],
            "cm_ads":       row[13],
            "roas_ex_inf":  row[14],
            "blended_roas": row[15],
        })
    return rows

def parse_channel_weekly(wb, sheet_name):
    """
    Returns list of dicts for Rakuten/Amazon:
      month, period, budget, ad_spend, impr, clicks, ctr, cpc,
      ad_sales, acos, total_sales, notes
    """
    ws = wb[sheet_name]
    rows = []
    current_month = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        month_raw = row[1]
        period    = row[2]

        if month_raw:
            current_month = str(month_raw).strip()

        if not period or not current_month:
            continue

        rows.append({
            "month":       current_month,
            "period":      str(period).strip(),
            "budget":      row[3],
            "ad_spend":    row[4],
            "impr":        row[5],
            "clicks":      row[6],
            "ctr":         row[7],
            "cpc":         row[8],
            "ad_sales":    row[9],
            "acos":        row[10],
            "total_sales": row[11],
            "notes":       row[12],
        })
    return rows

def parse_meta(wb):
    """
    Returns list of dicts:
      month, campaign, budget, ad_spend, impressions, reach,
      link_clicks, ctr_link, cpc, frequency, notes
    """
    ws = wb["META"]
    rows = []
    current_month = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        month_raw = row[1]
        campaign  = row[2]

        if month_raw:
            current_month = str(month_raw).strip()

        if not current_month:
            continue
        if not campaign:
            continue

        rows.append({
            "month":       current_month,
            "campaign":    str(campaign).strip(),
            "budget":      row[3],
            "ad_spend":    row[4],
            "impressions": row[5],
            "reach":       row[6],
            "link_clicks": row[7],
            "ctr_link":    row[8],
            "cpc":         row[9],
            "frequency":   row[10],
            "notes":       row[11],
        })
    return rows

def parse_cm(wb):
    """
    Returns list of dicts per product × channel:
      product, channel, price, cogs, logistics, ch_fee,
      ch_fee_adj, returns, other, cm, cm_pct
    """
    ws = wb["채널별 공헌이익"]
    rows = []
    current_product = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        # col B=product/channel, C=price, D=cogs, E=logistics,
        # F=ch_fee, G=ch_fee_adj, H=returns, I=other, J=cm, K=cm_pct
        label = row[1]
        if not label:
            continue

        label_str = str(label).strip()

        # detect product headers
        if label_str in ("PPSU 300", "STAINLESS 300", "가중평균"):
            current_product = label_str
            continue

        # detect channel rows
        if label_str in ("Rakuten", "Amazon", "Meta", "Influencer", "LINE"):
            if current_product is None:
                continue
            # compute CM = price - cogs - logistics - ch_fee + ch_fee_adj - returns - other
            price    = row[2] or 0
            cogs     = row[3] or 0
            logistics= row[4] or 0
            ch_fee   = row[5] or 0
            ch_fee_adj = row[6] or 0
            returns  = row[7] or 0
            other    = row[8] or 0
            cm_val   = row[9]
            cm_pct_v = row[10]

            if cm_val is None:
                cm_val = price - cogs - logistics - ch_fee + (ch_fee_adj or 0) - returns - other
            if cm_pct_v is None and price:
                cm_pct_v = cm_val / price if price else None

            rows.append({
                "product":    current_product,
                "channel":    label_str,
                "price":      price,
                "cogs":       cogs,
                "logistics":  logistics,
                "ch_fee":     ch_fee,
                "ch_fee_adj": ch_fee_adj,
                "returns":    returns,
                "other":      other,
                "cm":         cm_val,
                "cm_pct":     cm_pct_v,
            })

    return rows

# ── TAB 1: Overview_Monthly ────────────────────────────────────────────────

def build_overview_tab(ws, data):
    ws.title = "Overview_Monthly"

    # ── title row
    ws.merge_cells("A1:P1")
    hdr(ws, "A1", "GROSMIMI JAPAN  —  MARKETING PERFORMANCE OVERVIEW (Monthly)",
        bg=COLORS["header_dark"], size=13, align="center")
    ws.row_dimensions[1].height = 28

    # ── section headers row 2
    for col, label, bg in [
        ("A", "", COLORS["header_dark"]),
        ("B", "Month", COLORS["header_dark"]),
        ("C", "Channel", COLORS["header_dark"]),
        ("D", "Budget\n(JPY)", COLORS["header_mid"]),
        ("E", "Target\nROAS", COLORS["header_mid"]),
        ("F", "Target\nSales", COLORS["header_mid"]),
        ("G", "CM %", COLORS["header_mid"]),
        ("H", "Ad Spend\n(JPY)", COLORS["header_dark"]),
        ("I", "Exec %", COLORS["header_dark"]),
        ("J", "Ad Sales\n(Direct)", COLORS["header_dark"]),
        ("K", "Direct\nROAS", COLORS["header_dark"]),
        ("L", "Total Sales\n(JPY)", COLORS["header_dark"]),
        ("M", "CM\n(JPY)", COLORS["header_dark"]),
        ("N", "CM − Ads\n(JPY)", COLORS["header_dark"]),
        ("O", "ROAS\n(ex. Inf)", COLORS["header_dark"]),
        ("P", "Blended\nROAS", COLORS["header_dark"]),
    ]:
        hdr(ws, f"{col}2", label, bg=bg, wrap=True)
    ws.row_dimensions[2].height = 36

    # ── data rows
    MONTH_ORDER = ["Nov 2025", "Dec 2025", "Jan 2026", "Feb 2026",
                   "Mar 2026", "Apr 2026", "May 2026", "Jun 2026"]
    CHANNEL_DISPLAY = {
        "Rakuten": "Rakuten",
        "Amazon": "Amazon",
        "Meta (WL → Rakuten)": "Meta (WL→Rakuten)",
        "Influencer (→ Rakuten)": "Influencer (→Rakuten)",
        "TOTAL": "TOTAL",
    }

    row_num = 3
    prev_month = None

    for month in MONTH_ORDER:
        month_rows = [r for r in data if r["month"] and month in r["month"]]
        if not month_rows:
            continue

        month_start = row_num

        for r in month_rows:
            ch = r["channel"]
            ch_display = CHANNEL_DISPLAY.get(ch, ch)
            is_total = (ch == "TOTAL")
            ch_color = CHANNEL_COLORS.get(ch.split(" ")[0], COLORS["header_mid"])

            # alternate row bg
            row_bg = None if is_total else (COLORS["alt_row"] if row_num % 2 == 0 else COLORS["white"])
            total_bg = COLORS["section"] if is_total else None
            use_bg = total_bg or row_bg

            # col A: color bar
            ws[f"A{row_num}"].fill = _fill(ch_color if not is_total else COLORS["header_dark"])
            ws[f"A{row_num}"].border = _border()

            val(ws, f"B{row_num}", month if ch == "Rakuten" else "", bg=use_bg, align="center", bold=is_total)
            val(ws, f"C{row_num}", ch_display,  bg=use_bg, align="left",  bold=is_total,
                fg=ch_color if not is_total else "FFFFFF")
            if is_total:
                ws[f"C{row_num}"].fill = _fill(COLORS["header_dark"])

            val(ws, f"D{row_num}", r["budget"],      fmt="#,##0", bg=use_bg)
            val(ws, f"E{row_num}", r["target_roas"], fmt="0.00",  bg=use_bg)
            val(ws, f"F{row_num}", r["target_sales"],fmt="#,##0", bg=use_bg)
            val(ws, f"G{row_num}", r["cm_pct"],      fmt="0.0%",  bg=use_bg)
            val(ws, f"H{row_num}", r["ad_spend"],    fmt="#,##0", bg=use_bg, bold=is_total)
            val(ws, f"I{row_num}", r["exec_pct"],    fmt="0.0%",  bg=use_bg)
            val(ws, f"J{row_num}", r["ad_sales"],    fmt="#,##0", bg=use_bg)
            val(ws, f"K{row_num}", r["direct_roas"], fmt="0.00",  bg=use_bg)
            val(ws, f"L{row_num}", r["total_sales"], fmt="#,##0", bg=use_bg, bold=is_total)
            val(ws, f"M{row_num}", r["cm"],          fmt="#,##0", bg=use_bg)
            val(ws, f"N{row_num}", r["cm_ads"],      fmt="#,##0", bg=use_bg)
            val(ws, f"O{row_num}", r["roas_ex_inf"], fmt="0.00",  bg=use_bg)
            val(ws, f"P{row_num}", r["blended_roas"],fmt="0.00",  bg=use_bg)

            row_num += 1

        # blank separator row
        for col in "ABCDEFGHIJKLMNOP":
            ws[f"{col}{row_num}"].fill = _fill("D0D7DE")
        row_num += 1

    # ── column widths
    set_col_widths(ws, {
        "A": 2.5, "B": 10, "C": 22, "D": 12, "E": 8, "F": 12,
        "G": 7, "H": 12, "I": 7, "J": 12, "K": 9, "L": 14,
        "M": 12, "N": 12, "O": 10, "P": 10,
    })
    freeze(ws, "D3")

    # ── footnote
    ws[f"B{row_num+1}"].value = (
        "* Blended ROAS = Rakuten Total Sales / (Rak+Meta+Inf Spend).  "
        "ROAS(ex.Inf) = Influencer 비용 제외.  Amazon: Direct ROAS only.  "
        "TOTAL row: MER = All Sales / All Spend."
    )
    ws[f"B{row_num+1}"].font = Font(italic=True, size=8, color="666666")
    ws.merge_cells(f"B{row_num+1}:P{row_num+1}")


# ── TAB 2: Ad_Performance ──────────────────────────────────────────────────

def build_ad_tab(ws, rak_data, amz_data, meta_data):
    ws.title = "Ad_Performance"

    # title
    ws.merge_cells("A1:M1")
    hdr(ws, "A1", "AD PERFORMANCE — Rakuten / Amazon / Meta",
        bg=COLORS["header_dark"], size=13)
    ws.row_dimensions[1].height = 28

    row_num = 2

    # ── RAKUTEN section
    ws.merge_cells(f"A{row_num}:M{row_num}")
    hdr(ws, f"A{row_num}", "RAKUTEN — Weekly Performance",
        bg=COLORS["rakuten"], size=11)
    ws.row_dimensions[row_num].height = 22
    row_num += 1

    for label, col in [
        ("Month", "A"), ("Period", "B"), ("Budget (JPY)", "C"),
        ("Ad Spend", "D"), ("Impr", "E"), ("Clicks", "F"),
        ("CTR", "G"), ("CPC", "H"), ("Ad Sales", "I"),
        ("ACOS", "J"), ("Total Sales", "K"), ("Notes", "L"),
    ]:
        hdr(ws, f"{col}{row_num}", label, bg=COLORS["header_mid"], wrap=True)
    ws.row_dimensions[row_num].height = 32
    row_num += 1

    MONTH_ORDER = ["Nov 2025", "Dec 2025", "Jan 2026", "Feb 2026",
                   "Mar 2026", "Apr 2026", "May 2026", "Jun 2026"]

    for month in MONTH_ORDER:
        month_rows = [r for r in rak_data if r["month"] == month]
        if not month_rows:
            continue
        first = True
        for r in month_rows:
            is_total = r["period"] == "TOTAL"
            bg = COLORS["section"] if is_total else (COLORS["alt_row"] if row_num % 2 == 0 else COLORS["white"])
            val(ws, f"A{row_num}", month if first else "",    bg=bg, align="center")
            val(ws, f"B{row_num}", r["period"],               bg=bg, align="center", bold=is_total)
            val(ws, f"C{row_num}", r["budget"],    fmt="#,##0", bg=bg)
            val(ws, f"D{row_num}", r["ad_spend"],  fmt="#,##0", bg=bg, bold=is_total)
            val(ws, f"E{row_num}", r["impr"],      fmt="#,##0", bg=bg)
            val(ws, f"F{row_num}", r["clicks"],    fmt="#,##0", bg=bg)
            val(ws, f"G{row_num}", r["ctr"],       fmt="0.00%", bg=bg)
            val(ws, f"H{row_num}", r["cpc"],       fmt="#,##0", bg=bg)
            val(ws, f"I{row_num}", r["ad_sales"],  fmt="#,##0", bg=bg, bold=is_total)
            val(ws, f"J{row_num}", r["acos"],      fmt="0.0%",  bg=bg)
            val(ws, f"K{row_num}", r["total_sales"],fmt="#,##0", bg=bg, bold=is_total)
            val(ws, f"L{row_num}", r["notes"] or "", bg=bg, align="left")
            first = False
            row_num += 1
        # separator
        for col in "ABCDEFGHIJKL":
            ws[f"{col}{row_num}"].fill = _fill("D0D7DE")
        row_num += 1

    row_num += 1

    # ── AMAZON section
    ws.merge_cells(f"A{row_num}:M{row_num}")
    hdr(ws, f"A{row_num}", "AMAZON — Weekly Performance",
        bg=COLORS["amazon"], fg="000000", size=11)
    ws.row_dimensions[row_num].height = 22
    row_num += 1

    for label, col in [
        ("Month", "A"), ("Period", "B"), ("Budget (JPY)", "C"),
        ("Ad Spend", "D"), ("Impr", "E"), ("Clicks", "F"),
        ("CTR", "G"), ("CPC", "H"), ("Ad Sales", "I"),
        ("ACOS", "J"), ("Total Sales", "K"), ("Notes", "L"),
    ]:
        hdr(ws, f"{col}{row_num}", label, bg=COLORS["header_mid"], wrap=True)
    ws.row_dimensions[row_num].height = 32
    row_num += 1

    for month in MONTH_ORDER:
        month_rows = [r for r in amz_data if r["month"] == month]
        if not month_rows:
            continue
        first = True
        for r in month_rows:
            is_total = r["period"] == "TOTAL"
            bg = COLORS["section"] if is_total else (COLORS["alt_row"] if row_num % 2 == 0 else COLORS["white"])
            val(ws, f"A{row_num}", month if first else "",    bg=bg, align="center")
            val(ws, f"B{row_num}", r["period"],               bg=bg, align="center", bold=is_total)
            val(ws, f"C{row_num}", r["budget"],    fmt="#,##0", bg=bg)
            val(ws, f"D{row_num}", r["ad_spend"],  fmt="#,##0", bg=bg, bold=is_total)
            val(ws, f"E{row_num}", r["impr"],      fmt="#,##0", bg=bg)
            val(ws, f"F{row_num}", r["clicks"],    fmt="#,##0", bg=bg)
            val(ws, f"G{row_num}", r["ctr"],       fmt="0.00%", bg=bg)
            val(ws, f"H{row_num}", r["cpc"],       fmt="#,##0", bg=bg)
            val(ws, f"I{row_num}", r["ad_sales"],  fmt="#,##0", bg=bg, bold=is_total)
            val(ws, f"J{row_num}", r["acos"],      fmt="0.0%",  bg=bg)
            val(ws, f"K{row_num}", r["total_sales"],fmt="#,##0", bg=bg, bold=is_total)
            val(ws, f"L{row_num}", r["notes"] or "", bg=bg, align="left")
            first = False
            row_num += 1
        for col in "ABCDEFGHIJKL":
            ws[f"{col}{row_num}"].fill = _fill("D0D7DE")
        row_num += 1

    row_num += 1

    # ── META section
    ws.merge_cells(f"A{row_num}:M{row_num}")
    hdr(ws, f"A{row_num}", "META — Campaign Performance",
        bg=COLORS["meta"], size=11)
    ws.row_dimensions[row_num].height = 22
    row_num += 1

    for label, col in [
        ("Month", "A"), ("Campaign", "B"), ("Budget (JPY)", "C"),
        ("Ad Spend", "D"), ("Impressions", "E"), ("Reach", "F"),
        ("Link Clicks", "G"), ("CTR (Link)", "H"), ("CPC", "I"),
        ("Frequency", "J"), ("Notes", "K"),
    ]:
        hdr(ws, f"{col}{row_num}", label, bg=COLORS["header_mid"], wrap=True)
    ws.row_dimensions[row_num].height = 32
    row_num += 1

    for month in MONTH_ORDER:
        month_rows = [r for r in meta_data if r["month"] == month]
        if not month_rows:
            continue
        first = True
        for r in month_rows:
            is_total = r["campaign"] == "TOTAL"
            bg = COLORS["section"] if is_total else (COLORS["alt_row"] if row_num % 2 == 0 else COLORS["white"])
            val(ws, f"A{row_num}", month if first else "",    bg=bg, align="center")
            val(ws, f"B{row_num}", r["campaign"],             bg=bg, align="left", bold=is_total)
            val(ws, f"C{row_num}", r["budget"],    fmt="#,##0", bg=bg)
            val(ws, f"D{row_num}", r["ad_spend"],  fmt="#,##0", bg=bg, bold=is_total)
            val(ws, f"E{row_num}", r["impressions"],fmt="#,##0", bg=bg)
            val(ws, f"F{row_num}", r["reach"],     fmt="#,##0", bg=bg)
            val(ws, f"G{row_num}", r["link_clicks"],fmt="#,##0", bg=bg)
            val(ws, f"H{row_num}", r["ctr_link"],  fmt="0.00%", bg=bg)
            val(ws, f"I{row_num}", r["cpc"],       fmt="#,##0", bg=bg)
            val(ws, f"J{row_num}", r["frequency"], fmt="0.00",  bg=bg)
            val(ws, f"K{row_num}", r["notes"] or "", bg=bg, align="left")
            first = False
            row_num += 1
        for col in "ABCDEFGHIJK":
            ws[f"{col}{row_num}"].fill = _fill("D0D7DE")
        row_num += 1

    set_col_widths(ws, {
        "A": 11, "B": 38, "C": 12, "D": 11, "E": 11,
        "F": 10, "G": 11, "H": 10, "I": 9, "J": 10,
        "K": 11, "L": 40,
    })
    freeze(ws, "C4")


# ── TAB 3: CM_Analysis ────────────────────────────────────────────────────

def build_cm_tab(ws, cm_data):
    ws.title = "CM_Analysis"

    ws.merge_cells("A1:L1")
    hdr(ws, "A1", "CHANNEL CONTRIBUTION MARGIN ANALYSIS (채널별 공헌이익)",
        bg=COLORS["cm_header"], size=13)
    ws.row_dimensions[1].height = 28

    headers = [
        ("A", "Product"),
        ("B", "Channel"),
        ("C", "판매가\n(Price)"),
        ("D", "원가\n(COGS)"),
        ("E", "물류비\n(Logistics)"),
        ("F", "채널수수료\n(Ch. Fee)"),
        ("G", "수수료조정\n(Adj.)"),
        ("H", "리턴\n(Returns)"),
        ("I", "기타\n(Other)"),
        ("J", "공헌이익\n(CM)"),
        ("K", "공헌이익률\n(CM %)"),
    ]
    for col, label in headers:
        hdr(ws, f"{col}2", label, bg=COLORS["header_mid"], wrap=True)
    ws.row_dimensions[2].height = 40

    row_num = 3
    prev_product = None

    PRODUCT_COLORS = {
        "PPSU 300":      COLORS["ppsu"],
        "STAINLESS 300": COLORS["stainless"],
        "가중평균":        COLORS["section"],
    }

    for r in cm_data:
        product = r["product"]
        ch = r["channel"]

        # product header row
        if product != prev_product:
            ws.merge_cells(f"A{row_num}:K{row_num}")
            ws[f"A{row_num}"].value = f"▶  {product}"
            ws[f"A{row_num}"].font = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
            ws[f"A{row_num}"].fill = _fill(COLORS["cm_header"])
            ws[f"A{row_num}"].alignment = Alignment(horizontal="left", vertical="center")
            ws[f"A{row_num}"].border = _border()
            ws.row_dimensions[row_num].height = 22
            row_num += 1
            prev_product = product

        ch_color = CHANNEL_COLORS.get(ch, COLORS["header_mid"])
        bg = PRODUCT_COLORS.get(product, COLORS["white"])

        # color bar in col A
        ws[f"A{row_num}"].fill = _fill(ch_color)
        ws[f"A{row_num}"].border = _border()

        val(ws, f"B{row_num}", ch, bg=bg, align="left", bold=True, fg=ch_color)
        val(ws, f"C{row_num}", r["price"],      fmt="#,##0", bg=bg)
        val(ws, f"D{row_num}", r["cogs"],       fmt="#,##0", bg=bg)
        val(ws, f"E{row_num}", r["logistics"],  fmt="#,##0", bg=bg)
        val(ws, f"F{row_num}", r["ch_fee"],     fmt="#,##0", bg=bg)
        val(ws, f"G{row_num}", r["ch_fee_adj"], fmt="#,##0", bg=bg)
        val(ws, f"H{row_num}", r["returns"],    fmt="#,##0", bg=bg)
        val(ws, f"I{row_num}", r["other"],      fmt="#,##0", bg=bg)
        val(ws, f"J{row_num}", r["cm"],         fmt="#,##0", bg=bg, bold=True)
        # CM%: color green if >= 30%, red if < 15%
        cm_pct_v = r["cm_pct"]
        pct_fg = "27AE60" if (cm_pct_v and cm_pct_v >= 0.30) else ("E74C3C" if (cm_pct_v and cm_pct_v < 0.15) else "000000")
        c = ws[f"K{row_num}"]
        c.value = cm_pct_v
        c.number_format = "0.0%"
        c.font = Font(bold=True, color=pct_fg, size=10, name="Calibri")
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = _border()
        if bg:
            c.fill = _fill(bg)

        row_num += 1

    # CM waterfall explanation
    row_num += 1
    ws[f"A{row_num}"].value = "공헌이익 산식: CM = 판매가 − 원가 − 물류비 − 채널수수료 + 수수료조정 − 리턴 − 기타"
    ws[f"A{row_num}"].font = Font(italic=True, size=8, color="666666")
    ws.merge_cells(f"A{row_num}:K{row_num}")

    row_num += 1
    ws[f"A{row_num}"].value = "※ 가중평균 가정: PPSU 80% / STAINLESS 20% | 리턴율 6% | 기타 10%"
    ws[f"A{row_num}"].font = Font(italic=True, size=8, color="666666")
    ws.merge_cells(f"A{row_num}:K{row_num}")

    set_col_widths(ws, {
        "A": 2.5, "B": 14, "C": 10, "D": 10, "E": 12,
        "F": 12, "G": 10, "H": 10, "I": 10, "J": 12, "K": 12,
    })
    freeze(ws, "C3")


# ── MAIN ───────────────────────────────────────────────────────────────────

def main():
    print(f"[1/5] Reading source: {SOURCE}")
    wb_src = read_source()

    print("[2/5] Parsing data...")
    overview_data = parse_overview(wb_src)
    rak_data      = parse_channel_weekly(wb_src, "RAKUTEN")
    amz_data      = parse_channel_weekly(wb_src, "AMAZON")
    meta_data     = parse_meta(wb_src)
    cm_data       = parse_cm(wb_src)
    wb_src.close()

    print(f"       Overview rows:  {len(overview_data)}")
    print(f"       Rakuten rows:   {len(rak_data)}")
    print(f"       Amazon rows:    {len(amz_data)}")
    print(f"       Meta rows:      {len(meta_data)}")
    print(f"       CM rows:        {len(cm_data)}")

    print("[3/5] Building dashboard workbook...")
    wb_out = Workbook()
    ws1 = wb_out.active
    ws2 = wb_out.create_sheet()
    ws3 = wb_out.create_sheet()

    print("[4/5] Writing tabs...")
    build_overview_tab(ws1, overview_data)
    build_ad_tab(ws2, rak_data, amz_data, meta_data)
    build_cm_tab(ws3, cm_data)

    print(f"[5/5] Saving → {OUTFILE}")
    wb_out.save(OUTFILE)
    print(f"\nDone! Dashboard saved: {OUTFILE}")


if __name__ == "__main__":
    main()
