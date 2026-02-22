"""Klaviyo Email Flow & Campaign Performance Dashboard.

Usage:
    python tools/klaviyo_email_dashboard.py

Reads JSON files from .tmp/polar_data/ and generates
.tmp/klaviyo_email_dashboard_YYYY-MM-DD_HHMM.xlsx
"""
import calendar
import json, os, sys
from collections import defaultdict
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.join(SCRIPT_DIR, "..")
DATA_DIR = os.path.join(ROOT_DIR, ".tmp", "polar_data")
from output_utils import get_output_path
OUTPUT = get_output_path("marketing", "klaviyo_dashboard")

# ── Styles ───────────────────────────────────────────────────────────────────
DBLUE = PatternFill("solid", fgColor="002060")
MBLUE = PatternFill("solid", fgColor="4472C4")
LGRAY = PatternFill("solid", fgColor="D9E2F3")
LLGRAY = PatternFill("solid", fgColor="F2F2F2")
LGREEN = PatternFill("solid", fgColor="E2EFDA")
LRED = PatternFill("solid", fgColor="FCE4D6")
ORANGE_TAB = "ED7D31"
GREEN_TAB = "548235"
PURPLE_TAB = "7030A0"
GRAY_TAB = "808080"
BLUE_TAB = "4472C4"

WF = Font(bold=True, size=10, color="FFFFFF")
HF = Font(bold=True, size=10)
BF = Font(bold=True)
NF = Font(size=9)
TITLE_F = Font(bold=True, size=14)
SUB_F = Font(italic=True, size=10, color="FF0000")

EST_F = Font(italic=True, size=9, color="808080")  # Gray italic for estimates

DOL = "$#,##0.00"
INT = "#,##0"
PCT = "0.0%"
THIN = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)

# ── Category Rules ───────────────────────────────────────────────────────────
FLOW_CATEGORY_RULES = [
    ("welcome", "Welcome/Onboarding"),
    ("pop-up", "Welcome/Onboarding"),
    ("popup", "Welcome/Onboarding"),
    ("abandoned cart", "Cart/Browse Recovery"),
    ("browse abandon", "Cart/Browse Recovery"),
    ("first purchase", "Post-Purchase/Retention"),
    ("bounce back", "Post-Purchase/Retention"),
    ("replenishment", "Post-Purchase/Retention"),
    ("reminder", "Post-Purchase/Retention"),
    ("winback", "Win-back"),
    ("win back", "Win-back"),
    ("back in stock", "Back in Stock"),
]
FLOW_CATEGORIES = [
    "Welcome/Onboarding",
    "Cart/Browse Recovery",
    "Post-Purchase/Retention",
    "Win-back",
    "Back in Stock",
    "Other",
]

CAMPAIGN_CATEGORY_RULES = [
    # Promotional/Sales
    ("bfcm", "Promotional/Sales"),
    ("black friday", "Promotional/Sales"),
    ("cyber monday", "Promotional/Sales"),
    ("prime day", "Promotional/Sales"),
    ("% off", "Promotional/Sales"),
    ("sale", "Promotional/Sales"),
    ("deal", "Promotional/Sales"),
    ("discount", "Promotional/Sales"),
    ("promo", "Promotional/Sales"),
    ("flash", "Promotional/Sales"),
    ("save", "Promotional/Sales"),
    ("saving", "Promotional/Sales"),
    ("free shipping", "Promotional/Sales"),
    ("limited time", "Promotional/Sales"),
    ("bundle", "Promotional/Sales"),
    ("last chance", "Promotional/Sales"),
    ("last day", "Promotional/Sales"),
    ("reminder", "Promotional/Sales"),
    ("ends today", "Promotional/Sales"),
    ("ends tonight", "Promotional/Sales"),
    ("off for", "Promotional/Sales"),
    ("gift card", "Promotional/Sales"),
    ("free gift", "Promotional/Sales"),
    ("free ", "Promotional/Sales"),
    ("gwp", "Promotional/Sales"),
    ("labor day", "Promotional/Sales"),
    ("memorial day", "Promotional/Sales"),
    ("4th of july", "Promotional/Sales"),
    ("presidents day", "Promotional/Sales"),
    ("mlk", "Promotional/Sales"),
    ("friends and family", "Promotional/Sales"),
    ("best seller", "Promotional/Sales"),
    ("love&care", "Promotional/Sales"),
    ("love & care", "Promotional/Sales"),
    ("sip-sational", "Promotional/Sales"),
    # Product Launch
    ("new arrival", "Product Launch"),
    ("launch", "Product Launch"),
    ("introducing", "Product Launch"),
    ("just dropped", "Product Launch"),
    ("pre-order", "Product Launch"),
    ("preorder", "Product Launch"),
    ("pre order", "Product Launch"),
    ("now available", "Product Launch"),
    ("sneak peek", "Product Launch"),
    ("announcing", "Product Launch"),
    ("new season", "Product Launch"),
    ("new style", "Product Launch"),
    ("collection", "Product Launch"),
    ("new arrivals", "Product Launch"),
    # Content/Educational
    ("guide", "Content/Educational"),
    ("tip", "Content/Educational"),
    ("how to", "Content/Educational"),
    ("toddler acne", "Content/Educational"),
    ("screen time", "Content/Educational"),
    ("blog", "Content/Educational"),
    ("review", "Content/Educational"),
    ("awards", "Content/Educational"),
    ("rewards program", "Content/Educational"),
    ("skincare", "Content/Educational"),
    ("skin care", "Content/Educational"),
    ("dry season", "Content/Educational"),
    ("3-step", "Content/Educational"),
    # Seasonal/Holiday
    ("holiday", "Seasonal/Holiday"),
    ("christmas", "Seasonal/Holiday"),
    ("easter", "Seasonal/Holiday"),
    ("valentine", "Seasonal/Holiday"),
    ("mother", "Seasonal/Holiday"),
    ("father", "Seasonal/Holiday"),
    ("new year", "Seasonal/Holiday"),
    ("thanksgiving", "Seasonal/Holiday"),
    ("lunar", "Seasonal/Holiday"),
    ("chuseok", "Seasonal/Holiday"),
    ("halloween", "Seasonal/Holiday"),
    ("back to school", "Seasonal/Holiday"),
    ("festive", "Seasonal/Holiday"),
    ("spring", "Seasonal/Holiday"),
    ("summer", "Seasonal/Holiday"),
    ("sunny", "Seasonal/Holiday"),
    ("winter", "Seasonal/Holiday"),
    ("fall ", "Seasonal/Holiday"),
    ("cozy", "Seasonal/Holiday"),
    # Back in Stock
    ("back in stock", "Back in Stock"),
    ("restock", "Back in Stock"),
]
CAMPAIGN_CATEGORIES = [
    "Promotional/Sales",
    "Product Launch",
    "Content/Educational",
    "Seasonal/Holiday",
    "Back in Stock",
    "Other",
]

# ── Metric Thresholds for conditional formatting ────────────────────────────
METRIC_THRESHOLDS = {
    "open_rate": {"good": 0.50, "bad": 0.30, "inverted": False},
    "click_rate": {"good": 0.03, "bad": 0.01, "inverted": False},
    "order_rate": {"good": 0.02, "bad": 0.005, "inverted": False},
    "bounce_rate": {"good": 0.005, "bad": 0.02, "inverted": True},
    "unsub_rate": {"good": 0.001, "bad": 0.005, "inverted": True},
}

# ── Metric Key Mapping ──────────────────────────────────────────────────────
FLOW_MAP = {
    "klaviyo_sales_main.raw.flow_revenue": "revenue",
    "klaviyo_sales_main.raw.flow_orders": "orders",
    "klaviyo_sales_main.raw.flow_send": "sends",
    "klaviyo_sales_main.raw.flow_unique_open": "unique_opens",
    "klaviyo_sales_main.raw.flow_unique_click_excl_bot": "unique_clicks",
    "klaviyo_sales_main.computed.flow_unique_open_rate": "open_rate",
    "klaviyo_sales_main.computed.flow_unique_click_rate_excl_bot": "click_rate",
    "klaviyo_sales_main.computed.flow_placed_order_rate": "order_rate",
    "klaviyo_sales_main.computed.flow_revenue_per_subscriber": "rev_per_recip",
    "klaviyo_sales_main.computed.flow_bounce_rate": "bounce_rate",
    "klaviyo_sales_main.computed.flow_unsubscribe_rate": "unsub_rate",
}
CAMP_MAP = {
    "klaviyo_sales_main.raw.campaign_revenue": "revenue",
    "klaviyo_sales_main.raw.campaign_orders": "orders",
    "klaviyo_sales_main.raw.campaign_send": "sends",
    "klaviyo_sales_main.raw.campaign_unique_open": "unique_opens",
    "klaviyo_sales_main.raw.campaign_unique_click_excl_bot": "unique_clicks",
    "klaviyo_sales_main.computed.campaign_unique_open_rate": "open_rate",
    "klaviyo_sales_main.computed.campaign_unique_click_rate_excl_bot": "click_rate",
    "klaviyo_sales_main.computed.campaign_placed_order_rate": "order_rate",
    "klaviyo_sales_main.computed.campaign_revenue_per_subscriber": "rev_per_recip",
    "klaviyo_sales_main.computed.campaign_bounce_rate": "bounce_rate",
    "klaviyo_sales_main.computed.campaign_unsubscribe_rate": "unsub_rate",
}


# ── Helpers ──────────────────────────────────────────────────────────────────
def load(name):
    with open(os.path.join(DATA_DIR, name), encoding="utf-8") as f:
        return json.load(f)


def detect_store(name):
    if "onzenna" in name.lower():
        return "Onzenna"
    return "ZeZeBaeBae"


def classify_flow(name):
    lower = name.lower()
    for kw, cat in FLOW_CATEGORY_RULES:
        if kw in lower:
            return cat
    return "Other"


def classify_campaign(name, subject=""):
    text = f"{name} {subject}".lower()
    for kw, cat in CAMPAIGN_CATEGORY_RULES:
        if kw in text:
            return cat
    return "Other"


def remap(row, mapping):
    out = {}
    for polar_key, short_key in mapping.items():
        out[short_key] = row.get(polar_key, 0) or 0
    return out


def safe_div(num, den):
    return num / den if den else 0


def month_label_and_multiplier(month_str, last_data_date):
    """Return (label, multiplier) for a month.

    If the month is partial (last_data_date falls inside it), label shows
    '2026-02 (21/28d)' and multiplier = total_days / elapsed_days.
    Full months get plain label and multiplier = 1.0.
    """
    year, mon = int(month_str[:4]), int(month_str[5:7])
    total_days = calendar.monthrange(year, mon)[1]

    if last_data_date.year == year and last_data_date.month == mon:
        elapsed = last_data_date.day
        label = f"{month_str} ({elapsed}/{total_days}d)"
        mult = total_days / elapsed if elapsed > 0 else 1.0
        return label, mult, True
    return month_str, 1.0, False


def detect_data_cutoff(flows, camps):
    """Find the latest date in the dataset to determine the partial month cutoff."""
    last = "2024-01"
    for fd in flows.values():
        for m in fd["months"]:
            if m > last:
                last = m
    for cd in camps.values():
        for m in cd["months"]:
            if m > last:
                last = m
    # The last month in the data is likely partial; use (now - 1 day) as cutoff
    # since Polar syncs with a ~1 day lag
    yesterday = date.today() - __import__("datetime").timedelta(days=1)
    last_year, last_mon = int(last[:4]), int(last[5:7])
    if yesterday.year == last_year and yesterday.month == last_mon:
        return yesterday
    # If the last month in data is not the current month, assume it's complete
    return date(last_year, last_mon, calendar.monthrange(last_year, last_mon)[1])


def weighted_totals(items):
    """Compute aggregate totals with weighted-average rates from a list of item dicts."""
    t = defaultdict(float)
    for it in items:
        for k in ("sends", "unique_opens", "unique_clicks", "orders", "revenue"):
            t[k] += it.get(k, 0)
    t["open_rate"] = safe_div(t["unique_opens"], t["sends"])
    t["click_rate"] = safe_div(t["unique_clicks"], t["sends"])
    t["order_rate"] = safe_div(t["orders"], t["sends"])
    t["rev_per_recip"] = safe_div(t["revenue"], t["sends"])
    t["bounce_rate"] = 0  # not aggregatable from raw counts we have
    t["unsub_rate"] = 0
    return dict(t)


# ── Excel Writing Helpers ────────────────────────────────────────────────────
def bar(ws, row, text, fill=MBLUE, cols=14):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = THIN
    ws.cell(row=row, column=1, value=text).font = WF
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    return row + 1


def hdr(ws, row, labels, start=1):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=lbl)
        c.font = HF
        c.fill = LGRAY
        c.border = THIN
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    return row + 1


def nc(ws, row, col, val, fmt=DOL, bold=False):
    c = ws.cell(row=row, column=col, value=val)
    c.number_format = fmt
    c.border = THIN
    c.font = BF if bold else NF
    c.alignment = Alignment(horizontal="right")
    return c


def lc(ws, row, col, val, bold=False, fill=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = BF if bold else NF
    c.border = THIN
    if fill:
        c.fill = fill
    return c


def rate_cell(ws, row, col, val, metric_key, bold=False):
    c = nc(ws, row, col, val, PCT, bold)
    th = METRIC_THRESHOLDS.get(metric_key)
    if th and val is not None:
        if th["inverted"]:
            if val <= th["good"]:
                c.fill = LGREEN
            elif val >= th["bad"]:
                c.fill = LRED
        else:
            if val >= th["good"]:
                c.fill = LGREEN
            elif val <= th["bad"]:
                c.fill = LRED
    return c


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


SUMMARY_HEADERS = [
    "Category", "Name", "Status", "Sends", "Open Rate",
    "Click Rate", "Orders", "Order Rate", "Revenue",
    "Rev/Recip", "Bounce%", "Unsub%",
]
CAMPAIGN_HEADERS = [
    "Category", "Campaign", "Subject", "Sends", "Open Rate",
    "Click Rate", "Orders", "Order Rate", "Revenue",
    "Rev/Recip", "Bounce%", "Unsub%",
]
SUMMARY_WIDTHS = {1: 22, 2: 50, 3: 10, 4: 10, 5: 11, 6: 11, 7: 9,
                  8: 11, 9: 14, 10: 12, 11: 10, 12: 10}
CAMPAIGN_WIDTHS = {1: 20, 2: 48, 3: 44, 4: 10, 5: 11, 6: 11, 7: 9,
                   8: 11, 9: 14, 10: 12, 11: 10, 12: 10}


def write_metric_row(ws, row, item, col_start=4, bold=False, fill=None):
    """Write metric columns: Sends, Open Rate, Click Rate, Orders, Order Rate,
    Revenue, Rev/Recip, Bounce%, Unsub%."""
    c = col_start
    nc(ws, row, c, item.get("sends", 0), INT, bold); c += 1
    rate_cell(ws, row, c, item.get("open_rate", 0), "open_rate", bold); c += 1
    rate_cell(ws, row, c, item.get("click_rate", 0), "click_rate", bold); c += 1
    nc(ws, row, c, item.get("orders", 0), INT, bold); c += 1
    rate_cell(ws, row, c, item.get("order_rate", 0), "order_rate", bold); c += 1
    nc(ws, row, c, item.get("revenue", 0), DOL, bold); c += 1
    nc(ws, row, c, item.get("rev_per_recip", 0), DOL, bold); c += 1
    rate_cell(ws, row, c, item.get("bounce_rate", 0), "bounce_rate", bold); c += 1
    rate_cell(ws, row, c, item.get("unsub_rate", 0), "unsub_rate", bold)
    if fill:
        for cc in range(1, 13):
            ws.cell(row=row, column=cc).fill = fill
            ws.cell(row=row, column=cc).border = THIN


# ── Data Processing ──────────────────────────────────────────────────────────
def process_flows():
    """Load flow JSON, aggregate by flow name, compute totals."""
    raw = load("kl1_flow_monthly.json")["tableData"]
    flows = {}  # {flow_name: {store, category, months: {month: metrics}, totals: metrics}}
    for row in raw:
        name = row.get("flow", "")
        if not name or name == "∅":
            continue
        month = row.get("date", "")[:7]  # "2025-01"
        m = remap(row, FLOW_MAP)
        m["month"] = month
        if name not in flows:
            flows[name] = {
                "store": detect_store(name),
                "category": classify_flow(name),
                "months": {},
                "first_send": None,
            }
        flows[name]["months"][month] = m
        sends = m.get("sends", 0)
        if sends > 0:
            if flows[name]["first_send"] is None or month < flows[name]["first_send"]:
                flows[name]["first_send"] = month

    # Compute totals per flow
    for name, fd in flows.items():
        all_months = list(fd["months"].values())
        fd["totals"] = weighted_totals(all_months)
    return flows


def process_campaigns():
    """Load campaign JSON, aggregate by campaign name, compute totals."""
    raw = load("kl2_campaign_monthly.json")["tableData"]
    camps = {}
    for row in raw:
        name = row.get("campaign", "")
        if not name or name == "∅":
            continue
        subject = row.get("subject", "")
        month = row.get("date", "")[:7]
        m = remap(row, CAMP_MAP)
        m["month"] = month
        if name not in camps:
            camps[name] = {
                "store": detect_store(name),
                "category": classify_campaign(name, subject),
                "subject": subject,
                "months": {},
                "first_send": None,
            }
        # Keep the most informative subject
        if subject and subject != "∅" and (not camps[name]["subject"] or camps[name]["subject"] == "∅"):
            camps[name]["subject"] = subject
        camps[name]["months"][month] = m
        sends = m.get("sends", 0)
        if sends > 0:
            if camps[name]["first_send"] is None or month < camps[name]["first_send"]:
                camps[name]["first_send"] = month

    for name, cd in camps.items():
        all_months = list(cd["months"].values())
        cd["totals"] = weighted_totals(all_months)
    return camps


# ── Sheet Builders ───────────────────────────────────────────────────────────
def build_flow_summary(wb, flows):
    ws = wb.create_sheet("Flow_Summary")
    ws.sheet_properties.tabColor = BLUE_TAB
    set_widths(ws, SUMMARY_WIDTHS)

    r = 1
    ws.cell(row=r, column=1, value="KLAVIYO FLOW PERFORMANCE").font = TITLE_F
    r += 1
    ws.cell(row=r, column=1, value=f"Data through {now.strftime('%b %d, %Y')}").font = SUB_F
    r += 2

    for store in ["ZeZeBaeBae", "Onzenna"]:
        store_flows = {n: f for n, f in flows.items() if f["store"] == store}
        if not store_flows:
            continue
        r = bar(ws, r, f"{store} — Flow Performance by Category")

        r = hdr(ws, r, SUMMARY_HEADERS)

        grand_items = []
        for cat in FLOW_CATEGORIES:
            cat_flows = {n: f for n, f in store_flows.items() if f["category"] == cat}
            if not cat_flows:
                continue
            # Sort by revenue DESC
            sorted_flows = sorted(cat_flows.items(), key=lambda x: x[1]["totals"].get("revenue", 0), reverse=True)
            for name, fd in sorted_flows:
                latest_month = now.strftime("%Y-%m")
                status = "New" if fd["first_send"] == latest_month else "Existing"
                lc(ws, r, 1, cat)
                lc(ws, r, 2, name)
                lc(ws, r, 3, status)
                write_metric_row(ws, r, fd["totals"], col_start=4)
                grand_items.append(fd["totals"])
                r += 1

            # Category subtotal
            cat_total = weighted_totals([f["totals"] for f in cat_flows.values()])
            lc(ws, r, 1, f"  {cat} Subtotal", bold=True)
            lc(ws, r, 2, "", bold=True)
            lc(ws, r, 3, "", bold=True)
            write_metric_row(ws, r, cat_total, col_start=4, bold=True, fill=LGRAY)
            r += 1

        # Grand total
        if grand_items:
            grand = weighted_totals(grand_items)
            lc(ws, r, 1, f"{store} TOTAL", bold=True)
            lc(ws, r, 2, "", bold=True)
            lc(ws, r, 3, "", bold=True)
            write_metric_row(ws, r, grand, col_start=4, bold=True, fill=MBLUE)
            # White font for TOTAL row
            for cc in range(1, 13):
                ws.cell(row=r, column=cc).font = WF
            r += 2

    return ws


def build_campaign_summary(wb, camps):
    ws = wb.create_sheet("Campaign_Summary")
    ws.sheet_properties.tabColor = ORANGE_TAB
    set_widths(ws, CAMPAIGN_WIDTHS)

    r = 1
    ws.cell(row=r, column=1, value="KLAVIYO CAMPAIGN PERFORMANCE").font = TITLE_F
    r += 1
    ws.cell(row=r, column=1, value=f"Data through {now.strftime('%b %d, %Y')}").font = SUB_F
    r += 2

    for store in ["ZeZeBaeBae", "Onzenna"]:
        store_camps = {n: c for n, c in camps.items() if c["store"] == store}
        if not store_camps:
            continue
        r = bar(ws, r, f"{store} — Campaign Performance by Category")
        r = hdr(ws, r, CAMPAIGN_HEADERS)

        grand_items = []
        for cat in CAMPAIGN_CATEGORIES:
            cat_camps = {n: c for n, c in store_camps.items() if c["category"] == cat}
            if not cat_camps:
                continue
            sorted_camps = sorted(cat_camps.items(), key=lambda x: x[1]["totals"].get("revenue", 0), reverse=True)
            for name, cd in sorted_camps:
                latest_month = now.strftime("%Y-%m")
                status_label = ""  # For campaigns we use subject column instead of status
                lc(ws, r, 1, cat)
                lc(ws, r, 2, name)
                lc(ws, r, 3, cd.get("subject", ""))
                write_metric_row(ws, r, cd["totals"], col_start=4)
                grand_items.append(cd["totals"])
                r += 1

            cat_total = weighted_totals([c["totals"] for c in cat_camps.values()])
            lc(ws, r, 1, f"  {cat} Subtotal", bold=True)
            lc(ws, r, 2, "")
            lc(ws, r, 3, "")
            write_metric_row(ws, r, cat_total, col_start=4, bold=True, fill=LGRAY)
            r += 1

        if grand_items:
            grand = weighted_totals(grand_items)
            lc(ws, r, 1, f"{store} TOTAL", bold=True)
            lc(ws, r, 2, "")
            lc(ws, r, 3, "")
            write_metric_row(ws, r, grand, col_start=4, bold=True, fill=MBLUE)
            for cc in range(1, 13):
                ws.cell(row=r, column=cc).font = WF
            r += 2

    # Section C: Top 20 by Revenue
    r = bar(ws, r, "Top 20 Campaigns by Revenue (All Stores)")
    r = hdr(ws, r, CAMPAIGN_HEADERS)
    all_sorted = sorted(camps.items(), key=lambda x: x[1]["totals"].get("revenue", 0), reverse=True)[:20]
    for name, cd in all_sorted:
        lc(ws, r, 1, cd["category"])
        lc(ws, r, 2, name)
        lc(ws, r, 3, cd.get("subject", ""))
        write_metric_row(ws, r, cd["totals"], col_start=4)
        r += 1
    r += 1

    # Section D: Bottom by Open Rate (min 100 sends)
    r = bar(ws, r, "Low Open Rate Campaigns (100+ sends, sorted ascending)")
    r = hdr(ws, r, CAMPAIGN_HEADERS)
    qualified = [(n, c) for n, c in camps.items() if c["totals"].get("sends", 0) >= 100]
    bottom = sorted(qualified, key=lambda x: x[1]["totals"].get("open_rate", 0))[:15]
    for name, cd in bottom:
        lc(ws, r, 1, cd["category"])
        lc(ws, r, 2, name)
        lc(ws, r, 3, cd.get("subject", ""))
        write_metric_row(ws, r, cd["totals"], col_start=4)
        r += 1

    return ws


def build_monthly_trends(wb, flows, camps):
    ws = wb.create_sheet("Monthly_Trends")
    ws.sheet_properties.tabColor = GREEN_TAB
    set_widths(ws, {1: 20, 2: 14, 3: 14, 4: 14, 5: 10, 6: 10, 7: 10,
                    8: 16, 9: 12})

    r = 1
    ws.cell(row=r, column=1, value="MONTHLY EMAIL TRENDS").font = TITLE_F
    r += 2

    # Detect data cutoff for partial month
    cutoff = detect_data_cutoff(flows, camps)

    # Collect all months
    all_months = set()
    for fd in flows.values():
        all_months.update(fd["months"].keys())
    for cd in camps.values():
        all_months.update(cd["months"].keys())
    months_sorted = sorted(all_months)

    # Pre-compute labels & multipliers
    month_info = {}
    for m in months_sorted:
        label, mult, partial = month_label_and_multiplier(m, cutoff)
        month_info[m] = {"label": label, "mult": mult, "partial": partial}

    # Section A: Flow Revenue by Month
    r = bar(ws, r, "A. Flow Revenue by Month", cols=9)
    headers_a = ["Month", "ZeZe Revenue", "Onzenna Revenue", "Total Revenue",
                 "ZeZe Orders", "Onzenna Orders", "Total Orders",
                 "Full Mo. Est. Rev", "Full Mo. Est. Ord"]
    r = hdr(ws, r, headers_a, start=1)
    for month in months_sorted:
        mi = month_info[month]
        zeze_rev = zeze_ord = onz_rev = onz_ord = 0
        for fd in flows.values():
            m = fd["months"].get(month, {})
            if fd["store"] == "ZeZeBaeBae":
                zeze_rev += m.get("revenue", 0)
                zeze_ord += m.get("orders", 0)
            else:
                onz_rev += m.get("revenue", 0)
                onz_ord += m.get("orders", 0)
        total_rev = zeze_rev + onz_rev
        total_ord = zeze_ord + onz_ord
        lc(ws, r, 1, mi["label"])
        nc(ws, r, 2, zeze_rev, DOL)
        nc(ws, r, 3, onz_rev, DOL)
        nc(ws, r, 4, total_rev, DOL, bold=True)
        nc(ws, r, 5, zeze_ord, INT)
        nc(ws, r, 6, onz_ord, INT)
        nc(ws, r, 7, total_ord, INT, bold=True)
        if mi["partial"]:
            c8 = nc(ws, r, 8, total_rev * mi["mult"], DOL)
            c8.font = EST_F
            c9 = nc(ws, r, 9, round(total_ord * mi["mult"]), INT)
            c9.font = EST_F
        else:
            nc(ws, r, 8, total_rev, DOL)
            nc(ws, r, 9, total_ord, INT)
        r += 1
    r += 1

    # Section B: Campaign Revenue by Month
    r = bar(ws, r, "B. Campaign Revenue by Month", cols=9)
    r = hdr(ws, r, headers_a, start=1)
    for month in months_sorted:
        mi = month_info[month]
        zeze_rev = zeze_ord = onz_rev = onz_ord = 0
        for cd in camps.values():
            m = cd["months"].get(month, {})
            if cd["store"] == "ZeZeBaeBae":
                zeze_rev += m.get("revenue", 0)
                zeze_ord += m.get("orders", 0)
            else:
                onz_rev += m.get("revenue", 0)
                onz_ord += m.get("orders", 0)
        total_rev = zeze_rev + onz_rev
        total_ord = zeze_ord + onz_ord
        lc(ws, r, 1, mi["label"])
        nc(ws, r, 2, zeze_rev, DOL)
        nc(ws, r, 3, onz_rev, DOL)
        nc(ws, r, 4, total_rev, DOL, bold=True)
        nc(ws, r, 5, zeze_ord, INT)
        nc(ws, r, 6, onz_ord, INT)
        nc(ws, r, 7, total_ord, INT, bold=True)
        if mi["partial"]:
            c8 = nc(ws, r, 8, total_rev * mi["mult"], DOL)
            c8.font = EST_F
            c9 = nc(ws, r, 9, round(total_ord * mi["mult"]), INT)
            c9.font = EST_F
        else:
            nc(ws, r, 8, total_rev, DOL)
            nc(ws, r, 9, total_ord, INT)
        r += 1
    r += 1

    # Section C: New vs Existing Monthly Count
    r = bar(ws, r, "C. New vs Existing Items by Month", cols=6)
    r = hdr(ws, r, ["Month", "New Flows", "Existing Flows", "New Campaigns", "Existing Campaigns", "Total New"], start=1)
    for month in months_sorted:
        mi = month_info[month]
        new_flows = sum(1 for fd in flows.values() if fd["first_send"] == month)
        existing_flows = sum(1 for fd in flows.values()
                            if fd["first_send"] and fd["first_send"] < month
                            and month in fd["months"])
        new_camps = sum(1 for cd in camps.values() if cd["first_send"] == month)
        existing_camps = sum(1 for cd in camps.values()
                            if cd["first_send"] and cd["first_send"] < month
                            and month in cd["months"])
        lc(ws, r, 1, mi["label"])
        nc(ws, r, 2, new_flows, INT)
        nc(ws, r, 3, existing_flows, INT)
        nc(ws, r, 4, new_camps, INT)
        nc(ws, r, 5, existing_camps, INT)
        nc(ws, r, 6, new_flows + new_camps, INT, bold=True)
        r += 1
    r += 1

    # Section H: New Items Log
    r = bar(ws, r, "D. New Items Log (chronological)", cols=5)
    r = hdr(ws, r, ["First Send Month", "Type", "Store", "Name", "Category"], start=1)
    new_items = []
    for name, fd in flows.items():
        if fd["first_send"]:
            new_items.append((fd["first_send"], "Flow", fd["store"], name, fd["category"]))
    for name, cd in camps.items():
        if cd["first_send"]:
            new_items.append((cd["first_send"], "Campaign", cd["store"], name, cd["category"]))
    new_items.sort(key=lambda x: x[0], reverse=True)
    for month, typ, store, name, cat in new_items:
        lc(ws, r, 1, month)
        lc(ws, r, 2, typ)
        lc(ws, r, 3, store)
        lc(ws, r, 4, name)
        lc(ws, r, 5, cat)
        r += 1

    return ws


def build_category_analysis(wb, flows, camps):
    ws = wb.create_sheet("Category_Analysis")
    ws.sheet_properties.tabColor = PURPLE_TAB
    set_widths(ws, {1: 26, 2: 8, 3: 10, 4: 11, 5: 11, 6: 9, 7: 11, 8: 14, 9: 12})

    r = 1
    ws.cell(row=r, column=1, value="CATEGORY ANALYSIS").font = TITLE_F
    r += 2

    cat_headers = ["Category", "# Items", "Sends", "Open Rate", "Click Rate",
                   "Orders", "Order Rate", "Revenue", "Rev/Recip"]

    # Section A: Flow Categories
    r = bar(ws, r, "A. Flow Category Performance (All Time)", cols=9)
    r = hdr(ws, r, cat_headers, start=1)
    all_flow_items = []
    for cat in FLOW_CATEGORIES:
        cat_flows = [f for f in flows.values() if f["category"] == cat]
        if not cat_flows:
            continue
        totals_list = [f["totals"] for f in cat_flows]
        agg = weighted_totals(totals_list)
        lc(ws, r, 1, cat)
        nc(ws, r, 2, len(cat_flows), INT)
        nc(ws, r, 3, agg["sends"], INT)
        rate_cell(ws, r, 4, agg["open_rate"], "open_rate")
        rate_cell(ws, r, 5, agg["click_rate"], "click_rate")
        nc(ws, r, 6, agg["orders"], INT)
        rate_cell(ws, r, 7, agg["order_rate"], "order_rate")
        nc(ws, r, 8, agg["revenue"], DOL)
        nc(ws, r, 9, agg["rev_per_recip"], DOL)
        all_flow_items.extend(totals_list)
        r += 1

    # Total
    if all_flow_items:
        grand = weighted_totals(all_flow_items)
        lc(ws, r, 1, "TOTAL", bold=True)
        nc(ws, r, 2, len(flows), INT, bold=True)
        nc(ws, r, 3, grand["sends"], INT, bold=True)
        rate_cell(ws, r, 4, grand["open_rate"], "open_rate", bold=True)
        rate_cell(ws, r, 5, grand["click_rate"], "click_rate", bold=True)
        nc(ws, r, 6, grand["orders"], INT, bold=True)
        rate_cell(ws, r, 7, grand["order_rate"], "order_rate", bold=True)
        nc(ws, r, 8, grand["revenue"], DOL, bold=True)
        nc(ws, r, 9, grand["rev_per_recip"], DOL, bold=True)
        for cc in range(1, 10):
            ws.cell(row=r, column=cc).fill = MBLUE
            ws.cell(row=r, column=cc).font = WF
        r += 2

    # Section B: Campaign Categories
    r = bar(ws, r, "B. Campaign Category Performance (All Time)", cols=9)
    r = hdr(ws, r, cat_headers, start=1)
    all_camp_items = []
    for cat in CAMPAIGN_CATEGORIES:
        cat_camps = [c for c in camps.values() if c["category"] == cat]
        if not cat_camps:
            continue
        totals_list = [c["totals"] for c in cat_camps]
        agg = weighted_totals(totals_list)
        lc(ws, r, 1, cat)
        nc(ws, r, 2, len(cat_camps), INT)
        nc(ws, r, 3, agg["sends"], INT)
        rate_cell(ws, r, 4, agg["open_rate"], "open_rate")
        rate_cell(ws, r, 5, agg["click_rate"], "click_rate")
        nc(ws, r, 6, agg["orders"], INT)
        rate_cell(ws, r, 7, agg["order_rate"], "order_rate")
        nc(ws, r, 8, agg["revenue"], DOL)
        nc(ws, r, 9, agg["rev_per_recip"], DOL)
        all_camp_items.extend(totals_list)
        r += 1

    if all_camp_items:
        grand = weighted_totals(all_camp_items)
        lc(ws, r, 1, "TOTAL", bold=True)
        nc(ws, r, 2, len(camps), INT, bold=True)
        nc(ws, r, 3, grand["sends"], INT, bold=True)
        rate_cell(ws, r, 4, grand["open_rate"], "open_rate", bold=True)
        rate_cell(ws, r, 5, grand["click_rate"], "click_rate", bold=True)
        nc(ws, r, 6, grand["orders"], INT, bold=True)
        rate_cell(ws, r, 7, grand["order_rate"], "order_rate", bold=True)
        nc(ws, r, 8, grand["revenue"], DOL, bold=True)
        nc(ws, r, 9, grand["rev_per_recip"], DOL, bold=True)
        for cc in range(1, 10):
            ws.cell(row=r, column=cc).fill = MBLUE
            ws.cell(row=r, column=cc).font = WF
        r += 2

    return ws


def build_data_notes(wb):
    ws = wb.create_sheet("Data_Notes")
    ws.sheet_properties.tabColor = GRAY_TAB
    set_widths(ws, {1: 80})

    notes = [
        "DATA NOTES & METHODOLOGY",
        "",
        "═══ Data Source ═══",
        "Klaviyo data via Polar Analytics MCP connector",
        f"Report generated: {now.strftime('%Y-%m-%d %H:%M')}",
        "",
        "═══ Store Mapping ═══",
        "ZeZeBaeBae: Flows/Campaigns without [Onzenna] prefix",
        "Onzenna: Flows/Campaigns with [Onzenna] prefix",
        "",
        "═══ Flow Categorization Rules ═══",
        "Welcome/Onboarding: welcome, pop-up",
        "Cart/Browse Recovery: abandoned cart, browse abandon",
        "Post-Purchase/Retention: first purchase, bounce back, replenishment, reminder",
        "Win-back: winback",
        "Back in Stock: back in stock",
        "",
        "═══ Campaign Categorization Rules ═══",
        "Promotional/Sales: bfcm, black friday, sale, deal, prime day, free, bundle, etc.",
        "Product Launch: launch, new arrival, introducing, pre-order, etc.",
        "Content/Educational: guide, tip, how to, toddler acne, screen time, etc.",
        "Seasonal/Holiday: holiday, christmas, easter, valentine, etc.",
        "Back in Stock: back in stock, restock",
        "",
        "═══ New vs Existing Logic ═══",
        "An item is 'New' for a given month if it first sent email in that month.",
        "From the following month onward, it becomes 'Existing'.",
        "",
        "═══ Rate Calculation ═══",
        "Subtotal/Total rates use weighted averages: SUM(numerator) / SUM(sends)",
        "Click Rate uses bot-excluded counts (flow_unique_click_excl_bot)",
        "Bounce Rate and Unsub Rate in subtotals are set to 0 (not aggregatable from raw data)",
        "",
        "═══ Conditional Formatting ═══",
        "Open Rate: Green >= 50%, Red < 30%",
        "Click Rate: Green >= 3%, Red < 1%",
        "Order Rate: Green >= 2%, Red < 0.5%",
        "Bounce Rate: Green < 0.5%, Red >= 2%",
        "Unsub Rate: Green < 0.1%, Red >= 0.5%",
        "",
        "═══ Known Limitations ═══",
        "1. Campaign categories use keyword matching — some miscategorization possible",
        "2. Data freshness depends on Polar connector sync schedule",
        "3. Revenue attribution follows Klaviyo's default attribution window",
    ]
    for i, line in enumerate(notes, 1):
        c = ws.cell(row=i, column=1, value=line)
        if line.startswith("═"):
            c.font = BF
        elif i == 1:
            c.font = TITLE_F
    return ws


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    print("Loading Klaviyo data from JSON...")

    flows = process_flows()
    camps = process_campaigns()

    print(f"  Flows: {len(flows)} ({sum(1 for f in flows.values() if f['store']=='ZeZeBaeBae')} ZeZe + "
          f"{sum(1 for f in flows.values() if f['store']=='Onzenna')} Onzenna)")
    print(f"  Campaigns: {len(camps)} ({sum(1 for c in camps.values() if c['store']=='ZeZeBaeBae')} ZeZe + "
          f"{sum(1 for c in camps.values() if c['store']=='Onzenna')} Onzenna)")

    # Show category breakdown
    print("\nFlow categories:")
    for cat in FLOW_CATEGORIES:
        cnt = sum(1 for f in flows.values() if f["category"] == cat)
        if cnt:
            print(f"  {cat}: {cnt}")

    print("\nCampaign categories:")
    for cat in CAMPAIGN_CATEGORIES:
        cnt = sum(1 for c in camps.values() if c["category"] == cat)
        if cnt:
            print(f"  {cat}: {cnt}")

    print("\nBuilding Excel dashboard...")
    wb = Workbook()
    wb.remove(wb.active)

    build_flow_summary(wb, flows)
    build_campaign_summary(wb, camps)
    build_monthly_trends(wb, flows, camps)
    build_category_analysis(wb, flows, camps)
    build_data_notes(wb)

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")
    print(f"Tabs: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
