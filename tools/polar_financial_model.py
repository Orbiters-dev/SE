"""Polar Financial Model v2 — formula-driven, hierarchy-based financial model.

Layout: Col A=section label, B=Channel/Total, C=Brand, D=Product, E+=monthly data.
All parent totals use =SUM(explicit child refs). Ratios use =IF(den=0,0,num/den).
"""
import json, os, re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────────────────────
DIR = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(DIR, "..")
DATA = os.path.join(ROOT, ".tmp", "polar_data")
OUT = os.path.join(ROOT, "Data Storage", "Polar data", "Polar_Financial_Model.xlsx")

# ── Constants ────────────────────────────────────────────────────────────────

# Bill Pay influencer payments not captured by PayPal API.
# Add new entries here as they are identified (amount must be positive).
MANUAL_INF_PAYMENTS = [
    {"payer_name": "Emily Krausz",              "date": "2026-01-05", "amount": -4500.00},
    {"payer_name": "Emily Krausz",              "date": "2025-10-14", "amount": -1000.00},
    {"payer_name": "Kathlyn Marie Sanga Flores","date": "2025-12-05", "amount":  -275.00},
    {"payer_name": "Ehwa Lindsay",              "date": "2025-11-07", "amount":  -300.00},
    {"payer_name": "Ehwa Lindsay",              "date": "2025-07-22", "amount":  -100.00},
    {"payer_name": "Ehwa Lindsay",              "date": "2025-07-17", "amount":  -100.00},
    {"payer_name": "Jessica Lim",               "date": "2025-01-21", "amount":  -500.00},
]

BRANDS = [
    "Grosmimi", "Alpremio", "Comme Moi", "BabyRabbit", "Naeiae",
    "Bamboobebe", "Hattung", "CHA&MOM", "Beemymagic", "Nature Love Mere",
]
MODEL_CH = ["Onzenna", "Amazon", "TargetPlus", "TikTokShop", "B2B"]
AD_PLATS = ["Amazon Ads", "Facebook Ads", "Google Ads", "TikTok Ads"]
CH_MAP = {
    "D2C": "Onzenna", "Target+": "TargetPlus", "B2B": "B2B",
    "TikTok": "TikTokShop", "FBM": "Amazon",
    "Amazon - Grosmimi USA": "Amazon", "Amazon - Fleeters": "Amazon",
    "Amazon - Orbitool": "Amazon", "MCF/Removal Order": "Amazon", "PR": "PR_Sample",
}
PROD_CATS = [
    "PPSU Straw Cup", "Flip Top Cup", "Stainless Cup", "Tumbler",
    "Baby Bottle", "Bundles", "Replacement Parts", "Accessories",
    "Skincare", "Food & Snacks", "Baby Carrier", "Apparel",
    "Tableware", "Educational Toys", "Baby Care", "Bamboo Products",
    "Wholesale", "Other",
]

# ── Product classification ───────────────────────────────────────────────────
PROD_RULES = [
    ("PPSU Straw Cup", "Grosmimi", "PPSU Straw Cup"),
    ("Flip Top", "Grosmimi", "Flip Top Cup"),
    ("KNOTTED", "Grosmimi", "Flip Top Cup"),
    ("Stainless Steel Straw", "Grosmimi", "Stainless Cup"),
    ("Tumbler", "Grosmimi", "Tumbler"),
    ("Baby Bottle", "Grosmimi", "Baby Bottle"),
    ("Easy Baby Bottle", "Grosmimi", "Baby Bottle"),
    ("2-pack", "Grosmimi", "Bundles"),
    ("Multi Accessory", "Grosmimi", "Bundles"),
    ("Replacement", "Grosmimi", "Replacement Parts"),
    ("Replacements", "Grosmimi", "Replacement Parts"),
    ("One Touch Cap", "Grosmimi", "Replacement Parts"),
    ("Weighted Kit", "Grosmimi", "Replacement Parts"),
    ("Strap", "Grosmimi", "Accessories"),
    ("Brush", "Grosmimi", "Accessories"),
    ("Teether", "Grosmimi", "Accessories"),
    ("Silicone Plate", "Grosmimi", "Accessories"),
    ("CHA&MOM", "CHA&MOM", "Skincare"),
    ("Naeiae", "Naeiae", "Food & Snacks"),
    ("Alpremio", "Alpremio", "Baby Carrier"),
    ("BabyRabbit", "BabyRabbit", "Apparel"),
    ("Bamboobebe", "Bamboobebe", "Bamboo Products"),
    ("Beemeal", "Beemymagic", "Tableware"),
    ("Heart Tray", "Beemymagic", "Tableware"),
    ("Comme Moi", "Comme Moi", "Educational Toys"),
    ("Nature Love Mere", "Nature Love Mere", "Baby Care"),
    ("Hattung", "Hattung", "Other"),
    ("B2B Wholesale", "Other", "Wholesale"),
]

# ── B2B Wholesale override: Feb 2026 one-time order mapped to Grosmimi products ──
WHOLESALE_FEB2026 = [
    ("Grosmimi Beige 10oz",    120, 13.5),   # $1,620
    ("Grosmimi Beige 6oz",      90, 12.7),   # $1,143
    ("Grosmimi Charcoal 6oz",   90, 12.7),   # $1,143
    ("Grosmimi Straw (2110)",  240,  7.4),   # $1,776
    ("Grosmimi White 10oz",    120, 13.5),   # $1,620
]  # Total = $7,302

def classify_product(name):
    for kw, brand, cat in PROD_RULES:
        if kw.lower() in name.lower():
            return brand, cat
    return "Other", "Other"

# ── Campaign parsing ─────────────────────────────────────────────────────────
AMZ_BR = [("cha&mom", "CHA&MOM"), ("naeiae", "Naeiae"), ("alpremio", "Alpremio"), ("comme", "Comme Moi")]
FB_BR = [
    ("alpremio", "Alpremio"), ("naeiae", "Naeiae"),
    ("cha&mom", "CHA&MOM"), ("love&care", "CHA&MOM"),
    ("| cm |", "CHA&MOM"), ("_cm_", "CHA&MOM"),
    ("| gm |", "Grosmimi"), ("_gm_", "Grosmimi"),
    ("grosmimi", "Grosmimi"), ("dental mom", "Grosmimi"), ("dentalmom", "Grosmimi"),
    ("livfuselli", "Grosmimi"), ("tumbler", "Grosmimi"), ("stainless", "Grosmimi"),
    ("sls", "Grosmimi"), ("laurence", "Grosmimi"), ("lauren", "Grosmimi"),
    ("asc campaign", "Grosmimi"),
]
FB_LAND_AMZ = ["amz_traffic", "amz | traffic", "amz| traffic",
               "asc | amz | traffic", "asc i amz i traffic", "tof | amz |"]
FB_LAND_TARGET = ["target | traffic", "target |traffic"]
AD_PROD = [
    ("ppsu", "PPSU Straw Cup"), ("flip top", "Flip Top Cup"), ("fliptop", "Flip Top Cup"),
    ("knotted", "Flip Top Cup"), ("stainless", "Stainless Cup"),
    ("tumbler", "Tumbler"), ("stage1", "Replacement Parts"), ("stage2", "PPSU Straw Cup"),
    ("replacements", "Replacement Parts"), ("wash", "Skincare"), ("lotion", "Skincare"),
    ("cream", "Skincare"), ("naeiae", "Food & Snacks"), ("alpremio", "Baby Carrier"),
]

def ad_brand(camp, plat):
    c = camp.lower()
    if plat == "Amazon Ads":
        for k, b in AMZ_BR:
            if k in c: return b
        return "Grosmimi"
    if plat == "Facebook Ads":
        for k, b in FB_BR:
            if k in c: return b
        return "Other"
    if plat == "Google Ads": return "Grosmimi"
    return "Other"

def ad_landing(camp, plat):
    """Determine which sales channel an ad lands on."""
    c = camp.lower()
    if plat == "Amazon Ads": return "Amazon"
    if plat == "Google Ads": return "Onzenna"
    if plat == "Facebook Ads":
        for pat in FB_LAND_AMZ:
            if pat in c: return "Amazon"
        for pat in FB_LAND_TARGET:
            if pat in c: return "TargetPlus"
        if "shopify" in c: return "Onzenna"
        return "Onzenna"  # default: D2C is primary channel
    if plat == "TikTok Ads":
        if ("amz" in c or "amazon" in c) and "traffic" in c: return "Amazon"
        return "Onzenna"
    return "Onzenna"

def ad_prod(camp):
    c = camp.lower()
    for k, cat in AD_PROD:
        if k in c: return cat
    return "General"

def get_campaign_type(camp, plat):
    """Classify campaign type.
    Amazon: SP/SB/SD.  Google: CVR.  Facebook/TikTok: CVR/Traffic/Other.
    """
    if plat == "Google Ads":
        return "CVR"
    if plat == "Amazon Ads":
        c = camp.lower()
        if c.startswith("sb") or "sb_" in c or "sb-" in c:
            return "SB"
        if c.startswith("sd") or "sd_" in c:
            return "SD"
        return "SP"
    c = camp.lower()
    if "traffic" in c:
        return "Traffic"
    if any(k in c for k in ("cvr", "conversion", "purchase", "sales", "shopping")):
        return "CVR"
    return "Other"

# ── Styles ───────────────────────────────────────────────────────────────────
DBLUE = PatternFill("solid", fgColor="002060")
MBLUE = PatternFill("solid", fgColor="4472C4")
LGRAY = PatternFill("solid", fgColor="D9E2F3")
LLGRAY = PatternFill("solid", fgColor="F2F2F2")
LGREEN = PatternFill("solid", fgColor="E2EFDA")
LYELLOW = PatternFill("solid", fgColor="FFF2CC")
LRED = PatternFill("solid", fgColor="FCE4D6")
SKYBLUE = PatternFill("solid", fgColor="DAEEF3")
PINK = PatternFill("solid", fgColor="F2DCDB")
WF = Font(bold=True, color="FFFFFF", size=10)
WF9 = Font(bold=True, color="FFFFFF", size=9)
HF = Font(bold=True, size=9)
NF = Font(size=9)
BF = Font(bold=True, size=9)
BF_G = Font(bold=True, size=9, color="006100")
BF_R = Font(bold=True, size=9, color="9C0006")
NUM = '#,##0;-#,##0;"-"'
DEC = '#,##0.00;-#,##0.00;"-"'
DOL = '$#,##0;-$#,##0;"-"'
DOL2 = '$#,##0.00;-$#,##0.00;"-"'
PCT = '0.0%;-0.0%;"-"'
THIN = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
TOTAL_BORDER = Border(top=Side("thin"), bottom=Side("thin"))

def load(name):
    with open(os.path.join(DATA, name), encoding="utf-8") as f:
        return json.load(f)

# ══════════════════════════════════════════════════════════════════════════════
# HIERARCHY ENGINE
# ══════════════════════════════════════════════════════════════════════════════

class Node:
    """Tree node: label, level, children list, data dict {month: {metric: value}}."""
    __slots__ = ("label", "level", "children", "data")
    def __init__(self, label, level):
        self.label = label
        self.level = level
        self.children = []
        self.data = defaultdict(lambda: defaultdict(float))


def build_generic_tree(raw, depth, ytd_months, sort_key):
    """Build tree from raw dict {tuple_key: {month: {metric: val}}}.

    depth=3 means keys are (ch, brand, prod) -> levels 1,2,3 under a TOTAL root (level 0).
    depth=2 means keys are (ch, brand) -> levels 1,2.
    depth=1 means keys are (cat,) -> level 1 (flat).
    """
    root = Node("TOTAL", 0)
    # Index nodes by partial key
    nodes_by_key = {}  # tuple -> Node

    for key_tuple, month_data in raw.items():
        parent = root
        for i in range(depth):
            partial = key_tuple[:i + 1]
            if partial not in nodes_by_key:
                nd = Node(key_tuple[i], i + 1)
                nodes_by_key[partial] = nd
                parent.children.append(nd)
            node = nodes_by_key[partial]
            parent = node

        # Leaf: store data
        leaf = nodes_by_key[key_tuple[:depth]]
        for mo, metrics in month_data.items():
            for mk, mv in metrics.items():
                leaf.data[mo][mk] += mv

    sort_tree(root, ytd_months, sort_key)
    return root


def _ytd_val(node, ytd_months, sort_key):
    """Compute the YTD value for sorting — sum of sort_key across ytd_months, recursively."""
    if node.children:
        return sum(_ytd_val(c, ytd_months, sort_key) for c in node.children)
    return sum(node.data.get(m, {}).get(sort_key, 0) for m in ytd_months)


def sort_tree(node, ytd_months, sort_key):
    """Sort children by YTD desc, 'Other' always last. Recurse."""
    for c in node.children:
        sort_tree(c, ytd_months, sort_key)
    def sk(child):
        is_other = 1 if child.label == "Other" else 0
        return (is_other, -_ytd_val(child, ytd_months, sort_key))
    node.children.sort(key=sk)


def flatten(node):
    """DFS pre-order: parent before children."""
    result = [node]
    for c in node.children:
        result.extend(flatten(c))
    return result


# ══════════════════════════════════════════════════════════════════════════════
# SECTION WRITING ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def _cell_ref(col_idx, row_idx):
    """Return absolute cell reference like E5."""
    return f"{get_column_letter(col_idx)}{row_idx}"


from datetime import date as _date
_today = _date.today()
PARTIAL_DAY = _today.day if _today.year == 2026 and _today.month == 2 else _today.day
FULL_DAYS = 28  # Feb 2026 has 28 days


def calc_total_cols(months, level_headers):
    """Compute total column count including data, % of Total, and YoY blocks."""
    col_start = len(level_headers) + 1
    has_partial = months[-1] == "2026-02" if months else False
    num_reg = len(months) - 1 if has_partial else len(months)
    extra = 2 if has_partial else 0  # feb_full + feb_actual
    data_width = num_reg + extra + 1  # months + feb cols + ytd
    # layout: labels | data | gap | pct | gap | yoy
    return (col_start - 1) + 3 * data_width + 2


def write_sections(ws, start_row, flat_nodes, months, ytd_months, configs, level_headers, total_cols):
    """Write metric sections with Feb extrapolation, % of Total, and YoY blocks.

    Returns (next_row, sec_map).
    """
    row = start_row
    sec_map = {}
    num_nodes = len(flat_nodes)
    col_start = len(level_headers) + 1

    # ── Detect partial month ──
    has_partial = months[-1] == "2026-02" if months else False
    if has_partial:
        reg_months = months[:-1]
        partial_m = months[-1]
    else:
        reg_months = list(months)
        partial_m = None
    num_reg = len(reg_months)

    # ── Column positions: DATA block ──
    extra = 2 if has_partial else 0
    feb_full_col = col_start + num_reg if has_partial else None
    feb_actual_col = col_start + num_reg + 1 if has_partial else None
    ytd_col = col_start + num_reg + extra
    data_width = num_reg + extra + 1  # total columns in one block

    month_to_col = {m: col_start + i for i, m in enumerate(reg_months)}
    if has_partial:
        month_to_col[partial_m] = feb_actual_col

    # YTD refs: Jan 2026 + Feb Full-month (not partial)
    ytd_ref_cols = []
    for m in ytd_months:
        if m == partial_m and has_partial:
            ytd_ref_cols.append(feb_full_col)
        elif m in month_to_col:
            ytd_ref_cols.append(month_to_col[m])

    # ── Column positions: % of Total block ──
    pct_gap_col = col_start + data_width
    pct_start = pct_gap_col + 1

    # ── Column positions: YoY block ──
    yoy_gap_col = pct_start + data_width
    yoy_start = yoy_gap_col + 1

    # Helper: map data-block offset to pct/yoy column
    def pct_col(data_col):
        return pct_start + (data_col - col_start)

    def yoy_col(data_col):
        return yoy_start + (data_col - col_start)

    # YoY: find prior-year column in data block
    def prior_year_col(m):
        parts = m.split("-")
        prior_m = f"{int(parts[0]) - 1}-{parts[1]}"
        return month_to_col.get(prior_m)

    # Prior-year YTD month columns (e.g. 2025-01, 2025-02 for YTD 2026)
    prior_ytd_cols = []
    for m in ytd_months:
        parts = m.split("-")
        prior_m = f"{int(parts[0]) - 1}-{parts[1]}"
        if prior_m in month_to_col:
            prior_ytd_cols.append(month_to_col[prior_m])

    # Pre-build children indices
    node_to_idx = {id(n): i for i, n in enumerate(flat_nodes)}
    children_indices = {}
    for i, n in enumerate(flat_nodes):
        children_indices[i] = [node_to_idx[id(c)] for c in n.children]

    for cfg in configs:
        sec_key = cfg["key"]
        title = cfg["title"]
        fmt = cfg["fmt"]
        sec_type = cfg["type"]

        # ── Section header bar ──
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = DBLUE
            cell.font = WF
        ws.cell(row=row, column=1, value=title)
        row += 1

        # ── Column headers ──
        hdrs = list(level_headers)
        for m in reg_months:
            hdrs.append(_month_label(m))
        if has_partial:
            hdrs.append(f"Feb 2026\n(Full-mo)")
            hdrs.append(f"Feb {PARTIAL_DAY}\n2026")
        hdrs.append("YTD 2026")
        # gap
        hdrs.append("")
        # pct headers
        for m in reg_months:
            hdrs.append(_month_label(m))
        if has_partial:
            hdrs.append(f"Feb 2026\n(Full-mo)")
            hdrs.append(f"Feb {PARTIAL_DAY}\n2026")
        hdrs.append("YTD 2026")
        # gap
        hdrs.append("")
        # yoy headers
        for m in reg_months:
            hdrs.append(_month_label(m))
        if has_partial:
            hdrs.append(f"Feb 2026\n(Full-mo)")
            hdrs.append(f"Feb {PARTIAL_DAY}\n2026")
        hdrs.append("YTD 2026")

        for ci, lbl in enumerate(hdrs):
            cell = ws.cell(row=row, column=1 + ci, value=lbl)
            cell.font = HF
            cell.fill = LGRAY
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Color pct/yoy header cells
        for dc in range(pct_start, pct_start + data_width):
            ws.cell(row=row, column=dc).fill = LGREEN
        for dc in range(yoy_start, yoy_start + data_width):
            ws.cell(row=row, column=dc).fill = LYELLOW

        # Label gap columns
        g1 = ws.cell(row=row, column=pct_gap_col, value="% of\nTotal")
        g1.font = Font(bold=True, size=8, color="FFFFFF")
        g1.fill = PatternFill("solid", fgColor="548235")
        g1.alignment = Alignment(horizontal="center", wrap_text=True)
        g2 = ws.cell(row=row, column=yoy_gap_col, value="YoY\nGrowth")
        g2.font = Font(bold=True, size=8, color="FFFFFF")
        g2.fill = PatternFill("solid", fgColor="BF8F00")
        g2.alignment = Alignment(horizontal="center", wrap_text=True)
        row += 1

        # ── Data rows ──
        first_data_row = row
        sec_map[sec_key] = first_data_row
        total_row = first_data_row  # TOTAL node is flat_nodes[0]

        for ni, node in enumerate(flat_nodes):
            node_row = first_data_row + ni
            label_col = max(2, 1 + node.level)
            cell = ws.cell(row=node_row, column=label_col, value=node.label)

            if node.level == 0:
                _style_row(ws, node_row, total_cols, fill=MBLUE,
                           font=Font(bold=True, color="FFFFFF", size=9), border=TOTAL_BORDER)
            elif node.level == 1:
                _style_row(ws, node_row, total_cols, fill=LLGRAY, font=BF)
            else:
                _style_row(ws, node_row, total_cols, fill=None, font=NF)

            rf = Font(bold=True, color="FFFFFF", size=9) if node.level == 0 else (BF if node.level == 1 else NF)
            cell.font = rf

            child_rows = [first_data_row + ci for ci in children_indices[ni]]

            # ──────── DATA BLOCK ────────
            if sec_type == "data":
                metric = cfg["metric"]
                for m in reg_months:
                    col = month_to_col[m]
                    if node.children:
                        refs = ",".join(_cell_ref(col, cr) for cr in child_rows)
                        c = ws.cell(row=node_row, column=col, value=f"=SUM({refs})")
                    else:
                        c = ws.cell(row=node_row, column=col, value=node.data.get(m, {}).get(metric, 0))
                    c.number_format = fmt; c.font = rf

                if has_partial:
                    col = feb_actual_col
                    if node.children:
                        refs = ",".join(_cell_ref(col, cr) for cr in child_rows)
                        c = ws.cell(row=node_row, column=col, value=f"=SUM({refs})")
                    else:
                        c = ws.cell(row=node_row, column=col, value=node.data.get(partial_m, {}).get(metric, 0))
                    c.number_format = fmt; c.font = rf
                    # Feb Full-month: extrapolate unless no_extrapolate flag set
                    feb_ref = _cell_ref(feb_actual_col, node_row)
                    if cfg.get("no_extrapolate"):
                        c = ws.cell(row=node_row, column=feb_full_col,
                                    value=f"={feb_ref}")
                    else:
                        c = ws.cell(row=node_row, column=feb_full_col,
                                    value=f"={feb_ref}*{FULL_DAYS}/{PARTIAL_DAY}")
                    c.number_format = fmt; c.font = rf

                # YTD
                if ytd_ref_cols:
                    refs = ",".join(_cell_ref(c_, node_row) for c_ in ytd_ref_cols)
                    c = ws.cell(row=node_row, column=ytd_col, value=f"=SUM({refs})")
                else:
                    c = ws.cell(row=node_row, column=ytd_col, value=0)
                c.number_format = fmt; c.font = rf

            elif sec_type == "ratio":
                num_start = sec_map[cfg["num_key"]]
                den_start = sec_map[cfg["den_key"]]
                for m in reg_months:
                    col = month_to_col[m]
                    nr = _cell_ref(col, num_start + ni)
                    dr = _cell_ref(col, den_start + ni)
                    c = ws.cell(row=node_row, column=col, value=f"=IF({dr}=0,0,{nr}/{dr})")
                    c.number_format = fmt; c.font = rf

                if has_partial:
                    for fcol in (feb_actual_col, feb_full_col):
                        nr = _cell_ref(fcol, num_start + ni)
                        dr = _cell_ref(fcol, den_start + ni)
                        c = ws.cell(row=node_row, column=fcol, value=f"=IF({dr}=0,0,{nr}/{dr})")
                        c.number_format = fmt; c.font = rf

                nr = _cell_ref(ytd_col, num_start + ni)
                dr = _cell_ref(ytd_col, den_start + ni)
                c = ws.cell(row=node_row, column=ytd_col, value=f"=IF({dr}=0,0,{nr}/{dr})")
                c.number_format = fmt; c.font = rf

            elif sec_type == "diff":
                a_start = sec_map[cfg["a_key"]]
                b_start = sec_map[cfg["b_key"]]
                for m in reg_months:
                    col = month_to_col[m]
                    ar = _cell_ref(col, a_start + ni)
                    br = _cell_ref(col, b_start + ni)
                    c = ws.cell(row=node_row, column=col, value=f"={ar}-{br}")
                    c.number_format = fmt; c.font = rf

                if has_partial:
                    for fcol in (feb_actual_col, feb_full_col):
                        ar = _cell_ref(fcol, a_start + ni)
                        br = _cell_ref(fcol, b_start + ni)
                        c = ws.cell(row=node_row, column=fcol, value=f"={ar}-{br}")
                        c.number_format = fmt; c.font = rf

                ar = _cell_ref(ytd_col, a_start + ni)
                br = _cell_ref(ytd_col, b_start + ni)
                c = ws.cell(row=node_row, column=ytd_col, value=f"={ar}-{br}")
                c.number_format = fmt; c.font = rf

            # ──────── % OF TOTAL BLOCK ────────
            # Each cell = data_cell / TOTAL_data_cell (same column)
            for di in range(data_width):
                dc = col_start + di
                pc = pct_col(dc)
                d_ref = _cell_ref(dc, node_row)
                t_ref = _cell_ref(dc, total_row)
                c = ws.cell(row=node_row, column=pc,
                            value=f"=IF({t_ref}=0,0,{d_ref}/{t_ref})")
                c.number_format = PCT; c.font = rf

            # ──────── YOY GROWTH BLOCK ────────
            # Regular months: compare to same month prior year
            for m in reg_months:
                dc = month_to_col[m]
                yc = yoy_col(dc)
                pcol = prior_year_col(m)
                if pcol is not None:
                    curr = _cell_ref(dc, node_row)
                    prev = _cell_ref(pcol, node_row)
                    c = ws.cell(row=node_row, column=yc,
                                value=f"=IF(ABS({prev})=0,0,({curr}-{prev})/ABS({prev}))")
                    c.number_format = PCT; c.font = rf

            if has_partial:
                # Feb Full-month & Feb actual YoY vs Feb prior year
                prior_feb = month_to_col.get("2025-02")
                if prior_feb is not None:
                    for fcol in (feb_full_col, feb_actual_col):
                        yc = yoy_col(fcol)
                        curr = _cell_ref(fcol, node_row)
                        prev = _cell_ref(prior_feb, node_row)
                        c = ws.cell(row=node_row, column=yc,
                                    value=f"=IF(ABS({prev})=0,0,({curr}-{prev})/ABS({prev}))")
                        c.number_format = PCT; c.font = rf

            # YTD YoY
            yc_ytd = yoy_col(ytd_col)
            if prior_ytd_cols:
                curr_ytd = _cell_ref(ytd_col, node_row)
                if sec_type == "ratio":
                    # For ratios: compute prior YTD ratio from underlying sections
                    nk = cfg.get("num_key"); dk = cfg.get("den_key")
                    if nk and dk:
                        ns = sec_map[nk]; ds = sec_map[dk]
                        n_refs = "+".join(_cell_ref(pc_, ns + ni) for pc_ in prior_ytd_cols)
                        d_refs = "+".join(_cell_ref(pc_, ds + ni) for pc_ in prior_ytd_cols)
                        prior_ratio = f"IF(({d_refs})=0,0,({n_refs})/({d_refs}))"
                        c = ws.cell(row=node_row, column=yc_ytd,
                                    value=f"=IF(ABS({prior_ratio})=0,0,({curr_ytd}-{prior_ratio})/ABS({prior_ratio}))")
                        c.number_format = PCT; c.font = rf
                else:
                    # data/diff: sum prior year's YTD months
                    if len(prior_ytd_cols) == 1:
                        prev_sum = _cell_ref(prior_ytd_cols[0], node_row)
                    else:
                        prev_sum = "+".join(_cell_ref(pc_, node_row) for pc_ in prior_ytd_cols)
                        prev_sum = f"({prev_sum})"
                    c = ws.cell(row=node_row, column=yc_ytd,
                                value=f"=IF(ABS({prev_sum})=0,0,({curr_ytd}-{prev_sum})/ABS({prev_sum}))")
                    c.number_format = PCT; c.font = rf

        row = first_data_row + num_nodes + 1  # blank row between sections

    return row, sec_map


def _month_label(m):
    """Convert '2024-01-01' or '2024-01' to 'Jan 2024'."""
    parts = m.split("-")
    yr = parts[0]
    mo = int(parts[1])
    names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
             "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{names[mo - 1]} {yr}"


def _style_row(ws, row, total_cols, fill=None, font=None, border=None):
    """Apply fill, font, and optional border to all cells in a row."""
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=row, column=c)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if border:
            cell.border = border


def _set_col_widths(ws, total_cols, months, level_headers):
    """Set column widths — label cols wider, data cols narrow, gap cols minimal."""
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    col_start = len(level_headers) + 1
    has_partial = months[-1] == "2026-02" if months else False
    num_reg = len(months) - 1 if has_partial else len(months)
    extra = 2 if has_partial else 0
    data_width = num_reg + extra + 1
    pct_gap = col_start + data_width
    yoy_gap = pct_gap + 1 + data_width
    for i in range(col_start, total_cols + 1):
        cl = get_column_letter(i)
        if i == pct_gap or i == yoy_gap:
            ws.column_dimensions[cl].width = 6  # narrow gap
        else:
            ws.column_dimensions[cl].width = 12


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING & PREPROCESSING
# ══════════════════════════════════════════════════════════════════════════════

def preprocess():
    """Load all JSON data files and return structured dicts."""
    q1 = load("q1_channel_brand_product.json")["tableData"]
    q2 = load("q2_shopify_brand.json")["tableData"]
    q3 = load("q3_amazon_brand.json")["tableData"]
    q5 = load("q5_amazon_ads_campaign.json")["tableData"]
    q6 = load("q6_facebook_ads_campaign.json")["tableData"]
    q7 = load("q7_google_ads_campaign.json")["tableData"]
    q8 = load("q8_tiktok_ads_campaign.json")["tableData"]

    # ── Optional: Meta campaign IDs for direct links ──
    meta_id_path = os.path.join(DATA, "q9_meta_campaign_ids.json")
    meta_camp_ids = {}
    meta_account_id = ""
    if os.path.exists(meta_id_path):
        with open(meta_id_path, encoding="utf-8") as f:
            q9 = json.load(f)
        meta_camp_ids = q9.get("campaign_map", {})
        meta_account_id = q9.get("account_id", "")

    # ── Optional: Influencer orders (Shopify PR tag) ──
    q10_path = os.path.join(DATA, "q10_influencer_orders.json")
    q11_path = os.path.join(DATA, "q11_paypal_transactions.json")
    inf_orders = []
    paypal_txns = []
    if os.path.exists(q10_path):
        with open(q10_path, encoding="utf-8") as f:
            inf_orders = json.load(f).get("orders", [])
    if os.path.exists(q11_path):
        with open(q11_path, encoding="utf-8") as f:
            paypal_txns = json.load(f).get("transactions", [])
    # Merge manually identified Bill Pay influencer payments
    paypal_txns = paypal_txns + MANUAL_INF_PAYMENTS

    # ── Revenue: (channel, brand, product) -> {month -> {gross, disc, orders, net}} ──
    rev_cbp = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    all_months = set()

    for r in q1:
        raw_ch = r.get("custom_5005", "") or ""
        raw_brand = r.get("custom_5036", "") or ""
        raw_prod = r.get("custom_5037", "") or ""
        dt = r.get("date", "")
        if not dt:
            continue
        # Normalize date to YYYY-MM
        mo = dt[:7]
        all_months.add(mo)

        ch = CH_MAP.get(raw_ch, "Other")
        _, prod_cat = classify_product(raw_prod)
        brand = raw_brand if raw_brand in BRANDS else "Other"

        gross = abs(r.get("blended_gross_sales", 0) or 0)
        disc = abs(r.get("blended_discounts", 0) or 0)
        orders = r.get("blended_total_orders", 0) or 0
        net = r.get("blended_total_sales", 0) or 0

        if ch == "PR_Sample":
            # PR: force net to 0, use channel "PR"
            rev_cbp[("PR", brand, prod_cat)][mo]["gross"] += gross
            rev_cbp[("PR", brand, prod_cat)][mo]["disc"] += disc
            rev_cbp[("PR", brand, prod_cat)][mo]["orders"] += orders
            rev_cbp[("PR", brand, prod_cat)][mo]["net"] += 0
        else:
            rev_cbp[(ch, brand, prod_cat)][mo]["gross"] += gross
            rev_cbp[(ch, brand, prod_cat)][mo]["disc"] += disc
            rev_cbp[(ch, brand, prod_cat)][mo]["orders"] += orders
            rev_cbp[(ch, brand, prod_cat)][mo]["net"] += net

    # ── B2B Wholesale override: remap Feb 2026 Wholesale to Grosmimi products ──
    ws_key = ("B2B", "Other", "Wholesale")
    if ws_key in rev_cbp and "2026-02" in rev_cbp[ws_key]:
        del rev_cbp[ws_key]["2026-02"]
        if not rev_cbp[ws_key]:
            del rev_cbp[ws_key]
        for prod_name, qty, price in WHOLESALE_FEB2026:
            net = qty * price
            key = ("B2B", "Grosmimi", prod_name)
            rev_cbp[key]["2026-02"]["gross"] += net
            rev_cbp[key]["2026-02"]["orders"] += qty
            rev_cbp[key]["2026-02"]["net"] += net

    # ── Revenue aggregated: (brand, product) -> for Section 1B ──
    rev_bp = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for (ch, brand, prod), mo_data in rev_cbp.items():
        for mo, metrics in mo_data.items():
            for mk, mv in metrics.items():
                rev_bp[(brand, prod)][mo][mk] += mv

    # ── Revenue by product only: (product,) -> for Section 1C ──
    rev_p = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for (ch, brand, prod), mo_data in rev_cbp.items():
        for mo, metrics in mo_data.items():
            for mk, mv in metrics.items():
                rev_p[(prod,)][mo][mk] += mv

    # ── Revenue by (channel, brand) -> for Organic and CM ──
    rev_cb = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for (ch, brand, prod), mo_data in rev_cbp.items():
        for mo, metrics in mo_data.items():
            for mk, mv in metrics.items():
                rev_cb[(ch, brand)][mo][mk] += mv

    # ── Shopify costs by brand ──
    shopify = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in q2:
        b = r.get("custom_5036", "") or ""
        dt = r.get("date", "")
        if not dt: continue
        mo = dt[:7]
        all_months.add(mo)
        shopify[b][mo]["sales"] += (r.get("shopify_sales_main.computed.total_sales", 0) or 0)
        shopify[b][mo]["orders"] += (r.get("shopify_sales_main.raw.total_orders", 0) or 0)
        shopify[b][mo]["cogs"] += abs(r.get("shopify_sales_main.raw.cost_of_products_custom", 0) or 0)
        shopify[b][mo]["txn_fees"] += abs(r.get("shopify_sales_main.raw.transaction_fees", 0) or 0)
        shopify[b][mo]["gross"] += (r.get("shopify_sales_main.raw.gross_sales", 0) or 0)
        shopify[b][mo]["disc"] += abs(r.get("shopify_sales_main.raw.discounts", 0) or 0)

    # ── Amazon costs by brand ──
    amz = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in q3:
        b = r.get("custom_5036", "") or ""
        dt = r.get("date", "")
        if not dt: continue
        mo = dt[:7]
        all_months.add(mo)
        amz[b][mo]["sales"] += (r.get("amazonsp_order_items.computed.total_sales_amazon", 0) or 0)
        amz[b][mo]["orders"] += (r.get("amazonsp_order_items.raw.total_orders_amazon", 0) or 0)
        amz[b][mo]["cogs"] += abs(r.get("amazonsp_order_items.raw.cost_of_products_amazon", 0) or 0)
        amz[b][mo]["fees"] += abs(r.get("amazonsp_order_items.raw.total_fees_amazon", 0) or 0)
        amz[b][mo]["gross"] += (r.get("amazonsp_order_items.raw.gross_sales_amazon", 0) or 0)
        amz[b][mo]["promo"] += abs(r.get("amazonsp_order_items.raw.promotion_discounts_amazon", 0) or 0)

    # ── Ads by campaign ──
    plat_cfg = {
        "Amazon Ads": (q5, "amazonads_campaign.raw.cost", "amazonads_campaign.raw.attributed_sales",
                       "amazonads_campaign.raw.clicks", "amazonads_campaign.raw.impressions"),
        "Facebook Ads": (q6, "facebookads_ad_platform_and_device.raw.spend",
                        "facebookads_ad_platform_and_device.raw.purchases_conversion_value",
                        "facebookads_ad_platform_and_device.raw.clicks",
                        "facebookads_ad_platform_and_device.raw.impressions"),
        "Google Ads": (q7, "googleads_campaign_and_device.raw.cost",
                      "googleads_campaign_and_device.raw.conversion_value",
                      "googleads_campaign_and_device.raw.clicks",
                      "googleads_campaign_and_device.raw.impressions"),
        "TikTok Ads": (q8, "tiktokads_campaign_and_platform.raw.spend",
                      "tiktokads_campaign_and_platform.raw.purchases_conversion_value",
                      "tiktokads_campaign_and_platform.raw.clicks",
                      "tiktokads_campaign_and_platform.raw.impressions"),
    }

    # Ads by (platform, brand, product) -> {mo -> {spend, revenue, clicks}}
    ads_pbp = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    # Ads by (landing, brand) -> {mo -> {spend, revenue, clicks}}
    ads_lb = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    # Ads by (brand,) -> {mo -> {spend, revenue, clicks}}
    ads_b = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    # NEW: Ads by (platform, campaign_type, brand)
    ads_ptb = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    # NEW: Ads by (brand, campaign_type)
    ads_bt = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    # NEW: Ads by (campaign_type,)
    ads_t = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    # Campaign-level for vintage tab
    camp_monthly = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    camp_meta = {}  # (plat, camp_name) -> {"brand", "type"}

    for plat, (rows, sk, rk, ck, ik) in plat_cfg.items():
        for r in rows:
            camp = r.get("campaign", "") or ""
            dt = r.get("date", "")
            if not dt: continue
            mo = dt[:7]
            all_months.add(mo)
            sp = r.get(sk, 0) or 0
            rv = r.get(rk, 0) or 0
            cl = r.get(ck, 0) or 0
            br = ad_brand(camp, plat)
            land = ad_landing(camp, plat)
            pcat = ad_prod(camp)
            ctype = get_campaign_type(camp, plat)

            ads_pbp[(plat, br, pcat)][mo]["spend"] += sp
            ads_pbp[(plat, br, pcat)][mo]["revenue"] += rv
            ads_pbp[(plat, br, pcat)][mo]["clicks"] += cl

            ads_lb[(land, br)][mo]["spend"] += sp
            ads_lb[(land, br)][mo]["revenue"] += rv
            ads_lb[(land, br)][mo]["clicks"] += cl

            ads_b[(br,)][mo]["spend"] += sp
            ads_b[(br,)][mo]["revenue"] += rv
            ads_b[(br,)][mo]["clicks"] += cl

            ads_ptb[(plat, ctype, br)][mo]["spend"] += sp
            ads_ptb[(plat, ctype, br)][mo]["revenue"] += rv
            ads_ptb[(plat, ctype, br)][mo]["clicks"] += cl

            ads_bt[(br, ctype)][mo]["spend"] += sp
            ads_bt[(br, ctype)][mo]["revenue"] += rv
            ads_bt[(br, ctype)][mo]["clicks"] += cl

            ads_t[(ctype,)][mo]["spend"] += sp
            ads_t[(ctype,)][mo]["revenue"] += rv
            ads_t[(ctype,)][mo]["clicks"] += cl

            if camp:
                camp_monthly[(plat, camp)][mo]["spend"] += sp
                camp_monthly[(plat, camp)][mo]["revenue"] += rv
                camp_monthly[(plat, camp)][mo]["clicks"] += cl
                camp_meta[(plat, camp)] = {"brand": br, "type": ctype}

    # Derive months list
    months = sorted(all_months)
    ytd_months = [m for m in months if m.startswith("2026-")]

    # ── Build vintage data (two-pass) ──
    ads_vintage = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for (plat, camp_name), mo_data in camp_monthly.items():
        meta = camp_meta.get((plat, camp_name), {"brand": "Other", "type": "Other"})
        brand, ctype = meta["brand"], meta["type"]
        months_with_spend = sorted(m for m in mo_data if mo_data[m].get("spend", 0) > 0)
        if not months_with_spend:
            continue
        vintage_mo = months_with_spend[0]
        camp_label = f"{camp_name[:60]} [{ctype}]"
        for mo, metrics in mo_data.items():
            ads_vintage[(plat, brand, vintage_mo, camp_label)][mo]["spend"] += metrics.get("spend", 0)
            ads_vintage[(plat, brand, vintage_mo, camp_label)][mo]["revenue"] += metrics.get("revenue", 0)
            ads_vintage[(plat, brand, vintage_mo, camp_label)][mo]["clicks"] += metrics.get("clicks", 0)
            ads_vintage[(plat, brand, vintage_mo, camp_label)][mo]["new"] = 1.0 if mo == vintage_mo else 0.0

    # ── Running campaign count: campaigns with spend > 0 per month ──
    running_camps = defaultdict(int)
    for (plat, camp_name), mo_data in camp_monthly.items():
        for mo, metrics in mo_data.items():
            if metrics.get("spend", 0) > 0:
                running_camps[mo] += 1

    # ── Influencer data processing ──
    INF_KEYWORDS = ("pr", "supporter", "sample", "influencer", "giveaway", "collab")
    # PayPal: filter to influencer-only payments
    PP_INF_KW = ("collab", "influencer", "supporter", "paid", "commission",
                 "content", "video", "whitelisting")
    PP_EXCLUDE_KW = ("ads", "marketing", "missing item", "invoice")

    def _is_inf_payment(txn):
        """Check if PayPal txn is an influencer payment."""
        name = txn.get("payer_name", "").strip()
        note = (txn.get("note", "") or "").lower()
        subj = (txn.get("subject", "") or "").lower()
        text = f"{note} {subj}"
        # Exclude ad/platform payments
        for kw in PP_EXCLUDE_KW:
            if kw in text:
                return False
        # Named individual → influencer payment
        if name:
            return True
        # Empty name → check note for influencer keywords
        if not text.strip():
            return False
        for kw in PP_INF_KW:
            if kw in text:
                return True
        if re.search(r'\bpr\b', text):
            return True
        return False

    paypal_names = set()
    paypal_emails = set()
    paypal_monthly = defaultdict(lambda: {"count": 0, "amount": 0.0})
    paypal_people = defaultdict(lambda: defaultdict(lambda: {"count": 0, "amount": 0.0}))
    paypal_ambiguous = []  # for reporting
    for txn in paypal_txns:
        amt = txn.get("amount", 0)
        if amt >= 0:
            continue
        if not _is_inf_payment(txn):
            continue
        name = txn.get("payer_name", "").strip()
        email = txn.get("payer_email", "").lower().strip()
        mo = txn.get("date", "")[:7]
        display_name = name or "(Unknown)"
        txn_id = txn.get("transaction_id", "")
        if name:
            paypal_names.add(name.lower())
        if email:
            paypal_emails.add(email)
        if mo:
            paypal_monthly[mo]["count"] += 1
            paypal_monthly[mo]["amount"] += abs(amt)
            paypal_people[display_name][mo]["count"] += 1
            paypal_people[display_name][mo]["amount"] += abs(amt)

    # Helper: check if order qualifies as influencer
    def _is_inf_order(order):
        tags = [t.strip().lower() for t in order.get("tags", "").split(",")]
        note = (order.get("note") or "").lower()
        for tag in tags:
            for kw in INF_KEYWORDS:
                if kw in tag:
                    return True
        # Note: word boundary match to avoid "product"→"pr" false positives
        if note:
            for kw in ("pr", "sample", "supporter", "influencer", "giveaway", "collab"):
                if re.search(rf'\b{re.escape(kw)}\b', note):
                    return True
        return False

    # Pass 1: collect all influencer people (person-level)
    inf_people = {}  # (name, email) -> set of order months
    for order in inf_orders:
        fs = order.get("fulfillment_status") or ""
        if fs not in ("fulfilled", "shipped"):
            continue
        if not _is_inf_order(order):
            continue
        created = order.get("created_at", "")[:7]
        if not created:
            continue
        cust_name = order.get("customer_name", "").lower().strip()
        cust_email = order.get("customer_email", "").lower().strip()
        key = (cust_name, cust_email)
        if key not in inf_people:
            inf_people[key] = set()
        inf_people[key].add(created)

    # Pass 2: determine paid people (anyone who EVER got PayPal payment)
    # Match by: exact name, exact email, or first+last name fuzzy match
    paypal_last_first = {}  # last_name -> set of first_names (from PayPal)
    for pn in paypal_names:
        parts = pn.split()
        if len(parts) >= 2:
            paypal_last_first.setdefault(parts[-1], set()).add(parts[0])

    paid_people = set()
    for (name, email) in inf_people:
        # Exact name or email match
        if name in paypal_names or email in paypal_emails:
            paid_people.add((name, email))
            continue
        # Fuzzy: same last name + first name starts with same 3 chars
        parts = name.split()
        if len(parts) >= 2:
            last = parts[-1]
            first = parts[0]
            pp_firsts = paypal_last_first.get(last, set())
            for pf in pp_firsts:
                if pf == first or (len(pf) >= 3 and len(first) >= 3 and pf[:3] == first[:3]):
                    paid_people.add((name, email))
                    break

    # Pass 3: classify orders and aggregate
    inf_shipped = defaultdict(int)
    inf_products = defaultdict(lambda: defaultdict(int))
    inf_prod_meta = {}  # {product_title: (brand, cat)} — for Section B hierarchy
    inf_paid = defaultdict(lambda: {"count": 0, "amount": 0.0})
    inf_nonpaid = defaultdict(int)

    for order in inf_orders:
        fs = order.get("fulfillment_status") or ""
        if fs not in ("fulfilled", "shipped"):
            continue
        if not _is_inf_order(order):
            continue
        created = order.get("created_at", "")[:7]
        if not created:
            continue

        inf_shipped[created] += 1
        for li in order.get("line_items", []):
            title = li.get("title", "Unknown")
            inf_products[created][title] += li.get("quantity", 0)
            if title not in inf_prod_meta:
                inf_prod_meta[title] = classify_product(title)

        cust_name = order.get("customer_name", "").lower().strip()
        cust_email = order.get("customer_email", "").lower().strip()
        key = (cust_name, cust_email)
        if key in paid_people:
            inf_paid[created]["count"] += 1
        else:
            inf_nonpaid[created] += 1

    # ── Campaign totals for Campaign Summary tab ──
    camp_totals = {}
    for (plat, camp_name), mo_data in camp_monthly.items():
        meta = camp_meta.get((plat, camp_name), {"brand": "Other", "type": "Other"})
        total_spend = sum(m.get("spend", 0) for m in mo_data.values())
        total_rev = sum(m.get("revenue", 0) for m in mo_data.values())
        total_clicks = sum(m.get("clicks", 0) for m in mo_data.values())
        roas = total_rev / total_spend if total_spend > 0 else 0
        camp_totals[(plat, camp_name)] = {
            "brand": meta["brand"], "type": meta["type"],
            "spend": total_spend, "revenue": total_rev,
            "clicks": total_clicks, "roas": roas,
        }

    # ── Time-window aggregation for Summary tab ──
    WINDOWS = {
        "14d": ["2026-02"],
        "30d": ["2026-02", "2026-01"],
        "90d": ["2026-02", "2026-01", "2025-12", "2025-11"],
    }

    # Campaign window totals: {window: {(plat, camp): metrics}}
    camp_window = {}
    for wname, wmonths in WINDOWS.items():
        ct = {}
        for (plat, camp_name), mo_data in camp_monthly.items():
            meta = camp_meta.get((plat, camp_name), {"brand": "Other", "type": "Other"})
            ws_ = sum(mo_data.get(m, {}).get("spend", 0) for m in wmonths)
            wr = sum(mo_data.get(m, {}).get("revenue", 0) for m in wmonths)
            wc = sum(mo_data.get(m, {}).get("clicks", 0) for m in wmonths)
            if ws_ <= 0:
                continue
            ct[(plat, camp_name)] = {
                "brand": meta["brand"], "type": meta["type"],
                "spend": ws_, "revenue": wr, "clicks": wc,
                "roas": wr / ws_ if ws_ > 0 else 0,
                "cpc": ws_ / wc if wc > 0 else 0,
            }
        camp_window[wname] = ct

    # Ad pivot window: {window: {(plat, brand, type): {spend, revenue, clicks}}}
    ad_pivot_window = {}
    for wname, wmonths in WINDOWS.items():
        pivot = defaultdict(lambda: {"spend": 0, "revenue": 0, "clicks": 0})
        for (plat, camp_name), mo_data in camp_monthly.items():
            meta = camp_meta.get((plat, camp_name), {"brand": "Other", "type": "Other"})
            for m in wmonths:
                d = mo_data.get(m, {})
                key = (plat, meta["brand"], meta["type"])
                pivot[key]["spend"] += d.get("spend", 0)
                pivot[key]["revenue"] += d.get("revenue", 0)
                pivot[key]["clicks"] += d.get("clicks", 0)
        ad_pivot_window[wname] = dict(pivot)

    # Revenue window: {window: {(channel, brand): {gross, disc, orders, net}}}
    rev_window = {}
    for wname, wmonths in WINDOWS.items():
        agg = defaultdict(lambda: {"gross": 0, "disc": 0, "orders": 0, "net": 0})
        for (ch, brand), mo_data in rev_cb.items():
            if ch == "PR":
                continue
            for m in wmonths:
                d = mo_data.get(m, {})
                agg[(ch, brand)]["gross"] += d.get("gross", 0)
                agg[(ch, brand)]["disc"] += d.get("disc", 0)
                agg[(ch, brand)]["orders"] += d.get("orders", 0)
                agg[(ch, brand)]["net"] += d.get("net", 0)
        rev_window[wname] = dict(agg)

    # Organic window: {window: {(channel, brand): {total_rev, ad_rev}}}
    org_window = {}
    for wname, wmonths in WINDOWS.items():
        agg = defaultdict(lambda: {"total_rev": 0, "ad_rev": 0})
        for (ch, brand), mo_data in rev_cb.items():
            if ch == "PR":
                continue
            for m in wmonths:
                d = mo_data.get(m, {})
                agg[(ch, brand)]["total_rev"] += d.get("net", 0)
        for (land, brand), mo_data in ads_lb.items():
            for m in wmonths:
                d = mo_data.get(m, {})
                agg[(land, brand)]["ad_rev"] += d.get("revenue", 0)
        org_window[wname] = dict(agg)

    # ── Search volume data (optional) ──
    search_vol = {}
    try:
        sv_data = load("q12_search_volume.json")
        search_vol = sv_data
    except FileNotFoundError:
        pass

    return {
        "rev_cbp": dict(rev_cbp), "rev_bp": dict(rev_bp), "rev_p": dict(rev_p),
        "rev_cb": dict(rev_cb),
        "shopify": dict(shopify), "amz": dict(amz),
        "ads_pbp": dict(ads_pbp), "ads_lb": dict(ads_lb), "ads_b": dict(ads_b),
        "ads_ptb": dict(ads_ptb), "ads_bt": dict(ads_bt), "ads_t": dict(ads_t),
        "ads_vintage": dict(ads_vintage),
        "camp_monthly": dict(camp_monthly), "camp_meta": dict(camp_meta),
        "camp_totals": camp_totals,
        "running_camps": dict(running_camps),
        "meta_camp_ids": meta_camp_ids, "meta_account_id": meta_account_id,
        "inf_shipped": dict(inf_shipped), "inf_products": {k: dict(v) for k, v in inf_products.items()},
        "inf_prod_meta": inf_prod_meta,
        "inf_paid": dict(inf_paid), "inf_nonpaid": dict(inf_nonpaid),
        "paypal_monthly": dict(paypal_monthly),
        "paypal_people": {k: dict(v) for k, v in paypal_people.items()},
        "inf_paid_people": len(paid_people),
        "has_influencer_data": len(inf_orders) > 0,
        "months": months, "ytd_months": ytd_months,
        "camp_window": camp_window, "ad_pivot_window": ad_pivot_window,
        "rev_window": rev_window, "org_window": org_window,
        "search_vol": search_vol,
    }


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1: REVENUE
# ══════════════════════════════════════════════════════════════════════════════

def build_revenue(wb, D):
    """Single Revenue tab with 3 sub-sections: 1A (Ch x Brand x Prod), 1B (Brand x Prod), 1C (Prod)."""
    ws = wb.create_sheet("Revenue")
    ws.sheet_properties.tabColor = "002060"
    months = D["months"]
    ytd_months = D["ytd_months"]
    level_headers_3 = ["", "Channel", "Brand", "Product"]
    level_headers_2 = ["", "Brand", "Product", ""]
    level_headers_1 = ["", "Product", "", ""]
    total_cols = calc_total_cols(months, level_headers_3)

    rev_configs = [
        {"type": "data", "key": "gross", "title": "GROSS SALES", "fmt": DOL, "metric": "gross"},
        {"type": "data", "key": "disc", "title": "DISCOUNTS", "fmt": DOL, "metric": "disc"},
        {"type": "data", "key": "orders", "title": "# OF ORDERS", "fmt": NUM, "metric": "orders"},
        {"type": "ratio", "key": "aov", "title": "AOV (Gross / Orders)", "fmt": DOL2, "num_key": "gross", "den_key": "orders"},
        {"type": "ratio", "key": "disc_rate", "title": "DISCOUNT RATE (Disc / Gross)", "fmt": PCT, "num_key": "disc", "den_key": "gross"},
        {"type": "data", "key": "net", "title": "NET SALES", "fmt": DOL, "metric": "net"},
    ]

    row = 1
    # ── Section 1A: Channel x Brand x Product ──
    row = _write_section_title(ws, row, "SECTION 1A — CHANNEL x BRAND x PRODUCT", total_cols)
    tree_1a = build_generic_tree(D["rev_cbp"], 3, ytd_months, "net")
    flat_1a = flatten(tree_1a)
    row, sec_map_1a = write_sections(ws, row, flat_1a, months, ytd_months, rev_configs,
                                      level_headers_3, total_cols)
    row += 1

    # ── Section 1B: Brand x Product ──
    row = _write_section_title(ws, row, "SECTION 1B — BRAND x PRODUCT", total_cols)
    tree_1b = build_generic_tree(D["rev_bp"], 2, ytd_months, "net")
    flat_1b = flatten(tree_1b)
    # Adjust level_headers for 2-deep: level 0=TOTAL->col B, 1=Brand->col B, 2=Product->col C
    row, sec_map_1b = write_sections(ws, row, flat_1b, months, ytd_months, rev_configs,
                                      level_headers_2, total_cols)
    row += 1

    # ── Section 1C: Product only (flat) ──
    row = _write_section_title(ws, row, "SECTION 1C — PRODUCT CATEGORY", total_cols)
    tree_1c = build_generic_tree(D["rev_p"], 1, ytd_months, "net")
    flat_1c = flatten(tree_1c)
    row, sec_map_1c = write_sections(ws, row, flat_1c, months, ytd_months, rev_configs,
                                      level_headers_1, total_cols)

    # Column widths
    _set_col_widths(ws, total_cols, months, level_headers_3)
    # Column grouping: collapse/expand hierarchy levels
    ws.sheet_properties.outlinePr.summaryRight = False
    ws.column_dimensions['C'].outlineLevel = 1
    ws.column_dimensions['D'].outlineLevel = 2
    ws.freeze_panes = "E3"

    print(f"  Revenue: {row} rows, {total_cols} cols")
    return sec_map_1a, flat_1a


def _write_section_title(ws, row, title, total_cols):
    """Write a full-width dark blue section title bar."""
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = DBLUE
        cell.font = WF
    ws.cell(row=row, column=1, value=title)
    return row + 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2: ADS
# ══════════════════════════════════════════════════════════════════════════════

def build_ads(wb, D):
    """Ads tab with 4 sub-sections:
    2A: Platform x Campaign Type x Brand
    2B: Landing Channel x Brand
    2C: Brand x Campaign Type
    2D: Campaign Type (flat)
    """
    ws = wb.create_sheet("Ads")
    ws.sheet_properties.tabColor = "ED7D31"
    months = D["months"]
    ytd_months = D["ytd_months"]
    lh_3 = ["", "Platform", "Camp Type", "Brand"]
    lh_2a = ["", "Landing Ch", "Brand", ""]
    lh_2b = ["", "Brand", "Camp Type", ""]
    lh_1 = ["", "Camp Type", "", ""]
    total_cols = calc_total_cols(months, lh_3)

    ads_configs = [
        {"type": "data", "key": "spend", "title": "AD SPEND", "fmt": DOL, "metric": "spend"},
        {"type": "data", "key": "revenue", "title": "AD REVENUE", "fmt": DOL, "metric": "revenue"},
        {"type": "data", "key": "clicks", "title": "CLICKS", "fmt": NUM, "metric": "clicks"},
        {"type": "ratio", "key": "roas", "title": "ROAS (Revenue / Spend)", "fmt": DEC, "num_key": "revenue", "den_key": "spend"},
        {"type": "ratio", "key": "cpc", "title": "CPC (Spend / Clicks)", "fmt": DOL2, "num_key": "spend", "den_key": "clicks"},
    ]

    row = 1
    # ── 2A: Platform x Campaign Type x Brand ──
    row = _write_section_title(ws, row, "SECTION 2A — PLATFORM x CAMPAIGN TYPE x BRAND", total_cols)
    tree_2a = build_generic_tree(D["ads_ptb"], 3, ytd_months, "spend")
    flat_2a = flatten(tree_2a)
    row, _ = write_sections(ws, row, flat_2a, months, ytd_months, ads_configs, lh_3, total_cols)
    row += 1

    # ── 2B: Landing Channel x Brand ──
    row = _write_section_title(ws, row, "SECTION 2B — LANDING CHANNEL x BRAND", total_cols)
    tree_2b = build_generic_tree(D["ads_lb"], 2, ytd_months, "spend")
    flat_2b = flatten(tree_2b)
    row, _ = write_sections(ws, row, flat_2b, months, ytd_months, ads_configs, lh_2a, total_cols)
    row += 1

    # ── 2C: Brand x Campaign Type ──
    row = _write_section_title(ws, row, "SECTION 2C — BRAND x CAMPAIGN TYPE", total_cols)
    tree_2c = build_generic_tree(D["ads_bt"], 2, ytd_months, "spend")
    flat_2c = flatten(tree_2c)
    row, _ = write_sections(ws, row, flat_2c, months, ytd_months, ads_configs, lh_2b, total_cols)
    row += 1

    # ── 2D: Campaign Type (flat) ──
    row = _write_section_title(ws, row, "SECTION 2D — BY CAMPAIGN TYPE", total_cols)
    tree_2d = build_generic_tree(D["ads_t"], 1, ytd_months, "spend")
    flat_2d = flatten(tree_2d)
    row, _ = write_sections(ws, row, flat_2d, months, ytd_months, ads_configs, lh_1, total_cols)
    row += 1

    # ── Running Campaigns (Spend > $0) ──
    row = _write_section_title(ws, row, "RUNNING CAMPAIGNS (Spend > $0)", total_cols)
    col_start = len(lh_3) + 1
    has_partial = months[-1] == "2026-02" if months else False
    reg_months = months[:-1] if has_partial else list(months)
    # Header row
    for i, m in enumerate(reg_months):
        c = col_start + i
        ws.cell(row=row, column=c, value=m[2:].replace("-", "/"))
        ws.cell(row=row, column=c).font = BF
        ws.cell(row=row, column=c).fill = LGREEN
        ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")
    if has_partial:
        pc = col_start + len(reg_months)
        ws.cell(row=row, column=pc, value=f"26/02 ({PARTIAL_DAY}d)")
        ws.cell(row=row, column=pc).font = BF
        ws.cell(row=row, column=pc).fill = LGREEN
        ws.cell(row=row, column=pc + 1, value="26/02 Full")
        ws.cell(row=row, column=pc + 1).font = BF
        ws.cell(row=row, column=pc + 1).fill = LGREEN
    row += 1
    # Data row
    ws.cell(row=row, column=1, value="# Running")
    ws.cell(row=row, column=1).font = BF
    running = D["running_camps"]
    for i, m in enumerate(reg_months):
        c = col_start + i
        ws.cell(row=row, column=c, value=running.get(m, 0))
        ws.cell(row=row, column=c).number_format = NUM
    if has_partial:
        pc = col_start + len(reg_months)
        ws.cell(row=row, column=pc, value=running.get("2026-02", 0))
        ws.cell(row=row, column=pc).number_format = NUM
        ws.cell(row=row, column=pc + 1, value=running.get("2026-02", 0))
        ws.cell(row=row, column=pc + 1).number_format = NUM
    row += 1

    _set_col_widths(ws, total_cols, months, lh_3)
    ws.sheet_properties.outlinePr.summaryRight = False
    ws.column_dimensions['C'].outlineLevel = 1
    ws.column_dimensions['D'].outlineLevel = 2
    ws.freeze_panes = "E3"

    print(f"  Ads: {row} rows")


# ══════════════════════════════════════════════════════════════════════════════
# TAB: CAMPAIGN VINTAGE
# ══════════════════════════════════════════════════════════════════════════════

def build_vintage(wb, D):
    """Campaign Details tab: Platform x Brand x Vintage Month x Campaign [Type].
    4-level hierarchy showing campaign cohort analysis by first month of spend.
    """
    ws = wb.create_sheet("Campaign Details")
    ws.sheet_properties.tabColor = "7030A0"
    months = D["months"]
    ytd_months = D["ytd_months"]
    lh = ["", "Platform", "Brand", "Vintage", "Campaign"]
    total_cols = calc_total_cols(months, lh)

    vintage_configs = [
        {"type": "data",  "key": "new",     "title": "# NEW CAMPAIGNS",        "fmt": NUM, "metric": "new", "no_extrapolate": True},
        {"type": "data",  "key": "spend",   "title": "AD SPEND",               "fmt": DOL, "metric": "spend"},
        {"type": "data",  "key": "revenue", "title": "AD REVENUE",             "fmt": DOL, "metric": "revenue"},
        {"type": "ratio", "key": "roas",    "title": "ROAS (Revenue / Spend)", "fmt": DEC, "num_key": "revenue", "den_key": "spend"},
        {"type": "data",  "key": "clicks",  "title": "CLICKS",                 "fmt": NUM, "metric": "clicks"},
        {"type": "ratio", "key": "cpc",     "title": "CPC (Spend / Clicks)",   "fmt": DOL2, "num_key": "spend", "den_key": "clicks"},
    ]

    row = 1
    row = _write_section_title(ws, row, "CAMPAIGN VINTAGE — PLATFORM x BRAND x VINTAGE x CAMPAIGN", total_cols)
    tree = build_generic_tree(D["ads_vintage"], 4, ytd_months, "spend")
    flat = flatten(tree)
    row, sec_map = write_sections(ws, row, flat, months, ytd_months, vintage_configs, lh, total_cols)

    # ── Meta Ads Manager hyperlinks on campaign names ──
    meta_ids = D.get("meta_camp_ids", {})
    meta_acct = D.get("meta_account_id", "").replace("act_", "")
    if meta_ids and meta_acct:
        first_row = sec_map.get("new")  # first data row of the first section
        if first_row:
            link_font = Font(size=9, color="0563C1", underline="single")
            current_plat = ""
            for ni, node in enumerate(flat):
                if node.level == 1:
                    current_plat = node.label
                if node.level == 4 and current_plat == "Facebook Ads":
                    # Extract campaign name from label (strip " [Type]" suffix)
                    camp_label = node.label
                    camp_name = camp_label.rsplit(" [", 1)[0] if " [" in camp_label else camp_label
                    camp_id = meta_ids.get(camp_name)
                    if camp_id:
                        url = f"https://www.facebook.com/adsmanager/manage/campaigns/edit?act={meta_acct}&selected_campaign_ids={camp_id}"
                        cell = ws.cell(row=first_row + ni, column=5)
                        cell.hyperlink = url
                        cell.font = link_font

    # ── Conditional formatting: full-row coloring for highlighted campaigns ──
    col_start = len(lh) + 1
    has_partial = months[-1] == "2026-02" if months else False
    num_reg = len(months) - 1 if has_partial else len(months)
    extra = 2 if has_partial else 0
    data_width = num_reg + extra + 1

    vintage_camp_rows = {}
    roas_start = sec_map.get("roas")
    cpc_start = sec_map.get("cpc")

    # Determine which campaigns need highlighting based on ROAS/CPC values
    highlighted_camps = {}  # node_index -> "skyblue" or "pink"
    if roas_start and cpc_start:
        for ni, node in enumerate(flat):
            if node.level == 4:
                vintage_camp_rows[node.label] = roas_start + ni
            if node.level < 4:
                continue

            # Check latest actual data month for ROAS and CPC
            latest = months[-1] if months else None
            roas_val = node.data.get(latest, {}).get("revenue", 0) / max(node.data.get(latest, {}).get("spend", 0), 0.01) if latest else 0
            cpc_val = node.data.get(latest, {}).get("spend", 0) / max(node.data.get(latest, {}).get("clicks", 0), 1) if latest else 0
            total_spend = sum(m.get("spend", 0) for m in node.data.values())
            if total_spend < 50:
                continue

            if roas_val > 0 and roas_val < 2.0:
                highlighted_camps[ni] = SKYBLUE
            elif roas_val >= 4.0:
                highlighted_camps[ni] = PINK
            elif cpc_val > 0.15:
                highlighted_camps[ni] = SKYBLUE
            elif 0 < cpc_val < 0.10:
                highlighted_camps[ni] = PINK

        # Apply full-row coloring across ALL sections for highlighted campaigns
        for ni, fill in highlighted_camps.items():
            for sec_key in sec_map:
                sec_start = sec_map[sec_key]
                r = sec_start + ni
                for ci in range(1, total_cols + 1):
                    cell = ws.cell(row=r, column=ci)
                    if cell.fill == MBLUE or cell.fill == LLGRAY:
                        continue  # Don't override hierarchy styling
                    cell.fill = fill

    D["vintage_camp_rows"] = vintage_camp_rows

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 40
    pct_gap = col_start + data_width
    yoy_gap = pct_gap + 1 + data_width
    for i in range(col_start, total_cols + 1):
        cl = get_column_letter(i)
        if i == pct_gap or i == yoy_gap:
            ws.column_dimensions[cl].width = 6
        else:
            ws.column_dimensions[cl].width = 12

    ws.sheet_properties.outlinePr.summaryRight = False
    ws.column_dimensions['C'].outlineLevel = 1
    ws.column_dimensions['D'].outlineLevel = 2
    ws.column_dimensions['E'].outlineLevel = 3
    ws.freeze_panes = "F3"
    print(f"  Campaign Details: {row} rows")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3: ORGANIC
# ══════════════════════════════════════════════════════════════════════════════

def build_organic(wb, D):
    """Organic tab: Channel x Brand hierarchy with Total Rev, Ad Rev, Organic Rev, Organic %.

    Uses pre-aggregated rev_cb and ads_lb data to build trees, then writes sections.
    """
    ws = wb.create_sheet("Organic")
    ws.sheet_properties.tabColor = "548235"
    months = D["months"]
    ytd_months = D["ytd_months"]
    level_headers = ["", "Channel", "Brand", ""]
    total_cols = calc_total_cols(months, level_headers)

    # Build organic data: (channel, brand) -> {mo -> {total_rev, ad_rev, organic_rev}}
    # Map landing channels to sales channels for ad attribution
    LAND_MAP = {"Onzenna": "Onzenna", "Amazon": "Amazon", "TargetPlus": "TargetPlus"}

    # Total rev by (channel, brand)
    rev_cb = D["rev_cb"]
    ads_lb = D["ads_lb"]

    # Build organic tree data: (channel, brand) -> {mo -> metrics}
    # total_rev comes from rev_cb, ad_rev from ads_lb mapped by channel
    organic_raw = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    for (ch, brand), mo_data in rev_cb.items():
        if ch == "PR":
            continue  # Exclude PR from organic calc
        for mo, metrics in mo_data.items():
            organic_raw[(ch, brand)][mo]["total_rev"] += metrics.get("net", 0)

    # Map ad landing revenue to channels
    for (land, brand), mo_data in ads_lb.items():
        ch = land  # Landing channel maps directly to sales channel for known ones
        for mo, metrics in mo_data.items():
            organic_raw[(ch, brand)][mo]["ad_rev"] += metrics.get("revenue", 0)

    # Build tree
    tree = build_generic_tree(dict(organic_raw), 2, ytd_months, "total_rev")
    flat_nodes = flatten(tree)

    organic_configs = [
        {"type": "data", "key": "total_rev", "title": "TOTAL REVENUE (Net Sales)", "fmt": DOL, "metric": "total_rev"},
        {"type": "data", "key": "ad_rev", "title": "AD REVENUE (Attributed)", "fmt": DOL, "metric": "ad_rev"},
        {"type": "diff", "key": "organic_rev", "title": "ORGANIC REVENUE (Total - Ad)", "fmt": DOL, "a_key": "total_rev", "b_key": "ad_rev"},
        {"type": "ratio", "key": "organic_pct", "title": "ORGANIC % (Organic / Total)", "fmt": PCT, "num_key": "organic_rev", "den_key": "total_rev"},
    ]

    row = 1
    row = _write_section_title(ws, row, "SECTION 3 — ORGANIC REVENUE BACK-SOLVE", total_cols)
    row, sec_map = write_sections(ws, row, flat_nodes, months, ytd_months, organic_configs,
                                   level_headers, total_cols)

    # Column widths
    _set_col_widths(ws, total_cols, months, level_headers)
    ws.sheet_properties.outlinePr.summaryRight = False
    ws.column_dimensions['C'].outlineLevel = 1
    ws.column_dimensions['D'].outlineLevel = 2
    ws.freeze_panes = "E3"

    print(f"  Organic: {row} rows")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4: CONTRIBUTION MARGIN
# ══════════════════════════════════════════════════════════════════════════════

def build_cm(wb, D):
    """CM tab: Channel x Brand with Net Sales, COGS, Gross Profit, Channel Fees, CM Before Ads, Ad Spend, CM After Ads."""
    ws = wb.create_sheet("CM")
    ws.sheet_properties.tabColor = "FF0000"
    months = D["months"]
    ytd_months = D["ytd_months"]
    level_headers = ["", "Channel", "Brand", ""]
    total_cols = calc_total_cols(months, level_headers)

    rev_cb = D["rev_cb"]
    shopify = D["shopify"]
    amz = D["amz"]
    ads_lb = D["ads_lb"]

    # Build CM data: (channel, brand) -> {mo -> {net, cogs, fees, ad_spend}}
    cm_raw = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    for (ch, brand), mo_data in rev_cb.items():
        if ch == "PR":
            continue
        for mo, metrics in mo_data.items():
            cm_raw[(ch, brand)][mo]["net"] += metrics.get("net", 0)

    # COGS and Fees
    for (ch, brand), mo_data in rev_cb.items():
        if ch == "PR":
            continue
        for mo in months:
            net = rev_cb.get((ch, brand), {}).get(mo, {}).get("net", 0)
            if ch == "Amazon":
                cm_raw[(ch, brand)][mo]["cogs"] += amz.get(brand, {}).get(mo, {}).get("cogs", 0)
                cm_raw[(ch, brand)][mo]["fees"] += amz.get(brand, {}).get(mo, {}).get("fees", 0)
            else:
                # Allocate Shopify COGS proportionally
                total_non_amz = sum(
                    rev_cb.get((c, brand), {}).get(mo, {}).get("net", 0)
                    for c in MODEL_CH if c != "Amazon"
                )
                ratio = net / total_non_amz if total_non_amz else 0
                cm_raw[(ch, brand)][mo]["cogs"] += shopify.get(brand, {}).get(mo, {}).get("cogs", 0) * ratio
                cm_raw[(ch, brand)][mo]["fees"] += shopify.get(brand, {}).get(mo, {}).get("txn_fees", 0) * ratio

    # Ad spend: map landing to channel
    for (land, brand), mo_data in ads_lb.items():
        for mo, metrics in mo_data.items():
            cm_raw[(land, brand)][mo]["ad_spend"] += metrics.get("spend", 0)

    # Build tree
    tree = build_generic_tree(dict(cm_raw), 2, ytd_months, "net")
    flat_nodes = flatten(tree)

    cm_configs = [
        {"type": "data", "key": "net", "title": "NET SALES", "fmt": DOL, "metric": "net"},
        {"type": "data", "key": "cogs", "title": "COGS", "fmt": DOL, "metric": "cogs"},
        {"type": "diff", "key": "gp", "title": "GROSS PROFIT (Net - COGS)", "fmt": DOL, "a_key": "net", "b_key": "cogs"},
        {"type": "data", "key": "fees", "title": "CHANNEL FEES", "fmt": DOL, "metric": "fees"},
        {"type": "diff", "key": "cm_before", "title": "CM BEFORE ADS (GP - Fees)", "fmt": DOL, "a_key": "gp", "b_key": "fees"},
        {"type": "data", "key": "ad_spend", "title": "AD SPEND", "fmt": DOL, "metric": "ad_spend"},
        {"type": "diff", "key": "cm_after", "title": "CM AFTER ADS (CM Before - Ad Spend)", "fmt": DOL, "a_key": "cm_before", "b_key": "ad_spend"},
    ]

    row = 1
    row = _write_section_title(ws, row, "SECTION 4 — CONTRIBUTION MARGIN", total_cols)
    row, sec_map = write_sections(ws, row, flat_nodes, months, ytd_months, cm_configs,
                                   level_headers, total_cols)

    # Column widths
    _set_col_widths(ws, total_cols, months, level_headers)
    ws.sheet_properties.outlinePr.summaryRight = False
    ws.column_dimensions['C'].outlineLevel = 1
    ws.column_dimensions['D'].outlineLevel = 2
    ws.freeze_panes = "E3"

    print(f"  CM: {row} rows")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5: MODEL CHECK
# ══════════════════════════════════════════════════════════════════════════════

def build_model_check(wb, D):
    """Model Check tab: cross-sheet reference validation using formulas."""
    ws = wb.create_sheet("Model Check")
    ws.sheet_properties.tabColor = "808080"
    months = D["months"]
    ytd_months = D["ytd_months"]
    NC = 6
    row = 1

    # Section header
    for c in range(1, NC + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = DBLUE
        cell.font = WF
    ws.cell(row=row, column=1, value="MODEL CONSISTENCY CHECK")
    row += 2

    # Revenue check header
    for c in range(1, NC + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = MBLUE
        cell.font = WF
    ws.cell(row=row, column=1, value="REVENUE RECONCILIATION")
    row += 1

    headers = ["Check", "Section", "Sheet", "Row", "YTD Column", "Status"]
    for ci, h in enumerate(headers):
        cell = ws.cell(row=row, column=1 + ci, value=h)
        cell.font = HF
        cell.fill = LGRAY
        cell.border = THIN
    row += 1

    # Reference: Revenue tab TOTAL row (row 3 of first data section = first_data_row)
    # In Section 1A, the TOTAL node is always the first node in flat list -> first data row
    # Section 1B TOTAL is also first node in its flat list
    # We record these for informational purposes
    checks = [
        ("1A Total Net Sales", "Revenue", "TOTAL row in Section 1A Net Sales"),
        ("1B Total Net Sales", "Revenue", "TOTAL row in Section 1B Net Sales"),
        ("1C Total Net Sales", "Revenue", "TOTAL row in Section 1C Net Sales"),
    ]
    for label, sheet, desc in checks:
        cell = ws.cell(row=row, column=1, value=label)
        cell.border = THIN
        cell.font = NF
        cell = ws.cell(row=row, column=2, value=desc)
        cell.border = THIN
        cell.font = NF
        cell = ws.cell(row=row, column=3, value=sheet)
        cell.border = THIN
        cell.font = NF
        cell = ws.cell(row=row, column=6, value="Check in Excel")
        cell.border = THIN
        cell.font = NF
        row += 1

    row += 1

    # Ads check
    for c in range(1, NC + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = MBLUE
        cell.font = WF
    ws.cell(row=row, column=1, value="AD SPEND RECONCILIATION")
    row += 1

    checks_ads = [
        ("2A Total Spend", "Ads", "TOTAL row in Section 2A (Platform x Type x Brand)"),
        ("2B Total Spend", "Ads", "TOTAL row in Section 2B (Landing x Brand)"),
        ("2C Total Spend", "Ads", "TOTAL row in Section 2C (Brand x Type)"),
        ("2D Total Spend", "Ads", "TOTAL row in Section 2D (Campaign Type)"),
        ("Vintage Total Spend", "Campaign Details", "TOTAL in Campaign Details Ad Spend"),
    ]
    for label, sheet, desc in checks_ads:
        cell = ws.cell(row=row, column=1, value=label)
        cell.border = THIN
        cell.font = NF
        cell = ws.cell(row=row, column=2, value=desc)
        cell.border = THIN
        cell.font = NF
        cell = ws.cell(row=row, column=3, value=sheet)
        cell.border = THIN
        cell.font = NF
        cell = ws.cell(row=row, column=6, value="Check in Excel")
        cell.border = THIN
        cell.font = NF
        row += 1

    row += 1

    # Instructions
    for c in range(1, NC + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = LGREEN
        cell.font = BF
    ws.cell(row=row, column=1, value="All totals are formula-driven. Open in Excel to verify =SUM formulas propagate correctly.")
    row += 1
    ws.cell(row=row, column=1, value="Section 1A TOTAL should equal 1B TOTAL should equal 1C TOTAL for each metric.")
    ws.cell(row=row, column=1).font = NF
    row += 1
    ws.cell(row=row, column=1, value="Section 2A TOTAL should equal 2B TOTAL should equal 2C TOTAL should equal 2D TOTAL for Spend and Revenue.")
    ws.cell(row=row, column=1).font = NF
    row += 1
    ws.cell(row=row, column=1, value="Campaign Details TOTAL should equal Ads 2A TOTAL for Spend and Revenue (excludes zero-spend campaigns).")
    ws.cell(row=row, column=1).font = NF

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    print(f"  Model Check: {row} rows")


# ══════════════════════════════════════════════════════════════════════════════
# TAB: SUMMARY
# ══════════════════════════════════════════════════════════════════════════════

WINDOW_ORDER = ["90d", "30d", "14d"]
WINDOW_LABELS = {
    "90d": "LAST 90 DAYS (Nov 2025 – Feb 2026)",
    "30d": "LAST 30 DAYS (Jan – Feb 2026)",
    "14d": f"LAST 14 DAYS (Feb 1–{PARTIAL_DAY} 2026)",
}
MIN_SPEND_SUMMARY = 50


def _write_window_title(ws, row, title, total_cols):
    """Medium-blue sub-header for time window labels."""
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = MBLUE
        cell.font = WF9
    ws.cell(row=row, column=1, value=title)
    return row + 1


def _write_ranked_camps(ws, row, title, camps, total_cols, metric="roas", vintage_rows=None):
    """Write a ranked campaign section. metric='roas' for CVR, 'cpc' for Traffic.
    vintage_rows: dict mapping campaign labels to Campaign Details row numbers for hyperlinking.
    """
    row = _write_section_title(ws, row, title, total_cols)
    if metric == "cpc":
        headers = ["#", "Platform", "Campaign", "Brand", "Spend", "Clicks", "CPC", "Revenue", "ROAS"]
        cpc_idx, roas_idx = 6, 8  # 0-based column positions in vals
    else:
        headers = ["#", "Platform", "Campaign", "Brand", "Spend", "Revenue", "ROAS", "Clicks", "CPC"]
        cpc_idx, roas_idx = 8, 6
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = BF
        cell.fill = LGREEN
        cell.border = THIN
    row += 1
    for rank, ((plat, name), m) in enumerate(camps, 1):
        if metric == "cpc":
            vals = [rank, plat, name[:60], m["brand"],
                    m["spend"], m["clicks"], m["cpc"], m["revenue"], m["roas"]]
            fmts = [NUM, None, None, None, DOL, NUM, DOL2, DOL, DEC]
        else:
            vals = [rank, plat, name[:60], m["brand"],
                    m["spend"], m["revenue"], m["roas"], m["clicks"], m["cpc"]]
            fmts = [NUM, None, None, None, DOL, DOL, DEC, NUM, DOL2]
        for i, (v, f) in enumerate(zip(vals, fmts)):
            cell = ws.cell(row=row, column=i + 1, value=v)
            cell.font = NF
            cell.border = THIN
            if f:
                cell.number_format = f
            # Conditional coloring for CPC/ROAS values
            if i == cpc_idx and isinstance(v, (int, float)):
                if v > 0.15:
                    cell.fill = SKYBLUE
                elif 0 < v < 0.10:
                    cell.fill = PINK
            if i == roas_idx and isinstance(v, (int, float)):
                if 0 < v < 2.0:
                    cell.fill = SKYBLUE
                elif v >= 4.0:
                    cell.fill = PINK
        # Formula reference to Campaign Details tab (instead of hyperlink)
        if vintage_rows:
            for vl, vr in vintage_rows.items():
                vl_base = vl.rsplit(" [", 1)[0] if " [" in vl else vl
                if vl_base == name or name.startswith(vl_base[:50]):
                    camp_cell = ws.cell(row=row, column=3)
                    # Label column in Campaign Details is D (col 4)
                    camp_cell.value = f"='Campaign Details'!D{vr}"
                    camp_cell.font = Font(size=9, color="0563C1")
                    break
        row += 1
    if not camps:
        ws.cell(row=row, column=1, value="No qualifying campaigns").font = NF
        row += 1
    return row + 1


def _write_summary_row(ws, row, col1, col2, vals_dict, metrics_cfg, level=2):
    """Write a single hierarchy row with level-based styling.
    level: 0=TOTAL (blue), 1=parent (gray bold), 2=child (normal)
    """
    if level == 0:
        fill, font, border = MBLUE, Font(bold=True, color="FFFFFF", size=9), TOTAL_BORDER
    elif level == 1:
        fill, font, border = LLGRAY, BF, None
    else:
        fill, font, border = None, NF, None

    cell = ws.cell(row=row, column=1, value=col1)
    cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border

    cell = ws.cell(row=row, column=2, value=col2)
    cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border

    for i, mc in enumerate(metrics_cfg):
        cell = ws.cell(row=row, column=3 + i, value=vals_dict.get(mc["key"], 0))
        cell.font = font
        cell.number_format = mc["fmt"]
        if fill: cell.fill = fill
        if border: cell.border = border


def _write_ad_pivot_table(ws, row, title, pivot_data, orientation, total_cols):
    """2-level hierarchy ad pivot: Channel>Brand or Brand>Channel."""
    row = _write_section_title(ws, row, title, total_cols)

    if orientation == "channel_first":
        h1, h2 = "Channel", "Brand"
    else:
        h1, h2 = "Brand", "Channel"

    metrics_cfg = [
        {"key": "spend", "header": "Spend", "fmt": DOL},
        {"key": "revenue", "header": "Revenue", "fmt": DOL},
        {"key": "roas", "header": "ROAS", "fmt": DEC},
        {"key": "clicks", "header": "Clicks", "fmt": NUM},
        {"key": "cpc", "header": "CPC", "fmt": DOL2},
    ]
    headers = [h1, h2] + [m["header"] for m in metrics_cfg]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = BF
        cell.fill = LGREEN
        cell.border = THIN
    row += 1

    # Build hierarchy
    hierarchy = defaultdict(lambda: defaultdict(lambda: {"spend": 0, "revenue": 0, "clicks": 0}))
    grand = {"spend": 0, "revenue": 0, "clicks": 0}
    for (plat, brand, ctype), m in pivot_data.items():
        l1, l2 = (plat, brand) if orientation == "channel_first" else (brand, plat)
        for k in grand:
            hierarchy[l1][l2][k] += m[k]
            grand[k] += m[k]

    # TOTAL row
    grand["roas"] = grand["revenue"] / grand["spend"] if grand["spend"] > 0 else 0
    grand["cpc"] = grand["spend"] / grand["clicks"] if grand["clicks"] > 0 else 0
    _write_summary_row(ws, row, "TOTAL", "", grand, metrics_cfg, level=0)
    row += 1

    # Sort level 1 by spend desc
    for l1 in sorted(hierarchy, key=lambda k: sum(v["spend"] for v in hierarchy[k].values()), reverse=True):
        l1_agg = {"spend": 0, "revenue": 0, "clicks": 0}
        for v in hierarchy[l1].values():
            for k in l1_agg:
                l1_agg[k] += v[k]
        l1_agg["roas"] = l1_agg["revenue"] / l1_agg["spend"] if l1_agg["spend"] > 0 else 0
        l1_agg["cpc"] = l1_agg["spend"] / l1_agg["clicks"] if l1_agg["clicks"] > 0 else 0
        _write_summary_row(ws, row, l1, "", l1_agg, metrics_cfg, level=1)
        row += 1

        for l2 in sorted(hierarchy[l1], key=lambda k: hierarchy[l1][k]["spend"], reverse=True):
            d = hierarchy[l1][l2].copy()
            d["roas"] = d["revenue"] / d["spend"] if d["spend"] > 0 else 0
            d["cpc"] = d["spend"] / d["clicks"] if d["clicks"] > 0 else 0
            _write_summary_row(ws, row, "", l2, d, metrics_cfg, level=2)
            row += 1

    return row + 1


def _write_rev_summary_table(ws, row, rev_data, total_cols):
    """Channel > Brand revenue summary table."""
    metrics_cfg = [
        {"key": "gross", "header": "Gross Sales", "fmt": DOL},
        {"key": "disc", "header": "Discounts", "fmt": DOL},
        {"key": "orders", "header": "Orders", "fmt": NUM},
        {"key": "aov", "header": "AOV", "fmt": DOL2},
        {"key": "disc_rate", "header": "Disc Rate", "fmt": PCT},
        {"key": "net", "header": "Net Sales", "fmt": DOL},
    ]
    headers = ["Channel", "Brand"] + [m["header"] for m in metrics_cfg]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = BF
        cell.fill = LGREEN
        cell.border = THIN
    row += 1

    # Build hierarchy
    hierarchy = defaultdict(lambda: defaultdict(lambda: {"gross": 0, "disc": 0, "orders": 0, "net": 0}))
    grand = {"gross": 0, "disc": 0, "orders": 0, "net": 0}
    for (ch, brand), m in rev_data.items():
        for k in grand:
            hierarchy[ch][brand][k] += m.get(k, 0)
            grand[k] += m.get(k, 0)

    def _enrich(d):
        d["aov"] = d["gross"] / d["orders"] if d["orders"] > 0 else 0
        d["disc_rate"] = d["disc"] / d["gross"] if d["gross"] > 0 else 0
        return d

    # TOTAL row
    _write_summary_row(ws, row, "TOTAL", "", _enrich(grand.copy()), metrics_cfg, level=0)
    row += 1

    for ch in sorted(hierarchy, key=lambda k: (0 if k != "Other" else 1,
                     -sum(v["net"] for v in hierarchy[k].values()))):
        ch_agg = {"gross": 0, "disc": 0, "orders": 0, "net": 0}
        for v in hierarchy[ch].values():
            for k in ch_agg:
                ch_agg[k] += v[k]
        _write_summary_row(ws, row, ch, "", _enrich(ch_agg), metrics_cfg, level=1)
        row += 1

        for brand in sorted(hierarchy[ch], key=lambda k: (0 if k != "Other" else 1,
                            -hierarchy[ch][k]["net"])):
            _write_summary_row(ws, row, "", brand, _enrich(hierarchy[ch][brand].copy()), metrics_cfg, level=2)
            row += 1

    return row + 1


def _write_organic_summary_table(ws, row, org_data, total_cols):
    """Channel > Brand organic summary table."""
    metrics_cfg = [
        {"key": "total_rev", "header": "Total Rev", "fmt": DOL},
        {"key": "ad_rev", "header": "Ad Rev", "fmt": DOL},
        {"key": "organic_rev", "header": "Organic Rev", "fmt": DOL},
        {"key": "organic_pct", "header": "Organic %", "fmt": PCT},
    ]
    headers = ["Channel", "Brand"] + [m["header"] for m in metrics_cfg]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = BF
        cell.fill = LGREEN
        cell.border = THIN
    row += 1

    # Build hierarchy
    hierarchy = defaultdict(lambda: defaultdict(lambda: {"total_rev": 0, "ad_rev": 0}))
    grand = {"total_rev": 0, "ad_rev": 0}
    for (ch, brand), m in org_data.items():
        for k in grand:
            hierarchy[ch][brand][k] += m.get(k, 0)
            grand[k] += m.get(k, 0)

    def _enrich(d):
        d["organic_rev"] = d["total_rev"] - d["ad_rev"]
        d["organic_pct"] = d["organic_rev"] / d["total_rev"] if d["total_rev"] > 0 else 0
        return d

    # TOTAL row
    _write_summary_row(ws, row, "TOTAL", "", _enrich(grand.copy()), metrics_cfg, level=0)
    row += 1

    for ch in sorted(hierarchy, key=lambda k: (0 if k != "Other" else 1,
                     -sum(v["total_rev"] for v in hierarchy[k].values()))):
        ch_agg = {"total_rev": 0, "ad_rev": 0}
        for v in hierarchy[ch].values():
            for k in ch_agg:
                ch_agg[k] += v[k]
        _write_summary_row(ws, row, ch, "", _enrich(ch_agg), metrics_cfg, level=1)
        row += 1

        for brand in sorted(hierarchy[ch], key=lambda k: (0 if k != "Other" else 1,
                            -hierarchy[ch][k]["total_rev"])):
            _write_summary_row(ws, row, "", brand, _enrich(hierarchy[ch][brand].copy()), metrics_cfg, level=2)
            row += 1

    return row + 1


def build_summary(wb, D):
    """Summary tab: multi-period campaign rankings, ad pivots, revenue and organic summaries."""
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "FFC000"
    total_cols = 9

    camp_window = D["camp_window"]
    ad_pivot_window = D["ad_pivot_window"]
    rev_window = D["rev_window"]
    org_window = D["org_window"]

    row = 1

    # ═══════════════════════════════════════════════════
    # SECTION 1: CAMPAIGN PERFORMANCE RANKINGS
    # ═══════════════════════════════════════════════════
    row = _write_section_title(ws, row, "CAMPAIGN PERFORMANCE RANKINGS", total_cols)

    for wname in WINDOW_ORDER:
        ct = camp_window[wname]
        row = _write_window_title(ws, row, WINDOW_LABELS[wname], total_cols)

        # CVR campaigns — ranked by ROAS
        cvr = sorted(
            [(k, v) for k, v in ct.items() if v["type"] == "CVR" and v["spend"] >= MIN_SPEND_SUMMARY],
            key=lambda x: (-x[1]["roas"], -x[1]["spend"]),
        )
        vr = D.get("vintage_camp_rows", {})
        row = _write_ranked_camps(ws, row, "Top 5 ROAS — CVR Campaigns", cvr[:5], total_cols, "roas", vintage_rows=vr)
        worst_cvr = sorted(
            [(k, v) for k, v in ct.items() if v["type"] == "CVR" and v["spend"] >= MIN_SPEND_SUMMARY],
            key=lambda x: (x[1]["roas"], -x[1]["spend"]),
        )
        row = _write_ranked_camps(ws, row, "Worst 5 ROAS — CVR Campaigns", worst_cvr[:5], total_cols, "roas", vintage_rows=vr)

        # Traffic campaigns — ranked by CPC
        traffic = sorted(
            [(k, v) for k, v in ct.items()
             if v["type"] == "Traffic" and v["spend"] >= MIN_SPEND_SUMMARY and v["clicks"] > 0],
            key=lambda x: (x[1]["cpc"], -x[1]["spend"]),
        )
        row = _write_ranked_camps(ws, row, "Top 5 CPC — Traffic Campaigns (Lowest)", traffic[:5], total_cols, "cpc", vintage_rows=vr)
        worst_traffic = sorted(
            [(k, v) for k, v in ct.items()
             if v["type"] == "Traffic" and v["spend"] >= MIN_SPEND_SUMMARY and v["clicks"] > 0],
            key=lambda x: (-x[1]["cpc"], -x[1]["spend"]),
        )
        row = _write_ranked_camps(ws, row, "Worst 5 CPC — Traffic Campaigns (Highest)", worst_traffic[:5], total_cols, "cpc", vintage_rows=vr)

    # ═══════════════════════════════════════════════════
    # SECTION 2: AD SPEND SUMMARY
    # ═══════════════════════════════════════════════════
    row = _write_section_title(ws, row, "AD SPEND SUMMARY", total_cols)

    for wname in WINDOW_ORDER:
        pivot = ad_pivot_window[wname]
        row = _write_window_title(ws, row, WINDOW_LABELS[wname], total_cols)
        row = _write_ad_pivot_table(ws, row, "Channel > Brand", pivot, "channel_first", total_cols)
        row = _write_ad_pivot_table(ws, row, "Brand > Channel", pivot, "brand_first", total_cols)

    # ═══════════════════════════════════════════════════
    # SECTION 3: REVENUE SUMMARY
    # ═══════════════════════════════════════════════════
    row = _write_section_title(ws, row, "REVENUE SUMMARY", total_cols)

    for wname in WINDOW_ORDER:
        rev = rev_window[wname]
        row = _write_window_title(ws, row, WINDOW_LABELS[wname], total_cols)
        row = _write_rev_summary_table(ws, row, rev, total_cols)

    # ═══════════════════════════════════════════════════
    # SECTION 4: ORGANIC REVENUE SUMMARY
    # ═══════════════════════════════════════════════════
    row = _write_section_title(ws, row, "ORGANIC REVENUE SUMMARY", total_cols)

    for wname in WINDOW_ORDER:
        org = org_window[wname]
        row = _write_window_title(ws, row, WINDOW_LABELS[wname], total_cols)
        row = _write_organic_summary_table(ws, row, org, total_cols)

    # ═══════════════════════════════════════════════════
    # SECTION 5: SEARCH VOLUME SNAPSHOT
    # ═══════════════════════════════════════════════════
    sv = D.get("search_vol", {})
    if sv:
        row = _write_section_title(ws, row, "SEARCH VOLUME SNAPSHOT", total_cols)
        row = _write_search_volume_summary(ws, row, sv, total_cols)

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 12

    ws.freeze_panes = "A2"
    print(f"  Summary: {row} rows")


# ══════════════════════════════════════════════════════════════════════════════
# TAB: SEARCH VOLUME
# ══════════════════════════════════════════════════════════════════════════════

# Display order and label mapping for search volume keywords
SV_KEYWORDS = [
    ("Onzenna", ["onzenna"]),
    ("zezebaebae", ["zezebaebae"]),
    ("Grosmimi", ["grosmimi"]),
    ("Alpremio", ["alpremio"]),
    ("Cha&Mom", ["cha and mom", "cha&mom"]),
    ("Comme Moi", ["comme moi"]),
    ("BabyRabbit", ["babyrabbit", "baby rabbit"]),
    ("Naeiae", ["naeiae"]),
    ("Bamboobebe", ["bamboobebe"]),
    ("Hattung", ["hattung"]),
    ("Beemymagic", ["beemymagic"]),
    ("Nature Love Mere", ["nature love mere"]),
    ("PPSU", ["ppsu"]),
    ("PPSU Bottle", ["ppsu bottle"]),
    ("PPSU Baby Bottle", ["ppsu baby bottle"]),
    ("Phyto Seline", ["phyto seline", "phytoseline"]),
]


def _sv_merge(data_dict, api_keys, field="search_volume"):
    """Sum values across multiple API keyword keys for a merged display keyword."""
    total = 0
    for k in api_keys:
        entry = data_dict.get(k, {})
        v = entry.get(field) if isinstance(entry, dict) else None
        if v is not None:
            total += v
    return total if total > 0 else None


def _sv_merge_monthly(data_dict, api_keys, month):
    """Sum monthly values across merged keyword keys for Google Ads."""
    total = 0
    found = False
    for k in api_keys:
        entry = data_dict.get(k, {})
        monthly = entry.get("monthly", {})
        v = monthly.get(month)
        if v is not None:
            total += v
            found = True
    return total if found else None


def _sv_merge_trends(trends_dict, api_keys, month):
    """Average Google Trends values across merged keyword keys."""
    vals = []
    for k in api_keys:
        entry = trends_dict.get(k, {})
        v = entry.get(month)
        if v is not None:
            vals.append(v)
    return round(sum(vals) / len(vals), 1) if vals else None


def build_search_volume(wb, D):
    """Search Volume tab: unified month columns across all sections."""
    sv = D.get("search_vol", {})
    if not sv:
        return

    google_ads = sv.get("google_ads", {})
    amazon = sv.get("amazon", {})
    trends = sv.get("google_trends", {})

    ws = wb.create_sheet("Search Volume")
    ws.sheet_properties.tabColor = "7030A0"

    # Build unified month list (union of all sources, from 2024-01)
    all_months = set()
    for entry in google_ads.values():
        all_months.update(entry.get("monthly", {}).keys())
    for entry in trends.values():
        if isinstance(entry, dict):
            all_months.update(k for k in entry.keys() if k >= "2024-01")
    all_months = sorted(m for m in all_months if m >= "2024-01")

    row = 1
    meta_cols = 3  # Keyword, Avg Volume, CPC
    total_cols = meta_cols + len(all_months)

    # ═══════════════════════════════════════════════════
    # SECTION 1: GOOGLE SEARCH VOLUME (Monthly)
    # ═══════════════════════════════════════════════════
    row = _write_section_title(ws, row, "GOOGLE SEARCH VOLUME (Monthly, Absolute)", total_cols)

    headers = ["Keyword", "Avg Volume", "CPC"] + [_month_label(m) for m in all_months]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = BF
        cell.fill = LGRAY
        cell.border = THIN
    row += 1

    for label, api_keys in SV_KEYWORDS:
        vol = _sv_merge(google_ads, api_keys, "search_volume")
        cpc_val = None
        for k in api_keys:
            c = google_ads.get(k, {}).get("cpc")
            if c is not None:
                cpc_val = (cpc_val or 0) + c
        ws.cell(row=row, column=1, value=label).font = NF
        c = ws.cell(row=row, column=2, value=vol if vol else None)
        c.font = NF; c.number_format = NUM
        if cpc_val is not None:
            c = ws.cell(row=row, column=3, value=cpc_val)
            c.font = NF; c.number_format = DOL2
        for j, mo in enumerate(all_months):
            v = _sv_merge_monthly(google_ads, api_keys, mo)
            cell = ws.cell(row=row, column=meta_cols + 1 + j, value=v)
            cell.font = NF
            if v is not None:
                cell.number_format = NUM
        row += 1

    row += 1

    # ═══════════════════════════════════════════════════
    # SECTION 2: AMAZON SEARCH VOLUME (Current snapshot)
    # ═══════════════════════════════════════════════════
    row = _write_section_title(ws, row, "AMAZON SEARCH VOLUME (Current Snapshot — no monthly breakdown)", total_cols)

    headers = ["Keyword", "Current Vol", ""] + [_month_label(m) for m in all_months]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = BF
        cell.fill = LGRAY
        cell.border = THIN
    row += 1

    for label, api_keys in SV_KEYWORDS:
        vol = _sv_merge(amazon, api_keys, "search_volume")
        ws.cell(row=row, column=1, value=label).font = NF
        c = ws.cell(row=row, column=2, value=vol)
        c.font = NF
        if vol is not None:
            c.number_format = NUM
        row += 1

    row += 1

    # ═══════════════════════════════════════════════════
    # SECTION 3: GOOGLE TRENDS (Monthly, Relative 0-100)
    # ═══════════════════════════════════════════════════
    row = _write_section_title(ws, row, "GOOGLE TRENDS (Monthly, Relative Index 0-100)", total_cols)

    headers = ["Keyword", "", ""] + [_month_label(m) for m in all_months]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = BF
        cell.fill = LGRAY
        cell.border = THIN
    row += 1

    for label, api_keys in SV_KEYWORDS:
        ws.cell(row=row, column=1, value=label).font = NF
        for j, mo in enumerate(all_months):
            v = _sv_merge_trends(trends, api_keys, mo)
            cell = ws.cell(row=row, column=meta_cols + 1 + j, value=v)
            cell.font = NF
            if v is not None:
                cell.number_format = DEC
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    for i in range(meta_cols + 1, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 10

    ws.freeze_panes = "D2"
    print(f"  Search Volume: {row} rows")


def _write_search_volume_summary(ws, row, sv_data, total_cols):
    """Write search volume summary section for the Summary tab (Google Ads latest month + Amazon)."""
    google_ads = sv_data.get("google_ads", {})
    amazon = sv_data.get("amazon", {})

    headers = ["Keyword", "Google (Monthly)", "Amazon (Current)"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = BF
        cell.fill = LGREEN
        cell.border = THIN
    row += 1

    for label, api_keys in SV_KEYWORDS:
        # Google: latest available month
        ga_vol = None
        for k in api_keys:
            monthly = google_ads.get(k, {}).get("monthly", {})
            if monthly:
                latest = max(monthly.keys())
                v = monthly[latest]
                if v is not None:
                    ga_vol = (ga_vol or 0) + v
        # Amazon
        amz_vol = _sv_merge(amazon, api_keys, "search_volume")

        ws.cell(row=row, column=1, value=label).font = NF
        cell = ws.cell(row=row, column=2, value=ga_vol if ga_vol is not None else 0)
        cell.font = NF
        cell.number_format = NUM
        cell = ws.cell(row=row, column=3, value=amz_vol if amz_vol is not None else 0)
        cell.font = NF
        cell.number_format = NUM
        row += 1

    return row + 1


# ══════════════════════════════════════════════════════════════════════════════
# TAB: INFLUENCER DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════

def build_influencer_dashboard(wb, D):
    """Influencer Dashboard: PR orders, product breakdown, paid/non-paid, PayPal payments."""
    ws = wb.create_sheet("Influencer Dashboard")
    ws.sheet_properties.tabColor = "FF69B4"
    # Row grouping: expand button above the detail rows
    ws.sheet_properties.outlinePr.summaryBelow = False

    months = D["months"]
    inf_shipped = D["inf_shipped"]
    inf_products = D["inf_products"]
    inf_paid = D["inf_paid"]
    inf_nonpaid = D["inf_nonpaid"]

    row = 1
    col_start = 5  # data starts at col E (A=indicator, B=brand, C=cat, D=product)
    total_cols = len(months) + col_start - 1

    tot_col = col_start + len(months)  # Total column
    first_cl = get_column_letter(col_start)
    last_cl = get_column_letter(col_start + len(months) - 1)

    def _month_hdr(r):
        """Write month header row with Level 1 styling (light fill + borders)."""
        for i, m in enumerate(months):
            c = col_start + i
            cell = ws.cell(row=r, column=c, value=m[2:].replace("-", "/"))
            cell.font = BF
            cell.fill = LGRAY
            cell.border = THIN
            cell.alignment = Alignment(horizontal="center")
        # Total header
        cell = ws.cell(row=r, column=tot_col, value="Total")
        cell.font = BF; cell.fill = LGRAY; cell.border = THIN
        cell.alignment = Alignment(horizontal="center")

    def _section_banner(r, title):
        """Write section banner row (DBLUE full width)."""
        for c in range(1, tot_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = DBLUE
            cell.font = WF
        ws.cell(row=r, column=1, value=title)

    # ── Section A: Monthly Shipped PR Order Count ──
    _section_banner(row, "SECTION A — PR ORDERS: MONTHLY SHIPPED COUNT")
    row += 1

    # Month headers (Level 1)
    cell = ws.cell(row=row, column=2, value="Month")
    cell.font = BF; cell.fill = LGRAY; cell.border = THIN
    _month_hdr(row)
    row += 1

    # Shipped count row (Level 0: dark fill = aggregate total)
    cell = ws.cell(row=row, column=2, value="Shipped Orders")
    cell.font = WF9; cell.fill = MBLUE; cell.border = THIN
    for i, m in enumerate(months):
        c = col_start + i
        cell = ws.cell(row=row, column=c, value=inf_shipped.get(m, 0))
        cell.number_format = NUM; cell.font = WF9; cell.fill = MBLUE; cell.border = THIN
    cell = ws.cell(row=row, column=tot_col, value=f"=SUM({first_cl}{row}:{last_cl}{row})")
    cell.number_format = NUM; cell.font = WF9; cell.fill = MBLUE; cell.border = THIN
    row += 2

    # ── Section B: Monthly Shipped by Product (Grand Total → Brand → Category → Product) ──
    _section_banner(row, "SECTION B — PR ORDERS: MONTHLY SHIPPED BY PRODUCT")
    row += 1

    inf_prod_meta = D.get("inf_prod_meta", {})

    # Collect all product titles across all months
    all_products = set()
    for mo_prods in inf_products.values():
        all_products.update(mo_prods.keys())

    # Group: brand → cat → [product titles]
    brand_cat_prods = {}
    for title in all_products:
        brand, cat = inf_prod_meta.get(title, classify_product(title))
        brand_cat_prods.setdefault(brand, {}).setdefault(cat, []).append(title)

    # Sorting helpers
    def _ptot(t):
        return sum(inf_products.get(m, {}).get(t, 0) for m in months)
    def _ctot(br, ca):
        return sum(_ptot(t) for t in brand_cat_prods[br][ca])
    def _btot(br):
        return sum(_ctot(br, ca) for ca in brand_cat_prods[br])

    sorted_brands = sorted(brand_cat_prods, key=_btot, reverse=True)

    # Column header row: A blank | B=Brand | C=Category | D=Product | E+=months
    for c in range(1, tot_col + 1):
        ws.cell(row=row, column=c).fill = LGRAY
    ws.cell(row=row, column=2, value="Brand").font = BF
    ws.cell(row=row, column=2).fill = LGRAY; ws.cell(row=row, column=2).border = THIN
    ws.cell(row=row, column=3, value="Category").font = BF
    ws.cell(row=row, column=3).fill = LGRAY; ws.cell(row=row, column=3).border = THIN
    ws.cell(row=row, column=4, value="Product").font = BF
    ws.cell(row=row, column=4).fill = LGRAY; ws.cell(row=row, column=4).border = THIN
    _month_hdr(row)
    row += 1

    tot_letter = get_column_letter(tot_col)

    # Grand Total placeholder row (MBLUE — darkest, filled after brand rows)
    grand_total_row = row
    row += 1
    all_brand_rows = []

    for brand in sorted_brands:
        brand_row = row
        row += 1  # placeholder — written after children

        sorted_cats = sorted(brand_cat_prods[brand], key=lambda c: _ctot(brand, c), reverse=True)
        cat_rows = []

        for cat in sorted_cats:
            cat_row = row
            row += 1  # placeholder

            sorted_prods = sorted(brand_cat_prods[brand][cat], key=_ptot, reverse=True)
            prod_rows = []

            for title in sorted_prods:
                # Level 3 — Product detail: col D, NO fill, normal font, row group level 2
                cell = ws.cell(row=row, column=4, value=title[:44])
                cell.font = NF
                for i, m in enumerate(months):
                    c = col_start + i
                    qty = inf_products.get(m, {}).get(title, 0)
                    if qty:
                        ws.cell(row=row, column=c, value=qty).number_format = NUM
                ws.cell(row=row, column=tot_col,
                        value=f"=SUM({first_cl}{row}:{last_cl}{row})").number_format = NUM
                ws.row_dimensions[row].outlineLevel = 2
                ws.row_dimensions[row].hidden = True
                prod_rows.append(row)
                row += 1

            # Level 2 — Category row: col C, LLGRAY fill, black bold, row group level 1
            cell = ws.cell(row=cat_row, column=3, value=cat)
            cell.font = BF; cell.fill = LLGRAY; cell.border = THIN
            for i, m in enumerate(months):
                c = col_start + i
                col_letter = get_column_letter(c)
                refs = ",".join(f"{col_letter}{pr}" for pr in prod_rows) if prod_rows else ""
                v = f"=SUM({refs})" if refs else 0
                cell2 = ws.cell(row=cat_row, column=c, value=v)
                cell2.number_format = NUM; cell2.font = BF; cell2.fill = LLGRAY; cell2.border = THIN
            refs = ",".join(f"{tot_letter}{pr}" for pr in prod_rows) if prod_rows else ""
            cell3 = ws.cell(row=cat_row, column=tot_col,
                            value=f"=SUM({refs})" if refs else 0)
            cell3.number_format = NUM; cell3.font = BF; cell3.fill = LLGRAY; cell3.border = THIN
            ws.row_dimensions[cat_row].outlineLevel = 1
            ws.row_dimensions[cat_row].hidden = True

            cat_rows.append(cat_row)

        # Level 1 — Brand row: col B, LGRAY fill, black bold
        cell = ws.cell(row=brand_row, column=2, value=brand)
        cell.font = BF; cell.fill = LGRAY; cell.border = THIN
        for i, m in enumerate(months):
            c = col_start + i
            col_letter = get_column_letter(c)
            refs = ",".join(f"{col_letter}{cr}" for cr in cat_rows) if cat_rows else ""
            v = f"=SUM({refs})" if refs else 0
            cell2 = ws.cell(row=brand_row, column=c, value=v)
            cell2.number_format = NUM; cell2.font = BF; cell2.fill = LGRAY; cell2.border = THIN
        refs = ",".join(f"{tot_letter}{cr}" for cr in cat_rows) if cat_rows else ""
        cell3 = ws.cell(row=brand_row, column=tot_col,
                        value=f"=SUM({refs})" if refs else 0)
        cell3.number_format = NUM; cell3.font = BF; cell3.fill = LGRAY; cell3.border = THIN

        all_brand_rows.append(brand_row)

    # Level 0 — Grand Total row: MBLUE (darkest), white bold, SUM of all brand rows
    cell = ws.cell(row=grand_total_row, column=2, value="TOTAL — All Brands")
    cell.font = WF9; cell.fill = MBLUE; cell.border = THIN
    for i, m in enumerate(months):
        c = col_start + i
        col_letter = get_column_letter(c)
        refs = ",".join(f"{col_letter}{br}" for br in all_brand_rows) if all_brand_rows else ""
        v = f"=SUM({refs})" if refs else 0
        cell2 = ws.cell(row=grand_total_row, column=c, value=v)
        cell2.number_format = NUM; cell2.font = WF9; cell2.fill = MBLUE; cell2.border = THIN
    refs = ",".join(f"{tot_letter}{br}" for br in all_brand_rows) if all_brand_rows else ""
    cell3 = ws.cell(row=grand_total_row, column=tot_col,
                    value=f"=SUM({refs})" if refs else 0)
    cell3.number_format = NUM; cell3.font = WF9; cell3.fill = MBLUE; cell3.border = THIN

    row += 1

    # ── Section C: Paid vs Non-paid Split ──
    _section_banner(row, "SECTION C — PAID vs NON-PAID SPLIT (Shipped Orders)")
    row += 1

    cell = ws.cell(row=row, column=2, value="Category")
    cell.font = BF; cell.fill = LGRAY; cell.border = THIN
    _month_hdr(row)
    row += 1

    # Paid row (Level 2: lighter fill)
    cell = ws.cell(row=row, column=2, value="Paid (PayPal matched)")
    cell.font = BF; cell.fill = LLGRAY; cell.border = THIN
    for i, m in enumerate(months):
        c = col_start + i
        val = inf_paid.get(m, {}).get("count", 0) if isinstance(inf_paid.get(m), dict) else 0
        cell = ws.cell(row=row, column=c, value=val)
        cell.number_format = NUM; cell.fill = LLGRAY; cell.border = THIN
    cell = ws.cell(row=row, column=tot_col, value=f"=SUM({first_cl}{row}:{last_cl}{row})")
    cell.number_format = NUM; cell.fill = LLGRAY; cell.border = THIN
    row += 1

    # Non-paid row (Level 2: lighter fill)
    cell = ws.cell(row=row, column=2, value="Non-paid")
    cell.font = BF; cell.fill = LLGRAY; cell.border = THIN
    for i, m in enumerate(months):
        c = col_start + i
        cell = ws.cell(row=row, column=c, value=inf_nonpaid.get(m, 0))
        cell.number_format = NUM; cell.fill = LLGRAY; cell.border = THIN
    cell = ws.cell(row=row, column=tot_col, value=f"=SUM({first_cl}{row}:{last_cl}{row})")
    cell.number_format = NUM; cell.fill = LLGRAY; cell.border = THIN
    row += 2

    # ── Section D: PayPal Influencer Payments ──
    paypal_monthly = D.get("paypal_monthly", {})
    paypal_people = D.get("paypal_people", {})

    _section_banner(row, "SECTION D — PAYPAL INFLUENCER PAYMENTS")
    row += 1

    # Month headers (Level 1)
    cell = ws.cell(row=row, column=2, value="Metric")
    cell.font = BF; cell.fill = LGRAY; cell.border = THIN
    _month_hdr(row)
    row += 1

    # Total # Payments (Level 0: dark fill)
    cell = ws.cell(row=row, column=2, value="# Payments (Total)")
    cell.font = WF9; cell.fill = MBLUE; cell.border = THIN
    for i, m in enumerate(months):
        c = col_start + i
        val = paypal_monthly.get(m, {}).get("count", 0) if isinstance(paypal_monthly.get(m), dict) else 0
        cell = ws.cell(row=row, column=c, value=val)
        cell.number_format = NUM; cell.font = WF9; cell.fill = MBLUE; cell.border = THIN
    cell = ws.cell(row=row, column=tot_col, value=f"=SUM({first_cl}{row}:{last_cl}{row})")
    cell.number_format = NUM; cell.font = WF9; cell.fill = MBLUE; cell.border = THIN
    row += 1

    # Total $ Amount (Level 0: dark fill)
    cell = ws.cell(row=row, column=2, value="$ Amount (Total)")
    cell.font = WF9; cell.fill = MBLUE; cell.border = THIN
    for i, m in enumerate(months):
        c = col_start + i
        val = paypal_monthly.get(m, {}).get("amount", 0) if isinstance(paypal_monthly.get(m), dict) else 0
        cell = ws.cell(row=row, column=c, value=val)
        cell.number_format = DOL; cell.font = WF9; cell.fill = MBLUE; cell.border = THIN
    cell = ws.cell(row=row, column=tot_col, value=f"=SUM({first_cl}{row}:{last_cl}{row})")
    cell.number_format = DOL; cell.font = WF9; cell.fill = MBLUE; cell.border = THIN
    row += 1

    # BY PERSON sub-header (Level 1)
    cell = ws.cell(row=row, column=2, value="BY PERSON (expand to view)")
    cell.font = BF; cell.fill = LGRAY; cell.border = THIN
    for i in range(len(months)):
        c = col_start + i
        ws.cell(row=row, column=c).fill = LGRAY
        ws.cell(row=row, column=c).border = THIN
    ws.cell(row=row, column=tot_col).fill = LGRAY
    ws.cell(row=row, column=tot_col).border = THIN
    row += 1

    # Person rows (Level 3: grouped, collapsed by default)
    person_start = row
    people_totals = {}
    for name, mo_data in paypal_people.items():
        total = sum(v.get("amount", 0) for v in mo_data.values() if isinstance(v, dict))
        people_totals[name] = total
    sorted_people = sorted(people_totals, key=lambda x: people_totals[x], reverse=True)

    for person in sorted_people:
        mo_data = paypal_people[person]
        ws.cell(row=row, column=2, value=person).font = BF
        for i, m in enumerate(months):
            c = col_start + i
            d = mo_data.get(m)
            amt = d.get("amount", 0) if isinstance(d, dict) else 0
            if amt:
                ws.cell(row=row, column=c, value=amt).number_format = DOL
        # Total formula
        ws.cell(row=row, column=tot_col, value=f"=SUM({first_cl}{row}:{last_cl}{row})").number_format = DOL
        row += 1

    person_end = row - 1

    # Group person detail rows (collapsed by default)
    if person_start <= person_end:
        for r in range(person_start, person_end + 1):
            ws.row_dimensions[r].outlineLevel = 1
            ws.row_dimensions[r].hidden = True

    row += 1

    # Column widths
    ws.column_dimensions["A"].width = 4   # level indicator
    ws.column_dimensions["B"].width = 20  # Brand
    ws.column_dimensions["C"].width = 20  # Category
    ws.column_dimensions["D"].width = 38  # Product
    for i in range(len(months)):
        cl = get_column_letter(col_start + i)
        ws.column_dimensions[cl].width = 10
    ws.column_dimensions[get_column_letter(tot_col)].width = 12

    # Column outline levels (C=Category collapses, D=Product collapses deeper)
    ws.sheet_properties.outlinePr.summaryRight = False
    ws.column_dimensions["C"].outlineLevel = 1
    ws.column_dimensions["D"].outlineLevel = 2

    ws.freeze_panes = "E3"
    print(f"  Influencer Dashboard: {row} rows, {len(sorted_people)} PayPal people")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("Loading data...")
    D = preprocess()
    print(f"  Months: {D['months'][0]} to {D['months'][-1]} ({len(D['months'])} months)")
    print(f"  YTD months: {D['ytd_months']}")

    print("Building financial model...")
    wb = Workbook()
    wb.remove(wb.active)

    # Tab 1: Revenue
    sec_map_1a, flat_1a = build_revenue(wb, D)

    # Tab 2: Ads
    build_ads(wb, D)

    # Tab 3: Campaign Details
    build_vintage(wb, D)

    # Tab 4: Organic
    build_organic(wb, D)

    # Tab 5: CM
    build_cm(wb, D)

    # Tab 6: Model Check
    build_model_check(wb, D)

    # Tab 7: Summary
    build_summary(wb, D)

    # Tab 8: Search Volume (if data available)
    if D.get("search_vol"):
        build_search_volume(wb, D)

    # Tab 9: Influencer Dashboard (if data available)
    if D.get("has_influencer_data"):
        build_influencer_dashboard(wb, D)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    try:
        wb.save(OUT)
        print(f"\nModel saved: {OUT}")
    except PermissionError:
        fallback = OUT.replace(".xlsx", "_new.xlsx")
        wb.save(fallback)
        print(f"\nOriginal file locked — saved to: {fallback}")
        print("Close the Excel file and rename/replace manually.")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
