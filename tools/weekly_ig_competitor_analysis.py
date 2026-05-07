"""
WAT Tool: Weekly IG competitor brand analysis (automated).
Scrapes JP baby brand Instagram profiles via imginn.com (Firecrawl),
parses engagement data, generates analysis Excel, and emails to se.heo.

Runs every Friday 14:00 KST alongside weekly_ig_planner.py.

Pipeline:
  1. Scrape ~12 JP baby brand IG profiles via Firecrawl (imginn.com)
  2. Parse posts, engagement, content types
  3. Build 4-sheet analysis Excel
  4. Email Excel to se.heo@orbiters.co.kr
  5. Save competitor insights JSON for content planner to consume

Usage:
    python tools/weekly_ig_competitor_analysis.py             # full run
    python tools/weekly_ig_competitor_analysis.py --dry-run   # scrape + Excel, skip email
    python tools/weekly_ig_competitor_analysis.py --skip-scrape  # use cached data
"""

import os
import re
import sys
import json
import time
import base64
import argparse
import logging
import subprocess
from datetime import datetime, date
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Config ───────────────────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).parent.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
BRANDS_DIR = PROJECT_ROOT / ".firecrawl" / "ig-baby-brands"
INSIGHTS_PATH = TMP_DIR / "competitor_insights.json"

GMAIL_CREDENTIALS_PATH = os.getenv("GMAIL_OAUTH_CREDENTIALS_PATH", "credentials/gmail_oauth_credentials.json")
GMAIL_TOKEN_PATH = os.getenv("GMAIL_TOKEN_PATH",
                              os.getenv("ZEZEBAEBAE_GMAIL_TOKEN_PATH", "credentials/gmail_token.json"))
GMAIL_SENDER = os.getenv("GMAIL_SENDER", "orbiters11@gmail.com")
SEEUN_EMAIL = "se.heo@orbiters.co.kr"

# Target brands to scrape (imginn.com handles)
TARGET_BRANDS = [
    "bboxforkidsjapan",
    "pigeon_official_jp",
    "richell_baby_official",
    "thermos_kk",
    "stokkejapan",
    "combi_babylifedesign",
    "adenandanaisjp",
    "nuna_jpn",
    "ergobabyjapan",
    "mushie",
    "cybex_japan",
    "aprica_official",
    "babybjorn_japan",
    "joie_japan",
    "bettababy_official",
    "eddisonmama",
    "alobaby_official",
    "10mois_official",
    "andobaby_official",
    "naomiito_official",
]

RATE_LIMIT_DELAY = 3  # seconds between Firecrawl requests


# ── Step 1: Scrape via Firecrawl ─────────────────────────────────────────────

def scrape_brands(brands: list[str]) -> int:
    """Scrape brands in parallel via Firecrawl CLI. Returns count of scraped files."""
    BRANDS_DIR.mkdir(parents=True, exist_ok=True)

    # Build parallel scrape commands
    commands = []
    for handle in brands:
        url = f"https://imginn.com/{handle}/"
        output = BRANDS_DIR / f"{handle}.md"
        commands.append(f'firecrawl scrape "{url}" -o "{output}"')

    # Run in batches of 5 (respect rate limits)
    batch_size = 5
    scraped = 0
    for i in range(0, len(commands), batch_size):
        batch = commands[i:i + batch_size]
        cmd = " & ".join(batch)
        if sys.platform == "win32":
            cmd = " & ".join(batch)
        else:
            cmd = " & ".join(batch) + " & wait"

        logger.info(f"Scraping batch {i // batch_size + 1}/{(len(commands) + batch_size - 1) // batch_size}...")
        try:
            subprocess.run(cmd, shell=True, timeout=120, capture_output=True)
        except subprocess.TimeoutExpired:
            logger.warning(f"Batch {i // batch_size + 1} timed out")

        # Count results
        for handle in brands[i:i + batch_size]:
            md_file = BRANDS_DIR / f"{handle}.md"
            if md_file.exists() and md_file.stat().st_size > 500:
                scraped += 1

        if i + batch_size < len(commands):
            time.sleep(RATE_LIMIT_DELAY)

    logger.info(f"Scraped {scraped}/{len(brands)} brands")
    return scraped


# ── Step 2: Parse scraped data ───────────────────────────────────────────────

def parse_brand(md_file: Path) -> dict | None:
    """Parse a scraped imginn.com markdown file into structured data."""
    text = md_file.read_text(encoding="utf-8", errors="replace")
    if len(text.split("\n")) < 50:
        return None

    account_match = re.search(r"\*\*@(\w[\w.]+)\*\*", text)
    account = account_match.group(1) if account_match else md_file.stem

    name_match = re.search(r"\*\*([^@*]+)\*\*", text)
    display_name = name_match.group(1).strip() if name_match else ""

    followers_match = re.search(r"([\d,.]+[KkMm]?)\s*\n\s*followers", text)
    followers = followers_match.group(1) if followers_match else "N/A"

    posts_count_match = re.search(r"([\d,.]+)\s*\n\s*posts", text)
    posts_count = posts_count_match.group(1) if posts_count_match else "N/A"

    # Bio
    bio = ""
    lines = text.split("\n")
    for i, line in enumerate(lines):
        if "**@" in line and i + 1 < len(lines):
            bio_parts = []
            for j in range(i + 2, min(i + 8, len(lines))):
                l = lines[j].strip()
                if l and not l.isdigit() and "posts" not in l.lower() and "followers" not in l.lower() and not l.startswith("[") and not l.startswith("*"):
                    bio_parts.append(l)
                if l.isdigit() or "posts" in l.lower():
                    break
            bio = " ".join(bio_parts)[:300]
            break

    # Posts
    post_pattern = re.compile(
        r"\[!\[(.+?)\]\(https://s\d+\.imginn\.com.+?\)\]\(.+?\)\s*\n\s*(\d+)\s*\n\s*(\d+)\s*\n\s*(.+?)(?:\n|$)",
        re.DOTALL,
    )

    posts = []
    for m in post_pattern.finditer(text):
        caption_raw = m.group(1)
        caption = re.sub(r"\s*by @\w[\w.]+ at .+$", "", caption_raw).strip()
        likes = int(m.group(2))
        comments = int(m.group(3))
        time_ago = m.group(4).strip()
        hashtags = re.findall(r"#[\w\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF]+", caption)

        is_video = "\U0001f3a5" in caption or "ライブ" in caption or "LIVE" in caption
        is_collab = "コラボ" in caption or "PR" in caption
        is_campaign = "キャンペーン" in caption or "プレゼント" in caption or "セール" in caption
        is_product = any(w in caption for w in ["発売", "新作", "新色", "新登場", "商品"])
        is_lifestyle = any(w in caption for w in ["暮らし", "おでかけ", "子育て", "ママ", "育児"])

        content_type = (
            "ライブ/動画" if is_video
            else "コラボ" if is_collab
            else "キャンペーン" if is_campaign
            else "商品紹介" if is_product
            else "ライフスタイル" if is_lifestyle
            else "一般"
        )

        posts.append({
            "caption": caption[:300],
            "likes": likes,
            "comments": comments,
            "engagement": likes + comments,
            "time": time_ago,
            "hashtags": ", ".join(hashtags[:5]),
            "content_type": content_type,
        })

    if not posts:
        return None

    return {
        "account": account,
        "display_name": display_name,
        "followers": followers,
        "posts_count": posts_count,
        "bio": bio,
        "posts": sorted(posts, key=lambda p: p["engagement"], reverse=True),
    }


def parse_all_brands() -> list[dict]:
    """Parse all scraped brand markdown files."""
    files = [f for f in sorted(BRANDS_DIR.glob("*.md"))
             if not f.name.startswith("test-") and not f.name.startswith("search-")]

    brands = []
    for f in files:
        result = parse_brand(f)
        if result and result["posts"]:
            brands.append(result)
            logger.info(f"  @{result['account']}: {result['followers']} followers, {len(result['posts'])} posts")

    logger.info(f"Total brands with post data: {len(brands)}")
    return brands


# ── Step 3: Build Excel ──────────────────────────────────────────────────────

def build_analysis_excel(brands: list[dict], output_path: Path) -> None:
    """Create 4-sheet competitor analysis Excel."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="405DE6", end_color="405DE6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_headers(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin

    # --- Sheet 1: Brand Overview ---
    ws1 = wb.active
    ws1.title = "브랜드 개요"
    headers1 = ["브랜드", "계정", "팔로워", "총 게시물", "스크래핑 수", "평균 좋아요", "평균 댓글", "최고 좋아요", "소개"]
    write_headers(ws1, headers1)

    for row, b in enumerate(brands, 2):
        avg_likes = sum(p["likes"] for p in b["posts"]) / len(b["posts"])
        avg_comments = sum(p["comments"] for p in b["posts"]) / len(b["posts"])
        max_likes = max(p["likes"] for p in b["posts"])
        vals = [b["display_name"], f"@{b['account']}", b["followers"], b["posts_count"],
                len(b["posts"]), round(avg_likes, 1), round(avg_comments, 1), max_likes, b["bio"][:100]]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=col, value=v)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 25
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["I"].width = 50

    # --- Sheet 2: Top Posts ---
    ws2 = wb.create_sheet("인기 게시물 TOP50")
    headers2 = ["순위", "브랜드", "좋아요", "댓글", "게시 시점", "콘텐츠 유형", "캡션 요약", "해시태그"]
    write_headers(ws2, headers2)

    all_posts = []
    for b in brands:
        for p in b["posts"]:
            all_posts.append({**p, "brand": b["display_name"], "account": b["account"]})
    all_posts.sort(key=lambda x: x["engagement"], reverse=True)

    gold_fill = PatternFill(start_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", fill_type="solid")

    for rank, p in enumerate(all_posts[:50], 1):
        vals = [rank, f"{p['brand']} (@{p['account']})", p["likes"], p["comments"],
                p["time"], p["content_type"], p["caption"][:150], p["hashtags"]]
        for col, v in enumerate(vals, 1):
            cell = ws2.cell(row=rank + 1, column=col, value=v)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if rank == 1:
                cell.fill = gold_fill
            elif rank == 2:
                cell.fill = silver_fill
            elif rank == 3:
                cell.fill = bronze_fill

    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["G"].width = 60
    ws2.column_dimensions["H"].width = 30

    # --- Sheet 3: Content Type Analysis ---
    ws3 = wb.create_sheet("콘텐츠 유형별 분석")
    headers3 = ["콘텐츠 유형", "게시물 수", "평균 좋아요", "평균 댓글", "최고 좋아요 게시물"]
    write_headers(ws3, headers3)

    type_stats = {}
    for p in all_posts:
        ct = p["content_type"]
        if ct not in type_stats:
            type_stats[ct] = {"count": 0, "likes": [], "comments": [], "best": "", "best_likes": 0}
        type_stats[ct]["count"] += 1
        type_stats[ct]["likes"].append(p["likes"])
        type_stats[ct]["comments"].append(p["comments"])
        if p["likes"] > type_stats[ct]["best_likes"]:
            type_stats[ct]["best_likes"] = p["likes"]
            type_stats[ct]["best"] = f"{p['brand']}: {p['caption'][:80]}"

    for row, (ct, stats) in enumerate(
        sorted(type_stats.items(), key=lambda x: sum(x[1]["likes"]) / len(x[1]["likes"]), reverse=True), 2
    ):
        avg_l = sum(stats["likes"]) / len(stats["likes"])
        avg_c = sum(stats["comments"]) / len(stats["comments"])
        vals = [ct, stats["count"], round(avg_l, 1), round(avg_c, 1), stats["best"]]
        for col, v in enumerate(vals, 1):
            cell = ws3.cell(row=row, column=col, value=v)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["E"].width = 50

    # --- Sheet 4: All posts per brand ---
    ws4 = wb.create_sheet("브랜드별 전체 포스트")
    headers4 = ["브랜드", "좋아요", "댓글", "게시 시점", "유형", "캡션", "해시태그"]
    write_headers(ws4, headers4)

    row = 2
    for b in brands:
        for p in b["posts"]:
            vals = [f"@{b['account']}", p["likes"], p["comments"], p["time"],
                    p["content_type"], p["caption"][:200], p["hashtags"]]
            for col, v in enumerate(vals, 1):
                cell = ws4.cell(row=row, column=col, value=v)
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1

    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["F"].width = 60
    ws4.column_dimensions["G"].width = 30

    wb.save(output_path)
    logger.info(f"Analysis Excel saved: {output_path}")


# ── Step 4: Save insights for content planner ────────────────────────────────

def save_competitor_insights(brands: list[dict]) -> Path:
    """Save top-performing content insights for the content planner to consume."""
    all_posts = []
    for b in brands:
        for p in b["posts"]:
            all_posts.append({**p, "brand": b["display_name"], "account": b["account"]})
    all_posts.sort(key=lambda x: x["engagement"], reverse=True)

    # Top 20 posts
    top_posts = []
    for p in all_posts[:20]:
        top_posts.append({
            "brand": p["brand"],
            "likes": p["likes"],
            "comments": p["comments"],
            "content_type": p["content_type"],
            "caption_snippet": p["caption"][:200],
            "hashtags": p["hashtags"],
        })

    # Content type performance
    type_perf = {}
    for p in all_posts:
        ct = p["content_type"]
        if ct not in type_perf:
            type_perf[ct] = {"count": 0, "total_likes": 0, "total_comments": 0}
        type_perf[ct]["count"] += 1
        type_perf[ct]["total_likes"] += p["likes"]
        type_perf[ct]["total_comments"] += p["comments"]

    for ct in type_perf:
        n = type_perf[ct]["count"]
        type_perf[ct]["avg_likes"] = round(type_perf[ct]["total_likes"] / n, 1)
        type_perf[ct]["avg_comments"] = round(type_perf[ct]["total_comments"] / n, 1)

    insights = {
        "generated_at": datetime.now().isoformat(),
        "brands_analyzed": len(brands),
        "total_posts": len(all_posts),
        "top_posts": top_posts,
        "content_type_performance": type_perf,
        "brand_summary": [
            {
                "account": b["account"],
                "display_name": b["display_name"],
                "followers": b["followers"],
                "avg_likes": round(sum(p["likes"] for p in b["posts"]) / len(b["posts"]), 1),
                "top_content_types": list(set(p["content_type"] for p in b["posts"][:5])),
            }
            for b in brands
        ],
    }

    TMP_DIR.mkdir(parents=True, exist_ok=True)
    with open(INSIGHTS_PATH, "w", encoding="utf-8") as f:
        json.dump(insights, f, ensure_ascii=False, indent=2)

    logger.info(f"Competitor insights saved: {INSIGHTS_PATH}")
    return INSIGHTS_PATH


# ── Step 5: Email ────────────────────────────────────────────────────────────

def get_gmail_service():
    """Build Gmail API service using OAuth2 credentials."""
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    scopes = ["https://www.googleapis.com/auth/gmail.send"]
    creds = None

    if os.path.exists(GMAIL_TOKEN_PATH):
        creds = Credentials.from_authorized_user_file(GMAIL_TOKEN_PATH, scopes)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(GMAIL_CREDENTIALS_PATH):
                logger.error(f"Gmail OAuth credentials not found: {GMAIL_CREDENTIALS_PATH}")
                return None
            flow = InstalledAppFlow.from_client_secrets_file(GMAIL_CREDENTIALS_PATH, scopes)
            creds = flow.run_local_server(port=0)

        os.makedirs(os.path.dirname(GMAIL_TOKEN_PATH), exist_ok=True)
        with open(GMAIL_TOKEN_PATH, "w") as f:
            f.write(creds.to_json())

    return build("gmail", "v1", credentials=creds)


def send_analysis_email(excel_path: Path, brands_count: int, posts_count: int) -> bool:
    """Send competitor analysis Excel to se.heo@orbiters.co.kr."""
    try:
        service = get_gmail_service()
        if not service:
            logger.error("Gmail service unavailable — skipping email")
            return False

        today = date.today().strftime("%Y-%m-%d")
        week_num = (date.today().day - 1) // 7 + 1
        week_label = f"{date.today().month}월_W{week_num}"

        subject = f"[경쟁사 IG 분석] {week_label} 일본 유아 브랜드 인스타 분석 ({today})"

        body_html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px;">
            <h2 style="color: #405DE6;">📊 일본 유아 브랜드 인스타그램 경쟁사 분석</h2>
            <table style="border-collapse: collapse; width: 100%; margin: 16px 0;">
                <tr><td style="padding: 8px; border: 1px solid #ddd; font-weight: bold;">분석 기간</td>
                    <td style="padding: 8px; border: 1px solid #ddd;">{week_label} ({today})</td></tr>
                <tr><td style="padding: 8px; border: 1px solid #ddd; font-weight: bold;">분석 브랜드</td>
                    <td style="padding: 8px; border: 1px solid #ddd;">{brands_count}개 브랜드</td></tr>
                <tr><td style="padding: 8px; border: 1px solid #ddd; font-weight: bold;">총 게시물</td>
                    <td style="padding: 8px; border: 1px solid #ddd;">{posts_count}개 포스트</td></tr>
            </table>
            <h3 style="color: #333;">📋 엑셀 시트 구성</h3>
            <ul>
                <li><b>브랜드 개요</b> — 팔로워, 평균 좋아요/댓글, 최고 좋아요</li>
                <li><b>인기 게시물 TOP50</b> — 전체 브랜드 인기 게시물 순위</li>
                <li><b>콘텐츠 유형별 분석</b> — 어떤 유형이 반응이 좋은지</li>
                <li><b>브랜드별 전체 포스트</b> — 각 브랜드 상세 데이터</li>
            </ul>
            <p>첨부된 엑셀을 참고하여 이번 주 콘텐츠 기획에 활용해 주세요 ✅</p>
            <p style="color: #888; font-size: 12px;">이 리포트는 매주 금요일 14:00에 자동 생성됩니다.</p>
        </div>
        """

        msg = MIMEMultipart("mixed")
        msg["Subject"] = subject
        msg["From"] = GMAIL_SENDER
        msg["To"] = SEEUN_EMAIL
        msg.attach(MIMEText(body_html, "html", "utf-8"))

        with open(excel_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{excel_path.name}"')
        msg.attach(part)

        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
        result = service.users().messages().send(userId="me", body={"raw": raw}).execute()

        logger.info(f"Email sent to {SEEUN_EMAIL} (Message ID: {result.get('id')})")
        return True

    except Exception as e:
        logger.error(f"Email send failed: {e}")
        return False


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Weekly IG competitor brand analysis")
    parser.add_argument("--dry-run", action="store_true", help="Skip email sending")
    parser.add_argument("--skip-scrape", action="store_true", help="Use cached scraped data")
    parser.add_argument("--output", type=str, default=None, help="Custom Excel output path")
    args = parser.parse_args()

    logger.info("=== Weekly IG Competitor Analysis ===")

    # Step 1: Scrape brands
    if not args.skip_scrape:
        logger.info("Step 1: Scraping brand profiles via Firecrawl...")
        scrape_brands(TARGET_BRANDS)
    else:
        logger.info("Step 1: Skipped (using cached data)")

    # Step 2: Parse all brands
    logger.info("Step 2: Parsing scraped data...")
    brands = parse_all_brands()

    if not brands:
        logger.error("No brand data found. Check .firecrawl/ig-baby-brands/ directory.")
        return

    # Step 3: Build Excel
    ts = datetime.now().strftime("%Y%m%d")
    output_path = Path(args.output) if args.output else TMP_DIR / f"jp_baby_brands_ig_analysis_{ts}.xlsx"
    logger.info("Step 3: Building analysis Excel...")
    build_analysis_excel(brands, output_path)

    total_posts = sum(len(b["posts"]) for b in brands)

    # Step 4: Save insights for content planner
    logger.info("Step 4: Saving competitor insights...")
    save_competitor_insights(brands)

    # Step 5: Email
    if not args.dry_run:
        logger.info("Step 5: Sending email...")
        send_analysis_email(output_path, len(brands), total_posts)
    else:
        logger.info("Step 5: Dry-run — email skipped")

    print(f"\n{'='*55}")
    print(f"  Competitor Analysis Complete")
    print(f"{'='*55}")
    print(f"  Brands: {len(brands)}")
    print(f"  Posts:  {total_posts}")
    print(f"  Excel:  {output_path}")
    print(f"  Insights: {INSIGHTS_PATH}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
