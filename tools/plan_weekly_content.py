"""
WAT Tool: Generate weekly content plan (20 ideas) as Excel file.
Pipeline: Firecrawl trend scraping + competitor insights -> Claude analysis -> Excel output.

Sources: bbox IG, Pigeon IG, JP parenting IG hashtags, Twitter, MamaStar, Ameblo
         + competitor_insights.json (from weekly_ig_competitor_analysis.py)
Output: .tmp/weekly_content_plan_YYYYMMDD.xlsx

Usage:
    python tools/plan_weekly_content.py                              # 20 ideas (tips:8, brand:6, k_babyfood:3, knowledge:3)
    python tools/plan_weekly_content.py --distribution "tips:8,brand:6,k_babyfood:3,knowledge:3"  # explicit
    python tools/plan_weekly_content.py --dry-run                    # scrape trends only, no Excel
"""

import os
import sys
import re
import json
import time
import argparse
import logging
from datetime import datetime
from pathlib import Path

# Fix Windows console encoding for Korean/Japanese output
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)

import anthropic
from firecrawl import FirecrawlApp
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TMP_DIR = Path(__file__).parent.parent / ".tmp"
TRENDS_CACHE = TMP_DIR / "weekly_trends_raw.json"
INSIGHTS_PATH = TMP_DIR / "competitor_insights.json"
MODEL = "claude-sonnet-4-20250514"
RATE_LIMIT_DELAY = 3

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Color palette ────────────────────────────────────────────────────────────

NAVY = "1F3864"
SLATE_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
OFF_WHITE = "F5F7FA"
DARK_GRAY = "343A40"
MID_GRAY = "6C757D"

CATEGORY_COLORS = {
    "tips": "C8E6C9",       # green light
    "brand": "BBDEFB",      # blue light
    "k_babyfood": "F8BBD0", # pink light
    "knowledge": "E1BEE7",  # purple light
}
CATEGORY_LABELS = {
    "tips": "육아 정보",
    "brand": "브랜드 제품",
    "k_babyfood": "K-유아식",
    "knowledge": "육아 지식",
}

# ── Trend Sources ────────────────────────────────────────────────────────────

INSTAGRAM_ACCOUNTS = [
    # bbox (picnob.com — picuki.com is dead since 2026-03)
    "https://www.picnob.com/profile/bboxforkidsjapan/",
    # Pigeon
    "https://www.picnob.com/profile/pigeon_official.jp/",
]

INSTAGRAM_HASHTAGS = [
    "https://inflact.com/profiles/tag/%E8%82%B2%E5%85%90/",           # 育児
    "https://inflact.com/profiles/tag/%E3%83%9E%E3%83%9E/",           # ママ
    "https://inflact.com/profiles/tag/%E5%AD%90%E8%82%B2%E3%81%A6/",  # 子育て
    "https://inflact.com/profiles/tag/%E9%9B%A2%E4%B9%B3%E9%A3%9F/",  # 離乳食
    "https://inflact.com/profiles/tag/%E8%82%B2%E5%85%90%E3%81%82%E3%82%8B%E3%81%82%E3%82%8B/",  # 育児あるある
]

TWITTER_URLS = [
    "https://nitter.poast.org/search?f=tweets&q=%23%E8%82%B2%E5%85%90%E3%81%82%E3%82%8B%E3%81%82%E3%82%8B",
    "https://nitter.poast.org/search?f=tweets&q=%23%E8%82%B2%E5%85%90%E6%BC%AB%E7%94%BB",
    "https://nitter.poast.org/search?f=tweets&q=%23%E3%83%AF%E3%83%B3%E3%82%AA%E3%83%9A%E8%82%B2%E5%85%90",
]

COMMUNITY_URLS = [
    "https://mamastar.jp/bbs/ranking",
    "https://blogger.ameba.jp/genres/baby",
]


# ── Scraping ─────────────────────────────────────────────────────────────────

def scrape_page(app: FirecrawlApp, url: str, retries: int = 2) -> str:
    """Scrape a single page via Firecrawl, return markdown."""
    for attempt in range(retries):
        try:
            result = app.scrape(
                url,
                formats=["markdown"],
                wait_for=3000,
                only_main_content=True,
                timeout=60000,
            )
            md = ""
            if hasattr(result, "markdown"):
                md = result.markdown or ""
            elif isinstance(result, dict):
                md = result.get("markdown", "")
            logger.info(f"  -> {len(md)} chars from {url[:70]}")
            return md
        except Exception as e:
            logger.warning(f"  Attempt {attempt+1}/{retries} failed for {url[:70]}: {e}")
            if attempt < retries - 1:
                time.sleep(RATE_LIMIT_DELAY * (attempt + 1))
    return ""


def scrape_all_trends(app: FirecrawlApp, max_pages: int) -> list[dict]:
    """Scrape trends from all sources, return list of trend items."""
    all_items = []

    # Instagram accounts (bbox, Pigeon)
    logger.info("Scraping Instagram accounts (bbox, Pigeon)...")
    for url in INSTAGRAM_ACCOUNTS[:max_pages]:
        md = scrape_page(app, url)
        if md:
            items = _parse_instagram(md, url)
            all_items.extend(items)
        time.sleep(RATE_LIMIT_DELAY)

    # Instagram hashtags
    logger.info("Scraping Instagram hashtags...")
    for url in INSTAGRAM_HASHTAGS[:max_pages]:
        md = scrape_page(app, url)
        if md:
            items = _parse_instagram(md, url)
            all_items.extend(items)
        time.sleep(RATE_LIMIT_DELAY)

    # Twitter
    logger.info("Scraping Twitter/X...")
    for url in TWITTER_URLS[:max_pages]:
        md = scrape_page(app, url)
        if md:
            items = _parse_twitter(md, url)
            all_items.extend(items)
        time.sleep(RATE_LIMIT_DELAY)

    # Communities
    logger.info("Scraping communities (MamaStar, Ameblo)...")
    for url in COMMUNITY_URLS[:max_pages]:
        md = scrape_page(app, url)
        if md:
            items = _parse_community(md, url)
            all_items.extend(items)
        time.sleep(RATE_LIMIT_DELAY)

    # Deduplicate
    seen = set()
    unique = []
    for item in all_items:
        key = item.get("snippet", "")[:80]
        if key and key not in seen:
            seen.add(key)
            unique.append(item)

    logger.info(f"Total trends: {len(all_items)} -> Unique: {len(unique)}")
    return unique


def _is_japanese(text: str) -> bool:
    return bool(re.search(r"[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF]", text))


def _parse_instagram(md: str, url: str) -> list[dict]:
    items = []
    desc_pattern = re.compile(r"([^\n]{20,}(?:#\w+[^\n]*))", re.MULTILINE)
    for desc in desc_pattern.findall(md):
        hashtags = [f"#{h}" for h in re.findall(r"#(\w+)", desc) if _is_japanese(h)]
        if not hashtags and not _is_japanese(desc):
            continue
        items.append({
            "source": "instagram",
            "url": url,
            "snippet": re.sub(r"\s+", " ", desc).strip()[:400],
            "hashtags": hashtags[:10],
        })
    # Fallback: image alt texts
    if not items:
        for alt, _ in re.findall(r"\!\[([^\]]{10,})\]\((https?://[^\)]+)\)", md):
            if _is_japanese(alt):
                items.append({
                    "source": "instagram",
                    "url": url,
                    "snippet": alt.strip()[:400],
                    "hashtags": [],
                })
    return items[:20]


def _parse_twitter(md: str, url: str) -> list[dict]:
    items = []
    for block in re.split(r"\n(?=[@\[])", md):
        content = re.sub(r"\[([^\]]*)\]\([^\)]+\)", r"\1", block)
        content = re.sub(r"\!\[[^\]]*\]\([^\)]+\)", "", content)
        content = re.sub(r"[#*_`>]+", "", content)
        content = re.sub(r"\s+", " ", content).strip()
        if len(content) < 30:
            continue
        hashtags = [f"#{h}" for h in re.findall(r"#(\w+)", block) if _is_japanese(h)]
        items.append({
            "source": "twitter",
            "url": url,
            "snippet": content[:400],
            "hashtags": hashtags,
        })
    return items[:20]


def _parse_community(md: str, url: str) -> list[dict]:
    items = []
    for title, link in re.findall(r"\[([^\]]{10,})\]\((https?://[^\)]+)\)", md):
        if _is_japanese(title) and len(title) > 5:
            items.append({
                "source": "community",
                "url": link,
                "snippet": title.strip()[:400],
                "hashtags": [],
            })
    return items[:20]


# ── Claude Planning ──────────────────────────────────────────────────────────

SYSTEM_PROMPT = """あなたは日本の育児用品ブランド「Grosmimi」のInstagramコンテンツストラテジストです。
ターゲット: 6〜24ヶ月の子を育てるママ。週20件の情報カルーセル企画案を作成します。

## カテゴリ（4種類のみ — 配分厳守）
1. tips — 育児情報: 月齢別ガイド、季節対策、便利グッズ比較、方法解説
2. brand — ブランド製品: Grosmimi スペック・使い方・他社比較・お手入れ
3. k_babyfood — K-離乳食: 韓国式離乳食レシピ、食材の違い、作り方
4. knowledge — 育児知識: 栄養素、安全基準、素材の違い等の専門知識

## 絶対禁止カテゴリ（トーン領域分離）
- ミーム / 共感あるある / ぼやき / 独白 = Twitter領域。Instagram企画案では絶対作らない
- 「わかる〜」「それな」系の共感投稿は禁止
- Instagramは「保存したい」「友達にシェアしたい」と思わせる情報メディアトーン

## トーンガイド（「隣の街のお姉さんが教える育児情報」）
情報伝達性 + 親しみ の中間。堅いビジネス×、過度なカジュアル×。

OK 表現（情報+親しみ）:
- 語彙: 日焼け止め / お水 / お出かけ / お顔
- 語尾(JP): 〜のがおすすめ / 〜が基本 / 〜が目安 / 〜が◎ / 〜が安心
- 語尾(KO): "~해주세요" "~좋아요" "~정도가 적당해요" "~안심"
- 表現: こまめに / ちょうどいい / 目安 / 安心 / 딱 좋아요

【絶対NG語彙 — トピック名・タイトル・本文・キャプションすべてで使用禁止】
NG① 堅い・ビジネス・教科書語:
- 紫外線対策 → 日焼け止め
- 水分補給 → お水 / 水分
- 外出 → お出かけ
- 顔 (単独) → お顔
- 黄金ルール → ◯◯のコツ / おすすめ
- 必須 → 大切 / おすすめ
- 推奨 → おすすめ
- 適切に → ちょうどいい
- 基本原則 → 基本 / 目安
- 〜すること / 〜の必要がある / 〜を推奨 → 〜のがおすすめ / 〜が目安
- 韓国語: "~할 것" → "~해주세요" / "권장·필수" → "추천·중요" / "기본 원칙" → "기본·기준"

NG② 過度にカジュアル（友達トーン）:
- 〜してね / 〜していこう / 〜してあげて / 〜してあげて♡
- 韓国語: "~해봐~" "~하자!" "~하기♡"
- 過剰絵文字・テンション系

※ トピック名 (topic / topic_ko) に NG語彙が入っていたら全企画案を作り直すレベルの違反。
※ NG語彙が3個以上含まれた企画案は失格扱い。代替語彙 (OK表現) を必ず使うこと。

## 必須ルール
- 漢字比率を下げる。ひらがな + やわらかい外来語を活用
- オノマトペ（ポンポン・シュッ・トントン等）は カルーセル全体で 0〜1個 のみ
- 韓国語も「톡톡·칙·박박」のような擬声語の乱用禁止。JPと対応する1個のみ
- 「黄金ルール」「必須」「推奨」「適切に」など教科書・マーケティング表現は禁止
- 医療的断定・論争系育児アドバイス・競合批判は禁止

## カルーセル5スライド標準構造（必須）
| Slide | 役割 | 内容 |
| S1表紙 | 後킹 | メイン写真 + 텍스트オーバーレイ（subtitle + title）|
| S2훅 | 問題提起 | 「○○知らずに○○してない？」「これって正解？」式の疑問 |
| S3情報1 | データ | 月齢別比較表 / 一目でわかる表 / 写真比較 |
| S4情報2 | 実行 | ステップガイド / チェックリスト / NG vs OK |
| S5 CTA | アクション | 保存・フォロー・プロフィール誘導。代表CTAキャッチコピー1個 |

「テキストだけ並べたスライド」禁止。各スライドに写真/図解/比較表を含めること。

## サムネイル（S1表紙）必須フォーマット
- subtitle (소제목): カテゴリタグ。背景色付きの短いラベル。例:「育児のコツ」「ママのそれ知りたかった！」「知っておきたい」「意外と知らない！」
- title (대제목): 核心キーワード2〜3単語、大きく太字。1単語をピンク/コーラル色強調
  例:「赤ちゃんの**お昼寝**チャート」「**夜泣き**対策6選」「**マグ投げ期**、どう乗り越えた?」

## CTA キャッチコピー候補5個（必須）
S5 CTAとキャプション末で使用。3トーンのみ:
1. 호기심 + 프로필 유도 (액션=プロフィール): 例「"結局どれがいいの…？" の答え、プロフィールに全部📌」
3. 저장 유도 (액션=저장+팔로우): 例「迷ったら保存🔖 月齢別ガイドはフォローで」
4. 공감 소속감 (액션=팔로우 or 정보형이면 공유): 例「同じ悩みのママ友いたら、そっとシェアしてあげて🤍」

メインアクション = フォロー OR プロフィールリンク（基本）。
クーポンコード・DM・サイト外部訪問など 他の単独アクションは禁止（보조だけ可）。
5個分布: 1톤 2個 / 3톤 1〜2個 / 4톤 1〜2個。
カテゴリ別 대표1個 선정ルール: brand→1톤 / tips・knowledge→3톤 / Q&A공감→4톤 / 친구공유 정보형→4톤(공유 변형) 可。
各候補は JP原文 + KO翻訳 併記。

## 競合 출처 활용 (베끼기 금지)
- 競合 게시물(ピジョン・b.box・Combi 등)은 출발점일 뿐。그대로 옮겨 적기 금지
- 반드시 보강 리서치 추가: 厚労省 / 小児科学会 / 公的기관 가이드라인 / 학술 논문 / 다른 브랜드 사례 / K-육아 인사이트
- Grosmimi 자체 가치(PPSU・역지밸브・360° 핸들・K-육아) 녹여 차별화
- 출처는 1차(경쟁사) + 보강(공식기관/논문/다른브랜드) 둘 다 명시

## 출력 필드
- category: tips / brand / k_babyfood / knowledge
- topic: トピック (日本語) / topic_ko: 한국어 번역
- slides[5]: 各 {title_jp, title_ko, body_jp, body_ko, visual} — body는 3〜5행 ・불릿
- subtitle_jp / subtitle_ko: S1 표지 소제목 (배경라벨용)
- title_jp / title_ko: S1 표지 대제목 (강조 키워드 **bold**)
- image_concept: 전체 비주얼 방향 1〜2문
- emphasis: 디자인 핵심 메시지 1〜2문
- caption: 본문 5〜8行 (후크1 + 팩트1〜2 + ✔체크3〜4 + CTA1). 해시태그는 본문 외 별도. ※ 본문 행수 8행 초과 X
- caption_ko: 본문 5〜8행 한국어 번역 (해시태그 별도)
- hashtags: 15〜25개. 「#グロミミ #grosmimi #ストローマグ #スマートマグ」 통일 포함
- cta_candidates[5]: 각 {tone(1/3/4), action(profile/save_follow/follow_or_share), jp, ko}
- cta_representative: 5개 중 대표 1개 (S5 CTA 대제목용) {tone, jp, ko}
- source_primary: 경쟁사 게시물 출처 (브랜드명 + 게시물 주제)
- source_supplementary: 보강 출처 (공식기관/논문/다른브랜드)
- trend_ref: 인스피레이션 트렌드 키워드"""


def _load_competitor_insights() -> str:
    """Load competitor insights JSON and format as text for Claude prompt."""
    if not INSIGHTS_PATH.exists():
        logger.info("No competitor insights found — skipping")
        return ""

    try:
        with open(INSIGHTS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        logger.warning(f"Failed to load competitor insights: {e}")
        return ""

    parts = []
    parts.append("## 競合ブランド分析（直近の人気コンテンツ）\n")

    # Content type performance
    type_perf = data.get("content_type_performance", {})
    if type_perf:
        parts.append("### コンテンツタイプ別パフォーマンス")
        for ct, stats in sorted(type_perf.items(), key=lambda x: x[1].get("avg_likes", 0), reverse=True):
            parts.append(f"- {ct}: 平均いいね {stats.get('avg_likes', 0)}, "
                         f"平均コメント {stats.get('avg_comments', 0)}, "
                         f"投稿数 {stats.get('count', 0)}")
        parts.append("")

    # Top posts
    top_posts = data.get("top_posts", [])[:10]
    if top_posts:
        parts.append("### 人気投稿TOP10")
        for i, p in enumerate(top_posts, 1):
            parts.append(f"{i}. [{p.get('brand', '')}] いいね{p.get('likes', 0)} / "
                         f"コメント{p.get('comments', 0)} / タイプ: {p.get('content_type', '')}")
            parts.append(f"   内容: {p.get('caption_snippet', '')[:120]}")
        parts.append("")

    return "\n".join(parts)


def _call_claude_for_plans(client: anthropic.Anthropic, prompt: str, batch_size: int) -> list[dict]:
    """Call Claude once and parse JSON plans. Returns list of plan dicts."""
    for attempt in range(3):
        try:
            response = client.messages.create(
                model=MODEL,
                max_tokens=16384,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": prompt}],
            )
            text = response.content[0].text

            if "```json" in text:
                json_str = text.split("```json")[1].split("```")[0].strip()
            elif "```" in text:
                json_str = text.split("```")[1].split("```")[0].strip()
            else:
                json_str = text.strip()

            result = json.loads(json_str)
            plans = result.get("plans", [])
            logger.info(f"Claude returned {len(plans)} plan(s)")
            return plans

        except json.JSONDecodeError as e:
            logger.warning(f"JSON parse error (attempt {attempt+1}/3): {e}")
            if attempt < 2:
                prompt += "\n\n重要: 有効なJSONのみを出力してください。captionは200文字以内に短縮してください。"
                time.sleep(RATE_LIMIT_DELAY)
        except Exception as e:
            logger.error(f"Claude API error (attempt {attempt+1}/3): {e}")
            if attempt < 2:
                time.sleep(2 * (attempt + 1))

    raise RuntimeError("Failed to get valid plans from Claude after all retries")


def generate_plans(client: anthropic.Anthropic, trends: list[dict], count: int,
                   distribution: dict | None = None) -> list[dict]:
    """Use Claude to generate content plan ideas from trends + competitor insights.
    Splits into batches of 5 to avoid token overflow."""
    # Summarize trends
    trends_text = ""
    for i, t in enumerate(trends[:60]):
        trends_text += (
            f"{i+1}. [{t['source']}] {t['snippet'][:200]}\n"
            f"   Tags: {', '.join(t.get('hashtags', []))}\n\n"
        )

    # Load competitor insights
    competitor_text = _load_competitor_insights()

    # Split distribution into batches of 3 (carousel 5-slide responses overflow at 5)
    BATCH_SIZE = 3
    batches = []  # list of (batch_count, batch_dist_instruction)

    if distribution:
        for cat, cat_count in distribution.items():
            label = CATEGORY_LABELS.get(cat, cat)
            remaining = cat_count
            while remaining > 0:
                batch_n = min(remaining, BATCH_SIZE)
                batch_dist = (
                    f"【絶対厳守】このバッチで作成する {batch_n}個の企画案は、すべて category=\"{cat}\" のみ。\n"
                    f"- 「{label}」({cat}) 専用バッチ。他カテゴリ ({', '.join(k for k in CATEGORY_LABELS if k != cat)}) は絶対に作らない\n"
                    f"- 全 {batch_n}件の plan の category フィールドは必ず文字列 \"{cat}\""
                )
                batches.append((batch_n, (batch_dist, cat)))
                remaining -= batch_n
    else:
        for start in range(0, count, BATCH_SIZE):
            batch_n = min(BATCH_SIZE, count - start)
            batch_dist = "4つのカテゴリ（tips, brand, k_babyfood, knowledge）をバランスよく配分"
            batches.append((batch_n, (batch_dist, None)))

    all_plans = []
    for batch_idx, (batch_count, dist_pair) in enumerate(batches):
        dist_instruction, expected_cat = dist_pair if isinstance(dist_pair, tuple) else (dist_pair, None)
        logger.info(f"Generating batch {batch_idx+1}/{len(batches)} ({batch_count} plans, cat={expected_cat or 'any'})...")

        prompt = f"""以下のトレンドと競合分析を元に、{batch_count}個の Grosmimi Instagram カルーセル企画案を作成してください。

## トレンドデータ
{trends_text}

{competitor_text}

## 競合 출처 활용 룰 (必)
- 競合 게시물 그대로 베끼기 금지。반드시 厚労省 / 소아과학회 / 학술 논문 / 다른 브랜드(b.box, Combi, Richell 등) / K-육아 인사이트로 보강
- 출처는 source_primary (경쟁사) + source_supplementary (보강) 둘 다 채울 것
- Grosmimi 차별점 (PPSU / 역지밸브 / 360° 핸들 / K-육아) 1개 이상 녹여 차별화

## 작성 룰
- {dist_instruction}
- 톤: 「동네 언니가 알려주는 육아 정보」. 「必須/推奨/黄金ルール/適切に」 등 교과서 표현 금지. 「〜してね/〜してあげて/♡」 등 과도 친구톤 금지. 「〜のがおすすめ/〜が目安/〜が◎/〜が安心」 채택
- 漢字 비율 ↓, 히라가나·やわらかい外来語 ↑
- オノマトペ는 카루셀 전체에서 0〜1個만
- ミーム/あるある/ぼやき/공감/わかる系 = 트위터 영역 절대 금지
- 5枚 카루셀 (S1 표지 / S2 훅 / S3 정보1 / S4 정보2 / S5 CTA) 필수. 각 slide의 body_jp/body_ko는 3〜5행 ・불릿
- S1 표지에는 subtitle_jp/ko (소제목 라벨) + title_jp/ko (대제목, 핵심 키워드 1개 **bold**) 별도 채우기
- 캡션은 5〜8行: 후크1행 + 팩트1〜2행 + ✔체크 3〜4개 + CTA1행. 200〜400文字 X (행수 기준). 해시태그는 캡션 끝
- 해시태그 15〜25개. 「#グロミミ」「#grosmimi」「#ストローマグ」「#スマートマグ」 4개 통일 포함
- CTA 캐치프레이즈 候補 5個 필수: tone={{1=호기심+프로필, 3=저장+팔로우, 4=공감+팔로우/공유}}, 분포는 1톤 2 / 3톤 1〜2 / 4톤 1〜2. 메인액션은 follow/profile (DM/쿠폰 단독 X). 각 JP+KO 병기
- cta_representative 1개 선정: brand→1톤 / tips·knowledge→3톤 / 공감형→4톤

## 출력 형식 (JSON 만, 余計한 텍스트 X)
```json
{{
  "plans": [
    {{
      "category": "tips",
      "topic": "トピック",
      "topic_ko": "주제 한국어 번역",
      "subtitle_jp": "知っておきたい",
      "subtitle_ko": "알아두면 좋은",
      "title_jp": "赤ちゃんの**お昼寝**チャート",
      "title_ko": "아기 **낮잠** 차트",
      "slides": [
        {{"title_jp":"S1表紙(サムネ)","title_ko":"S1 표지","body_jp":"・○○○\\n・○○○","body_ko":"・○○○\\n・○○○","visual":"메인 사진 구상"}},
        {{"title_jp":"S2 훅","title_ko":"S2 훅","body_jp":"・疑問1\\n・疑問2","body_ko":"・의문1\\n・의문2","visual":"비주얼"}},
        {{"title_jp":"S3 情報1","title_ko":"S3 정보1","body_jp":"・データ1\\n・データ2","body_ko":"・데이터1\\n・데이터2","visual":"비교표/그래프"}},
        {{"title_jp":"S4 情報2","title_ko":"S4 정보2","body_jp":"・手順1\\n・手順2","body_ko":"・단계1\\n・단계2","visual":"체크리스트"}},
        {{"title_jp":"S5 CTA (대표 캐치 1個)","title_ko":"S5 CTA","body_jp":"・CTA本文","body_ko":"・CTA 본문","visual":"CTA 시각 요소"}}
      ],
      "image_concept": "全体ビジュアル方向 1〜2문",
      "emphasis": "디자인 핵심 메시지 1〜2문",
      "caption": "후크\\n팩트1〜2行\\n✔ポイント1\\n✔ポイント2\\n✔ポイント3\\nCTA 1行\\n\\n#グロミミ #grosmimi #ストローマグ #スマートマグ ...",
      "caption_ko": "후크\\n팩트1〜2행\\n✔포인트1\\n✔포인트2\\n✔포인트3\\nCTA 1행\\n\\n#그로미미 ...",
      "hashtags": ["#グロミミ","#grosmimi","#ストローマグ","#スマートマグ","..."],
      "cta_candidates": [
        {{"tone":1,"action":"profile","jp":"\\"結局どれがいいの…？\\" の答え、プロフィールに全部📌","ko":"\\"결국 뭐가 좋은데?\\" 답, 프로필에 다 모아뒀어요📌"}},
        {{"tone":1,"action":"profile","jp":"...","ko":"..."}},
        {{"tone":3,"action":"save_follow","jp":"迷ったら保存🔖 月齢別ガイドはフォローで","ko":"고민되면 저장🔖 월령별 가이드는 팔로우로"}},
        {{"tone":4,"action":"follow","jp":"同じ悩みのママ友いたら、そっとシェア🤍","ko":"같은 고민의 엄마 친구 있으면, 살짝 공유해줘🤍"}},
        {{"tone":4,"action":"share","jp":"...","ko":"..."}}
      ],
      "cta_representative": {{"tone":3,"jp":"迷ったら保存🔖 月齢別ガイドはフォローで","ko":"고민되면 저장🔖 월령별 가이드는 팔로우로"}},
      "source_primary": "ピジョン IG「赤ちゃんのおやつはいつから？」",
      "source_supplementary": "厚労省 授乳・離乳の支援ガイド + b.box 関連投稿",
      "trend_ref": "트렌드 키워드 + 출처 (1줄)"
    }}
  ]
}}
```"""

        batch_plans = _call_claude_for_plans(client, prompt, batch_count)

        # Distribution 강제: 기대 카테고리 외 plan은 카테고리 덮어쓰기 (drift 방지)
        if expected_cat:
            for p in batch_plans:
                if p.get("category") != expected_cat:
                    logger.warning(
                        f"Category drift detected: got '{p.get('category')}', forcing '{expected_cat}' "
                        f"(topic={p.get('topic','')[:40]})"
                    )
                    p["category"] = expected_cat

        all_plans.extend(batch_plans)

        if batch_idx < len(batches) - 1:
            time.sleep(RATE_LIMIT_DELAY)

    logger.info(f"Total plans generated: {len(all_plans)}")
    return all_plans


# ── Excel Output ─────────────────────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


COLUMNS = [
    ("#",                       5,  False),
    ("카테고리",                14, False),
    ("주제",                    28, True),
    ("주제(한국어)",            28, True),
    ("S1 표지",                 40, True),
    ("S2 훅",                   40, True),
    ("S3 정보1",                45, True),
    ("S4 정보2",                45, True),
    ("S5 CTA",                  40, True),
    ("이미지 구상",             40, True),
    ("강조 포인트",             28, True),
    ("캡션",                    50, True),
    ("캡션(한국어)",            50, True),
    ("해시태그",                35, True),
    ("CTA 캐치프레이즈 후보5",  50, True),
    ("출처(1차+보강)",          40, True),
    ("참고 트렌드",             28, True),
]


SLIDE_ROLES = ["표지", "훅", "정보1", "정보2", "CTA"]


def _format_slide_cell(slide: dict, role: str, extra_header: dict | None = None) -> str:
    """Render a single slide dict into a multi-line cell value.
    extra_header: 표지 슬라이드의 경우 subtitle/title 라벨 추가용."""
    if not slide and not extra_header:
        return ""
    slide = slide or {}
    title_jp = slide.get("title_jp", "").strip()
    title_ko = slide.get("title_ko", "").strip()
    body_jp = slide.get("body_jp", "").strip()
    body_ko = slide.get("body_ko", "").strip()
    visual = slide.get("visual", "").strip()

    parts = [f"[역할] {role}"]

    if extra_header:
        sub_jp = (extra_header.get("subtitle_jp") or "").strip()
        sub_ko = (extra_header.get("subtitle_ko") or "").strip()
        title_main_jp = (extra_header.get("title_jp") or "").strip()
        title_main_ko = (extra_header.get("title_ko") or "").strip()
        if sub_jp or sub_ko:
            parts.append("[소제목 라벨]")
            if sub_jp:
                parts.append(f"JP: {sub_jp}")
            if sub_ko:
                parts.append(f"KO: {sub_ko}")
        if title_main_jp or title_main_ko:
            parts.append("[대제목 (강조 키워드 **bold**)]")
            if title_main_jp:
                parts.append(f"JP: {title_main_jp}")
            if title_main_ko:
                parts.append(f"KO: {title_main_ko}")

    if title_jp or title_ko:
        parts.append("[슬라이드 타이틀]")
        if title_jp:
            parts.append(f"JP: {title_jp}")
        if title_ko:
            parts.append(f"KO: {title_ko}")
    if body_jp or body_ko:
        parts.append("[본문]")
        if body_jp:
            parts.append(f"JP:\n{body_jp}")
        if body_ko:
            parts.append(f"KO:\n{body_ko}")
    if visual:
        parts.append(f"[비주얼]\n{visual}")
    return "\n".join(parts)


def _format_cta_candidates(plan: dict) -> str:
    """CTA 5 후보 + 대표 1개를 한 셀에 렌더."""
    cands = plan.get("cta_candidates") or []
    rep = plan.get("cta_representative") or {}
    parts = []
    if rep:
        parts.append("[대표 CTA]")
        if rep.get("jp"):
            parts.append(f"JP: {rep.get('jp','')}")
        if rep.get("ko"):
            parts.append(f"KO: {rep.get('ko','')}")
        parts.append("")
    parts.append("[후보 5개]")
    tone_label = {1: "1.호기심+프로필", 3: "3.저장+팔로우", 4: "4.공감+팔로우/공유"}
    for i, c in enumerate(cands[:5], start=1):
        tone = c.get("tone")
        action = c.get("action", "")
        label = tone_label.get(tone, f"tone={tone}")
        parts.append(f"{i}. ({label} / action={action})")
        if c.get("jp"):
            parts.append(f"   JP: {c.get('jp','')}")
        if c.get("ko"):
            parts.append(f"   KO: {c.get('ko','')}")
    return "\n".join(parts)


def _format_sources(plan: dict) -> str:
    primary = (plan.get("source_primary") or "").strip()
    suppl = (plan.get("source_supplementary") or "").strip()
    parts = []
    if primary:
        parts.append(f"1차: {primary}")
    if suppl:
        parts.append(f"보강: {suppl}")
    return "\n".join(parts)


def build_excel(plans: list[dict], output_path: Path) -> None:
    """Write plans to a styled Excel file."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "주간 기획안"
    ws.sheet_view.showGridLines = False

    # ── Header row ───────────────────────────────────────────────────────
    for ci, (header, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Data rows ────────────────────────────────────────────────────────
    for ri, plan in enumerate(plans):
        row_idx = ri + 2
        cat_key = plan.get("category", "tips")
        cat_label = CATEGORY_LABELS.get(cat_key, cat_key)
        cat_bg = CATEGORY_COLORS.get(cat_key, OFF_WHITE)
        row_bg = OFF_WHITE if ri % 2 == 0 else WHITE

        slides = plan.get("slides") or []
        slide_cells = []
        for i in range(5):
            s = slides[i] if i < len(slides) else {}
            extra = plan if i == 0 else None  # S1 표지에만 subtitle/title 헤더 추가
            slide_cells.append(_format_slide_cell(s, SLIDE_ROLES[i], extra_header=extra))

        values = [
            ri + 1,
            cat_label,
            plan.get("topic", ""),
            plan.get("topic_ko", ""),
            slide_cells[0],
            slide_cells[1],
            slide_cells[2],
            slide_cells[3],
            slide_cells[4],
            plan.get("image_concept", ""),
            plan.get("emphasis", ""),
            plan.get("caption", ""),
            plan.get("caption_ko", ""),
            " ".join(plan.get("hashtags", [])),
            _format_cta_candidates(plan),
            _format_sources(plan),
            plan.get("trend_ref", ""),
        ]

        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=ci, value=val)
            _, _, wrap = COLUMNS[ci - 1]

            # Category column gets special coloring
            bg = cat_bg if ci == 2 else row_bg

            cell.font = Font(name="Calibri", size=10, color=DARK_GRAY)
            cell.fill = PatternFill(fill_type="solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal="center" if ci <= 2 else "left",
                vertical="top",
                wrap_text=wrap,
            )
            cell.border = thin_border()

        ws.row_dimensions[row_idx].height = 280

    # ── Summary row ──────────────────────────────────────────────────────
    summary_row = len(plans) + 3
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(
        name="Calibri", bold=True, size=10, color=NAVY
    )

    # Category counts
    from collections import Counter
    cat_counts = Counter(p.get("category", "?") for p in plans)
    summary_parts = []
    for key, label in CATEGORY_LABELS.items():
        cnt = cat_counts.get(key, 0)
        if cnt > 0:
            summary_parts.append(f"{label}: {cnt}")
    ws.cell(row=summary_row, column=2, value=" / ".join(summary_parts)).font = Font(
        name="Calibri", size=10, color=MID_GRAY
    )

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.cell(row=summary_row + 1, column=2, value=f"Generated: {generated_at}").font = Font(
        name="Calibri", size=9, color=MID_GRAY, italic=True
    )

    wb.save(output_path)
    logger.info(f"Excel saved: {output_path}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate weekly content plan as Excel")
    parser.add_argument("--count", type=int, default=20,
                        help="Number of content ideas (default 20)")
    parser.add_argument("--max-pages", type=int, default=3,
                        help="Max pages per source for scraping (default 3)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Scrape trends only, skip Claude + Excel")
    parser.add_argument("--output", type=str, default=None,
                        help="Custom output path for Excel file")
    parser.add_argument("--distribution", type=str,
                        default="tips:8,brand:6,k_babyfood:3,knowledge:3",
                        help="Category distribution. Default: tips:8,brand:6,k_babyfood:3,knowledge:3")
    args = parser.parse_args()

    TMP_DIR.mkdir(parents=True, exist_ok=True)

    # ── Step 1: Scrape trends ────────────────────────────────────────────
    fc_key = os.getenv("FIRECRAWL_API_KEY")
    if not fc_key:
        raise EnvironmentError("FIRECRAWL_API_KEY not found. Check .env file.")

    app = FirecrawlApp(api_key=fc_key)
    logger.info("Starting trend collection...")
    trends = scrape_all_trends(app, max_pages=args.max_pages)

    # Cache trends
    with open(TRENDS_CACHE, "w", encoding="utf-8") as f:
        json.dump({
            "scraped_at": datetime.now().isoformat(),
            "total_items": len(trends),
            "items": trends,
        }, f, ensure_ascii=False, indent=2)
    logger.info(f"Trends cached: {TRENDS_CACHE}")

    if args.dry_run:
        print(f"\n{'='*55}")
        print(f"  Dry run complete. {len(trends)} trends collected.")
        print(f"  Cached: {TRENDS_CACHE}")
        print(f"{'='*55}\n")
        return

    # ── Step 2: Generate plans with Claude ───────────────────────────────
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise EnvironmentError("ANTHROPIC_API_KEY not found. Check .env file.")

    client = anthropic.Anthropic(api_key=api_key)
    logger.info(f"Generating {args.count} content plans with Claude...")
    # Parse distribution if provided
    distribution = None
    if args.distribution:
        distribution = {}
        for part in args.distribution.split(","):
            k, v = part.strip().split(":")
            distribution[k.strip()] = int(v.strip())
        # Override count to match distribution total
        args.count = sum(distribution.values())

    plans = generate_plans(client, trends, args.count, distribution=distribution)

    # ── Step 3: Write Excel ──────────────────────────────────────────────
    ts = datetime.now().strftime("%Y%m%d")
    output_path = Path(args.output) if args.output else TMP_DIR / f"weekly_content_plan_{ts}.xlsx"
    build_excel(plans, output_path)

    # ── Summary ──────────────────────────────────────────────────────────
    print(f"\n{'='*55}")
    print(f"  Weekly Content Plan Generated")
    print(f"{'='*55}")
    print(f"  Ideas:  {len(plans)}")
    from collections import Counter
    cat_counts = Counter(p.get("category", "?") for p in plans)
    for key, label in CATEGORY_LABELS.items():
        cnt = cat_counts.get(key, 0)
        if cnt > 0:
            print(f"    {label}: {cnt}")
    print(f"  Output: {output_path}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
