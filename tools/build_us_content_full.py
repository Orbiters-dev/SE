"""Social Enricher — Creator Pool + Content Pool from Syncly sheets.

Joins Creators_updated + Output_updated, applies view/like thresholds,
removes blacklisted, groups content per creator, enriches via Apify,
and exports to Excel or syncs to PostgreSQL.

Usage:
    python tools/build_us_content_full.py
    python tools/build_us_content_full.py --dry-run
    python tools/build_us_content_full.py --enrich
    python tools/build_us_content_full.py --min-views 5000 --min-likes 50 --min-er 0.03
    python tools/build_us_content_full.py --pg-sync
    python tools/build_us_content_full.py --platform instagram --language en --since 2025-01-01
"""

import os, sys, re, json, io
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# Fix Windows cp949
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

DIR = Path(__file__).resolve().parent
PROJECT_ROOT = DIR.parent
sys.path.insert(0, str(DIR))

try:
    from env_loader import load_env
    load_env()
except Exception:
    pass

SHEET_ID = "1dIAhP8wCEdFulSAai3K-RoZTvLBIaWxAK7hzInBsF0o"
CACHE_DIR = PROJECT_ROOT / ".tmp" / "data_crawler" / "cache"
ENRICH_CACHE = PROJECT_ROOT / ".tmp" / "data_crawler" / "enrich_cache.json"
ILLEGAL_CHARS = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

# Apify actors (same as fetch_apify_content.py)
IG_SCRAPER = "apify/instagram-scraper"
TT_SCRAPER = "clockworks/free-tiktok-scraper"


# ── Helpers ──────────────────────────────────────────────────────────

def safe_num(val):
    """Parse numeric value from string (handles commas, blanks, #DIV/0!)."""
    if not val or val in ("", "#DIV/0!", "#REF!", "#N/A", "#VALUE!"):
        return 0
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0


def clean(val):
    """Clean cell value for Excel."""
    if val is None:
        return ""
    s = str(val).strip()
    return ILLEGAL_CHARS.sub("", s)


def load_cached_tab(tab_name):
    """Load tab from data_crawler cache. Returns list of dicts."""
    safe_tab = tab_name.replace(" ", "_")
    cache_file = CACHE_DIR / f"cache_{SHEET_ID}_{safe_tab}.json"
    if not cache_file.exists():
        print(f"  [WARN] Cache not found for '{tab_name}'. Run data_crawler first.")
        print(f"  Running: python tools/data_crawler.py --source sheet:{SHEET_ID} --sheet-name \"{tab_name}\" --dry-run")
        return []
    with open(cache_file, "r", encoding="utf-8") as f:
        raw = json.load(f)
    if not raw:
        return []
    headers = [h.strip() for h in raw[0]]
    rows = []
    for row in raw[1:]:
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        rows.append(dict(zip(headers, padded[:len(headers)])))
    return rows


def detect_platform_from_url(url):
    """Detect platform from profile/post URL."""
    if not url:
        return "unknown"
    url_lower = url.lower()
    if "tiktok" in url_lower:
        return "tiktok"
    if "instagram" in url_lower:
        return "instagram"
    return "unknown"


# ── Apify Enrichment ─────────────────────────────────────────────────

def load_enrich_cache():
    """Load cached post metrics from previous Apify runs."""
    if ENRICH_CACHE.exists():
        with open(ENRICH_CACHE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_enrich_cache(cache):
    """Save post metrics cache."""
    ENRICH_CACHE.parent.mkdir(parents=True, exist_ok=True)
    with open(ENRICH_CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)


def enrich_post_metrics(content_map, chunk_size=50):
    """Fetch per-post views/likes/comments via Apify. Caches results."""
    from apify_client import ApifyClient
    token = os.environ.get("APIFY_API_TOKEN")
    if not token:
        print("  [ERROR] APIFY_API_TOKEN not set. Skipping enrichment.")
        return {}

    client = ApifyClient(token)
    cache = load_enrich_cache()
    print(f"  Enrich cache: {len(cache):,} posts already cached")

    # Collect all unique post URLs by platform (dedup!)
    seen = set()
    ig_urls = []
    tt_urls = []
    for contents in content_map.values():
        for ct in contents:
            url = ct.get("post_url", "")
            # Re-fetch if cache entry is legacy (3-field, missing 'caption')
            cached = cache.get(url)
            is_full = isinstance(cached, dict) and "caption" in cached
            if not url or is_full or url in seen:
                continue
            seen.add(url)
            platform = detect_platform_from_url(url)
            if platform == "instagram":
                ig_urls.append(url)
            elif platform == "tiktok":
                tt_urls.append(url)

    print(f"  URLs to scrape: {len(ig_urls):,} IG + {len(tt_urls):,} TikTok "
          f"(skipped {len(cache):,} cached)")

    # ── IG chunks ──
    ig_scraped = 0
    for i in range(0, len(ig_urls), chunk_size):
        chunk = ig_urls[i:i + chunk_size]
        print(f"  [IG] Chunk {i//chunk_size + 1}/{(len(ig_urls)-1)//chunk_size + 1} "
              f"({len(chunk)} URLs)...", flush=True)
        try:
            run = client.actor(IG_SCRAPER).call(
                run_input={
                    "directUrls": chunk,
                    "resultsLimit": len(chunk),
                    "resultsType": "posts",
                },
                timeout_secs=600,
            )
            items = list(client.dataset(run["defaultDatasetId"]).iterate_items())
            for item in items:
                post_url = item.get("url") or item.get("inputUrl", "")
                # Try to match by shortCode
                short_code = item.get("shortCode", "")
                matched_url = None
                for u in chunk:
                    if short_code and short_code in u:
                        matched_url = u
                        break
                    if post_url and (post_url in u or u in post_url):
                        matched_url = u
                        break
                if not matched_url and post_url:
                    matched_url = post_url
                if matched_url:
                    views = item.get("videoViewCount") or item.get("videoPlayCount", 0)
                    likes = item.get("likesCount", 0)
                    comments = item.get("commentsCount", 0)
                    hashtags_raw = item.get("hashtags") or []
                    cache[matched_url] = {
                        "views": views,
                        "likes": likes,
                        "comments": comments,
                        "caption": item.get("caption", ""),
                        "hashtags": ",".join(h.get("name", "") for h in hashtags_raw if isinstance(h, dict)),
                        "post_date": item.get("timestamp", ""),
                        "engagement_rate": round((likes + comments) / views, 4) if views else 0,
                        "media_type": item.get("type", ""),
                        "shortcode": item.get("shortCode", ""),
                        "owner_followers": item.get("ownerFollowersCount", ""),
                        "thumbnail_url": item.get("displayUrl", ""),
                    }
                    ig_scraped += 1
            print(f"    Got {len(items)} results, total cached: {len(cache):,}")
        except Exception as e:
            print(f"    [ERROR] IG chunk failed: {e}")
        # Save after each chunk in case of interruption
        save_enrich_cache(cache)

    # ── TikTok chunks ──
    tt_scraped = 0
    for i in range(0, len(tt_urls), chunk_size):
        chunk = tt_urls[i:i + chunk_size]
        print(f"  [TT] Chunk {i//chunk_size + 1}/{(len(tt_urls)-1)//chunk_size + 1} "
              f"({len(chunk)} URLs)...", flush=True)
        try:
            run = client.actor(TT_SCRAPER).call(
                run_input={
                    "postURLs": chunk,
                    "shouldDownloadCovers": False,
                    "shouldDownloadVideos": False,
                    "shouldDownloadSubtitles": False,
                },
                timeout_secs=300,
            )
            items = list(client.dataset(run["defaultDatasetId"]).iterate_items())
            for item in items:
                web_url = item.get("webVideoUrl", "")
                matched_url = None
                for u in chunk:
                    if web_url and (web_url in u or u in web_url):
                        matched_url = u
                        break
                if not matched_url and web_url:
                    matched_url = web_url
                if matched_url:
                    views = item.get("playCount", 0)
                    likes = item.get("diggCount", 0)
                    comments = item.get("commentCount", 0)
                    hashtags_raw = item.get("hashtags") or []
                    author = item.get("authorMeta") or {}
                    cache[matched_url] = {
                        "views": views,
                        "likes": likes,
                        "comments": comments,
                        "caption": item.get("text", ""),
                        "hashtags": ",".join(h.get("name", "") for h in hashtags_raw if isinstance(h, dict)),
                        "post_date": item.get("createTimeISO", ""),
                        "engagement_rate": round((likes + comments) / views, 4) if views else 0,
                        "media_type": "Video",
                        "shortcode": str(item.get("id", "")),
                        "owner_followers": author.get("fans", ""),
                        "thumbnail_url": (item.get("videoMeta") or {}).get("coverUrl", ""),
                    }
                    tt_scraped += 1
            print(f"    Got {len(items)} results, total cached: {len(cache):,}")
        except Exception as e:
            print(f"    [ERROR] TT chunk failed: {e}")
        save_enrich_cache(cache)

    print(f"\n  Enrichment complete: {ig_scraped} IG + {tt_scraped} TikTok scraped")
    print(f"  Total cache: {len(cache):,} posts")
    return cache


# ── Main Logic ───────────────────────────────────────────────────────

def build_export(min_views=2000, min_likes=30, dry_run=False, enrich=False,
                 min_er=0, min_followers=0, max_followers=0,
                 platform_filter="", language_filter="", theme_filter="",
                 since_filter="", pg_sync=False, max_content=5):
    print("=" * 60)
    print("US Content Full Export Builder (Social Enricher)")
    print(f"  Thresholds: views >= {min_views:,} OR likes >= {min_likes:,}")
    filters_active = []
    if min_er > 0:
        filters_active.append(f"ER >= {min_er}")
    if min_followers > 0:
        filters_active.append(f"followers >= {min_followers:,}")
    if max_followers > 0:
        filters_active.append(f"followers <= {max_followers:,}")
    if platform_filter:
        filters_active.append(f"platform = {platform_filter}")
    if language_filter:
        filters_active.append(f"language = {language_filter}")
    if theme_filter:
        filters_active.append(f"theme in [{theme_filter}]")
    if since_filter:
        filters_active.append(f"since {since_filter}")
    if filters_active:
        print(f"  Filters: {', '.join(filters_active)}")
    if pg_sync:
        print(f"  PG Sync: enabled")
    print("=" * 60)

    # 1. Load all tabs from cache
    print("\n[1/5] Loading cached tabs...")
    creators_raw = load_cached_tab("Creators_updated")
    output_raw = load_cached_tab("Output_updated")
    blacklist_raw = load_cached_tab("Mgmt_black_list")
    print(f"  Creators: {len(creators_raw):,} rows")
    print(f"  Output:   {len(output_raw):,} rows")
    print(f"  Blacklist: {len(blacklist_raw):,} entries")

    if not creators_raw or not output_raw:
        print("\n[ERROR] Cache missing. Run these first:")
        print(f'  python tools/data_crawler.py --source "sheet:{SHEET_ID}" --sheet-name "Creators_updated" --dry-run')
        print(f'  python tools/data_crawler.py --source "sheet:{SHEET_ID}" --sheet-name "Output_updated" --dry-run')
        return

    # 2. Build blacklist set
    blacklist = set()
    for row in blacklist_raw:
        bl_id = row.get("ID", "").strip().lower()
        if bl_id:
            blacklist.add(bl_id)
    print(f"\n[2/5] Blacklist: {len(blacklist)} usernames")

    # 3. Filter creators: views >= min_views OR likes >= min_likes, NOT blacklisted
    print(f"\n[3/5] Filtering creators (views >= {min_views:,} OR likes >= {min_likes:,})...")
    eligible_creators = {}  # username_lower -> creator dict
    stats = {"total": 0, "blacklisted": 0, "no_data": 0, "below_threshold": 0, "passed": 0}

    for c in creators_raw:
        stats["total"] += 1
        username = c.get("Username (id)", "").strip()
        username_lower = username.lower()

        # Blacklist check
        if username_lower in blacklist:
            stats["blacklisted"] += 1
            continue
        if c.get("Blacklist 여부", "").strip().upper() == "TRUE":
            stats["blacklisted"] += 1
            continue

        views_30d = safe_num(c.get("최근 30일 조회 수 총합", ""))
        likes_30d = safe_num(c.get("최근 30일 좋아요 수 총합", ""))

        # OR condition
        if not (views_30d >= min_views or likes_30d >= min_likes):
            if views_30d == 0 and likes_30d == 0:
                stats["no_data"] += 1
            else:
                stats["below_threshold"] += 1
            continue

        # Advanced filters
        comments_30d = safe_num(c.get("최근 30일 댓글 수 총합", ""))
        creator_er = (likes_30d + comments_30d) / views_30d if views_30d > 0 else 0
        followers_num = safe_num(c.get("Followers", ""))
        platform = c.get("Platform", "").strip().lower()
        language = c.get("Language", "").strip().lower()

        if min_er > 0 and creator_er < min_er:
            stats["below_threshold"] += 1
            continue
        if min_followers > 0 and followers_num < min_followers:
            stats["below_threshold"] += 1
            continue
        if max_followers > 0 and followers_num > max_followers:
            stats["below_threshold"] += 1
            continue
        if platform_filter and platform != platform_filter.lower():
            stats["below_threshold"] += 1
            continue
        if language_filter and language != language_filter.lower():
            stats["below_threshold"] += 1
            continue

        stats["passed"] += 1
        eligible_creators[username_lower] = {
            "username": username,
            "email": c.get("Email", "").strip(),
            "platform": platform,
            "profile_url": c.get("Profile URL", "").strip(),
            "first_discovered": c.get("최초 발견 일자", "").strip(),
            "followers": c.get("Followers", "").strip(),
            "views_30d": views_30d,
            "likes_30d": likes_30d,
            "comments_30d": comments_30d,
            "engagement_rate": creator_er,
            "language": language,
            "age": c.get("Age", "").strip(),
            "gender": c.get("Gender", "").strip(),
            "race": c.get("Race", "").strip(),
            "location": c.get("Location", "").strip(),
            "collab_status": c.get("제휴 상태", "").strip(),
        }

    print(f"  Total creators:     {stats['total']:,}")
    print(f"  Blacklisted:        {stats['blacklisted']:,}")
    print(f"  No 30d data:        {stats['no_data']:,}")
    print(f"  Below threshold:    {stats['below_threshold']:,}")
    print(f"  >>> Eligible:       {stats['passed']:,}")

    # 4. Match content from Output_updated to eligible creators
    print(f"\n[4/5] Matching content to eligible creators...")
    # content_map: username_lower -> [content_dicts]
    content_map = defaultdict(list)
    output_matched = 0
    output_skipped_bl = 0
    output_skipped_noeligible = 0

    for o in output_raw:
        out_username = o.get("Username (id)", "").strip()
        out_username_lower = out_username.lower()

        # Skip blacklisted
        if out_username_lower in blacklist:
            output_skipped_bl += 1
            continue
        if o.get("블랙리스트 여부", "").strip().upper() == "TRUE":
            output_skipped_bl += 1
            continue

        # Only match to eligible creators
        if out_username_lower not in eligible_creators:
            output_skipped_noeligible += 1
            continue

        # Content-level filters
        content_theme = o.get("Theme", "").strip().lower()
        content_date = o.get("date", "").strip()
        if theme_filter:
            themes = [t.strip().lower() for t in theme_filter.split(",")]
            if not any(t in content_theme for t in themes):
                continue
        if since_filter and content_date and content_date < since_filter:
            continue

        post_url = o.get("Post URL", "").strip()

        # Dedup: skip if same post_url already exists for this creator
        if any(c["post_url"] == post_url for c in content_map[out_username_lower]):
            continue

        output_matched += 1
        content_map[out_username_lower].append({
            "post_url": post_url,
            "platform": detect_platform_from_url(post_url),
            "post_date": o.get("date", "").strip(),
            "bio_text": o.get("Bio_text", "").strip(),
            "caption": o.get("Caption", "").strip(),
            "transcript": o.get("Transcript", "").strip(),
            "email_from_output": o.get("Email", "").strip(),
            "summary": o.get("Summary", "").strip(),
            "theme": o.get("Theme", "").strip(),
            "level": o.get("Level", "").strip(),
            "score": o.get("Score", "").strip(),
            "keyword_1": o.get("Keyword 1", "").strip(),
        })

    print(f"  Output rows matched:      {output_matched:,}")
    print(f"  Output rows (blacklisted): {output_skipped_bl:,}")
    print(f"  Output rows (no creator):  {output_skipped_noeligible:,}")

    # ── Trim: keep top N per creator, prioritize data completeness ──
    pre_trim = sum(len(v) for v in content_map.values())
    for uname_lower in content_map:
        posts = content_map[uname_lower]
        if len(posts) <= max_content:
            continue
        # Score each post: transcript > caption > summary > theme
        def completeness(p):
            score = 0
            if p.get("transcript", "").strip():
                score += 4
            if p.get("caption", "").strip():
                score += 2
            if p.get("summary", "").strip():
                score += 1
            if p.get("theme", "").strip():
                score += 1
            return score
        posts.sort(key=lambda p: completeness(p), reverse=True)
        content_map[uname_lower] = posts[:max_content]
    post_trim = sum(len(v) for v in content_map.values())
    output_matched = post_trim
    print(f"  Dedup trim (max {max_content}/creator): {pre_trim:,} -> {post_trim:,} ({pre_trim - post_trim:,} removed)")

    creators_with_content = sum(1 for u in eligible_creators if u in content_map)
    creators_no_content = len(eligible_creators) - creators_with_content
    print(f"  Creators with content:     {creators_with_content:,}")
    print(f"  Creators without content:  {creators_no_content:,}")

    # 5. Merge emails: prefer creator email, fallback to output email
    for uname_lower, creator in eligible_creators.items():
        if not creator["email"] and uname_lower in content_map:
            for c in content_map[uname_lower]:
                if c["email_from_output"]:
                    creator["email"] = c["email_from_output"]
                    break
        # Also grab bio from output if creator doesn't have it
        if uname_lower in content_map:
            for c in content_map[uname_lower]:
                if c["bio_text"]:
                    creator["bio_text"] = c["bio_text"]
                    break
            else:
                creator["bio_text"] = ""
        else:
            creator["bio_text"] = ""

    # ── Summary for PG structure ──
    print(f"\n{'='*60}")
    print(f"[SUMMARY] PG-ready estimates:")
    print(f"  Creator Pool rows:  {len(eligible_creators):,}")
    print(f"  Content Pool rows:  {output_matched:,}")
    max_content = max((len(v) for v in content_map.values()), default=0)
    avg_content = output_matched / max(creators_with_content, 1)
    print(f"  Max content/creator: {max_content}")
    print(f"  Avg content/creator: {avg_content:.1f}")
    print(f"{'='*60}")

    if dry_run:
        print("\n[DRY RUN] No file written.")
        for uname_lower, creator in list(eligible_creators.items())[:3]:
            contents = content_map.get(uname_lower, [])
            print(f"\n  Creator: {creator['username']} ({creator['platform']}) "
                  f"| F:{creator['followers']} V30:{creator['views_30d']:,.0f} L30:{creator['likes_30d']:,.0f}")
            print(f"    Email: {creator['email']} | Content count: {len(contents)}")
            for i, ct in enumerate(contents[:2]):
                print(f"    Content {i+1}: {ct['post_url'][:60]}...")
        return

    # ── 5a. Apify Enrichment (optional) ──
    enrich_data = {}
    if enrich:
        print(f"\n[5a/6] Social Enricher — fetching per-post metrics via Apify...")
        enrich_data = enrich_post_metrics(content_map, chunk_size=50)
    else:
        # Still load cache if available (from previous --enrich runs)
        enrich_data = load_enrich_cache()
        if enrich_data:
            print(f"\n[5a/6] Using cached enrichment: {len(enrich_data):,} posts")
        else:
            print(f"\n[5a/6] No enrichment (use --enrich to fetch post metrics via Apify)")

    # Helper to get post metrics (11 fields)
    def get_post_metrics(post_url):
        m = enrich_data.get(post_url, {})
        return {
            "views": m.get("views", ""),
            "likes": m.get("likes", ""),
            "comments": m.get("comments", ""),
            "caption": m.get("caption", ""),
            "hashtags": m.get("hashtags", ""),
            "post_date": m.get("post_date", ""),
            "engagement_rate": m.get("engagement_rate", ""),
            "media_type": m.get("media_type", ""),
            "shortcode": m.get("shortcode", ""),
            "owner_followers": m.get("owner_followers", ""),
            "thumbnail_url": m.get("thumbnail_url", ""),
        }

    # ── 5b. Write Excel ──
    print(f"\n[6/6] Writing Excel...")
    import openpyxl

    output_dir = PROJECT_ROOT / "SYNCLY"
    output_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    output_path = output_dir / f"US_Content_Full_{today}.xlsx"

    wb = openpyxl.Workbook()

    # ── Sheet 1: Creator Pool ──
    ws_creators = wb.active
    ws_creators.title = "Creator Pool"
    creator_headers = [
        "username", "email", "platform", "profile_url",
        "first_discovered", "followers", "views_30d", "likes_30d",
        "comments_30d", "engagement_rate",
        "language", "age", "gender", "race",
        "location", "collab_status", "bio_text", "content_count",
    ]
    ws_creators.append(creator_headers)

    # Sort by views descending
    sorted_creators = sorted(
        eligible_creators.items(),
        key=lambda x: x[1]["views_30d"],
        reverse=True,
    )

    for uname_lower, cr in sorted_creators:
        content_count = len(content_map.get(uname_lower, []))
        ws_creators.append([
            clean(cr["username"]),
            clean(cr["email"]),
            clean(cr["platform"]),
            clean(cr["profile_url"]),
            clean(cr["first_discovered"]),
            clean(cr["followers"]),
            cr["views_30d"],
            cr["likes_30d"],
            cr["comments_30d"],
            round(cr.get("engagement_rate", 0), 4),
            clean(cr.get("language", "")),
            clean(cr["age"]),
            clean(cr["gender"]),
            clean(cr["race"]),
            clean(cr["location"]),
            clean(cr["collab_status"]),
            clean(cr.get("bio_text", "")),
            content_count,
        ])

    # ── Sheet 2: Content Pool (with post-level metrics) ──
    ws_content = wb.create_sheet("Content Pool")
    content_headers = [
        "username", "content_num", "post_url", "platform",
        "post_date", "post_views", "post_likes", "post_comments",
        "engagement_rate", "media_type", "hashtags", "shortcode",
        "thumbnail_url",
        "caption", "transcript", "summary",
        "theme", "level", "score", "keyword_1", "bio_text",
    ]
    ws_content.append(content_headers)

    enriched_count = 0
    for uname_lower, cr in sorted_creators:
        contents = content_map.get(uname_lower, [])
        for i, ct in enumerate(contents):
            pm = get_post_metrics(ct["post_url"])
            if pm["views"] != "":
                enriched_count += 1
            # Use enriched post_date if available, fallback to Syncly
            post_date = pm["post_date"] or ct["post_date"]
            # Use enriched caption if available, fallback to Syncly
            caption = pm["caption"] or ct["caption"]
            ws_content.append([
                clean(cr["username"]),
                i + 1,
                clean(ct["post_url"]),
                clean(ct["platform"]),
                clean(post_date),
                pm["views"], pm["likes"], pm["comments"],
                pm["engagement_rate"], clean(pm["media_type"]),
                clean(pm["hashtags"]), clean(pm["shortcode"]),
                clean(pm["thumbnail_url"]),
                clean(caption),
                clean(ct["transcript"]),
                clean(ct["summary"]),
                clean(ct["theme"]),
                clean(ct["level"]),
                clean(ct["score"]),
                clean(ct["keyword_1"]),
                clean(ct["bio_text"]),
            ])

    # ── Sheet 3: Flat View (Creator + Content merged) ──
    ws_flat = wb.create_sheet("Flat View")
    flat_headers = [
        "username", "email", "platform", "followers",
        "views_30d", "likes_30d", "comments_30d",
        "content_num", "post_url", "post_date",
        "post_views", "post_likes", "post_comments",
        "engagement_rate", "media_type", "hashtags",
        "caption", "transcript", "bio_text",
    ]
    ws_flat.append(flat_headers)

    for uname_lower, cr in sorted_creators:
        contents = content_map.get(uname_lower, [])
        if not contents:
            ws_flat.append([
                clean(cr["username"]),
                clean(cr["email"]),
                clean(cr["platform"]),
                clean(cr["followers"]),
                cr["views_30d"],
                cr["likes_30d"],
                cr["comments_30d"],
                0, "", "", "", "", "", "", "", "", "", "", "",
            ])
        else:
            for i, ct in enumerate(contents):
                pm = get_post_metrics(ct["post_url"])
                post_date = pm["post_date"] or ct["post_date"]
                caption = pm["caption"] or ct["caption"]
                ws_flat.append([
                    clean(cr["username"]),
                    clean(cr["email"]),
                    clean(cr["platform"]),
                    clean(cr["followers"]),
                    cr["views_30d"],
                    cr["likes_30d"],
                    cr["comments_30d"],
                    i + 1,
                    clean(ct["post_url"]),
                    clean(post_date),
                    pm["views"], pm["likes"], pm["comments"],
                    pm["engagement_rate"], clean(pm["media_type"]),
                    clean(pm["hashtags"]),
                    clean(caption),
                    clean(ct["transcript"]),
                    clean(ct["bio_text"]),
                ])

    wb.save(str(output_path))
    print(f"\n  Saved: {output_path}")
    print(f"  Sheets: Creator Pool ({len(eligible_creators):,}), "
          f"Content Pool ({output_matched:,}), "
          f"Flat View")
    if enrich_data:
        print(f"  Enriched posts: {enriched_count:,}/{output_matched:,}")

    # ── PG Sync ──
    if pg_sync:
        print(f"\n[PG SYNC] Pushing to PostgreSQL...")
        try:
            from push_content_to_pg import push_posts, push_metrics
        except ImportError:
            sys.path.insert(0, str(DIR))
            from push_content_to_pg import push_posts, push_metrics

        today_str = datetime.now().strftime("%Y-%m-%d")
        pg_posts = []
        pg_metrics = []
        seen_post_ids = set()
        skipped_no_id = 0
        skipped_dup = 0
        for uname_lower, cr in sorted_creators:
            contents = content_map.get(uname_lower, [])
            for ct in contents:
                pm = enrich_data.get(ct["post_url"], {})
                shortcode = pm.get("shortcode", "")
                if not shortcode:
                    # Extract from URL
                    url_parts = ct["post_url"].rstrip("/").split("/")
                    shortcode = url_parts[-1] if url_parts else ""
                if not shortcode:
                    skipped_no_id += 1
                    continue
                if shortcode in seen_post_ids:
                    skipped_dup += 1
                    continue
                seen_post_ids.add(shortcode)
                post_date = pm.get("post_date", "") or ct.get("post_date", "")
                caption = pm.get("caption", "") or ct.get("caption", "")
                hashtags = pm.get("hashtags", "")

                pg_posts.append({
                    "post_id": shortcode,
                    "url": ct["post_url"],
                    "platform": ct["platform"],
                    "username": cr["username"],
                    "nickname": cr["username"],
                    "followers": int(safe_num(cr["followers"])),
                    "caption": caption or "",
                    "transcript": ct.get("transcript", ""),
                    "hashtags": hashtags,
                    "tagged_account": "",
                    "post_date": post_date[:10] if post_date else today_str,
                    "brand": "",
                    "region": "us",
                    "source": "social_enricher",
                    "bio_text": cr.get("bio_text", ""),
                    "views_30d": int(cr["views_30d"]),
                    "likes_30d": int(cr["likes_30d"]),
                    "text": ct.get("summary", ""),
                    "engagement_rate": pm.get("engagement_rate", 0) or 0,
                })

                views = pm.get("views", 0)
                likes = pm.get("likes", 0)
                comments = pm.get("comments", 0)
                if views or likes or comments:
                    pg_metrics.append({
                        "post_id": shortcode,
                        "date": today_str,
                        "views": int(views) if views else 0,
                        "likes": int(likes) if likes else 0,
                        "comments": int(comments) if comments else 0,
                    })

        if skipped_no_id or skipped_dup:
            print(f"  PG dedup: {skipped_no_id} skipped (no post_id), {skipped_dup} skipped (duplicate)")
        push_posts(pg_posts)
        if pg_metrics:
            push_metrics(pg_metrics)
        print(f"  PG sync complete: {len(pg_posts)} posts, {len(pg_metrics)} metrics")

    print(f"\nDone!")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(
        description="Social Enricher - Build US Content Full export with Apify post metrics"
    )
    # Threshold filters
    parser.add_argument("--min-views", type=int, default=2000,
        help="Min 30-day views threshold (default: 2000)")
    parser.add_argument("--min-likes", type=int, default=30,
        help="Min 30-day likes threshold (default: 30)")
    # Advanced filters
    parser.add_argument("--min-er", type=float, default=0,
        help="Min engagement rate filter (e.g. 0.03 = 3%%)")
    parser.add_argument("--min-followers", type=int, default=0,
        help="Min followers filter")
    parser.add_argument("--max-followers", type=int, default=0,
        help="Max followers filter (0 = no limit)")
    parser.add_argument("--platform", type=str, default="",
        help="Filter by platform: instagram, tiktok")
    parser.add_argument("--language", type=str, default="",
        help="Filter by language (e.g. en, ja)")
    parser.add_argument("--theme", type=str, default="",
        help="Filter content by theme (comma-separated, e.g. baby,family)")
    parser.add_argument("--since", type=str, default="",
        help="Filter content posted since date (YYYY-MM-DD)")
    parser.add_argument("--max-content", type=int, default=5,
        help="Max content posts per creator (default: 5, prioritizes completeness)")
    # Actions
    parser.add_argument("--dry-run", action="store_true",
        help="Show counts + samples without writing Excel")
    parser.add_argument("--enrich", action="store_true",
        help="Fetch per-post views/likes/comments via Apify (costs credits)")
    parser.add_argument("--pg-sync", action="store_true",
        help="Sync results to PostgreSQL via orbitools API")
    args = parser.parse_args()

    build_export(
        min_views=args.min_views,
        min_likes=args.min_likes,
        dry_run=args.dry_run,
        enrich=args.enrich,
        min_er=args.min_er,
        min_followers=args.min_followers,
        max_followers=args.max_followers,
        platform_filter=args.platform,
        language_filter=args.language,
        theme_filter=args.theme,
        since_filter=args.since,
        pg_sync=args.pg_sync,
        max_content=args.max_content,
    )
