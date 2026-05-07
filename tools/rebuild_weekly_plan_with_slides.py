"""
기존 주간 기획안 엑셀(소제목/대제목/비주얼 방향 형식)을 새 15열 포맷
(S1~S5 슬라이드별 본문 셀)으로 재작성한다.

각 행마다 Claude API 호출:
- 입력: 소제목/대제목/주제/캡션/비주얼 방향
- 출력: slides 5개 객체 (title_jp/title_ko/body_jp/body_ko/visual)

Usage:
  python tools/rebuild_weekly_plan_with_slides.py --input <path> [--output <path>]
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path

import anthropic
import openpyxl
from dotenv import load_dotenv

sys.path.insert(0, str(Path(__file__).resolve().parent))
from plan_weekly_content import build_excel, MODEL, RATE_LIMIT_DELAY  # type: ignore

load_dotenv()


def read_existing(input_path: Path) -> list[dict]:
    """기존 엑셀의 각 행을 dict로 읽어 반환 (Summary 행 제외)."""
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        first_cell = ws.cell(r, 1).value
        if first_cell is None or str(first_cell).strip().lower() in ("summary", ""):
            continue
        row = {}
        for ci, h in enumerate(headers, start=1):
            row[h] = ws.cell(r, ci).value
        rows.append(row)
    return rows


SYSTEM_PROMPT = """あなたはGrosmimi Japan（グロミミ）のInstagram担当エディターです。
6〜24か月の赤ちゃんを育てるママ向けに、5枚カルーセルの各スライド本文を作成します。

トーン:
- 「隣の頼れる年上ママ」が語りかけるような情報×親しみの中間トーン
- 難しい漢字語・教科書的表現は避ける（NG: 黄金ルール / 必須 / 推奨）
- オノマトペは全体で0〜1個まで（多用禁止）
- 各スライドのbodyは・を頭に付けたポイント箇条書き（3〜5行）"""


def build_slide_prompt(row: dict) -> str:
    cat = row.get("카테고리", "")
    subtitle_jp = row.get("소제목", "")
    subtitle_ko = row.get("소제목(한국어)", "")
    title_jp = row.get("대제목", "")
    title_ko = row.get("대제목(한국어)", "")
    topic_jp = row.get("주제", "")
    topic_ko = row.get("주제(한국어)", "")
    visual_jp = row.get("비주얼 방향", "")
    visual_ko = row.get("비주얼 방향(한국어)", "")
    caption = row.get("캡션", "")

    return f"""以下の企画案について、5枚カルーセルの各スライド本文をJSONで生成してください。

## 入力
- カテゴリ: {cat}
- 小タイトル(JP): {subtitle_jp}
- 小タイトル(KO): {subtitle_ko}
- メインタイトル(JP): {title_jp}
- メインタイトル(KO): {title_ko}
- トピック(JP): {topic_jp}
- トピック(KO): {topic_ko}
- 既存の構成案（参考）: {visual_jp}
- 既存の構成案 韓国語: {visual_ko}
- キャプション(JP): {caption[:300]}

## 出力要件
5枚構成: S1=表紙 / S2=훅・問題提起 / S3=核心情報1（比較表/データ） / S4=核心情報2（手順/チェックリスト） / S5=CTA
各スライドに以下を埋める:
- title_jp / title_ko: スライド上のテキストオーバーレイ（短く、キーワード強調）
- body_jp / body_ko: ポイント箇条書き 3〜5行（・○○○ 形式）。空欄禁止
- visual: 写真/イラスト/比較表など視覚要素の具体的構想（1〜2文）

S1のtitleは既存「{title_jp}」/「{title_ko}」を活用OK。

## 出力形式（JSONのみ、余計なテキスト禁止）
```json
{{
  "slides": [
    {{"title_jp":"...","title_ko":"...","body_jp":"・...\\n・...\\n・...","body_ko":"・...\\n・...\\n・...","visual":"..."}},
    {{"title_jp":"...","title_ko":"...","body_jp":"...","body_ko":"...","visual":"..."}},
    {{"title_jp":"...","title_ko":"...","body_jp":"...","body_ko":"...","visual":"..."}},
    {{"title_jp":"...","title_ko":"...","body_jp":"...","body_ko":"...","visual":"..."}},
    {{"title_jp":"...","title_ko":"...","body_jp":"...","body_ko":"...","visual":"..."}}
  ]
}}
```"""


def call_claude_for_slides(client: anthropic.Anthropic, row: dict) -> list[dict]:
    prompt = build_slide_prompt(row)
    for attempt in range(3):
        try:
            resp = client.messages.create(
                model=MODEL,
                max_tokens=4096,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": prompt}],
            )
            text = resp.content[0].text
            if "```json" in text:
                json_str = text.split("```json")[1].split("```")[0].strip()
            elif "```" in text:
                json_str = text.split("```")[1].split("```")[0].strip()
            else:
                json_str = text.strip()
            data = json.loads(json_str)
            slides = data.get("slides", [])
            if len(slides) >= 5:
                return slides[:5]
            print(f"  [WARN] slides count {len(slides)} < 5, retrying...")
        except json.JSONDecodeError as e:
            print(f"  [WARN] JSON parse fail ({attempt+1}/3): {e}")
        except Exception as e:
            print(f"  [WARN] API error ({attempt+1}/3): {e}")
        time.sleep(2 * (attempt + 1))
    raise RuntimeError("Claude API failed after 3 retries")


CATEGORY_REVERSE = {
    "육아 정보": "tips",
    "브랜드 제품": "brand",
    "K-유아식": "k_babyfood",
    "K유아식": "k_babyfood",
    "육아 지식": "knowledge",
    "지식": "knowledge",
    "tips": "tips",
    "brand": "brand",
    "k_babyfood": "k_babyfood",
    "knowledge": "knowledge",
}


def row_to_plan(row: dict, slides: list[dict]) -> dict:
    """기존 row + 새 slides를 새 plan 스키마로 합성."""
    cat_label = (row.get("카테고리") or "tips").strip()
    cat_key = CATEGORY_REVERSE.get(cat_label, "tips")

    hashtags = row.get("해시태그", "") or ""
    if isinstance(hashtags, str):
        hashtags_list = [h for h in hashtags.split() if h.startswith("#")]
    else:
        hashtags_list = []

    return {
        "category": cat_key,
        "topic": row.get("주제", "") or "",
        "topic_ko": row.get("주제(한국어)", "") or "",
        "slides": slides,
        "image_concept": (row.get("비주얼 방향", "") or "")[:300],
        "emphasis": "",
        "caption": row.get("캡션", "") or "",
        "caption_ko": row.get("캡션(한국어)", "") or "",
        "hashtags": hashtags_list,
        "trend_ref": row.get("참고 트렌드", "") or "",
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="기존 기획안 엑셀 경로")
    parser.add_argument("--output", help="출력 경로(미지정 시 입력 파일 덮어쓰기, 백업 자동)")
    parser.add_argument("--limit", type=int, help="처리할 행 수 제한 (테스트용)")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: 파일 없음: {input_path}")
        sys.exit(1)

    output_path = Path(args.output) if args.output else input_path

    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY 미설정")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    print(f"[1] 입력 읽기: {input_path.name}")
    rows = read_existing(input_path)
    if args.limit:
        rows = rows[: args.limit]
    print(f"  → {len(rows)}건")

    print(f"\n[2] Claude API로 슬라이드 본문 생성")
    plans: list[dict] = []
    for i, row in enumerate(rows):
        topic = row.get("주제", "")
        print(f"  [{i+1}/{len(rows)}] {str(topic)[:50]}")
        slides = call_claude_for_slides(client, row)
        plans.append(row_to_plan(row, slides))
        if i < len(rows) - 1:
            time.sleep(RATE_LIMIT_DELAY)

    if output_path == input_path:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = input_path.with_name(input_path.stem + f"_BEFORE_REWRITE_{ts}" + input_path.suffix)
        shutil.copy2(input_path, backup)
        print(f"\n[3] 원본 백업: {backup.name}")

    print(f"\n[4] 새 포맷 엑셀 저장: {output_path.name}")
    build_excel(plans, output_path)
    print(f"  ✓ 완료 — {len(plans)}건, 15열")


if __name__ == "__main__":
    main()
