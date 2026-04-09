"""
Discover Content — Apify 키워드 검색 → 필터 → HT 풀 파이프라인 분석
====================================================================

Google Trends 기반 키워드로 TikTok + IG 영상을 발견하고,
뷰 기준 필터 → HT 30프레임 분석 → 풀 데이터 JSON 출력.

Usage:
    # 기본: toddler cup 키워드, 100K+ 뷰, 50개
    python tools/discover_content.py --category "toddler cup"

    # 커스텀
    python tools/discover_content.py --keywords "sippy cup,straw cup baby" --min-views 50000 --limit 30

    # 드라이런 (검색만, 분석 안 함)
    python tools/discover_content.py --category "toddler cup" --dry-run

    # 분석만 (이전 검색 결과에서)
    python tools/discover_content.py --from-discovery .tmp/discovery/toddler_cup_discovery.json --limit 50

    # TikTok만
    python tools/discover_content.py --category "toddler cup" --platform tiktok
"""

import argparse
import io
import json
import os
import sys
import tempfile
import time
from datetime import datetime, timezone
from pathlib import Path

# Fix Windows cp949
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

DIR = Path(__file__).resolve().parent
PROJECT_ROOT = DIR.parent
sys.path.insert(0, str(DIR))
sys.path.insert(0, str(DIR / "ci"))

try:
    from env_loader import load_env
    load_env()
except Exception:
    pass

APIFY_TOKEN = os.getenv("APIFY_API_TOKEN", "") or os.getenv("APIFY_TOKEN", "")
OUTPUT_DIR = PROJECT_ROOT / ".tmp" / "discovery"

# ── Category → Keywords mapping (fallback when Google Trends unavailable) ──

CATEGORY_KEYWORDS = {
    "toddler cup": {
        "tiktok": [
            "toddler cup", "sippy cup", "straw cup baby", "toddler drinking",
            "baby straw cup", "transition cup", "toddler cup review",
        ],
        "instagram": [
            "toddlercup", "sippycup", "strawcupbaby", "toddlerdrinking",
            "babycup", "transitioncup", "ppsu", "babystrawcup",
        ],
    },
}


# ── Google Trends 실시간 키워드 확장 ──

def expand_keywords_from_trends(
    seed: str,
    geo: str = "US",
    timeframe: str = "today 3-m",
    max_keywords: int = 15,
) -> list[str]:
    """
    Google Trends related queries + suggestions로 키워드 자동 확장.

    Args:
        seed: 시드 키워드 (e.g., "toddler cup")
        geo: 국가 코드 (US, JP, etc.)
        timeframe: "today 3-m", "today 1-m", "today 12-m"
        max_keywords: 최대 리턴 키워드 수

    Returns:
        확장된 키워드 리스트 (시드 포함, 중복 제거, 트렌드 순)
    """
    try:
        from pytrends.request import TrendReq
    except ImportError:
        print("  [trends] pytrends not installed, using fallback keywords")
        return [seed]

    keywords = [seed]  # 시드는 항상 포함

    try:
        pytrends = TrendReq(hl="en-US", tz=360, timeout=(10, 25))

        # 1) Related queries (top + rising)
        pytrends.build_payload([seed], timeframe=timeframe, geo=geo)
        related = pytrends.related_queries()

        if seed in related:
            # Rising queries (급상승 = 최신 트렌드)
            rising = related[seed].get("rising")
            if rising is not None and not rising.empty:
                for _, row in rising.head(8).iterrows():
                    q = row["query"].strip().lower()
                    if q not in keywords:
                        keywords.append(q)

            # Top queries (꾸준한 인기)
            top = related[seed].get("top")
            if top is not None and not top.empty:
                for _, row in top.head(8).iterrows():
                    q = row["query"].strip().lower()
                    if q not in keywords:
                        keywords.append(q)

        # 2) Suggestions (자동완성)
        suggestions = pytrends.suggestions(seed)
        for s in suggestions[:5]:
            title = s.get("title", "").strip().lower()
            if title and title not in keywords and len(title) < 60:
                keywords.append(title)

    except Exception as e:
        print(f"  [trends] Google Trends error: {e}")
        print(f"  [trends] Falling back to seed keyword only")

    final = keywords[:max_keywords]
    print(f"  [trends] {len(final)} keywords from Google Trends for '{seed}':")
    for kw in final:
        tag = "SEED" if kw == seed else "TREND"
        print(f"    [{tag}] {kw}")

    return final

# ── Apify actors ──

# TikTok search: clockworks/free-tiktok-scraper supports searchQueries
TT_SEARCH_ACTOR = "clockworks/free-tiktok-scraper"
# IG hashtag search
IG_HASHTAG_ACTOR = "apify/instagram-hashtag-scraper"
# IG post scraper (for direct URLs)
IG_POST_ACTOR = "apify/instagram-scraper"


def _apify_run(actor_id: str, run_input: dict, timeout_secs: int = 300) -> list:
    """Run Apify actor and return dataset items."""
    import requests
    if not APIFY_TOKEN:
        print("[ERROR] APIFY_API_TOKEN not set")
        return []

    # Apify REST API uses tilde (~) instead of slash for actor IDs
    api_actor_id = actor_id.replace("/", "~")
    url = f"https://api.apify.com/v2/acts/{api_actor_id}/runs"
    headers = {"Authorization": f"Bearer {APIFY_TOKEN}", "Content-Type": "application/json"}

    resp = requests.post(url, json=run_input, headers=headers, timeout=30)
    if not resp.ok:
        print(f"  [apify] {actor_id} HTTP {resp.status_code}: {resp.text[:300]}")
        return []
    run_data = resp.json()["data"]
    run_id = run_data["id"]
    dataset_id = run_data.get("defaultDatasetId")

    status_url = f"https://api.apify.com/v2/actor-runs/{run_id}"
    print(f"  [apify] Run started: {run_id} — polling...")
    status = "RUNNING"
    start = time.time()
    while time.time() - start < timeout_secs:
        time.sleep(10)
        try:
            sr = requests.get(status_url, headers=headers, timeout=60)
            status = sr.json()["data"]["status"]
            elapsed = int(time.time() - start)
            print(f"  [apify] {elapsed}s — status: {status}")
            if status in ("SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"):
                break
        except requests.exceptions.Timeout:
            elapsed = int(time.time() - start)
            print(f"  [apify] {elapsed}s — poll timeout, retrying...")
            continue

    if status != "SUCCEEDED":
        print(f"  [apify] {actor_id} run {run_id} ended: {status}")
        return []

    # Get dataset ID from run result
    try:
        final = requests.get(status_url, headers=headers, timeout=30)
        dataset_id = final.json()["data"].get("defaultDatasetId", dataset_id)
    except Exception:
        pass

    items_url = f"https://api.apify.com/v2/datasets/{dataset_id}/items?format=json"
    items_resp = requests.get(items_url, headers=headers, timeout=120)
    return items_resp.json() if items_resp.ok else []


# ── Phase 1: Discovery ──

def search_tiktok(keywords: list[str], max_results: int = 200) -> list[dict]:
    """Search TikTok for keyword videos via Apify."""
    print(f"\n[TikTok Search] Keywords: {keywords}")
    items = _apify_run(TT_SEARCH_ACTOR, {
        "searchQueries": keywords,
        "resultsPerPage": max_results,
    }, timeout_secs=300)

    results = []
    for item in items:
        # TikTok search returns video-level items
        views = int(item.get("playCount", 0) or item.get("plays", 0) or 0)
        likes = int(item.get("diggCount", 0) or item.get("likes", 0) or 0)
        comments = int(item.get("commentCount", 0) or item.get("comments", 0) or 0)
        shares = int(item.get("shareCount", 0) or item.get("shares", 0) or 0)

        username = (item.get("authorMeta", {}).get("name", "")
                    or item.get("author", {}).get("uniqueId", "")
                    or item.get("authorName", ""))
        post_url = item.get("webVideoUrl", "") or item.get("url", "")

        if not username or not post_url:
            continue

        results.append({
            "platform": "tiktok",
            "username": username.lower(),
            "post_url": post_url,
            "views": views,
            "likes": likes,
            "comments": comments,
            "shares": shares,
            "caption": (item.get("text", "") or item.get("desc", ""))[:200],
            "duration": item.get("videoMeta", {}).get("duration", 0) or item.get("duration", 0),
            "video_url": item.get("videoUrl", ""),
            "followers": item.get("authorMeta", {}).get("fans", 0) or 0,
            "raw": item,  # 풀 데이터 보존
        })

    print(f"  Found {len(results)} TikTok videos")
    return results


def search_instagram(hashtags: list[str], max_results: int = 200) -> list[dict]:
    """Search Instagram hashtags for reels/posts via Apify."""
    print(f"\n[Instagram Search] Hashtags: {hashtags}")
    items = _apify_run(IG_HASHTAG_ACTOR, {
        "hashtags": hashtags,
        "resultsLimit": max_results,
        "resultsType": "posts",
    }, timeout_secs=300)

    results = []
    for item in items:
        views = int(item.get("videoViewCount", 0) or item.get("viewCount", 0) or 0)
        likes = int(item.get("likesCount", 0) or item.get("likes", 0) or 0)
        comments_count = int(item.get("commentsCount", 0) or item.get("comments", 0) or 0)

        username = item.get("ownerUsername", "") or item.get("owner", {}).get("username", "")
        post_url = item.get("url", "") or item.get("displayUrl", "")
        shortcode = item.get("shortCode", "") or item.get("id", "")

        # 비디오/릴만
        is_video = item.get("isVideo", False) or item.get("type", "") == "Video"
        if not is_video:
            continue
        if not username:
            continue

        if not post_url and shortcode:
            post_url = f"https://www.instagram.com/reel/{shortcode}/"

        results.append({
            "platform": "instagram",
            "username": username.lower(),
            "post_url": post_url,
            "views": views,
            "likes": likes,
            "comments": comments_count,
            "shares": 0,
            "caption": (item.get("caption", "") or "")[:200],
            "duration": item.get("videoDuration", 0) or 0,
            "video_url": item.get("videoUrl", ""),
            "followers": item.get("ownerFollowerCount", 0) or 0,
            "raw": item,
        })

    print(f"  Found {len(results)} Instagram reels")
    return results


def discover(
    keywords_tt: list[str],
    hashtags_ig: list[str],
    platform: str = "both",
    max_results: int = 200,
) -> list[dict]:
    """Run Apify discovery on both platforms."""
    all_results = []

    if platform in ("tiktok", "both"):
        all_results.extend(search_tiktok(keywords_tt, max_results))
    if platform in ("instagram", "both"):
        all_results.extend(search_instagram(hashtags_ig, max_results))

    print(f"\n[Discovery Total] {len(all_results)} raw results")
    return all_results


def filter_and_rank(
    results: list[dict],
    min_views: int = 100_000,
    limit: int = 50,
) -> list[dict]:
    """Filter by views, deduplicate by creator, rank by views."""
    # 뷰 필터
    filtered = [r for r in results if r["views"] >= min_views]
    print(f"  100K+ views: {len(filtered)}")

    # 크리에이터별 최고 뷰 포스트만 (1인 1영상)
    best_by_creator = {}
    for r in filtered:
        key = f"{r['platform']}:{r['username']}"
        if key not in best_by_creator or r["views"] > best_by_creator[key]["views"]:
            best_by_creator[key] = r

    deduped = list(best_by_creator.values())
    deduped.sort(key=lambda x: x["views"], reverse=True)

    top = deduped[:limit]
    print(f"  Deduplicated: {len(deduped)} creators → Top {len(top)}")

    return top


# ── Phase 2: HT Full Pipeline Analysis ──

def analyze_ht(posts: list[dict], source_keyword: str = "") -> list[dict]:
    """
    Run DY dynamic-frame full pipeline on discovered posts.

    For each post:
    1. Download video (yt-dlp for TikTok, Apify CDN for IG)
    2. Extract DY keyframes (20/30/40 based on comments) + audio
    3. Vision DY analysis (GPT-4o) with product/child/question tagging
    4. Whisper transcription → hook_caption extraction
    5. Gemini audio tone analysis
    6. Score calculator v2.1
    7. Top 5 comments fetch (Apify)
    """
    from ci.frame_extractor import extract_frames, extract_audio
    from ci.media_cache import MediaCache
    from ci.downloader import download_video, get_cdn_url
    from ci.vision_tagger import analyze_frames
    from ci.score_calculator import calculate_scores_v2

    # Optional imports
    try:
        from ci.whisper_transcriber import transcribe, analyze_script
        has_whisper = True
    except ImportError:
        has_whisper = False
        print("  [WARN] Whisper not available")

    try:
        from ci.audio_analyzer import analyze_audio
        has_gemini_audio = True
    except ImportError:
        has_gemini_audio = False
        print("  [WARN] Gemini audio analyzer not available")

    # Comment fetcher (for top 5 comments)
    try:
        from ci.enricher import fetch_ig_comments, fetch_tt_comments
        has_comments = True
    except ImportError:
        has_comments = False
        print("  [WARN] Comment fetcher not available")

    media_cache = MediaCache()
    all_results = []

    for idx, post in enumerate(posts):
        username = post["username"]
        post_url = post["post_url"]
        platform = post["platform"]
        post_id = post_url.rstrip("/").split("/")[-1] if post_url else f"disc_{idx}"

        print(f"\n--- [{idx+1}/{len(posts)}] @{username} ({platform}) ---")
        print(f"  URL: {post_url[:80]}")
        print(f"  Views: {post['views']:,} | Likes: {post['likes']:,}")

        comment_count = post.get("comments", 0)

        # Check cache
        if media_cache.has_frames(username, post_id, tier="HT"):
            print(f"  [CACHE HIT] Already have frames")
            frames = media_cache.load_frames(username, post_id)
            audio_path = media_cache.get_audio_path(username, post_id)
            if not audio_path.exists():
                audio_path = None
            duration_sec = 0
            meta = media_cache.load_meta(username, post_id)
            if meta:
                duration_sec = meta.get("duration_sec", 0)
        else:
            # Download + extract
            with tempfile.TemporaryDirectory() as tmpdir:
                tmp = Path(tmpdir)

                # Get video
                video_path = None
                if platform == "tiktok":
                    # TikTok: use yt-dlp directly
                    from ci.downloader import _download_tiktok_video
                    video_path = _download_tiktok_video(post_url, tmp)
                else:
                    # IG: get CDN URL first
                    cdn_url = post.get("video_url", "")
                    if not cdn_url:
                        cdn_url = get_cdn_url(post_url, "instagram") or ""
                    if cdn_url:
                        video_path = tmp / "video.mp4"
                        dl = download_video(cdn_url, video_path, post_url, "instagram")
                        if not dl:
                            video_path = None

                if not video_path or not video_path.exists():
                    print(f"  [SKIP] Download failed")
                    all_results.append({**post, "status": "download_failed", "analysis": None})
                    continue

                try:
                    # Duration
                    try:
                        from moviepy import VideoFileClip
                        clip = VideoFileClip(str(video_path))
                        duration_sec = clip.duration
                        clip.close()
                    except Exception:
                        duration_sec = 0

                    # Extract DY frames (dynamic based on comments)
                    frames_dir = tmp / "frames"
                    frames = extract_frames(video_path, frames_dir, tier="DY", comments=comment_count)
                    print(f"  [FRAMES] Extracted {len(frames)} DY keyframes (comments={comment_count})")

                    # Extract audio
                    audio_out = tmp / "audio.mp3"
                    audio_path = extract_audio(video_path, audio_out)

                    # Save to cache (use HT tier key for compatibility)
                    media_cache.save_frames(username, post_id, frames, tier="HT",
                                            duration_sec=duration_sec, platform=platform,
                                            comments=comment_count)
                    if audio_path:
                        media_cache.save_audio(username, post_id, audio_path)

                    # Reload from cache
                    frames = media_cache.load_frames(username, post_id)
                    cached_audio = media_cache.get_audio_path(username, post_id)
                    audio_path = cached_audio if cached_audio.exists() else None
                except Exception as e:
                    print(f"  [SKIP] Frame extraction failed: {e}")
                    all_results.append({**post, "status": "extraction_failed", "analysis": None})
                    continue

        # Vision DY analysis
        print(f"  [VISION] DY analysis ({len(frames)} frames)")
        vision_result = analyze_frames(frames, tier="DY")

        # Whisper
        whisper_result = {}
        if has_whisper and audio_path and audio_path.exists():
            try:
                transcript = transcribe(audio_path, language="en")
                script_analysis = analyze_script(transcript) if transcript else {}
                whisper_result = {"transcript": transcript, **script_analysis}
                print(f"  [WHISPER] Transcribed ({len(transcript or '')} chars)")
            except Exception as e:
                print(f"  [WHISPER] Error: {e}")

        # Gemini audio tone
        audio_analysis = None
        if has_gemini_audio and audio_path and audio_path.exists():
            try:
                print(f"  [AUDIO] Gemini tone analysis...")
                audio_analysis = analyze_audio(audio_path, duration_sec)
            except Exception as e:
                print(f"  [AUDIO] Error: {e}")

        # Merge CI results
        ci_results = {**vision_result, **whisper_result}
        if audio_analysis:
            ci_results["audio_analysis"] = audio_analysis

        # Score v2.1
        v2_scores = calculate_scores_v2(
            ci_results,
            followers=post.get("followers", 0),
            views=post["views"],
            likes=post["likes"],
            comments=post["comments"],
            enrichment={"duration_seconds": duration_sec},
            audio_analysis=audio_analysis,
        )

        # ── Hook Caption (0-3초 자막 추출) ──
        hook_caption = ""
        transcript_text = whisper_result.get("transcript", "") or ""
        if transcript_text and duration_sec > 0:
            hook_caption = _extract_hook_caption(transcript_text, duration_sec)

        # ── Top 5 Comments (likes 기준) ──
        top_5_comments = []
        if has_comments and post_url:
            try:
                if platform == "tiktok":
                    raw_comments = fetch_tt_comments(post_url, max_comments=100)
                else:
                    raw_comments = fetch_ig_comments(post_url, max_comments=100)
                # Sort by likes, take top 5
                raw_comments.sort(key=lambda c: c.get("likes", 0), reverse=True)
                top_5_comments = [c["text"] for c in raw_comments[:5]]
                print(f"  [COMMENTS] Fetched {len(raw_comments)}, top 5 selected")
            except Exception as e:
                print(f"  [COMMENTS] Error: {e}")

        # ── Raw Data (팩트, 불변) ──
        raw_data = {
            "username": username,
            "platform": platform,
            "post_url": post_url,
            "post_id": post_id,
            "views": post["views"],
            "likes": post["likes"],
            "comments": post["comments"],
            "shares": post.get("shares", 0),
            "caption": post.get("caption", ""),
            "followers": post.get("followers", 0),
            "duration_sec": round(duration_sec, 1),
            "engagement_rate": round((post["likes"] + post["comments"]) / post["views"] * 100, 2) if post["views"] > 0 else 0,
            "virality_coeff": round(post["views"] / post["followers"], 2) if post.get("followers", 0) > 0 else 0,
        }

        # ── Evaluator (AI 판단, 프롬프트/가중치 변경 시 재생성) ──
        evaluator = {
            "vision": vision_result,
            "whisper": whisper_result if whisper_result.get("transcript") else {},
            "audio": audio_analysis if audio_analysis else {},
            "scores": v2_scores,
        }

        result = {
            **raw_data,
            "frame_count": len(frames),
            "media_dir": media_cache.get_media_dir(username, post_id),
            "status": "analyzed",
            # 분리된 구조
            "raw_data": raw_data,
            "evaluator": evaluator,
            # 편의용 최상위 키 (정렬/필터용)
            "composite_v2_score": v2_scores.get("composite_v2_score", 0),
            "tier_scores": v2_scores.get("tier_scores", {}),
            "audio_tone": v2_scores.get("audio_tone", {}),
            "audio_bonus": v2_scores.get("audio_bonus", 0),
            # ── Content Tagging (7개 새 컬럼) ──
            "product_center_pct": vision_result.get("product_center_pct", 0),
            "product_first_appearance_pct": vision_result.get("product_first_appearance_pct", -1),
            "child_appearance_pct": vision_result.get("child_appearance_pct", 0),
            "main_question": vision_result.get("main_question", ""),
            "hook_caption": hook_caption,
            "top_5_comments": top_5_comments,
            "source_keyword": source_keyword,
        }
        all_results.append(result)

        print(f"  → composite={v2_scores.get('composite_v2_score', 0)} | "
              f"audio_bonus={v2_scores.get('audio_bonus', 0)} | "
              f"tiers={v2_scores.get('tier_scores', {})}")

    return all_results


# ── Phase 2.5: Profile Enrichment ──

def enrich_profiles(results: list[dict]) -> list[dict]:
    """
    2차 크롤링: 각 크리에이터 프로필 메트릭 + 최근 포스트 통계 추가.

    Apify로 프로필 데이터(bio, 총 팔로워, 총 포스트, 인증여부) +
    최근 포스트(평균 뷰, 평균 좋아요, 포스팅 빈도) 수집.
    """
    try:
        from ci.enricher import (
            fetch_tt_profile, fetch_ig_profile,
            fetch_tt_recent_posts, fetch_ig_recent_posts,
            calc_posting_stats,
        )
    except ImportError as e:
        print(f"  [enrich] enricher import failed: {e}")
        return results

    # 크리에이터 단위로 그룹 (같은 핸들 중복 호출 방지)
    seen = set()
    unique_creators = []
    for r in results:
        key = f"{r['platform']}:{r['username']}"
        if key not in seen:
            seen.add(key)
            unique_creators.append((r["platform"], r["username"]))

    print(f"\n[Profile Enrichment] {len(unique_creators)} unique creators")
    profile_cache = {}

    for idx, (platform, username) in enumerate(unique_creators):
        key = f"{platform}:{username}"
        print(f"  [{idx+1}/{len(unique_creators)}] @{username} ({platform})...", end=" ")

        try:
            if platform == "tiktok":
                profile = fetch_tt_profile(username)
                recent = fetch_tt_recent_posts(username, limit=20)
            else:
                profile = fetch_ig_profile(username)
                recent = fetch_ig_recent_posts(username, limit=20)

            # 최근 포스트 통계
            stats = {}
            if recent:
                views_list = [p.get("views", 0) for p in recent if p.get("views", 0) > 0]
                likes_list = [p.get("likes", 0) for p in recent]
                stats["recent_posts_count"] = len(recent)
                stats["avg_views"] = int(sum(views_list) / len(views_list)) if views_list else 0
                stats["median_views"] = int(sorted(views_list)[len(views_list)//2]) if views_list else 0
                stats["avg_likes"] = int(sum(likes_list) / len(likes_list)) if likes_list else 0
                try:
                    posting_stats = calc_posting_stats(recent)
                    stats.update(posting_stats)
                except Exception:
                    pass

            profile_cache[key] = {"profile": profile, "recent_stats": stats}
            print(f"followers={profile.get('followers', '?'):,} posts={profile.get('posts_count', '?')}")

        except Exception as e:
            print(f"ERROR: {e}")
            profile_cache[key] = {"profile": {}, "recent_stats": {}}

    # 결과에 프로필 데이터 병합
    for r in results:
        key = f"{r['platform']}:{r['username']}"
        enriched = profile_cache.get(key, {})
        r["profile"] = enriched.get("profile", {})
        r["recent_stats"] = enriched.get("recent_stats", {})
        # followers 업데이트 (프로필에서 가져온 게 더 정확)
        if enriched.get("profile", {}).get("followers"):
            r["followers"] = enriched["profile"]["followers"]

    return results


# ── Phase 3: Output ──

def print_summary(results: list[dict]):
    """Print results summary table."""
    analyzed = [r for r in results if r.get("status") == "analyzed"]
    failed = [r for r in results if r.get("status") != "analyzed"]

    print(f"\n{'='*80}")
    print(f"DISCOVERY + HT ANALYSIS RESULTS")
    print(f"{'='*80}")
    print(f"Total: {len(results)} | Analyzed: {len(analyzed)} | Failed: {len(failed)}")

    if not analyzed:
        return

    print(f"\n{'Username':20s} {'Plat':5s} {'Views':>10s} {'Score':>6s} {'Audio':>6s} "
          f"{'Content':>8s} {'Fit':>8s} {'Audience':>8s} {'Perf':>8s}")
    print("-" * 90)

    for r in sorted(analyzed, key=lambda x: x["composite_v2_score"], reverse=True):
        ts = r.get("tier_scores", {})
        print(f"{r['username'][:20]:20s} {r['platform'][:5]:5s} "
              f"{r['views']:>10,} {r['composite_v2_score']:>6} "
              f"{r.get('audio_bonus', 0):>6} "
              f"{ts.get('content', 0):>8} {ts.get('fit', 0):>8} "
              f"{ts.get('audience', 0):>8} {ts.get('performance', 0):>8}")

    # Audio analysis summary
    audio_analyzed = [r for r in analyzed if r.get("audio_tone", {}).get("audio_tone_tier") not in (None, "no_audio")]
    print(f"\nAudio Analyzed: {len(audio_analyzed)}/{len(analyzed)}")
    if audio_analyzed:
        tiers = {}
        for r in audio_analyzed:
            tier = r.get("audio_tone", {}).get("audio_tone_tier", "unknown")
            tiers[tier] = tiers.get(tier, 0) + 1
        print(f"  Audio Tiers: {tiers}")


def _extract_hook_caption(transcript: str, duration_sec: float) -> str:
    """Extract text from the first 3 seconds of a transcript.

    Whisper transcript is plain text (no timestamps). We estimate the first 3s
    as a proportion of the total text based on duration.
    """
    if not transcript or duration_sec <= 0:
        return ""
    # Estimate: first 3 seconds ≈ first (3/duration) fraction of transcript
    ratio = min(3.0 / duration_sec, 1.0)
    char_count = max(int(len(transcript) * ratio), 50)  # at least 50 chars
    hook_text = transcript[:char_count].strip()
    # Trim to last word boundary
    if len(hook_text) < len(transcript):
        last_space = hook_text.rfind(" ")
        if last_space > 20:
            hook_text = hook_text[:last_space]
    return hook_text


def save_excel(results: list[dict], category: str) -> Path:
    """Save results to Excel with content tagging columns."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    slug = category.replace(" ", "_").lower()
    ts = datetime.now().strftime("%Y%m%d")

    # Output path with versioning
    out_dir = PROJECT_ROOT / "Data Storage" / "discovery"
    out_dir.mkdir(parents=True, exist_ok=True)

    # Find next version
    existing = list(out_dir.glob(f"{slug}_discovery_{ts}_v*.xlsx"))
    version = len(existing) + 1
    out_path = out_dir / f"{slug}_discovery_{ts}_v{version}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Discovery Results"

    # Headers
    headers = [
        "Username", "Platform", "Post URL", "Views", "Likes", "Comments",
        "Duration (s)", "Caption", "Composite Score", "Brand Fit", "Hook Score",
        "Product Center %", "Product First Appear %", "Child Appear %",
        "Main Question", "Hook Caption", "Top 5 Comments", "Source Keyword",
        "Frame Count", "Engagement Rate", "Followers",
    ]

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    analyzed = [r for r in results if r.get("status") == "analyzed"]
    analyzed.sort(key=lambda x: x.get("composite_v2_score", 0), reverse=True)

    for row_idx, r in enumerate(analyzed, 2):
        vision = r.get("evaluator", {}).get("vision", {})
        scores = r.get("evaluator", {}).get("scores", {})
        top_comments = r.get("top_5_comments", [])

        row_data = [
            r.get("username", ""),
            r.get("platform", ""),
            r.get("post_url", ""),
            r.get("views", 0),
            r.get("likes", 0),
            r.get("comments", 0),
            r.get("duration_sec", 0),
            (r.get("caption", "") or "")[:200],
            r.get("composite_v2_score", 0),
            vision.get("brand_fit_score", scores.get("brand_fit_score", 0)),
            vision.get("hook_score", 0),
            r.get("product_center_pct", vision.get("product_center_pct", 0)),
            r.get("product_first_appearance_pct", vision.get("product_first_appearance_pct", -1)),
            r.get("child_appearance_pct", vision.get("child_appearance_pct", 0)),
            r.get("main_question", vision.get("main_question", "")),
            r.get("hook_caption", ""),
            "\n".join(top_comments) if top_comments else "",
            r.get("source_keyword", ""),
            r.get("frame_count", 0),
            r.get("engagement_rate", 0),
            r.get("followers", 0),
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=col in (8, 15, 16, 17))

    # Column widths
    col_widths = [18, 8, 40, 12, 10, 10, 8, 30, 10, 8, 8,
                  12, 14, 12, 35, 35, 45, 20, 8, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(str(out_path))
    print(f"\nExcel saved: {out_path} ({len(analyzed)} rows)")
    return out_path


def save_results(results: list[dict], category: str) -> Path:
    """Save full results JSON."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    slug = category.replace(" ", "_").lower()

    # Full results (with raw Apify data removed for size)
    clean_results = []
    for r in results:
        cr = {k: v for k, v in r.items() if k != "raw"}
        clean_results.append(cr)

    out_path = OUTPUT_DIR / f"{slug}_ht_results_{ts}.json"
    out_path.write_text(
        json.dumps(clean_results, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8"
    )
    print(f"\nResults saved: {out_path}")
    return out_path


def main():
    parser = argparse.ArgumentParser(description="Discover + HT Analyze content by keyword")
    parser.add_argument("--category", default="toddler cup",
                        help="Predefined category (e.g., 'toddler cup')")
    parser.add_argument("--keywords", help="Custom TikTok keywords (comma-separated)")
    parser.add_argument("--hashtags", help="Custom IG hashtags (comma-separated)")
    parser.add_argument("--platform", choices=["tiktok", "instagram", "both"], default="both")
    parser.add_argument("--min-views", type=int, default=100_000)
    parser.add_argument("--limit", type=int, default=50)
    parser.add_argument("--max-search", type=int, default=200,
                        help="Max results per Apify search")
    parser.add_argument("--dry-run", action="store_true",
                        help="Discovery only, no HT analysis")
    parser.add_argument("--from-discovery", type=str,
                        help="Load previous discovery JSON instead of re-searching")
    parser.add_argument("--skip-analysis", action="store_true",
                        help="Save discovery results without running HT pipeline")
    parser.add_argument("--trends", action="store_true",
                        help="Use Google Trends to auto-expand keywords from seed")
    parser.add_argument("--trends-geo", default="US",
                        help="Google Trends geo (default: US)")
    parser.add_argument("--trends-timeframe", default="today 3-m",
                        help="Google Trends timeframe (default: today 3-m)")
    parser.add_argument("--enrich", action="store_true",
                        help="2차 Apify 크롤링: 프로필 메트릭 + 최근 포스트 통계 추가")
    args = parser.parse_args()

    print(f"\n{'='*60}")
    print(f"[Discover Content] Category: {args.category}")
    print(f"  Platform: {args.platform} | Min views: {args.min_views:,} | Limit: {args.limit}")
    if args.trends:
        print(f"  Google Trends: ON (geo={args.trends_geo}, {args.trends_timeframe})")
    print(f"{'='*60}")

    # ── Phase 1: Discovery ──
    if args.from_discovery:
        print(f"\n[Phase 1] Loading previous discovery: {args.from_discovery}")
        disc_path = Path(args.from_discovery)
        disc_data = json.loads(disc_path.read_text(encoding="utf-8"))
        # If it's already filtered results, use as-is
        raw_results = disc_data if isinstance(disc_data, list) else disc_data.get("results", [])
    else:
        # Get keywords
        if args.keywords:
            keywords_tt = [k.strip() for k in args.keywords.split(",")]
        elif args.trends:
            # Google Trends 실시간 키워드 확장
            print(f"\n[Phase 0] Google Trends keyword expansion...")
            keywords_tt = expand_keywords_from_trends(
                seed=args.category,
                geo=args.trends_geo,
                timeframe=args.trends_timeframe,
            )
        elif args.category in CATEGORY_KEYWORDS:
            keywords_tt = CATEGORY_KEYWORDS[args.category]["tiktok"]
        else:
            keywords_tt = [args.category]

        if args.hashtags:
            hashtags_ig = [h.strip() for h in args.hashtags.split(",")]
        elif args.trends:
            # Trends 키워드에서 공백 제거하여 IG 해시태그로 변환
            hashtags_ig = [kw.replace(" ", "") for kw in keywords_tt]
        elif args.category in CATEGORY_KEYWORDS:
            hashtags_ig = CATEGORY_KEYWORDS[args.category]["instagram"]
        else:
            hashtags_ig = [args.category.replace(" ", "")]

        print(f"\n[Phase 1] Apify Discovery...")
        print(f"  TikTok keywords: {keywords_tt}")
        print(f"  IG hashtags: {hashtags_ig}")

        raw_results = discover(keywords_tt, hashtags_ig, args.platform, args.max_search)

        # Save raw discovery
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        slug = args.category.replace(" ", "_").lower()
        disc_path = OUTPUT_DIR / f"{slug}_discovery.json"
        disc_path.write_text(
            json.dumps(raw_results, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8"
        )
        print(f"  Raw discovery saved: {disc_path}")

    # ── Phase 2: Filter ──
    print(f"\n[Phase 2] Filtering (min_views={args.min_views:,}, limit={args.limit})...")
    top_posts = filter_and_rank(raw_results, min_views=args.min_views, limit=args.limit)

    if args.dry_run or args.skip_analysis:
        print(f"\n{'='*60}")
        print(f"[DRY RUN] Top {len(top_posts)} posts ready for HT analysis")
        print(f"{'='*60}")
        for i, p in enumerate(top_posts[:20]):
            print(f"  {i+1:3d}. @{p['username']:20s} {p['platform']:5s} "
                  f"views={p['views']:>10,} {p['post_url'][:60]}")
        if len(top_posts) > 20:
            print(f"  ... +{len(top_posts) - 20} more")

        # Save filtered list
        filtered_path = OUTPUT_DIR / f"{args.category.replace(' ', '_')}_filtered.json"
        filtered_path.write_text(
            json.dumps(top_posts, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8"
        )
        print(f"\n  Filtered list saved: {filtered_path}")
        return

    # ── Phase 3: DY Analysis ──
    print(f"\n[Phase 3] DY Dynamic-Frame Analysis ({len(top_posts)} posts)...")
    results = analyze_ht(top_posts, source_keyword=args.category)

    # ── Phase 3.5: Profile Enrichment ──
    if args.enrich:
        print(f"\n[Phase 3.5] Profile Enrichment (Apify 2차 크롤링)...")
        results = enrich_profiles(results)

    # ── Phase 4: Output ──
    print_summary(results)
    out_path = save_results(results, args.category)
    save_excel(results, args.category)

    # Platform breakdown
    platforms = {}
    for r in results:
        p = r.get("platform", "?")
        platforms[p] = platforms.get(p, 0) + 1
    print(f"\nPlatform breakdown: {platforms}")

    return results


if __name__ == "__main__":
    main()
