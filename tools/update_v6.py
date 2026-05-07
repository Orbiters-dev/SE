"""
Update V6 Excel with live API data.

Pulls data from:
  - DataForSEO        → Search Volume tab (Google + Amazon keyword volumes)
  - Rakuten RMS API   → RAKUTEN tab (weekly total sales)
  - Amazon SP-API     → AMAZON tab (weekly total sales)
  - Meta Marketing API → META tab (campaign-level monthly data)

Usage:
  python tools/update_v6.py                    # update current month (all channels)
  python tools/update_v6.py 2026-01            # update specific month
  python tools/update_v6.py --week 3           # update specific week of current month
  python tools/update_v6.py --search-volume    # rebuild Search Volume tab only
  python tools/update_v6.py --all              # update all months (Nov 2025 ~ current)
"""

import os
os.environ["PYTHONIOENCODING"] = "utf-8"

import sys, argparse, calendar
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

load_dotenv()

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# Import our API modules
sys.path.insert(0, str(Path(__file__).parent))
from rakuten_api import weekly_sales as rakuten_weekly_sales
from amazon_sp_api import weekly_sales as amazon_weekly_sales
import meta_api

# Override Meta credentials for Japan account
meta_api.AD_ACCOUNT_ID = os.getenv("META_JP_AD_ACCOUNT_ID")
meta_api.ACCESS_TOKEN = os.getenv("META_JP_ACCESS_TOKEN")
from meta_api import get_campaign_insights

EXCEL_PATH = Path(__file__).parent.parent / "Japan_Marketing Plan_Monthly_V6.xlsx"

# ── Month / Week mapping ──────────────────────────────────────────────
# V6 structure: 8 months (Nov 2025 ~ Jun 2026), 6 rows per month
# Data starts at row 3, each month = 5 data rows + 1 TOTAL row
DATA_START = 3
ROWS_PER_MONTH = 6  # 5 weeks/campaigns + 1 TOTAL

MONTH_INDEX = {
    "2025-11": 0, "2025-12": 1, "2026-01": 2, "2026-02": 3,
    "2026-03": 4, "2026-04": 5, "2026-05": 6, "2026-06": 7,
}

# ── Styling constants ────────────────────────────────────────────────
NAVY      = "1B2A4A"
DARK_BLUE = "2C3E6B"
LIGHT_BLU = "D6E4F0"
LIGHT_GRY = "F5F5F5"
WHITE     = "FFFFFF"
TOTAL_BG  = "E1E5ED"
BORDER_C  = "C0C0C0"
GREEN_SEC = "2E7D32"
GREEN_BG  = "4CAF50"
ORANGE_BG = "E65100"
BLUE_SEC  = "1565C0"
GREEN_CF  = "C6EFCE"
AMBER_CF  = "FFEB9C"
RED_CF    = "FFC7CE"

# Styling helpers
def _brd(left="thin", right="thin", top="thin", bottom="thin",
         lc=BORDER_C, rc=BORDER_C, tc=BORDER_C, bc=BORDER_C):
    return Border(
        left=Side(left, color=lc), right=Side(right, color=rc),
        top=Side(top, color=tc), bottom=Side(bottom, color=bc))

thin_border = _brd()
bot_thick   = _brd(bottom="medium", bc=NAVY)
hdr_brd     = _brd(bottom="medium", bc=NAVY)

f_title   = Font(name="Aptos", size=14, bold=True, color=WHITE)
f_sec     = Font(name="Aptos", size=11, bold=True, color=WHITE)
f_source  = Font(name="Aptos", size=9, italic=True, color="999999")
f_hdr     = Font(name="Aptos", size=10, bold=True, color=NAVY)
f_data    = Font(name="Aptos", size=10, color="333333")
f_data_b  = Font(name="Aptos", size=10, bold=True, color="333333")
f_tot_lbl = Font(name="Aptos", size=10, bold=True, color=NAVY)
f_tot_dat = Font(name="Aptos", size=10, bold=True, color="1B2A4A")
f_note    = Font(name="Aptos", size=9, italic=True, color="999999")
f_metric  = Font(name="Aptos", size=10, bold=True, color="333333")

fl_navy   = PatternFill("solid", fgColor=NAVY)
fl_green  = PatternFill("solid", fgColor=GREEN_BG)
fl_orange = PatternFill("solid", fgColor=ORANGE_BG)
fl_blue   = PatternFill("solid", fgColor=BLUE_SEC)
fl_lblue  = PatternFill("solid", fgColor=LIGHT_BLU)
fl_lgray  = PatternFill("solid", fgColor=LIGHT_GRY)
fl_white  = PatternFill("solid", fgColor=WHITE)
fl_total  = PatternFill("solid", fgColor=TOTAL_BG)

ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
ar = Alignment(horizontal="right", vertical="center")
al = Alignment(horizontal="left", vertical="center")

YEN = '[$\u00a5-411]#,##0'
NUM = '#,##0'
PCT = "0.0%"

# Search Volume keywords
SV_KEYWORDS = ["grosmimi", "\u30b0\u30ed\u30df\u30df", "\u30b9\u30c8\u30ed\u30fc\u30de\u30b0", "\u30b9\u30de\u30fc\u30c8\u30de\u30b0"]

# Month columns for Search Volume: Nov 2025 ~ Jun 2026 (8 months)
SV_MONTHS = [
    (2025, 11), (2025, 12), (2026, 1), (2026, 2),
    (2026, 3), (2026, 4), (2026, 5), (2026, 6),
]
SV_MONTH_LABELS = [
    "Nov\n2025", "Dec\n2025", "Jan\n2026", "Feb\n2026",
    "Mar\n2026", "Apr\n2026", "May\n2026", "Jun\n2026",
]


def _block_start(mi):
    return DATA_START + mi * ROWS_PER_MONTH

def _total_row(mi):
    return DATA_START + ROWS_PER_MONTH * (mi + 1) - 1


def _week_dates(year, month, week_num):
    """Get start/end dates for a week within a month.
    W1 = 1st~7th, W2 = 8th~14th, W3 = 15th~21st, W4 = 22nd~28th, W5 = 29th~end
    """
    last_day = calendar.monthrange(year, month)[1]
    starts = [1, 8, 15, 22, 29]
    ends = [7, 14, 21, 28, last_day]

    if week_num < 1 or week_num > 5:
        raise ValueError(f"Week must be 1-5, got {week_num}")

    start = starts[week_num - 1]
    end = min(ends[week_num - 1], last_day)

    if start > last_day:
        return None, None  # W5 doesn't exist for short months

    return (f"{year}-{month:02d}-{start:02d}",
            f"{year}-{month:02d}-{end:02d}")


def sc(ws, r, c, val=None, font=None, fill=None, fmt=None, align=None, brd=None):
    """Set cell with styling."""
    cell = ws.cell(row=r, column=c)
    if val is not None:
        cell.value = val
    if font:  cell.font = font
    if fill:  cell.fill = fill
    if fmt:   cell.number_format = fmt
    if align: cell.alignment = align
    cell.border = brd or thin_border
    return cell


# ===================================================================
#  SEARCH VOLUME TAB
# ===================================================================

def update_search_volume(wb):
    """Rebuild Search Volume tab with DataForSEO data.

    Section A: Google Search Volume (monthly breakdown)
    Section B: Amazon Search Volume
    Section C: Search Volume vs Marketing Performance (refs to other tabs)
    """
    from fetch_keyword_volume import fetch_google_volume_historical, fetch_amazon_volume

    tab_name = "Search Volume"

    # Delete existing tab if present
    if tab_name in wb.sheetnames:
        idx = wb.sheetnames.index(tab_name)
        del wb[tab_name]
        ws = wb.create_sheet(tab_name, idx)
    else:
        ws = wb.create_sheet(tab_name)

    ws.sheet_properties.tabColor = GREEN_BG

    # Column widths
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 20   # Keyword / Metric
    ws.column_dimensions["C"].width = 12   # Avg Volume
    ws.column_dimensions["D"].width = 10   # CPC
    for i, _ in enumerate(SV_MONTHS):
        col_letter = get_column_letter(5 + i)  # E onwards
        ws.column_dimensions[col_letter].width = 11
    last_data_col = 4 + len(SV_MONTHS)  # L (col 12)

    # ── Row 1: Title bar ──────────────────────────────────────────
    ws.row_dimensions[1].height = 36
    sc(ws, 1, 2, "SEARCH VOLUME \u2014 \u30ad\u30fc\u30ef\u30fc\u30c9\u691c\u7d22\u30dc\u30ea\u30e5\u30fc\u30e0\u5206\u6790",
       f_title, fl_navy, align=Alignment(horizontal="left", vertical="center"),
       brd=thin_border)
    for c in range(3, last_data_col + 1):
        sc(ws, 1, c, fill=fl_navy, font=f_title, brd=thin_border)
    ws.merge_cells(f"B1:{get_column_letter(last_data_col)}1")

    # ── Section A: GOOGLE SEARCH VOLUME ───────────────────────────
    row = 3
    sc(ws, row, 2, "GOOGLE SEARCH VOLUME", f_sec, fl_green, align=al, brd=thin_border)
    for c in range(3, last_data_col + 1):
        sc(ws, row, c, fill=fl_green, font=f_sec, brd=thin_border)
    ws.merge_cells(f"B{row}:{get_column_letter(last_data_col)}{row}")
    ws.row_dimensions[row].height = 26

    # Source line
    row = 4
    sc(ws, row, 2, "Source: DataForSEO Google Ads API | Location: JP | Language: ja",
       f_source, fl_white, align=al, brd=Border())
    for c in range(3, last_data_col + 1):
        sc(ws, row, c, font=f_source, fill=fl_white, brd=Border())
    ws.merge_cells(f"B{row}:{get_column_letter(last_data_col)}{row}")

    # Column headers
    row = 5
    ws.row_dimensions[row].height = 32
    sc(ws, row, 2, "Keyword", f_hdr, fl_lblue, align=ac, brd=hdr_brd)
    sc(ws, row, 3, "Avg\nVolume", f_hdr, fl_lblue, align=ac, brd=hdr_brd)
    sc(ws, row, 4, "CPC\n(\u00a5)", f_hdr, fl_lblue, align=ac, brd=hdr_brd)
    for i, label in enumerate(SV_MONTH_LABELS):
        sc(ws, row, 5 + i, label, f_hdr, fl_lblue, align=ac, brd=hdr_brd)

    # Fetch Google volume data
    print("\n  Fetching Google search volumes...")
    google_data = {}
    try:
        google_data = fetch_google_volume_historical(
            SV_KEYWORDS, location="JP",
            date_from="2025-01-01", date_to="2026-02-28"
        )
        print(f"  Got data for {len(google_data)} keywords")
    except Exception as e:
        print(f"  [ERROR] Google volume fetch failed: {e}")

    # Data rows
    google_data_start = 6
    for ki, kw in enumerate(SV_KEYWORDS):
        row = google_data_start + ki
        bg = fl_white if ki % 2 == 0 else fl_lgray

        sc(ws, row, 2, kw, f_data_b, bg, align=al, brd=thin_border)

        kw_data = google_data.get(kw.lower(), {})
        avg_vol = kw_data.get("avg")
        cpc_val = kw_data.get("cpc")
        monthly = kw_data.get("monthly", {})

        sc(ws, row, 3, avg_vol, f_data_b, bg, NUM, ar, thin_border)

        cpc_cell = sc(ws, row, 4, cpc_val, f_data, bg, align=ar, brd=thin_border)
        if cpc_val is not None:
            cpc_cell.number_format = '\u00a5#,##0.00'

        for mi, (yr, mo) in enumerate(SV_MONTHS):
            vol = monthly.get((yr, mo))
            sc(ws, row, 5 + mi, vol, f_data, bg, NUM, ar, thin_border)

        if avg_vol:
            print(f"    {kw}: avg={avg_vol}, cpc={cpc_val}")
        else:
            print(f"    {kw}: no data")

    google_data_end = google_data_start + len(SV_KEYWORDS) - 1

    # ── Section B: AMAZON SEARCH VOLUME ───────────────────────────
    row = google_data_end + 2
    sc(ws, row, 2, "AMAZON SEARCH VOLUME", f_sec,
       PatternFill("solid", fgColor=ORANGE_BG), align=al, brd=thin_border)
    for c in range(3, last_data_col + 1):
        sc(ws, row, c, fill=PatternFill("solid", fgColor=ORANGE_BG),
           font=f_sec, brd=thin_border)
    ws.merge_cells(f"B{row}:{get_column_letter(last_data_col)}{row}")
    ws.row_dimensions[row].height = 26

    # Source line
    row += 1
    sc(ws, row, 2, "Source: DataForSEO Amazon Labs API | Location: US (JP not available)",
       f_source, fl_white, align=al, brd=Border())
    for c in range(3, last_data_col + 1):
        sc(ws, row, c, font=f_source, fill=fl_white, brd=Border())
    ws.merge_cells(f"B{row}:{get_column_letter(last_data_col)}{row}")

    # Column headers
    row += 1
    ws.row_dimensions[row].height = 28
    sc(ws, row, 2, "Keyword", f_hdr, fl_lblue, align=ac, brd=hdr_brd)
    sc(ws, row, 3, "Search\nVolume", f_hdr, fl_lblue, align=ac, brd=hdr_brd)
    for c in range(4, last_data_col + 1):
        sc(ws, row, c, fill=fl_lblue, font=f_hdr, brd=hdr_brd)

    # Fetch Amazon volume data (JP not supported, fallback to US)
    print("\n  Fetching Amazon search volumes (US market)...")
    amazon_results = []
    try:
        amazon_results = fetch_amazon_volume(SV_KEYWORDS, location="US")
        print(f"  Got {len(amazon_results)} results (US market, JP not available)")
    except Exception as e:
        print(f"  [ERROR] Amazon volume fetch failed: {e}")

    amazon_map = {r["keyword"].lower(): r for r in amazon_results}

    # Data rows
    amz_data_start = row + 1
    for ki, kw in enumerate(SV_KEYWORDS):
        row = amz_data_start + ki
        bg = fl_white if ki % 2 == 0 else fl_lgray

        sc(ws, row, 2, kw, f_data_b, bg, align=al, brd=thin_border)

        amz = amazon_map.get(kw.lower(), {})
        vol = amz.get("search_volume")
        sc(ws, row, 3, vol, f_data_b, bg, NUM, ar, thin_border)

        for c in range(4, last_data_col + 1):
            sc(ws, row, c, fill=bg, brd=thin_border)

        print(f"    {kw}: {vol if vol else 'no data'}")

    amz_data_end = amz_data_start + len(SV_KEYWORDS) - 1

    # ── Section C: COMPARISON (Search Volume vs Marketing Performance) ──
    row = amz_data_end + 2
    sc(ws, row, 2, "SEARCH VOLUME vs MARKETING PERFORMANCE", f_sec,
       fl_blue, align=al, brd=thin_border)
    for c in range(3, last_data_col + 1):
        sc(ws, row, c, fill=fl_blue, font=f_sec, brd=thin_border)
    ws.merge_cells(f"B{row}:{get_column_letter(last_data_col)}{row}")
    ws.row_dimensions[row].height = 26

    # Source line
    row += 1
    sc(ws, row, 2, "Google Volume = grosmimi monthly | Spend/Sales from subsidiary tabs",
       f_source, fl_white, align=al, brd=Border())
    for c in range(3, last_data_col + 1):
        sc(ws, row, c, font=f_source, fill=fl_white, brd=Border())
    ws.merge_cells(f"B{row}:{get_column_letter(last_data_col)}{row}")

    # Column headers (matching the 8 month columns)
    row += 1
    ws.row_dimensions[row].height = 32
    sc(ws, row, 2, "Metric", f_hdr, fl_lblue, align=ac, brd=hdr_brd)
    sc(ws, row, 3, "", f_hdr, fl_lblue, align=ac, brd=hdr_brd)
    sc(ws, row, 4, "", f_hdr, fl_lblue, align=ac, brd=hdr_brd)
    for i, label in enumerate(SV_MONTH_LABELS):
        sc(ws, row, 5 + i, label, f_hdr, fl_lblue, align=ac, brd=hdr_brd)

    comp_hdr_row = row

    # Comparison data rows with MoM change
    # Each metric gets: value row + MoM Δ% row
    grosmimi_row = google_data_start  # row 6

    # Font/fill for MoM change rows
    f_mom = Font(name="Aptos", size=9, italic=True, color="666666")
    fl_mom = PatternFill("solid", fgColor="F0F4F8")
    MOM_FMT = '+0.0%;-0.0%;\u2014'

    metrics = [
        ("Google Vol (grosmimi)", NUM),
        ("Google Vol (\u30b9\u30c8\u30ed\u30fc\u30de\u30b0)", NUM),
        ("Meta Ad Spend", YEN),
        ("Rakuten Total Sales", YEN),
        ("Amazon Total Sales", YEN),
        ("Total Spend (All)", YEN),
    ]

    straw_row = google_data_start + 2  # ストローマグ = 3rd keyword (row 8)
    cur_row = comp_hdr_row + 1

    for metric_name, fmt in metrics:
        # ── Value row ──
        row = cur_row
        bg = fl_white

        sc(ws, row, 2, metric_name, f_metric, bg, align=al, brd=thin_border)
        sc(ws, row, 3, "", f_data, bg, brd=thin_border)
        sc(ws, row, 4, "", f_data, bg, brd=thin_border)

        for col_i in range(len(SV_MONTHS)):
            col = 5 + col_i
            month_mi = col_i
            tr = _total_row(month_mi)

            if metric_name == "Google Vol (grosmimi)":
                ref = f"={get_column_letter(col)}{grosmimi_row}"
                sc(ws, row, col, ref, f_data, bg, fmt, ar, thin_border)
            elif metric_name == "Google Vol (\u30b9\u30c8\u30ed\u30fc\u30de\u30b0)":
                ref = f"={get_column_letter(col)}{straw_row}"
                sc(ws, row, col, ref, f_data, bg, fmt, ar, thin_border)
            elif metric_name == "Meta Ad Spend":
                sc(ws, row, col, f"=META!E{tr}", f_data, bg, fmt, ar, thin_border)
            elif metric_name == "Rakuten Total Sales":
                sc(ws, row, col, f"=RAKUTEN!L{tr}", f_data, bg, fmt, ar, thin_border)
            elif metric_name == "Amazon Total Sales":
                sc(ws, row, col, f"=AMAZON!L{tr}", f_data, bg, fmt, ar, thin_border)
            elif metric_name == "Total Spend (All)":
                sc(ws, row, col,
                   f"=RAKUTEN!E{tr}+AMAZON!E{tr}+META!E{tr}",
                   f_tot_dat, fl_total, fmt, ar, thin_border)

        cur_row += 1

        # ── MoM Δ% row ──
        mom_row = cur_row
        sc(ws, mom_row, 2, "  MoM \u0394%", f_mom, fl_mom, align=al, brd=thin_border)
        sc(ws, mom_row, 3, "", f_mom, fl_mom, brd=thin_border)
        sc(ws, mom_row, 4, "", f_mom, fl_mom, brd=thin_border)

        for col_i in range(len(SV_MONTHS)):
            col = 5 + col_i
            if col_i == 0:
                # First month: no previous to compare
                sc(ws, mom_row, col, "\u2014", f_mom, fl_mom, align=ar, brd=thin_border)
            else:
                prev_cl = get_column_letter(col - 1)
                cur_cl = get_column_letter(col)
                formula = f'=IFERROR(({cur_cl}{row}-{prev_cl}{row})/{prev_cl}{row},"\u2014")'
                sc(ws, mom_row, col, formula, f_mom, fl_mom, MOM_FMT, ar, thin_border)

        cur_row += 1

    last_row = cur_row - 1

    # ── Conditional formatting on MoM rows ────────────────────────
    gf_pos = PatternFill("solid", fgColor="E8F5E9")
    gn_pos = Font(name="Aptos", size=9, bold=True, italic=True, color="2E7D32")
    rf_neg = PatternFill("solid", fgColor="FFEBEE")
    rn_neg = Font(name="Aptos", size=9, bold=True, italic=True, color="C62828")

    # Apply to all MoM rows
    mom_range = f"E{comp_hdr_row+2}:{get_column_letter(last_data_col)}{last_row}"
    ws.conditional_formatting.add(mom_range,
        CellIsRule(operator="greaterThan", formula=["0"], fill=gf_pos, font=gn_pos))
    ws.conditional_formatting.add(mom_range,
        CellIsRule(operator="lessThan", formula=["0"], fill=rf_neg, font=rn_neg))

    # ── Conditional formatting on volume columns ──────────────────
    vol_range = f"E{google_data_start}:{get_column_letter(last_data_col)}{google_data_end}"
    gf = PatternFill("solid", fgColor=GREEN_CF)
    gn = Font(name="Aptos", size=10, bold=True, color="2E7D32")
    af = PatternFill("solid", fgColor=AMBER_CF)
    an = Font(name="Aptos", size=10, bold=True, color="F57F17")

    ws.conditional_formatting.add(vol_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["1000"], fill=gf, font=gn))
    ws.conditional_formatting.add(vol_range,
        CellIsRule(operator="between", formula=["100", "999"], fill=af, font=an))

    # ── Footnote ──────────────────────────────────────────────────
    nr = last_row + 2
    sc(ws, nr, 2, None, f_note, brd=Border())
    for c in range(3, last_data_col + 1):
        sc(ws, nr, c, font=f_note, brd=Border())
    ws.merge_cells(f"B{nr}:{get_column_letter(last_data_col)}{nr}")
    ws.cell(row=nr, column=2).value = (
        "Google Volume: DataForSEO Google Ads API (monthly search volume). "
        "Amazon Volume: DataForSEO Amazon Labs API. "
        "Comparison section references TOTAL rows from RAKUTEN/AMAZON/META tabs. "
        "Updated: " + datetime.now().strftime("%Y-%m-%d %H:%M"))
    ws.cell(row=nr, column=2).font = f_note
    ws.cell(row=nr, column=2).alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[nr].height = 28

    # ── Setup ─────────────────────────────────────────────────────
    ws.freeze_panes = "C6"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    print(f"\n  [OK] Search Volume tab rebuilt")


# ===================================================================
#  RAKUTEN: Fill weekly total sales
# ===================================================================

def update_rakuten(wb, year, month, week=None):
    """Fill RAKUTEN tab with API sales data."""
    ws = wb["RAKUTEN"]
    mi = MONTH_INDEX.get(f"{year}-{month:02d}")
    if mi is None:
        print(f"  [SKIP] {year}-{month:02d} not in dashboard range")
        return

    bs = _block_start(mi)
    weeks = range(1, 6) if week is None else [week]

    for w in weeks:
        date_from, date_to = _week_dates(year, month, w)
        if date_from is None:
            continue

        r = bs + (w - 1)

        print(f"\n  RAKUTEN W{w}: {date_from} ~ {date_to}")
        result = rakuten_weekly_sales(date_from, date_to)

        # L column (col 12) = Total Sales
        ws.cell(row=r, column=12).value = result["total_sales"]
        print(f"    \u2192 L{r} = \u00a5{result['total_sales']:,} ({result['order_count']} orders)")

    print(f"\n  [OK] RAKUTEN {year}-{month:02d} updated")


# ===================================================================
#  AMAZON: Fill weekly total sales (SP-API)
# ===================================================================

def update_amazon(wb, year, month, week=None):
    """Fill AMAZON tab with SP-API order sales data.

    Same structure as RAKUTEN: weekly rows, L column = Total Sales.
    """
    ws = wb["AMAZON"]
    mi = MONTH_INDEX.get(f"{year}-{month:02d}")
    if mi is None:
        print(f"  [SKIP] {year}-{month:02d} not in dashboard range")
        return

    bs = _block_start(mi)
    weeks = range(1, 6) if week is None else [week]

    for w in weeks:
        date_from, date_to = _week_dates(year, month, w)
        if date_from is None:
            continue

        r = bs + (w - 1)

        print(f"\n  AMAZON W{w}: {date_from} ~ {date_to}")
        result = amazon_weekly_sales(date_from, date_to)

        # L column (col 12) = Total Sales
        ws.cell(row=r, column=12).value = result["total_sales"]
        print(f"    \u2192 L{r} = \u00a5{result['total_sales']:,} ({result['order_count']} orders)")

    print(f"\n  [OK] AMAZON {year}-{month:02d} updated")


# ===================================================================
#  META: Fill campaign-level monthly data
# ===================================================================

def update_meta(wb, year, month):
    """Fill META tab with campaign-level data from Meta API (JP account)."""
    ws = wb["META"]
    mi = MONTH_INDEX.get(f"{year}-{month:02d}")
    if mi is None:
        print(f"  [SKIP] {year}-{month:02d} not in dashboard range")
        return

    bs = _block_start(mi)

    last_day = calendar.monthrange(year, month)[1]
    today = datetime.now()

    if year == today.year and month == today.month:
        end_day = today.day
    else:
        end_day = last_day

    date_from = f"{year}-{month:02d}-01"
    date_to = f"{year}-{month:02d}-{end_day:02d}"

    print(f"\n  META campaigns: {date_from} ~ {date_to}")
    campaigns = get_campaign_insights(date_from, date_to)

    if not campaigns:
        print("    No campaign data found")
        return

    campaigns.sort(key=lambda c: float(c.get("spend", 0)), reverse=True)

    for ci, camp in enumerate(campaigns[:5]):
        r = bs + ci
        name = camp.get("campaign_name", "Unknown")
        spend = float(camp.get("spend", 0))
        impr = int(camp.get("impressions", 0))
        reach = int(camp.get("reach", 0)) if camp.get("reach") else 0
        clicks = int(camp.get("clicks", 0))
        freq = float(camp.get("frequency", 0)) if camp.get("frequency") else 0

        ws.cell(row=r, column=3).value = name     # C: Campaign name
        ws.cell(row=r, column=5).value = spend     # E: Ad Spend
        ws.cell(row=r, column=6).value = impr      # F: Impressions
        ws.cell(row=r, column=7).value = reach     # G: Reach
        ws.cell(row=r, column=8).value = clicks    # H: Link Clicks
        ws.cell(row=r, column=11).value = freq     # K: Frequency

        print(f"    [{ci+1}] {name}: \u00a5{spend:,.0f} | {impr:,} impr | {clicks:,} clicks")

    # Clear remaining campaign slots
    for ci in range(len(campaigns), 5):
        r = bs + ci
        for col in [3, 5, 6, 7, 8, 11]:
            ws.cell(row=r, column=col).value = None

    print(f"\n  [OK] META {year}-{month:02d} updated ({len(campaigns[:5])} campaigns)")


# ===================================================================
#  MAIN
# ===================================================================

def update(year=None, month=None, week=None, search_volume=False, all_months=False):
    """Run the full update."""
    if year is None or month is None:
        now = datetime.now()
        year, month = now.year, now.month

    print("=" * 60)
    print(f"V6 Dashboard Update \u2014 {year}-{month:02d}")
    if all_months:
        print("  Mode: ALL MONTHS (Nov 2025 ~ current)")
    if search_volume:
        print("  + Search Volume tab rebuild")
    print("=" * 60)

    if not EXCEL_PATH.exists():
        print(f"[ERROR] V6 file not found: {EXCEL_PATH}")
        print("  Run 'python tools/rebuild_v6.py' first.")
        return

    wb = load_workbook(str(EXCEL_PATH))

    # ── Search Volume tab ──
    if search_volume:
        print(f"\n{'─'*40}")
        print("SEARCH VOLUME \u2014 Keyword Analysis")
        print(f"{'─'*40}")
        try:
            update_search_volume(wb)
        except Exception as e:
            print(f"  [ERROR] Search Volume failed: {e}")
            import traceback
            traceback.print_exc()

    # Determine months to update
    if all_months:
        months_to_update = []
        for key in MONTH_INDEX:
            y, m = key.split("-")
            y, m = int(y), int(m)
            # Only update months up to current month
            if (y < year) or (y == year and m <= month):
                months_to_update.append((y, m))
    else:
        months_to_update = [(year, month)]

    for upd_year, upd_month in months_to_update:
        upd_week = week if (upd_year == year and upd_month == month) else None

        # ── Rakuten: weekly sales ──
        print(f"\n{'─'*40}")
        print(f"RAKUTEN \u2014 {upd_year}-{upd_month:02d}")
        print(f"{'─'*40}")
        try:
            update_rakuten(wb, upd_year, upd_month, upd_week)
        except Exception as e:
            print(f"  [ERROR] Rakuten failed: {e}")

        # ── Amazon: weekly sales ──
        print(f"\n{'─'*40}")
        print(f"AMAZON \u2014 {upd_year}-{upd_month:02d}")
        print(f"{'─'*40}")
        try:
            update_amazon(wb, upd_year, upd_month, upd_week)
        except Exception as e:
            print(f"  [ERROR] Amazon failed: {e}")

        # ── Meta: campaign data ──
        print(f"\n{'─'*40}")
        print(f"META \u2014 {upd_year}-{upd_month:02d}")
        print(f"{'─'*40}")
        try:
            update_meta(wb, upd_year, upd_month)
        except Exception as e:
            print(f"  [ERROR] Meta failed: {e}")

    # ── Save ──
    wb.save(str(EXCEL_PATH))
    print(f"\n{'='*60}")
    print(f"[OK] Saved: {EXCEL_PATH.name}")
    print(f"{'='*60}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Update V6 dashboard with API data")
    parser.add_argument("month", nargs="?", help="YYYY-MM (default: current month)")
    parser.add_argument("--week", "-w", type=int, help="Update specific week only (1-5)")
    parser.add_argument("--search-volume", "-s", action="store_true",
                        help="Rebuild Search Volume tab")
    parser.add_argument("--all", "-a", action="store_true",
                        help="Update all months (Nov 2025 ~ current)")
    args = parser.parse_args()

    year, month = None, None
    if args.month:
        parts = args.month.split("-")
        year, month = int(parts[0]), int(parts[1])

    update(year, month, args.week,
           search_volume=args.search_volume,
           all_months=args.all)
