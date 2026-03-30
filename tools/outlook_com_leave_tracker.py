"""
연차 자동 추적 시스템 (Outlook COM 기반)

Outlook 데스크탑 앱을 직접 제어합니다 - Azure AD / OAuth 불필요.
로그인된 Outlook 계정을 그대로 사용합니다.

Prerequisites:
    uv pip install pywin32 anthropic openpyxl holidays

    ~/.wat_secrets:
        OUTLOOK_EMAIL=dk.shin@orbiters.co.kr
        LEAVE_TEAM_EMAILS=dk.shin@orbiters.co.kr,hr@orbiters.co.kr
        ANTHROPIC_API_KEY=sk-ant-...

Usage:
    python tools/outlook_com_leave_tracker.py --setup
    python tools/outlook_com_leave_tracker.py --sync
    python tools/outlook_com_leave_tracker.py --report [--dry-run]
    python tools/outlook_com_leave_tracker.py --auto
"""

import argparse
import json
import os
import re
import sys
from datetime import date as date_cls, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from env_loader import load_env
load_env()

# ── Config ────────────────────────────────────────────────────────────────────
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
OUTLOOK_EMAIL     = os.getenv("OUTLOOK_EMAIL", "dk.shin@orbiters.co.kr")
LEAVE_TEAM_EMAILS = [
    e for e in {
        *[e.strip() for e in os.getenv("LEAVE_TEAM_EMAILS", "").split(",") if e.strip()],
        "wj.choi@orbiters.co.kr",
        "mj.lee@orbiters.co.kr",
    }
    if e.lower() != "dk.shin@orbiters.co.kr"
]

BASE_DIR          = Path(__file__).parent.parent
PROCESSED_PATH    = BASE_DIR / ".tmp" / "leave_com_processed.json"
SENT_NOTIF_PATH   = BASE_DIR / ".tmp" / "leave_sent_notifications.json"
INTERNAL_DOMAIN   = "orbiters.co.kr"  # 이 도메인 외 발송 금지
EXCEL_DIR         = Path(r"//Orbiters/경영지원/연차관리")
ANNUAL_LEAVE_DAYS = 15  # 기본값 (입사일 미등록 시 fallback)
EMPLOYEE_DATA_PATH = Path(__file__).parent / "employee_data.json"
MONTHS_KR         = ["1월","2월","3월","4월","5월","6월","7월","8월","9월","10월","11월","12월"]
LEAVE_KEYWORDS    = ["연차", "휴가", "반차", "병가", "공가", "예비군", "보건휴가", "생리휴가",
                     "결혼", "경조사", "장례", "출산", "난임", "사망",
                     "취소", "철회"]

# 연차에서 차감되는 유형
ANNUAL_DEDUCT_TYPES = {"연차", "반차"}

# 연간 한도가 있는 비연차 휴가 (취업규칙 기준)
# annual_limit: 연간 최대 일수 (None=무제한)
# paid_days: 유급 처리되는 일수 (None=전부 유급, 0=전부 무급)
LEAVE_POLICY = {
    "본인결혼":     {"label": "본인 결혼",         "annual_limit": 5,    "paid_days": 5},
    "배우자출산":   {"label": "배우자 출산",        "annual_limit": 10,   "paid_days": 10},
    "배우자사망":   {"label": "배우자/부모 사망",   "annual_limit": 5,    "paid_days": 5},
    "조부모사망":   {"label": "조부모/외조부모 사망","annual_limit": 3,   "paid_days": 3},
    "자녀사망":     {"label": "자녀/자녀배우자 사망","annual_limit": 3,   "paid_days": 3},
    "형제사망":     {"label": "형제자매 사망",      "annual_limit": 1,    "paid_days": 1},
    "난임치료":     {"label": "난임치료휴가",        "annual_limit": 6,    "paid_days": 2},
    "병가":         {"label": "병가",               "annual_limit": 30,   "paid_days": 0},
    "생리휴가":     {"label": "생리휴가",            "annual_limit": None, "paid_days": 0},
    "공가":         {"label": "공가",               "annual_limit": None, "paid_days": None},
    "예비군":       {"label": "예비군/민방위",       "annual_limit": None, "paid_days": None},
    "경조사":       {"label": "경조사",             "annual_limit": None, "paid_days": None},  # 세분화 불가 시 fallback
}

# ── 법정 연차 발생 계산 (근로기준법 제60조) ───────────────────────────────────
def _load_employee_data() -> dict:
    if EMPLOYEE_DATA_PATH.exists():
        return json.loads(EMPLOYEE_DATA_PATH.read_text(encoding="utf-8"))
    return {}

def _load_hire_dates() -> dict:
    return _load_employee_data().get("hire_dates", {})

def get_person_email(name: str) -> str:
    """이름으로 등록된 이메일 반환. 미등록 시 빈 문자열."""
    return _load_employee_data().get("emails", {}).get(name, "")


def calc_entitlement(hire_date_str: str, as_of: "date_cls") -> float:
    """
    근로기준법 기준 발생 연차 계산.
      - 1년 미만: 완성된 근속 월 수 (최대 11일)
      - 1년 이상: 15 + floor((완성연수 - 1) / 2), 최대 25일
    """
    hire = datetime.strptime(hire_date_str, "%Y-%m-%d").date()
    if as_of < hire:
        return 0.0
    full_months = (as_of.year - hire.year) * 12 + (as_of.month - hire.month)
    if as_of.day < hire.day:
        full_months -= 1
    full_months = max(full_months, 0)
    full_years  = full_months // 12
    if full_years < 1:
        return float(min(full_months, 11))
    return float(min(15 + (full_years - 1) // 2, 25))


def get_person_annual_leave(name: str, ref_year: "int | None" = None) -> float:
    """
    회계연도 기준 발생 연차 반환.
    ref_year 의 12월 31일을 기준으로 계산 → 해당 연도 내 최대 발생치(고정값).
    미등록 시 기본값(ANNUAL_LEAVE_DAYS) 반환.
    """
    if ref_year is None:
        ref_year = date_cls.today().year
    as_of = date_cls(ref_year, 12, 31)
    hire_dates = _load_hire_dates()
    hire_str   = hire_dates.get(name)
    if hire_str:
        return calc_entitlement(hire_str, as_of)
    return float(ANNUAL_LEAVE_DAYS)


# ── Excel 스타일 ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TOTAL_FILL  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
CANCEL_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
ALT_FILL    = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
NO_FILL     = PatternFill(fill_type=None)
HEADER_FONT = Font(name="Calibri", size=10, bold=True)
DATA_FONT   = Font(name="Calibri", size=10)
TITLE_FONT  = Font(name="Calibri", size=12, bold=True)
THIN        = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ── Outlook COM ───────────────────────────────────────────────────────────────
def get_outlook():
    try:
        import win32com.client
        return win32com.client.Dispatch("Outlook.Application")
    except ImportError:
        print("ERROR: pywin32 미설치 → uv pip install pywin32")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: Outlook 실행 실패 → {e}")
        sys.exit(1)


def fetch_leave_emails(year: int) -> list:
    """Outlook COM으로 받은편지함에서 연차 키워드 이메일 수집."""
    outlook   = get_outlook()
    namespace = outlook.GetNamespace("MAPI")
    inbox     = namespace.GetDefaultFolder(6)  # olFolderInbox

    emails = []
    items  = inbox.Items
    items.Sort("[ReceivedTime]", True)  # 최신순

    item = items.GetFirst()
    while item:
        try:
            received = item.ReceivedTime
            recv_year = received.year if hasattr(received, "year") else int(str(received)[:4])

            # 내림차순 정렬이므로 target year보다 오래된 항목은 이후 없음
            if recv_year < year:
                break
            if recv_year != year:
                item = items.GetNext()
                continue

            subject = item.Subject or ""
            body    = item.Body    or ""

            if any(kw in (subject + body) for kw in LEAVE_KEYWORDS):
                recv_str = received.strftime("%Y-%m-%d") if hasattr(received, "strftime") else str(received)[:10]

                # Exchange DN → SMTP 변환 시도
                raw_email = item.SenderEmailAddress or ""
                smtp_email = raw_email
                if raw_email.upper().startswith("/O="):
                    try:
                        ex_user = item.Sender.GetExchangeUser()
                        if ex_user:
                            smtp_email = ex_user.PrimarySmtpAddress or raw_email
                    except Exception:
                        pass

                emails.append({
                    "id":           item.EntryID,
                    "subject":      subject,
                    "sender_name":  item.SenderName or "",
                    "sender_email": smtp_email,
                    "body":         body,
                    "received_at":  recv_str,
                })
        except Exception:
            pass
        item = items.GetNext()

    return emails


SENDER_EMAIL = "dk.shin@orbiters.co.kr"


def _is_internal(email: str) -> bool:
    """오비터스 내부 도메인(@orbiters.co.kr) 여부 확인."""
    return email.strip().lower().endswith(f"@{INTERNAL_DOMAIN}")


def save_draft_email(subject: str, html_body: str, recipients: list,
                     cc: list | None = None) -> bool:
    """Outlook 임시저장함에 초안으로 저장. 즉시 발송하지 않음.
    recipients → To (직원 본인)
    cc         → CC 참조 (팀 메일 등)

    제한사항:
    - 발신자: dk.shin@orbiters.co.kr 고정
    - 수신자/참조: @orbiters.co.kr 내부 직원만 허용, 외부 주소 자동 차단
    - 자동 발송 없음 - 사용자가 임시저장함에서 직접 확인 후 발송
    """
    # ── 내부 직원 필터링 ────────────────────────────────────────────
    safe_to = [r for r in recipients if _is_internal(r)]
    safe_cc = [r for r in (cc or []) if _is_internal(r)]

    blocked = [r for r in recipients if not _is_internal(r)]
    blocked += [r for r in (cc or []) if not _is_internal(r)]
    if blocked:
        print(f"  [외부 차단] 내부 도메인 아닌 주소 제외: {', '.join(blocked)}")

    if not safe_to:
        print(f"  [임시저장 차단] 유효한 내부 수신자 없음 - 건너뜀")
        return False

    try:
        outlook = get_outlook()
        mail    = outlook.CreateItem(0)  # olMailItem
        mail.Subject  = subject
        mail.HTMLBody = html_body
        mail.To       = "; ".join(safe_to)
        if safe_cc:
            mail.CC = "; ".join(safe_cc)

        # 발신 계정 고정: dk.shin@orbiters.co.kr
        ns = outlook.GetNamespace("MAPI")
        for account in ns.Accounts:
            if account.SmtpAddress.lower() == SENDER_EMAIL.lower():
                mail.SendUsingAccount = account
                break

        mail.Save()  # 임시저장함 저장 (발송 안 함)
        print(f"  임시저장 완료: {subject}")
        print(f"  → Outlook 임시저장함에서 확인 후 직접 발송하세요.")
        return True
    except Exception as e:
        print(f"  [임시저장 오류] {e}")
        return False


def _build_add_email_html(name: str, parsed_list: list,
                          alloc: float = 0, total_used: float = 0, remaining: float = 0) -> str:
    """연차 신청 확인 이메일 HTML 생성."""
    p          = parsed_list[0]
    leave_type = p["leave_type"]
    note       = p.get("note", "")
    total_days = sum(seg["days"] for seg in parsed_list)
    days_val   = int(total_days) if total_days == int(total_days) else total_days

    def _fmt(v: float) -> str:
        return str(int(v)) if v == int(v) else str(v)

    date_rows = ""
    for seg in parsed_list:
        date_rows += _date_rows_html(seg["start_date"], seg["end_date"], seg["days"])

    note_row = ""
    if note:
        note_row = (
            f'<tr><td style="padding:7px 14px;border:1px solid #E5E7EB;'
            f'color:#555;width:130px;">비고</td>'
            f'<td style="padding:7px 14px;border:1px solid #E5E7EB;">{note}</td></tr>'
        )

    # 잔여 현황 테이블 (연차/반차일 때만 표시)
    remaining_section = ""
    if leave_type in ANNUAL_DEDUCT_TYPES and alloc > 0:
        before_remaining = remaining + days_val  # 이번 신청 전 잔여
        remaining_section = (
            f'<h3 style="color:#1565C0;margin-top:20px;font-size:13px;">잔여 현황</h3>'
            f'<table style="border-collapse:collapse;font-size:13px;">'
            f'<tr>'
            f'<th style="padding:7px 20px;border:1px solid #E5E7EB;background:#D9E2F3;text-align:center;">현재 잔여 연차</th>'
            f'<th style="padding:7px 20px;border:1px solid #E5E7EB;background:#D9E2F3;text-align:center;">사용 연차</th>'
            f'<th style="padding:7px 20px;border:1px solid #E5E7EB;background:#D9E2F3;text-align:center;">최종 잔여 연차</th>'
            f'</tr>'
            f'<tr>'
            f'<td style="padding:7px 20px;border:1px solid #E5E7EB;text-align:center;">{_fmt(before_remaining)}일</td>'
            f'<td style="padding:7px 20px;border:1px solid #E5E7EB;text-align:center;">{_fmt(days_val)}일</td>'
            f'<td style="padding:7px 20px;border:1px solid #E5E7EB;text-align:center;background:#E2EFDA;"><strong>{_fmt(remaining)}일</strong></td>'
            f'</tr>'
            f'</table>'
        )

    return (
        f'<!DOCTYPE html><html><head><meta charset="utf-8"></head><body '
        f'style="font-family:Malgun Gothic,Arial,sans-serif;font-size:14px;color:#333;margin:20px;">'
        f'<h2 style="color:#1565C0;">연차 신청 접수 확인</h2>'
        f'<p>안녕하세요, <strong>{name}</strong>님.<br>'
        f'아래 내용으로 연차 신청이 접수되었습니다.</p>'
        f'<table style="border-collapse:collapse;font-size:13px;margin-top:10px;">'
        f'<tr><td style="padding:7px 14px;border:1px solid #E5E7EB;color:#555;width:130px;">신청자</td>'
        f'<td style="padding:7px 14px;border:1px solid #E5E7EB;"><strong>{name}</strong></td></tr>'
        f'<tr><td style="padding:7px 14px;border:1px solid #E5E7EB;background:#F9FAFB;'
        f'color:#555;width:130px;">휴가 유형</td>'
        f'<td style="padding:7px 14px;border:1px solid #E5E7EB;background:#F9FAFB;">'
        f'<strong>{leave_type}</strong></td></tr>'
        f'{date_rows}'
        f'<tr><td style="padding:7px 14px;border:1px solid #E5E7EB;background:#F9FAFB;'
        f'color:#555;width:130px;">총 사용일수</td>'
        f'<td style="padding:7px 14px;border:1px solid #E5E7EB;background:#F9FAFB;">'
        f'<strong>{days_val}일</strong></td></tr>'
        f'{note_row}'
        f'</table>'
        f'{remaining_section}'
        f'<p style="margin-top:16px;font-size:12px;color:#999;">'
        f'※ 본 메일은 연차 관리 시스템에서 자동 생성된 초안입니다.</p>'
        f'</body></html>'
    )


def _build_cancel_email_html(name: str, cancelled_dates: list, remaining: float = 0) -> str:
    """연차 취소 처리 확인 이메일 HTML 생성."""
    date_items = "".join(
        f'<li>{_fmt_date_kr(d)}</li>' for d in sorted(cancelled_dates)
    ) if cancelled_dates else "<li>(날짜 정보 없음)</li>"

    def _fmt(v: float) -> str:
        return str(int(v)) if v == int(v) else str(v)

    remaining_row = (
        f'<p style="margin-top:14px;font-size:13px;">'
        f'현재 잔여 연차: <strong>{_fmt(remaining)}일</strong> (취소분 복구 반영)</p>'
    ) if remaining > 0 else ""

    return (
        f'<!DOCTYPE html><html><head><meta charset="utf-8"></head><body '
        f'style="font-family:Malgun Gothic,Arial,sans-serif;font-size:14px;color:#333;margin:20px;">'
        f'<h2 style="color:#C62828;">연차 취소 처리 완료</h2>'
        f'<p>안녕하세요, <strong>{name}</strong>님.<br>'
        f'아래 날짜의 연차가 취소 처리되었습니다.</p>'
        f'<ul style="font-size:13px;line-height:1.8;">{date_items}</ul>'
        f'{remaining_row}'
        f'<p style="margin-top:12px;font-size:13px;">해당 일정이 정상적으로 복구되었음을 알려드립니다.</p>'
        f'<p style="margin-top:16px;font-size:12px;color:#999;">'
        f'※ 본 메일은 연차 관리 시스템에서 자동 생성된 초안입니다.</p>'
        f'</body></html>'
    )


# ── 처리 기록 ─────────────────────────────────────────────────────────────────
def load_processed_ids() -> set:
    PROCESSED_PATH.parent.mkdir(parents=True, exist_ok=True)
    if PROCESSED_PATH.exists():
        return set(json.loads(PROCESSED_PATH.read_text(encoding="utf-8")))
    return set()


def save_processed_ids(ids: set):
    PROCESSED_PATH.write_text(
        json.dumps(list(ids), ensure_ascii=False, indent=2), encoding="utf-8"
    )


# ── 발송 이력 (중복 발송 방지) ────────────────────────────────────────────────
def _notif_key(name: str, start: str, end: str, leave_type: str, action: str) -> str:
    """승인/취소 알림 이력 키 생성. 동일 키면 이미 발송된 것."""
    return f"{name}|{start[:10]}|{end[:10]}|{leave_type}|{action}"


def load_sent_notifications() -> set:
    SENT_NOTIF_PATH.parent.mkdir(parents=True, exist_ok=True)
    if SENT_NOTIF_PATH.exists():
        return set(json.loads(SENT_NOTIF_PATH.read_text(encoding="utf-8")))
    return set()


def save_sent_notifications(notifs: set):
    SENT_NOTIF_PATH.write_text(
        json.dumps(sorted(notifs), ensure_ascii=False, indent=2), encoding="utf-8"
    )


# ── 한국 공휴일 ───────────────────────────────────────────────────────────────
def get_kr_holidays(year: int) -> set:
    try:
        import holidays as hol_lib
        return set(hol_lib.KR(years=year).keys())
    except ImportError:
        # 패키지 없으면 고정 공휴일(음력 제외)만 사용
        return {
            date_cls(year,  1,  1), date_cls(year,  3,  1),
            date_cls(year,  5,  5), date_cls(year,  6,  6),
            date_cls(year,  8, 15), date_cls(year, 10,  3),
            date_cls(year, 10,  9), date_cls(year, 12, 25),
        }


def is_working_day(d: date_cls, kr_holidays: set) -> bool:
    return d.weekday() < 5 and d not in kr_holidays


def expand_date_range(from_str: str, to_str: str) -> list:
    start = datetime.strptime(from_str[:10], "%Y-%m-%d").date()
    end   = datetime.strptime(to_str[:10],   "%Y-%m-%d").date()
    return [start + timedelta(days=i) for i in range(max((end - start).days + 1, 1))]


def group_leave_dates(leave_dates: list, kr_holidays: set) -> list:
    """연속 영업일을 하나의 구간으로 묶음."""
    sorted_dates = sorted(set(d for d in leave_dates if is_working_day(d, kr_holidays)))
    if not sorted_dates:
        return []
    segments = []
    seg_start = seg_end = sorted_dates[0]
    count = 1
    for d in sorted_dates[1:]:
        gap_has_workday = any(
            is_working_day(seg_end + timedelta(days=i), kr_holidays)
            for i in range(1, (d - seg_end).days)
        )
        if not gap_has_workday:
            seg_end = d
            count += 1
        else:
            segments.append({"start_date": seg_start.isoformat(), "end_date": seg_end.isoformat(), "days": float(count)})
            seg_start = seg_end = d
            count = 1
    segments.append({"start_date": seg_start.isoformat(), "end_date": seg_end.isoformat(), "days": float(count)})
    return segments


# ── 키워드 폴백 파서 (API 불필요) ────────────────────────────────────────────
def _extract_dates_from_text(text: str) -> list:
    """텍스트에서 날짜를 추출. 여러 포맷 지원."""
    cur_year = datetime.now().year
    dates = []

    # ISO / 점 구분: 2026-03-12, 2026.03.12
    for m in re.finditer(r'(\d{4})[-./](\d{1,2})[-./](\d{1,2})', text):
        try:
            dates.append(date_cls(int(m.group(1)), int(m.group(2)), int(m.group(3))))
        except ValueError:
            pass

    # 한국식: (2026년) 3월 12일
    for m in re.finditer(r'(?:(\d{4})년\s*)?(\d{1,2})월\s*(\d{1,2})일', text):
        yr = int(m.group(1)) if m.group(1) else cur_year
        try:
            dates.append(date_cls(yr, int(m.group(2)), int(m.group(3))))
        except ValueError:
            pass

    # 슬래시: 3/12  (ISO 날짜의 일부인 경우 제외 - 앞뒤에 숫자/슬래시 없을 때만)
    for m in re.finditer(r'(?<![0-9/])(\d{1,2})/(\d{1,2})(?![0-9/])', text):
        try:
            dates.append(date_cls(cur_year, int(m.group(1)), int(m.group(2))))
        except ValueError:
            pass

    return sorted(set(dates))


def fallback_parse_leave(email: dict):
    """Claude API 없이 키워드+정규식으로 파싱 (폴백)."""
    text = (email["subject"] + " " + email["body"])

    # 휴가 관련 이메일인지 확인 (키워드 최소 1개 포함)
    LEAVE_KW = ["연차", "휴가", "반차", "병가", "공가", "예비군", "민방위",
                "보건휴가", "생리휴가", "결혼", "신혼", "경조사", "장례",
                "출산", "사망", "난임", "취소", "철회"]
    if not any(kw in text for kw in LEAVE_KW):
        return None

    # action 판별: 취소/철회 워딩 있으면 cancel
    action = "cancel" if any(kw in text for kw in ["취소", "철회"]) else "add"

    # leave_type 판별 (구체적인 것 먼저 - 취업규칙 기준)
    TYPE_MAP = [
        # 경조사 세분화 (사망 종류)
        ("배우자 출산", "배우자출산"),
        ("배우자출산",  "배우자출산"),
        ("난임",        "난임치료"),
        # 사망 구분 (조부모 > 배우자/부모 > 자녀 > 형제 순)
        ("조부모",      "조부모사망"),
        ("외조부모",    "조부모사망"),
        ("자녀",        "자녀사망"),   # "자녀 사망" 우선
        ("형제",        "형제사망"),
        ("자매",        "형제사망"),
        # 결혼/사망 일반
        ("신혼",        "본인결혼"),
        ("결혼",        "본인결혼"),
        ("사망",        "배우자사망"),
        ("장례",        "배우자사망"),
        ("경조사",      "경조사"),
        # 기타
        ("반차",        "반차"),
        ("병가",        "병가"),
        ("공가",        "공가"),
        ("민방위",      "예비군"),
        ("예비군",      "예비군"),
        ("생리휴가",    "생리휴가"),
        ("보건휴가",    "생리휴가"),
        ("연차",        "연차"),
        ("휴가",        "연차"),
    ]
    leave_type = "기타"
    for kw, lt in TYPE_MAP:
        if kw in text:
            leave_type = lt
            break

    # 반차 오전/오후 note
    note = ""
    if leave_type == "반차":
        if "오전" in text:
            note = "오전 반차"
        elif "오후" in text:
            note = "오후 반차"

    # 날짜 추출
    all_dates = _extract_dates_from_text(text)
    if not all_dates:
        return None

    # min~max 범위를 단일 date_range로 사용
    date_ranges = [{"from": all_dates[0].isoformat(), "to": all_dates[-1].isoformat()}]

    # 이후 로직은 Claude 파서와 동일
    all_expanded, mentioned_years = [], set()
    for dr in date_ranges:
        expanded = expand_date_range(dr["from"], dr["to"])
        all_expanded.extend(expanded)
        mentioned_years.update(d.year for d in expanded)

    kr_holidays = set()
    for yr in mentioned_years:
        kr_holidays |= get_kr_holidays(yr)

    name = email["sender_name"]

    if leave_type == "반차":
        working = [d for d in sorted(set(all_expanded)) if is_working_day(d, kr_holidays)]
        if not working:
            return None
        d = working[0]
        return [_make_entry(name, d.isoformat(), d.isoformat(), 0.5, leave_type, note, email, action)]

    segments = group_leave_dates(all_expanded, kr_holidays)
    if not segments:
        return None
    return [_make_entry(name, s["start_date"], s["end_date"], s["days"], leave_type, note, email, action)
            for s in segments]


# ── Claude AI 파싱 ────────────────────────────────────────────────────────────
def parse_leave_request(email: dict):
    """이메일에서 연차 정보를 추출. Claude API 우선, 실패 시 키워드 폴백."""
    cur_year = datetime.now().year

    # API 키 없으면 바로 폴백
    if not ANTHROPIC_API_KEY:
        print("  [폴백] API 키 미설정 → 키워드 파싱")
        return fallback_parse_leave(email)

    body_text = re.sub(r"\s+", " ", email["body"]).strip()[:3000]

    prompt = f"""다음 이메일이 휴가/부재 신청 이메일인지 분석해주세요.

발신자: {email['sender_name']} ({email['sender_email']})
제목: {email['subject']}
수신일: {email['received_at']}
본문:
{body_text}

휴가/부재 신청 또는 취소이면 아래 JSON만 응답 (다른 텍스트 없이):
{{
  "is_leave_request": true,
  "action": "add",
  "name": "신청자 이름 (발신자 이름 사용)",
  "leave_type": "연차",
  "note": "특이사항 (없으면 빈 문자열)",
  "date_ranges": [
    {{"from": "YYYY-MM-DD", "to": "YYYY-MM-DD"}}
  ]
}}

휴가/부재와 무관한 이메일이면:
{{"is_leave_request": false}}

action 기준:
- "add": 신규 휴가 신청
- "cancel": 기존 휴가 취소/철회 (취소, 철회, 변경 등의 워딩 포함)

leave_type 분류 기준 (반드시 아래 중 하나):
- "연차": 일반 연차휴가
- "반차": 반일 휴가 (note에 오전/오후 명시)
- "병가": 질병으로 인한 휴가
- "공가": 공무, 회사 공식 업무로 인한 부재
- "예비군": 예비군 훈련
- "보건휴가": 여성 보건휴가
- "결혼": 본인 결혼 및 신혼여행
- "경조사": 가족 경조사 (결혼 제외), 장례
- "기타": 위에 해당하지 않는 휴가

규칙:
- date_ranges: 이메일 언급 날짜 그대로 (공휴일/주말 포함 가능, 필터링은 시스템이 함)
- 단일 날짜면 from == to
- 연도 불명확 시 {cur_year}년 기준
- JSON만 응답, 마크다운 코드블록 없이"""

    try:
        import anthropic
        client   = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=500,
            messages=[{"role": "user", "content": prompt}],
        )
        text   = response.content[0].text.strip()
        text   = re.sub(r"^```[a-z]*\n?", "", text)
        text   = re.sub(r"\n?```$",        "", text)
        parsed = json.loads(text)
    except ImportError:
        print("  [폴백] anthropic 미설치 → 키워드 파싱")
        return fallback_parse_leave(email)
    except Exception as e:
        print(f"  [Claude 오류] {e} → 키워드 폴백")
        return fallback_parse_leave(email)

    if not parsed.get("is_leave_request"):
        return None

    name        = parsed.get("name") or email["sender_name"]
    action      = parsed.get("action", "add")
    leave_type  = parsed.get("leave_type", "연차")
    note        = parsed.get("note", "")
    date_ranges = parsed.get("date_ranges", [])
    if not date_ranges:
        return None

    all_dates, mentioned_years = [], set()
    for dr in date_ranges:
        expanded = expand_date_range(dr.get("from", ""), dr.get("to", dr.get("from", "")))
        all_dates.extend(expanded)
        mentioned_years.update(d.year for d in expanded)

    kr_holidays = set()
    for yr in mentioned_years:
        kr_holidays |= get_kr_holidays(yr)

    if leave_type == "반차":
        working = [d for d in sorted(set(all_dates)) if is_working_day(d, kr_holidays)]
        if not working:
            return None
        d = working[0]
        return [_make_entry(name, d.isoformat(), d.isoformat(), 0.5, leave_type, note, email, action)]

    segments = group_leave_dates(all_dates, kr_holidays)
    if not segments:
        return None
    return [_make_entry(name, s["start_date"], s["end_date"], s["days"], leave_type, note, email, action)
            for s in segments]


def _make_entry(name, start, end, days, leave_type, note, email, action="add"):
    return {
        "name":          name,
        "action":        action,
        "start_date":    start,
        "end_date":      end,
        "days":          days,
        "leave_type":    leave_type,
        "note":          note,
        "sender_email":  email["sender_email"],
        "sender_name":   email["sender_name"],
        "email_subject": email["subject"],
        "received_at":   email["received_at"],
        "date_ranges":   email.get("date_ranges", []),
    }


# ── Excel ─────────────────────────────────────────────────────────────────────
PERSON_HEADERS    = ["신청일", "이메일 제목", "휴가 시작", "휴가 종료", "사용일수", "유형", "비고", "상태"]
PERSON_COL_WIDTHS = [12, 40, 12, 12, 10, 8, 25, 8]
PERSON_DATA_ROW   = 4


def get_excel_path(year: int) -> Path:
    return EXCEL_DIR / f"연차관리_{year}.xlsx"


def safe_sheet_name(name: str) -> str:
    return re.sub(r'[\\/*?:\[\]]', "", name)[:31]


def setup_excel(year: int) -> Path:
    path = get_excel_path(year)
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    if path.exists():
        print(f"이미 존재: {path}")
        return path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    headers = ["이름", "이메일", "발생연차"] + MONTHS_KR + ["총사용", "잔여"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.border = THIN;      cell.alignment = CTR

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    for i in range(4, 16):
        ws.column_dimensions[get_column_letter(i)].width = 7
    ws.column_dimensions[get_column_letter(16)].width = 8
    ws.column_dimensions[get_column_letter(17)].width = 8
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"생성 완료: {path}")
    return path


def ensure_person_sheet(wb, name: str, email: str, year: int | None = None):
    if year is None:
        year = datetime.now().year
    sheet_name = safe_sheet_name(name)
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    ws["A1"] = f"{name} ({email})"
    ws["A1"].font = TITLE_FONT; ws["A1"].alignment = LFT
    ws.merge_cells("A1:H1")

    alloc = get_person_annual_leave(name, year)
    alloc_label = int(alloc) if alloc == int(alloc) else alloc
    ws["A2"] = f"{year}년 연차 사용 내역 (발생: {alloc_label}일, 회계연도 기준)"
    ws["A2"].font = DATA_FONT; ws["A2"].alignment = LFT
    ws.merge_cells("A2:H2")

    for col, h in enumerate(PERSON_HEADERS, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.border = THIN;      cell.alignment = CTR

    for i, w in enumerate(PERSON_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"
    return ws


def get_person_entries(ws) -> list:
    entries = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=PERSON_DATA_ROW, values_only=True), start=PERSON_DATA_ROW):
        if row[0] is None:
            continue
        entries.append({
            "row":        row_idx,
            "start_date": str(row[2]) if row[2] else "",
            "end_date":   str(row[3]) if row[3] else "",
            "leave_type": str(row[5]) if row[5] else "",
            "status":     str(row[7]) if len(row) > 7 and row[7] else "",
        })
    return entries


def is_duplicate(entries: list, parsed: dict) -> bool:
    return any(
        e["start_date"] == parsed.get("start_date", "") and
        e["end_date"]   == parsed.get("end_date",   "") and
        e["status"]     != "취소"
        for e in entries
    )


def append_person_entry(ws, parsed: dict):
    next_row = max(ws.max_row + 1, PERSON_DATA_ROW)
    row_fill = ALT_FILL if (next_row % 2 == 0) else NO_FILL
    values = [
        parsed.get("received_at",    ""),
        parsed.get("email_subject",  ""),
        parsed.get("start_date",     ""),
        parsed.get("end_date",       ""),
        parsed.get("days",      1.0),
        parsed.get("leave_type", "연차"),
        parsed.get("note",          ""),
        "",  # 상태 (정상)
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = DATA_FONT; cell.border = THIN; cell.fill = row_fill
        cell.alignment = CTR if col in (1, 3, 4, 5, 6, 8) else LFT


def cancel_person_entries(ws, date_ranges: list, kr_holidays: set) -> int:
    """date_ranges에 해당하는 활성 행을 '취소' 처리. 취소된 행 수 반환."""
    cancelled = 0
    cancel_dates = set()
    for dr in date_ranges:
        expanded = expand_date_range(dr.get("from", ""), dr.get("to", dr.get("from", "")))
        cancel_dates.update(d.isoformat() for d in expanded if is_working_day(d, kr_holidays))

    for row in ws.iter_rows(min_row=PERSON_DATA_ROW):
        start_val  = row[2].value
        status_val = row[7].value if len(row) > 7 else None
        if not start_val or str(status_val) == "취소":
            continue
        if str(start_val)[:10] in cancel_dates:
            for cell in row:
                cell.fill = CANCEL_FILL
            row[7].value = "취소"
            row[7].alignment = CTR
            cancelled += 1
    return cancelled


def rebuild_summary(wb, year: int):
    ws = wb["Summary"]
    # 병합셀 해제 후 초기화
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None; cell.fill = NO_FILL
            cell.border = Border(); cell.font = DATA_FONT

    cur_row = 2
    for sheet_name in wb.sheetnames:
        if sheet_name == "Summary":
            continue
        person_ws    = wb[sheet_name]
        info_text    = str(person_ws["A1"].value or "")
        name         = info_text.split("(")[0].strip()
        em           = re.search(r"\(([^)]+)\)", info_text)
        person_email = em.group(1) if em else ""

        monthly = [0.0] * 12
        for row_data in person_ws.iter_rows(min_row=PERSON_DATA_ROW, values_only=True):
            if not row_data[0]:
                continue
            try:
                start_dt   = datetime.strptime(str(row_data[2])[:10], "%Y-%m-%d")
                leave_type = str(row_data[5] or "")
                status = str(row_data[7]) if len(row_data) > 7 and row_data[7] else ""
                if start_dt.year == year and leave_type in ANNUAL_DEDUCT_TYPES and status != "취소":
                    monthly[start_dt.month - 1] += float(row_data[4] or 0)
            except (ValueError, TypeError):
                pass

        alloc      = get_person_annual_leave(name, year)
        total_used = sum(monthly)
        remaining  = alloc - total_used
        for col, val in enumerate([name, person_email, alloc] + monthly + [total_used, remaining], 1):
            cell = ws.cell(row=cur_row, column=col, value=val)
            cell.font = DATA_FONT; cell.border = THIN
            cell.alignment = LFT if col == 2 else CTR
            if col in (16, 17):
                cell.fill = TOTAL_FILL
        cur_row += 1


# ── 리포트 HTML ───────────────────────────────────────────────────────────────
def build_report_html(wb, year: int) -> str:
    ws    = wb["Summary"]
    today = datetime.now()
    rows_html = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name, email, alloc = row[0], row[1], row[2]
        monthly    = list(row[3:15])
        total_used = row[15] or 0
        remaining  = row[16] or 0
        monthly_td = "".join(f"<td>{v if v else ''}</td>" for v in monthly)
        rows_html += (
            f"<tr><td><strong>{name or ''}</strong></td>"
            f"<td style='font-size:12px;color:#555;'>{email or ''}</td>"
            f"<td>{int(alloc or 0)}</td>{monthly_td}"
            f"<td><strong>{total_used}</strong></td>"
            f"<td><strong>{remaining}</strong></td></tr>"
        )

    month_ths = "".join(f"<th>{m}</th>" for m in MONTHS_KR)
    return (
        f'<!DOCTYPE html><html><head><meta charset="utf-8"><style>'
        f'body{{font-family:Malgun Gothic,Arial,sans-serif;font-size:14px;color:#333;margin:20px;}}'
        f'h2{{color:#1565C0;}}'
        f'table{{border-collapse:collapse;width:100%;font-size:12px;margin-top:12px;}}'
        f'th{{background:#D9E2F3;padding:8px 5px;border:1px solid #aaa;text-align:center;white-space:nowrap;}}'
        f'td{{padding:6px 5px;border:1px solid #ddd;text-align:center;}}'
        f'tr:nth-child(even){{background:#F8F9FA;}}'
        f'.footer{{margin-top:16px;font-size:11px;color:#999;}}'
        f'</style></head><body>'
        f'<h2>{year}년 연차 현황 리포트</h2>'
        f'<p>{today.strftime("%Y년 %m월 %d일")} 기준 | 법정 연차 개인별 적용</p>'
        f'<table><thead><tr><th>이름</th><th>이메일</th><th>발생</th>'
        f'{month_ths}<th>총사용</th><th>잔여</th></tr></thead>'
        f'<tbody>{rows_html}</tbody></table>'
        f'<p class="footer">* 엑셀 원본: \\\\Orbiters\\경영지원\\연차관리\\연차관리_{year}.xlsx (Z 드라이브 불필요)</p>'
        f'</body></html>'
    )


# ── 이메일 날짜 포맷 헬퍼 ─────────────────────────────────────────────────────
_KR_DOW = ["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"]

def _fmt_date_kr(date_str: str) -> str:
    """'2026-03-20' → '2026년 3월 20일 (금요일)'"""
    d = datetime.strptime(date_str[:10], "%Y-%m-%d").date()
    return f"{d.year}년 {d.month}월 {d.day}일 ({_KR_DOW[d.weekday()]})"

def _date_rows_html(start_str: str, end_str: str, days: float,
                    label: str = "일자/기간", days_label: str = "사용 일수") -> str:
    """이메일 테이블용 날짜 행 HTML. 단일일이면 1행, 다중일이면 시작/종료/일수 3행."""
    TD  = 'style="padding:7px 14px;border:1px solid #E5E7EB;color:#555;width:130px;"'
    TD2 = 'style="padding:7px 14px;border:1px solid #E5E7EB;"'
    ALT = 'style="padding:7px 14px;border:1px solid #E5E7EB;background:#F9FAFB;color:#555;width:130px;"'
    ALT2= 'style="padding:7px 14px;border:1px solid #E5E7EB;background:#F9FAFB;"'
    if start_str[:10] == end_str[:10]:
        return (
            f'<tr><td {TD}>{label}</td>'
            f'<td {TD2}><strong>{_fmt_date_kr(start_str)}</strong></td></tr>'
        )
    days_val = int(days) if days == int(days) else days
    return (
        f'<tr><td {TD}>시작일</td>'
        f'<td {TD2}><strong>{_fmt_date_kr(start_str)}</strong></td></tr>'
        f'<tr><td {ALT}>종료일</td>'
        f'<td {ALT2}><strong>{_fmt_date_kr(end_str)}</strong></td></tr>'
        f'<tr><td {TD}>{days_label}</td>'
        f'<td {TD2}><strong>{days_val}일</strong></td></tr>'
    )


# ── 명령 ──────────────────────────────────────────────────────────────────────
def cmd_setup(year: int):
    path = setup_excel(year)
    print(f"\n엑셀 준비 완료: {path}")


def cmd_sync(year: int):
    print(f"\n{year}년 연차/휴가 이메일 검색 중...")
    emails = fetch_leave_emails(year)
    print(f"  검색된 이메일: {len(emails)}건")

    processed_ids  = load_processed_ids()
    new_emails     = [e for e in emails if e["id"] not in processed_ids]
    print(f"  미처리 이메일: {len(new_emails)}건")

    if not new_emails:
        print("  새로 처리할 이메일이 없습니다.")
        return

    excel_path = get_excel_path(year)
    if not excel_path.exists():
        setup_excel(year)
    wb = openpyxl.load_workbook(excel_path)

    added = skipped_dup = skipped_not_leave = 0
    needs_rebuild = False

    for i, email in enumerate(new_emails, 1):
        print(f"\n[{i}/{len(new_emails)}] {email['subject'][:60]}")
        print(f"  발신: {email['sender_name']} <{email['sender_email']}>")

        parsed_list = parse_leave_request(email)
        processed_ids.add(email["id"])

        if not parsed_list:
            print("  → 연차 신청 아님, 스킵")
            skipped_not_leave += 1
            continue

        name       = parsed_list[0]["name"]
        total_days = sum(p["days"] for p in parsed_list)
        leave_type = parsed_list[0]["leave_type"]
        action     = parsed_list[0].get("action", "add")
        print(f"  → {name} / 총 {total_days}일 ({leave_type}) / action={action}")

        # sender_email이 Exchange DN이면 employee_data.json으로 fallback
        raw_se = parsed_list[0]["sender_email"]
        sender_email = raw_se if not raw_se.upper().startswith("/O=") else (get_person_email(name) or raw_se)
        for seg in parsed_list:
            seg["sender_email"] = sender_email

        person_ws = ensure_person_sheet(wb, name, sender_email, year)
        p = parsed_list[0]

        if action == "cancel":
            # ── 취소 처리 ──────────────────────────────────────────────
            date_ranges = [{"from": seg["start_date"], "to": seg["end_date"]} for seg in parsed_list]
            kr_holidays = set()
            for seg in parsed_list:
                for yr in {datetime.strptime(seg["start_date"][:10], "%Y-%m-%d").year,
                           datetime.strptime(seg["end_date"][:10],   "%Y-%m-%d").year}:
                    kr_holidays |= get_kr_holidays(yr)

            n_cancelled = cancel_person_entries(person_ws, date_ranges, kr_holidays)
            if n_cancelled == 0:
                print("  → 취소할 일치 항목 없음, 스킵")
                continue

            added += n_cancelled
            rebuild_summary(wb, year)  # 취소 즉시 Summary 갱신
            print(f"  → Excel 취소 처리 완료 ({n_cancelled}건)")

            # 취소 확인 초안 저장 (발송 안 함 - 임시저장함 확인 후 직접 발송)
            cancel_dates = []
            for seg in parsed_list:
                cancel_dates.append(seg["start_date"])
            _alloc      = get_person_annual_leave(name, year)
            _used_after = sum(
                float(r[4] or 0)
                for r in person_ws.iter_rows(min_row=PERSON_DATA_ROW, values_only=True)
                if r[0] and str(r[5] or "") in ANNUAL_DEDUCT_TYPES and str(r[7] or "") != "취소"
            )
            _remaining  = _alloc - _used_after
            html_body = _build_cancel_email_html(name, cancel_dates, remaining=_remaining)
            save_draft_email(
                subject    = f"[연차 취소] {name} - {cancel_dates[0][:10]}",
                html_body  = html_body,
                recipients = [sender_email],
                cc         = LEAVE_TEAM_EMAILS,
            )

        else:
            # ── 신규 등록 처리 ─────────────────────────────────────────
            entries   = get_person_entries(person_ws)
            seg_added = 0
            for parsed in parsed_list:
                if is_duplicate(entries, parsed):
                    skipped_dup += 1
                    continue
                append_person_entry(person_ws, parsed)
                seg_added += 1

            if seg_added == 0:
                print("  → 모든 구간 중복, 스킵")
                continue

            added += seg_added
            needs_rebuild = True
            print(f"  → Excel 기록 완료 ({seg_added}건)")

            # 신청 확인 초안 저장 (발송 안 함 - 임시저장함 확인 후 직접 발송)
            start_str   = parsed_list[0]["start_date"]
            _alloc      = get_person_annual_leave(name, year)
            _used_after = sum(
                float(r[4] or 0)
                for r in person_ws.iter_rows(min_row=PERSON_DATA_ROW, values_only=True)
                if r[0] and str(r[5] or "") in ANNUAL_DEDUCT_TYPES and str(r[7] or "") != "취소"
            )
            _remaining  = _alloc - _used_after
            html_body = _build_add_email_html(name, parsed_list,
                                              alloc=_alloc, total_used=_used_after, remaining=_remaining)
            save_draft_email(
                subject    = f"[연차 신청 확인] {name} - {start_str[:10]}",
                html_body  = html_body,
                recipients = [sender_email],
                cc         = LEAVE_TEAM_EMAILS,
            )

    if needs_rebuild:
        rebuild_summary(wb, year)

    wb.save(excel_path)
    save_processed_ids(processed_ids)

    print(f"\n동기화 완료!")
    print(f"  신규 기록: {added}건 | 중복 스킵: {skipped_dup}건 | 비연차 스킵: {skipped_not_leave}건")
    print(f"  엑셀: {excel_path}")


def cmd_report(year: int, dry_run: bool = False):
    excel_path = get_excel_path(year)
    if not excel_path.exists():
        print("ERROR: 엑셀 파일 없음. 먼저 --sync를 실행하세요.")
        sys.exit(1)

    wb   = openpyxl.load_workbook(excel_path)
    html = build_report_html(wb, year)
    out  = BASE_DIR / ".tmp" / "leave_report_preview.html"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(html, encoding="utf-8")
    print(f"리포트 저장: {out}")

    if not dry_run:
        # 리포트 초안 저장 (발송 안 함 - 임시저장함 확인 후 직접 발송)
        today_str  = datetime.now().strftime("%Y년 %m월")
        recipients = LEAVE_TEAM_EMAILS if LEAVE_TEAM_EMAILS else [SENDER_EMAIL]
        save_draft_email(
            subject    = f"[연차 현황] {year}년 {today_str} 기준 리포트",
            html_body  = html,
            recipients = recipients,
        )
    else:
        print("※ dry-run 모드 - 임시저장 건너뜀")


def cmd_auto(year: int):
    today = datetime.now()

    # 주말(토·일) 및 공휴일이면 실행 건너뜀
    kr_holidays = get_kr_holidays(today.year)
    if today.weekday() >= 5 or today.date() in kr_holidays:
        day_label = "주말" if today.weekday() >= 5 else "공휴일"
        print(f"오늘은 {day_label}({today.strftime('%Y-%m-%d')}) → 실행 건너뜀")
        return

    cmd_sync(year)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="연차 자동 추적 (Outlook COM)")
    parser.add_argument("--setup",   action="store_true", help="Excel 초기 파일 생성")
    parser.add_argument("--sync",    action="store_true", help="이메일 스캔 & Excel 업데이트")
    parser.add_argument("--report",  action="store_true", help="연차 현황 HTML 파일 생성 (.tmp/leave_report_preview.html)")
    parser.add_argument("--auto",    action="store_true", help="주말/공휴일 제외 자동 sync")
    parser.add_argument("--year",    type=int, default=datetime.now().year)
    args = parser.parse_args()

    if   args.setup:  cmd_setup(args.year)
    elif args.sync:   cmd_sync(args.year)
    elif args.report: cmd_report(args.year)
    elif args.auto:   cmd_auto(args.year)
    else:             parser.print_help()


if __name__ == "__main__":
    main()
