"""
연차 자동 추적 시스템 (Microsoft Graph API 기반)

Graph API로 받은편지함 이메일을 읽어 연차 내용을 Z드라이브 엑셀에 기록하고
팀 이메일 알림을 발송합니다. Outlook 실행 불필요.

Prerequisites (최초 1회 설정):
    1. Azure AD 앱 등록 후 ~/.wat_secrets에 추가:
       OUTLOOK_CLIENT_ID=<Azure 앱 Client ID>
       OUTLOOK_TENANT_ID=<Azure 앱 Tenant ID>
       OUTLOOK_EMAIL=dk.shin@orbiters.co.kr
       LEAVE_TEAM_EMAILS=team1@company.com,team2@company.com
       ANTHROPIC_API_KEY=<Claude API key>
    2. Azure AD 앱 권한: Mail.Read, Mail.Send, User.Read (Delegated)
    3. 최초 실행 시 브라우저 1회 인증 -> 이후 자동 갱신

Usage:
    python tools/outlook_leave_tracker.py --setup
    python tools/outlook_leave_tracker.py --sync
    python tools/outlook_leave_tracker.py --report --dry-run
    python tools/outlook_leave_tracker.py --auto
"""

import argparse
import json
import os
import re
import sys
from datetime import date as date_cls, datetime, timedelta
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import anthropic

sys.path.insert(0, str(Path(__file__).parent))
from env_loader import load_env
load_env()

# ── Config ────────────────────────────────────────────────────────────────────
OUTLOOK_EMAIL     = os.getenv("OUTLOOK_EMAIL", "")
LEAVE_TEAM_EMAILS = [e.strip() for e in os.getenv("LEAVE_TEAM_EMAILS", "").split(",") if e.strip()]
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
OUTLOOK_CLIENT_ID = os.getenv("OUTLOOK_CLIENT_ID", "")
OUTLOOK_TENANT_ID = os.getenv("OUTLOOK_TENANT_ID", "common")

BASE_DIR         = Path(__file__).parent.parent
PROCESSED_PATH   = BASE_DIR / ".tmp" / "leave_tracker_processed.json"
GRAPH_TOKEN_PATH = BASE_DIR / "credentials" / "leave_tracker_token.json"
EXCEL_DIR        = Path(r"Z:\경영지원\연차관리")
ANNUAL_LEAVE_DAYS = 15
MONTHS_KR = ["1월","2월","3월","4월","5월","6월","7월","8월","9월","10월","11월","12월"]
GRAPH_SCOPES = ["Mail.Read", "Mail.Send"]
LEAVE_DOMAIN  = os.getenv("LEAVE_DOMAIN", "")  # 예: orbiters.co.lr

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TOTAL_FILL  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
WARN_FILL   = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
ALT_FILL    = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
NO_FILL     = PatternFill(fill_type=None)
HEADER_FONT = Font(name="Calibri", size=10, bold=True)
DATA_FONT   = Font(name="Calibri", size=10)
TITLE_FONT  = Font(name="Calibri", size=12, bold=True)
THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"),  bottom=Side(style="thin"))
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ── Graph API 인증 (MSAL) ──────────────────────────────────────────────────────
def get_graph_token() -> str:
    """최초 1회 브라우저 인증 후 토큰 캐싱. 이후 자동 갱신."""
    if not OUTLOOK_CLIENT_ID:
        print("ERROR: OUTLOOK_CLIENT_ID 미설정 -> ~/.wat_secrets에 추가하세요.")
        sys.exit(1)
    try:
        import msal
    except ImportError:
        print("ERROR: msal 미설치 -> pip install msal")
        sys.exit(1)

    GRAPH_TOKEN_PATH.parent.mkdir(parents=True, exist_ok=True)
    cache = msal.SerializableTokenCache()
    if GRAPH_TOKEN_PATH.exists():
        cache.deserialize(GRAPH_TOKEN_PATH.read_text(encoding="utf-8"))

    authority = f"https://login.microsoftonline.com/{OUTLOOK_TENANT_ID}"
    app = msal.PublicClientApplication(OUTLOOK_CLIENT_ID, authority=authority, token_cache=cache)

    accounts = app.get_accounts()
    result = None
    if accounts:
        result = app.acquire_token_silent(GRAPH_SCOPES, account=accounts[0])

    if not result:
        flow = app.initiate_device_flow(scopes=GRAPH_SCOPES)
        print("\n[Microsoft 계정 인증 필요 - 최초 1회만]")
        print(f"  1. 브라우저 접속: {flow['verification_uri']}")
        print(f"  2. 코드 입력: {flow['user_code']}")
        print("  3. 회사 Microsoft 계정으로 로그인")
        print("  (완료될 때까지 대기 중...)\n")
        result = app.acquire_token_by_device_flow(flow)

    if "access_token" in result:
        GRAPH_TOKEN_PATH.write_text(cache.serialize(), encoding="utf-8")
        return result["access_token"]

    print(f"ERROR: 인증 실패 - {result.get('error_description', result)}")
    sys.exit(1)


# ── 임시저장 초안 생성 (Graph API POST /me/messages) ──────────────────────────
def create_draft_email(parsed: dict, recipients: list) -> bool:
    """Graph API로 Outlook 임시저장 초안 생성. 자동 발송 없이 검토 후 수동 발송."""
    name       = parsed.get("name", "")
    start_date = parsed.get("start_date", "")
    end_date   = parsed.get("end_date", start_date)
    days       = parsed.get("days", 1.0)
    leave_type = parsed.get("leave_type", "연차")
    note       = parsed.get("note", "")

    label = "[반차-오후]" if (days == 0.5 and "오후" in (note or "")) else \
            "[반차-오전]" if days == 0.5 else f"[{leave_type}]"
    subject = f"{label} {name} ({start_date}~{end_date}) 연차 기록 완료"

    html_body = (
        f'<p style="font-family:Malgun Gothic,Arial;font-size:14px;">'
        f'<strong>{name}</strong>님의 <strong>{leave_type}</strong>가 엑셀에 기록되었습니다.</p>'
        f'<table style="border-collapse:collapse;font-size:13px;">'
        f'<tr><td style="padding:4px 12px 4px 0;color:#555;">기간</td>'
        f'<td><strong>{start_date} ~ {end_date}</strong></td></tr>'
        f'<tr><td style="padding:4px 12px 4px 0;color:#555;">일수</td><td>{days}일</td></tr>'
        f'<tr><td style="padding:4px 12px 4px 0;color:#555;">유형</td><td>{leave_type}</td></tr>'
        f'</table>'
        f'<p style="font-size:12px;color:#888;margin-top:16px;">* 임시저장 초안 - LeaveTracker (검토 후 발송)</p>'
    )

    token = get_graph_token()
    # sendMail 대신 POST /me/messages → 임시저장(Drafts) 폴더에 저장
    payload = {
        "subject": subject,
        "body": {"contentType": "HTML", "content": html_body},
        "toRecipients": [{"emailAddress": {"address": a}} for a in recipients],
    }
    try:
        resp = requests.post(
            "https://graph.microsoft.com/v1.0/me/messages",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=payload, timeout=15,
        )
        if resp.status_code == 201:
            draft_id = resp.json().get("id", "")[:20]
            print(f"  임시저장 생성: {subject} (id: {draft_id}...)")
            return True
        print(f"  [임시저장 오류] HTTP {resp.status_code}: {resp.text[:200]}")
        return False
    except Exception as e:
        print(f"  [임시저장 오류] {e}")
        return False


# ── Processed IDs ─────────────────────────────────────────────────────────────
def load_processed_ids() -> set:
    PROCESSED_PATH.parent.mkdir(parents=True, exist_ok=True)
    if PROCESSED_PATH.exists():
        return set(json.loads(PROCESSED_PATH.read_text(encoding="utf-8")))
    return set()


def save_processed_ids(ids: set):
    PROCESSED_PATH.write_text(json.dumps(list(ids), ensure_ascii=False, indent=2), encoding="utf-8")


# ── 이메일 수집 (Graph API) ────────────────────────────────────────────────────
LEAVE_KEYWORDS = ["연차", "휴가", "반차", "병가"]


def fetch_leave_emails(year: int) -> list:
    """Graph API로 받은편지함에서 해당 연도 연차 키워드 이메일 수집. Outlook 실행 불필요."""
    token = get_graph_token()
    url = (
        "https://graph.microsoft.com/v1.0/me/mailFolders/inbox/messages"
        f"?$filter=receivedDateTime ge {year}-01-01T00:00:00Z"
        f" and receivedDateTime le {year}-12-31T23:59:59Z"
        "&$select=id,subject,body,sender,receivedDateTime&$top=999"
    )
    emails = []
    while url:
        resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
        if resp.status_code != 200:
            print(f"ERROR: 이메일 조회 실패 HTTP {resp.status_code}: {resp.text[:200]}")
            break
        data = resp.json()
        for msg in data.get("value", []):
            subject      = msg.get("subject", "") or ""
            body_content = msg.get("body", {}).get("content", "") or ""
            if not any(kw in (subject + body_content) for kw in LEAVE_KEYWORDS):
                continue
            sender   = msg.get("sender", {}).get("emailAddress", {})
            received = (msg.get("receivedDateTime", "") or "")[:10]
            emails.append({
                "id":           msg["id"],
                "subject":      subject,
                "sender_name":  sender.get("name", ""),
                "sender_email": sender.get("address", ""),
                "body":         body_content,
                "received_at":  received,
            })
        url = data.get("@odata.nextLink")
    return emails


# ── Korean Holiday Calendar ───────────────────────────────────────────────────
def get_kr_holidays(year: int) -> set:
    try:
        import holidays as hol_lib
        return set(hol_lib.KR(years=year).keys())
    except ImportError:
        return {
            date_cls(year,  1,  1), date_cls(year,  3,  1),
            date_cls(year,  5,  5), date_cls(year,  6,  6),
            date_cls(year,  8, 15), date_cls(year, 10,  3),
            date_cls(year, 10,  9), date_cls(year, 12, 25),
        }


def is_working_day(d: date_cls, kr_holidays: set) -> bool:
    return d.weekday() < 5 and d not in kr_holidays


def expand_date_range(from_str: str, to_str: str) -> list:
    start = datetime.strptime(from_str[:10], "%Y-%m-%d").date()
    end   = datetime.strptime(to_str[:10],   "%Y-%m-%d").date()
    return [start + timedelta(days=i) for i in range(max((end - start).days + 1, 1))]


def group_leave_dates(leave_dates: list, kr_holidays: set) -> list:
    sorted_dates = sorted(set(d for d in leave_dates if is_working_day(d, kr_holidays)))
    if not sorted_dates:
        return []
    segments = []
    seg_start = seg_end = sorted_dates[0]
    workday_count = 1
    for d in sorted_dates[1:]:
        gap_has_workday = any(
            is_working_day(seg_end + timedelta(days=i), kr_holidays)
            for i in range(1, (d - seg_end).days)
        )
        if not gap_has_workday:
            seg_end = d
            workday_count += 1
        else:
            segments.append({"start_date": seg_start.isoformat(), "end_date": seg_end.isoformat(), "days": float(workday_count)})
            seg_start = seg_end = d
            workday_count = 1
    segments.append({"start_date": seg_start.isoformat(), "end_date": seg_end.isoformat(), "days": float(workday_count)})
    return segments


def build_holiday_context(dates_hint: list) -> str:
    years = set()
    for d in dates_hint:
        try: years.add(int(d[:4]))
        except: pass
    if not years:
        years = {datetime.now().year}
    lines = [f"  {h.isoformat()}" for yr in sorted(years) for h in sorted(get_kr_holidays(yr))]
    return "\n".join(lines) if lines else "  (공휴일 정보 없음)"


# ── Claude AI Parsing ─────────────────────────────────────────────────────────
def parse_leave_request(email: dict):
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    body_text = re.sub(r"<[^>]+>", " ", email["body"])
    body_text = re.sub(r"\s+", " ", body_text).strip()[:3000]
    current_year = datetime.now().year
    holiday_ctx  = build_holiday_context([email["received_at"], f"{current_year}-01-01"])

    prompt = f"""다음 이메일이 연차/휴가/반차/병가 신청 이메일인지 분석해주세요.

발신자: {email['sender_name']} ({email['sender_email']})
제목: {email['subject']}
수신일: {email['received_at']}
본문:
{body_text}

연차/휴가 신청이면 아래 JSON만 응답하세요 (다른 텍스트 없이):
{{
  "is_leave_request": true,
  "name": "신청자 이름 (발신자 이름 사용)",
  "leave_type": "연차",
  "note": "특이사항 (없으면 빈 문자열)",
  "date_ranges": [
    {{"from": "YYYY-MM-DD", "to": "YYYY-MM-DD"}}
  ]
}}

연차/휴가 신청이 아니면:
{{"is_leave_request": false}}

규칙:
- date_ranges: 이메일 언급 날짜를 그대로 (공휴일/주말 포함 가능, 필터링은 시스템이 함)
- 단일 날짜면 from == to
- 반차는 leave_type="반차", note에 오전/오후 명시
- 연도 불명확 시 {current_year}년 기준
- JSON만 응답, 마크다운 코드블록 없이

참고 - 한국 공휴일:
{holiday_ctx}"""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-6", max_tokens=500,
            messages=[{"role": "user", "content": prompt}],
        )
        text = response.content[0].text.strip()
        text = re.sub(r"^```[a-z]*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
        parsed = json.loads(text)
    except Exception as e:
        print(f"  [파싱 오류] {e}")
        return None

    if not parsed.get("is_leave_request"):
        return None

    name        = parsed.get("name") or email["sender_name"]
    leave_type  = parsed.get("leave_type", "연차")
    note        = parsed.get("note", "")
    date_ranges = parsed.get("date_ranges", [])
    if not date_ranges:
        return None

    all_dates = []
    mentioned_years = set()
    for dr in date_ranges:
        expanded = expand_date_range(dr.get("from", ""), dr.get("to", dr.get("from", "")))
        all_dates.extend(expanded)
        mentioned_years.update(d.year for d in expanded)

    kr_holidays = set()
    for yr in mentioned_years:
        kr_holidays |= get_kr_holidays(yr)

    if leave_type == "반차":
        working = [d for d in sorted(set(all_dates)) if is_working_day(d, kr_holidays)]
        if not working:
            return None
        d = working[0]
        return [{"name": name, "start_date": d.isoformat(), "end_date": d.isoformat(),
                 "days": 0.5, "leave_type": leave_type, "note": note,
                 "sender_email": email["sender_email"], "sender_name": email["sender_name"],
                 "email_subject": email["subject"], "received_at": email["received_at"]}]

    segments = group_leave_dates(all_dates, kr_holidays)
    if not segments:
        return None

    return [{"name": name, "start_date": s["start_date"], "end_date": s["end_date"],
             "days": s["days"], "leave_type": leave_type, "note": note,
             "sender_email": email["sender_email"], "sender_name": email["sender_name"],
             "email_subject": email["subject"], "received_at": email["received_at"]}
            for s in segments]


# ── Excel ─────────────────────────────────────────────────────────────────────
def get_excel_path(year: int) -> Path:
    return EXCEL_DIR / f"연차관리_{year}.xlsx"


def safe_sheet_name(name: str) -> str:
    return re.sub(r'[\\/*?:\[\]]', "", name)[:31]


def setup_excel(year: int) -> Path:
    path = get_excel_path(year)
    EXCEL_DIR.mkdir(parents=True, exist_ok=True)
    if path.exists():
        print(f"이미 존재: {path}")
        return path
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    for col, h in enumerate(["이름", "이메일", "발생연차"] + MONTHS_KR + ["총사용", "잔여"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.border = THIN; cell.alignment = CTR
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    for i in range(4, 16):
        ws.column_dimensions[get_column_letter(i)].width = 7
    ws.column_dimensions[get_column_letter(16)].width = 8
    ws.column_dimensions[get_column_letter(17)].width = 8
    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"생성 완료: {path}")
    return path


PERSON_HEADERS    = ["신청일", "이메일 제목", "휴가 시작", "휴가 종료", "사용일수", "유형", "비고"]
PERSON_COL_WIDTHS = [12, 38, 12, 12, 10, 8, 25]
PERSON_DATA_ROW   = 4


def ensure_person_sheet(wb, name: str, email: str):
    sheet_name = safe_sheet_name(name)
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws["A1"] = f"{name} ({email})"
    ws["A1"].font = TITLE_FONT; ws["A1"].alignment = LFT
    ws.merge_cells("A1:G1")
    ws["A2"] = f"{datetime.now().year}년 연차 사용 내역 (발생: {ANNUAL_LEAVE_DAYS}일)"
    ws["A2"].font = DATA_FONT; ws["A2"].alignment = LFT
    ws.merge_cells("A2:G2")
    for col, h in enumerate(PERSON_HEADERS, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.border = THIN; cell.alignment = CTR
    for i, w in enumerate(PERSON_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"
    return ws


def get_person_entries(ws) -> list:
    entries = []
    for row in ws.iter_rows(min_row=PERSON_DATA_ROW, values_only=True):
        if row[0] is None:
            continue
        entries.append({
            "received_at": str(row[0]) if row[0] else "",
            "subject":     str(row[1]) if row[1] else "",
            "start_date":  str(row[2]) if row[2] else "",
            "end_date":    str(row[3]) if row[3] else "",
            "days":        row[4],
            "leave_type":  str(row[5]) if row[5] else "",
            "note":        str(row[6]) if row[6] else "",
        })
    return entries


def is_duplicate(entries: list, parsed: dict) -> bool:
    return any(
        str(e["start_date"]) == str(parsed.get("start_date", "")) and
        str(e["end_date"])   == str(parsed.get("end_date", ""))
        for e in entries
    )


def append_person_entry(ws, parsed: dict):
    next_row = max(ws.max_row + 1, PERSON_DATA_ROW)
    row_fill = ALT_FILL if (next_row % 2 == 0) else NO_FILL
    values = [parsed.get("received_at",""), parsed.get("email_subject",""),
              parsed.get("start_date",""), parsed.get("end_date",""),
              parsed.get("days",1.0), parsed.get("leave_type","연차"), parsed.get("note","")]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=next_row, column=col, value=val)
        cell.font = DATA_FONT; cell.border = THIN; cell.fill = row_fill
        cell.alignment = CTR if col in (1, 3, 4, 5, 6) else LFT


def rebuild_summary(wb, year: int):
    ws = wb["Summary"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None; cell.fill = NO_FILL
            cell.border = Border(); cell.font = DATA_FONT
    current_row = 2
    for sheet_name in wb.sheetnames:
        if sheet_name == "Summary":
            continue
        person_ws    = wb[sheet_name]
        info_text    = str(person_ws["A1"].value or "")
        name         = info_text.split("(")[0].strip()
        em           = re.search(r"\(([^)]+)\)", info_text)
        person_email = em.group(1) if em else ""
        monthly = [0.0] * 12
        for entry in get_person_entries(person_ws):
            try:
                start_dt = datetime.strptime(str(entry["start_date"])[:10], "%Y-%m-%d")
                if start_dt.year == year:
                    monthly[start_dt.month - 1] += float(entry["days"] or 0)
            except (ValueError, TypeError):
                pass
        total_used = sum(monthly)
        remaining  = ANNUAL_LEAVE_DAYS - total_used
        for col, val in enumerate([name, person_email, ANNUAL_LEAVE_DAYS] + monthly + [total_used, remaining], 1):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.font = DATA_FONT; cell.border = THIN
            if col == 17:
                cell.fill = WARN_FILL if remaining < 3 else TOTAL_FILL; cell.alignment = CTR
            elif col == 16:
                cell.fill = TOTAL_FILL; cell.alignment = CTR
            elif col == 2:
                cell.alignment = LFT
            else:
                cell.alignment = CTR
        current_row += 1


def build_report_html(wb, year: int) -> str:
    ws    = wb["Summary"]
    today = datetime.now()
    rows_html = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name, email, alloc = row[0], row[1], row[2]
        monthly    = list(row[3:15])
        total_used = row[15] or 0
        remaining  = row[16] or 0
        color      = "#C62828" if remaining < 3 else "#2E7D32"
        monthly_td = "".join(f'<td>{v if v else ""}</td>' for v in monthly)
        rows_html += (
            f'<tr><td><strong>{name or ""}</strong></td>'
            f'<td style="font-size:12px;color:#555;">{email or ""}</td>'
            f'<td>{int(alloc or 0)}</td>{monthly_td}'
            f'<td><strong>{total_used}</strong></td>'
            f'<td style="color:{color};font-weight:bold;">{remaining}</td></tr>'
        )
    month_ths = "".join(f"<th>{m}</th>" for m in MONTHS_KR)
    return (
        f'<!DOCTYPE html><html><head><meta charset="utf-8"><style>'
        f'body{{font-family:Malgun Gothic,Arial,sans-serif;font-size:14px;color:#333;margin:20px;}}'
        f'h2{{color:#1565C0;}}table{{border-collapse:collapse;width:100%;font-size:12px;margin-top:12px;}}'
        f'th{{background:#D9E2F3;padding:8px 5px;border:1px solid #aaa;text-align:center;white-space:nowrap;}}'
        f'td{{padding:6px 5px;border:1px solid #ddd;text-align:center;}}'
        f'tr:nth-child(even){{background:#F8F9FA;}}'
        f'.footer{{margin-top:16px;font-size:11px;color:#999;}}'
        f'</style></head><body>'
        f'<h2>{year}년 연차 현황 리포트</h2>'
        f'<p>{today.strftime("%Y년 %m월 %d일")} 기준 | 발생 연차: {ANNUAL_LEAVE_DAYS}일/연</p>'
        f'<table><thead><tr><th>이름</th><th>이메일</th><th>발생</th>{month_ths}'
        f'<th>총사용</th><th>잔여</th></tr></thead>'
        f'<tbody>{rows_html}</tbody></table>'
        f'<p class="footer">* 잔여 3일 미만: 빨간색<br>'
        f'* 엑셀 원본: Z:\\경영지원\\연차관리\\연차관리_{year}.xlsx</p>'
        f'</body></html>'
    )


# ── Commands ──────────────────────────────────────────────────────────────────
def cmd_setup(year: int):
    print(f"\n연차 관리 엑셀 준비 완료: {setup_excel(year)}")


def cmd_sync(year: int, test_only: bool = False):
    if test_only:
        print(f"[테스트 모드] 알림 수신: {OUTLOOK_EMAIL or '(미설정)'}\n")

    print(f"{year}년 연차/휴가 이메일 검색 중...")
    emails = fetch_leave_emails(year)
    print(f"  검색된 이메일: {len(emails)}건")

    processed_ids = load_processed_ids()
    new_emails    = [e for e in emails if e["id"] not in processed_ids]
    print(f"  미처리 이메일: {len(new_emails)}건")

    if not new_emails:
        print("\n새로 처리할 이메일이 없습니다.")
        return

    excel_path = get_excel_path(year)
    if not excel_path.exists():
        setup_excel(year)
    wb = openpyxl.load_workbook(excel_path)

    added = skipped_dup = skipped_not_leave = 0

    for i, email in enumerate(new_emails, 1):
        print(f"\n[{i}/{len(new_emails)}] {email['subject'][:55]}")
        print(f"  발신: {email['sender_name']} <{email['sender_email']}>")

        parsed_list = parse_leave_request(email)
        processed_ids.add(email["id"])

        if not parsed_list:
            print("  -> 연차 신청 아님, 스킵")
            skipped_not_leave += 1
            continue

        name       = parsed_list[0]["name"]
        total_days = sum(p["days"] for p in parsed_list)
        seg_info   = ", ".join(f"{p['start_date']}~{p['end_date']}({p['days']}일)" for p in parsed_list)
        print(f"  -> {name} / {len(parsed_list)}개 구간 / 총 {total_days}일 ({parsed_list[0]['leave_type']})")
        print(f"     {seg_info}")

        person_ws = ensure_person_sheet(wb, name, parsed_list[0]["sender_email"])
        entries   = get_person_entries(person_ws)

        seg_added = 0
        for parsed in parsed_list:
            if is_duplicate(entries, parsed):
                skipped_dup += 1
                continue
            append_person_entry(person_ws, parsed)
            seg_added += 1

        if seg_added == 0:
            print("  -> 모든 구간 중복, 스킵")
            continue

        added += seg_added

        notify_recipients = (
            [r for r in [OUTLOOK_EMAIL] if r] if test_only
            else list({parsed_list[0]["sender_email"]} | set(LEAVE_TEAM_EMAILS))
        )
        if notify_recipients:
            create_draft_email(parsed_list[0], notify_recipients)

    if added > 0:
        rebuild_summary(wb, year)

    wb.save(excel_path)
    save_processed_ids(processed_ids)

    print(f"\n동기화 완료!")
    print(f"  신규 기록: {added}건 | 중복 스킵: {skipped_dup}건 | 비연차 스킵: {skipped_not_leave}건")
    print(f"  엑셀: {excel_path}")


# ── 도메인 전체 유저 조회 (Graph API) ─────────────────────────────────────────
def fetch_domain_user_emails(domain: str) -> list:
    """Graph API로 조직 내 해당 도메인 유저 이메일 목록 반환."""
    if not domain:
        return list(LEAVE_TEAM_EMAILS)
    token = get_graph_token()
    url = "https://graph.microsoft.com/v1.0/users?=mail,userPrincipalName&=999"
    result = []
    while url:
        resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
        if resp.status_code != 200:
            print(f"[유저 조회 오류] HTTP {resp.status_code}: {resp.text[:200]}")
            break
        data = resp.json()
        for u in data.get("value", []):
            email = u.get("mail") or u.get("userPrincipalName", "")
            if email and email.lower().endswith("@" + domain.lower()):
                result.append(email)
        url = data.get("@odata.nextLink")
    return result


def cmd_report(year: int, dry_run: bool = False):
    excel_path = get_excel_path(year)
    if not excel_path.exists():
        print("ERROR: 엑셀 파일 없음. 먼저 --sync를 실행하세요.")
        sys.exit(1)

    wb      = openpyxl.load_workbook(excel_path)
    html    = build_report_html(wb, year)
    today   = datetime.now()
    subject = f"[연차현황] {year}년 {today.month}월 연차 사용 현황"

    if dry_run:
        out = BASE_DIR / ".tmp" / "leave_report_preview.html"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(html, encoding="utf-8")
        print(f"[DRY RUN] 미리보기 저장: {out}")
        rcp = fetch_domain_user_emails(LEAVE_DOMAIN) if LEAVE_DOMAIN else list(LEAVE_TEAM_EMAILS)
        print(f"발송 예정: {len(rcp)}명 ({LEAVE_DOMAIN or '개별 설정'})")
        return

    recipients = fetch_domain_user_emails(LEAVE_DOMAIN) if LEAVE_DOMAIN else list(LEAVE_TEAM_EMAILS)
    if not recipients:
        print("ERROR: LEAVE_DOMAIN 또는 LEAVE_TEAM_EMAILS 미설정.")
        sys.exit(1)
    print(f"  발송 대상: {len(recipients)}명")

    token   = get_graph_token()
    payload = {
        "message": {
            "subject": subject,
            "body": {"contentType": "HTML", "content": html},
            "toRecipients": [{"emailAddress": {"address": a}} for a in recipients],
        },
        "saveToSentItems": True,
    }
    try:
        resp = requests.post(
            "https://graph.microsoft.com/v1.0/me/sendMail",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=payload, timeout=15,
        )
        if resp.status_code == 202:
            print(f"\n리포트 발송 완료 -> {len(LEAVE_TEAM_EMAILS)}개 주소")
        else:
            print(f"ERROR: 리포트 발송 실패 HTTP {resp.status_code}: {resp.text[:200]}")
    except Exception as e:
        print(f"ERROR: 리포트 발송 실패 - {e}")


def cmd_auto(year: int, test_only: bool = False):
    cmd_sync(year, test_only=test_only)
    today = datetime.now()
    if today.day == 1:
        print("\n매월 1일 -> 리포트 발송")
        cmd_report(year)
    else:
        print(f"\n리포트 발송 스킵 (매월 1일에만, 오늘은 {today.day}일)")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="연차 자동 추적 시스템 (Graph API)")
    parser.add_argument("--setup",     action="store_true")
    parser.add_argument("--sync",      action="store_true")
    parser.add_argument("--report",    action="store_true")
    parser.add_argument("--auto",      action="store_true")
    parser.add_argument("--dry-run",   action="store_true")
    parser.add_argument("--test-only", action="store_true")
    parser.add_argument("--year",      type=int, default=datetime.now().year)
    args = parser.parse_args()

    if args.setup:
        cmd_setup(args.year)
    elif args.sync:
        cmd_sync(args.year, test_only=args.test_only)
    elif args.report:
        cmd_report(args.year, dry_run=args.dry_run)
    elif args.auto:
        cmd_auto(args.year, test_only=args.test_only)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()