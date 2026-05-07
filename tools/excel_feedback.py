"""
WAT Tool: Read team feedback from Excel on Teams/SharePoint.

New Excel format (v2 — dropdown-based):
    Column F (승인): Dropdown — "Confirmed" or "Declined"
    Column G (대안 텍스트): If Declined, team writes alternative text here

Logic:
    F = "Confirmed"              → approve, post as-is
    F = "Declined" + G = text    → use G column text as the tweet instead
    F = "Declined" + G = empty   → skip (cancel)
    F = empty                    → pending (not yet reviewed)

Usage:
    python tools/excel_feedback.py --check              # download & check for feedback
    python tools/excel_feedback.py --check --apply      # check + apply changes
    python tools/excel_feedback.py --poll               # poll every 60s (daemon mode)
"""

import os
import sys
import json
import argparse
import logging
import tempfile
from pathlib import Path
from datetime import datetime, timedelta, timezone

from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

sys.path.insert(0, str(Path(__file__).parent))

import requests
from openpyxl import load_workbook
from teams_upload import (
    _get_access_token,
    _graph_headers,
    _get_channel_drive_folder,
    GRAPH_BASE,
    upload_file,
)

PROJECT_ROOT = Path(__file__).parent.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
def _plan_path(date_str: str = None) -> Path:
    """Get date-specific plan file path."""
    if not date_str:
        jst = timezone(timedelta(hours=9))
        date_str = datetime.now(jst).strftime("%Y-%m-%d")
    return TMP_DIR / f"daily_tweet_plan_{date_str}.json"

def download_plan_excel() -> str:
    """Download the latest tweet_plan Excel from Teams/SharePoint.

    Returns: local file path, or empty string on failure.
    """
    try:
        drive_id, folder_id = _get_channel_drive_folder()

        # Find the tweet plan file
        url = f"{GRAPH_BASE}/drives/{drive_id}/items/{folder_id}/children"
        resp = requests.get(url, headers=_graph_headers(), timeout=15)
        resp.raise_for_status()

        # Find the most relevant tweet_plan file (today's date first, then newest)
        from datetime import timezone
        jst = timezone(timedelta(hours=9))
        today_str = datetime.now(jst).strftime("%Y-%m-%d")

        candidates = []
        for item in resp.json().get("value", []):
            name = item.get("name", "")
            if name.startswith("tweet_plan") and name.endswith(".xlsx"):
                candidates.append(item)

        if not candidates:
            logger.warning("No tweet_plan Excel found on Teams")
            return ""

        # Prefer today's file
        target = None
        for item in candidates:
            if today_str in item["name"]:
                target = item
                break

        # Fallback: newest by lastModifiedDateTime
        if not target:
            candidates.sort(
                key=lambda x: x.get("lastModifiedDateTime", ""),
                reverse=True,
            )
            target = candidates[0]

        logger.info(f"Selected Excel: {target['name']}")

        # Download content
        download_url = f"{GRAPH_BASE}/drives/{drive_id}/items/{target['id']}/content"
        resp = requests.get(download_url, headers=_graph_headers(), timeout=30)
        resp.raise_for_status()

        TMP_DIR.mkdir(parents=True, exist_ok=True)
        local_path = str(TMP_DIR / f"feedback_{target['name']}")
        with open(local_path, "wb") as f:
            f.write(resp.content)

        logger.info(f"Downloaded: {target['name']} ({len(resp.content):,} bytes)")
        return local_path

    except Exception as e:
        logger.error(f"Download failed: {e}")
        return ""


def read_feedback(excel_path: str) -> list[dict]:
    """Read feedback from both sheets.

    Sheet 1 (내 트윗) — fixed rows, one per slot:
        A: 시간, F: 승인 (Confirmed/Declined), G: 대안 텍스트

    Sheet 2 (리플 계획) — variable rows, one per reply target:
        A: 시간, B: 타겟 계정, F: 승인, G: 대안 텍스트
        Multiple rows can share the same slot time.
        Rows with "(리플 없음)" are info rows — skip them.

    Returns: list of feedback dicts
        For tweets: {slot, action, alt_text, row, sheet}
        For replies: {slot, action, alt_text, row, sheet, reply_index, target_username}
        action: "approve" | "replace" | "cancel"
    """
    wb = load_workbook(excel_path, data_only=True)
    feedbacks = []

    for sheet_name in ["내 트윗", "리플 계획"]:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # Find columns by header text (row 3)
        col_map = {}
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=3, column=col).value or "").strip()
            if "시간" in val:
                col_map["time"] = col
            elif "승인" in val:
                col_map["approval"] = col
            elif "대안" in val:
                col_map["alt"] = col
            elif "타겟 계정" in val:
                col_map["target"] = col

        if "approval" not in col_map:
            logger.warning(f"승인 column not found in '{sheet_name}' sheet")
            continue

        if sheet_name == "내 트윗":
            # Fixed rows — scan until empty time column
            for row in range(4, ws.max_row + 1):
                time_val = ws.cell(row=row, column=col_map.get("time", 1)).value
                if not time_val:
                    break
                slot = _parse_slot(time_val)
                if slot == 0:
                    continue

                approval = str(ws.cell(row=row, column=col_map["approval"]).value or "").strip()
                alt_text = str(ws.cell(row=row, column=col_map.get("alt", 99)).value or "").strip() if "alt" in col_map else ""

                action = _resolve_action(approval, alt_text)
                if not action:
                    continue

                feedbacks.append({
                    "slot": slot,
                    "action": action,
                    "alt_text": alt_text,
                    "row": row,
                    "sheet": sheet_name,
                })

        else:
            # 리플 계획 — variable rows, track reply_index per slot
            reply_counters = {}  # {slot: count}

            for row in range(4, ws.max_row + 1):
                time_val = ws.cell(row=row, column=col_map.get("time", 1)).value
                if not time_val:
                    break
                slot = _parse_slot(time_val)
                if slot == 0:
                    continue

                # Skip info rows "(리플 없음)"
                target_col = col_map.get("target")
                if target_col:
                    target_val = str(ws.cell(row=row, column=target_col).value or "").strip()
                    if target_val == "—" or "리플 없음" in str(ws.cell(row=row, column=target_col + 1).value or ""):
                        continue

                approval = str(ws.cell(row=row, column=col_map["approval"]).value or "").strip()
                alt_text = str(ws.cell(row=row, column=col_map.get("alt", 99)).value or "").strip() if "alt" in col_map else ""

                action = _resolve_action(approval, alt_text)
                if not action:
                    continue

                reply_idx = reply_counters.get(slot, 0)
                reply_counters[slot] = reply_idx + 1

                target_username = ""
                if target_col:
                    target_username = str(ws.cell(row=row, column=target_col).value or "").strip().lstrip("@")

                feedbacks.append({
                    "slot": slot,
                    "action": action,
                    "alt_text": alt_text,
                    "row": row,
                    "sheet": sheet_name,
                    "reply_index": reply_idx,
                    "target_username": target_username,
                })

    wb.close()
    return feedbacks


def _resolve_action(approval: str, alt_text: str) -> str:
    """Convert approval dropdown value to action string.

    Logic (opt-out model):
        Confirmed              → approve
        Declined + alt text    → replace with alt text
        Declined (no alt)      → cancel/skip
        (empty) + alt text     → replace (team wrote alt but forgot dropdown)
        (empty) + no alt       → "" (pending — auto-approve after timeout)

    Returns: "approve" | "replace" | "cancel" | "" (empty = pending)
    """
    approval_lower = (approval or "").strip().lower()

    if approval_lower == "confirmed":
        return "approve"
    elif approval_lower == "declined":
        return "replace" if alt_text else "cancel"
    elif not approval_lower and alt_text:
        # Team wrote alt text but didn't select dropdown → treat as replace
        return "replace"
    return ""


def _parse_slot(time_val) -> int:
    """Parse time value like '9:00' or '15:00' into slot number."""
    if not time_val:
        return 0
    s = str(time_val).strip().replace(":00", "")
    try:
        return int(s)
    except ValueError:
        return 0


def apply_feedback(feedbacks: list[dict]) -> dict:
    """Apply feedback to the daily plan.

    Handles both tweet feedback (Sheet 1) and reply feedback (Sheet 2).

    Tweet actions:
        approve  → mark tweet as approved
        replace  → use alt_text as the new tweet
        cancel   → cancel the tweet slot

    Reply actions:
        approve  → mark reply as approved
        replace  → use alt_text as the new reply
        cancel   → cancel this specific reply

    Returns: summary dict
    """
    if not PLAN_FILE.exists():
        logger.error("No plan file to modify")
        return {"error": "no plan file"}

    with open(PLAN_FILE, "r", encoding="utf-8") as f:
        plan = json.load(f)

    results = {"cancelled": [], "replaced": [], "approved": []}

    for fb in feedbacks:
        slot = fb["slot"]
        action = fb["action"]
        alt_text = fb.get("alt_text", "")
        slot_key = str(slot)
        sheet = fb.get("sheet", "내 트윗")

        if slot_key not in plan.get("slots", {}):
            continue

        slot_data = plan["slots"][slot_key]

        if sheet == "내 트윗":
            # Tweet feedback
            label = f"[{slot}:00 tweet]"

            if action == "cancel":
                slot_data["cancelled"] = True
                results["cancelled"].append(f"{slot}:tweet")
                logger.info(f"{label} CANCELLED")

            elif action == "replace":
                from twitter_utils import count_weighted_chars
                slot_data["tweet_jp"] = alt_text
                slot_data["chars"] = count_weighted_chars(alt_text)
                slot_data["replaced"] = True
                slot_data["approved"] = True
                results["replaced"].append(f"{slot}:tweet")
                logger.info(f"{label} REPLACED: {alt_text[:50]}...")

            elif action == "approve":
                slot_data["approved"] = True
                results["approved"].append(f"{slot}:tweet")
                logger.info(f"{label} CONFIRMED")

        else:
            # Reply feedback
            reply_idx = fb.get("reply_index", 0)
            replies = slot_data.get("replies", [])
            label = f"[{slot}:00 reply#{reply_idx}]"

            if reply_idx >= len(replies):
                logger.warning(f"{label} reply_index out of range (have {len(replies)} replies)")
                continue

            reply = replies[reply_idx]

            if action == "cancel":
                reply["cancelled"] = True
                results["cancelled"].append(f"{slot}:reply#{reply_idx}")
                logger.info(f"{label} CANCELLED")

            elif action == "replace":
                from twitter_utils import count_weighted_chars
                reply["reply_jp"] = alt_text
                reply["reply_chars"] = count_weighted_chars(alt_text)
                reply["replaced"] = True
                reply["approved"] = True
                results["replaced"].append(f"{slot}:reply#{reply_idx}")
                logger.info(f"{label} REPLACED: {alt_text[:50]}...")

            elif action == "approve":
                reply["approved"] = True
                results["approved"].append(f"{slot}:reply#{reply_idx}")
                logger.info(f"{label} CONFIRMED")

    # Save updated plan
    with open(PLAN_FILE, "w", encoding="utf-8") as f:
        json.dump(plan, f, ensure_ascii=False, indent=2)

    return results


def check_and_apply(apply: bool = False) -> list[dict]:
    """Full cycle: download → read → optionally apply."""
    excel_path = download_plan_excel()
    if not excel_path:
        return []

    feedbacks = read_feedback(excel_path)

    if not feedbacks:
        logger.info("No feedback found in Excel")
        return []

    logger.info(f"Found {len(feedbacks)} feedback(s):")
    for fb in feedbacks:
        logger.info(f"  [{fb['slot']}:00] {fb['action']}: {fb.get('alt_text', '')[:50]}")

    if apply:
        results = apply_feedback(feedbacks)
        logger.info(f"Applied: {results}")

        # Re-generate Excel and re-upload
        if results["replaced"] or results["cancelled"]:
            from generate_daily_excel import create_daily_excel, load_plan
            updated_plan = load_plan()
            new_excel = create_daily_excel(updated_plan)
            upload_file(new_excel, notify=True, message="트윗 플랜 업데이트 (팀 피드백 반영)")
            logger.info("Updated Excel re-uploaded to Teams")

    return feedbacks


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Read team feedback from Excel on Teams")
    parser.add_argument("--check", action="store_true", help="Download and check for feedback")
    parser.add_argument("--apply", action="store_true", help="Apply feedback changes")
    parser.add_argument("--poll", action="store_true", help="Poll every 60s")
    args = parser.parse_args()

    if args.poll:
        import time
        logger.info("Polling mode: checking every 60s...")
        while True:
            check_and_apply(apply=True)
            time.sleep(60)
    elif args.check:
        check_and_apply(apply=args.apply)
    else:
        parser.print_help()
