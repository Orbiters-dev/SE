"""
kpi_style_engine.py - 포맷이 Core: Centralized style constants and formatting utilities.

Part of 골만이 Squad (formatter role).
All KPI Excel style definitions are centralized here to prevent duplication.

Usage:
    from kpi_style_engine import STYLES, apply_header_style, format_number_cell
"""

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ── Color Palette ─────────────────────────────────────────────────────────────

class Colors:
    """ORBI KPI report color palette."""
    HEADER_BG     = "002060"    # Dark blue (headers, grand totals)
    SECTION_BG    = "D6DCE4"    # Light grey (section headers)
    TOTAL_BG      = "FFF2CC"    # Light yellow (subtotals)
    GTOTAL_BG     = "002060"    # Dark blue (grand total)
    NM_BG         = "595959"    # Dark grey (n.m = not measured)
    OK_BG         = "C6EFCE"    # Green (status OK)
    WARN_BG       = "FFF2CC"    # Yellow (status WARN)
    FAIL_BG       = "FFC7CE"    # Red (status FAIL)
    SUBTOTAL_BG   = "D6E4F0"    # Light blue (brand subtotals)
    WHITE         = "FFFFFF"
    BLACK         = "000000"
    GREY          = "595959"
    LIGHT_GREY    = "808080"
    BORDER_GREY   = "D9D9D9"


# ── Pre-built Styles ─────────────────────────────────────────────────────────

class Fills:
    HEADER   = PatternFill("solid", fgColor=Colors.HEADER_BG)
    SECTION  = PatternFill("solid", fgColor=Colors.SECTION_BG)
    TOTAL    = PatternFill("solid", fgColor=Colors.TOTAL_BG)
    GTOTAL   = PatternFill("solid", fgColor=Colors.GTOTAL_BG)
    NM       = PatternFill("solid", fgColor=Colors.NM_BG)
    OK       = PatternFill("solid", fgColor=Colors.OK_BG)
    WARN     = PatternFill("solid", fgColor=Colors.WARN_BG)
    FAIL     = PatternFill("solid", fgColor=Colors.FAIL_BG)
    SUBTOTAL = PatternFill("solid", fgColor=Colors.SUBTOTAL_BG)


class Fonts:
    WHITE       = Font(bold=True, color=Colors.WHITE)
    WHITE_SM    = Font(bold=True, color=Colors.WHITE, size=11)
    WHITE_LG    = Font(bold=True, color=Colors.WHITE, size=14)
    BOLD        = Font(bold=True)
    BOLD_11     = Font(bold=True, size=11)
    NORMAL      = Font(size=11)
    ITALIC_GREY = Font(italic=True, color=Colors.GREY)
    NM          = Font(color=Colors.WHITE, size=8)
    SMALL_GREY  = Font(size=9, color=Colors.LIGHT_GREY)


class Aligns:
    CENTER    = Alignment(horizontal="center", vertical="center")
    CENTER_W  = Alignment(horizontal="center", wrap_text=True)
    LEFT      = Alignment(horizontal="left", vertical="center")
    RIGHT     = Alignment(horizontal="right", vertical="center")
    NM        = Alignment(horizontal="center")


class Borders:
    THIN = Border(
        left=Side(style="thin", color=Colors.BORDER_GREY),
        right=Side(style="thin", color=Colors.BORDER_GREY),
        top=Side(style="thin", color=Colors.BORDER_GREY),
        bottom=Side(style="thin", color=Colors.BORDER_GREY),
    )


# ── Convenience Bundle ────────────────────────────────────────────────────────

STYLES = {
    "fills": Fills,
    "fonts": Fonts,
    "aligns": Aligns,
    "borders": Borders,
    "colors": Colors,
}


# ── Helper Functions ──────────────────────────────────────────────────────────

def apply_header_style(ws, n_cols: int, row: int = 1):
    """Apply dark blue header style to a row."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = Fills.HEADER
        cell.font = Fonts.WHITE
        cell.alignment = Aligns.CENTER


def apply_nm_style(cell):
    """Apply n.m (not measured) style to a cell."""
    cell.fill = Fills.NM
    cell.font = Fonts.NM
    cell.alignment = Aligns.NM


def format_number_cell(cell, value, fmt: str = '#,##0'):
    """Set cell value and number format."""
    cell.value = value
    if isinstance(value, (int, float)):
        cell.number_format = fmt


def apply_status_fill(cell, status: str):
    """Apply OK/WARN/FAIL fill based on status string."""
    status_map = {
        "OK": Fills.OK,
        "PASS": Fills.OK,
        "WARN": Fills.WARN,
        "STALE": Fills.WARN,
        "FAIL": Fills.FAIL,
        "NO DATA": Fills.FAIL,
    }
    fill = status_map.get(status.upper(), Fills.WARN)
    cell.fill = fill


def set_column_widths(ws, label_width: int = 24, data_width: int = 14, max_cols: int = 20):
    """Set standard column widths: wide label column + uniform data columns."""
    from openpyxl.utils import get_column_letter
    ws.column_dimensions["A"].width = label_width
    for i in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = data_width
