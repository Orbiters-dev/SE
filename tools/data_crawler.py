"""Data Crawler — Extract filtered data from large spreadsheets to Excel.

Handles Google Sheets (50K+ rows via gspread) and local Excel files
(true streaming via openpyxl read_only). Supports structured filters,
column selection, caching, and dry-run mode.

Usage:
    # Google Sheet with filter
    python tools/data_crawler.py --source "sheet:SHEET_ID" --filter "region=JP" --columns "name,email"

    # Local Excel, dry run
    python tools/data_crawler.py --source "file:data.xlsx" --filter "sales>1000000" --dry-run

    # Full Google Sheets URL
    python tools/data_crawler.py --source "sheet:https://docs.google.com/spreadsheets/d/1ABC.../edit"

    # Clear cache
    python tools/data_crawler.py --clear-cache
"""

import os, sys, re, json, time, argparse, shutil, io
from pathlib import Path
from datetime import datetime

# Fix Windows cp949 encoding issues
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

DIR = Path(__file__).resolve().parent
PROJECT_ROOT = DIR.parent
sys.path.insert(0, str(DIR))

try:
    from env_loader import load_env
    load_env()
except Exception:
    pass

CACHE_DIR = PROJECT_ROOT / ".tmp" / "data_crawler" / "cache"
OUTPUT_DIR = PROJECT_ROOT / ".tmp" / "data_crawler" / "output"
ILLEGAL_CHARS = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')
PROGRESS_INTERVAL = 5000


# ── Auth ─────────────────────────────────────────────────────────────

def get_gc():
    """Authenticate with Google Sheets via service account."""
    from google.oauth2.service_account import Credentials
    import gspread
    creds = Credentials.from_service_account_file(
        str(PROJECT_ROOT / "credentials" / "google_service_account.json"),
        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"],
    )
    return gspread.authorize(creds)


# ── Source Readers ───────────────────────────────────────────────────

def parse_sheet_id(url_or_id):
    """Extract Google Sheet ID from URL or return raw ID."""
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url_or_id)
    return m.group(1) if m else url_or_id


def download_sheet_data(gc, sheet_id, tab_name, max_retries=3):
    """Download sheet tab with retry for 503 errors."""
    sh = gc.open_by_key(sheet_id)
    # Resolve tab name
    if tab_name:
        ws = sh.worksheet(tab_name)
    else:
        ws = sh.get_worksheet(0)
        tab_name = ws.title
    for attempt in range(max_retries):
        try:
            print(f"  Downloading '{tab_name}' (attempt {attempt+1})...", flush=True)
            return ws.get_all_values(), tab_name
        except Exception as e:
            if "503" in str(e) and attempt < max_retries - 1:
                wait = 30 * (attempt + 1)
                print(f"  503 error, retrying in {wait}s...", flush=True)
                time.sleep(wait)
            else:
                raise


def iter_sheet_rows(sheet_id_or_url, tab_name=None, use_cache=True):
    """Generator: yields (headers, row_dict) from Google Sheet. Caches to JSON."""
    sheet_id = parse_sheet_id(sheet_id_or_url)
    cache_safe_tab = (tab_name or "Sheet1").replace(" ", "_")
    cache_file = CACHE_DIR / f"cache_{sheet_id}_{cache_safe_tab}.json"

    if use_cache and cache_file.exists():
        print(f"  Using cache: {cache_file.name}", flush=True)
        with open(cache_file, "r", encoding="utf-8") as f:
            raw = json.load(f)
        actual_tab = tab_name or "Sheet1"
    else:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        gc = get_gc()
        raw, actual_tab = download_sheet_data(gc, sheet_id, tab_name)
        with open(cache_file, "w", encoding="utf-8") as f:
            json.dump(raw, f, ensure_ascii=False)
        print(f"  Cached {len(raw)} rows to {cache_file.name}", flush=True)

    if not raw:
        return
    headers = [h.strip() for h in raw[0]]
    for i, row in enumerate(raw[1:]):
        padded = list(row) + [""] * max(0, len(headers) - len(row))
        yield headers, dict(zip(headers, padded[:len(headers)]))
        if (i + 1) % PROGRESS_INTERVAL == 0:
            print(f"  ... {i+1} rows processed", flush=True)


def iter_excel_rows(filepath, sheet_name=None):
    """Generator: yields (headers, row_dict) using openpyxl read_only (true streaming)."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    print(f"  Reading Excel: {filepath} (tab: {ws.title})", flush=True)
    headers = None
    row_count = 0
    try:
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [
                    str(h).strip() if h is not None else f"col_{i}"
                    for i, h in enumerate(row)
                ]
                continue
            padded = list(row) + [None] * max(0, len(headers) - len(row))
            row_count += 1
            yield headers, dict(zip(headers, padded[:len(headers)]))
            if row_count % PROGRESS_INTERVAL == 0:
                print(f"  ... {row_count} rows processed", flush=True)
    finally:
        wb.close()
    print(f"  Total rows in file: {row_count}", flush=True)


def list_tabs(source_type, source_val):
    """List available tab names for a source."""
    if source_type == "sheet":
        gc = get_gc()
        sheet_id = parse_sheet_id(source_val)
        sh = gc.open_by_key(sheet_id)
        tabs = [ws.title for ws in sh.worksheets()]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(source_val, read_only=True)
        tabs = wb.sheetnames
        wb.close()
    return tabs


# ── Filter Engine ────────────────────────────────────────────────────

def parse_filters(filter_str):
    """Parse 'region=JP,sales>1000000' into [(col, op, val), ...]."""
    if not filter_str:
        return []
    specs = []
    # Split on comma, but respect quoted values
    parts = [p.strip() for p in filter_str.split(",")]
    for part in parts:
        if not part:
            continue
        matched = False
        for op in ["!=", ">=", "<=", ">", "<", "~", "^", "="]:
            idx = part.find(op)
            if idx > 0:
                col = part[:idx].strip()
                val = part[idx + len(op):].strip()
                specs.append((col, op, val))
                matched = True
                break
        if not matched:
            print(f"  [WARN] Could not parse filter: {part}", flush=True)
    return specs


def _coerce_numeric(val):
    """Try to parse val as float."""
    if val is None or val == "":
        return None
    try:
        cleaned = str(val).replace(",", "").replace(" ", "").strip()
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def apply_filters(row_dict, filter_specs):
    """Returns True if row passes ALL filters (AND logic)."""
    for col, op, target in filter_specs:
        cell_val = row_dict.get(col)
        if cell_val is None:
            cell_val = ""
        cell_str = str(cell_val).strip()

        if op == "=":
            if cell_str.lower() != target.lower():
                return False
        elif op == "!=":
            if cell_str.lower() == target.lower():
                return False
        elif op in (">", "<", ">=", "<="):
            num = _coerce_numeric(cell_val)
            tgt = _coerce_numeric(target)
            if num is None or tgt is None:
                return False
            if op == ">" and not (num > tgt):
                return False
            if op == "<" and not (num < tgt):
                return False
            if op == ">=" and not (num >= tgt):
                return False
            if op == "<=" and not (num <= tgt):
                return False
        elif op == "~":
            if target.lower() not in cell_str.lower():
                return False
        elif op == "^":
            if not re.search(target, cell_str, re.IGNORECASE):
                return False
    return True


# ── Column Selector ──────────────────────────────────────────────────

def resolve_columns(headers, columns_arg):
    """Resolve --columns arg to actual header list. Returns list of column names."""
    if not columns_arg or columns_arg.lower() == "all":
        return headers[:]
    selected = [c.strip() for c in columns_arg.split(",")]
    headers_lower = {h.lower(): h for h in headers}
    resolved = []
    for col in selected:
        canonical = headers_lower.get(col.lower())
        if canonical:
            resolved.append(canonical)
        else:
            print(f"  [WARN] Column not found: '{col}' (available: {', '.join(headers[:10])}...)", flush=True)
    return resolved


def select_row_values(row_dict, col_list):
    """Extract values from row_dict in col_list order."""
    return [row_dict.get(c, "") for c in col_list]


# ── Excel Writer ─────────────────────────────────────────────────────

def clean_cell(val):
    """Remove illegal XML characters for Excel safety."""
    if isinstance(val, str):
        return ILLEGAL_CHARS.sub("", val)
    if val is None:
        return ""
    return val


def write_excel(output_path, headers, row_value_iter):
    """Stream-write filtered rows to Excel using write_only mode. Returns row count."""
    import openpyxl
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Data")
    ws.append([clean_cell(h) for h in headers])
    count = 0
    for row_vals in row_value_iter:
        ws.append([clean_cell(v) for v in row_vals])
        count += 1
        if count % PROGRESS_INTERVAL == 0:
            print(f"  ... {count} rows written", flush=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return count


# ── Output Path Resolution ───────────────────────────────────────────

def resolve_output_path(output_arg):
    """Resolve output path with NAS Z: drive fallback per CLAUDE.md."""
    if not output_arg:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return OUTPUT_DIR / f"output_{ts}.xlsx"
    p = Path(output_arg)
    if not p.is_absolute():
        # Ensure .xlsx extension
        if not p.suffix:
            p = p.with_suffix(".xlsx")
        return OUTPUT_DIR / p
    # NAS Z: drive fallback
    p_str = str(p)
    if p_str.upper().startswith("Z:"):
        if p.exists() or p.parent.exists():
            return p
        fallback = Path(p_str.replace("Z:", r"C:\SynologyDrive", 1).replace("z:", r"C:\SynologyDrive", 1))
        if fallback.parent.exists():
            print(f"  [NAS] Z: unavailable, using SynologyDrive fallback", flush=True)
            return fallback
        print(f"  [WARN] NAS path unavailable, falling back to .tmp/", flush=True)
        return OUTPUT_DIR / p.name
    return p


# ── Source Parsing ───────────────────────────────────────────────────

def parse_source(source_str):
    """Parse --source arg into (type, value). Returns ('sheet'|'file', value)."""
    if source_str.startswith("sheet:"):
        return "sheet", source_str[6:]
    if source_str.startswith("file:"):
        return "file", source_str[5:]
    # Auto-detect
    if "docs.google.com" in source_str or "spreadsheets" in source_str:
        return "sheet", source_str
    if os.path.exists(source_str) or source_str.endswith((".xlsx", ".xls")):
        return "file", source_str
    # Default: assume sheet ID
    return "sheet", source_str


# ── Main ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Data Crawler — extract filtered data from spreadsheets to Excel"
    )
    parser.add_argument("--source",
        help='Source: "sheet:SHEET_ID_OR_URL" or "file:path.xlsx"')
    parser.add_argument("--filter", dest="filter_str", default="",
        help='Structured filters: "region=JP,sales>1000000"')
    parser.add_argument("--columns", default="all",
        help='Column selection: "name,email,sales" or "all"')
    parser.add_argument("--output", default=None,
        help="Output filename (.tmp/data_crawler/output/ unless absolute path)")
    parser.add_argument("--sheet-name", default=None,
        help="Tab name for multi-tab sheets (default: first tab)")
    parser.add_argument("--dry-run", action="store_true",
        help="Show row count + sample without writing Excel")
    parser.add_argument("--no-cache", action="store_true",
        help="Force re-download (skip cache)")
    parser.add_argument("--clear-cache", action="store_true",
        help="Delete all cached data and exit")
    parser.add_argument("--list-tabs", action="store_true",
        help="List available tab names and exit")
    args = parser.parse_args()

    # Handle clear-cache
    if args.clear_cache:
        if CACHE_DIR.exists():
            shutil.rmtree(CACHE_DIR)
            print(f"Cache cleared: {CACHE_DIR}")
        else:
            print("No cache to clear.")
        return

    if not args.source:
        parser.error("--source is required (unless using --clear-cache)")

    source_type, source_val = parse_source(args.source)
    print(f"Source: {source_type} = {source_val}", flush=True)

    # Handle list-tabs
    if args.list_tabs:
        tabs = list_tabs(source_type, source_val)
        print(f"Available tabs ({len(tabs)}):")
        for i, t in enumerate(tabs):
            print(f"  {i+1}. {t}")
        return

    # Parse filters
    filter_specs = parse_filters(args.filter_str)
    if filter_specs:
        print(f"Filters ({len(filter_specs)}): {filter_specs}", flush=True)
    else:
        print("Filters: none (all rows)", flush=True)

    # Create row iterator
    if source_type == "sheet":
        row_gen = iter_sheet_rows(source_val, args.sheet_name, use_cache=not args.no_cache)
    else:
        row_gen = iter_excel_rows(source_val, sheet_name=args.sheet_name)

    # ── DRY RUN ──
    if args.dry_run:
        headers = None
        total_read = 0
        total_matched = 0
        samples = []
        for hdrs, row_dict in row_gen:
            if headers is None:
                headers = hdrs
            total_read += 1
            if apply_filters(row_dict, filter_specs):
                total_matched += 1
                if len(samples) < 5:
                    sel_cols = resolve_columns(hdrs, args.columns)
                    samples.append(dict(zip(sel_cols, select_row_values(row_dict, sel_cols))))

        print(f"\n{'='*50}")
        print(f"[DRY RUN] Total rows read: {total_read:,}")
        print(f"[DRY RUN] Rows matching filters: {total_matched:,}")
        if headers:
            sel_cols = resolve_columns(headers, args.columns)
            print(f"[DRY RUN] Selected columns ({len(sel_cols)}): {', '.join(sel_cols)}")
        if samples:
            print(f"[DRY RUN] Sample ({len(samples)} rows):")
            for i, s in enumerate(samples):
                print(f"  Row {i+1}: {s}")
        print(f"{'='*50}")
        return

    # ── LIVE RUN ──
    output_path = resolve_output_path(args.output)
    print(f"Output: {output_path}", flush=True)

    # We need headers before writing. Peek first row.
    headers = None
    sel_cols = None
    total_read = 0
    total_matched = 0

    def filtered_stream():
        nonlocal headers, sel_cols, total_read, total_matched
        for hdrs, row_dict in row_gen:
            if headers is None:
                headers = hdrs
                sel_cols = resolve_columns(headers, args.columns)
            total_read += 1
            if apply_filters(row_dict, filter_specs):
                total_matched += 1
                yield select_row_values(row_dict, sel_cols)

    # Buffer first match to get headers
    stream = filtered_stream()
    first_row = None
    try:
        first_row = next(stream)
    except StopIteration:
        pass

    if sel_cols is None:
        print("\n[ERROR] No data found in source.", flush=True)
        return

    def full_stream():
        if first_row is not None:
            yield first_row
        yield from stream

    count = write_excel(output_path, sel_cols, full_stream())

    print(f"\n{'='*50}")
    print(f"Done!")
    print(f"  Rows read:    {total_read:,}")
    print(f"  Rows matched: {total_matched:,}")
    print(f"  Rows written: {count:,}")
    print(f"  Output:       {output_path}")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
