#!/usr/bin/env python3
"""
US Weekly Baby/Toddler Content Report
======================================
Independent pipeline (Syncly 무관):
  1. Apify IG reel-scraper → US baby creators' recent reels
  2. CI scoring (Whisper + Vision + Composite Score)
  3. Cross-check with Syncly (gk_content_posts) for comparison
  4. JOIN pipeline_creators for email / TikTok
  5. Export Excel → Email

Usage:
  python tools/us_weekly_report.py                    # full run
  python tools/us_weekly_report.py --dry-run           # no CI, no email
  python tools/us_weekly_report.py --skip-scrape        # use existing DB data only
  python tools/us_weekly_report.py --skip-ci            # scrape but skip Whisper/Vision
  python tools/us_weekly_report.py --days 14            # last 14 days instead of 7
  python tools/us_weekly_report.py --to me@email.com    # custom recipient
"""
import argparse
import json
import os
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "tools"))
from env_loader import load_env

CACHE_DIR = PROJECT_ROOT / ".tmp" / "us_weekly"
TODAY = datetime.now().strftime("%Y-%m-%d")

# Apify actor
IG_REEL_SCRAPER = "apify/instagram-reel-scraper"

# US baby/toddler hashtags for caption filtering
BABY_KEYWORDS = [
    "baby", "toddler", "momlife", "mom life", "momhack", "toddlermom",
    "sippy", "straw cup", "feeding", "weaning", "nursery", "diaper",
    "breastfeed", "formula", "highchair", "baby food", "babyfood",
    "toddlerlife", "newmom", "firsttimemom", "motherhood", "mamalife",
    "babytok", "momtok", "toddlerfood", "blw", "babyledweaning",
    "ppsu", "grosmimi", "onzenna", "bottle", "baby bottle",
]

# Brand/store accounts to skip
EXCLUDE_ACCOUNTS = {
    "onzenna.official", "grosmimi_usa", "grosmimi_japan", "grosmimi_official",
    "grosmimi_korea", "onzenna", "grosmimi", "zezebaebae",
    "grosmimithailand", "grosmimi_thailand",
}


def _get_db():
    import psycopg2
    return psycopg2.connect(
        host=os.getenv("DB_HOST", "172.31.13.240"),
        dbname=os.getenv("DB_NAME", "export_calculator_db"),
        user=os.getenv("DB_USER", "es_db_user"),
        password=os.getenv("DB_PASSWORD", "orbit1234"),
    )


def _caption_is_baby(caption: str) -> bool:
    """Check if caption contains baby/toddler keywords."""
    low = (caption or "").lower()
    return any(kw in low for kw in BABY_KEYWORDS)


import re
_ILLEGAL_XML_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f'
    r'\ud800-\udfff\ufdd0-\ufdef\ufffe\uffff]'
)


def _clean_for_excel(text):
    """Remove illegal XML characters that openpyxl rejects."""
    if not text:
        return text
    if not isinstance(text, str):
        return text
    return _ILLEGAL_XML_RE.sub('', text)


# ---------------------------------------------------------------------------
# Step 1: Get US baby creators from DB
# ---------------------------------------------------------------------------
def get_us_baby_creators(min_followers=5000, limit=150):
    """Get US creators who post baby/toddler content with good reach."""
    creators = []
    try:
        conn = _get_db()
        cur = conn.cursor()
        # Creators with baby-keyword posts and 5K+ views
        cur.execute("""
            SELECT username, MAX(followers) as max_f, COUNT(*) as cnt,
                   MAX(views_30d) as max_views
            FROM gk_content_posts
            WHERE region = 'us'
              AND username != '' AND username IS NOT NULL
              AND views_30d >= 5000
              AND (LOWER(caption) LIKE '%%baby%%' OR LOWER(caption) LIKE '%%toddler%%'
                   OR LOWER(caption) LIKE '%%momlife%%' OR LOWER(caption) LIKE '%%feeding%%'
                   OR LOWER(caption) LIKE '%%sippy%%' OR LOWER(caption) LIKE '%%weaning%%'
                   OR LOWER(caption) LIKE '%%diaper%%' OR LOWER(caption) LIKE '%%nursery%%'
                   OR LOWER(caption) LIKE '%%babyfood%%' OR LOWER(caption) LIKE '%%blw%%'
                   OR LOWER(caption) LIKE '%%motherhood%%' OR LOWER(caption) LIKE '%%momtok%%')
            GROUP BY username
            HAVING MAX(followers) >= %s OR MAX(views_30d) >= 50000
            ORDER BY MAX(views_30d) DESC NULLS LAST
            LIMIT %s
        """, (min_followers, limit))
        for row in cur.fetchall():
            creators.append({
                "username": row[0], "followers": row[1] or 0,
                "posts": row[2], "max_views": row[3] or 0,
            })
        conn.close()
        print(f"[DB] Found {len(creators)} US baby creators")
    except Exception as e:
        print(f"[DB WARN] {e}")
    return creators


# ---------------------------------------------------------------------------
# Step 2: Apify IG Reel Scraper
# ---------------------------------------------------------------------------
def scrape_ig_reels(client, creators, min_views=5000, max_per_creator=5, days=7):
    """Fetch recent reels from US baby creators."""
    usernames = [c["username"] for c in creators
                 if c["username"].lower() not in EXCLUDE_ACCOUNTS]
    print(f"\n[IG REELS] Scraping {len(usernames)} US creators (max {max_per_creator}/each)...")

    all_reels = []
    cutoff = datetime.now() - timedelta(days=days)

    # Batch in chunks of 5
    for i in range(0, len(usernames), 5):
        chunk = usernames[i:i+5]
        print(f"  Batch {i//5+1}/{(len(usernames)+4)//5}: {', '.join(f'@{u}' for u in chunk)}")
        try:
            run = client.actor(IG_REEL_SCRAPER).call(
                run_input={"username": chunk, "resultsLimit": max_per_creator},
                timeout_secs=300,
            )
            items = list(client.dataset(run["defaultDatasetId"]).iterate_items())

            for item in items:
                username = item.get("ownerUsername", "") or ""
                views = (item.get("videoPlayCount", 0) or item.get("videoViewCount", 0)
                         or item.get("video_play_count", 0) or 0)
                if views < min_views:
                    continue

                caption = (item.get("caption", "") or "")[:500]
                if not _caption_is_baby(caption):
                    continue

                shortcode = item.get("shortCode", "") or ""
                taken = (item.get("timestamp", "") or "")[:10]
                post_url = f"https://www.instagram.com/reel/{shortcode}/" if shortcode else ""

                # Date filter
                try:
                    post_date = datetime.strptime(taken, "%Y-%m-%d")
                    if post_date < cutoff:
                        continue
                except ValueError:
                    pass

                all_reels.append({
                    "platform": "instagram",
                    "username": username,
                    "shortcode": shortcode,
                    "post_url": post_url,
                    "views": views,
                    "likes": item.get("likesCount", 0) or 0,
                    "comments": item.get("commentsCount", 0) or 0,
                    "caption": caption,
                    "date": taken,
                    "followers": item.get("ownerFullName", ""),  # will enrich later
                    "source": "us_weekly_report",
                })

            print(f"    -> {len(items)} reels, {len(all_reels)} passed filter so far")
        except Exception as e:
            print(f"    FAIL: {e}")

        time.sleep(2)

    print(f"[IG REELS] {len(all_reels)} US baby reels with {min_views:,}+ views (last {days}d)")
    return all_reels


# ---------------------------------------------------------------------------
# Step 3: Push to DB
# ---------------------------------------------------------------------------
def push_to_db(reels):
    """Insert new reels into gk_content_posts, skip duplicates."""
    if not reels:
        return 0

    conn = _get_db()
    cur = conn.cursor()

    # Get existing URLs
    cur.execute("""SELECT url FROM gk_content_posts
                   WHERE region='us' AND source='us_weekly_report'""")
    existing = {r[0] for r in cur.fetchall() if r[0]}

    inserted = 0
    for r in reels:
        if r["post_url"] in existing:
            continue
        try:
            cur.execute("""
                INSERT INTO gk_content_posts
                (post_id, url, platform, username, caption, post_date, region, source,
                 collected_at, views_30d, likes_30d, comments_30d)
                VALUES (%s, %s, %s, %s, %s, %s, 'us', 'us_weekly_report', %s, %s, %s, %s)
                ON CONFLICT DO NOTHING
            """, (
                r["shortcode"], r["post_url"], r["platform"], r["username"],
                r["caption"], r["date"] if r["date"] else None,
                datetime.now().isoformat(),
                r["views"], r["likes"], r["comments"],
            ))
            inserted += 1
        except Exception as e:
            print(f"  ! @{r['username']} INSERT error: {e}")
            conn.rollback()
            conn = _get_db()
            cur = conn.cursor()

    conn.commit()
    conn.close()
    print(f"[DB] Inserted {inserted} new US reels")
    return inserted


# ---------------------------------------------------------------------------
# Step 4: Run CI scoring (calls analyze_video_content.py logic)
# ---------------------------------------------------------------------------
def run_ci_scoring(max_posts=100, min_views=5000):
    """Run CI on US posts without transcript."""
    print(f"\n[CI] Running Whisper+Vision+Score on US posts (max {max_posts})...")
    try:
        from analyze_video_content import run_ci_pipeline
        results = run_ci_pipeline(region="us", max_posts=max_posts, min_views=min_views)
        print(f"[CI] Completed: {len(results)} posts scored")
        return results
    except ImportError:
        # Fallback: call as subprocess
        import subprocess
        cmd = [
            sys.executable, str(PROJECT_ROOT / "tools" / "analyze_video_content.py"),
            "--region", "us", "--max", str(max_posts), "--min-views", str(min_views),
        ]
        print(f"[CI] Running subprocess: {' '.join(cmd)}")
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=1800)
        print(proc.stdout[-2000:] if proc.stdout else "(no output)")
        if proc.returncode != 0:
            print(f"[CI WARN] exit code {proc.returncode}")
            if proc.stderr:
                print(proc.stderr[-500:])
        return []


# ---------------------------------------------------------------------------
# Step 5: Export Excel
# ---------------------------------------------------------------------------
def export_excel(days=7, output_path=None):
    """Query DB and build Excel report with Syncly cross-check."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[ERROR] openpyxl not installed. pip install openpyxl")
        return None

    conn = _get_db()
    cur = conn.cursor()

    # Main query: US baby posts from last N days, joined with pipeline_creators
    cur.execute("""
        SELECT
            cp.username,
            cp.url,
            cp.post_date,
            cp.views_30d,
            cp.likes_30d,
            cp.comments_30d,
            cp.followers,
            cp.caption,
            cp.transcript,
            cp.content_quality_score,
            cp.creator_fit_score,
            cp.brand_fit_score,
            cp.engagement_rate,
            cp.source,
            pc.email,
            pc.tiktok_handle,
            pc.pipeline_status,
            pc.country
        FROM gk_content_posts cp
        LEFT JOIN onz_pipeline_creators pc
            ON LOWER(cp.username) = LOWER(pc.ig_handle)
        WHERE cp.region = 'us'
          AND cp.views_30d >= 5000
          AND cp.post_date >= CURRENT_DATE - INTERVAL '%s days'
          AND (LOWER(cp.caption) LIKE '%%baby%%' OR LOWER(cp.caption) LIKE '%%toddler%%'
               OR LOWER(cp.caption) LIKE '%%momlife%%' OR LOWER(cp.caption) LIKE '%%feeding%%'
               OR LOWER(cp.caption) LIKE '%%sippy%%' OR LOWER(cp.caption) LIKE '%%weaning%%'
               OR LOWER(cp.caption) LIKE '%%diaper%%' OR LOWER(cp.caption) LIKE '%%nursery%%'
               OR LOWER(cp.caption) LIKE '%%babyfood%%' OR LOWER(cp.caption) LIKE '%%blw%%'
               OR LOWER(cp.caption) LIKE '%%motherhood%%' OR LOWER(cp.caption) LIKE '%%momtok%%'
               OR LOWER(cp.caption) LIKE '%%baby bottle%%' OR LOWER(cp.caption) LIKE '%%highchair%%')
        ORDER BY cp.views_30d DESC
    """ % days)
    rows = cur.fetchall()

    # Syncly cross-check: which posts also exist in Syncly source
    cur.execute("""
        SELECT url FROM gk_content_posts
        WHERE region='us' AND source IN ('syncly', 'apify_daily', 'content_sync')
    """)
    syncly_urls = {r[0] for r in cur.fetchall() if r[0]}
    conn.close()

    if not rows:
        print(f"[EXCEL] No US baby posts found in last {days} days")
        return None

    print(f"[EXCEL] Building report: {len(rows)} posts")

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"US Baby Content {TODAY}"

    # Header style
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = [
        "Creator (IG)", "TikTok", "Email", "Followers", "Views",
        "Likes", "Comments", "ER%", "Upload Date", "Post URL",
        "Quality Score", "Fit Score", "Brand Fit",
        "Transcript", "Caption (100ch)",
        "Syncly Match", "Pipeline Status", "Source",
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Score color coding
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_idx, r in enumerate(rows, 2):
        username, url, post_date, views, likes, comments, followers, caption, \
            transcript, quality, fit, brand_fit, er, source, \
            email, tiktok, pipeline_status, country = r

        # Syncly cross-check
        in_syncly = "YES" if url in syncly_urls else "NO"

        # ER calculation
        if not er and followers and followers > 0:
            er = round(((likes or 0) + (comments or 0)) / followers * 100, 2)

        vals = [
            f"@{username}",
            f"@{tiktok}" if tiktok else "",
            email or "",
            followers or 0,
            views or 0,
            likes or 0,
            comments or 0,
            er or 0,
            str(post_date) if post_date else "",
            url or "",
            quality or 0,
            fit or 0,
            brand_fit or 0,
            (transcript or "")[:300],
            (caption or "")[:100],
            in_syncly,
            pipeline_status or "",
            source or "",
        ]

        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=_clean_for_excel(v))
            cell.border = thin_border

        # Color code scores
        q_cell = ws.cell(row=row_idx, column=11)
        if quality and quality >= 60:
            q_cell.fill = green_fill
        elif quality and quality >= 30:
            q_cell.fill = yellow_fill
        elif quality:
            q_cell.fill = red_fill

        f_cell = ws.cell(row=row_idx, column=12)
        if fit and fit >= 60:
            f_cell.fill = green_fill
        elif fit and fit >= 30:
            f_cell.fill = yellow_fill
        elif fit:
            f_cell.fill = red_fill

    # Column widths
    widths = [18, 15, 25, 10, 12, 10, 10, 8, 12, 40, 12, 10, 10, 50, 30, 12, 15, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    summaries = [
        ["US Baby Content Weekly Report", ""],
        ["Date", TODAY],
        ["Period", f"Last {days} days"],
        ["Total Posts", len(rows)],
        ["Unique Creators", len(set(r[0] for r in rows))],
        ["Posts with Transcript", sum(1 for r in rows if r[8])],
        ["Posts with Score", sum(1 for r in rows if r[9])],
        ["Avg Quality Score", round(sum(r[9] or 0 for r in rows) / max(len(rows), 1), 1)],
        ["Avg Fit Score", round(sum(r[10] or 0 for r in rows) / max(len(rows), 1), 1)],
        ["Syncly Matches", sum(1 for r in rows if r[1] in syncly_urls)],
        ["Posts with Email", sum(1 for r in rows if r[14])],
        ["Posts with TikTok", sum(1 for r in rows if r[15])],
        ["Top Views", max(r[3] or 0 for r in rows)],
    ]
    for row_idx, (k, v) in enumerate(summaries, 1):
        ws2.cell(row=row_idx, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=row_idx, column=2, value=v)
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 20

    # Save
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    if not output_path:
        output_path = CACHE_DIR / f"US_Baby_Report_{TODAY}.xlsx"
    wb.save(output_path)
    print(f"[EXCEL] Saved: {output_path}")
    return str(output_path)


# ---------------------------------------------------------------------------
# Step 6: Send email
# ---------------------------------------------------------------------------
def send_report_email(excel_path, to_email="wj.choi@orbiters.co.kr"):
    """Send Excel report via Gmail."""
    if not excel_path or not Path(excel_path).exists():
        print("[EMAIL] No Excel file to send")
        return False

    try:
        import subprocess
        cmd = [
            sys.executable, str(PROJECT_ROOT / "tools" / "send_gmail.py"),
            "--to", to_email,
            "--subject", f"[Social Deep Crawl] US Baby Content Weekly Report {TODAY}",
            "--body", f"""<h2>US Baby/Toddler Content Weekly Report</h2>
<p>Date: {TODAY}</p>
<p>Attached: Excel report with creator data, transcripts, and scores.</p>
<ul>
  <li>Views 5,000+ baby/toddler content from the past week</li>
  <li>CI scored (Content Quality + Creator-Brand Fit)</li>
  <li>Syncly cross-check included</li>
  <li>Pipeline creator email + TikTok matched</li>
</ul>
<p style="color:#666;font-size:12px;">Auto-generated by Social Deep Crawling pipeline</p>""",
            "--attachment", str(excel_path),
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        if result.returncode == 0:
            print(f"[EMAIL] Sent to {to_email}")
            return True
        else:
            print(f"[EMAIL ERROR] {result.stderr[:500]}")
            return False
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")
        return False


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="US Weekly Baby Content Report")
    parser.add_argument("--days", type=int, default=7, help="Look back N days (default 7)")
    parser.add_argument("--min-views", type=int, default=5000)
    parser.add_argument("--max-creators", type=int, default=100, help="Max creators to scrape")
    parser.add_argument("--max-ci", type=int, default=50, help="Max posts for CI scoring")
    parser.add_argument("--to", default="wj.choi@orbiters.co.kr", help="Email recipient")
    parser.add_argument("--dry-run", action="store_true", help="Skip CI and email")
    parser.add_argument("--skip-scrape", action="store_true", help="Skip Apify scraping, use DB only")
    parser.add_argument("--skip-ci", action="store_true", help="Skip Whisper/Vision CI")
    parser.add_argument("--skip-email", action="store_true", help="Skip email, just generate Excel")
    args = parser.parse_args()

    print(f"{'='*60}")
    print(f"  US Baby Content Weekly Report — {TODAY}")
    print(f"  Days: {args.days} | Min views: {args.min_views:,} | Max CI: {args.max_ci}")
    print(f"{'='*60}\n")

    load_env()

    # Step 1-3: Scrape + Push
    if not args.skip_scrape and not args.dry_run:
        from apify_client import ApifyClient
        token = os.environ.get("APIFY_API_TOKEN", "")
        if not token:
            print("[WARN] No APIFY_API_TOKEN, skipping scrape")
        else:
            client = ApifyClient(token)
            creators = get_us_baby_creators(limit=args.max_creators)
            if creators:
                reels = scrape_ig_reels(
                    client, creators, args.min_views,
                    max_per_creator=5, days=args.days,
                )
                if reels:
                    push_to_db(reels)
    elif args.skip_scrape:
        print("[SKIP] Apify scraping")

    # Step 4: CI scoring
    if not args.skip_ci and not args.dry_run:
        run_ci_scoring(max_posts=args.max_ci, min_views=args.min_views)
    elif args.skip_ci or args.dry_run:
        print("[SKIP] CI scoring")

    # Step 5: Export Excel
    excel_path = export_excel(days=args.days)

    # Step 6: Email
    if excel_path and not args.skip_email and not args.dry_run:
        send_report_email(excel_path, args.to)
    elif args.dry_run:
        print(f"[DRY RUN] Would email to {args.to}")

    print(f"\n{'='*60}")
    print(f"  Done! Report: {excel_path or 'N/A'}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
