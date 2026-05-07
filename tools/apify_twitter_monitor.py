"""
Apify Twitter Monitor — 競合ブランドのX(Twitter)アカウントを自動モニタリング

Apifyの twitter-scraper を使い、競合ブランドの最新ツイート・フォロワー数・
エンゲージメントを構造化データとして収集する。
既存の twitter_research.py (Firecrawl版) を補完・代替する安定版。

出力:
  - JSON: .tmp/apify_twitter_monitor_{date}.json
  - Excel: .tmp/apify_twitter_monitor_{date}.xlsx (Teams送信用)
  - Teams: Webhook通知 (オプション)

Usage:
    python tools/apify_twitter_monitor.py                     # 全競合スキャン
    python tools/apify_twitter_monitor.py --dry-run            # クロールのみ(保存なし)
    python tools/apify_twitter_monitor.py --brand ピジョン      # 特定ブランドのみ
    python tools/apify_twitter_monitor.py --json               # JSON出力
    python tools/apify_twitter_monitor.py --notify             # Teams通知付き
"""

import argparse
import json
import os
import sys
import logging
from pathlib import Path
from datetime import datetime, timezone, timedelta

from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).parent.parent
TMP_DIR = PROJECT_ROOT / ".tmp"
TMP_DIR.mkdir(exist_ok=True)
JST = timezone(timedelta(hours=9))

# ═══════════════════════════════════════════════════════════════════════
# CONFIG: 調査対象ブランド (twitter_research.py と同じリスト)
# ═══════════════════════════════════════════════════════════════════════

COMPETITOR_BRANDS = [
    # 競合 (ベビー用品)
    {"name": "ピジョン",      "handle": "pigeon_official_jp", "category": "競合"},
    {"name": "コンビ",        "handle": "combi_jp",           "category": "競合"},
    {"name": "リッチェル",    "handle": "richell_jp",         "category": "競合"},
    {"name": "NUKジャパン",   "handle": "nuk_japan",          "category": "競合"},
    {"name": "マンチキン",    "handle": "munchkin_japan",     "category": "競合"},
    {"name": "ファルスカ",    "handle": "farska_jp",          "category": "競合"},
    {"name": "カトージ",      "handle": "katoji_official",    "category": "競合"},
    {"name": "Joie",          "handle": "joie_japan",         "category": "競合"},
    # 参考ブランド
    {"name": "アカチャンホンポ", "handle": "akachanhonpo",    "category": "参考"},
    {"name": "西松屋",          "handle": "nishimatsuya_com", "category": "参考"},
    {"name": "無印良品",        "handle": "muji_net",         "category": "参考"},
]

# Apify actor for Twitter/X scraping
TWITTER_ACTOR = "apidojo/tweet-scraper"
TWEETS_PER_BRAND = 10
TIMEOUT_SECS = 180


# ═══════════════════════════════════════════════════════════════════════
# CRAWL: Apify Twitter Scraper
# ═══════════════════════════════════════════════════════════════════════

def crawl_brand(client, brand: dict) -> dict:
    """Crawl a single brand's X account via Apify."""
    handle = brand["handle"]
    result = {
        "brand": brand["name"],
        "handle": handle,
        "category": brand["category"],
        "crawled_at": datetime.now(JST).isoformat(),
        "tweets": [],
        "profile": {},
        "error": None,
    }

    try:
        run = client.actor(TWITTER_ACTOR).call(
            run_input={
                "twitterHandles": [handle],
                "maxItems": TWEETS_PER_BRAND,
                "addUserInfo": True,
                "proxyConfig": {"useApifyProxy": True},
            },
            timeout_secs=TIMEOUT_SECS,
        )
        items = list(client.dataset(run["defaultDatasetId"]).iterate_items())

        if not items:
            result["error"] = "No data returned"
            return result

        # Extract profile info from first item with user data
        for item in items:
            user = item.get("author", item.get("user", {}))
            if user and user.get("followers"):
                result["profile"] = {
                    "followers": user.get("followers", 0),
                    "following": user.get("following", 0),
                    "tweets_count": user.get("statusesCount", user.get("tweetsCount", 0)),
                    "display_name": user.get("name", user.get("displayName", "")),
                    "bio": (user.get("description", user.get("bio", "")) or "")[:200],
                    "verified": user.get("isVerified", user.get("verified", False)),
                }
                break

        # Extract tweets
        for item in items:
            tweet = {
                "text": (item.get("text", item.get("full_text", "")) or "")[:280],
                "created_at": item.get("createdAt", item.get("created_at", "")),
                "likes": item.get("likeCount", item.get("likes", 0)),
                "retweets": item.get("retweetCount", item.get("retweets", 0)),
                "replies": item.get("replyCount", item.get("replies", 0)),
                "views": item.get("viewCount", item.get("views", 0)),
                "hashtags": item.get("hashtags", []),
                "is_retweet": item.get("isRetweet", False),
            }
            result["tweets"].append(tweet)

        logger.info(f"  [OK] @{handle} ({brand['name']}): "
                     f"{len(result['tweets'])} tweets, "
                     f"{result['profile'].get('followers', 'N/A')} followers")

    except Exception as e:
        result["error"] = str(e)[:200]
        logger.error(f"  [ERR] @{handle}: {e}")

    return result


def crawl_all(brands: list[dict]) -> list[dict]:
    """Crawl all brands via Apify."""
    token = os.environ.get("APIFY_API_TOKEN")
    if not token:
        logger.error("APIFY_API_TOKEN not set. Set it in .env or environment.")
        return []

    try:
        from apify_client import ApifyClient
    except ImportError:
        logger.error("apify-client not installed. Run: pip install apify-client")
        return []

    client = ApifyClient(token)
    results = []

    for brand in brands:
        result = crawl_brand(client, brand)
        results.append(result)

    return results


# ═══════════════════════════════════════════════════════════════════════
# ANALYZE: Compute engagement metrics
# ═══════════════════════════════════════════════════════════════════════

def analyze_results(results: list[dict]) -> list[dict]:
    """Add computed metrics to each brand result."""
    for r in results:
        if r.get("error") or not r["tweets"]:
            r["metrics"] = {"status": "error", "detail": r.get("error", "no tweets")}
            continue

        tweets = r["tweets"]
        original = [t for t in tweets if not t.get("is_retweet")]

        total_likes = sum(t.get("likes", 0) for t in original) if original else 0
        total_rts = sum(t.get("retweets", 0) for t in original) if original else 0
        total_replies = sum(t.get("replies", 0) for t in original) if original else 0
        count = len(original) or 1

        # Collect all hashtags
        all_hashtags = []
        for t in tweets:
            all_hashtags.extend(t.get("hashtags", []))
        hashtag_freq = {}
        for h in all_hashtags:
            h_lower = h.lower()
            hashtag_freq[h_lower] = hashtag_freq.get(h_lower, 0) + 1
        top_hashtags = sorted(hashtag_freq.items(), key=lambda x: -x[1])[:5]

        r["metrics"] = {
            "total_tweets": len(tweets),
            "original_tweets": len(original),
            "avg_likes": round(total_likes / count, 1),
            "avg_retweets": round(total_rts / count, 1),
            "avg_replies": round(total_replies / count, 1),
            "top_hashtags": [{"tag": h, "count": c} for h, c in top_hashtags],
            "engagement_score": round((total_likes + total_rts * 2 + total_replies * 3) / count, 1),
        }

    return results


# ═══════════════════════════════════════════════════════════════════════
# OUTPUT: Save JSON + Excel
# ═══════════════════════════════════════════════════════════════════════

def save_json(results: list[dict], date_str: str) -> Path:
    """Save results to JSON."""
    path = TMP_DIR / f"apify_twitter_monitor_{date_str}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    logger.info(f"  Saved: {path}")
    return path


def save_excel(results: list[dict], date_str: str) -> Path | None:
    """Save results to Excel for Teams upload."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("  openpyxl not installed — skipping Excel output")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "競合モニタリング"

    # Header
    headers = [
        "カテゴリ", "ブランド", "ハンドル", "フォロワー",
        "ツイート数", "平均いいね", "平均RT", "平均返信",
        "エンゲージメントスコア", "人気ハッシュタグ", "最新ツイート",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data
    for row, r in enumerate(results, 2):
        metrics = r.get("metrics", {})
        profile = r.get("profile", {})
        latest = r["tweets"][0]["text"][:100] if r.get("tweets") else ""
        top_tags = ", ".join(
            f"#{h['tag']}" for h in metrics.get("top_hashtags", [])
        )

        ws.cell(row=row, column=1, value=r.get("category", ""))
        ws.cell(row=row, column=2, value=r.get("brand", ""))
        ws.cell(row=row, column=3, value=f"@{r.get('handle', '')}")
        ws.cell(row=row, column=4, value=profile.get("followers", "N/A"))
        ws.cell(row=row, column=5, value=metrics.get("total_tweets", 0))
        ws.cell(row=row, column=6, value=metrics.get("avg_likes", 0))
        ws.cell(row=row, column=7, value=metrics.get("avg_retweets", 0))
        ws.cell(row=row, column=8, value=metrics.get("avg_replies", 0))
        ws.cell(row=row, column=9, value=metrics.get("engagement_score", 0))
        ws.cell(row=row, column=10, value=top_tags)
        ws.cell(row=row, column=11, value=latest)

    path = TMP_DIR / f"apify_twitter_monitor_{date_str}.xlsx"
    wb.save(str(path))
    logger.info(f"  Saved: {path}")
    return path


# ═══════════════════════════════════════════════════════════════════════
# NOTIFY: Teams Webhook
# ═══════════════════════════════════════════════════════════════════════

def notify_teams(results: list[dict]):
    """Send summary to Teams webhook."""
    webhook = os.environ.get("TEAMS_WEBHOOK_URL")
    if not webhook:
        logger.info("  TEAMS_WEBHOOK_URL not set — skipping notification")
        return

    import requests

    # Build summary
    lines = ["## 🔍 Apify 競合Xモニタリングレポート\n"]
    ts = datetime.now(JST).strftime("%Y-%m-%d %H:%M JST")
    lines.append(f"**日時:** {ts}\n")

    for r in results:
        metrics = r.get("metrics", {})
        profile = r.get("profile", {})
        if r.get("error"):
            lines.append(f"- **{r['brand']}** (@{r['handle']}): ❌ エラー")
        else:
            followers = profile.get("followers", "N/A")
            eng = metrics.get("engagement_score", 0)
            lines.append(
                f"- **{r['brand']}** (@{r['handle']}): "
                f"{followers} followers, "
                f"Eng Score: {eng}"
            )

    body = {"text": "\n".join(lines)}

    try:
        resp = requests.post(webhook, json=body, timeout=10)
        if resp.status_code == 200:
            logger.info("  Teams notification sent")
        else:
            logger.warning(f"  Teams webhook returned {resp.status_code}")
    except Exception as e:
        logger.warning(f"  Teams notification failed: {e}")


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Apify Twitter Monitor — 競合Xアカウントモニタリング"
    )
    parser.add_argument("--dry-run", action="store_true",
                        help="クロールのみ（ファイル保存なし）")
    parser.add_argument("--brand", type=str, default=None,
                        help="特定ブランド名のみ調査 (例: ピジョン)")
    parser.add_argument("--json", action="store_true",
                        help="結果をJSON出力")
    parser.add_argument("--notify", action="store_true",
                        help="Teams Webhook通知")
    args = parser.parse_args()

    now = datetime.now(JST)
    date_str = now.strftime("%Y%m%d_%H%M")

    logger.info("=== Apify Twitter Monitor ===")
    logger.info(f"  Timestamp: {now.strftime('%Y-%m-%d %H:%M JST')}")
    logger.info(f"  Dry run: {args.dry_run}")

    # Select brands
    brands = COMPETITOR_BRANDS
    if args.brand:
        brands = [b for b in brands if args.brand in b["name"]]
        if not brands:
            logger.error(f"Brand '{args.brand}' not found")
            sys.exit(1)

    logger.info(f"  Brands: {len(brands)}")
    print()

    # Crawl
    logger.info("[1/3] Crawling X accounts via Apify...")
    results = crawl_all(brands)
    if not results:
        logger.error("No results. Check APIFY_API_TOKEN and apify-client.")
        sys.exit(1)

    # Analyze
    logger.info("[2/3] Analyzing engagement metrics...")
    results = analyze_results(results)

    success = sum(1 for r in results if not r.get("error"))
    errors = len(results) - success
    logger.info(f"  Success: {success}, Errors: {errors}")
    print()

    # Output
    if not args.dry_run:
        logger.info("[3/3] Saving results...")
        json_path = save_json(results, date_str)
        excel_path = save_excel(results, date_str)

        if args.notify:
            notify_teams(results)
    else:
        logger.info("[3/3] DRY RUN — skipping save")

    # Print summary or JSON
    if args.json:
        print(json.dumps(results, ensure_ascii=False, indent=2))
    else:
        print("\n=== Summary ===")
        for r in results:
            m = r.get("metrics", {})
            p = r.get("profile", {})
            status = "✓" if not r.get("error") else "✗"
            print(f"  {status} {r['brand']:12s} @{r['handle']:20s} "
                  f"Followers: {str(p.get('followers', 'N/A')):>8s}  "
                  f"EngScore: {m.get('engagement_score', 'N/A')}")
        print(f"\n  Total: {success} OK, {errors} errors")


if __name__ == "__main__":
    main()
