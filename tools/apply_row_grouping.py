"""
Apply row-based grouping to Search Volume and Campaign Details sheets.

Row outline strategy:
  Search Volume:
    - Section header + column header rows → level 0 (always visible)
    - Keyword data rows                  → level 1 (collapsible)

  Campaign Details:
    - Section header (#002060) + column header (#D9E2F3) → level 0
    - TOTAL (#4472C4) + Platform header (#F2F2F2)        → level 1
    - Brand / data rows (#DAEEF3, #F2DCDB, other)        → level 2

Column groupings are cleared on both sheets (already removed manually).
summaryBelow=False  → collapse button sits ABOVE the grouped rows (next to the header).
"""

import sys, os
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.worksheet.properties import Outline

from output_utils import get_latest_file, get_output_path
EXCEL_PATH = get_latest_file("polar", "financial_model")
if not EXCEL_PATH:
    raise FileNotFoundError("No financial_model file found in Data Storage/polar/")


# ── Color constants ────────────────────────────────────────────────────────────
DARK_BLUE    = "FF002060"   # section header  → level 0
COL_HEADER   = "FFD9E2F3"  # column header   → level 0
TOTAL_BLUE   = "FF4472C4"  # TOTAL row       → level 1
PLATFORM_GRY = "FFF2F2F2"  # platform header → level 1
DATA_BLUE    = "FFDAEEF3"  # data row        → level 2
DATA_PINK    = "FFF2DCDB"  # data row        → level 2

# Section headers in Search Volume (any colour not listed above)
GSC_HDR    = "FF375623"
GOOGLE_HDR = "FF2E75B6"
AMAZON_HDR = "FFC55A11"


def get_bg(ws, row: int, col: int = 1) -> str:
    cell = ws.cell(row=row, column=col)
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.type == "rgb":
        return fill.fgColor.rgb
    return "00000000"


def clear_col_groups(ws):
    for col_dim in ws.column_dimensions.values():
        col_dim.outline_level = 0


def set_summary_above(ws):
    """Place expand/collapse button ABOVE the group (summary row is the header)."""
    if ws.sheet_properties.outlinePr is None:
        ws.sheet_properties.outlinePr = Outline()
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False


# ── Search Volume ──────────────────────────────────────────────────────────────
def group_search_volume(ws):
    """
    Hardcoded section layout (16 keywords per section):
      Section 1 — GSC       : header=row 1, col_hdr=row 2,  data=rows  3-18
      Section 2 — Google Ads: header=row 20, col_hdr=row 21, data=rows 22-37
      Section 3 — Amazon    : header=row 39, col_hdr=row 40, data=rows 41-56
      Section 4 — Trends    : header=row 58, col_hdr=row 59, data=rows 60-75
    Separator rows (19, 38, 57) stay at level 0.
    """
    print("  [Search Volume] applying row grouping …")
    set_summary_above(ws)
    clear_col_groups(ws)

    # Reset all rows to level 0
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].outline_level = 0

    # Hardcoded data ranges for each section
    DATA_RANGES = [
        range(3, 19),   # GSC data
        range(22, 38),  # Google Ads data
        range(41, 57),  # Amazon data
        range(60, 76),  # Google Trends data
    ]
    for data_range in DATA_RANGES:
        for r in data_range:
            ws.row_dimensions[r].outline_level = 1

    grouped = sum(len(dr) for dr in DATA_RANGES)
    print(f"    → {grouped} data rows grouped at level 1 (4 sections × 16 keywords)")


# ── Campaign Details ───────────────────────────────────────────────────────────
def group_campaign_details(ws):
    print("  [Campaign Details] applying row grouping …")
    set_summary_above(ws)
    clear_col_groups(ws)

    # Reset all row outline levels first
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].outline_level = 0

    # Walk rows and assign levels
    in_section  = False
    in_platform = False
    plain_level = 0  # level for rows with no special color

    for r in range(1, ws.max_row + 1):
        bg = get_bg(ws, r)

        if bg == DARK_BLUE:
            ws.row_dimensions[r].outline_level = 0
            in_section  = True
            in_platform = False
            plain_level = 1   # next plain rows are inside section (level 1)

        elif bg == COL_HEADER:
            ws.row_dimensions[r].outline_level = 0
            # Column header marks section is open; data starts after this
            # (don't change in_platform)

        elif bg == TOTAL_BLUE:
            # TOTAL summary row → level 1 (visible when section shown)
            ws.row_dimensions[r].outline_level = 1

        elif bg == PLATFORM_GRY:
            ws.row_dimensions[r].outline_level = 1
            in_platform = True
            plain_level = 2   # rows after a platform header go deeper

        elif bg in (DATA_BLUE, DATA_PINK):
            ws.row_dimensions[r].outline_level = 2

        else:
            # No background (00000000) — use context
            ws.row_dimensions[r].outline_level = plain_level if in_section else 0

    # Report
    c = {0: 0, 1: 0, 2: 0}
    for r in range(1, ws.max_row + 1):
        lvl = ws.row_dimensions[r].outline_level if ws.row_dimensions.get(r) else 0
        c[lvl] = c.get(lvl, 0) + 1
    print(f"    → level 0: {c[0]} rows | level 1: {c[1]} rows | level 2: {c[2]} rows")


# ── Main ───────────────────────────────────────────────────────────────────────
def run():
    print(f"Loading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    if "Search Volume" in wb.sheetnames:
        group_search_volume(wb["Search Volume"])
    else:
        print("  ⚠ 'Search Volume' sheet not found")

    if "Campaign Details" in wb.sheetnames:
        group_campaign_details(wb["Campaign Details"])
    else:
        print("  ⚠ 'Campaign Details' sheet not found")

    # Save as new version
    out_path = get_output_path("polar", "financial_model")
    wb.save(out_path)
    print(f"\n✓ Saved: {out_path}")
    print("Done. Row grouping applied — column groupings cleared.")


if __name__ == "__main__":
    run()
