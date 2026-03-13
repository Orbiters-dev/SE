"""
run_kpi_monthly.py - Monthly KPI analysis -> Excel (kpis_model)

Sections:
  1) Discount rate by brand x channel (Shopify) — flat table with list price
  2) Ad spend by platform (Amazon Ads, Meta, Google)
  3) Influencer seeding cost (PayPal + COGS + shipping)

Data sources:
  - PG Data Keeper: shopify_orders_daily, amazon_ads_daily, meta_ads_daily, google_ads_daily
  - Shopify API: influencer/PR orders (fetch_influencer_orders.py)
  - PayPal API: influencer payments (fetch_paypal_transactions.py)
  - COGS by SKU.xlsx

Output:
  - Loads latest kpis_model_*.xlsx from Data Storage/kpi_reports/
  - Adds/replaces 3 tabs: KPI_할인율 / KPI_광고비 / KPI_시딩비용
  - Saves as next version (v+1)

Usage:
    python tools/run_kpi_monthly.py
    python tools/run_kpi_monthly.py --no-sheet   # console only
    python tools/run_kpi_monthly.py --from 2024-01 --to 2026-02
"""

import os
import sys
import json
from collections import defaultdict
from pathlib import Path

TOOLS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(TOOLS_DIR))
from env_loader import load_env
load_env()

ROOT = TOOLS_DIR.parent
SEEDING_CACHE = ROOT / ".tmp" / "polar_data"  # fetch_influencer_orders.py / fetch_paypal_transactions.py output
COGS_PATH = ROOT.parent / "Shared" / "NoPolar KPIs" / "Data config sheet" / "COGS by SKU.xlsx"
# ORBI KPIs output (WJ Test1 Data Storage)
OUTPUT_DIR = ROOT / "Data Storage" / "kpi_reports"


# ── Data Keeper ───────────────────────────────────────────────────────────────

def load_dk(table, days=800):
    sys.path.insert(0, str(TOOLS_DIR))
    from data_keeper_client import DataKeeper
    dk = DataKeeper(prefer_cache=False)
    return dk.get(table, days=days)


def compute_through_date():
    """Consistent through-date = min of latest full day across all main channels (PST)."""
    from datetime import datetime, timedelta, timezone
    PST = timezone(timedelta(hours=-8))
    yesterday_pst = (datetime.now(PST).date() - timedelta(days=1)).isoformat()

    main_tables = ["shopify_orders_daily", "amazon_ads_daily", "meta_ads_daily", "google_ads_daily"]
    latest = []
    for t in main_tables:
        rows = load_dk(t, days=30)
        dates = [r.get("date", "") for r in rows if r.get("date")]
        if dates:
            latest.append(max(dates))

    # min across channels = date ALL channels have data for; cap at yesterday PST
    through = min(latest) if latest else yesterday_pst
    return min(through, yesterday_pst)


# ── COGS ─────────────────────────────────────────────────────────────────────

def load_cogs_map():
    try:
        import openpyxl
        wb = openpyxl.load_workbook(COGS_PATH, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        sku_col = next((i for i, h in enumerate(headers) if "sku" in h), None)
        cost_col = next((i for i, h in enumerate(headers) if "cost" in h and "type" not in h), None)
        if sku_col is None or cost_col is None:
            return {}
        out = {}
        for row in rows[1:]:
            sku = str(row[sku_col] or "").strip().lower()
            try:
                cost = float(row[cost_col] or 0)
            except (TypeError, ValueError):
                continue
            if sku and cost > 0:
                out[sku] = cost
        wb.close()
        return out
    except Exception as e:
        print(f"  [WARN] COGS load failed: {e}")
        return {}


# ── 0. DATA STATUS TAB ────────────────────────────────────────────────────────

DATA_SOURCES = [
    # (display_name, table_name, date_field, spend_or_revenue_field)
    ("Shopify Orders",     "shopify_orders_daily", "date", "net_sales"),
    ("Amazon Sales (SP-API)", "amazon_sales_daily", "date", "net_sales"),
    ("Amazon Ads",         "amazon_ads_daily",     "date", "spend"),
    ("Meta Ads",           "meta_ads_daily",       "date", "spend"),
    ("Google Ads",         "google_ads_daily",     "date", "spend"),
    ("GA4",                "ga4_daily",            "date", "sessions"),
    ("Klaviyo",            "klaviyo_daily",        "date", "revenue"),
]


def add_data_status_tab(wb, through_date):
    """Add a 'Data Status' tab at position 0 showing data freshness per platform."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from datetime import datetime, timezone, timedelta

    tab_name = "Data Status"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(title=tab_name, index=0)

    PST = timezone(timedelta(hours=-8))
    now_pst = datetime.now(PST)
    generated = now_pst.strftime("%Y-%m-%d %H:%M PST")

    # Styles
    FILL_HEADER = PatternFill("solid", fgColor="002060")
    FILL_OK     = PatternFill("solid", fgColor="C6EFCE")
    FILL_WARN   = PatternFill("solid", fgColor="FFF2CC")
    FILL_FAIL   = PatternFill("solid", fgColor="FFC7CE")
    FILL_TITLE  = PatternFill("solid", fgColor="002060")
    FONT_WHITE  = Font(bold=True, color="FFFFFF", size=11)
    FONT_TITLE  = Font(bold=True, color="FFFFFF", size=14)
    FONT_BOLD   = Font(bold=True, size=11)
    FONT_NORMAL = Font(size=11)
    FONT_SMALL  = Font(size=9, color="808080")
    ALIGN_CTR   = Alignment(horizontal="center", vertical="center")
    ALIGN_LEFT  = Alignment(horizontal="left", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "ORBI KPI -- Data Source Status"
    title_cell.fill = FILL_TITLE
    title_cell.font = FONT_TITLE
    title_cell.alignment = ALIGN_CTR
    for col in range(1, 8):
        ws.cell(row=1, column=col).fill = FILL_TITLE

    # Subtitle row
    ws.merge_cells("A2:G2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated: {generated}  |  Through-date: {through_date}"
    sub_cell.font = FONT_SMALL
    sub_cell.alignment = ALIGN_CTR

    # Header row
    headers = ["Platform", "Table", "Rows", "Date From", "Date To", "Days Gap", "Status"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c_idx, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_WHITE
        cell.alignment = ALIGN_CTR

    # Data rows
    row_idx = 5
    for display_name, table, date_field, _ in DATA_SOURCES:
        try:
            rows = load_dk(table, days=800)
        except Exception:
            rows = []

        total_rows = len(rows)
        dates = [r.get(date_field, "") for r in rows if r.get(date_field)]
        if dates:
            date_from = min(dates)
            date_to = max(dates)
            # Days gap from through_date
            from datetime import date as dt_date
            try:
                latest_dt = dt_date.fromisoformat(date_to)
                through_dt = dt_date.fromisoformat(through_date)
                gap = (through_dt - latest_dt).days
            except (ValueError, TypeError):
                gap = None
        else:
            date_from = "-"
            date_to = "-"
            gap = None

        # Status logic
        if total_rows == 0 or date_to == "-":
            status = "NO DATA"
            fill = FILL_FAIL
        elif gap is not None and gap <= 1:
            status = "OK"
            fill = FILL_OK
        elif gap is not None and gap <= 3:
            status = "STALE"
            fill = FILL_WARN
        else:
            status = "STALE"
            fill = FILL_FAIL

        vals = [display_name, table, total_rows, date_from, date_to,
                f"{gap}d" if gap is not None else "-", status]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=c_idx, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CTR if c_idx >= 3 else ALIGN_LEFT
            cell.border = THIN_BORDER
            if c_idx == 7:  # Status column
                cell.fill = fill
                cell.font = FONT_BOLD
        row_idx += 1

    # Summary row
    row_idx += 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
    note = ws.cell(row=row_idx, column=1)
    note.value = (
        "Through-date = MIN(latest date across Shopify, Amazon Ads, Meta Ads, Google Ads), "
        "capped at yesterday PST. All KPI tabs use this date as cutoff."
    )
    note.font = FONT_SMALL
    note.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 11
    ws.column_dimensions["G"].width = 12

    ws.sheet_properties.tabColor = "002060"
    print(f"  [Data Status] tab added (through_date={through_date})")


# ── 1. DISCOUNT RATE (Golmani wide format, brand x channel, PR excluded) ──────

CHANNEL_ORDER = ["ONZ", "Amazon", "B2B", "TikTok", "Unknown"]  # PR excluded; D2C in PG = ONZ
D2C_RAW_CHANNELS = {"D2C", "Amazon", "TikTok"}  # user's "D2C" = ONZ+Amazon+TikTok
BRAND_ORDER   = ["Grosmimi", "Naeiae", "CHA&MOM", "Onzenna", "Alpremio", "Unknown"]

# Avg COGS per unit by brand (from COGS by SKU.xlsx x Product Variant)
AVG_COGS = {
    "Grosmimi": 8.41, "Naeiae": 5.35, "CHA&MOM": 7.53,
    "Onzenna": 5.35,  "Alpremio": 12.57, "Unknown": 8.00,
}

# Avg selling price per unit by brand (used to estimate units when DB has units=0 from backfill)
AVG_PRICE = {
    "Grosmimi": 28.0, "Naeiae": 18.0, "CHA&MOM": 32.0,
    "Onzenna": 22.0,  "Alpremio": 38.0, "Unknown": 25.0,
}

def load_brand_avg_cogs():
    """Recompute avg COGS per brand from files (fallback: AVG_COGS constant)."""
    try:
        import openpyxl
        sku_brand = {}
        wb = openpyxl.load_workbook(str(COGS_PATH.parent / "Product Variant by SKU.xlsx"),
                                     read_only=True, data_only=True)
        for r in list(wb.active.iter_rows(values_only=True))[1:]:
            sku, brand = str(r[0] or "").strip().lower(), str(r[3] or "").strip()
            if sku and brand:
                sku_brand[sku] = brand
        wb.close()

        wb2 = openpyxl.load_workbook(str(COGS_PATH), read_only=True, data_only=True)
        brand_costs = defaultdict(list)
        for r in list(wb2.active.iter_rows(values_only=True))[1:]:
            sku = str(r[2] or "").strip().lower()
            cost = float(r[4] or 0)
            brand = sku_brand.get(sku)
            if brand and cost > 0:
                brand_costs[brand].append(cost)
        wb2.close()
        return {b: sum(v)/len(v) for b, v in brand_costs.items() if v}
    except Exception as e:
        print(f"  [WARN] COGS map load failed: {e}, using constants")
        return AVG_COGS


def analyze_discounts(date_from, date_to, through_date=None):
    """Wide format (Golmani-style):
       Rows: Section header + TOTAL + Brand subtotal + Channel detail
       Cols: [Label, Brand, Channel] + [Jan 2024, Feb 2024, ..., YTD]
       Sections: Gross Sales / Net Sales / Discounts / Discount % /
                 Units / Avg List Price / COGS (est.) / GM / GM%
       PR channel excluded (goes to seeding cost section).
       through_date: only include daily data up to this date (consistent ceiling).
    """
    import openpyxl
    brand_cogs = load_brand_avg_cogs()

    # ── collect PG data ──────────────────────────────────────────────────────
    rows = load_dk("shopify_orders_daily")
    latest_date = "0000-00-00"
    agg = defaultdict(lambda: {"gross":0.0,"disc":0.0,"net":0.0,"orders":0,"units":0})

    for r in rows:
        raw_date = r.get("date") or ""
        if through_date and raw_date > through_date:
            continue  # skip data beyond consistent through_date
        if raw_date > latest_date:
            latest_date = raw_date
        date = raw_date[:7]
        channel = r.get("channel") or "Unknown"
        if channel == "PR":          # PR → seeding, exclude here
            continue
        if not date or date < date_from or date > date_to:
            continue
        brand = r.get("brand") or "Unknown"
        key = (date, brand, channel)
        agg[key]["gross"]  += float(r.get("gross_sales") or 0)
        agg[key]["disc"]   += float(r.get("discounts")   or 0)
        agg[key]["net"]    += float(r.get("net_sales")    or 0)
        agg[key]["orders"] += int(r.get("orders") or 0)
        agg[key]["units"]  += int(r.get("units")  or 0)

    # ── month list ───────────────────────────────────────────────────────────
    months = sorted(set(k[0] for k in agg))
    # Use passed-in through_date (consistent across sections), or compute fallback
    from datetime import datetime as _dtt, timedelta as _td, timezone as _tz
    PST = _tz(_td(hours=-8))
    today_pst = _dtt.now(PST).date()
    if not through_date:
        yesterday_pst = (today_pst - _td(days=1)).isoformat()
        through_date  = min(latest_date, yesterday_pst) if latest_date > "0000-00-00" else yesterday_pst
    partial_note = f"(through {through_date})"

    # Month labels: current month gets partial note
    def month_label(m):
        import calendar
        y, mo = int(m[:4]), int(m[5:7])
        if y == today_pst.year and mo == today_pst.month:
            return f"{calendar.month_abbr[mo]} {y}\n{partial_note}"
        return f"{calendar.month_abbr[mo]} {y}"

    col_labels = [month_label(m) for m in months] + ["YTD"]

    # ── helper: get metric dict for a set of keys ────────────────────────────
    def get_vals(keys):
        v = {"gross":0.0,"disc":0.0,"net":0.0,"units":0}
        for k in keys:
            if k in agg:
                for f in v:
                    v[f] += agg[k][f]
        return v

    def ytd_keys(base_keys):
        return [k for k in base_keys if k[0] >= f"{months[-1][:4]}-01"]

    # ── build metric rows per entity ─────────────────────────────────────────
    brands_present = sorted(
        set(k[1] for k in agg if agg[k]["gross"] > 50),
        key=lambda b: BRAND_ORDER.index(b) if b in BRAND_ORDER else 99
    )

    def entity_rows():
        """Yield (label, brand, channel, {month: vals}, ytd_vals) tuples."""
        # TOTAL
        all_keys = list(agg.keys())
        monthly = {}
        for m in months:
            monthly[m] = get_vals([k for k in all_keys if k[0] == m])
        ytd = get_vals([k for k in all_keys if k[0] >= f"{months[-1][:4]}-01"])
        yield ("TOTAL", "", "", monthly, ytd)

        for brand in brands_present:
            brand_keys = [k for k in agg if k[1] == brand]
            monthly_b = {}
            for m in months:
                monthly_b[m] = get_vals([k for k in brand_keys if k[0] == m])
            ytd_b = get_vals([k for k in brand_keys if k[0] >= f"{months[-1][:4]}-01"])
            yield (brand, brand, "", monthly_b, ytd_b)

            def ch_display(c):
                return "ONZ" if c == "D2C" else c

            channels = sorted(
                set(k[2] for k in brand_keys if agg[k]["gross"] > 10),
                key=lambda c: CHANNEL_ORDER.index(ch_display(c)) if ch_display(c) in CHANNEL_ORDER else 99
            )
            for ch in channels:
                ch_keys = [k for k in brand_keys if k[2] == ch]
                monthly_c = {}
                for m in months:
                    monthly_c[m] = get_vals([k for k in ch_keys if k[0] == m])
                ytd_c = get_vals([k for k in ch_keys if k[0] >= f"{months[-1][:4]}-01"])
                yield (f"  {ch_display(ch)}", brand, ch, monthly_c, ytd_c)

    entities = list(entity_rows())

    # ── console summary ──────────────────────────────────────────────────────
    print("=" * 80)
    print(f"1) DISCOUNT RATE BY BRAND x CHANNEL  (data through {latest_date}, PR excluded)")
    print("=" * 80)
    for label, brand, ch, monthly, ytd in entities:
        if ch == "" and brand != "":  # brand subtotals only
            g, d = ytd["gross"], ytd["disc"]
            print(f"  {brand:<12} gross=${g:>10,.0f}  disc=${d:>8,.0f}  rate={d/g*100:.1f}%" if g else "")

    # ── build wide sheet ─────────────────────────────────────────────────────
    METRIC_SECTIONS = [
        ("GROSS SALES ($)",        "gross",  lambda v: round(v["gross"],2)),
        ("NET SALES ($)",          "net",    lambda v: round(v["net"],2)),
        ("DISCOUNTS ($)",          "disc",   lambda v: round(v["disc"],2)),
        ("DISCOUNT RATE",          "rate",   lambda v: round(v["disc"]/v["gross"],4) if v["gross"] else 0),
        ("UNITS",                  "units",  lambda v: v["units"]),
        ("AVG LIST PRICE ($/unit)","price",  lambda v: round(v["gross"]/v["units"],2) if v["units"] else 0),
        ("COGS est. ($)",          "cogs",   None),   # special: units × avg_cogs[brand]
        ("GM ($)",                 "gm",     None),   # net - cogs
        ("GM %",                   "gm_pct", None),   # gm / net
    ]

    out = []  # list of rows

    # Title
    out.append([f"DISCOUNT ANALYSIS — BRAND x CHANNEL  |  {partial_note}  |  PR excluded (see Seeding tab)"])
    out.append([])

    # Header row
    hdr = ["Metric", "Brand", "Channel"] + col_labels
    out.append(hdr)

    # ── build total_monthly and d2c_monthly for Executive Summary / Summary tab ─
    brand_month = defaultdict(lambda: {"gross":0.0,"net":0.0,"disc":0.0,"units":0,"orders":0})
    for (date, brand, ch), v in agg.items():
        k = (date, brand)
        for f in ("gross","net","disc","units","orders"):
            brand_month[k][f] += v[f]

    total_monthly = {}
    d2c_monthly   = {}
    for m in months:
        tm = {"gross":0.0,"net":0.0,"disc":0.0,"units":0,"orders":0,"cogs":0.0}
        dm = {"gross":0.0,"net":0.0,"disc":0.0,"units":0,"orders":0,"cogs":0.0}
        for brand in brands_present:
            bp = AVG_PRICE.get(brand, AVG_PRICE["Unknown"])
            bc = brand_cogs.get(brand, AVG_COGS.get(brand, 8.0))
            # total (all channels)
            bv = brand_month.get((m, brand), {})
            u = bv.get("units",0) or (bv.get("gross",0) / bp if bv.get("gross",0) else 0)
            tm["gross"] += bv.get("gross",0);  tm["net"] += bv.get("net",0)
            tm["disc"]  += bv.get("disc",0);   tm["units"] += bv.get("units",0)
            tm["orders"] += bv.get("orders",0); tm["cogs"]  += u * bc
            # D2C = ONZ(raw:D2C) + Amazon + TikTok
            for ch in D2C_RAW_CHANNELS:
                dv = agg.get((m, brand, ch), {})
                if not dv:
                    continue
                du = dv.get("units",0) or (dv.get("gross",0) / bp if dv.get("gross",0) else 0)
                dm["gross"] += dv.get("gross",0);  dm["net"] += dv.get("net",0)
                dm["disc"]  += dv.get("disc",0);   dm["units"] += dv.get("units",0)
                dm["orders"] += dv.get("orders",0); dm["cogs"]  += du * bc
        total_monthly[m] = tm
        d2c_monthly[m]   = dm

    for section_name, _, getter in METRIC_SECTIONS:
        # Section title row
        out.append([section_name])

        for label, brand, ch, monthly, ytd in entities:
            avg_cogs_brand  = brand_cogs.get(brand, AVG_COGS.get(brand, 8.0))
            avg_price_brand = AVG_PRICE.get(brand, AVG_PRICE["Unknown"])

            def est_units(v):
                """Units from DB; if 0 (backfill data), estimate from gross_sales / avg_price."""
                if v["units"] > 0:
                    return v["units"]
                if v["gross"] > 0 and avg_price_brand > 0:
                    return v["gross"] / avg_price_brand
                return 0

            row = [label, brand if ch else "", ch]
            ytd_val = None

            for m in months:
                v = monthly[m]
                if getter is None:
                    u = est_units(v)
                    if section_name == "COGS est. ($)":
                        val = round(u * avg_cogs_brand, 2)
                    elif section_name == "GM ($)":
                        val = round(v["net"] - u * avg_cogs_brand, 2)
                    else:  # GM %
                        cogs = u * avg_cogs_brand
                        gm   = v["net"] - cogs
                        val  = round(gm / v["net"], 4) if v["net"] else 0
                else:
                    val = getter(v)
                row.append(val)

            # YTD
            vy = ytd
            if getter is None:
                u = est_units(vy)
                if section_name == "COGS est. ($)":
                    ytd_val = round(u * avg_cogs_brand, 2)
                elif section_name == "GM ($)":
                    ytd_val = round(vy["net"] - u * avg_cogs_brand, 2)
                else:
                    cogs = u * avg_cogs_brand
                    gm   = vy["net"] - cogs
                    ytd_val = round(gm / vy["net"], 4) if vy["net"] else 0
            else:
                ytd_val = getter(vy)
            row.append(ytd_val)

            out.append(row)

    # Channel-level monthly discounts for MKT Spend section
    channel_disc_monthly = defaultdict(lambda: defaultdict(float))
    for (date, brand, channel), v in agg.items():
        ch = "ONZ" if channel == "D2C" else channel
        channel_disc_monthly[ch][date] += v["disc"]

    print(f"  Rows in tab: {len(out)}")
    return out, total_monthly, d2c_monthly, channel_disc_monthly


# ── 2. AD SPEND ───────────────────────────────────────────────────────────────

def analyze_ad_spend(date_from, date_to, through_date=None):
    """Wide format (months as columns), Amazon Ads split by brand."""
    import calendar as _cal
    from datetime import datetime as _dtt2, timedelta as _td2, timezone as _tz2

    # Amazon Ads: per brand per month
    amz_brand = defaultdict(lambda: defaultdict(float))
    for r in load_dk("amazon_ads_daily"):
        d = r.get("date") or ""
        if through_date and d > through_date:
            continue
        date = d[:7]
        if date_from <= date <= date_to:
            brand = r.get("brand") or "Unknown"
            amz_brand[brand][date] += float(r.get("spend") or 0)

    meta_monthly = defaultdict(float)
    for r in load_dk("meta_ads_daily"):
        d = r.get("date") or ""
        if through_date and d > through_date:
            continue
        date = d[:7]
        if date_from <= date <= date_to:
            meta_monthly[date] += float(r.get("spend") or 0)

    google_monthly = defaultdict(float)
    for r in load_dk("google_ads_daily"):
        d = r.get("date") or ""
        if through_date and d > through_date:
            continue
        date = d[:7]
        if date_from <= date <= date_to:
            google_monthly[date] += float(r.get("spend") or 0)

    # All months across all platforms
    all_months_set = set(meta_monthly) | set(google_monthly)
    for bd in amz_brand.values():
        all_months_set |= set(bd)
    all_months = sorted(m for m in all_months_set if date_from <= m <= date_to)

    _today_pst = _dtt2.now(_tz2(_td2(hours=-8))).date()
    _partial_note = f"(through {through_date})" if through_date else ""
    def _mlabel(m):
        y, mo = int(m[:4]), int(m[5:7])
        if y == _today_pst.year and mo == _today_pst.month and _partial_note:
            return f"{_cal.month_abbr[mo]} {y}\n{_partial_note}"
        return f"{_cal.month_abbr[mo]} {y}"

    ytd_year = all_months[-1][:4] if all_months else "2026"

    def _ytd(monthly_dict, brand_dict=None):
        if brand_dict is not None:
            return sum(v for m, v in brand_dict.items() if m >= f"{ytd_year}-01")
        return sum(v for m, v in monthly_dict.items() if m >= f"{ytd_year}-01")

    # ── Console ──────────────────────────────────────────────────────────────
    print("\n" + "=" * 80)
    print("2) AD SPEND BY PLATFORM (Monthly)")
    print("=" * 80)
    print(f"\n{'Month':<10}  {'Amazon Ads':>12}  {'Meta Ads':>12}  {'Google Ads':>12}  {'TOTAL':>12}")
    print("-" * 65)
    grand_by_month = defaultdict(float)
    for month in all_months:
        amz = sum(bd.get(month, 0) for bd in amz_brand.values())
        meta = meta_monthly.get(month, 0)
        goog = google_monthly.get(month, 0)
        total = amz + meta + goog
        grand_by_month[month] = total
        print(f"{month:<10}  ${amz:>10,.0f}  ${meta:>10,.0f}  ${goog:>10,.0f}  ${total:>10,.0f}")
    print("-" * 65)
    gt_amz  = sum(sum(bd.values()) for bd in amz_brand.values())
    gt_meta = sum(meta_monthly.values())
    gt_goog = sum(google_monthly.values())
    print(f"{'TOTAL':<10}  ${gt_amz:>10,.0f}  ${gt_meta:>10,.0f}  ${gt_goog:>10,.0f}  ${gt_amz+gt_meta+gt_goog:>10,.0f}")

    # ── Data availability windows (for n.m marking) ──────────────────────────
    NM = "n.m"  # sentinel for "not measured / no data collected"

    # Detect partial first month: if the earliest daily date doesn't start on the 1st,
    # that month is partial and should be n.m.
    amz_daily_dates = sorted(set(r.get("date","") for r in load_dk("amazon_ads_daily")
                                 if r.get("date") and (not through_date or r["date"] <= through_date)))
    meta_daily_dates = sorted(set(r.get("date","") for r in load_dk("meta_ads_daily")
                                  if r.get("date") and (not through_date or r["date"] <= through_date)))
    goog_daily_dates = sorted(set(r.get("date","") for r in load_dk("google_ads_daily")
                                  if r.get("date") and (not through_date or r["date"] <= through_date)))

    def _first_full_month(daily_dates):
        """Return the first YYYY-MM where the data starts on day 01 (full month)."""
        if not daily_dates:
            return "9999-99"
        first_date = daily_dates[0]
        first_month = first_date[:7]
        if first_date.endswith("-01"):
            return first_month
        # Partial first month — next month is the first full month
        y, m = int(first_month[:4]), int(first_month[5:7])
        m += 1
        if m > 12:
            m = 1; y += 1
        return f"{y}-{m:02d}"

    amz_data_start  = _first_full_month(amz_daily_dates)
    meta_data_start = _first_full_month(meta_daily_dates)
    goog_data_start = _first_full_month(goog_daily_dates)
    print(f"  Ad data start (first full month): Amazon={amz_data_start}, Meta={meta_data_start}, Google={goog_data_start}")

    def amz_val(month, brand_dict):
        if month < amz_data_start:
            return NM
        return round(brand_dict.get(month, 0), 2)

    def meta_val(month):
        if month < meta_data_start:
            return NM
        return round(meta_monthly.get(month, 0), 2)

    def goog_val(month):
        if month < goog_data_start:
            return NM
        return round(google_monthly.get(month, 0), 2)

    # ── Wide-format sheet rows ────────────────────────────────────────────────
    hdr = ["Platform / Brand"] + [_mlabel(m) for m in all_months] + ["YTD"]
    sheet_rows = [hdr]

    # Amazon Ads section
    sheet_rows.append(["Amazon Ads"])
    amz_brands_sorted = sorted(
        amz_brand.keys(),
        key=lambda b: BRAND_ORDER.index(b) if b in BRAND_ORDER else 99
    )
    amz_month_totals = defaultdict(float)  # only sums real values
    for brand in amz_brands_sorted:
        bd = amz_brand[brand]
        ytd_sum = _ytd(None, bd)
        row = [f"  {brand}"]
        for m in all_months:
            v = amz_val(m, bd)
            if v != NM:
                amz_month_totals[m] += v
            row.append(v)
        row.append(round(ytd_sum, 2))
        sheet_rows.append(row)
    # Amazon subtotal: "n.m" for months before data start
    amz_ytd = sum(v for m, v in amz_month_totals.items() if m >= f"{ytd_year}-01")
    amz_sub = [NM if m < amz_data_start else round(amz_month_totals.get(m, 0), 2)
               for m in all_months]
    sheet_rows.append(["  TOTAL Amazon"] + amz_sub + [round(amz_ytd, 2)])

    # Meta Ads
    sheet_rows.append(["Meta Ads"])
    meta_ytd = _ytd(meta_monthly)
    sheet_rows.append(["  Meta"] + [meta_val(m) for m in all_months] + [round(meta_ytd, 2)])

    # Google Ads
    sheet_rows.append(["Google Ads"])
    goog_ytd = _ytd(google_monthly)
    sheet_rows.append(["  Google"] + [goog_val(m) for m in all_months] + [round(goog_ytd, 2)])

    # Grand total: sum real values only, n.m if ALL platforms are n.m
    grand_ytd = amz_ytd + meta_ytd + goog_ytd
    total_row = ["TOTAL"]
    for m in all_months:
        av = amz_month_totals.get(m, NM if m < amz_data_start else 0)
        mv = meta_monthly.get(m, 0) if m >= meta_data_start else NM
        gv = google_monthly.get(m, 0) if m >= goog_data_start else NM
        nums = [x for x in [av, mv, gv] if x != NM]
        total_row.append(NM if not nums else round(sum(nums), 2))
    total_row.append(round(grand_ytd, 2))
    sheet_rows.append(total_row)

    return sheet_rows


# ── 3. SEEDING COST ───────────────────────────────────────────────────────────

def _ensure_seeding_data():
    """Fetch influencer orders & PayPal data if cache is missing or stale (>24h)."""
    import importlib, time as _t
    for script, outfile in [
        ("fetch_influencer_orders", "q10_influencer_orders.json"),
        ("fetch_paypal_transactions", "q11_paypal_transactions.json"),
    ]:
        path = SEEDING_CACHE / outfile
        if path.exists():
            age_h = (_t.time() - path.stat().st_mtime) / 3600
            if age_h < 24:
                continue
            print(f"  {outfile} is {age_h:.0f}h old, refreshing...")
        else:
            print(f"  {outfile} not found, fetching...")
        try:
            mod = importlib.import_module(script)
            mod.main()
        except Exception as e:
            print(f"  [WARN] {script} fetch failed: {e}")


def analyze_seeding_cost(date_from, date_to):
    cogs_map = load_cogs_map()
    print(f"\n  COGS map: {len(cogs_map)} SKUs")

    _ensure_seeding_data()

    # PayPal transactions (direct API via fetch_paypal_transactions.py)
    paypal_monthly = defaultdict(float)
    try:
        d11 = json.loads((SEEDING_CACHE / "q11_paypal_transactions.json").read_text(encoding="utf-8"))
        txns = d11.get("transactions", d11 if isinstance(d11, list) else [])
        for t in txns:
            amt = float(t.get("amount", 0) or 0)
            if amt >= 0:
                continue
            m = (t.get("date") or "")[:7]
            if date_from <= m <= date_to:
                paypal_monthly[m] += abs(amt)
    except Exception as e:
        print(f"  [WARN] PayPal load failed: {e}")

    # PR / influencer orders (direct Shopify API via fetch_influencer_orders.py)
    sample_cogs = defaultdict(float)
    shipping    = defaultdict(float)
    units       = defaultdict(int)
    unmatched   = set()
    try:
        d10 = json.loads((SEEDING_CACHE / "q10_influencer_orders.json").read_text(encoding="utf-8"))
        orders = d10 if isinstance(d10, list) else d10.get("orders", [])
        for order in orders:
            m = (order.get("created_at") or "")[:7]
            if not m or m < date_from or m > date_to:
                continue
            for li in order.get("line_items", []):
                sku = (li.get("sku") or "").strip().lower()
                qty = int(li.get("quantity", 1) or 1)
                cost = cogs_map.get(sku, 0)
                if cost > 0:
                    sample_cogs[m] += cost * qty
                elif sku:
                    unmatched.add(sku)
                shipping[m] += 10.0 * qty
                units[m]    += qty
        print(f"  PR orders loaded, unmatched SKUs: {len(unmatched)}")
    except Exception as e:
        print(f"  [WARN] PR orders load failed: {e}")

    all_months = sorted(set(list(paypal_monthly) + list(sample_cogs) + list(shipping)))
    all_months = [m for m in all_months if date_from <= m <= date_to]

    print("\n" + "=" * 80)
    print("3) INFLUENCER SEEDING COST (Monthly)")
    print("=" * 80)
    print(f"\n{'Month':<10}  {'PayPal':>13}  {'Sample COGS':>13}  {'Shipping':>13}  {'TOTAL':>13}  {'Units':>8}")
    print("-" * 75)

    totals = defaultdict(float)
    total_units = 0
    for m in all_months:
        pp   = paypal_monthly.get(m, 0)
        cogs = sample_cogs.get(m, 0)
        ship = shipping.get(m, 0)
        tot  = pp + cogs + ship
        u    = units.get(m, 0)
        totals["pp"] += pp; totals["cogs"] += cogs
        totals["ship"] += ship; totals["total"] += tot
        total_units += u
        print(f"{m:<10}  ${pp:>12,.0f}  ${cogs:>12,.0f}  ${ship:>12,.0f}  ${tot:>12,.0f}  {u:>8,}")

    print("-" * 75)
    print(f"{'TOTAL':<10}  ${totals['pp']:>12,.0f}  ${totals['cogs']:>12,.0f}  ${totals['ship']:>12,.0f}  ${totals['total']:>12,.0f}  {total_units:>8,}")

    # ── Wide-format sheet rows (months as columns) ────────────────────────────
    import calendar as _cal2
    from datetime import datetime as _dtt3, timedelta as _td3, timezone as _tz3
    _today3 = _dtt3.now(_tz3(_td3(hours=-8))).date()
    _pnote  = f"(through {through_date})" if hasattr(analyze_seeding_cost, '_through') else ""

    def _slabel(m):
        y, mo = int(m[:4]), int(m[5:7])
        return f"{_cal2.month_abbr[mo]} {y}"

    ytd_year = all_months[-1][:4] if all_months else "2026"
    hdr = ["Item"] + [_slabel(m) for m in all_months] + ["YTD"]

    def _row(label, monthly_dict, fmt="$"):
        ytd = sum(v for m, v in monthly_dict.items() if m >= f"{ytd_year}-01")
        return [label] + [round(monthly_dict.get(m, 0), 2) for m in all_months] + [round(ytd, 2)]

    tot_monthly = {m: paypal_monthly.get(m,0)+sample_cogs.get(m,0)+shipping.get(m,0) for m in all_months}
    ytd_units = sum(v for m, v in units.items() if m >= f"{ytd_year}-01")

    sheet_rows = [
        hdr,
        _row("PayPal", paypal_monthly),
        _row("Sample COGS", sample_cogs),
        _row("Shipping", shipping),
        ["TOTAL"] + [round(tot_monthly.get(m,0), 2) for m in all_months] + [round(sum(v for m, v in tot_monthly.items() if m >= f"{ytd_year}-01"), 2)],
        ["Units"] + [units.get(m, 0) for m in all_months] + [ytd_units],
    ]

    return sheet_rows


# ── Excel Output ─────────────────────────────────────────────────────────────

def find_latest_model():
    """Find latest kpis_model_*.xlsx by version number (v1, v2, ... v10, v11...)."""
    import re
    files = list(OUTPUT_DIR.glob("kpis_model_*.xlsx"))
    files = [f for f in files if not f.name.startswith("~")]  # skip lock files
    if not files:
        raise FileNotFoundError(f"No kpis_model_*.xlsx in {OUTPUT_DIR}")
    def _ver(p):
        m = re.search(r'_v(\d+)\.xlsx$', p.name)
        return int(m.group(1)) if m else 0
    return max(files, key=_ver)


def next_version_path(src: Path) -> Path:
    """financial_model_2026-03-06_v1.xlsx → _v2.xlsx"""
    import re
    stem = src.stem  # financial_model_2026-03-06_v1
    m = re.match(r"(.+_v)(\d+)$", stem)
    if m:
        return src.parent / f"{m.group(1)}{int(m.group(2))+1}.xlsx"
    return src.parent / f"{stem}_v2.xlsx"


def style_header(ws, n_cols):
    from openpyxl.styles import PatternFill, Font, Alignment
    fill = PatternFill("solid", fgColor="002060")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def write_wide_tab(wb, tab_name, rows):
    """Wide-format tab: 1 label column + month columns.
    Row types detected by content:
      - Section header: row has only 1 cell (len==1) or all others empty → grey
      - TOTAL row: first cell == 'TOTAL' or 'TOTAL ...' → yellow
      - Indent row: first cell starts with '  ' → italic (channel-level)
      - Header row: row[0] index == 0 → dark blue
    Numbers formatted from column 2 onward.
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    FILL_HEADER  = PatternFill("solid", fgColor="002060")
    FILL_SECTION = PatternFill("solid", fgColor="D6DCE4")
    FILL_TOTAL   = PatternFill("solid", fgColor="FFF2CC")
    FILL_GTOTAL  = PatternFill("solid", fgColor="002060")
    FILL_NM      = PatternFill("solid", fgColor="595959")  # dark grey for n.m
    FONT_WHITE   = Font(bold=True, color="FFFFFF")
    FONT_BOLD    = Font(bold=True)
    FONT_ITALIC  = Font(italic=True, color="595959")
    FONT_NM      = Font(color="FFFFFF", size=8)
    ALIGN_CTR    = Alignment(horizontal="center", wrap_text=True)
    ALIGN_NM     = Alignment(horizontal="center")

    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(title=tab_name)

    max_cols = max((len(r) for r in rows if r), default=1)

    for r_idx, row in enumerate(rows, 1):
        if not row:
            continue
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

        first = str(row[0] or "") if row else ""
        is_header  = (r_idx == 1)
        is_section = (len([v for v in row if v not in (None, "", 0)]) == 1
                      and not first.startswith("  "))
        is_grand   = (first.strip() == "TOTAL")
        is_total   = (not is_grand and (first.strip().startswith("TOTAL") or
                      first.strip().startswith("  TOTAL")))
        is_indent  = first.startswith("  ") and not is_total

        for c_idx in range(1, max_cols + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if is_header:
                cell.fill = FILL_HEADER
                cell.font = FONT_WHITE
                cell.alignment = ALIGN_CTR
            elif is_section:
                cell.fill = FILL_SECTION
                cell.font = FONT_BOLD
            elif is_grand:
                cell.fill = FILL_GTOTAL
                cell.font = FONT_WHITE
            elif is_total:
                cell.fill = FILL_TOTAL
                cell.font = FONT_BOLD
            elif is_indent:
                cell.font = FONT_ITALIC

            # n.m cells: dark grey background, white text, centered
            if cell.value == "n.m":
                cell.fill = FILL_NM
                cell.font = FONT_NM
                cell.alignment = ALIGN_NM
                continue

            # Format numbers (column 2+)
            if c_idx >= 2 and cell.value is not None:
                v = cell.value
                if isinstance(v, float):
                    cell.number_format = '#,##0'
                elif isinstance(v, int):
                    cell.number_format = '#,##0'

    # Freeze col 2, row 2
    ws.freeze_panes = ws.cell(row=2, column=2)
    ws.row_dimensions[1].height = 42

    # Column widths
    from openpyxl.utils import get_column_letter
    ws.column_dimensions["A"].width = 24
    for i in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    print(f"  -> Tab '{tab_name}' written ({len(rows)} rows, wide format)")


def write_tab(wb, tab_name, rows, header_row=None):
    """Add or replace a sheet tab in workbook.
    header_row: 1-based index of the column-header row (styled dark blue).
    Section rows (only col A filled, no B/C) get grey background.
    Brand subtotal rows (col C empty) get light blue.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    FILL_HEADER  = PatternFill("solid", fgColor="002060")
    FILL_SECTION = PatternFill("solid", fgColor="D6DCE4")
    FILL_BRAND   = PatternFill("solid", fgColor="E2EFDA")
    FILL_TOTAL   = PatternFill("solid", fgColor="FFF2CC")
    FONT_WHITE   = Font(bold=True, color="FFFFFF")
    FONT_BOLD    = Font(bold=True)
    FONT_ITALIC  = Font(italic=True, color="595959")
    ALIGN_CTR    = Alignment(horizontal="center", wrap_text=True)

    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(title=tab_name)

    max_cols = max((len(r) for r in rows if r), default=1)

    current_section = ""  # tracks the active metric section for format decisions

    for r_idx, row in enumerate(rows, 1):
        if not row:
            continue
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

        # Detect row type
        first = str(row[0] or "") if row else ""
        is_header = (header_row and r_idx == header_row)
        is_section = (len(row) <= 1 and first and not first.startswith("  "))
        is_total   = (first == "TOTAL")
        is_brand   = (len(row) > 2 and not first.startswith("  ") and
                      row[1] and not row[2] and first not in ("TOTAL", "Metric", "Brand"))
        is_channel = first.startswith("  ")

        if is_section:
            current_section = first.upper()

        # Percent sections: DISCOUNT RATE, GM %
        is_pct_section = ("RATE" in current_section or
                          "GM %" in current_section or
                          current_section.endswith("%"))

        for c_idx in range(1, max_cols + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            if is_header:
                cell.fill = FILL_HEADER
                cell.font = FONT_WHITE
                cell.alignment = ALIGN_CTR
            elif is_section:
                cell.fill = FILL_SECTION
                cell.font = FONT_BOLD
            elif is_total:
                cell.fill = FILL_TOTAL
                cell.font = FONT_BOLD
            elif is_brand:
                cell.fill = FILL_BRAND
                cell.font = FONT_BOLD
            elif is_channel:
                cell.font = FONT_ITALIC

            # Format numbers
            if c_idx > 3 and cell.value is not None:
                v = cell.value
                if isinstance(v, float) and is_pct_section:
                    cell.number_format = '0.0%'
                elif isinstance(v, float):
                    cell.number_format = '#,##0'
                elif isinstance(v, int):
                    cell.number_format = '#,##0'

    # Freeze first 3 cols + header row
    if header_row:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=4)
        ws.row_dimensions[header_row].height = 42  # tall enough for 2-line month labels

    # Column widths — uniform
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 10
    for i in range(4, max_cols + 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = 14

    print(f"  -> Tab '{tab_name}' written ({len(rows)} rows)")


# ── Executive Summary: expand month columns ──────────────────────────────────

def expand_exec_summary_months(wb, target_start="2025-01"):
    """Ensure Executive Summary has month columns starting from target_start.
    Inserts any missing month columns before the current first month.
    New cells in existing data rows get 'n.m' with dark grey fill.
    """
    import calendar, re
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from copy import copy

    if "Executive Summary" not in wb.sheetnames:
        return

    ws = wb["Executive Summary"]

    # Parse current month columns from row 2
    def _norm(s):
        return re.sub(r'[\s\n]*[\(\[].*', '', str(s)).strip()

    existing_months = {}   # "YYYY-MM" -> col_index
    ytd_col = None
    for c in range(2, ws.max_column + 1):
        v = ws.cell(row=2, column=c).value
        if not v:
            continue
        norm = _norm(v)
        if norm == "YTD":
            ytd_col = c
            continue
        try:
            parts = norm.split()
            mon = list(calendar.month_abbr).index(parts[0])
            existing_months[f"{parts[1]}-{mon:02d}"] = c
        except (ValueError, IndexError):
            pass

    if not existing_months:
        return

    first_existing = min(existing_months.keys())
    if first_existing <= target_start:
        print(f"  -> Exec Summary already starts at {first_existing}, no expansion needed")
        return

    # Build list of months to insert
    from datetime import date
    ty, tm = int(target_start[:4]), int(target_start[5:7])
    fy, fm = int(first_existing[:4]), int(first_existing[5:7])
    months_to_add = []
    y, m = ty, tm
    while (y, m) < (fy, fm):
        months_to_add.append(f"{y}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1

    n_insert = len(months_to_add)
    if n_insert == 0:
        return

    # Insert columns at position 2 (before current first month column)
    ws.insert_cols(2, n_insert)

    # Styles
    FILL_NM = PatternFill("solid", fgColor="595959")
    FONT_NM = Font(color="FFFFFF", size=8)
    ALIGN_NM = Alignment(horizontal="center")
    # Copy header style from first existing month (now shifted)
    first_data_col = 2 + n_insert  # the shifted first existing month
    hdr_cell_ref = ws.cell(row=2, column=first_data_col)

    # Set headers for new columns (row 2)
    for i, mk in enumerate(months_to_add):
        y, mo = int(mk[:4]), int(mk[5:7])
        label = f"{calendar.month_abbr[mo]} {y}"
        cell = ws.cell(row=2, column=2 + i)
        cell.value = label
        # Copy header style from existing header
        if hdr_cell_ref.has_style:
            cell.font = copy(hdr_cell_ref.font)
            cell.fill = copy(hdr_cell_ref.fill)
            cell.alignment = copy(hdr_cell_ref.alignment)
            cell.number_format = hdr_cell_ref.number_format

    # Row 1: extend title fill across all new columns (copy from the row 1 existing cell)
    r1_ref = ws.cell(row=1, column=first_data_col)
    for i in range(n_insert):
        cell = ws.cell(row=1, column=2 + i)
        if r1_ref.has_style:
            cell.font = copy(r1_ref.font)
            cell.fill = copy(r1_ref.fill)
            cell.alignment = copy(r1_ref.alignment)

    # Fill n.m for all data rows (R3 through the row before MKT SPEND section)
    # Find where MKT SPEND starts (we write it fresh each run anyway)
    mkt_start = ws.max_row + 1
    for rx in range(3, ws.max_row + 1):
        label = str(ws.cell(row=rx, column=1).value or "").strip()
        if label == "MKT SPEND":
            mkt_start = rx
            break

    # Section header rows (no data, just label+fill) - skip these
    SECTION_LABELS = {"REVENUE", "PROFITABILITY", "ADVERTISING EFFICIENCY"}

    for rx in range(3, mkt_start):
        label = str(ws.cell(row=rx, column=1).value or "").strip()
        if not label or label in SECTION_LABELS:
            continue  # section header or blank row
        for i in range(n_insert):
            cell = ws.cell(row=rx, column=2 + i)
            cell.value = "n.m"
            cell.fill = FILL_NM
            cell.font = FONT_NM
            cell.alignment = ALIGN_NM

    # Set column widths for new columns
    for i in range(n_insert):
        cl = get_column_letter(2 + i)
        ws.column_dimensions[cl].width = 13

    print(f"  -> Exec Summary expanded: added {months_to_add} (cols B~{get_column_letter(2+n_insert-1)})")


# ── Executive Summary COGS update ─────────────────────────────────────────────

def update_exec_summary(wb, total_monthly):
    """Update Revenue, COGS, Gross Profit, GM%, and other rows in Executive Summary."""
    import calendar
    from openpyxl.styles import PatternFill, Font, Alignment

    if "Executive Summary" not in wb.sheetnames:
        print("  [WARN] Executive Summary tab not found, skipping")
        return

    ws = wb["Executive Summary"]

    # Build month -> column index map from header row (R2)
    import re as _re
    def _norm_hdr(s):
        return _re.sub(r'[\s\n]*[\(\[].*', '', str(s)).strip()

    hdr = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    col_map = {_norm_hdr(v): i + 1 for i, v in enumerate(hdr) if v}

    # Find key rows by label (search col A)
    ROW_MAP = {}  # label -> row number
    SEARCH_LABELS = {
        "Net Revenue", "Gross Sales", "Discounts", "Discount Rate",
        "Total Orders", "AOV (Net/Orders)",
        "COGS", "Gross Profit", "Gross Margin %", "GM %", "Gross Margin",
    }
    for r in range(1, ws.max_row + 1):
        v = str(ws.cell(row=r, column=1).value or "").strip()
        if v in SEARCH_LABELS:
            ROW_MAP[v] = r

    cogs_row = ROW_MAP.get("COGS")
    gp_row   = ROW_MAP.get("Gross Profit")
    gm_row   = ROW_MAP.get("Gross Margin %") or ROW_MAP.get("GM %") or ROW_MAP.get("Gross Margin")
    net_row  = ROW_MAP.get("Net Revenue")
    gross_row = ROW_MAP.get("Gross Sales")
    disc_row  = ROW_MAP.get("Discounts")
    drate_row = ROW_MAP.get("Discount Rate")
    orders_row = ROW_MAP.get("Total Orders")
    aov_row    = ROW_MAP.get("AOV (Net/Orders)")

    if not cogs_row:
        print("  [WARN] COGS row not found in Executive Summary, skipping")
        return

    # Compute YTD year from latest month
    all_months = sorted(total_monthly.keys())
    ytd_year = all_months[-1][:4] if all_months else str(__import__("datetime").date.today().year)

    # n.m sentinel styles
    FILL_NM = PatternFill("solid", fgColor="595959")
    FONT_NM = Font(color="FFFFFF", size=8)
    ALIGN_NM = Alignment(horizontal="center")

    ytd_vals = {"cogs": 0.0, "net": 0.0, "gross": 0.0, "disc": 0.0, "units": 0, "orders": 0}

    for m, v in total_monthly.items():
        y, mo = int(m[:4]), int(m[5:7])
        col_label = f"{calendar.month_abbr[mo]} {y}"
        col = col_map.get(col_label)
        if not col:
            continue

        cogs  = round(v["cogs"], 0)
        net   = round(v["net"], 0)
        gross = round(v["gross"], 0)
        disc  = round(v["disc"], 0)
        gp    = round(net - cogs, 0)
        gm    = round(gp / net, 4) if net else 0.0
        drate = round(disc / gross, 4) if gross else 0.0
        orders = v.get("orders", 0)  # order count (not units/line-items)
        aov   = round(net / orders, 0) if orders else 0

        def _set(row, val, fmt='#,##0'):
            if row:
                cell = ws.cell(row=row, column=col)
                # Clear n.m if we have real data
                if cell.value == "n.m":
                    cell.fill = PatternFill()
                    cell.font = Font()
                    cell.alignment = Alignment()
                cell.value = val
                cell.number_format = fmt

        _set(net_row, net)
        _set(gross_row, gross)
        _set(disc_row, disc)
        _set(drate_row, drate, '0.0%')
        _set(orders_row, orders, '#,##0')
        _set(aov_row, aov)
        _set(cogs_row, cogs)
        _set(gp_row, gp)
        _set(gm_row, gm, '0.0%')

        if m[:4] == ytd_year:
            for k in ytd_vals:
                ytd_vals[k] += v.get(k, 0)

    # YTD column
    ytd_col = col_map.get("YTD")
    if ytd_col:
        ytd_net  = round(ytd_vals["net"], 0)
        ytd_gross = round(ytd_vals["gross"], 0)
        ytd_disc = round(ytd_vals["disc"], 0)
        ytd_cogs = round(ytd_vals["cogs"], 0)
        ytd_gp   = round(ytd_net - ytd_cogs, 0)
        ytd_gm   = round(ytd_gp / ytd_net, 4) if ytd_net else 0.0
        ytd_drate = round(ytd_disc / ytd_gross, 4) if ytd_gross else 0.0
        ytd_orders = ytd_vals["orders"]
        ytd_aov  = round(ytd_net / ytd_orders, 0) if ytd_orders else 0

        def _setytd(row, val, fmt='#,##0'):
            if row:
                cell = ws.cell(row=row, column=ytd_col)
                if cell.value == "n.m":
                    cell.fill = PatternFill()
                    cell.font = Font()
                    cell.alignment = Alignment()
                cell.value = val
                cell.number_format = fmt

        _setytd(net_row, ytd_net)
        _setytd(gross_row, ytd_gross)
        _setytd(disc_row, ytd_disc)
        _setytd(drate_row, ytd_drate, '0.0%')
        _setytd(orders_row, ytd_orders, '#,##0')
        _setytd(aov_row, ytd_aov)
        _setytd(cogs_row, ytd_cogs)
        _setytd(gp_row, ytd_gp)
        _setytd(gm_row, ytd_gm, '0.0%')

    print(f"  -> Executive Summary updated: Revenue/COGS/GP/GM% + all metric rows")


# ── Executive Summary MKT SPEND section ───────────────────────────────────────

def add_mkt_spend_to_exec_summary(wb):
    """Add MKT SPEND section to Executive Summary using Excel FORMULAS
    that reference detail tabs (KPI_광고비, KPI_시딩비용, KPI_할인율).
    Must be called AFTER those tabs are written to the workbook.
    """
    import calendar, re
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    TAB_AD   = 'KPI_광고비'
    TAB_SEED = 'KPI_시딩비용'
    TAB_DISC = 'KPI_할인율'

    for tab in ("Executive Summary", TAB_AD, TAB_SEED, TAB_DISC):
        if tab not in wb.sheetnames:
            print(f"  [WARN] '{tab}' not found, skipping MKT Spend")
            return

    ws = wb["Executive Summary"]
    max_col = ws.max_column
    GROSS_ROW = 5  # Gross Sales row

    # ── Styles ────────────────────────────────────────────────────────────────
    FILL_SECTION = PatternFill("solid", fgColor="D6DCE4")
    FILL_TOTAL   = PatternFill("solid", fgColor="FFF2CC")
    FILL_GRAND   = PatternFill("solid", fgColor="002060")
    FONT_BOLD    = Font(bold=True)
    FONT_WHITE   = Font(bold=True, color="FFFFFF")
    FONT_ITALIC  = Font(italic=True, color="595959")

    # ── Remove previous MKT SPEND section ────────────────────────────────────
    for rx in range(1, ws.max_row + 1):
        if str(ws.cell(row=rx, column=1).value or "").strip() == "MKT SPEND":
            start_del = max(rx - 1, rx)
            ws.delete_rows(start_del, ws.max_row - start_del + 1)
            break

    # ── Exec Summary: month -> col index ─────────────────────────────────────
    def _norm(s):
        return re.sub(r'[\s\n]*[\(\[].*', '', str(s)).strip()

    exec_months = {}   # "YYYY-MM" -> col_index
    exec_ytd = None
    for c in range(2, max_col + 1):
        v = ws.cell(row=2, column=c).value
        if not v:
            continue
        norm = _norm(v)
        if norm == "YTD":
            exec_ytd = c
            continue
        try:
            parts = norm.split()
            mon = list(calendar.month_abbr).index(parts[0])
            exec_months[f"{parts[1]}-{mon:02d}"] = c
        except (ValueError, IndexError):
            pass

    all_months = sorted(exec_months.keys())

    # ── Detail tab column maps: {YYYY-MM: col_letter, "YTD": col_letter} ────
    def _tab_cols(ws_tab, hdr_row, start_col):
        result = {}
        for c in range(start_col, ws_tab.max_column + 1):
            h = str(ws_tab.cell(row=hdr_row, column=c).value or "")
            h = h.split('\n')[0].strip()
            if h == "YTD":
                result["YTD"] = get_column_letter(c)
                continue
            try:
                parts = h.split()
                mon = list(calendar.month_abbr).index(parts[0])
                result[f"{parts[1]}-{mon:02d}"] = get_column_letter(c)
            except (ValueError, IndexError):
                pass
        return result

    ws_ad   = wb[TAB_AD]
    ws_seed = wb[TAB_SEED]
    ws_disc = wb[TAB_DISC]
    ad_cols   = _tab_cols(ws_ad,   1, 2)   # header row 1, data from col 2
    seed_cols = _tab_cols(ws_seed, 1, 2)
    disc_cols = _tab_cols(ws_disc, 3, 4)   # header row 3, data from col 4

    # ── Find key rows in detail tabs ─────────────────────────────────────────
    # KPI_광고비: find Amazon/Meta/Google/TOTAL rows
    ad_rows = {}
    for r in range(1, ws_ad.max_row + 1):
        label = str(ws_ad.cell(row=r, column=1).value or "").strip()
        if label in ("TOTAL Amazon", "Meta", "Google", "TOTAL"):
            ad_rows[label] = r

    # KPI_시딩비용: find TOTAL row
    seed_total_row = None
    for r in range(1, ws_seed.max_row + 1):
        if str(ws_seed.cell(row=r, column=1).value or "").strip() == "TOTAL":
            seed_total_row = r

    # KPI_할인율: find rows in DISCOUNTS ($) section
    DISC_SECTIONS = {"GROSS SALES ($)", "NET SALES ($)", "DISCOUNTS ($)",
                     "DISCOUNT RATE", "UNITS", "AVG LIST PRICE ($/unit)",
                     "COGS est. ($)", "GM ($)", "GM %"}
    in_disc_section = False
    disc_total_row = None
    disc_ch_rows = defaultdict(list)   # channel_display -> [row_numbers]
    for r in range(1, ws_disc.max_row + 1):
        raw = str(ws_disc.cell(row=r, column=1).value or "")
        label = raw.strip()

        if label == "DISCOUNTS ($)":
            in_disc_section = True
            continue

        if in_disc_section:
            # Detect next section header by known name
            if label in DISC_SECTIONS:
                in_disc_section = False
                continue

            if label == "TOTAL":
                disc_total_row = r
            elif raw.startswith("  ") and label:
                disc_ch_rows[label].append(r)

    # Order channels
    ch_order = [c for c in CHANNEL_ORDER if c in disc_ch_rows]
    for c in sorted(disc_ch_rows.keys()):
        if c not in ch_order:
            ch_order.append(c)

    # ── Formula builders ─────────────────────────────────────────────────────
    def _ref(tab, cols, row, month):
        """='tab'!col_letter+row for a month (or 'YTD')."""
        cl = cols.get(month)
        if cl and row:
            return f"='{tab}'!{cl}{row}"
        return 0

    def _sumrefs(tab, cols, rows_list, month):
        """=SUM('tab'!col+r1, ...) for multiple rows."""
        cl = cols.get(month)
        if not cl or not rows_list:
            return 0
        if len(rows_list) == 1:
            return f"='{tab}'!{cl}{rows_list[0]}"
        refs = ",".join(f"'{tab}'!{cl}{r}" for r in rows_list)
        return f"=SUM({refs})"

    # ── YTD year column range ────────────────────────────────────────────────
    ytd_year = all_months[-1][:4] if all_months else "2026"
    ytd_month_cols = sorted([exec_months[m] for m in all_months if m[:4] == ytd_year])
    ytd_first_cl = get_column_letter(min(ytd_month_cols)) if ytd_month_cols else None
    ytd_last_cl  = get_column_letter(max(ytd_month_cols)) if ytd_month_cols else None

    # ── Writing helpers ──────────────────────────────────────────────────────
    r = ws.max_row + 2
    start_r = r

    def _style_row(style):
        if style == 'indent':
            ws.cell(row=r, column=1).font = FONT_ITALIC
        elif style == 'total':
            ws.cell(row=r, column=1).font = FONT_BOLD
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = FILL_TOTAL
        elif style == 'grand':
            ws.cell(row=r, column=1).font = FONT_WHITE
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = FILL_GRAND

    def _ytd_cell(row, style=None, fmt='#,##0'):
        """Write YTD as =SUM of year's months in same row."""
        if exec_ytd and ytd_first_cl and ytd_last_cl:
            cell = ws.cell(row=row, column=exec_ytd)
            cell.value = f"=SUM({ytd_first_cl}{row}:{ytd_last_cl}{row})"
            cell.number_format = fmt
            if style == 'total':
                cell.fill = FILL_TOTAL
                cell.font = FONT_BOLD
            elif style == 'grand':
                cell.fill = FILL_GRAND
                cell.font = FONT_WHITE

    def _section(label):
        nonlocal r
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = FILL_SECTION
        ws.cell(row=r, column=1, value=label).font = FONT_BOLD
        r += 1

    def _formula_row(label, mk_formula, style=None):
        """Write a row: month cells from mk_formula(m), YTD = SUM of year's months.
        Returns the written row number."""
        nonlocal r
        ws.cell(row=r, column=1, value=label)
        _style_row(style)
        for m in all_months:
            col = exec_months.get(m)
            if not col:
                continue
            cell = ws.cell(row=r, column=col)
            cell.value = mk_formula(m)
            cell.number_format = '#,##0'
            if style == 'grand':
                cell.font = FONT_WHITE
        _ytd_cell(r, style)
        written = r
        r += 1
        return written

    def _sum_row(label, from_row, to_row, style=None):
        """Write a row that SUMs a range of rows within the same sheet.
        Each month col: =SUM(col+from : col+to).  YTD = SUM of year's months.
        Returns the written row number."""
        nonlocal r
        ws.cell(row=r, column=1, value=label)
        _style_row(style)
        for m in all_months:
            col = exec_months.get(m)
            if not col:
                continue
            cl = get_column_letter(col)
            cell = ws.cell(row=r, column=col)
            cell.value = f"=SUM({cl}{from_row}:{cl}{to_row})"
            cell.number_format = '#,##0'
            if style == 'grand':
                cell.font = FONT_WHITE
        _ytd_cell(r, style)
        written = r
        r += 1
        return written

    def _pct_row(label, numerator_row):
        """=IFERROR(numerator/Gross, 0) for each month + YTD."""
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = FONT_ITALIC
        for col in list(exec_months.values()) + ([exec_ytd] if exec_ytd else []):
            cl = get_column_letter(col)
            cell = ws.cell(row=r, column=col)
            cell.value = f"=IFERROR({cl}{numerator_row}/{cl}{GROSS_ROW},0)"
            cell.number_format = '0.0%'
            cell.font = FONT_ITALIC
        r += 1

    FILL_NM = PatternFill("solid", fgColor="595959")
    FONT_NM_CELL = Font(color="FFFFFF", size=8)
    ALIGN_NM = Alignment(horizontal="center")

    def _nm_row(label, style='indent'):
        """Write a row filled with 'n.m' (not measured) for all month + YTD cells.
        Returns the written row number."""
        nonlocal r
        ws.cell(row=r, column=1, value=label)
        _style_row(style)
        for col in list(exec_months.values()) + ([exec_ytd] if exec_ytd else []):
            cell = ws.cell(row=r, column=col)
            cell.value = "n.m"
            cell.fill = FILL_NM
            cell.font = FONT_NM_CELL
            cell.alignment = ALIGN_NM
        written = r
        r += 1
        return written

    # ── Write MKT SPEND section ──────────────────────────────────────────────
    _section("MKT SPEND")

    # -- Ad Spend: individual rows link to KPI_광고비, total = in-page SUM --
    _section("Ad Spend")
    amz_r = _formula_row("  Amazon Ads",
        lambda m: _ref(TAB_AD, ad_cols, ad_rows.get("TOTAL Amazon"), m), 'indent')
    meta_r = _formula_row("  Meta Ads",
        lambda m: _ref(TAB_AD, ad_cols, ad_rows.get("Meta"), m), 'indent')
    goog_r = _formula_row("  Google Ads",
        lambda m: _ref(TAB_AD, ad_cols, ad_rows.get("Google"), m), 'indent')
    tiktok_r = _nm_row("  TikTok Ads")
    ad_r = _sum_row("Total Ad Spend", amz_r, tiktok_r, 'total')
    _pct_row("  % of Gross Sales", ad_r)

    # -- Seeding Cost: link to KPI_시딩비용 --
    _section("Seeding Cost")
    seed_r = _formula_row("Total Seeding",
        lambda m: _ref(TAB_SEED, seed_cols, seed_total_row, m), 'total')
    _pct_row("  % of Gross Sales", seed_r)

    # -- Discounts by Channel: individual rows link to KPI_할인율, total = in-page SUM --
    _section("Discounts (by Channel)")
    ch_first_r = None
    ch_last_r = None
    for ch in ch_order:
        rows_for_ch = disc_ch_rows.get(ch, [])
        ch_r = _formula_row(f"  {ch}",
            lambda m, _rows=rows_for_ch: _sumrefs(TAB_DISC, disc_cols, _rows, m),
            'indent')
        if ch_first_r is None:
            ch_first_r = ch_r
        ch_last_r = ch_r
    disc_r = _sum_row("Total Discounts", ch_first_r, ch_last_r, 'total') if ch_first_r else r
    _pct_row("  Disc Rate (% of Gross)", disc_r)

    # -- Grand Total: =Ad + Seed + Disc (in-page formula) --
    def _grand(m):
        col = exec_months.get(m)
        if not col:
            return 0
        cl = get_column_letter(col)
        return f"={cl}{ad_r}+{cl}{seed_r}+{cl}{disc_r}"
    grand_r = _formula_row("TOTAL MKT SPEND", _grand, 'grand')
    _pct_row("  % of Gross Sales", grand_r)

    # ── Link R19 (Ad Spend) and R24 (Total Ad Spend) to MKT SPEND ────────
    for target_row in (19, 24):
        label = str(ws.cell(row=target_row, column=1).value or "").strip()
        if label in ("Ad Spend", "Total Ad Spend"):
            for col in list(exec_months.values()) + ([exec_ytd] if exec_ytd else []):
                cl = get_column_letter(col)
                ws.cell(row=target_row, column=col).value = f"={cl}{ad_r}"
                ws.cell(row=target_row, column=col).number_format = '#,##0'
            print(f"  -> R{target_row} '{label}' linked to MKT SPEND R{ad_r}")

    print(f"  -> Executive Summary MKT SPEND section added (rows {start_r} ~ {r - 1}, formula-linked)")


# ── Summary tab D2C section ────────────────────────────────────────────────────

def add_summary_d2c_section(wb, d2c_monthly, seeding_rows, adspend_rows):
    """Append a D2C KPI summary block to the existing Summary tab."""
    import calendar
    from openpyxl.styles import PatternFill, Font, Alignment

    if "Summary" not in wb.sheetnames:
        print("  [WARN] Summary tab not found, skipping")
        return

    ws = wb["Summary"]

    FILL_HEADER  = PatternFill("solid", fgColor="002060")
    FILL_SECTION = PatternFill("solid", fgColor="D6DCE4")
    FILL_TOTAL   = PatternFill("solid", fgColor="FFF2CC")
    FONT_WHITE   = Font(bold=True, color="FFFFFF")
    FONT_BOLD    = Font(bold=True)
    ALIGN_CTR    = Alignment(horizontal="center", wrap_text=True)

    # Parse wide-format rows: header row[0] has month labels, find TOTAL row
    def _parse_wide_total(rows):
        """Extract {month_str: value} from wide-format rows TOTAL row."""
        import calendar as _cal
        if not rows:
            return {}
        hdr = rows[0]  # ["Label/Platform", "Jan 2024", ..., "YTD"]
        # Build mapping: month_abbr_year -> "YYYY-MM"
        col_to_month = {}
        for ci, label in enumerate(hdr[1:], 1):
            if label == "YTD":
                continue
            try:
                parts = str(label).split()
                if len(parts) == 2:
                    mon_abbr, yr = parts
                    mon_num = list(_cal.month_abbr).index(mon_abbr)
                    col_to_month[ci] = f"{yr}-{mon_num:02d}"
            except (ValueError, IndexError):
                pass
        result = {}
        for row in rows[1:]:
            if row and str(row[0]).strip() == "TOTAL":
                for ci, month_str in col_to_month.items():
                    if ci < len(row):
                        v = row[ci]
                        result[month_str] = float(v) if v not in (None, "", "n.m") else 0.0
                break
        return result

    ad_dict = _parse_wide_total(adspend_rows)
    seed_dict = _parse_wide_total(seeding_rows)

    months = sorted(d2c_monthly.keys())
    if not months:
        print("  [WARN] No D2C monthly data, skipping Summary section")
        return

    # Month column headers
    col_labels = []
    for m in months:
        y, mo = int(m[:4]), int(m[5:7])
        col_labels.append(f"{calendar.month_abbr[mo]} {y}")
    col_labels.append("YTD")

    ytd_year = months[-1][:4]

    def ytd(d):
        return sum(v for m, v in d.items() if m[:4] == ytd_year)

    def d2c_val(field, m):
        return d2c_monthly[m].get(field, 0)

    # Build rows: [label] + [monthly values] + [YTD]
    metrics = [
        ("D2C Gross Sales ($)",   [round(d2c_val("gross", m), 0) for m in months],  round(ytd({m: d2c_val("gross",m) for m in months}), 0),  '#,##0', False),
        ("D2C Net Sales ($)",     [round(d2c_val("net",   m), 0) for m in months],  round(ytd({m: d2c_val("net",  m) for m in months}), 0),  '#,##0', False),
        ("D2C Discount ($)",      [round(d2c_val("disc",  m), 0) for m in months],  round(ytd({m: d2c_val("disc", m) for m in months}), 0),  '#,##0', False),
        ("D2C Discount Rate",     [round(d2c_val("disc",m)/d2c_val("gross",m), 4) if d2c_val("gross",m) else 0 for m in months],
                                   round(ytd({m: d2c_val("disc",m) for m in months}) / ytd({m: d2c_val("gross",m) for m in months}), 4)
                                   if ytd({m: d2c_val("gross",m) for m in months}) else 0,  '0.0%', True),
        ("Ad Spend (total $)",    [round(ad_dict.get(m, 0), 0) for m in months],    round(ytd({m: ad_dict.get(m,0) for m in months}), 0),    '#,##0', False),
        ("Seeding Cost ($)",      [round(seed_dict.get(m, 0), 0) for m in months],  round(ytd({m: seed_dict.get(m,0) for m in months}), 0),  '#,##0', False),
    ]

    # Start writing after last used row (+ 3 blank rows)
    start = ws.max_row + 3

    # Section title
    ws.cell(row=start, column=1).value = "D2C KPI SUMMARY  |  D2C = ONZ + Amazon + TikTok"
    for c in range(1, len(col_labels) + 2):
        ws.cell(row=start, column=c).fill = FILL_SECTION
        ws.cell(row=start, column=c).font = FONT_BOLD
    start += 1

    # Header row
    ws.cell(row=start, column=1).value = "Metric"
    for ci, label in enumerate(col_labels, 2):
        ws.cell(row=start, column=ci).value = label
        ws.cell(row=start, column=ci).fill = FILL_HEADER
        ws.cell(row=start, column=ci).font = FONT_WHITE
        ws.cell(row=start, column=ci).alignment = ALIGN_CTR
    ws.cell(row=start, column=1).fill = FILL_HEADER
    ws.cell(row=start, column=1).font = FONT_WHITE
    start += 1

    # Data rows
    for label, vals, ytd_val, fmt, is_pct in metrics:
        ws.cell(row=start, column=1).value = label
        ws.cell(row=start, column=1).font = FONT_BOLD
        for ci, v in enumerate(vals, 2):
            cell = ws.cell(row=start, column=ci)
            cell.value = v
            cell.number_format = fmt
        ytd_cell = ws.cell(row=start, column=len(col_labels) + 1)
        ytd_cell.value = ytd_val
        ytd_cell.number_format = fmt
        ytd_cell.fill = FILL_TOTAL
        start += 1

    # Column widths (only if narrower than needed)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 22)
    for i in range(2, len(col_labels) + 2):
        ltr = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[ltr].width = max(ws.column_dimensions[ltr].width, 12)

    print(f"  -> Summary D2C section appended (row {ws.max_row - len(metrics) + 1} ~)")


# ── Amazon Discount Debug Tab ─────────────────────────────────────────────────

def add_amazon_discount_tab(wb, through_date):
    """Add KPI_Amazon할인_상세 tab in wide format (months as columns, like other KPI tabs).
    Metrics: Gross Sales, Discounts, Disc Rate, Net Sales, Units — per brand + TOTAL.
    """
    import calendar
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dtt, timedelta as _td, timezone as _tz

    TAB = "KPI_Amazon할인_상세"
    if TAB in wb.sheetnames:
        del wb[TAB]
    ws = wb.create_sheet(TAB)

    FILL_HDR     = PatternFill("solid", fgColor="002060")
    FILL_SECTION = PatternFill("solid", fgColor="D6DCE4")
    FILL_TOTAL   = PatternFill("solid", fgColor="FFF2CC")
    FILL_GTOTAL  = PatternFill("solid", fgColor="002060")
    FONT_W       = Font(bold=True, color="FFFFFF")
    FONT_B       = Font(bold=True)
    FONT_IT      = Font(italic=True, color="595959")
    ALIGN_C      = Alignment(horizontal="center", wrap_text=True)

    # ── Load data ──────────────────────────────────────────────────────────────
    rows = load_dk("shopify_orders_daily", days=800)
    amz = [r for r in rows if r.get("channel") == "Amazon"
           and (r.get("date") or "") <= through_date]

    # Aggregate by brand x month
    brand_month = defaultdict(lambda: defaultdict(lambda: {"gross":0.0,"disc":0.0,"net":0.0,"units":0}))
    month_agg   = defaultdict(lambda: {"gross":0.0,"disc":0.0,"net":0.0,"units":0})
    for r in amz:
        m = (r.get("date") or "")[:7]
        brand = r.get("brand") or "Unknown"
        g = float(r.get("gross_sales",0) or 0)
        d = float(r.get("discounts",0) or 0)
        n = float(r.get("net_sales",0) or 0)
        u = int(r.get("units",0) or 0)
        brand_month[brand][m]["gross"] += g
        brand_month[brand][m]["disc"]  += d
        brand_month[brand][m]["net"]   += n
        brand_month[brand][m]["units"] += u
        month_agg[m]["gross"] += g
        month_agg[m]["disc"]  += d
        month_agg[m]["net"]   += n
        month_agg[m]["units"] += u

    sorted_months = sorted(month_agg.keys())
    _today = _dtt.now(_tz(_td(hours=-8))).date()
    ytd_year = sorted_months[-1][:4] if sorted_months else "2026"

    def _mlabel(m):
        y, mo = int(m[:4]), int(m[5:7])
        return f"{calendar.month_abbr[mo]} {y}"

    brands = sorted(brand_month.keys(),
                    key=lambda b: ["Grosmimi","Naeiae","CHA&MOM","Unknown"].index(b)
                    if b in ["Grosmimi","Naeiae","CHA&MOM","Unknown"] else 99)

    # ── Note row ───────────────────────────────────────────────────────────────
    ws.cell(row=1, column=1,
            value=f"Amazon Channel Discount Detail  |  Data through {through_date}  |  Shopify channel='Amazon' only (not FBA MCF)")
    ws.cell(row=1, column=1).font = FONT_B
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="FFF2CC")

    # ── Header row ─────────────────────────────────────────────────────────────
    hdr = ["Metric / Brand"] + [_mlabel(m) for m in sorted_months] + ["YTD"]
    for ci, h in enumerate(hdr, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = FILL_HDR
        cell.font = FONT_W
        cell.alignment = ALIGN_C
    ws.row_dimensions[2].height = 30

    # ── Write metric sections ──────────────────────────────────────────────────
    dr = 3
    METRICS = [
        ("Gross Sales ($)",  "gross", '#,##0'),
        ("Discounts ($)",    "disc",  '#,##0'),
        ("Net Sales ($)",    "net",   '#,##0'),
        ("Disc Rate (%)",    "rate",  '0.0%'),
        ("Units",            "units", '#,##0'),
    ]

    for metric_label, field, fmt in METRICS:
        # Section header
        for ci in range(1, len(hdr) + 1):
            ws.cell(row=dr, column=ci).fill = FILL_SECTION
        ws.cell(row=dr, column=1, value=metric_label).font = FONT_B
        dr += 1

        for brand in brands + ["TOTAL"]:
            is_total = (brand == "TOTAL")
            label = "TOTAL" if is_total else f"  {brand}"
            ws.cell(row=dr, column=1, value=label)

            if is_total:
                ws.cell(row=dr, column=1).font = FONT_B
            else:
                ws.cell(row=dr, column=1).font = FONT_IT

            ytd_g = ytd_d = ytd_n = 0.0
            ytd_u = 0
            for ci, m in enumerate(sorted_months, 2):
                if is_total:
                    mv = month_agg[m]
                else:
                    mv = brand_month[brand].get(m, {"gross":0,"disc":0,"net":0,"units":0})

                if field == "rate":
                    val = mv["disc"] / mv["gross"] if mv["gross"] else 0.0
                elif field == "units":
                    val = mv["units"]
                else:
                    val = mv[field]

                cell = ws.cell(row=dr, column=ci, value=round(val, 4) if field == "rate" else round(val, 0))
                cell.number_format = fmt
                if is_total:
                    cell.fill = FILL_TOTAL
                    cell.font = FONT_B

                if m[:4] == ytd_year:
                    ytd_g += mv["gross"]; ytd_d += mv["disc"]
                    ytd_n += mv["net"];   ytd_u += mv["units"]

            # YTD column
            ytd_col = len(sorted_months) + 2
            if field == "rate":
                ytd_val = round(ytd_d / ytd_g, 4) if ytd_g else 0.0
            elif field == "units":
                ytd_val = ytd_u
            elif field == "gross":
                ytd_val = round(ytd_g, 0)
            elif field == "disc":
                ytd_val = round(ytd_d, 0)
            else:
                ytd_val = round(ytd_n, 0)

            cell = ws.cell(row=dr, column=ytd_col, value=ytd_val)
            cell.number_format = fmt
            if is_total:
                cell.fill = FILL_TOTAL
                cell.font = FONT_B

            dr += 1

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 24
    for ci in range(2, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.freeze_panes = ws.cell(row=3, column=2)
    print(f"  -> Tab '{TAB}' written (wide format, {len(brands)} brands, {len(sorted_months)} months)")


# ── Amazon Marketplace Discount (SP-API actual sales vs Shopify ref price) ────

def add_amazon_marketplace_tab(wb, through_date):
    """Add KPI_Amazon_MP할인 tab: true Amazon Marketplace discount vs Shopify reference price.

    Uses amazon_sales_daily (SP-API) for actual Amazon selling prices,
    and shopify_orders_daily D2C gross_sales/units as the reference (compare_at_price basis).
    """
    import calendar
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    TAB = "KPI_Amazon_MP할인"
    if TAB in wb.sheetnames:
        del wb[TAB]
    ws = wb.create_sheet(TAB)

    FILL_HDR     = PatternFill("solid", fgColor="002060")
    FILL_SECTION = PatternFill("solid", fgColor="D6DCE4")
    FILL_TOTAL   = PatternFill("solid", fgColor="FFF2CC")
    FONT_W       = Font(bold=True, color="FFFFFF")
    FONT_B       = Font(bold=True)
    FONT_IT      = Font(italic=True, color="595959")
    ALIGN_C      = Alignment(horizontal="center", wrap_text=True)

    # ── Load Amazon Marketplace data ──────────────────────────────────────────
    amz_rows = load_dk("amazon_sales_daily", days=800)
    amz_rows = [r for r in amz_rows if (r.get("date") or "") <= through_date]

    # Aggregate by brand x month
    amz_bm = defaultdict(lambda: defaultdict(lambda: {"gross": 0.0, "units": 0, "orders": 0}))
    amz_month = defaultdict(lambda: {"gross": 0.0, "units": 0, "orders": 0})
    for r in amz_rows:
        m = (r.get("date") or "")[:7]
        brand = r.get("brand") or "Unknown"
        g = float(r.get("gross_sales", 0) or 0)
        u = int(r.get("units", 0) or 0)
        o = int(r.get("orders", 0) or 0)
        amz_bm[brand][m]["gross"] += g
        amz_bm[brand][m]["units"] += u
        amz_bm[brand][m]["orders"] += o
        amz_month[m]["gross"] += g
        amz_month[m]["units"] += u
        amz_month[m]["orders"] += o

    if not amz_month:
        print(f"  [WARN] No Amazon Marketplace data, skipping {TAB}")
        return

    # ── Build Shopify D2C reference ASP by brand x month ─────────────────────
    shopify_rows = load_dk("shopify_orders_daily", days=800)
    d2c_bm = defaultdict(lambda: defaultdict(lambda: {"gross": 0.0, "units": 0}))
    for r in shopify_rows:
        ch = r.get("channel") or ""
        if ch not in ("D2C", "ONZ"):
            continue
        m = (r.get("date") or "")[:7]
        brand = r.get("brand") or "Unknown"
        g = float(r.get("gross_sales", 0) or 0)
        u = int(r.get("units", 0) or 0)
        d2c_bm[brand][m]["gross"] += g
        d2c_bm[brand][m]["units"] += u

    def _get_ref_asp(brand, month):
        """Get Shopify D2C reference ASP for brand. Falls back to nearest month, then AVG_PRICE."""
        # 1. Same month
        dv = d2c_bm.get(brand, {}).get(month, {})
        if dv.get("units", 0) >= 5:  # need min 5 units for meaningful ASP
            return dv["gross"] / dv["units"]
        # 2. Nearest month with data (search backwards then forwards)
        all_months = sorted(d2c_bm.get(brand, {}).keys(), reverse=True)
        for m in all_months:
            if m <= month:
                dv2 = d2c_bm[brand][m]
                if dv2.get("units", 0) >= 5:
                    return dv2["gross"] / dv2["units"]
        for m in sorted(d2c_bm.get(brand, {}).keys()):
            dv2 = d2c_bm[brand][m]
            if dv2.get("units", 0) >= 5:
                return dv2["gross"] / dv2["units"]
        # 3. Fallback to AVG_PRICE
        return AVG_PRICE.get(brand, AVG_PRICE.get("Unknown", 25.0))

    sorted_months = sorted(amz_month.keys())
    ytd_year = sorted_months[-1][:4] if sorted_months else "2026"

    def _mlabel(m):
        y, mo = int(m[:4]), int(m[5:7])
        return f"{calendar.month_abbr[mo]} {y}"

    brands = sorted(amz_bm.keys(),
                    key=lambda b: ["Grosmimi", "Naeiae", "CHA&MOM"].index(b)
                    if b in ["Grosmimi", "Naeiae", "CHA&MOM"] else 99)

    # ── Note row ──────────────────────────────────────────────────────────────
    ws.cell(row=1, column=1,
            value=f"Amazon Marketplace Discount  |  Data through {through_date}  |  Ref = Shopify D2C gross ASP (compare_at_price basis)")
    ws.cell(row=1, column=1).font = FONT_B
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="FFF2CC")

    # ── Header row ────────────────────────────────────────────────────────────
    hdr = ["Metric / Brand"] + [_mlabel(m) for m in sorted_months] + ["YTD"]
    for ci, h in enumerate(hdr, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = FILL_HDR
        cell.font = FONT_W
        cell.alignment = ALIGN_C
    ws.row_dimensions[2].height = 30

    # ── Metric sections ───────────────────────────────────────────────────────
    METRICS = [
        ("Amazon Gross Sales ($)", "amz_gross", '#,##0'),
        ("Amazon Units",           "amz_units", '#,##0'),
        ("Amazon ASP ($)",         "amz_asp",   '$#,##0.00'),
        ("Reference ASP ($)",      "ref_asp",   '$#,##0.00'),
        ("Implied Discount ($)",   "disc",      '#,##0'),
        ("Implied Disc Rate (%)",  "rate",      '0.0%'),
    ]

    dr = 3
    for metric_label, field, fmt in METRICS:
        # Section header
        for ci in range(1, len(hdr) + 1):
            ws.cell(row=dr, column=ci).fill = FILL_SECTION
        ws.cell(row=dr, column=1, value=metric_label).font = FONT_B
        dr += 1

        for brand in brands + ["TOTAL"]:
            is_total = (brand == "TOTAL")
            label = "TOTAL" if is_total else f"  {brand}"
            ws.cell(row=dr, column=1, value=label)
            ws.cell(row=dr, column=1).font = FONT_B if is_total else FONT_IT

            # YTD accumulators
            ytd_amz_g = ytd_amz_u = 0.0
            ytd_ref_weighted = 0.0  # ref_asp * amz_units, for weighted avg

            for ci, m in enumerate(sorted_months, 2):
                if is_total:
                    mv = amz_month[m]
                else:
                    mv = amz_bm.get(brand, {}).get(m, {"gross": 0, "units": 0, "orders": 0})

                amz_g = mv["gross"]
                amz_u = mv["units"]
                amz_asp = amz_g / amz_u if amz_u else 0

                if is_total:
                    # Weighted ref ASP across brands
                    ref_total = 0.0
                    for b in brands:
                        bmv = amz_bm.get(b, {}).get(m, {"gross": 0, "units": 0})
                        if bmv["units"]:
                            ref_total += _get_ref_asp(b, m) * bmv["units"]
                    ref_asp = ref_total / amz_u if amz_u else 0
                else:
                    ref_asp = _get_ref_asp(brand, m) if amz_u else 0

                disc_dollar = (ref_asp - amz_asp) * amz_u if ref_asp and amz_u else 0
                disc_rate = (ref_asp - amz_asp) / ref_asp if ref_asp and amz_asp else 0
                # Floor at 0 — if Amazon price > Shopify ref, no negative discount
                disc_dollar = max(disc_dollar, 0)
                disc_rate = max(disc_rate, 0)

                if field == "amz_gross":
                    val = round(amz_g, 0)
                elif field == "amz_units":
                    val = amz_u
                elif field == "amz_asp":
                    val = round(amz_asp, 2)
                elif field == "ref_asp":
                    val = round(ref_asp, 2) if amz_u else 0
                elif field == "disc":
                    val = round(disc_dollar, 0)
                elif field == "rate":
                    val = round(disc_rate, 4)
                else:
                    val = 0

                cell = ws.cell(row=dr, column=ci, value=val)
                cell.number_format = fmt
                if is_total:
                    cell.fill = FILL_TOTAL
                    cell.font = FONT_B

                if m[:4] == ytd_year:
                    ytd_amz_g += amz_g
                    ytd_amz_u += amz_u
                    ytd_ref_weighted += ref_asp * amz_u if amz_u else 0

            # YTD column
            ytd_col = len(sorted_months) + 2
            ytd_amz_asp = ytd_amz_g / ytd_amz_u if ytd_amz_u else 0
            ytd_ref_asp = ytd_ref_weighted / ytd_amz_u if ytd_amz_u else 0
            ytd_disc = max((ytd_ref_asp - ytd_amz_asp) * ytd_amz_u, 0)
            ytd_rate = max((ytd_ref_asp - ytd_amz_asp) / ytd_ref_asp, 0) if ytd_ref_asp else 0

            if field == "amz_gross":
                ytd_val = round(ytd_amz_g, 0)
            elif field == "amz_units":
                ytd_val = int(ytd_amz_u)
            elif field == "amz_asp":
                ytd_val = round(ytd_amz_asp, 2)
            elif field == "ref_asp":
                ytd_val = round(ytd_ref_asp, 2)
            elif field == "disc":
                ytd_val = round(ytd_disc, 0)
            elif field == "rate":
                ytd_val = round(ytd_rate, 4)
            else:
                ytd_val = 0

            cell = ws.cell(row=dr, column=ytd_col, value=ytd_val)
            cell.number_format = fmt
            if is_total:
                cell.fill = FILL_TOTAL
                cell.font = FONT_B

            dr += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 24
    for ci in range(2, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    ws.freeze_panes = ws.cell(row=3, column=2)
    print(f"  -> Tab '{TAB}' written (wide format, {len(brands)} brands, {len(sorted_months)} months)")


# ── SUMMARY TABS (Sales Summary / Ads Summary / Unit Economics) ──────────────

def _summary_month_range(through_date):
    """Return 12 months ending at through_date's month, plus YTD year."""
    import calendar
    from datetime import date as _d
    td = _d.fromisoformat(through_date)
    months = []
    y, m = td.year, td.month
    for _ in range(12):
        months.append(f"{y}-{m:02d}")
        m -= 1
        if m == 0:
            m = 12; y -= 1
    months.reverse()
    ytd_year = str(td.year)
    return months, ytd_year


def _write_summary_tab(wb, tab_name, title, sections, months, ytd_year, through_date):
    """Write/overwrite a summary tab with wide-format sections.
    sections = [
        (section_title, [
            (row_label, {month: value}, ytd_value, fmt),
            ...
        ]),
        ...
    ]
    """
    import calendar
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import date as _d

    if tab_name in wb.sheetnames:
        idx = wb.sheetnames.index(tab_name)
        del wb[tab_name]
        ws = wb.create_sheet(tab_name, idx)
    else:
        ws = wb.create_sheet(tab_name)

    FILL_HDR = PatternFill("solid", fgColor="002060")
    FILL_SECTION = PatternFill("solid", fgColor="D6DCE4")
    FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")
    FONT_W = Font(bold=True, color="FFFFFF")
    FONT_B = Font(bold=True)
    FONT_IT = Font(italic=True, color="595959")
    ALIGN_C = Alignment(horizontal="center", wrap_text=True)

    td = _d.fromisoformat(through_date)

    # Title
    ws.cell(row=1, column=1, value=title).font = FONT_B

    # Header row
    r = 2
    ws.cell(row=r, column=1)  # blank
    for ci, m in enumerate(months, 2):
        y, mo = int(m[:4]), int(m[5:7])
        label = f"{calendar.month_abbr[mo]} {y}"
        if y == td.year and mo == td.month:
            label += f"\n(~{td.strftime('%b %d')})"
        ws.cell(row=r, column=ci, value=label).alignment = ALIGN_C
    ytd_col = len(months) + 2
    ws.cell(row=r, column=ytd_col, value="YTD").alignment = ALIGN_C

    r = 3
    for section_title, rows_data in sections:
        if section_title:
            for c in range(1, ytd_col + 1):
                ws.cell(row=r, column=c).fill = FILL_SECTION
            ws.cell(row=r, column=1, value=section_title).font = FONT_B
            r += 1

        for row_label, month_vals, ytd_val, fmt in rows_data:
            is_total = row_label.strip().startswith("Total") or row_label == "Total"
            ws.cell(row=r, column=1, value=row_label)
            if is_total:
                ws.cell(row=r, column=1).font = FONT_B
            elif row_label.startswith("  "):
                ws.cell(row=r, column=1).font = FONT_IT

            for ci, m in enumerate(months, 2):
                val = month_vals.get(m, 0)
                cell = ws.cell(row=r, column=ci, value=val)
                cell.number_format = fmt
                if is_total:
                    cell.fill = FILL_TOTAL
                    cell.font = FONT_B

            cell = ws.cell(row=r, column=ytd_col, value=ytd_val)
            cell.number_format = fmt
            if is_total:
                cell.fill = FILL_TOTAL
                cell.font = FONT_B
            r += 1

        r += 1  # blank row between sections

    # Column widths
    ws.column_dimensions["A"].width = 30
    for ci in range(2, ytd_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14
    ws.freeze_panes = ws.cell(row=3, column=2)
    return ws


def update_sales_summary(wb, through_date):
    """Sales Summary tab: Net Revenue by Channel + Brand, Channel Mix %, Brand Mix %."""
    months, ytd_year = _summary_month_range(through_date)

    shopify = load_dk("shopify_orders_daily", days=800)
    amz_sales = load_dk("amazon_sales_daily", days=800)

    # Aggregate shopify by channel x month
    ch_month = defaultdict(lambda: defaultdict(float))  # channel -> month -> net
    br_month = defaultdict(lambda: defaultdict(float))  # brand -> month -> net
    total_month = defaultdict(float)

    CHANNEL_MAP = {"D2C": "Onzenna", "PR": None}
    for r in shopify:
        d = r.get("date", "")
        if d > through_date:
            continue
        m = d[:7]
        if m not in months:
            continue
        ch_raw = r.get("channel", "Unknown")
        if ch_raw == "PR":
            continue
        ch = CHANNEL_MAP.get(ch_raw, ch_raw)
        net = float(r.get("net_sales", 0) or 0)
        brand = r.get("brand", "Unknown")
        ch_month[ch][m] += net
        br_month[brand][m] += net
        total_month[m] += net

    # Amazon Marketplace net (SP-API) — add to Amazon channel and brand totals
    for r in amz_sales:
        d = r.get("date", "")
        if d > through_date:
            continue
        m = d[:7]
        if m not in months:
            continue
        net = float(r.get("net_sales", 0) or r.get("gross_sales", 0) or 0)
        brand = r.get("brand", "Unknown")
        ch_month["Amazon"][m] += net
        br_month[brand][m] += net
        total_month[m] += net

    def _ytd(d):
        return sum(v for m, v in d.items() if m[:4] == ytd_year)

    # Build sections
    sections = []

    # 1) Total Net Revenue
    total_row = ("Total", total_month, _ytd(total_month), '#,##0')
    sections.append(("TOTAL NET REVENUE", [total_row]))

    # 2) Net Revenue by Channel
    ch_order = ["Amazon", "Onzenna", "TargetPlus", "B2B", "TikTok"]
    ch_rows = []
    for ch in ch_order:
        if ch in ch_month:
            ch_rows.append((ch, ch_month[ch], _ytd(ch_month[ch]), '#,##0'))
    # any others
    for ch in sorted(ch_month):
        if ch not in ch_order:
            ch_rows.append((ch, ch_month[ch], _ytd(ch_month[ch]), '#,##0'))
    sections.append(("NET REVENUE BY CHANNEL", ch_rows))

    # 3) Channel Mix %
    mix_rows = []
    for label, mv, _, _ in ch_rows:
        pct = {m: round(mv.get(m, 0) / total_month[m], 4) if total_month.get(m) else 0 for m in months}
        ytd_t = _ytd(total_month)
        ytd_p = round(_ytd(mv) / ytd_t, 4) if ytd_t else 0
        mix_rows.append((label, pct, ytd_p, '0.0%'))
    sections.append(("CHANNEL MIX % (of Net Revenue)", mix_rows))

    # 4) Net Revenue by Brand
    br_order = ["Grosmimi", "Naeiae", "CHA&MOM", "Onzenna", "Alpremio"]
    br_rows = []
    for br in br_order:
        if br in br_month:
            br_rows.append((br, br_month[br], _ytd(br_month[br]), '#,##0'))
    other_net = defaultdict(float)
    for br in br_month:
        if br not in br_order:
            for m, v in br_month[br].items():
                other_net[m] += v
    if any(other_net.values()):
        br_rows.append(("Other", other_net, _ytd(other_net), '#,##0'))
    sections.append(("NET REVENUE BY BRAND", br_rows))

    # 5) Brand Mix %
    bmix_rows = []
    for label, _, _, _ in br_rows:
        mv = br_month.get(label, other_net if label == "Other" else {})
        pct = {m: round(mv.get(m, 0) / total_month[m], 4) if total_month.get(m) else 0 for m in months}
        ytd_t = _ytd(total_month)
        ytd_p = round(_ytd(mv) / ytd_t, 4) if ytd_t else 0
        bmix_rows.append((label, pct, ytd_p, '0.0%'))
    sections.append(("BRAND MIX % (of Net Revenue)", bmix_rows))

    _write_summary_tab(wb, "Sales Summary", "SALES SUMMARY", sections, months, ytd_year, through_date)
    print(f"  -> Sales Summary updated ({len(months)} months)")


def update_ads_summary(wb, through_date):
    """Ads Summary tab: Blended metrics, Spend/ROAS by Platform, Mix %."""
    months, ytd_year = _summary_month_range(through_date)

    # Load ad data
    amz_ads = load_dk("amazon_ads_daily", days=800)
    meta_ads = load_dk("meta_ads_daily", days=800)
    google_ads = load_dk("google_ads_daily", days=800)

    # Aggregate by platform x month: spend, revenue, clicks, impressions
    plat = defaultdict(lambda: defaultdict(lambda: {"spend": 0.0, "revenue": 0.0, "clicks": 0, "impressions": 0}))

    for r in amz_ads:
        d = r.get("date", "")
        if d > through_date:
            continue
        m = d[:7]
        if m in months:
            plat["Amazon Ads"][m]["spend"] += float(r.get("spend", 0) or 0)
            plat["Amazon Ads"][m]["revenue"] += float(r.get("sales", 0) or r.get("revenue", 0) or 0)
            plat["Amazon Ads"][m]["clicks"] += int(r.get("clicks", 0) or 0)
            plat["Amazon Ads"][m]["impressions"] += int(r.get("impressions", 0) or 0)

    for r in meta_ads:
        d = r.get("date", "")
        if d > through_date:
            continue
        m = d[:7]
        if m in months:
            plat["Facebook Ads"][m]["spend"] += float(r.get("spend", 0) or 0)
            plat["Facebook Ads"][m]["revenue"] += float(r.get("purchase_value", 0) or r.get("revenue", 0) or 0)
            plat["Facebook Ads"][m]["clicks"] += int(r.get("clicks", 0) or r.get("link_clicks", 0) or 0)
            plat["Facebook Ads"][m]["impressions"] += int(r.get("impressions", 0) or 0)

    for r in google_ads:
        d = r.get("date", "")
        if d > through_date:
            continue
        m = d[:7]
        if m in months:
            plat["Google Ads"][m]["spend"] += float(r.get("spend", 0) or r.get("cost", 0) or 0)
            plat["Google Ads"][m]["revenue"] += float(r.get("conversions_value", 0) or r.get("revenue", 0) or 0)
            plat["Google Ads"][m]["clicks"] += int(r.get("clicks", 0) or 0)
            plat["Google Ads"][m]["impressions"] += int(r.get("impressions", 0) or 0)

    # Also load total orders for CAC
    shopify = load_dk("shopify_orders_daily", days=800)
    orders_month = defaultdict(int)
    net_month = defaultdict(float)
    for r in shopify:
        d = r.get("date", "")
        if d > through_date or r.get("channel") == "PR":
            continue
        m = d[:7]
        if m in months:
            orders_month[m] += int(r.get("orders", 0) or 0)
            net_month[m] += float(r.get("net_sales", 0) or 0)

    def _ytd(d):
        return sum(v for m, v in d.items() if m[:4] == ytd_year)

    # Totals
    total_spend = {m: sum(plat[p][m]["spend"] for p in plat) for m in months}
    total_rev = {m: sum(plat[p][m]["revenue"] for p in plat) for m in months}
    total_clicks = {m: sum(plat[p][m]["clicks"] for p in plat) for m in months}

    sections = []

    # 1) Blended Advertising Metrics
    blended = []
    blended.append(("Total Ad Spend", total_spend, _ytd(total_spend), '#,##0'))
    blended.append(("Ad-Attributed Revenue", total_rev, _ytd(total_rev), '#,##0'))

    roas_m = {m: round(total_rev[m] / total_spend[m], 2) if total_spend.get(m) else 0 for m in months}
    ytd_roas = round(_ytd(total_rev) / _ytd(total_spend), 2) if _ytd(total_spend) else 0
    blended.append(("Blended ROAS", roas_m, ytd_roas, '0.00'))

    spend_pct = {m: round(total_spend[m] / net_month[m], 4) if net_month.get(m) else 0 for m in months}
    ytd_sp = round(_ytd(total_spend) / _ytd(net_month), 4) if _ytd(net_month) else 0
    blended.append(("Ad Spend / Net Revenue", spend_pct, ytd_sp, '0.0%'))

    cac_m = {m: round(total_spend[m] / orders_month[m], 2) if orders_month.get(m) else 0 for m in months}
    ytd_cac = round(_ytd(total_spend) / sum(orders_month[m] for m in months if m[:4] == ytd_year), 2) if sum(orders_month[m] for m in months if m[:4] == ytd_year) else 0
    blended.append(("Blended CAC", cac_m, ytd_cac, '#,##0.00'))

    cpc_m = {m: round(total_spend[m] / total_clicks[m], 2) if total_clicks.get(m) else 0 for m in months}
    ytd_cpc = round(_ytd(total_spend) / _ytd(total_clicks), 2) if _ytd(total_clicks) else 0
    blended.append(("Avg CPC", cpc_m, ytd_cpc, '$#,##0.00'))
    sections.append(("BLENDED ADVERTISING METRICS", blended))

    # 2) Ad Spend by Platform
    plat_order = ["Amazon Ads", "Facebook Ads", "Google Ads", "TikTok Ads"]
    spend_rows = []
    for p in plat_order:
        if p in plat:
            s = {m: round(plat[p][m]["spend"], 2) for m in months}
            spend_rows.append((p, s, round(_ytd(s), 2), '#,##0'))
        else:
            spend_rows.append((p, {}, 0, '#,##0'))
    sections.append(("AD SPEND BY PLATFORM", spend_rows))

    # 3) ROAS by Platform
    roas_rows = []
    for p in plat_order:
        if p in plat:
            rm = {m: round(plat[p][m]["revenue"] / plat[p][m]["spend"], 2) if plat[p][m]["spend"] else 0 for m in months}
            ytd_s = sum(plat[p][m]["spend"] for m in months if m[:4] == ytd_year)
            ytd_r = sum(plat[p][m]["revenue"] for m in months if m[:4] == ytd_year)
            roas_rows.append((p, rm, round(ytd_r / ytd_s, 2) if ytd_s else 0, '0.00'))
        else:
            roas_rows.append((p, {}, 0, '0.00'))
    sections.append(("ROAS BY PLATFORM", roas_rows))

    # 4) Platform Spend Mix %
    mix_rows = []
    for p in plat_order:
        if p in plat:
            pct = {m: round(plat[p][m]["spend"] / total_spend[m], 4) if total_spend.get(m) else 0 for m in months}
            ytd_ts = _ytd(total_spend)
            ytd_ps = sum(plat[p][m]["spend"] for m in months if m[:4] == ytd_year)
            mix_rows.append((p, pct, round(ytd_ps / ytd_ts, 4) if ytd_ts else 0, '0.0%'))
        else:
            mix_rows.append((p, {}, 0, '0.0%'))
    sections.append(("PLATFORM SPEND MIX %", mix_rows))

    _write_summary_tab(wb, "Ads Summary", "ADS SUMMARY", sections, months, ytd_year, through_date)
    print(f"  -> Ads Summary updated ({len(months)} months)")


def update_unit_economics(wb, through_date):
    """Unit Economics tab: Per-order waterfall + margin stack + CAC by channel."""
    months, ytd_year = _summary_month_range(through_date)

    shopify = load_dk("shopify_orders_daily", days=800)
    brand_cogs = load_brand_avg_cogs()

    # Aggregate
    ch_month = defaultdict(lambda: defaultdict(lambda: {"net": 0.0, "orders": 0, "units": 0, "gross": 0.0}))
    total_month = defaultdict(lambda: {"net": 0.0, "orders": 0, "units": 0, "gross": 0.0})
    brand_month_u = defaultdict(lambda: defaultdict(lambda: {"units": 0, "gross": 0.0}))

    for r in shopify:
        d = r.get("date", "")
        if d > through_date or r.get("channel") == "PR":
            continue
        m = d[:7]
        if m not in months:
            continue
        ch = r.get("channel", "Unknown")
        if ch == "D2C":
            ch = "Onzenna"
        net = float(r.get("net_sales", 0) or 0)
        orders = int(r.get("orders", 0) or 0)
        units = int(r.get("units", 0) or 0)
        gross = float(r.get("gross_sales", 0) or 0)
        brand = r.get("brand", "Unknown")
        ch_month[ch][m]["net"] += net
        ch_month[ch][m]["orders"] += orders
        total_month[m]["net"] += net
        total_month[m]["orders"] += orders
        total_month[m]["units"] += units
        total_month[m]["gross"] += gross
        brand_month_u[brand][m]["units"] += units
        brand_month_u[brand][m]["gross"] += gross

    # COGS per order = sum(units_by_brand * avg_cogs_brand) / total_orders
    def _cogs_per_order(m):
        total_cogs = 0.0
        for br, md in brand_month_u.items():
            u = md[m]["units"]
            if u == 0 and md[m]["gross"] > 0:
                u = md[m]["gross"] / AVG_PRICE.get(br, 25.0)
            total_cogs += u * brand_cogs.get(br, AVG_COGS.get(br, 8.0))
        return total_cogs / total_month[m]["orders"] if total_month[m]["orders"] else 0

    # Load ad spend
    amz_ads = load_dk("amazon_ads_daily", days=800)
    meta_ads = load_dk("meta_ads_daily", days=800)
    google_ads = load_dk("google_ads_daily", days=800)

    ad_spend_month = defaultdict(float)
    ch_ad = defaultdict(lambda: defaultdict(float))  # "Amazon" -> month -> spend

    for r in amz_ads:
        d = r.get("date", "")
        if d > through_date:
            continue
        m = d[:7]
        if m in months:
            s = float(r.get("spend", 0) or 0)
            ad_spend_month[m] += s
            ch_ad["Amazon"][m] += s

    for r in meta_ads:
        d = r.get("date", "")
        if d > through_date:
            continue
        m = d[:7]
        if m in months:
            s = float(r.get("spend", 0) or 0)
            ad_spend_month[m] += s
            ch_ad["Onzenna"][m] += s  # Meta primarily drives Onzenna D2C

    for r in google_ads:
        d = r.get("date", "")
        if d > through_date:
            continue
        m = d[:7]
        if m in months:
            s = float(r.get("spend", 0) or 0)
            ad_spend_month[m] += s
            ch_ad["Onzenna"][m] += s  # Google primarily drives Onzenna D2C

    # Fees ~ 13% of net (Shopify + payment processing)
    FEE_RATE = 0.13

    def _ytd_sum(d):
        return sum(d[m] if isinstance(d[m], (int, float)) else d[m] for m in months if m[:4] == ytd_year and m in d)

    sections = []

    # 1) Per-Order Waterfall
    waterfall = []
    aov = {m: round(total_month[m]["net"] / total_month[m]["orders"], 2) if total_month[m]["orders"] else 0 for m in months}
    ytd_net = sum(total_month[m]["net"] for m in months if m[:4] == ytd_year)
    ytd_ord = sum(total_month[m]["orders"] for m in months if m[:4] == ytd_year)
    waterfall.append(("AOV (Net Revenue / Order)", aov, round(ytd_net / ytd_ord, 2) if ytd_ord else 0, '$#,##0.00'))

    cogs_po = {m: round(_cogs_per_order(m), 2) for m in months}
    ytd_cogs = sum(_cogs_per_order(m) * total_month[m]["orders"] for m in months if m[:4] == ytd_year)
    waterfall.append(("(-) COGS per Order", cogs_po, round(ytd_cogs / ytd_ord, 2) if ytd_ord else 0, '$#,##0.00'))

    gp_po = {m: round(aov.get(m, 0) - cogs_po.get(m, 0), 2) for m in months}
    waterfall.append(("= Gross Profit / Order", gp_po, round((ytd_net - ytd_cogs) / ytd_ord, 2) if ytd_ord else 0, '$#,##0.00'))

    fees_po = {m: round(total_month[m]["net"] * FEE_RATE / total_month[m]["orders"], 2) if total_month[m]["orders"] else 0 for m in months}
    ytd_fees = ytd_net * FEE_RATE
    waterfall.append(("(-) Fees per Order", fees_po, round(ytd_fees / ytd_ord, 2) if ytd_ord else 0, '$#,##0.00'))

    cac = {m: round(ad_spend_month[m] / total_month[m]["orders"], 2) if total_month[m]["orders"] else 0 for m in months}
    ytd_ads = sum(ad_spend_month[m] for m in months if m[:4] == ytd_year)
    waterfall.append(("(-) CAC (Ad Spend / Order)", cac, round(ytd_ads / ytd_ord, 2) if ytd_ord else 0, '$#,##0.00'))

    cm_po = {m: round(gp_po.get(m, 0) - fees_po.get(m, 0) - cac.get(m, 0), 2) for m in months}
    ytd_cm = (ytd_net - ytd_cogs - ytd_fees - ytd_ads)
    waterfall.append(("= CM / Order", cm_po, round(ytd_cm / ytd_ord, 2) if ytd_ord else 0, '$#,##0.00'))
    sections.append(("PER-ORDER WATERFALL (Blended)", waterfall))

    # 2) Margin Stack (% of Net Revenue)
    stack = []
    stack.append(("Gross Margin %", {m: round(1 - cogs_po.get(m, 0) / aov.get(m, 1), 4) for m in months},
                   round(1 - ytd_cogs / ytd_net, 4) if ytd_net else 0, '0.0%'))
    stack.append(("(-) Fee %", {m: round(FEE_RATE, 4) for m in months}, round(FEE_RATE, 4), '0.0%'))
    stack.append(("(-) Ad Spend %", {m: round(ad_spend_month[m] / total_month[m]["net"], 4) if total_month[m]["net"] else 0 for m in months},
                   round(ytd_ads / ytd_net, 4) if ytd_net else 0, '0.0%'))
    cm_pct = {m: round((total_month[m]["net"] - (_cogs_per_order(m) * total_month[m]["orders"]) - total_month[m]["net"] * FEE_RATE - ad_spend_month[m]) / total_month[m]["net"], 4) if total_month[m]["net"] else 0 for m in months}
    stack.append(("= CM %", cm_pct, round(ytd_cm / ytd_net, 4) if ytd_net else 0, '0.0%'))
    sections.append(("MARGIN STACK (% of Net Revenue)", stack))

    # 3) CAC by Channel
    ch_order = ["Amazon", "Onzenna", "TargetPlus", "B2B", "TikTok"]
    cac_rows = []
    for ch in ch_order:
        if ch in ch_month:
            cm = {m: round(ch_ad.get(ch, {}).get(m, 0) / ch_month[ch][m]["orders"], 2) if ch_month[ch][m]["orders"] else 0 for m in months}
            ytd_s = sum(ch_ad.get(ch, {}).get(m, 0) for m in months if m[:4] == ytd_year)
            ytd_o = sum(ch_month[ch][m]["orders"] for m in months if m[:4] == ytd_year)
            cac_rows.append((ch, cm, round(ytd_s / ytd_o, 2) if ytd_o else 0, '$#,##0.00'))
    sections.append(("CAC BY CHANNEL (Channel Ad Spend / Channel Orders)", cac_rows))

    _write_summary_tab(wb, "Unit Economics", "UNIT ECONOMICS", sections, months, ytd_year, through_date)
    print(f"  -> Unit Economics updated ({len(months)} months)")


def update_campaign_details(wb, through_date):
    """ADS Campaign Details tab: campaign-level monthly spend/sales."""
    import calendar
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    TAB = "ADS Campaign Details"
    months, ytd_year = _summary_month_range(through_date)

    amz = load_dk("amazon_ads_daily", days=800)
    meta = load_dk("meta_ads_daily", days=800)
    google = load_dk("google_ads_daily", days=800)

    # Aggregate: (platform, brand, campaign_name) -> month -> {spend, revenue, clicks}
    camp_data = defaultdict(lambda: defaultdict(lambda: {"spend": 0.0, "revenue": 0.0, "clicks": 0}))

    for r in amz:
        d = r.get("date", "")
        if d > through_date:
            continue
        m = d[:7]
        if m not in months:
            continue
        name = r.get("campaign_name") or r.get("name") or "Unknown"
        brand = r.get("brand", "Unknown")
        key = ("Amazon Ads", brand, name)
        camp_data[key][m]["spend"] += float(r.get("spend", 0) or 0)
        camp_data[key][m]["revenue"] += float(r.get("sales", 0) or r.get("revenue", 0) or 0)
        camp_data[key][m]["clicks"] += int(r.get("clicks", 0) or 0)

    for r in meta:
        d = r.get("date", "")
        if d > through_date:
            continue
        m = d[:7]
        if m not in months:
            continue
        name = r.get("campaign_name") or r.get("name") or "Unknown"
        brand = r.get("brand", "Unknown")
        key = ("Facebook Ads", brand, name)
        camp_data[key][m]["spend"] += float(r.get("spend", 0) or 0)
        camp_data[key][m]["revenue"] += float(r.get("purchase_value", 0) or r.get("revenue", 0) or 0)
        camp_data[key][m]["clicks"] += int(r.get("clicks", 0) or r.get("link_clicks", 0) or 0)

    for r in google:
        d = r.get("date", "")
        if d > through_date:
            continue
        m = d[:7]
        if m not in months:
            continue
        name = r.get("campaign_name") or r.get("name") or "Unknown"
        brand = r.get("brand", "Unknown")
        key = ("Google Ads", brand, name)
        camp_data[key][m]["spend"] += float(r.get("spend", 0) or r.get("cost", 0) or 0)
        camp_data[key][m]["revenue"] += float(r.get("conversions_value", 0) or r.get("revenue", 0) or 0)
        camp_data[key][m]["clicks"] += int(r.get("clicks", 0) or 0)

    if not camp_data:
        print(f"  [WARN] No campaign data, skipping {TAB}")
        return

    # Sort campaigns: platform > brand > total spend desc
    sorted_camps = sorted(camp_data.keys(),
                          key=lambda k: (k[0], k[1], -sum(camp_data[k][m]["spend"] for m in months)))

    # Write tab
    if TAB in wb.sheetnames:
        idx = wb.sheetnames.index(TAB)
        del wb[TAB]
        ws = wb.create_sheet(TAB, idx)
    else:
        ws = wb.create_sheet(TAB)

    FILL_HDR = PatternFill("solid", fgColor="002060")
    FILL_SECTION = PatternFill("solid", fgColor="D6DCE4")
    FILL_TOTAL = PatternFill("solid", fgColor="FFF2CC")
    FONT_W = Font(bold=True, color="FFFFFF")
    FONT_B = Font(bold=True)
    ALIGN_C = Alignment(horizontal="center", wrap_text=True)

    from datetime import date as _d
    td = _d.fromisoformat(through_date)

    # Metrics: Spend section, then Revenue section
    for section_idx, (section_label, field) in enumerate([("AD SPEND", "spend"), ("AD REVENUE", "revenue")]):
        start_row = 1 + section_idx * (len(sorted_camps) + 10)

        ws.cell(row=start_row, column=1, value=f"CAMPAIGN DETAIL — {section_label}").font = FONT_B
        r = start_row + 1

        # Header
        headers = ["Platform", "Brand", "Campaign"]
        for m in months:
            y, mo = int(m[:4]), int(m[5:7])
            label = f"{calendar.month_abbr[mo]} {y}"
            if y == td.year and mo == td.month:
                label += f"\n(~{td.strftime('%b %d')})"
            headers.append(label)
        headers.append("YTD")

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.fill = FILL_HDR
            cell.font = FONT_W
            cell.alignment = ALIGN_C
        r += 1

        # Total row
        ws.cell(row=r, column=1, value="TOTAL").font = FONT_B
        for ci, m in enumerate(months, 4):
            val = sum(camp_data[k][m][field] for k in sorted_camps)
            cell = ws.cell(row=r, column=ci, value=round(val, 0))
            cell.number_format = '#,##0'
            cell.fill = FILL_TOTAL
            cell.font = FONT_B
        ytd_val = sum(sum(camp_data[k][m][field] for m in months if m[:4] == ytd_year) for k in sorted_camps)
        cell = ws.cell(row=r, column=len(months) + 4, value=round(ytd_val, 0))
        cell.number_format = '#,##0'
        cell.fill = FILL_TOTAL
        cell.font = FONT_B
        r += 1

        # Campaign rows
        for plat, brand, name in sorted_camps:
            ws.cell(row=r, column=1, value=plat)
            ws.cell(row=r, column=2, value=brand)
            ws.cell(row=r, column=3, value=name)
            ytd = 0
            for ci, m in enumerate(months, 4):
                val = round(camp_data[(plat, brand, name)][m][field], 0)
                ws.cell(row=r, column=ci, value=val).number_format = '#,##0'
                if m[:4] == ytd_year:
                    ytd += val
            ws.cell(row=r, column=len(months) + 4, value=ytd).number_format = '#,##0'
            r += 1

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 40
    for ci in range(4, len(months) + 5):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.freeze_panes = ws.cell(row=3, column=4)
    print(f"  -> {TAB} updated ({len(sorted_camps)} campaigns, {len(months)} months)")


def update_influencer_dashboard(wb, through_date):
    """Influencer Dashboard tab: PR orders monthly by brand x product."""
    import calendar
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import date as _d

    TAB = "Influencer Dashboard"

    shopify = load_dk("shopify_orders_daily", days=800)
    # PR orders
    pr_month = defaultdict(lambda: defaultdict(int))  # brand -> month -> units
    pr_total = defaultdict(int)  # month -> units

    for r in shopify:
        d = r.get("date", "")
        if d > through_date:
            continue
        if r.get("channel") != "PR":
            continue
        m = d[:7]
        brand = r.get("brand", "Unknown")
        units = int(r.get("units", 0) or r.get("orders", 0) or 0)
        pr_month[brand][m] += units
        pr_total[m] += units

    if not pr_total:
        print(f"  [WARN] No PR order data, skipping {TAB}")
        return

    all_months = sorted(pr_total.keys())

    if TAB in wb.sheetnames:
        idx = wb.sheetnames.index(TAB)
        del wb[TAB]
        ws = wb.create_sheet(TAB, idx)
    else:
        ws = wb.create_sheet(TAB)

    FILL_HDR = PatternFill("solid", fgColor="002060")
    FILL_SECTION = PatternFill("solid", fgColor="D6DCE4")
    FONT_W = Font(bold=True, color="FFFFFF")
    FONT_B = Font(bold=True)
    ALIGN_C = Alignment(horizontal="center", wrap_text=True)

    td = _d.fromisoformat(through_date)

    # Section A: Monthly shipped orders
    ws.cell(row=1, column=1, value="SECTION A — PR ORDERS: MONTHLY SUMMARY").font = FONT_B
    r = 2
    headers = ["Metric"] + [f"{m[2:4]}/{m[5:7]}" for m in all_months]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=r, column=ci + 1, value=h if ci > 1 else None)
    ws.cell(row=r, column=2, value="Month")
    r = 3

    ws.cell(row=r, column=2, value="Shipped Orders")
    for ci, m in enumerate(all_months, 3):
        ws.cell(row=r, column=ci, value=pr_total[m]).number_format = '#,##0'
    r += 2

    # Section B: By brand
    ws.cell(row=r, column=1, value="SECTION B — PR ORDERS BY BRAND").font = FONT_B
    r += 1
    headers2 = ["Brand"] + [f"{m[2:4]}/{m[5:7]}" for m in all_months]
    for ci, h in enumerate(headers2, 2):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.fill = FILL_HDR
        cell.font = FONT_W
        cell.alignment = ALIGN_C
    r += 1

    # Total row
    ws.cell(row=r, column=2, value="TOTAL — All Brands").font = FONT_B
    for ci, m in enumerate(all_months, 3):
        ws.cell(row=r, column=ci, value=pr_total[m]).number_format = '#,##0'
    r += 1

    # Brand rows
    for brand in BRAND_ORDER:
        if brand in pr_month:
            ws.cell(row=r, column=2, value=f"  {brand}")
            for ci, m in enumerate(all_months, 3):
                ws.cell(row=r, column=ci, value=pr_month[brand].get(m, 0)).number_format = '#,##0'
            r += 1

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    for ci in range(3, len(all_months) + 3):
        ws.column_dimensions[get_column_letter(ci)].width = 8

    print(f"  -> {TAB} updated ({len(all_months)} months, {len(pr_month)} brands)")


# ── LEGACY TAB UPDATERS (fill March 2026+ from PG) ───────────────────────────

CHANNEL_MAP_PG_TO_LEGACY = {
    "D2C": "Onzenna",
    "B2B": "B2B",
    "TargetPlus": "TargetPlus",
    "TikTok": "TikTokShop",
    "TikTokShop": "TikTokShop",
    "PR": "PR",
    "Amazon": "Amazon",
}


def _find_partial_month_col(ws, through_date):
    """Find the partial-month data column (e.g. 'Mar 7\\n2026') in the FIRST panel.
    Returns the column number for the partial-period data column (typically col 32).
    Only searches the first panel (cols 5-40) to avoid matching repeated headers.
    """
    import re
    from datetime import date as _d
    td = _d.fromisoformat(through_date)
    # Only search first panel (cols 5 ~ 40 covers Jan 2024 through partial+YTD+%)
    for c in range(5, min(ws.max_column + 1, 45)):
        hdr = ws.cell(3, c).value
        if not hdr or "\n" not in str(hdr):
            continue
        hdr_s = str(hdr)
        if str(td.year) not in hdr_s:
            continue
        parts = hdr_s.split("\n")
        if len(parts) != 2:
            continue
        first = parts[0].strip()
        # Match "Mar 7", "Mar 11", "Jan 1-15" — month + day (1-31)
        # Reject "Mar 2026" — month + year (4 digits)
        m = re.match(r'^[A-Z][a-z]{2}\s+(\d{1,2})(?:\D|$)', first)
        if m and int(m.group(1)) <= 31:
            return c
    return 32  # default fallback


def _find_month_col(ws, month_str):
    """Find column for a month like 'Jan 2026'. Returns col number or None."""
    for c in range(5, ws.max_column + 1):
        hdr = ws.cell(3, c).value
        if hdr and str(hdr).strip() == month_str:
            return c
    return None


def _scan_legacy_sales_rows(ws):
    """Scan Sales tab section 1A to build (metric, channel, brand, 'Other') -> row map."""
    METRICS = ('GROSS SALES', 'DISCOUNTS', '# OF ORDERS', 'NET SALES')
    row_map = {}  # (metric, channel, brand) -> row of "Other" product
    current_metric = None
    current_channel = None
    current_brand = None

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value

        if a and 'SECTION 1B' in str(a):
            break
        if a and str(a) in METRICS:
            current_metric = str(a)
            current_channel = None
            current_brand = None
            continue
        if current_metric is None:
            continue
        if b and str(b) not in ('Channel',) and c is None and d is None:
            current_channel = str(b)
            current_brand = None
            continue
        if c and b is None and d is None:
            current_brand = str(c)
            continue
        if d and str(d) == 'Other' and b is None and c is None:
            if current_metric and current_channel and current_brand:
                row_map[(current_metric, current_channel, current_brand)] = r

    return row_map


def _scan_legacy_cm_rows(ws):
    """Scan CM tab to build (metric, channel, brand) -> row map for brand-level value rows."""
    METRICS = ('NET SALES', 'COGS', 'AD SPEND')
    row_map = {}
    current_metric = None
    current_channel = None

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value

        if a and str(a).split('(')[0].strip() in METRICS:
            current_metric = str(a).split('(')[0].strip()
            current_channel = None
            continue
        if a and str(a) not in METRICS and current_metric:
            # Next section with different name
            for m in ('GROSS PROFIT', 'CHANNEL FEES', 'CM BEFORE ADS', 'CM AFTER ADS'):
                if m in str(a):
                    current_metric = None
                    break
            continue

        if current_metric is None:
            continue
        if b and str(b) not in ('Channel',) and c is None:
            current_channel = str(b)
            continue
        if c and b is None:
            brand = str(c)
            if current_channel:
                row_map[(current_metric, current_channel, brand)] = r

    return row_map


def _scan_legacy_organic_rows(ws):
    """Scan Organic Sales tab: (metric, channel, brand) -> row for brand-level value rows."""
    METRICS = ('TOTAL REVENUE (Net Sales)', 'AD REVENUE (Attributed)')
    row_map = {}
    current_metric = None
    current_channel = None

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value

        if a:
            a_str = str(a).strip()
            for m in METRICS:
                if a_str.startswith(m.split('(')[0].strip()):
                    current_metric = m
                    current_channel = None
                    break
            else:
                if 'ORGANIC' in a_str:
                    current_metric = None
            continue

        if current_metric is None:
            continue
        if b and str(b) not in ('Channel',) and c is None:
            current_channel = str(b)
            continue
        if c and b is None:
            brand = str(c)
            if current_channel:
                row_map[(current_metric, current_channel, brand)] = r

    return row_map


def _scan_legacy_ads_rows(ws):
    """Scan Ads tab section 2A: (metric, platform, camp_type, brand) -> row."""
    METRICS = ('AD SPEND', 'AD REVENUE', 'CLICKS')
    row_map = {}
    current_metric = None
    current_platform = None
    current_camp_type = None

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value

        if a and 'SECTION 2B' in str(a):
            break
        if a and str(a) in METRICS:
            current_metric = str(a)
            current_platform = None
            current_camp_type = None
            continue
        if a and str(a) in ('ROAS (Revenue / Spend)', 'CPC (Spend / Clicks)'):
            current_metric = None  # these are formula sections
            continue

        if current_metric is None:
            continue
        if b and str(b) not in ('Platform', 'TOTAL') and c is None and d is None:
            current_platform = str(b)
            current_camp_type = None
            continue
        if c and b is None and d is None:
            current_camp_type = str(c)
            continue
        if d and b is None and c is None:
            brand = str(d)
            if current_platform and current_camp_type:
                row_map[(current_metric, current_platform, current_camp_type, brand)] = r

    return row_map


def update_legacy_sales(wb, through_date):
    """Fill March 2026+ data in legacy Sales tab from PG shopify_orders_daily."""
    from datetime import date as _d
    td = _d.fromisoformat(through_date)

    if "Sales" not in wb.sheetnames:
        print("  [SKIP] No 'Sales' tab found")
        return

    ws = wb["Sales"]
    row_map = _scan_legacy_sales_rows(ws)
    if not row_map:
        print("  [WARN] Could not scan Sales tab structure")
        return

    # Find the partial month column
    partial_col = _find_partial_month_col(ws, through_date)

    # Query PG for current month
    shopify = load_dk("shopify_orders_daily", days=45)
    current_month = td.strftime("%Y-%m")

    # Aggregate by channel x brand for current month
    agg = defaultdict(lambda: {"gross": 0.0, "disc": 0.0, "orders": 0, "net": 0.0})
    for r in shopify:
        d = r.get("date", "")
        if d > through_date:
            continue
        if d[:7] != current_month:
            continue
        ch_raw = r.get("channel", "Unknown")
        ch = CHANNEL_MAP_PG_TO_LEGACY.get(ch_raw, ch_raw)
        brand = r.get("brand", "Unknown")
        agg[(ch, brand)]["gross"] += float(r.get("gross_sales", 0) or 0)
        agg[(ch, brand)]["disc"] += float(r.get("discounts", 0) or 0)
        agg[(ch, brand)]["orders"] += int(r.get("orders", 0) or 0)
        agg[(ch, brand)]["net"] += float(r.get("net_sales", 0) or 0)

    METRIC_FIELD = {
        "GROSS SALES": "gross",
        "DISCOUNTS": "disc",
        "# OF ORDERS": "orders",
        "NET SALES": "net",
    }

    filled = 0
    for (metric, channel, brand), row in row_map.items():
        field = METRIC_FIELD.get(metric)
        if not field:
            continue
        key = (channel, brand)
        if key in agg:
            val = agg[key][field]
            if val:
                ws.cell(row=row, column=partial_col, value=round(val, 2) if isinstance(val, float) else val)
                filled += 1

    # Update partial column header
    import calendar
    month_abbr = calendar.month_abbr[td.month]
    ws.cell(row=3, column=partial_col,
            value=f"{month_abbr} 1-{td.day}\n{td.year}")

    # Update full-month projection formula factor in the column BEFORE partial
    days_elapsed = td.day
    days_in_month = calendar.monthrange(td.year, td.month)[1]
    full_mo_col = partial_col - 1
    from openpyxl.utils import get_column_letter
    partial_letter = get_column_letter(partial_col)

    for r in range(4, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=full_mo_col).value
        if cell_val and isinstance(cell_val, str) and f"={partial_letter}" in cell_val:
            # Replace any *N/M factor with *days_in_month/days_elapsed
            import re as _re
            new_formula = _re.sub(r'\*\d+/\d+', f'*{days_in_month}/{days_elapsed}', str(cell_val))
            ws.cell(row=r, column=full_mo_col, value=new_formula)

    print(f"  -> Sales tab updated ({filled} cells filled for {current_month}, {days_elapsed} days)")


def update_legacy_cm(wb, through_date):
    """Fill March 2026+ data in legacy CM tab from PG."""
    from datetime import date as _d
    td = _d.fromisoformat(through_date)

    if "CM" not in wb.sheetnames:
        print("  [SKIP] No 'CM' tab found")
        return

    ws = wb["CM"]
    row_map = _scan_legacy_cm_rows(ws)
    if not row_map:
        print("  [WARN] Could not scan CM tab structure")
        return

    partial_col = _find_partial_month_col(ws, through_date)
    current_month = td.strftime("%Y-%m")

    # Load shopify + amazon sales for NET SALES
    shopify = load_dk("shopify_orders_daily", days=45)
    amz_sales = load_dk("amazon_sales_daily", days=45)

    net_agg = defaultdict(lambda: defaultdict(float))  # channel -> brand -> net
    for r in shopify:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month:
            continue
        ch_raw = r.get("channel", "Unknown")
        if ch_raw == "PR":
            continue
        ch = CHANNEL_MAP_PG_TO_LEGACY.get(ch_raw, ch_raw)
        brand = r.get("brand", "Unknown")
        net_agg[ch][brand] += float(r.get("net_sales", 0) or 0)

    for r in amz_sales:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month:
            continue
        brand = r.get("brand", "Unknown")
        net_agg["Amazon"][brand] += float(r.get("net_sales", 0) or r.get("gross_sales", 0) or 0)

    # Load ads for AD SPEND
    amz_ads = load_dk("amazon_ads_daily", days=45)
    meta_ads = load_dk("meta_ads_daily", days=45)
    google_ads = load_dk("google_ads_daily", days=45)

    ad_agg = defaultdict(lambda: defaultdict(float))  # channel -> brand -> spend
    for r in amz_ads:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month:
            continue
        brand = r.get("brand", "Unknown")
        ad_agg["Amazon"][brand] += float(r.get("spend", 0) or 0)

    for r in meta_ads:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month:
            continue
        brand = r.get("brand", "Unknown")
        ad_agg["Onzenna"][brand] += float(r.get("spend", 0) or 0)

    for r in google_ads:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month:
            continue
        brand = r.get("brand", "Unknown")
        ad_agg["Onzenna"][brand] += float(r.get("spend", 0) or r.get("cost", 0) or 0)

    # Fill NET SALES rows
    filled = 0
    for (metric, channel, brand), row in row_map.items():
        if metric == "NET SALES":
            val = net_agg.get(channel, {}).get(brand, 0)
            if val:
                ws.cell(row=row, column=partial_col, value=round(val, 2))
                filled += 1
        elif metric == "AD SPEND":
            val = ad_agg.get(channel, {}).get(brand, 0)
            if val:
                ws.cell(row=row, column=partial_col, value=round(val, 2))
                filled += 1

    # Update header
    import calendar
    month_abbr = calendar.month_abbr[td.month]
    ws.cell(row=3, column=partial_col,
            value=f"{month_abbr} 1-{td.day}\n{td.year}")

    # Update full-month projection factor
    days_elapsed = td.day
    days_in_month = calendar.monthrange(td.year, td.month)[1]
    full_mo_col = partial_col - 1
    from openpyxl.utils import get_column_letter
    partial_letter = get_column_letter(partial_col)
    import re as _re
    for r in range(4, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=full_mo_col).value
        if cell_val and isinstance(cell_val, str) and f"={partial_letter}" in cell_val:
            new_formula = _re.sub(r'\*\d+/\d+', f'*{days_in_month}/{days_elapsed}', str(cell_val))
            ws.cell(row=r, column=full_mo_col, value=new_formula)

    print(f"  -> CM tab updated ({filled} cells filled for {current_month})")


def update_legacy_organic(wb, through_date):
    """Fill March 2026+ data in legacy Organic Sales tab from PG."""
    from datetime import date as _d
    td = _d.fromisoformat(through_date)

    if "Organic Sales" not in wb.sheetnames:
        print("  [SKIP] No 'Organic Sales' tab found")
        return

    ws = wb["Organic Sales"]
    row_map = _scan_legacy_organic_rows(ws)
    if not row_map:
        print("  [WARN] Could not scan Organic Sales tab structure")
        return

    partial_col = _find_partial_month_col(ws, through_date)
    current_month = td.strftime("%Y-%m")

    # TOTAL REVENUE = shopify net + amazon net
    shopify = load_dk("shopify_orders_daily", days=45)
    amz_sales = load_dk("amazon_sales_daily", days=45)

    rev_agg = defaultdict(lambda: defaultdict(float))
    for r in shopify:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month or r.get("channel") == "PR":
            continue
        ch = CHANNEL_MAP_PG_TO_LEGACY.get(r.get("channel", ""), r.get("channel", ""))
        brand = r.get("brand", "Unknown")
        rev_agg[ch][brand] += float(r.get("net_sales", 0) or 0)

    for r in amz_sales:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month:
            continue
        brand = r.get("brand", "Unknown")
        rev_agg["Amazon"][brand] += float(r.get("net_sales", 0) or r.get("gross_sales", 0) or 0)

    # AD REVENUE = ad-attributed revenue by channel
    amz_ads = load_dk("amazon_ads_daily", days=45)
    meta_ads = load_dk("meta_ads_daily", days=45)
    google_ads = load_dk("google_ads_daily", days=45)

    ad_rev_agg = defaultdict(lambda: defaultdict(float))
    for r in amz_ads:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month:
            continue
        brand = r.get("brand", "Unknown")
        ad_rev_agg["Amazon"][brand] += float(r.get("sales", 0) or r.get("revenue", 0) or 0)

    for r in meta_ads:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month:
            continue
        brand = r.get("brand", "Unknown")
        ad_rev_agg["Onzenna"][brand] += float(r.get("purchase_value", 0) or r.get("revenue", 0) or 0)

    for r in google_ads:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month:
            continue
        brand = r.get("brand", "Unknown")
        ad_rev_agg["Onzenna"][brand] += float(r.get("conversions_value", 0) or r.get("revenue", 0) or 0)

    filled = 0
    for (metric, channel, brand), row in row_map.items():
        if "TOTAL REVENUE" in metric:
            val = rev_agg.get(channel, {}).get(brand, 0)
        elif "AD REVENUE" in metric:
            val = ad_rev_agg.get(channel, {}).get(brand, 0)
        else:
            continue
        if val:
            ws.cell(row=row, column=partial_col, value=round(val, 2))
            filled += 1

    import calendar
    month_abbr = calendar.month_abbr[td.month]
    ws.cell(row=3, column=partial_col, value=f"{month_abbr} 1-{td.day}\n{td.year}")

    days_elapsed = td.day
    days_in_month = calendar.monthrange(td.year, td.month)[1]
    full_mo_col = partial_col - 1
    from openpyxl.utils import get_column_letter
    partial_letter = get_column_letter(partial_col)
    import re as _re
    for r in range(4, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=full_mo_col).value
        if cell_val and isinstance(cell_val, str) and f"={partial_letter}" in cell_val:
            new_formula = _re.sub(r'\*\d+/\d+', f'*{days_in_month}/{days_elapsed}', str(cell_val))
            ws.cell(row=r, column=full_mo_col, value=new_formula)

    print(f"  -> Organic Sales tab updated ({filled} cells filled for {current_month})")


def update_legacy_ads(wb, through_date):
    """Fill March 2026+ data in legacy Ads tab from PG campaign data."""
    from datetime import date as _d
    td = _d.fromisoformat(through_date)

    if "Ads" not in wb.sheetnames:
        print("  [SKIP] No 'Ads' tab found")
        return

    ws = wb["Ads"]
    partial_col = _find_partial_month_col(ws, through_date)
    current_month = td.strftime("%Y-%m")

    # Load all ads data
    amz_ads = load_dk("amazon_ads_daily", days=45)
    meta_ads = load_dk("meta_ads_daily", days=45)
    google_ads = load_dk("google_ads_daily", days=45)

    # Aggregate: platform -> brand -> {spend, revenue, clicks}
    plat_brand = defaultdict(lambda: defaultdict(lambda: {"spend": 0.0, "revenue": 0.0, "clicks": 0}))

    for r in amz_ads:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month:
            continue
        brand = r.get("brand", "Unknown")
        ad_type = r.get("ad_type", "SP")
        plat_brand[("Amazon Ads", ad_type)][brand]["spend"] += float(r.get("spend", 0) or 0)
        plat_brand[("Amazon Ads", ad_type)][brand]["revenue"] += float(r.get("sales", 0) or r.get("revenue", 0) or 0)
        plat_brand[("Amazon Ads", ad_type)][brand]["clicks"] += int(r.get("clicks", 0) or 0)

    for r in meta_ads:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month:
            continue
        brand = r.get("brand", "Unknown")
        plat_brand[("Facebook Ads", "CVR")][brand]["spend"] += float(r.get("spend", 0) or 0)
        plat_brand[("Facebook Ads", "CVR")][brand]["revenue"] += float(r.get("purchase_value", 0) or r.get("revenue", 0) or 0)
        plat_brand[("Facebook Ads", "CVR")][brand]["clicks"] += int(r.get("clicks", 0) or r.get("link_clicks", 0) or 0)

    for r in google_ads:
        d = r.get("date", "")
        if d > through_date or d[:7] != current_month:
            continue
        brand = r.get("brand", "Unknown")
        plat_brand[("Google Ads", "CVR")][brand]["spend"] += float(r.get("spend", 0) or r.get("cost", 0) or 0)
        plat_brand[("Google Ads", "CVR")][brand]["revenue"] += float(r.get("conversions_value", 0) or r.get("revenue", 0) or 0)
        plat_brand[("Google Ads", "CVR")][brand]["clicks"] += int(r.get("clicks", 0) or 0)

    # Scan Ads tab row structure and fill
    row_map = _scan_legacy_ads_rows(ws)

    METRIC_FIELD = {
        "AD SPEND": "spend",
        "AD REVENUE": "revenue",
        "CLICKS": "clicks",
    }

    filled = 0
    for (metric, platform, camp_type, brand), row in row_map.items():
        field = METRIC_FIELD.get(metric)
        if not field:
            continue
        key = (platform, camp_type)
        val = plat_brand.get(key, {}).get(brand, {}).get(field, 0)
        if val:
            ws.cell(row=row, column=partial_col, value=round(val, 2) if isinstance(val, float) else val)
            filled += 1

    import calendar
    month_abbr = calendar.month_abbr[td.month]
    ws.cell(row=3, column=partial_col, value=f"{month_abbr} 1-{td.day}\n{td.year}")

    days_elapsed = td.day
    days_in_month = calendar.monthrange(td.year, td.month)[1]
    full_mo_col = partial_col - 1
    from openpyxl.utils import get_column_letter
    partial_letter = get_column_letter(partial_col)
    import re as _re
    for r in range(4, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=full_mo_col).value
        if cell_val and isinstance(cell_val, str) and f"={partial_letter}" in cell_val:
            new_formula = _re.sub(r'\*\d+/\d+', f'*{days_in_month}/{days_elapsed}', str(cell_val))
            ws.cell(row=r, column=full_mo_col, value=new_formula)

    print(f"  -> Ads tab updated ({filled} cells filled for {current_month})")


def update_legacy_summary(wb, through_date):
    """Rebuild Summary tab with campaign performance rankings from PG."""
    from datetime import date as _d, timedelta
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    import calendar

    td = _d.fromisoformat(through_date)
    TAB = "Summary"

    if TAB not in wb.sheetnames:
        print(f"  [SKIP] No '{TAB}' tab found")
        return

    # Load last 90 days of campaign data
    d90 = (td - timedelta(days=90)).isoformat()
    amz_ads = load_dk("amazon_ads_daily", days=100)
    meta_ads = load_dk("meta_ads_daily", days=100)
    google_ads = load_dk("google_ads_daily", days=100)

    # Aggregate by campaign
    camps = defaultdict(lambda: {"spend": 0.0, "revenue": 0.0, "clicks": 0, "impressions": 0, "purchases": 0})

    for r in amz_ads:
        d = r.get("date", "")
        if d < d90 or d > through_date:
            continue
        name = r.get("campaign_name") or "Unknown"
        brand = r.get("brand", "Unknown")
        ad_type = r.get("ad_type", "SP")
        key = ("Amazon Ads", brand, name, ad_type)
        camps[key]["spend"] += float(r.get("spend", 0) or 0)
        camps[key]["revenue"] += float(r.get("sales", 0) or 0)
        camps[key]["clicks"] += int(r.get("clicks", 0) or 0)
        camps[key]["impressions"] += int(r.get("impressions", 0) or 0)
        camps[key]["purchases"] += int(r.get("purchases", 0) or 0)

    for r in meta_ads:
        d = r.get("date", "")
        if d < d90 or d > through_date:
            continue
        name = r.get("campaign_name") or "Unknown"
        brand = r.get("brand", "Unknown")
        key = ("Facebook Ads", brand, name, "CVR")
        camps[key]["spend"] += float(r.get("spend", 0) or 0)
        camps[key]["revenue"] += float(r.get("purchase_value", 0) or 0)
        camps[key]["clicks"] += int(r.get("clicks", 0) or r.get("link_clicks", 0) or 0)
        camps[key]["impressions"] += int(r.get("impressions", 0) or 0)
        camps[key]["purchases"] += int(r.get("purchases", 0) or 0)

    for r in google_ads:
        d = r.get("date", "")
        if d < d90 or d > through_date:
            continue
        name = r.get("campaign_name") or "Unknown"
        brand = r.get("brand", "Unknown")
        key = ("Google Ads", brand, name, "CVR")
        camps[key]["spend"] += float(r.get("spend", 0) or r.get("cost", 0) or 0)
        camps[key]["revenue"] += float(r.get("conversions_value", 0) or 0)
        camps[key]["clicks"] += int(r.get("clicks", 0) or 0)
        camps[key]["impressions"] += int(r.get("impressions", 0) or 0)

    if not camps:
        print(f"  [WARN] No campaign data for Summary tab")
        return

    # Filter campaigns with significant spend (> $10)
    active_camps = {k: v for k, v in camps.items() if v["spend"] >= 10}

    # Calculate derived metrics
    for k, v in active_camps.items():
        v["roas"] = round(v["revenue"] / v["spend"], 2) if v["spend"] else 0
        v["cpc"] = round(v["spend"] / v["clicks"], 2) if v["clicks"] else 0
        v["ctr"] = round(v["clicks"] / v["impressions"], 4) if v["impressions"] else 0

    # Rebuild tab
    idx = wb.sheetnames.index(TAB)
    del wb[TAB]
    ws = wb.create_sheet(TAB, idx)

    FILL_HDR = PatternFill("solid", fgColor="002060")
    FILL_SECTION = PatternFill("solid", fgColor="D6DCE4")
    FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
    FILL_RED = PatternFill("solid", fgColor="FFC7CE")
    FONT_W = Font(bold=True, color="FFFFFF")
    FONT_B = Font(bold=True)
    ALIGN_C = Alignment(horizontal="center")

    d90_fmt = _d.fromisoformat(d90).strftime("%b %Y")
    td_fmt = td.strftime("%b %Y")

    # Rankings to generate
    rankings = [
        ("Top 5 ROAS — CVR Campaigns", sorted(
            [(k, v) for k, v in active_camps.items() if k[3] in ("CVR",) and v["roas"] > 0],
            key=lambda x: -x[1]["roas"])[:5]),
        ("Top 5 ROAS — SP Campaigns", sorted(
            [(k, v) for k, v in active_camps.items() if k[3] == "SP" and v["roas"] > 0],
            key=lambda x: -x[1]["roas"])[:5]),
        ("Bottom 5 ROAS (>$50 spend)", sorted(
            [(k, v) for k, v in active_camps.items() if v["spend"] >= 50],
            key=lambda x: x[1]["roas"])[:5]),
        ("Top 10 Spend", sorted(
            active_camps.items(), key=lambda x: -x[1]["spend"])[:10]),
        ("Top 5 CTR", sorted(
            [(k, v) for k, v in active_camps.items() if v["impressions"] >= 100],
            key=lambda x: -x[1]["ctr"])[:5]),
    ]

    ws.cell(row=1, column=1, value="CAMPAIGN PERFORMANCE RANKINGS").font = FONT_B
    ws.cell(row=2, column=1, value=f"LAST 90 DAYS ({d90_fmt} – {td_fmt})").font = Font(italic=True, color="595959")

    r = 4
    headers = ["#", "Platform", "Campaign", "Brand", "Spend", "Revenue", "ROAS", "Clicks", "CPC", "CTR"]

    for title, camp_list in rankings:
        # Section title
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = FILL_SECTION
        ws.cell(row=r, column=1, value=title).font = FONT_B
        r += 1

        # Header
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.fill = FILL_HDR
            cell.font = FONT_W
            cell.alignment = ALIGN_C
        r += 1

        # Data rows
        for rank, (key, data) in enumerate(camp_list, 1):
            plat, brand, name, ad_type = key
            ws.cell(row=r, column=1, value=rank)
            ws.cell(row=r, column=2, value=plat)
            ws.cell(row=r, column=3, value=name[:50])
            ws.cell(row=r, column=4, value=brand)
            ws.cell(row=r, column=5, value=round(data["spend"], 2)).number_format = '#,##0.00'
            ws.cell(row=r, column=6, value=round(data["revenue"], 2)).number_format = '#,##0.00'
            ws.cell(row=r, column=7, value=data["roas"]).number_format = '0.00'
            ws.cell(row=r, column=8, value=data["clicks"]).number_format = '#,##0'
            ws.cell(row=r, column=9, value=data["cpc"]).number_format = '$#,##0.00'
            ws.cell(row=r, column=10, value=data["ctr"]).number_format = '0.00%'

            # Color code ROAS
            if data["roas"] >= 3:
                ws.cell(row=r, column=7).fill = FILL_GREEN
            elif data["roas"] < 1:
                ws.cell(row=r, column=7).fill = FILL_RED
            r += 1

        r += 2  # gap between sections

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 12
    for ci in range(5, 11):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    ws.freeze_panes = ws.cell(row=5, column=1)

    print(f"  -> Summary tab rebuilt ({len(active_camps)} active campaigns, {len(rankings)} rankings)")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--from", dest="date_from", default="2024-01")
    parser.add_argument("--to",   dest="date_to",   default="2099-12")
    parser.add_argument("--no-sheet", action="store_true", help="Skip Google Sheet write")
    args = parser.parse_args()

    print(f"\nKPI Monthly Analysis  [{args.date_from} ~ {args.date_to}]")
    print(f"Data source: PG Data Keeper + Polar JSON\n")

    # Compute consistent through_date from all main channels (PST)
    through_date = compute_through_date()
    print(f"Consistent through_date: {through_date}\n")

    rows_discount, total_monthly, d2c_monthly, channel_disc_monthly = analyze_discounts(args.date_from, args.date_to, through_date)
    rows_adspend  = analyze_ad_spend(args.date_from, args.date_to, through_date)
    rows_seeding  = analyze_seeding_cost(args.date_from, args.date_to)

    print("\n" + "=" * 80)

    if args.no_sheet:
        print("--no-sheet: Excel write skipped.")
        return

    import openpyxl
    src = find_latest_model()
    dst = next_version_path(src)
    print(f"\nLoading: {src.name}")
    wb = openpyxl.load_workbook(str(src))

    add_data_status_tab(wb, through_date)
    write_tab(wb, "KPI_할인율",   rows_discount, header_row=3)
    write_wide_tab(wb, "KPI_광고비",   rows_adspend)
    write_wide_tab(wb, "KPI_시딩비용", rows_seeding)

    expand_exec_summary_months(wb, target_start="2025-01")
    update_exec_summary(wb, total_monthly)
    add_mkt_spend_to_exec_summary(wb)
    add_summary_d2c_section(wb, d2c_monthly, rows_seeding, rows_adspend)
    add_amazon_discount_tab(wb, through_date)
    add_amazon_marketplace_tab(wb, through_date)

    # Summary tabs — powered by Data Keeper
    update_sales_summary(wb, through_date)
    update_ads_summary(wb, through_date)
    update_unit_economics(wb, through_date)
    update_campaign_details(wb, through_date)
    update_influencer_dashboard(wb, through_date)

    # Legacy Polar tabs — fill current month from Data Keeper PG
    update_legacy_sales(wb, through_date)
    update_legacy_cm(wb, through_date)
    update_legacy_organic(wb, through_date)
    update_legacy_ads(wb, through_date)
    update_legacy_summary(wb, through_date)

    wb.save(str(dst))
    print(f"\nSaved: {dst.name}")
    print("DONE")


if __name__ == "__main__":
    main()
