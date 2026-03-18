"""
kpi_formatter.py - 포맷이: KPI report formatting orchestrator.

Part of 골만이 Squad (formatter role).
Wraps run_kpi_monthly.py formatting functions and adds:
  - Validation status integration (from 검증이)
  - Client-ready template selection
  - Centralized style engine

Usage:
    python tools/kpi_formatter.py                        # Full format
    python tools/kpi_formatter.py --template executive   # Executive-only tabs
    python tools/kpi_formatter.py --skip-legacy           # Skip legacy Polar tabs
    python tools/kpi_formatter.py --validation-report .tmp/validation_report.json

Pipeline:
    검증이 (kpi_validator.py) → 골만이 (run_kpi_monthly.py compute) → 포맷이 (this file)
"""

import os
import sys
import json
import argparse
from pathlib import Path

TOOLS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(TOOLS_DIR))
from env_loader import load_env
load_env()

# Import formatting functions from run_kpi_monthly (gradual migration target)
from run_kpi_monthly import (
    add_data_status_tab,
    write_tab,
    write_wide_tab,
    expand_exec_summary_months,
    update_exec_summary,
    add_mkt_spend_to_exec_summary,
    add_summary_d2c_section,
    add_amazon_discount_tab,
    add_amazon_marketplace_tab,
    update_sales_summary,
    update_ads_summary,
    update_unit_economics,
    update_campaign_details,
    update_influencer_dashboard,
    update_legacy_sales,
    update_legacy_cm,
    update_legacy_organic,
    update_legacy_ads,
    update_legacy_summary,
    find_latest_model,
    next_version_path,
)


# ── Validation Status Tab ────────────────────────────────────────────────────

def add_validation_tab(wb, validation_report: dict):
    """Add a 'Validation' tab showing 검증이 results."""
    from kpi_style_engine import Fills, Fonts, Aligns, apply_status_fill

    tab_name = "Validation"
    if tab_name in wb.sheetnames:
        del wb[tab_name]
    ws = wb.create_sheet(title=tab_name, index=1)  # after Data Status

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"].value = "KPI Data Validation Report (검증이)"
    ws["A1"].fill = Fills.HEADER
    ws["A1"].font = Fonts.WHITE_LG
    ws["A1"].alignment = Aligns.CENTER
    for col in range(1, 7):
        ws.cell(row=1, column=col).fill = Fills.HEADER

    # Subtitle
    ws.merge_cells("A2:F2")
    ts = validation_report.get("timestamp", "N/A")
    through = validation_report.get("through_date", "N/A")
    overall = validation_report.get("overall_status", "N/A")
    ws["A2"].value = f"Generated: {ts}  |  Through: {through}  |  Overall: {overall}"
    ws["A2"].font = Fonts.SMALL_GREY
    ws["A2"].alignment = Aligns.CENTER

    # Headers
    headers = ["Table", "Rows", "Schema", "Identity", "Coverage", "Notes"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = Fills.HEADER
        cell.font = Fonts.WHITE
        cell.alignment = Aligns.CENTER

    # Table results
    row_idx = 5
    table_results = validation_report.get("table_results", {})
    for table_name, result in sorted(table_results.items()):
        ws.cell(row=row_idx, column=1, value=table_name)
        ws.cell(row=row_idx, column=2, value=result.get("rows", 0))

        schema_status = result.get("schema", {}).get("status", "N/A")
        identity_status = result.get("identity", {}).get("status", "N/A")
        coverage_status = result.get("coverage", {}).get("status", "N/A")

        schema_cell = ws.cell(row=row_idx, column=3, value=schema_status)
        apply_status_fill(schema_cell, schema_status)

        identity_cell = ws.cell(row=row_idx, column=4, value=identity_status)
        apply_status_fill(identity_cell, identity_status)

        coverage_cell = ws.cell(row=row_idx, column=5, value=coverage_status)
        apply_status_fill(coverage_cell, coverage_status)

        # Notes: combine warnings
        notes = []
        for w in result.get("coverage", {}).get("warnings", []):
            notes.append(w)
        schema_errors = result.get("schema", {}).get("schema_errors", 0)
        if schema_errors > 0:
            notes.append(f"{schema_errors} schema errors")
        ws.cell(row=row_idx, column=6, value="; ".join(notes) if notes else "")

        row_idx += 1

    # Cross-table section
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Cross-Table Checks")
    ws.cell(row=row_idx, column=1).font = Fonts.BOLD
    row_idx += 1

    cross = validation_report.get("cross_table", {})

    # Through-date
    td = cross.get("through_date", {})
    ws.cell(row=row_idx, column=1, value="Through-date alignment")
    td_cell = ws.cell(row=row_idx, column=3, value=td.get("status", "N/A"))
    apply_status_fill(td_cell, td.get("status", "N/A"))
    row_idx += 1

    # Reconciliation
    recon = cross.get("reconciliation", {})
    ws.cell(row=row_idx, column=1, value="Amazon reconciliation")
    recon_cell = ws.cell(row=row_idx, column=3, value=recon.get("status", "N/A"))
    apply_status_fill(recon_cell, recon.get("status", "N/A"))
    recon_warns = recon.get("warnings", [])
    if recon_warns:
        ws.cell(row=row_idx, column=6, value=recon_warns[0][:100])
    row_idx += 1

    # Anomalies
    anomaly = cross.get("anomalies", {})
    n_anom = len(anomaly.get("anomalies", []))
    n_spend = len(anomaly.get("spend_warnings", []))
    ws.cell(row=row_idx, column=1, value="Anomaly detection")
    anom_cell = ws.cell(row=row_idx, column=3, value=anomaly.get("status", "N/A"))
    apply_status_fill(anom_cell, anomaly.get("status", "N/A"))
    ws.cell(row=row_idx, column=6, value=f"{n_anom} anomalies, {n_spend} spend warnings")
    row_idx += 2

    # Anomaly details
    if anomaly.get("anomalies"):
        ws.cell(row=row_idx, column=1, value="Anomaly Details")
        ws.cell(row=row_idx, column=1).font = Fonts.BOLD
        row_idx += 1
        for a in anomaly["anomalies"][:10]:
            ws.cell(row=row_idx, column=1, value=a.get("metric", ""))
            ws.cell(row=row_idx, column=2, value=a.get("month", ""))
            ws.cell(row=row_idx, column=3, value=f"MoM {a.get('mom_change', 0):+.1f}%")
            ws.cell(row=row_idx, column=4, value=a.get("value", 0))
            ws.cell(row=row_idx, column=4).number_format = '#,##0'
            row_idx += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 60

    print(f"  -> Tab 'Validation' written (검증이 results)")


# ── Format Pipeline ──────────────────────────────────────────────────────────

def format_kpi_report(
    wb,
    through_date: str,
    rows_discount: list,
    rows_adspend: list,
    rows_seeding: list,
    total_monthly: dict,
    d2c_monthly: dict,
    validation_report: dict | None = None,
    template: str = "full",
    skip_legacy: bool = False,
):
    """
    Full KPI report formatting pipeline.

    Args:
        wb: openpyxl Workbook (loaded)
        through_date: PST-aligned cutoff date
        rows_discount: Discount analysis rows
        rows_adspend: Ad spend rows
        rows_seeding: Seeding cost rows
        total_monthly: Monthly totals dict
        d2c_monthly: D2C monthly dict
        validation_report: 검증이 JSON report (optional)
        template: "full" | "executive" (controls which tabs to write)
        skip_legacy: Skip legacy Polar tabs
    """
    print("\n  포맷이: Formatting KPI report...")

    # 0. Data Status
    add_data_status_tab(wb, through_date)

    # 0.5. Validation tab (if report available)
    if validation_report:
        add_validation_tab(wb, validation_report)

    # 1. Core KPI tabs
    write_tab(wb, "KPI_할인율", rows_discount, header_row=3)
    write_wide_tab(wb, "KPI_광고비", rows_adspend)
    write_wide_tab(wb, "KPI_시딩비용", rows_seeding)

    # 2. Executive Summary
    expand_exec_summary_months(wb, target_start="2025-01")
    update_exec_summary(wb, total_monthly)
    add_mkt_spend_to_exec_summary(wb)
    add_summary_d2c_section(wb, d2c_monthly, rows_seeding, rows_adspend)

    # 3. Amazon discount analysis
    add_amazon_discount_tab(wb, through_date)
    add_amazon_marketplace_tab(wb, through_date)

    if template == "executive":
        print("  포맷이: Executive template — skipping detail tabs.")
        return

    # 4. Summary tabs (Data Keeper powered)
    update_sales_summary(wb, through_date)
    update_ads_summary(wb, through_date)
    update_unit_economics(wb, through_date)
    update_campaign_details(wb, through_date)
    update_influencer_dashboard(wb, through_date)

    # 5. Legacy tabs
    if not skip_legacy:
        update_legacy_sales(wb, through_date)
        update_legacy_cm(wb, through_date)
        update_legacy_organic(wb, through_date)
        update_legacy_ads(wb, through_date)
        update_legacy_summary(wb, through_date)
    else:
        print("  포맷이: --skip-legacy: Legacy tabs skipped.")

    print("  포맷이: Formatting complete.")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    """Standalone formatter: loads latest workbook, applies formatting."""
    parser = argparse.ArgumentParser(description="포맷이: KPI Report Formatter")
    parser.add_argument("--template", choices=["full", "executive"], default="full")
    parser.add_argument("--skip-legacy", action="store_true")
    parser.add_argument("--validation-report", type=str, help="Path to validation_report.json")
    parser.add_argument("--from", dest="date_from", default="2024-01")
    parser.add_argument("--to", dest="date_to", default="2099-12")
    args = parser.parse_args()

    import openpyxl
    from run_kpi_monthly import (
        compute_through_date, analyze_discounts, analyze_ad_spend, analyze_seeding_cost,
    )

    print("\n포맷이: KPI Report Formatter")
    print("=" * 60)

    # Compute data
    through_date = compute_through_date()
    print(f"  Through-date: {through_date}")

    rows_discount, total_monthly, d2c_monthly, _ = analyze_discounts(
        args.date_from, args.date_to, through_date
    )
    rows_adspend = analyze_ad_spend(args.date_from, args.date_to, through_date)
    rows_seeding = analyze_seeding_cost(args.date_from, args.date_to)

    # Load validation report
    validation_report = None
    vr_path = args.validation_report or str(TOOLS_DIR.parent / ".tmp" / "validation_report.json")
    if os.path.exists(vr_path):
        with open(vr_path, "r", encoding="utf-8") as f:
            validation_report = json.load(f)
        print(f"  Validation report loaded: {validation_report.get('overall_status', 'N/A')}")
    else:
        print("  No validation report found. Run kpi_validator.py first.")

    # Load workbook
    src = find_latest_model()
    dst = next_version_path(src)
    print(f"  Loading: {src.name}")
    wb = openpyxl.load_workbook(str(src))

    # Format
    format_kpi_report(
        wb=wb,
        through_date=through_date,
        rows_discount=rows_discount,
        rows_adspend=rows_adspend,
        rows_seeding=rows_seeding,
        total_monthly=total_monthly,
        d2c_monthly=d2c_monthly,
        validation_report=validation_report,
        template=args.template,
        skip_legacy=args.skip_legacy,
    )

    wb.save(str(dst))
    print(f"\n  Saved: {dst.name}")
    print("DONE")


if __name__ == "__main__":
    main()
