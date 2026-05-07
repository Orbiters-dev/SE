#!/usr/bin/env python3
"""
매일 인플루언서 스카우팅 자동화
- 해시태그 8개 랜덤 선택 → 포스트 수집 → 프로필 → 릴스 뷰수
- 0~4K 팔로워 + 평균 뷰 5K+ + 컨텐츠 크리에이터 필터
- 제외: Google Sheet 컨택 리스트 + 이전 스카우팅 이력 + RAG
- 결과: Excel 저장 + Teams 알림

사용법:
  python tools/run_daily_scout.py              # 전체 실행 + Teams 알림
  python tools/run_daily_scout.py --dry-run    # 실행만, Teams 안 보냄
  python tools/run_daily_scout.py --status     # 현황만 출력
"""

import argparse
import json
import os
import sys
import io
import time
import random
import requests
from datetime import datetime, timezone, timedelta
from pathlib import Path

# Windows cp949 인코딩 문제 방지
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv

BASE_DIR = Path(__file__).resolve().parent.parent
load_dotenv(BASE_DIR / ".env")

JST = timezone(timedelta(hours=9))
SCOUT_DIR = BASE_DIR / ".tmp" / "influencer_scout"
OUTPUT_DIR = Path(r"C:\Users\orbit\Desktop\s\인플루언서 리스트")

APIFY_TOKEN = os.getenv("APIFY_API_TOKEN", "")
TEAMS_WEBHOOK = os.getenv("TEAMS_WEBHOOK_URL_SEEUN", "")

# Google Sheet 설정
SHEET_ID = "1wkue4G7FP_fiVeqSmMp7Z6IsIMmvOIc93TBb0NwcAmU"
SERVICE_ACCOUNT_PATH = BASE_DIR / "credentials" / "google_service_account.json"

# 해시태그 풀 (매일 8개 랜덤 선택)
HASHTAG_POOL = [
    "買ってよかった育児グッズ", "育児便利グッズ", "ワンオペ育児", "子育てママ",
    "離乳食グッズ", "ベビーグッズレビュー", "育児の味方", "ママおすすめ",
    "赤ちゃんのいる生活", "育児グッズ", "ベビー用品購入品", "子育て便利",
    "ベビーアイテム", "育児ハック", "新米ママ", "プチプラ育児",
    "0歳児ママ", "1歳児ママ", "育児の裏ワザ", "ママライフ",
    "子育てグッズ", "ベビー用品レビュー", "育児記録", "子育て奮闘中",
]

# 크리에이터 바이오 키워드
CREATOR_KEYWORDS = [
    "おすすめ", "購入品", "レビュー", "紹介", "情報", "発信", "グッズ", "アイテム",
    "育児", "子育て", "ママ", "パパ", "便利", "100均", "ダイソー", "セリア", "神",
    "愛用", "PR", "お仕事", "ご依頼", "コラボ", "DM", "離乳食", "知育", "ベビー用品",
    "買ってよかった", "裏ワザ", "ハック", "vlog", "ワンオペ", "おでかけ", "暮らし", "時短",
]

# 제외 패턴 (username/bio 키워드)
EXCLUDE_PATTERNS = [
    "clinic", "salon", "shop", "store", "coin", "離婚", "投資", "coach",
    "beaute", "美容皮膚", "不動産", "FX", "仮想通貨", "副業", "稼",
    "official", "公式", "株式会社", "inc.", "ltd", "co.", "llc",
    "pr代行", "広告代理", "agency", "メディア", "編集部", "事務局",
]

# 브랜드 블랙리스트 (PG에서 학습된 것 + 수동 추가)
BRAND_BLACKLIST_FILE = Path(__file__).parent.parent / ".tmp" / "brand_blacklist.json"


def load_brand_blacklist():
    """블랙리스트 로드 (세은이 신고한 브랜드 포함)"""
    if BRAND_BLACKLIST_FILE.exists():
        with open(BRAND_BLACKLIST_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"usernames": [], "bio_patterns": []}


def save_brand_blacklist(bl):
    BRAND_BLACKLIST_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(BRAND_BLACKLIST_FILE, "w", encoding="utf-8") as f:
        json.dump(bl, f, ensure_ascii=False, indent=2)


def calc_brand_score(profile):
    """계정이 브랜드/기업일 확률 점수 (0-100). 50+ = 제외"""
    score = 0
    bio = (profile.get("biography", "") or "").lower()
    username = (profile.get("username", "") or "").lower()
    followers = profile.get("followersCount", 0) or profile.get("followers", 0) or 0
    following = profile.get("followingCount", 0) or profile.get("following", 0) or 0
    is_biz = profile.get("isBusinessAccount", False)
    has_url = bool(profile.get("externalUrl") or profile.get("externalLinkUrl"))

    # 비즈니스 계정
    if is_biz:
        score += 30

    # bio에 브랜드/기업 키워드
    biz_keywords = ["公式", "official", "shop", "store", "株式会社", "inc.", "ltd",
                     "co.", "llc", "pr代行", "agency", "メディア", "編集部", "事務局",
                     "送料無料", "セール", "クーポン", "キャンペーン"]
    if any(kw in bio for kw in biz_keywords):
        score += 40

    # 외부 URL 있음
    if has_url:
        score += 10

    # 팔로워 많은데 팔로잉 극소 (브랜드 패턴)
    if followers > 5000 and following < 200:
        score += 20

    # 포스트 수 과다 (매일 올리는 브랜드)
    posts = profile.get("postsCount", 0) or 0
    if posts > 1000:
        score += 10

    # 블랙리스트 체크
    bl = load_brand_blacklist()
    if username in bl.get("usernames", []):
        score = 100
    for pat in bl.get("bio_patterns", []):
        if pat.lower() in bio:
            score += 30

    return min(score, 100)


def ensure_dirs():
    SCOUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def run_actor(actor_id, input_data, timeout_secs=180):
    """Apify actor 실행"""
    url = f"https://api.apify.com/v2/acts/{actor_id}/run-sync-get-dataset-items"
    params = {"token": APIFY_TOKEN, "timeout": timeout_secs}
    try:
        r = requests.post(url, json=input_data, params=params, timeout=timeout_secs + 30)
        if r.status_code in (200, 201):
            return r.json() if r.text else []
        else:
            print(f"  Error {r.status_code}: {r.text[:200]}")
            return []
    except Exception as e:
        print(f"  Exception: {e}")
        return []


def load_exclusion_list():
    """제외 리스트 로드 (Google Sheet + RAG + 이전 스카우팅)"""
    excluded = set()

    # 1. Google Sheet
    try:
        import gspread
        from google.oauth2.service_account import Credentials

        SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
        creds = Credentials.from_service_account_file(str(SERVICE_ACCOUNT_PATH), scopes=SCOPES)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(SHEET_ID)

        # influencer List tab (col 4 = @ID)
        ws1 = sh.get_worksheet_by_id(2007940503)
        for row in ws1.get_all_values()[1:]:
            if len(row) > 4 and row[4].strip():
                h = row[4].strip().lstrip("@").lower()
                if h and len(h) > 1:
                    excluded.add(h)

        # Influencer Search tab (col 2 = Instagram @ID)
        ws2 = sh.get_worksheet_by_id(1294482195)
        for row in ws2.get_all_values()[1:]:
            if len(row) > 2 and row[2].strip():
                h = row[2].strip().lstrip("@").lower()
                if h and len(h) > 1 and not h.startswith("http"):
                    excluded.add(h)

        print(f"  Google Sheet: {len(excluded)}명")
    except Exception as e:
        print(f"  Google Sheet 읽기 실패: {e}")
        # Fallback: 캐시된 제외 리스트
        cache_path = SCOUT_DIR / "exclusion_list.json"
        if cache_path.exists():
            data = json.loads(cache_path.read_text(encoding="utf-8"))
            excluded.update(data.get("handles", []))
            print(f"  캐시 제외 리스트: {len(excluded)}명")

    # 2. RAG 프로필
    rag_dir = BASE_DIR / ".tmp" / "influencer_rag" / "profiles"
    if rag_dir.exists():
        for f in rag_dir.glob("*.json"):
            try:
                p = json.loads(f.read_text(encoding="utf-8"))
                h = p.get("handle", "").strip().lstrip("@").lower()
                if h and len(h) > 1:
                    excluded.add(h)
            except Exception:
                pass

    # 3. 이전 스카우팅 이력
    scouted_path = SCOUT_DIR / "all_scouted.json"
    if scouted_path.exists():
        try:
            prev = json.loads(scouted_path.read_text(encoding="utf-8"))
            excluded.update(u.lower() for u in prev)
            print(f"  이전 스카우팅: {len(prev)}명")
        except Exception:
            pass

    # 캐시 저장
    exclusion = {
        "handles": sorted(excluded),
        "updated_at": datetime.now(JST).isoformat(),
        "count": len(excluded),
    }
    (SCOUT_DIR / "exclusion_list.json").write_text(
        json.dumps(exclusion, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    return excluded


def scrape_hashtags(hashtags, excluded):
    """해시태그에서 신규 유저명 수집"""
    all_usernames = set()
    for tag in hashtags:
        print(f"  #{tag} 수집 중...")
        items = run_actor(
            "apify~instagram-hashtag-scraper",
            {"hashtags": [tag], "resultsLimit": 100},
            timeout_secs=180,
        )
        valid = [p for p in items if not p.get("error")]
        usernames = {p.get("ownerUsername", "") for p in valid if p.get("ownerUsername")}
        new = {u for u in usernames if u.lower() not in excluded and u}
        all_usernames.update(new)
        print(f"    {len(valid)}건 → 신규 {len(new)}명")
        time.sleep(2)
    return list(all_usernames)


def scrape_profiles(usernames):
    """프로필 상세 수집"""
    all_profiles = []
    batch_size = 30
    for i in range(0, len(usernames), batch_size):
        batch = usernames[i : i + batch_size]
        print(f"  프로필 {i + 1}~{i + len(batch)}/{len(usernames)}...")
        items = run_actor(
            "apify~instagram-profile-scraper",
            {"usernames": batch},
            timeout_secs=120,
        )
        all_profiles.extend(items)
        if i + batch_size < len(usernames):
            time.sleep(3)
    return all_profiles


def scrape_reels(usernames):
    """릴스 뷰수 수집"""
    views_map = {}
    batch_size = 10
    for i in range(0, len(usernames), batch_size):
        batch = usernames[i : i + batch_size]
        print(f"  릴스 {i + 1}~{i + len(batch)}/{len(usernames)}...")
        items = run_actor(
            "apify~instagram-reel-scraper",
            {"username": batch, "resultsLimit": 6},
            timeout_secs=300,
        )
        user_views = {}
        for post in items:
            if post.get("error"):
                continue
            username = post.get("ownerUsername", "")
            views = post.get("videoPlayCount", 0) or post.get("videoViewCount", 0)
            if username and views:
                user_views.setdefault(username, []).append(views)
        for username, vlist in user_views.items():
            views_map[username] = {
                "avg_views": int(sum(vlist) / len(vlist)),
                "max_views": max(vlist),
                "min_views": min(vlist),
                "count": len(vlist),
            }
        if i + batch_size < len(usernames):
            time.sleep(3)
    return views_map


def filter_creators(profiles, views_map):
    """컨텐츠 크리에이터 필터링"""
    prof_map = {p.get("username", ""): p for p in profiles if not p.get("error")}
    creators = []

    for username, vdata in sorted(views_map.items(), key=lambda x: x[1]["avg_views"], reverse=True):
        if vdata["avg_views"] < 5000:
            continue
        p = prof_map.get(username, {})
        if not p:
            continue
        followers = p.get("followersCount", 0) or p.get("followers", 0) or 0
        if followers > 4000:
            continue
        bio = p.get("biography", "") or ""
        bio_lower = bio.lower()
        username_lower = username.lower()

        # 비크리에이터 제외 (키워드)
        if any(pat in bio_lower or pat in username_lower for pat in EXCLUDE_PATTERNS):
            continue

        # 브랜드 계정 자동 감지 (score 50+ = 제외)
        brand_score = calc_brand_score(p)
        if brand_score >= 50:
            continue

        matched = [kw for kw in CREATOR_KEYWORDS if kw in bio]
        ratio = vdata["avg_views"] / followers if followers > 0 else 0

        creators.append({
            "username": username,
            "full_name": p.get("fullName", "") or "",
            "followers": followers,
            "following": p.get("followingCount", 0) or 0,
            "posts": p.get("postsCount", 0) or 0,
            "avg_views": vdata["avg_views"],
            "max_views": vdata["max_views"],
            "min_views": vdata["min_views"],
            "reels_count": vdata["count"],
            "ratio": ratio,
            "bio": bio,
            "matched_kw": ", ".join(matched[:5]) if matched else "(bio確認)",
            "url": f"https://www.instagram.com/{username}/",
            "is_business": p.get("isBusinessAccount", False),
        })

    return creators


def export_excel(creators, output_path):
    """Excel 저장"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "スカウト候補"

    hfont = Font(name="Yu Gothic", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    dfont = Font(name="Yu Gothic", size=10)
    lfont = Font(name="Yu Gothic", size=10, color="0563C1", underline="single")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(vertical="top", wrap_text=True)
    alt = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

    headers = [
        "#", "IG Handle", "名前", "フォロワー", "フォロー中", "投稿数",
        "平均リール再生", "最大再生", "再生/フォロワー比", "リール数",
        "Bio", "キーワード", "IG URL", "ビジネス", "メモ",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = center
        cell.border = border

    for i, c in enumerate(creators, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).font = dfont
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2, value=f'@{c["username"]}').font = dfont
        ws.cell(row=row, column=3, value=c["full_name"]).font = dfont
        ws.cell(row=row, column=4, value=c["followers"]).font = dfont
        ws.cell(row=row, column=4).number_format = "#,##0"
        ws.cell(row=row, column=5, value=c["following"]).font = dfont
        ws.cell(row=row, column=5).number_format = "#,##0"
        ws.cell(row=row, column=6, value=c["posts"]).font = dfont
        ws.cell(row=row, column=7, value=c["avg_views"]).font = dfont
        ws.cell(row=row, column=7).number_format = "#,##0"
        ws.cell(row=row, column=8, value=c["max_views"]).font = dfont
        ws.cell(row=row, column=8).number_format = "#,##0"
        ws.cell(row=row, column=9, value=round(c["ratio"], 1)).font = dfont
        ws.cell(row=row, column=9).number_format = '0.0"x"'
        ws.cell(row=row, column=10, value=c["reels_count"]).font = dfont
        ws.cell(row=row, column=11, value=c["bio"][:200]).font = dfont
        ws.cell(row=row, column=11).alignment = wrap
        ws.cell(row=row, column=12, value=c["matched_kw"]).font = dfont
        ws.cell(row=row, column=13, value=c["url"]).font = lfont
        ws.cell(row=row, column=13).hyperlink = c["url"]
        ws.cell(row=row, column=14, value="Yes" if c["is_business"] else "").font = dfont
        ws.cell(row=row, column=15, value="").font = dfont
        for col in range(1, 16):
            ws.cell(row=row, column=col).border = border
            if i % 2 == 0:
                ws.cell(row=row, column=col).fill = alt

    widths = [4, 22, 25, 10, 10, 8, 14, 14, 12, 8, 50, 25, 35, 8, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:O{len(creators) + 1}"
    wb.save(output_path)


def update_scouted_history(creators):
    """스카우팅 이력 업데이트"""
    scouted_path = SCOUT_DIR / "all_scouted.json"
    prev = set()
    if scouted_path.exists():
        try:
            prev = set(json.loads(scouted_path.read_text(encoding="utf-8")))
        except Exception:
            pass
    updated = prev | {c["username"] for c in creators}
    scouted_path.write_text(json.dumps(sorted(updated), ensure_ascii=False), encoding="utf-8")
    return len(updated)


def send_teams_notification(creators, output_path, hashtags_used):
    """Teams 알림 발송"""
    if not TEAMS_WEBHOOK:
        print("  TEAMS_WEBHOOK_URL_SEEUN 미설정 — 알림 생략")
        return

    today = datetime.now(JST).strftime("%Y-%m-%d")
    top5 = creators[:5]
    top_text = "\n".join(
        f"  {i}. @{c['username']} | {c['followers']:,}팔 | avg {c['avg_views']:,}뷰 | {c['ratio']:.0f}x"
        for i, c in enumerate(top5, 1)
    )

    message = (
        f"📋 인플루언서 스카우팅 완료 ({today})\n\n"
        f"신규 후보: {len(creators)}명\n"
        f"해시태그: {len(hashtags_used)}개 사용\n"
        f"저장: 인플루언서 리스트 폴더\n\n"
        f"🏆 Top 5:\n{top_text}\n\n"
        f"엑셀에서 확인해 주세요!"
    )

    payload = {"text": message}
    try:
        r = requests.post(TEAMS_WEBHOOK, json=payload, timeout=10)
        if r.status_code in (200, 202):
            print("  Teams 알림 발송 완료")
        else:
            print(f"  Teams 알림 실패: {r.status_code}")
    except Exception as e:
        print(f"  Teams 알림 예외: {e}")


def show_status():
    """현황 출력"""
    scouted_path = SCOUT_DIR / "all_scouted.json"
    exclusion_path = SCOUT_DIR / "exclusion_list.json"

    print("\n📊 인플루언서 스카우팅 현황")

    if scouted_path.exists():
        scouted = json.loads(scouted_path.read_text(encoding="utf-8"))
        print(f"  스카우팅 이력: {len(scouted)}명")
    else:
        print("  스카우팅 이력: 없음")

    if exclusion_path.exists():
        exc = json.loads(exclusion_path.read_text(encoding="utf-8"))
        print(f"  제외 리스트: {exc['count']}명 (업데이트: {exc['updated_at'][:10]})")

    # 최근 Excel 파일
    if OUTPUT_DIR.exists():
        files = sorted(OUTPUT_DIR.glob("인플루언서_스카우트_*.xlsx"), reverse=True)
        if files:
            print(f"  최근 파일: {files[0].name}")
            print(f"  총 파일: {len(files)}개")


def main():
    ensure_dirs()

    parser = argparse.ArgumentParser(description="Daily Influencer Scout")
    parser.add_argument("--dry-run", action="store_true", help="Teams 알림 안 보냄")
    parser.add_argument("--status", action="store_true", help="현황만 출력")
    args = parser.parse_args()

    if args.status:
        show_status()
        return

    today = datetime.now(JST).strftime("%Y%m%d")
    print(f"\n🔍 인플루언서 스카우팅 시작 ({today})")

    # 1. 제외 리스트 로드
    print("\n[1/5] 제외 리스트 로드...")
    excluded = load_exclusion_list()
    print(f"  총 제외: {len(excluded)}명")

    # 2. 해시태그 랜덤 선택 + 수집
    hashtags = random.sample(HASHTAG_POOL, min(8, len(HASHTAG_POOL)))
    print(f"\n[2/5] 해시태그 수집 ({len(hashtags)}개)...")
    usernames = scrape_hashtags(hashtags, excluded)
    print(f"  신규 계정: {len(usernames)}명")

    if not usernames:
        print("  신규 계정 없음 — 종료")
        return

    # 3. 프로필 수집 + 0~4K 필터
    print(f"\n[3/5] 프로필 수집...")
    profiles = scrape_profiles(usernames)
    candidates = [
        p.get("username", "")
        for p in profiles
        if not p.get("error")
        and not p.get("isPrivate", False)
        and (p.get("followersCount", 0) or p.get("followers", 0) or 0) <= 4000
    ]
    print(f"  0~4K 공개 계정: {len(candidates)}명")

    if not candidates:
        print("  후보 없음 — 종료")
        return

    # 4. 릴스 뷰수 수집
    print(f"\n[4/5] 릴스 뷰수 수집...")
    views_map = scrape_reels(candidates)
    print(f"  뷰수 데이터: {len(views_map)}명")

    # 5. 필터 + Excel 저장
    print(f"\n[5/5] 필터링 + 저장...")
    creators = filter_creators(profiles, views_map)
    print(f"  최종 후보: {len(creators)}명")

    if creators:
        output_path = OUTPUT_DIR / f"인플루언서_스카우트_{today}.xlsx"
        # 동일 날짜 파일 있으면 v2, v3...
        version = 1
        while output_path.exists():
            version += 1
            output_path = OUTPUT_DIR / f"인플루언서_스카우트_{today}_v{version}.xlsx"

        export_excel(creators, str(output_path))
        total_scouted = update_scouted_history(creators)
        print(f"  저장: {output_path.name}")
        print(f"  누적 스카우팅: {total_scouted}명")

        # Teams 알림
        if not args.dry_run:
            send_teams_notification(creators, output_path, hashtags)
        else:
            print("  [dry-run] Teams 알림 생략")
    else:
        print("  조건 맞는 후보 없음")
        if not args.dry_run:
            # 결과 없음 알림도 보냄
            payload = {"text": f"📋 인플루언서 스카우팅 ({today}): 오늘은 조건에 맞는 신규 후보가 없었습니다."}
            try:
                requests.post(TEAMS_WEBHOOK, json=payload, timeout=10)
            except Exception:
                pass

    print("\n✅ 완료")


if __name__ == "__main__":
    main()
