"""Generate fin_data.js for Financial KPIs dashboard tab.

Pulls data from DataKeeper (Shopify, Amazon, Meta, Google, GA4) and computes:
  - Revenue by brand, channel, month
  - Ad spend by platform
  - Ad-attributed vs organic revenue
  - Top search queries
  - GA4 traffic sources
  - Marketing spend waterfall -> Contribution Margin

Usage:
    python tools/generate_fin_data.py
    python tools/generate_fin_data.py --months 9    # last 9 months (default)
    python tools/generate_fin_data.py --push         # git push after generation

Part of 골만이 Squad pipeline.
"""

import json
import os
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

TOOLS_DIR = Path(__file__).resolve().parent
ROOT = TOOLS_DIR.parent
sys.path.insert(0, str(TOOLS_DIR))

from env_loader import load_env
load_env()

from data_keeper_client import DataKeeper

OUTPUT = ROOT / "docs" / "financial-dashboard" / "fin_data.js"
PST = timezone(timedelta(hours=-8))

# Brand / channel config (reuse from run_kpi_monthly)
BRAND_ORDER = ["Grosmimi", "Naeiae", "CHA&MOM", "Alpremio", "Other"]
BRAND_COLORS = {
    "Grosmimi": "#8b5cf6", "Naeiae": "#eab308", "CHA&MOM": "#0ea5e9",
    "Alpremio": "#f97316", "Other": "#94a3b8",
}
CHANNEL_COLORS = {
    "Onzenna D2C": "#6366f1", "Amazon MP": "#f59e0b", "Amazon FBA MCF": "#fb923c",
    "TikTok Shop": "#ec4899", "Target+": "#ef4444", "B2B": "#10b981", "Other": "#94a3b8",
}
AD_COLORS = {
    "Amazon Ads": "#f59e0b", "Meta CVR": "#3b82f6", "Meta Traffic": "#93c5fd",
    "Google Ads": "#10b981",
}
# REMOVED: flat AVG_COGS — always use SKU-level or brand VW avg from NAS.
# If NAS unavailable, COGS = null (n.m.), never a guess.
AVG_PRICE = {
    "Grosmimi": 28.0, "Naeiae": 18.0, "CHA&MOM": 32.0,
    "Onzenna": 22.0, "Alpremio": 38.0,
}

# n.m. (not measured) — data not yet collected for these periods.
# Use None in JSON arrays; HTML renders as "n.m." in dark grey.
# Rule: 0 means genuinely zero activity; None means no data source.
_NM_RULES = {
    "amazon_ads":   lambda m: False,  # backfill covers all months now
    "fba_fulfill":  lambda m: True,   # FBA fee report not yet running
    "sku_cogs":     lambda m: m < "2025-10",  # SKU daily starts Oct 2025
}

# NAS paths for SKU-level COGS (Option B from run_kpi_monthly.py)
_NAS_COGS_DIR = Path(r"Z:\Orbiters\ORBI CLAUDE_0223\ORBITERS CLAUDE\ORBITERS CLAUDE\Shared\NoPolar KPIs\Data config sheet")

_POLAR_EXCEL = Path("Data Storage/_archive/Monthly Sales by brands_raw.xlsx")
_SEEDING_DIR = Path(".tmp/polar_data")


def load_amz_ads_backfill():
    """Load Amazon Ads monthly spend from Polar Excel 'IR 매출분석' row 119.
    Used for months before DataKeeper amazon_ads_daily (Dec 2025+)."""
    try:
        import openpyxl
        if not _POLAR_EXCEL.exists():
            print("  [WARN] Polar Excel not found, no Amazon Ads backfill")
            return {}
        wb = openpyxl.load_workbook(str(_POLAR_EXCEL), read_only=True, data_only=True)
        ws = wb["IR 매출분석"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # Row 2 has dates, Row 119 has Amazon Ads total
        date_row = rows[2]
        amz_row = rows[119]
        out = {}
        for i, d in enumerate(date_row):
            if d and hasattr(d, "year") and d.year >= 2025:
                m = f"{d.year}-{d.month:02d}"
                val = amz_row[i] if i < len(amz_row) else 0
                if val and float(val) > 0:
                    out[m] = round(float(val))
        return out
    except Exception as e:
        print(f"  [WARN] Amazon Ads backfill load failed: {e}")
        return {}


def load_influencer_costs():
    """Load influencer costs from PayPal + Shopify PR JSON files.
    Returns (paid_monthly, nonpaid_monthly) dicts {month: amount}."""
    cogs_map = load_sku_cogs_map()
    paid = defaultdict(float)
    nonpaid = defaultdict(float)

    # PAID: PayPal outbound payments
    pp_path = _SEEDING_DIR / "q11_paypal_transactions.json"
    if pp_path.exists():
        try:
            import json as _json
            data = _json.loads(pp_path.read_text(encoding="utf-8"))
            txns = data.get("transactions", data if isinstance(data, list) else [])
            for t in txns:
                amt = float(t.get("amount", 0) or 0)
                if amt < 0:
                    m = (t.get("date", "") or "")[:7]
                    if m:
                        paid[m] += abs(amt)
        except Exception as e:
            print(f"  [WARN] PayPal load failed: {e}")

    # NON-PAID: Shopify PR orders (Sample COGS + $10 shipping per unit)
    inf_path = _SEEDING_DIR / "q10_influencer_orders.json"
    if inf_path.exists():
        try:
            import json as _json
            data = _json.loads(inf_path.read_text(encoding="utf-8"))
            orders = data if isinstance(data, list) else data.get("orders", [])
            for order in orders:
                m = (order.get("created_at", "") or "")[:7]
                if not m:
                    continue
                for li in order.get("line_items", []):
                    sku = (li.get("sku") or "").strip().lower()
                    qty = int(li.get("quantity", 1) or 1)
                    cost = cogs_map.get(sku, 0)
                    nonpaid[m] += cost * qty + 10.0 * qty  # COGS + $10 shipping
        except Exception as e:
            print(f"  [WARN] PR orders load failed: {e}")

    return dict(paid), dict(nonpaid)


def load_sku_cogs_map():
    """Load SKU -> COGS from NAS Excel. Returns {sku_lower: cost}."""
    try:
        import openpyxl
        p = _NAS_COGS_DIR / "COGS by SKU.xlsx"
        if not p.exists():
            return {}
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        sku_col = next((i for i, h in enumerate(headers) if "sku" in h), None)
        cost_col = next((i for i, h in enumerate(headers) if "cost" in h and "type" not in h), None)
        if sku_col is None or cost_col is None:
            wb.close()
            return {}
        out = {}
        for row in rows[1:]:
            sku = str(row[sku_col] or "").strip().lower()
            try:
                cost = float(row[cost_col] or 0)
            except (TypeError, ValueError):
                continue
            if sku and cost > 0:
                out[sku] = cost
        wb.close()
        return out
    except Exception as e:
        print(f"  [WARN] SKU COGS load failed: {e}")
        return {}


def build_sku_cogs_monthly(amazon_sku_rows, shopify_sku_rows, cogs_map, through, brand_order):
    """Compute per-brand per-month COGS from SKU-level data.

    Returns:
        amz_cogs[brand][month] = total COGS (float)
        shop_cogs[brand][month] = total COGS (float)  -- Shopify D2C only (excl PR/Amazon)
        shop_cogs_by_channel[channel][brand][month] = total COGS (float)
    """
    SKIP_CH = {"PR", "Amazon"}
    amz_cogs = defaultdict(lambda: defaultdict(float))
    shop_cogs = defaultdict(lambda: defaultdict(float))
    shop_cogs_by_channel = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    # Volume-weighted avg COGS per brand (for unmatched SKUs)
    _brand_vw = defaultdict(lambda: {"cost_sum": 0.0, "units": 0})

    # Pass 1: compute brand vw avg from matched amazon SKUs
    for r in amazon_sku_rows:
        sku = str(r.get("sku", "")).strip().lower()
        cost = cogs_map.get(sku)
        if cost:
            b = r.get("brand", "Other")
            u = int(r.get("units", 0) or 0)
            _brand_vw[b]["cost_sum"] += cost * u
            _brand_vw[b]["units"] += u
    for r in shopify_sku_rows:
        sku = str(r.get("sku", "")).strip().lower()
        cost = cogs_map.get(sku)
        if cost:
            b = r.get("brand", "Other")
            u = int(r.get("units", 0) or 0)
            _brand_vw[b]["cost_sum"] += cost * u
            _brand_vw[b]["units"] += u

    brand_vw_avg = {}
    for b, v in _brand_vw.items():
        brand_vw_avg[b] = v["cost_sum"] / v["units"] if v["units"] else 0

    # Track SKU units per month (for partial-month detection)
    amz_sku_units = defaultdict(int)  # month -> total units from SKU data

    # Pass 2: compute monthly COGS
    for r in amazon_sku_rows:
        d = r.get("date", "")
        if not d or d > through:
            continue
        m = d[:7]
        b = r.get("brand", "Other")
        if b not in brand_order and b != "Other":
            b = "Other"
        u = int(r.get("units", 0) or 0)
        sku = str(r.get("sku", "")).strip().lower()
        cost = cogs_map.get(sku, brand_vw_avg.get(b, 0))
        amz_cogs[b][m] += u * cost
        amz_sku_units[m] += u

    shop_sku_units = defaultdict(int)
    for r in shopify_sku_rows:
        d = r.get("date", "")
        if not d or d > through:
            continue
        ch = r.get("channel", "")
        if ch in SKIP_CH:
            continue
        m = d[:7]
        b = r.get("brand", "Other")
        if b not in brand_order and b != "Other":
            b = "Other"
        u = int(r.get("units", 0) or 0)
        sku = str(r.get("sku", "")).strip().lower()
        cost = cogs_map.get(sku, brand_vw_avg.get(b, 0))
        shop_cogs[b][m] += u * cost
        shop_sku_units[m] += u
        # Map channel for channel P&L
        ch_mapped = "Onzenna D2C"
        if ch == "B2B" or ch == "Wholesale":
            ch_mapped = "B2B"
        elif ch == "TikTok":
            ch_mapped = "TikTok"
        shop_cogs_by_channel[ch_mapped][b][m] += u * cost

    return amz_cogs, shop_cogs, shop_cogs_by_channel, brand_vw_avg, amz_sku_units, shop_sku_units


CHANNEL_PNL_COLORS = {
    "Amazon MP": "#f59e0b",
    "Onzenna D2C": "#6366f1",
    "Target+": "#ef4444",
    "B2B": "#10b981",
    "Other": "#94a3b8",
}



def generate():
    import argparse
    parser = argparse.ArgumentParser(description="Generate Financial KPIs data")
    parser.add_argument("--months", type=int, default=15)
    parser.add_argument("--push", action="store_true")
    args = parser.parse_args()

    dk = DataKeeper(prefer_cache=False)
    now_pst = datetime.now(PST)
    today = now_pst.date()
    yesterday = (today - timedelta(days=1)).isoformat()

    # Date ranges
    days_back = args.months * 31 + 30  # extra buffer
    d7 = (today - timedelta(days=7)).isoformat()
    d30 = (today - timedelta(days=30)).isoformat()
    mtd_start = today.replace(day=1).isoformat()
    ytd_start = f"{today.year}-01-01"

    import calendar

    print("=== Generating Financial KPIs Data ===")
    print(f"  Today (PST): {today}")

    # ── 1. Load all DataKeeper tables ─────────────────────────────────────────
    print("\n[1/7] Loading DataKeeper tables...")
    shopify = dk.get("shopify_orders_daily", days=days_back)
    amazon_sales = dk.get("amazon_sales_daily", days=days_back)
    amazon_ads = dk.get("amazon_ads_daily", days=days_back)
    meta_ads = dk.get("meta_ads_daily", days=days_back)
    google_ads = dk.get("google_ads_daily", days=days_back)
    ga4 = dk.get("ga4_daily", days=days_back)
    search_terms = dk.get("amazon_ads_search_terms", days=30)
    gsc = dk.get("gsc_daily", days=30)
    shopify_sku = dk.get("shopify_orders_sku_daily", date_from="2025-06-01")
    amazon_sku = dk.get("amazon_sales_sku_daily", days=days_back)
    klaviyo = dk.get("klaviyo_daily", days=days_back)
    # Hero Products dependencies
    brand_analytics = dk.get("amazon_brand_analytics", days=180)  # override 30d with 180d for hero SFR history
    try:
        sqp_brand = dk.get("amazon_sqp_brand", days=365)
    except (ValueError, Exception):
        sqp_brand = []  # SQP table not yet in DK
    _cp_all = {}
    for _brand in ["Grosmimi", "CHA&MOM", "Naeiae", "Onzenna", "Babyrabbit", "Commemoi", "Goongbe", ""]:
        _cp = dk.get("content_posts", days=365, limit=50000, brand=_brand if _brand else None)
        for p in _cp:
            pid = p.get("post_id") or p.get("url", "")
            if pid and pid not in _cp_all:
                _cp_all[pid] = p
    content_posts = list(_cp_all.values())
    content_metrics = dk.get("content_metrics_daily", days=90, limit=50000)
    google_search_terms = dk.get("google_ads_search_terms", days=30)
    print(f"  Shopify: {len(shopify)} rows, Amazon Sales: {len(amazon_sales)}")
    print(f"  Amazon Ads: {len(amazon_ads)}, Meta: {len(meta_ads)}, Google: {len(google_ads)}")
    print(f"  GA4: {len(ga4)}, Search Terms: {len(search_terms)}, GSC: {len(gsc)}")
    print(f"  Brand Analytics: {len(brand_analytics)}, Shopify SKU: {len(shopify_sku)}, Amazon SKU: {len(amazon_sku)}")
    print(f"  Klaviyo: {len(klaviyo)}")
    print(f"  SQP Brand: {len(sqp_brand)}, Content Posts: {len(content_posts)}, Content Metrics: {len(content_metrics)}")

    # ── Load SKU-level COGS map ──────────────────────────────────────────────
    sku_cogs_map = load_sku_cogs_map()
    print(f"  SKU COGS map: {len(sku_cogs_map)} SKUs {'(NAS)' if sku_cogs_map else '(MISSING - COGS will be n.m.)'}")

    # ── Load Amazon Ads backfill + Influencer costs ─────────────────────────
    amz_ads_backfill = load_amz_ads_backfill()
    print(f"  Amazon Ads backfill: {len(amz_ads_backfill)} months {'(Polar Excel)' if amz_ads_backfill else '(none)'}")
    inf_paid_m, inf_nonpaid_m = load_influencer_costs()
    print(f"  Influencer: PAID {len(inf_paid_m)}m, NON-PAID {len(inf_nonpaid_m)}m")

    # ── Compute through-date ─────────────────────────────────────────────────
    def max_date(rows):
        dates = [r.get("date", "") for r in rows if r.get("date")]
        return max(dates) if dates else "0000"

    through = min(max_date(shopify), max_date(amazon_ads), yesterday)
    print(f"  Through-date: {through}")

    # ── Data source freshness metadata ────────────────────────────────────────
    def source_meta(name, rows, label, refresh="2x daily"):
        dates = [r.get("date", "") for r in rows if r.get("date")]
        return {
            "name": name,
            "label": label,
            "rows": len(rows),
            "min_date": min(dates) if dates else "",
            "max_date": max(dates) if dates else "",
            "refresh": refresh,
        }

    data_sources = [
        source_meta("shopify_orders_daily", shopify, "Shopify Orders"),
        source_meta("amazon_sales_daily", amazon_sales, "Amazon Sales (SP-API)"),
        source_meta("amazon_ads_daily", amazon_ads, "Amazon Ads"),
        source_meta("meta_ads_daily", meta_ads, "Meta Ads"),
        source_meta("google_ads_daily", google_ads, "Google Ads"),
        source_meta("ga4_daily", ga4, "GA4 Analytics"),
        source_meta("amazon_ads_search_terms", search_terms, "Amazon Search Terms", "1x daily"),
        source_meta("gsc_daily", gsc, "Google Search Console"),
        source_meta("klaviyo_daily", klaviyo, "Klaviyo Email"),
    ]

    # ── Partial month detection ───────────────────────────────────────────────
    through_date_obj = datetime.strptime(through, "%Y-%m-%d").date()
    current_month = through[:7]  # e.g. "2026-03"
    days_elapsed = through_date_obj.day
    days_in_month = calendar.monthrange(through_date_obj.year, through_date_obj.month)[1]
    is_partial = days_elapsed < days_in_month
    fm_multiplier = round(days_in_month / days_elapsed, 4) if days_elapsed > 0 else 1
    print(f"  Partial month: {current_month} ({days_elapsed}/{days_in_month}d) -> multiplier {fm_multiplier}x")

    partial_month_info = {
        "month": current_month,
        "days_elapsed": days_elapsed,
        "days_in_month": days_in_month,
        "is_partial": is_partial,
        "multiplier": fm_multiplier,
    }

    # ── Helper: date -> ISO week key ────────────────────────────────────────
    def _week_key(date_str):
        dt = datetime.strptime(date_str, "%Y-%m-%d").date()
        iy, iw, _ = dt.isocalendar()
        return f"{iy}-W{iw:02d}"

    # ── 2. Revenue by Brand (Shopify only, PR/Amazon excluded) ─────────────────
    # Amazon channel = stale FBA MCF rows in PG (now reclassified as D2C in current code)
    # True Amazon Marketplace sales come from amazon_sales_daily (SP-API)
    SKIP_CHANNELS = {"PR", "Amazon"}
    print("\n[2/7] Computing revenue by brand...")
    brand_monthly = defaultdict(lambda: defaultdict(lambda: {
        "gross": 0, "net": 0, "disc": 0, "orders": 0, "units": 0
    }))
    # Weekly mirrors
    brand_weekly = defaultdict(lambda: defaultdict(lambda: {
        "gross": 0, "net": 0, "disc": 0, "orders": 0, "units": 0
    }))
    amz_brand_weekly = defaultdict(lambda: defaultdict(lambda: {
        "gross": 0, "net": 0, "orders": 0, "units": 0
    }))
    channel_weekly = defaultdict(lambda: defaultdict(lambda: {"net": 0, "orders": 0}))
    ad_weekly = defaultdict(lambda: defaultdict(lambda: {
        "spend": 0, "sales": 0, "impressions": 0, "clicks": 0
    }))
    brand_ad_weekly = defaultdict(lambda: defaultdict(lambda: {"spend": 0, "sales": 0}))

    for r in shopify:
        d = r.get("date", "")
        if not d or d > through:
            continue
        if r.get("channel") in SKIP_CHANNELS:
            continue
        month = d[:7]
        brand = r.get("brand") or "Other"
        if brand not in BRAND_ORDER:
            brand = "Other"
        brand_monthly[brand][month]["gross"] += float(r.get("gross_sales") or 0)
        brand_monthly[brand][month]["net"] += float(r.get("net_sales") or 0)
        brand_monthly[brand][month]["disc"] += float(r.get("discounts") or 0)
        brand_monthly[brand][month]["orders"] += int(r.get("orders") or 0)
        brand_monthly[brand][month]["units"] += int(r.get("units") or 0)
        wk = _week_key(d)
        brand_weekly[brand][wk]["gross"] += float(r.get("gross_sales") or 0)
        brand_weekly[brand][wk]["net"] += float(r.get("net_sales") or 0)
        brand_weekly[brand][wk]["disc"] += float(r.get("discounts") or 0)
        brand_weekly[brand][wk]["orders"] += int(r.get("orders") or 0)
        brand_weekly[brand][wk]["units"] += int(r.get("units") or 0)

    # Amazon SP-API revenue by brand
    amz_brand_monthly = defaultdict(lambda: defaultdict(lambda: {
        "gross": 0, "net": 0, "orders": 0, "units": 0
    }))
    for r in amazon_sales:
        d = r.get("date", "")
        if not d or d > through:
            continue
        month = d[:7]
        brand = r.get("brand") or "Other"
        if brand not in BRAND_ORDER:
            brand = "Other"
        amz_brand_monthly[brand][month]["gross"] += float(r.get("gross_sales") or r.get("ordered_product_sales") or 0)
        amz_brand_monthly[brand][month]["net"] += float(r.get("net_sales") or 0)
        amz_brand_monthly[brand][month]["orders"] += int(r.get("orders") or 0)
        amz_brand_monthly[brand][month]["units"] += int(r.get("units") or 0)
        wk = _week_key(d)
        amz_brand_weekly[brand][wk]["gross"] += float(r.get("gross_sales") or r.get("ordered_product_sales") or 0)
        amz_brand_weekly[brand][wk]["net"] += float(r.get("net_sales") or 0)
        amz_brand_weekly[brand][wk]["orders"] += int(r.get("orders") or 0)
        amz_brand_weekly[brand][wk]["units"] += int(r.get("units") or 0)

    # ── 3. Revenue by Channel (Shopify channels + Amazon MP) ──────────────────
    print("\n[3/7] Computing revenue by channel...")
    channel_monthly = defaultdict(lambda: defaultdict(lambda: {"net": 0, "orders": 0}))

    CHANNEL_MAP = {
        "D2C": "Onzenna D2C",
        "Amazon": "Amazon FBA MCF",
        "TikTok": "TikTok Shop",
        "B2B": "B2B",
        "Target+": "Target+",
    }

    for r in shopify:
        d = r.get("date", "")
        if not d or d > through:
            continue
        if r.get("channel") in SKIP_CHANNELS:
            continue
        month = d[:7]
        raw_ch = r.get("channel") or "Other"
        ch = CHANNEL_MAP.get(raw_ch, "Other")
        channel_monthly[ch][month]["net"] += float(r.get("net_sales") or 0)
        channel_monthly[ch][month]["orders"] += int(r.get("orders") or 0)
        wk = _week_key(d)
        channel_weekly[ch][wk]["net"] += float(r.get("net_sales") or 0)
        channel_weekly[ch][wk]["orders"] += int(r.get("orders") or 0)

    # Amazon Marketplace (SP-API)
    for r in amazon_sales:
        d = r.get("date", "")
        if not d or d > through:
            continue
        month = d[:7]
        channel_monthly["Amazon MP"][month]["net"] += float(r.get("net_sales") or 0)
        channel_monthly["Amazon MP"][month]["orders"] += int(r.get("orders") or 0)
        wk = _week_key(d)
        channel_weekly["Amazon MP"][wk]["net"] += float(r.get("net_sales") or 0)
        channel_weekly["Amazon MP"][wk]["orders"] += int(r.get("orders") or 0)

    # ── 4. Ad Spend by Platform ───────────────────────────────────────────────
    print("\n[4/7] Computing ad spend by platform...")
    ad_monthly = defaultdict(lambda: defaultdict(lambda: {
        "spend": 0, "sales": 0, "impressions": 0, "clicks": 0
    }))

    for r in amazon_ads:
        d = r.get("date", "")
        if not d or d > through:
            continue
        month = d[:7]
        ad_monthly["Amazon Ads"][month]["spend"] += float(r.get("spend") or 0)
        ad_monthly["Amazon Ads"][month]["sales"] += float(r.get("sales") or 0)
        ad_monthly["Amazon Ads"][month]["impressions"] += int(r.get("impressions") or 0)
        ad_monthly["Amazon Ads"][month]["clicks"] += int(r.get("clicks") or 0)
        wk = _week_key(d)
        ad_weekly["Amazon Ads"][wk]["spend"] += float(r.get("spend") or 0)
        ad_weekly["Amazon Ads"][wk]["sales"] += float(r.get("sales") or 0)
        ad_weekly["Amazon Ads"][wk]["impressions"] += int(r.get("impressions") or 0)
        ad_weekly["Amazon Ads"][wk]["clicks"] += int(r.get("clicks") or 0)

    for r in meta_ads:
        d = r.get("date", "")
        if not d or d > through:
            continue
        month = d[:7]
        # Classify by landing destination: Amazon-landing = Traffic, else = CVR
        cname = (r.get("campaign_name") or "").lower()
        landing = (r.get("landing_url") or "").lower()
        is_amz = "amazon" in cname or "amz" in cname or "amazon" in landing
        label = "Meta Traffic" if is_amz else "Meta CVR"
        ad_monthly[label][month]["spend"] += float(r.get("spend") or 0)
        ad_monthly[label][month]["sales"] += float(r.get("purchase_value") or 0)
        ad_monthly[label][month]["impressions"] += int(r.get("impressions") or 0)
        ad_monthly[label][month]["clicks"] += int(r.get("clicks") or 0)
        wk = _week_key(d)
        ad_weekly[label][wk]["spend"] += float(r.get("spend") or 0)
        ad_weekly[label][wk]["sales"] += float(r.get("purchase_value") or 0)
        ad_weekly[label][wk]["impressions"] += int(r.get("impressions") or 0)
        ad_weekly[label][wk]["clicks"] += int(r.get("clicks") or 0)

    for r in google_ads:
        d = r.get("date", "")
        if not d or d > through:
            continue
        month = d[:7]
        ad_monthly["Google Ads"][month]["spend"] += float(r.get("spend") or 0)
        ad_monthly["Google Ads"][month]["sales"] += float(r.get("conversion_value") or 0)
        ad_monthly["Google Ads"][month]["impressions"] += int(r.get("impressions") or 0)
        ad_monthly["Google Ads"][month]["clicks"] += int(r.get("clicks") or 0)
        wk = _week_key(d)
        ad_weekly["Google Ads"][wk]["spend"] += float(r.get("spend") or 0)
        ad_weekly["Google Ads"][wk]["sales"] += float(r.get("conversion_value") or 0)
        ad_weekly["Google Ads"][wk]["impressions"] += int(r.get("impressions") or 0)
        ad_weekly["Google Ads"][wk]["clicks"] += int(r.get("clicks") or 0)

    # ── Brand-level ad performance ────────────────────────────────────────────
    brand_ad_monthly = defaultdict(lambda: defaultdict(lambda: {"spend": 0, "sales": 0}))

    for r in amazon_ads:
        d = r.get("date", "")
        if not d or d > through:
            continue
        month = d[:7]
        brand = r.get("brand") or "Other"
        if brand not in BRAND_ORDER:
            brand = "Other"
        brand_ad_monthly[brand][month]["spend"] += float(r.get("spend") or 0)
        brand_ad_monthly[brand][month]["sales"] += float(r.get("sales") or 0)
        wk = _week_key(d)
        brand_ad_weekly[brand][wk]["spend"] += float(r.get("spend") or 0)
        brand_ad_weekly[brand][wk]["sales"] += float(r.get("sales") or 0)

    for r in meta_ads:
        d = r.get("date", "")
        if not d or d > through:
            continue
        month = d[:7]
        brand = r.get("brand") or "Other"
        if brand not in BRAND_ORDER:
            brand = "Other"
        brand_ad_monthly[brand][month]["spend"] += float(r.get("spend") or 0)
        brand_ad_monthly[brand][month]["sales"] += float(r.get("purchase_value") or 0)
        wk = _week_key(d)
        brand_ad_weekly[brand][wk]["spend"] += float(r.get("spend") or 0)
        brand_ad_weekly[brand][wk]["sales"] += float(r.get("purchase_value") or 0)

    for r in google_ads:
        d = r.get("date", "")
        if not d or d > through:
            continue
        month = d[:7]
        # All Google Ads campaigns are Grosmimi (Mint | prefix)
        brand_ad_monthly["Grosmimi"][month]["spend"] += float(r.get("spend") or 0)
        brand_ad_monthly["Grosmimi"][month]["sales"] += float(r.get("conversion_value") or 0)
        wk = _week_key(d)
        brand_ad_weekly["Grosmimi"][wk]["spend"] += float(r.get("spend") or 0)
        brand_ad_weekly["Grosmimi"][wk]["sales"] += float(r.get("conversion_value") or 0)

    # ── Brand × Platform ad breakdown ─────────────────────────────────────────
    brand_ad_by_platform = defaultdict(lambda: defaultdict(lambda: defaultdict(
        lambda: {"spend": 0, "sales": 0, "impressions": 0, "clicks": 0}
    )))
    brand_ad_by_platform_wk = defaultdict(lambda: defaultdict(lambda: defaultdict(
        lambda: {"spend": 0, "sales": 0, "impressions": 0, "clicks": 0}
    )))

    # ── Brand × Platform DAILY (for Hero Products ad overlay) ─────────────────
    # brand_ad_daily_raw[brand][metric][date] = float
    # metrics: amz_spend, amz_clicks, amz_sales, mt_spend, mt_clicks, mt_sales,
    #          gads_spend, gads_clicks, gads_sales
    brand_ad_daily_raw = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    for r in amazon_ads:
        d = r.get("date", "")
        if not d or d > through:
            continue
        brand = r.get("brand") or "Other"
        if brand not in BRAND_ORDER:
            brand = "Other"
        month = d[:7]
        wk = _week_key(d)
        brand_ad_by_platform["Amazon Ads"][brand][month]["spend"] += float(r.get("spend") or 0)
        brand_ad_by_platform["Amazon Ads"][brand][month]["sales"] += float(r.get("sales") or 0)
        brand_ad_by_platform["Amazon Ads"][brand][month]["impressions"] += int(r.get("impressions") or 0)
        brand_ad_by_platform["Amazon Ads"][brand][month]["clicks"] += int(r.get("clicks") or 0)
        brand_ad_by_platform_wk["Amazon Ads"][brand][wk]["spend"] += float(r.get("spend") or 0)
        brand_ad_by_platform_wk["Amazon Ads"][brand][wk]["sales"] += float(r.get("sales") or 0)
        brand_ad_by_platform_wk["Amazon Ads"][brand][wk]["impressions"] += int(r.get("impressions") or 0)
        brand_ad_by_platform_wk["Amazon Ads"][brand][wk]["clicks"] += int(r.get("clicks") or 0)
        # Daily accumulation for Hero tab
        brand_ad_daily_raw[brand]["amz_spend"][d] += float(r.get("spend") or 0)
        brand_ad_daily_raw[brand]["amz_clicks"][d] += int(r.get("clicks") or 0)
        brand_ad_daily_raw[brand]["amz_sales"][d] += float(r.get("sales") or 0)

    for r in meta_ads:
        d = r.get("date", "")
        if not d or d > through:
            continue
        brand = r.get("brand") or "Other"
        if brand not in BRAND_ORDER:
            brand = "Other"
        cname = (r.get("campaign_name") or "").lower()
        landing = (r.get("landing_url") or "").lower()
        is_amz = "amazon" in cname or "amz" in cname or "amazon" in landing
        label = "Meta Traffic" if is_amz else "Meta CVR"
        month = d[:7]
        wk = _week_key(d)
        brand_ad_by_platform[label][brand][month]["spend"] += float(r.get("spend") or 0)
        brand_ad_by_platform[label][brand][month]["sales"] += float(r.get("purchase_value") or 0)
        brand_ad_by_platform[label][brand][month]["impressions"] += int(r.get("impressions") or 0)
        brand_ad_by_platform[label][brand][month]["clicks"] += int(r.get("clicks") or 0)
        brand_ad_by_platform_wk[label][brand][wk]["spend"] += float(r.get("spend") or 0)
        brand_ad_by_platform_wk[label][brand][wk]["sales"] += float(r.get("purchase_value") or 0)
        brand_ad_by_platform_wk[label][brand][wk]["impressions"] += int(r.get("impressions") or 0)
        brand_ad_by_platform_wk[label][brand][wk]["clicks"] += int(r.get("clicks") or 0)
        # Daily accumulation for Hero tab (Meta Traffic = AMZ landing)
        if is_amz:
            brand_ad_daily_raw[brand]["mt_spend"][d] += float(r.get("spend") or 0)
            brand_ad_daily_raw[brand]["mt_clicks"][d] += int(r.get("clicks") or 0)
            brand_ad_daily_raw[brand]["mt_sales"][d] += float(r.get("purchase_value") or 0)

    # NOTE: Attribution sales injection happens after attribution data loaded (see below)

    for r in google_ads:
        d = r.get("date", "")
        if not d or d > through:
            continue
        brand = r.get("brand") or "Grosmimi"
        if brand not in BRAND_ORDER:
            brand = "Grosmimi"
        month = d[:7]
        wk = _week_key(d)
        brand_ad_by_platform["Google Ads"][brand][month]["spend"] += float(r.get("spend") or 0)
        brand_ad_by_platform["Google Ads"][brand][month]["sales"] += float(r.get("conversion_value") or 0)
        brand_ad_by_platform["Google Ads"][brand][month]["impressions"] += int(r.get("impressions") or 0)
        brand_ad_by_platform["Google Ads"][brand][month]["clicks"] += int(r.get("clicks") or 0)
        brand_ad_by_platform_wk["Google Ads"][brand][wk]["spend"] += float(r.get("spend") or 0)
        brand_ad_by_platform_wk["Google Ads"][brand][wk]["sales"] += float(r.get("conversion_value") or 0)
        brand_ad_by_platform_wk["Google Ads"][brand][wk]["impressions"] += int(r.get("impressions") or 0)
        brand_ad_by_platform_wk["Google Ads"][brand][wk]["clicks"] += int(r.get("clicks") or 0)
        # Daily accumulation for Hero tab
        brand_ad_daily_raw[brand]["gads_spend"][d] += float(r.get("spend") or 0)
        brand_ad_daily_raw[brand]["gads_clicks"][d] += int(r.get("clicks") or 0)
        brand_ad_daily_raw[brand]["gads_sales"][d] += float(r.get("conversion_value") or 0)

    # ── Build month list ──────────────────────────────────────────────────────
    all_months_set = set()
    for d in [brand_monthly, amz_brand_monthly, channel_monthly]:
        for entity in d.values():
            all_months_set |= set(entity.keys())
    for platform in ad_monthly.values():
        all_months_set |= set(platform.keys())

    # Filter to last N months
    cutoff_month = (today - timedelta(days=args.months * 30)).strftime("%Y-%m")
    months = sorted(m for m in all_months_set if m >= cutoff_month)
    ad_start_idx = next((i for i, m in enumerate(months) if m >= "2026-01"), 0)
    ad_months = months[ad_start_idx:]
    print(f"  Months: {months[0]} -> {months[-1]} ({len(months)} months)")
    print(f"  Ad months: {ad_months[0] if ad_months else 'none'} -> {ad_months[-1] if ad_months else 'none'} (from idx {ad_start_idx})")

    # ── 4b. SKU-level COGS by brand/month ────────────────────────────────────
    amz_sku_cogs, shop_sku_cogs, shop_cogs_by_channel, brand_vw_cogs, amz_sku_units_m, shop_sku_units_m = build_sku_cogs_monthly(
        amazon_sku, shopify_sku, sku_cogs_map, through, BRAND_ORDER
    )
    if sku_cogs_map:
        print(f"  SKU COGS built - VW avg: {', '.join(f'{b}=${v:.2f}' for b, v in sorted(brand_vw_cogs.items()) if b in BRAND_ORDER[:4])}")
    else:
        print(f"  SKU COGS: NAS unavailable, using brand VW avg from matched SKUs")

    # ── 5. Compute summary KPIs (7d, 30d, MTD) ───────────────────────────────
    print("\n[5/7] Computing KPI summaries...")

    def compute_period_kpi(date_from, date_to):
        """Compute KPIs for a date range from raw daily data."""
        shopify_rev = 0
        shopify_gross = 0
        shopify_disc = 0
        shopify_orders = 0
        shopify_units = 0
        shopify_cogs = 0

        for r in shopify:
            d = r.get("date", "")
            if not d or d < date_from or d > date_to or r.get("channel") in SKIP_CHANNELS:
                continue
            net = float(r.get("net_sales") or 0)
            gross = float(r.get("gross_sales") or 0)
            disc = float(r.get("discounts") or 0)
            units = int(r.get("units") or 0)
            brand = r.get("brand") or "Other"
            cogs_rate = brand_vw_cogs.get(brand, 0)
            price_rate = AVG_PRICE.get(brand, 25.0)
            if units == 0 and gross > 0:
                units = int(gross / price_rate)
            shopify_rev += net
            shopify_gross += gross
            shopify_disc += disc
            shopify_orders += int(r.get("orders") or 0)
            shopify_units += units
            shopify_cogs += units * cogs_rate

        amz_rev = 0
        amz_orders = 0
        for r in amazon_sales:
            d = r.get("date", "")
            if not d or d < date_from or d > date_to:
                continue
            amz_rev += float(r.get("net_sales") or 0)
            amz_orders += int(r.get("orders") or 0)

        total_rev = shopify_rev + amz_rev
        total_orders = shopify_orders + amz_orders

        # Ad spend
        total_ad_spend = 0
        ad_attributed_sales = 0
        for r in amazon_ads:
            d = r.get("date", "")
            if d and date_from <= d <= date_to:
                total_ad_spend += float(r.get("spend") or 0)
                ad_attributed_sales += float(r.get("sales") or 0)
        for r in meta_ads:
            d = r.get("date", "")
            if d and date_from <= d <= date_to:
                total_ad_spend += float(r.get("spend") or 0)
                ad_attributed_sales += float(r.get("purchase_value") or 0)
        for r in google_ads:
            d = r.get("date", "")
            if d and date_from <= d <= date_to:
                total_ad_spend += float(r.get("spend") or 0)
                ad_attributed_sales += float(r.get("conversion_value") or 0)

        # Amazon COGS estimate (brand-level)
        amz_cogs = 0
        for r in amazon_sales:
            d = r.get("date", "")
            if not d or d < date_from or d > date_to:
                continue
            brand = r.get("brand") or "Other"
            units = int(r.get("units") or 0)
            if units == 0:
                net = float(r.get("net_sales") or 0)
                if net > 0:
                    units = int(net / AVG_PRICE.get(brand, 25))
            amz_cogs += units * brand_vw_cogs.get(brand, 0)

        total_cogs = shopify_cogs + amz_cogs
        gm = total_rev - total_cogs  # GM = Total Revenue - Total COGS
        total_discounts = abs(shopify_disc)
        mkt_total = total_ad_spend + total_discounts  # + seeding (not available)
        cm = gm - mkt_total  # CM = GM - MKT Total

        return {
            "total_revenue": round(total_rev),
            "shopify_revenue": round(shopify_rev),
            "amazon_revenue": round(amz_rev),
            "total_orders": total_orders,
            "total_ad_spend": round(total_ad_spend),
            "ad_attributed_sales": round(ad_attributed_sales),
            "organic_revenue": round(total_rev - ad_attributed_sales) if total_rev > ad_attributed_sales else 0,
            "gross_margin": round(gm),
            "gm_pct": round(gm / total_rev * 100, 1) if total_rev else 0,
            "contribution_margin": round(cm),
            "cm_pct": round(cm / total_rev * 100, 1) if total_rev else 0,
            "mer": round(total_rev / total_ad_spend, 2) if total_ad_spend else 0,
            "roas": round(ad_attributed_sales / total_ad_spend, 2) if total_ad_spend else 0,
            "discount_rate": round(abs(shopify_disc) / shopify_gross * 100, 1) if shopify_gross else 0,
        }

    summary = {
        "7d": compute_period_kpi(d7, through),
        "30d": compute_period_kpi(d30, through),
        "mtd": compute_period_kpi(mtd_start, through),
    }

    # ── 6. Search Queries & Traffic Sources ───────────────────────────────────
    print("\n[6/7] Computing search queries & traffic sources...")

    # Amazon search terms (top 30 by spend) — with brand
    st_agg = defaultdict(lambda: {"impressions": 0, "clicks": 0, "spend": 0, "sales": 0, "orders": 0, "brands": set()})
    for r in search_terms:
        q = (r.get("query") or r.get("search_term") or "").strip().lower()
        if not q:
            continue
        brand = r.get("brand") or "Other"
        st_agg[q]["impressions"] += int(r.get("impressions") or 0)
        st_agg[q]["clicks"] += int(r.get("clicks") or 0)
        st_agg[q]["spend"] += float(r.get("spend") or r.get("cost") or 0)
        st_agg[q]["sales"] += float(r.get("sales") or r.get("revenue") or 0)
        st_agg[q]["orders"] += int(r.get("orders") or r.get("conversions") or 0)
        st_agg[q]["brands"].add(brand)

    # Also aggregate per-brand for brand sub-tabs
    st_brand_agg = defaultdict(lambda: defaultdict(lambda: {"impressions": 0, "clicks": 0, "spend": 0, "sales": 0, "orders": 0}))
    for r in search_terms:
        q = (r.get("query") or r.get("search_term") or "").strip().lower()
        brand = r.get("brand") or "Other"
        if not q:
            continue
        st_brand_agg[brand][q]["impressions"] += int(r.get("impressions") or 0)
        st_brand_agg[brand][q]["clicks"] += int(r.get("clicks") or 0)
        st_brand_agg[brand][q]["spend"] += float(r.get("spend") or r.get("cost") or 0)
        st_brand_agg[brand][q]["sales"] += float(r.get("sales") or r.get("revenue") or 0)
        st_brand_agg[brand][q]["orders"] += int(r.get("orders") or r.get("conversions") or 0)

    def build_search_data(agg, limit=30):
        top = sorted(agg.items(), key=lambda x: x[1]["spend"], reverse=True)[:limit]
        result = []
        for q, v in top:
            acos = round(v["spend"] / v["sales"] * 100, 1) if v["sales"] else 0
            ctr = round(v["clicks"] / v["impressions"] * 100, 2) if v["impressions"] else 0
            cvr = round(v["orders"] / v["clicks"] * 100, 2) if v["clicks"] else 0
            entry = {
                "query": q,
                "impressions": v["impressions"],
                "clicks": v["clicks"],
                "ctr": ctr,
                "spend": round(v["spend"], 2),
                "sales": round(v["sales"], 2),
                "orders": v["orders"],
                "acos": acos,
                "cvr": cvr,
            }
            if "brands" in v:
                entry["brand"] = ", ".join(sorted(v["brands"]))
            result.append(entry)
        return result

    search_data = build_search_data(st_agg)
    search_by_brand = {}
    for brand, agg in st_brand_agg.items():
        search_by_brand[brand] = build_search_data(agg, 20)

    # ── Keyword Performance (brand × period) for Ads tab ──────────────────
    # search_terms date format: "2026-03-07~2026-03-13" (weekly chunks)
    # Period: latest 1 week, latest 2 weeks, all available (~30d)
    def _parse_st_end_date(date_str):
        """Extract end date from weekly range like '2026-03-07~2026-03-13'."""
        if "~" in date_str:
            return date_str.split("~")[1].strip()
        return date_str

    # Sort date ranges to identify latest
    st_dates = sorted(set(r.get("date", "") for r in search_terms if r.get("date")))
    st_end_dates = [_parse_st_end_date(d) for d in st_dates]
    latest_end = max(st_end_dates) if st_end_dates else through

    def _kw_agg_by_brand_period(rows, date_filter=None):
        """Aggregate search terms by brand -> keyword, optionally filtering dates."""
        brand_kw = defaultdict(lambda: defaultdict(lambda: {
            "spend": 0.0, "sales": 0.0, "clicks": 0, "impressions": 0, "purchases": 0
        }))
        for r in rows:
            d = r.get("date", "")
            if date_filter and not date_filter(d):
                continue
            brand = r.get("brand", "Other")
            kw = (r.get("search_term") or r.get("query") or "").strip().lower()
            if not kw:
                continue
            brand_kw[brand][kw]["spend"] += float(r.get("spend", 0) or 0)
            brand_kw[brand][kw]["sales"] += float(r.get("sales", 0) or 0)
            brand_kw[brand][kw]["clicks"] += int(r.get("clicks", 0) or 0)
            brand_kw[brand][kw]["impressions"] += int(r.get("impressions", 0) or 0)
            brand_kw[brand][kw]["purchases"] += int(r.get("purchases", 0) or 0)
        return brand_kw

    def _kw_top_list(brand_kw_agg, limit=50):
        """Convert brand -> kw agg into sorted list per brand."""
        out = {}
        for brand, kws in brand_kw_agg.items():
            items = []
            for kw, v in kws.items():
                roas = round(v["sales"] / v["spend"], 2) if v["spend"] > 0 else 0
                cpc = round(v["spend"] / v["clicks"], 2) if v["clicks"] > 0 else 0
                ctr = round(v["clicks"] / v["impressions"] * 100, 2) if v["impressions"] > 0 else 0
                items.append({
                    "keyword": kw,
                    "spend": round(v["spend"], 2),
                    "sales": round(v["sales"], 2),
                    "clicks": v["clicks"],
                    "impressions": v["impressions"],
                    "purchases": v["purchases"],
                    "roas": roas,
                    "cpc": cpc,
                    "ctr": ctr,
                })
            items.sort(key=lambda x: x["spend"], reverse=True)
            out[brand] = items[:limit]
        return out

    # 3 periods: latest week, latest 2 weeks, all (~30d)
    def _kw_top_worst(brand_kw_agg, top_n=100, worst_n=100):
        """Split into top (by spend, for overview) and worst (by ROAS, min spend $5)."""
        out = {}
        for brand, kws in brand_kw_agg.items():
            items = []
            for kw, v in kws.items():
                roas = round(v["sales"] / v["spend"], 2) if v["spend"] > 0 else 0
                cpc = round(v["spend"] / v["clicks"], 2) if v["clicks"] > 0 else 0
                ctr = round(v["clicks"] / v["impressions"] * 100, 2) if v["impressions"] > 0 else 0
                items.append({
                    "keyword": kw,
                    "spend": round(v["spend"], 2),
                    "sales": round(v["sales"], 2),
                    "clicks": v["clicks"],
                    "impressions": v["impressions"],
                    "purchases": v["purchases"],
                    "roas": roas, "cpc": cpc, "ctr": ctr,
                })
            top = sorted(items, key=lambda x: x["roas"], reverse=True)[:top_n]
            worst = sorted([i for i in items if i["spend"] >= 5], key=lambda x: x["roas"])[:worst_n]
            out[brand] = {"top": top, "worst": worst}
        return out

    kw_all = _kw_top_worst(_kw_agg_by_brand_period(search_terms))
    kw_latest = _kw_top_worst(_kw_agg_by_brand_period(
        search_terms,
        date_filter=lambda d: d == st_dates[-1] if st_dates else True
    ))
    kw_2w = _kw_top_worst(_kw_agg_by_brand_period(
        search_terms,
        date_filter=lambda d: d in st_dates[-2:] if len(st_dates) >= 2 else True
    ))

    # Brands: Grosmimi always first
    kw_brands = sorted(set(
        r.get("brand", "") for r in search_terms if r.get("brand") and r.get("brand") != "test"
    ))
    if "Grosmimi" in kw_brands:
        kw_brands.remove("Grosmimi")
        kw_brands.insert(0, "Grosmimi")

    keyword_performance = {
        "periods": {
            "7d": {"label": "Latest 7D", "data": kw_latest},
            "14d": {"label": "Latest 14D", "data": kw_2w},
            "30d": {"label": "All (~30D)", "data": kw_all},
        },
        "brands": kw_brands,
        "date_range": f"{st_dates[0] if st_dates else '?'} to {st_dates[-1] if st_dates else '?'}",
    }
    print(f"  Keyword performance: {sum(len(v) for v in kw_all.values())} keywords, {len(kw_all)} brands")

    # GSC queries (top 20 by clicks)
    gsc_agg = defaultdict(lambda: {"impressions": 0, "clicks": 0, "position": []})
    for r in gsc:
        q = (r.get("query") or "").strip().lower()
        if not q:
            continue
        gsc_agg[q]["impressions"] += int(r.get("impressions") or 0)
        gsc_agg[q]["clicks"] += int(r.get("clicks") or 0)
        pos = float(r.get("position") or r.get("avg_position") or 0)
        if pos > 0:
            gsc_agg[q]["position"].append(pos)

    top_gsc = sorted(gsc_agg.items(), key=lambda x: x[1]["clicks"], reverse=True)[:20]
    gsc_data = []
    for q, v in top_gsc:
        avg_pos = round(sum(v["position"]) / len(v["position"]), 1) if v["position"] else 0
        ctr = round(v["clicks"] / v["impressions"] * 100, 2) if v["impressions"] else 0
        gsc_data.append({
            "query": q,
            "impressions": v["impressions"],
            "clicks": v["clicks"],
            "ctr": ctr,
            "position": avg_pos,
        })

    # ── Keyword Rankings (GSC daily position tracking) ────────────────────────
    # Top keywords by clicks, with daily position history for sparklines
    gsc_full = dk.get("gsc_daily", days=90)
    kw_daily = defaultdict(lambda: defaultdict(lambda: {"clicks": 0, "impressions": 0, "position": []}))
    for r in gsc_full:
        q = (r.get("query") or "").strip().lower()
        d = r.get("date", "")
        if not q or not d:
            continue
        kw_daily[q][d]["clicks"] += int(r.get("clicks") or 0)
        kw_daily[q][d]["impressions"] += int(r.get("impressions") or 0)
        pos = float(r.get("position") or 0)
        if pos > 0:
            kw_daily[q][d]["position"].append(pos)

    kw_dates = sorted(set(d for days_data in kw_daily.values() for d in days_data.keys()))

    def build_kw_rankings(kw_daily, kw_dates, cutoff_date, period_label):
        """Build keyword rankings for a specific date range."""
        kw_totals = {}
        for q, days_data in kw_daily.items():
            filtered = {d: dd for d, dd in days_data.items() if d >= cutoff_date}
            if not filtered:
                continue
            total_clicks = sum(dd["clicks"] for dd in filtered.values())
            total_impr = sum(dd["impressions"] for dd in filtered.values())
            all_pos = [p for dd in filtered.values() for p in dd["position"]]
            avg_pos = round(sum(all_pos) / len(all_pos), 1) if all_pos else 0
            kw_totals[q] = {"clicks": total_clicks, "impressions": total_impr, "avg_position": avg_pos}

        top_kw = sorted(kw_totals.items(), key=lambda x: x[1]["clicks"], reverse=True)[:25]
        filtered_dates = sorted(d for d in kw_dates if d >= cutoff_date)

        rankings = []
        for q, totals in top_kw:
            # Weekly average positions for trend
            weekly_positions = []
            for i in range(0, len(filtered_dates), 7):
                week_dates = filtered_dates[i:i+7]
                week_pos = []
                for d in week_dates:
                    dd = kw_daily[q].get(d, {})
                    week_pos.extend(dd.get("position", []))
                if week_pos:
                    weekly_positions.append(round(sum(week_pos) / len(week_pos), 1))
                else:
                    weekly_positions.append(None)
            ctr = round(totals["clicks"] / totals["impressions"] * 100, 2) if totals["impressions"] else 0
            rankings.append({
                "query": q,
                "clicks": totals["clicks"],
                "impressions": totals["impressions"],
                "avg_position": totals["avg_position"],
                "ctr": ctr,
                "weekly_positions": weekly_positions,
            })
        return rankings

    # Build rankings for 7D, 30D, 90D
    cutoff_7d = (today - timedelta(days=7)).isoformat()
    cutoff_30d = (today - timedelta(days=30)).isoformat()
    cutoff_90d = (today - timedelta(days=90)).isoformat()

    keyword_rankings = {
        "7d": build_kw_rankings(kw_daily, kw_dates, cutoff_7d, "7D"),
        "30d": build_kw_rankings(kw_daily, kw_dates, cutoff_30d, "30D"),
        "90d": build_kw_rankings(kw_daily, kw_dates, cutoff_90d, "90D"),
    }

    # Collect all unique keywords across all periods for avg position summary
    all_kw_set = set()
    for period_data in keyword_rankings.values():
        for k in period_data:
            all_kw_set.add(k["query"])

    # Build avg position per keyword per period (for the fixed header)
    kw_positions_summary = []
    for q in sorted(all_kw_set):
        row = {"query": q}
        for period in ["7d", "30d", "90d"]:
            match = next((k for k in keyword_rankings[period] if k["query"] == q), None)
            row["pos_" + period] = match["avg_position"] if match else None
            row["clicks_" + period] = match["clicks"] if match else 0
            row["impressions_" + period] = match["impressions"] if match else 0
        # Sort key: use 90d clicks as primary
        row["_sort"] = row["clicks_90d"]
        kw_positions_summary.append(row)

    kw_positions_summary.sort(key=lambda x: x["_sort"], reverse=True)
    for r in kw_positions_summary:
        del r["_sort"]

    # ── DataForSEO keyword volumes ─────────────────────────────────────────────
    dfseo = dk.get("dataforseo_keywords", days=7)
    # Deduplicate: latest date per keyword
    latest_dfseo = {}
    for r in dfseo:
        kw = r.get("keyword", "")
        d = r.get("date", "")
        if kw and (kw not in latest_dfseo or d > latest_dfseo[kw].get("date", "")):
            latest_dfseo[kw] = r

    keyword_volumes = []
    for kw, r in sorted(latest_dfseo.items(), key=lambda x: int(x[1].get("search_volume") or 0), reverse=True):
        vol = int(r.get("search_volume") or 0)
        # Parse monthly_searches JSON
        monthly_raw = r.get("monthly_searches", "[]")
        if isinstance(monthly_raw, str):
            try:
                monthly_parsed = json.loads(monthly_raw)
            except (json.JSONDecodeError, TypeError):
                monthly_parsed = []
        else:
            monthly_parsed = monthly_raw or []
        # Extract last 6 months of search volume
        monthly_trend = [m.get("monthly_searches", 0) for m in monthly_parsed[-6:]] if monthly_parsed else []
        keyword_volumes.append({
            "keyword": kw,
            "brand": r.get("brand", ""),
            "search_volume": vol,
            "cpc": round(float(r.get("cpc") or 0), 2),
            "competition_index": int(r.get("competition_index") or 0),
            "monthly_trend": monthly_trend,
        })

    # ── Brand Analytics (Amazon) ──────────────────────────────────────────────
    # Group by brand, show top search terms with ranking + click/conversion share
    BABY_CATEGORY_TERMS = {
        "toddler cup", "toddler cups", "sippy cup", "straw cup", "baby cup",
        "baby bottle", "ppsu bottle", "ppsu cup", "ppsu straw",
        "toddler straw cup", "baby straw cup", "sippy cups for toddlers",
        "training cup", "transition cup", "weighted straw cup",
        "baby snack", "baby rice puff", "baby puffs", "toddler snack",
        "baby rice cracker", "baby teething wafer", "organic baby snack",
        "baby wipe", "baby wipes", "water wipes",
    }

    ba_by_brand = defaultdict(list)
    ba_category = []  # Category competitor keywords
    for r in brand_analytics:
        term = (r.get("search_term") or "").strip().lower()
        if not term:
            continue
        entry = {
            "term": term,
            "rank": int(r.get("search_frequency_rank") or 0),
            "asin_rank": int(r.get("asin_rank") or 0),
            "asin_name": (r.get("asin_name") or "")[:80],
            "click_share": round(float(r.get("click_share") or 0) * 100, 2),
            "conv_share": round(float(r.get("conversion_share") or 0) * 100, 2),
            "is_ours": bool(r.get("is_ours")),
        }
        if r.get("is_ours"):
            brand = r.get("brand") or "Other"
            ba_by_brand[brand].append(entry)
        else:
            # Only keep baby-product related category terms
            if any(cat in term for cat in BABY_CATEGORY_TERMS):
                ba_category.append(entry)

    # Deduplicate BA entries per brand (take latest/best per search_term + asin_rank)
    ba_data = {}
    for brand, entries in ba_by_brand.items():
        seen = {}
        for e in entries:
            key = (e["term"], e["asin_rank"])
            if key not in seen or e["click_share"] > seen[key]["click_share"]:
                seen[key] = e
        sorted_entries = sorted(seen.values(), key=lambda x: x["rank"])[:20]
        ba_data[brand] = sorted_entries

    # Category competitors (deduplicated, top 15)
    cat_seen = {}
    for e in ba_category:
        key = (e["term"], e["asin_rank"])
        if key not in cat_seen or e["click_share"] > cat_seen[key]["click_share"]:
            cat_seen[key] = e
    ba_category_top = sorted(cat_seen.values(), key=lambda x: x["rank"])[:15]

    # GA4 traffic sources (top 15 by sessions, last 30d)
    traffic_agg = defaultdict(lambda: {"sessions": 0, "users": 0, "revenue": 0, "conversions": 0})
    for r in ga4:
        d = r.get("date", "")
        if not d or d < d30 or d > through:
            continue
        src = r.get("source") or "direct"
        med = r.get("medium") or "none"
        key = f"{src} / {med}"
        traffic_agg[key]["sessions"] += int(r.get("sessions") or 0)
        traffic_agg[key]["users"] += int(r.get("users") or r.get("active_users") or 0)
        traffic_agg[key]["revenue"] += float(r.get("revenue") or r.get("purchase_revenue") or 0)
        traffic_agg[key]["conversions"] += int(r.get("conversions") or r.get("purchases") or 0)

    top_traffic = sorted(traffic_agg.items(), key=lambda x: x[1]["sessions"], reverse=True)[:15]
    traffic_data = []
    for src, v in top_traffic:
        conv_rate = round(v["conversions"] / v["sessions"] * 100, 2) if v["sessions"] else 0
        traffic_data.append({
            "source": src,
            "sessions": v["sessions"],
            "users": v["users"],
            "revenue": round(v["revenue"], 2),
            "conversions": v["conversions"],
            "conv_rate": conv_rate,
        })

    # ── Amazon Sessions (SP-API Sales & Traffic Report) ──────────────────
    def _fetch_amz_sessions(days=30):
        """Fetch Amazon sessions/pageViews from GET_SALES_AND_TRAFFIC_REPORT."""
        try:
            import requests as _req, gzip as _gz, time as _tm
            seller_cfg = {
                "refresh_token": os.getenv("AMZ_SP_REFRESH_TOKEN_GROSMIMI", ""),
                "client_id": os.getenv("AMZ_SP_GROSMIMI_CLIENT_ID") or os.getenv("AMZ_SP_CLIENT_ID", ""),
                "client_secret": os.getenv("AMZ_SP_GROSMIMI_CLIENT_SECRET") or os.getenv("AMZ_SP_CLIENT_SECRET", ""),
            }
            tok = _req.post("https://api.amazon.com/auth/o2/token", data={
                "grant_type": "refresh_token", **seller_cfg,
            }, timeout=15).json().get("access_token")
            if not tok:
                return {}
            hdrs = {"x-amz-access-token": tok, "Content-Type": "application/json"}
            start = (today - timedelta(days=days)).strftime("%Y-%m-%dT00:00:00Z")
            end = today.strftime("%Y-%m-%dT00:00:00Z")
            r = _req.post("https://sellingpartnerapi-na.amazon.com/reports/2021-06-30/reports",
                headers=hdrs, json={
                    "reportType": "GET_SALES_AND_TRAFFIC_REPORT",
                    "marketplaceIds": ["ATVPDKIKX0DER"],
                    "reportOptions": {"dateGranularity": "DAY", "asinGranularity": "SKU"},
                    "dataStartTime": start, "dataEndTime": end,
                }, timeout=30)
            if r.status_code >= 300:
                return {}
            rid = r.json().get("reportId")
            doc_id = None
            for _ in range(30):
                _tm.sleep(10)
                r2 = _req.get(f"https://sellingpartnerapi-na.amazon.com/reports/2021-06-30/reports/{rid}",
                    headers=hdrs, timeout=30)
                st = r2.json().get("processingStatus")
                if st == "DONE":
                    doc_id = r2.json().get("reportDocumentId")
                    break
                elif st in ("CANCELLED", "FATAL"):
                    return {}
            if not doc_id:
                return {}
            r3 = _req.get(f"https://sellingpartnerapi-na.amazon.com/reports/2021-06-30/documents/{doc_id}",
                headers=hdrs, timeout=30)
            doc = r3.json()
            r4 = _req.get(doc["url"], timeout=60)
            content = r4.content
            if doc.get("compressionAlgorithm") == "GZIP":
                content = _gz.decompress(content)
            data = json.loads(content.decode("utf-8"))
            # Aggregate daily
            result = {"sessions": 0, "pageViews": 0, "days": {}, "asin_days": defaultdict(dict)}
            for e in data.get("salesAndTrafficByDate", []):
                d = e.get("date", "")
                t = e.get("trafficByDate", {})
                sess = t.get("sessions", 0)
                pv = t.get("pageViews", 0)
                result["sessions"] += sess
                result["pageViews"] += pv
                result["days"][d] = {"sessions": sess, "pageViews": pv}
            # Per-ASIN daily for hero category sessions
            for e in data.get("salesAndTrafficByAsin", []):
                asin = e.get("parentAsin") or e.get("childAsin", "")
                d = e.get("date", "")
                t = e.get("trafficByAsin", {})
                if asin and d:
                    result["asin_days"][asin][d] = {
                        "sessions": t.get("sessions", 0),
                        "pageViews": t.get("pageViews", 0),
                    }
            return result
        except Exception as ex:
            print(f"  [WARN] Amazon sessions fetch failed: {ex}")
            return {}

    print("  Fetching Amazon sessions...")
    amz_sessions_30d = _fetch_amz_sessions(30)
    amz_sessions_7d = _fetch_amz_sessions(7)
    if amz_sessions_30d:
        print(f"  Amazon sessions (30d): {amz_sessions_30d['sessions']:,} sessions, {amz_sessions_30d['pageViews']:,} pageViews")
    amz_sessions_raw = _fetch_amz_sessions(90)
    if amz_sessions_raw:
        print(f"  Amazon sessions (90d): {amz_sessions_raw['sessions']:,} sessions, {len(amz_sessions_raw.get('asin_days',{}))} ASINs")

    # ── Ad Creative Breakdown (WL/Image/Video) ─────────────────────────────
    def _classify_creative(ad_name):
        al = (ad_name or "").lower()
        if "video" in al: return "Video"
        if "image" in al or "img" in al: return "Image"
        if "wl" in al or "whitelabel" in al or "white label" in al: return "Whitelabel"
        if "carousel" in al: return "Carousel"
        return "Other"

    # Build per-campaign creative breakdown
    camp_creative_agg = defaultdict(lambda: defaultdict(lambda: {
        "spend": 0.0, "clicks": 0, "impressions": 0, "purchases": 0,
        "ads": []  # list of {ad_id, ad_name}
    }))
    seen_ads = set()
    for r in meta_ads:
        cn = r.get("campaign_name", "")
        an = r.get("ad_name", "")
        aid = r.get("ad_id", "")
        ctype = _classify_creative(an)
        agg = camp_creative_agg[cn][ctype]
        agg["spend"] += float(r.get("spend", 0) or 0)
        agg["clicks"] += int(r.get("clicks", 0) or 0)
        agg["impressions"] += int(r.get("impressions", 0) or 0)
        agg["purchases"] += int(r.get("purchases", 0) or 0)
        ad_key = f"{cn}|{aid}"
        if ad_key not in seen_ads and an:
            seen_ads.add(ad_key)
            agg["ads"].append({"id": aid, "name": an[:80]})

    # Build output: top campaigns with creative breakdown
    ad_creative_data = []
    for cn in sorted(camp_creative_agg, key=lambda x: -sum(v["spend"] for v in camp_creative_agg[x].values()))[:20]:
        types = []
        for ctype, v in sorted(camp_creative_agg[cn].items(), key=lambda x: -x[1]["spend"]):
            cpc = round(v["spend"] / v["clicks"], 2) if v["clicks"] else 0
            types.append({
                "type": ctype,
                "spend": round(v["spend"]),
                "clicks": v["clicks"],
                "cpc": cpc,
                "purchases": v["purchases"],
                "ads": v["ads"][:5],  # top 5 ads per type
            })
        total_spend = sum(v["spend"] for v in camp_creative_agg[cn].values())
        brand = "Unknown"
        for r in meta_ads:
            if r.get("campaign_name") == cn:
                brand = r.get("brand", "Unknown")
                break
        ad_creative_data.append({
            "campaign": cn[:60],
            "campaign_id": next((r.get("campaign_id","") for r in meta_ads if r.get("campaign_name")==cn), ""),
            "brand": brand,
            "total_spend": round(total_spend),
            "types": types,
        })

    # ── Amazon Attribution Data ────────────────────────────────────────────
    # Fetch Attribution report: Meta Traffic -> Amazon conversion tracking
    def _load_attribution():
        try:
            import requests as _req
            tok = _req.post("https://api.amazon.com/auth/o2/token", data={
                "grant_type": "refresh_token",
                "client_id": os.getenv("AMZ_ADS_CLIENT_ID", ""),
                "client_secret": os.getenv("AMZ_ADS_CLIENT_SECRET", ""),
                "refresh_token": os.getenv("AMZ_ADS_REFRESH_TOKEN", ""),
            }, timeout=15).json().get("access_token")
            if not tok:
                return {}
            hdrs = {
                "Authorization": f"Bearer {tok}",
                "Amazon-Advertising-API-ClientId": os.getenv("AMZ_ADS_CLIENT_ID", ""),
                "Amazon-Advertising-API-Scope": "1094731557245186",  # Grosmimi Attribution
                "Content-Type": "application/json",
            }
            all_rpts = []
            cursor = ""
            for _ in range(10):
                r = _req.post("https://advertising-api.amazon.com/attribution/report",
                    headers=hdrs, json={
                        "reportType": "PERFORMANCE", "count": 300,
                        "startDate": (today - timedelta(days=60)).strftime("%Y%m%d"),
                        "endDate": today.strftime("%Y%m%d"),
                        "groupBy": "CAMPAIGN", "cursorId": cursor,
                    }, timeout=30)
                if r.status_code != 200:
                    break
                data = r.json()
                rpts = data.get("reports", [])
                all_rpts.extend(rpts)
                cursor = data.get("cursorId", "")
                if not cursor or not rpts:
                    break

            # Aggregate by campaign
            camp = defaultdict(lambda: {"clicks": 0, "dpv": 0, "atc": 0,
                                         "purchases": 0, "sales": 0.0, "brb": 0.0})
            for rr in all_rpts:
                cid = rr.get("campaignId", "")
                camp[cid]["clicks"] += int(rr.get("Click-throughs", 0) or 0)
                camp[cid]["dpv"] += int(rr.get("attributedDetailPageViewsClicks14d", 0) or 0)
                camp[cid]["atc"] += int(rr.get("attributedAddToCartClicks14d", 0) or 0)
                camp[cid]["purchases"] += int(rr.get("attributedPurchases14d", 0) or 0)
                camp[cid]["sales"] += float(rr.get("attributedSales14d", 0) or 0)
                camp[cid]["brb"] += float(rr.get("brb_bonus_amount", 0) or 0)
            return dict(camp)
        except Exception as e:
            print(f"  [WARN] Attribution load failed: {e}")
            return {}

    attribution = _load_attribution()
    if attribution:
        attr_total = {k: sum(v[k] for v in attribution.values()) for k in ["clicks", "purchases", "sales", "brb"]}
        print(f"  Attribution: {len(attribution)} campaigns, {attr_total['clicks']:,} clicks, ${attr_total['sales']:,.0f} sales, ${attr_total['brb']:,.0f} BRB")

        # Inject Attribution sales into Meta Traffic platform breakdown (Grosmimi)
        mt_gros = brand_ad_by_platform.get("Meta Traffic", {}).get("Grosmimi", {})
        total_mt_spend = sum(v.get("spend", 0) for v in mt_gros.values())
        if total_mt_spend > 0 and attr_total["sales"] > 0:
            for m, v in mt_gros.items():
                ratio = v.get("spend", 0) / total_mt_spend
                v["sales"] = round(attr_total["sales"] * ratio, 2)
            print(f"  Attribution -> Meta Traffic Grosmimi (monthly): ${attr_total['sales']:,.0f} across {len(mt_gros)} months")

        # Weekly too
        mt_gros_wk = brand_ad_by_platform_wk.get("Meta Traffic", {}).get("Grosmimi", {})
        total_mt_wk_spend = sum(v.get("spend", 0) for v in mt_gros_wk.values())
        if total_mt_wk_spend > 0 and attr_total["sales"] > 0:
            for w, v in mt_gros_wk.items():
                ratio = v.get("spend", 0) / total_mt_wk_spend
                v["sales"] = round(attr_total["sales"] * ratio, 2)
            print(f"  Attribution -> Meta Traffic Grosmimi (weekly): across {len(mt_gros_wk)} weeks")
    else:
        print("  Attribution: no data")

    # ── Channel Traffic Breakdown (Amazon vs Onzenna) ────────────────────────
    # GA4 = Onzenna D2C traffic. Amazon = Ads clicks + Meta Traffic clicks + organic (estimated).
    def _build_channel_traffic(ga4_rows, amz_ads_rows, meta_rows, google_rows, days, amz_sessions_data=None, amz_sales_rows=None):
        cutoff = (today - timedelta(days=days)).isoformat()

        # Onzenna (GA4 channel groupings)
        onz = defaultdict(lambda: {"sessions": 0, "purchases": 0})
        for r in ga4_rows:
            if r.get("date", "") < cutoff:
                continue
            ch = r.get("channel_grouping", "Other")
            onz[ch]["sessions"] += int(r.get("sessions", 0) or 0)
            onz[ch]["purchases"] += int(r.get("purchases", 0) or 0)

        # Amazon Ads: clicks, spend, sales, purchases
        amz_ad = {"clicks": 0, "spend": 0.0, "sales": 0.0, "purchases": 0, "impressions": 0}
        for r in amz_ads_rows:
            if r.get("date", "") < cutoff:
                continue
            amz_ad["clicks"] += int(r.get("clicks", 0) or 0)
            amz_ad["spend"] += float(r.get("spend", 0) or 0)
            amz_ad["sales"] += float(r.get("sales", 0) or 0)
            amz_ad["purchases"] += int(r.get("purchases", 0) or 0)
            amz_ad["impressions"] += int(r.get("impressions", 0) or 0)

        # Meta: split traffic vs CVR
        meta_traffic = {"clicks": 0, "spend": 0.0, "purchases": 0, "purchase_value": 0.0}
        meta_cvr = {"clicks": 0, "spend": 0.0, "purchases": 0, "purchase_value": 0.0}
        meta_traffic_by_camp = defaultdict(lambda: {"clicks": 0, "spend": 0.0, "ads": defaultdict(lambda: {"spend": 0.0, "clicks": 0, "id": "", "name": ""})})
        for r in meta_rows:
            if r.get("date", "") < cutoff:
                continue
            cn = (r.get("campaign_name", "") or "")
            cl = cn.lower()
            if "traffic" in cl or "amz" in cl:
                meta_traffic["clicks"] += int(r.get("clicks", 0) or 0)
                meta_traffic["spend"] += float(r.get("spend", 0) or 0)
                meta_traffic["purchases"] += int(r.get("purchases", 0) or 0)
                meta_traffic["purchase_value"] += float(r.get("purchase_value", 0) or 0)
                meta_traffic_by_camp[cn]["clicks"] += int(r.get("clicks", 0) or 0)
                meta_traffic_by_camp[cn]["spend"] += float(r.get("spend", 0) or 0)
                # Ad-level for asset breakdown
                aid = r.get("ad_id", "")
                an = r.get("ad_name", "")
                if aid:
                    meta_traffic_by_camp[cn]["ads"][aid]["spend"] += float(r.get("spend", 0) or 0)
                    meta_traffic_by_camp[cn]["ads"][aid]["clicks"] += int(r.get("clicks", 0) or 0)
                    meta_traffic_by_camp[cn]["ads"][aid]["id"] = aid
                    meta_traffic_by_camp[cn]["ads"][aid]["name"] = an
            else:
                meta_cvr["clicks"] += int(r.get("clicks", 0) or 0)
                meta_cvr["spend"] += float(r.get("spend", 0) or 0)
                meta_cvr["purchases"] += int(r.get("purchases", 0) or 0)
                meta_cvr["purchase_value"] += float(r.get("purchase_value", 0) or 0)

        # Google Ads
        gads = {"clicks": 0, "spend": 0.0, "conversions": 0.0, "conversion_value": 0.0}
        for r in google_rows:
            if r.get("date", "") < cutoff:
                continue
            gads["clicks"] += int(r.get("clicks", 0) or 0)
            gads["spend"] += float(r.get("spend", 0) or 0)
            gads["conversions"] += float(r.get("conversions", 0) or 0)
            gads["conversion_value"] += float(r.get("conversion_value", 0) or 0)

        GA4_MAP = {
            "Cross-network": "Meta + Google (Cross-network)",
            "Paid Social": "Meta Ads (Paid Social)",
            "Paid Search": "Google Ads (Paid Search)",
            "Direct": "Direct",
            "Organic Search": "Organic Search (SEO)",
            "Organic Social": "Organic Social (IG/TikTok/etc)",
            "Referral": "Referral",
            "Email": "Email (Klaviyo)",
            "Unassigned": "Unassigned",
            "Display": "Display Ads",
        }
        onz_mapped = {}
        for ch, v in onz.items():
            onz_mapped[GA4_MAP.get(ch, ch)] = v

        onz_total = sum(v["sessions"] for v in onz.values())
        onz_purchases = sum(v["purchases"] for v in onz.values())
        amz_total_clicks = amz_ad["clicks"] + meta_traffic["clicks"]

        # Total Amazon orders/sales for organic estimation
        _amz_total_orders = 0
        _amz_total_sales = 0.0
        if amz_sales_rows:
            for r in amz_sales_rows:
                if r.get("date", "") >= cutoff:
                    _amz_total_orders += int(r.get("orders", 0) or 0)
                    _amz_total_sales += float(r.get("gross_sales", 0) or r.get("net_sales", 0) or 0)

        def _build_attr_campaigns(attr_data, meta_spend_by_camp):
            """Build attribution campaign list with matched Meta spend."""
            camps = []
            for attr_name, v in sorted(attr_data.items(), key=lambda x: -x[1]["sales"]):
                if v["sales"] == 0 and v["clicks"] < 100:
                    continue
                # Find matching Meta campaign spend
                matched_spend = 0
                best_meta = _match_attribution_reverse(attr_name, meta_spend_by_camp)
                if best_meta:
                    matched_spend = best_meta["spend"]
                camp_roas = round(v["sales"] / matched_spend, 1) if matched_spend > 0 else 0
                camp_roas_adj = round((v["sales"] + v["brb"]) / matched_spend, 1) if matched_spend > 0 else 0
                # Ad asset breakdown from matched Meta campaign
                assets = []
                if best_meta and best_meta.get("ads"):
                    type_agg = defaultdict(lambda: {"spend": 0.0, "clicks": 0, "ads": []})
                    for aid, ad in best_meta["ads"].items():
                        atype = _classify_creative(ad.get("name", ""))
                        type_agg[atype]["spend"] += ad["spend"]
                        type_agg[atype]["clicks"] += ad["clicks"]
                        type_agg[atype]["ads"].append({
                            "id": ad["id"],
                            "name": ad["name"][:70],
                        })
                    for atype, ta in sorted(type_agg.items(), key=lambda x: -x[1]["spend"]):
                        cpc = round(ta["spend"] / ta["clicks"], 2) if ta["clicks"] else 0
                        assets.append({
                            "type": atype,
                            "spend": round(ta["spend"]),
                            "clicks": ta["clicks"],
                            "cpc": cpc,
                            "ads": ta["ads"][:5],
                        })

                camps.append({
                    "name": attr_name[:60],
                    "sales": round(v["sales"]),
                    "purchases": v["purchases"],
                    "clicks": v["clicks"],
                    "brb": round(v["brb"], 2),
                    "spend": round(matched_spend),
                    "roas": camp_roas,
                    "roas_adj": camp_roas_adj,
                    "assets": assets,
                })
            return camps

        def _match_attribution_reverse(attr_name, meta_camps):
            """Match Attribution campaign name to Meta campaign spend."""
            def _norm(s): return s.lower().replace("|"," ").replace("_"," ").replace("-"," ").replace("\t"," ").strip()
            an = _norm(attr_name)
            STOP = {"amz", "traffic", "wl", "meta", "target", "the", ""}
            attr_tokens = set(an.split()) - STOP
            best = None
            best_score = 0
            for meta_name, meta_data in meta_camps.items():
                mn = _norm(meta_name)
                meta_tokens = set(mn.split()) - STOP
                overlap = attr_tokens & meta_tokens
                score = len(overlap)
                if any(t for t in overlap if t[:4].isdigit() and len(t) >= 6):
                    score += 3
                if any(t for t in overlap if t in ("dental", "dentalmom", "stainless", "strawcup",
                                                     "spring", "sale", "livfuselli")):
                    score += 2
                if score > best_score:
                    best_score = score
                    best = meta_data
            return best if best_score >= 3 else None

        # Attribution data (all-time from API, not date-filtered — snapshot)
        attr_total_sales = sum(v["sales"] for v in attribution.values())
        attr_total_purchases = sum(v["purchases"] for v in attribution.values())
        attr_total_brb = sum(v["brb"] for v in attribution.values())
        meta_spend_for_attr = meta_traffic["spend"]
        attr_roas = round(attr_total_sales / meta_spend_for_attr, 1) if meta_spend_for_attr else 0
        attr_roas_adj = round((attr_total_sales + attr_total_brb) / meta_spend_for_attr, 1) if meta_spend_for_attr else 0
        attr_summary = {
            "sales": round(attr_total_sales),
            "purchases": attr_total_purchases,
            "brb": round(attr_total_brb),
            "roas": attr_roas,
            "roas_with_brb": attr_roas_adj,
            "campaigns": _build_attr_campaigns(attribution, meta_traffic_by_camp),
        }

        def _src(name, clicks, spend, sales, purchases, pct):
            cpc = round(spend / clicks, 2) if clicks else 0
            roas = round(sales / spend, 1) if spend else 0
            return {"source": name, "clicks": clicks, "spend": round(spend),
                    "sales": round(sales), "purchases": purchases,
                    "cpc": cpc, "roas": roas, "pct": pct}

        return {
            "onzenna": {
                "total_sessions": onz_total,
                "total_purchases": onz_purchases,
                "total_ad_spend": round(meta_cvr["spend"] + gads["spend"]),
                "sources": sorted(
                    [{"source": k, "sessions": v["sessions"], "purchases": v["purchases"],
                      "pct": round(v["sessions"] / onz_total * 100, 1) if onz_total else 0}
                     for k, v in onz_mapped.items() if v["sessions"] > 0],
                    key=lambda x: -x["sessions"]
                ),
                "ad_detail": [
                    _src("Meta CVR", meta_cvr["clicks"], meta_cvr["spend"],
                         meta_cvr["purchase_value"], meta_cvr["purchases"],
                         round(meta_cvr["clicks"] / (meta_cvr["clicks"] + gads["clicks"]) * 100, 1) if (meta_cvr["clicks"] + gads["clicks"]) else 0),
                    _src("Google Ads", gads["clicks"], gads["spend"],
                         gads["conversion_value"], int(gads["conversions"]),
                         round(gads["clicks"] / (meta_cvr["clicks"] + gads["clicks"]) * 100, 1) if (meta_cvr["clicks"] + gads["clicks"]) else 0),
                ],
            },
            "amazon": {
                "total_clicks": amz_total_clicks,
                "total_sessions": (amz_sessions_30d if days >= 30 else amz_sessions_7d).get("sessions", 0),
                "total_pageviews": (amz_sessions_30d if days >= 30 else amz_sessions_7d).get("pageViews", 0),
                "total_spend": round(amz_ad["spend"] + meta_traffic["spend"]),
                "total_sales": round(amz_ad["sales"]),
                "total_purchases": amz_ad["purchases"],
                "sources": [
                    _src("Amazon Ads (Sponsored)", amz_ad["clicks"], amz_ad["spend"],
                         amz_ad["sales"], amz_ad["purchases"],
                         round(amz_ad["clicks"] / amz_total_clicks * 100, 1) if amz_total_clicks else 0),
                    _src("Meta Traffic (AMZ landing)", meta_traffic["clicks"], meta_traffic["spend"],
                         attr_total_sales, attr_total_purchases,
                         round(meta_traffic["clicks"] / amz_total_clicks * 100, 1) if amz_total_clicks else 0),
                    (lambda internal_sess, organic_orders, organic_sales: {
                        "source": "Amazon Internal (Organic + Ads)",
                        "clicks": internal_sess, "spend": 0,
                        "sales": round(organic_sales),
                        "purchases": organic_orders, "cpc": 0, "roas": 0,
                        "pct": round(internal_sess / (amz_sessions_data or {}).get("sessions", 1) * 100, 1) if internal_sess > 0 else 0,
                        "note_text": "Sessions from Amazon search + direct. Ads clicks happen within these sessions.",
                    })(
                        max(0, (amz_sessions_data or {}).get("sessions", 0) - meta_traffic["clicks"]),
                        max(0, _amz_total_orders - amz_ad["purchases"] - attr_total_purchases),
                        max(0, _amz_total_sales - amz_ad["sales"] - attr_total_sales),
                    ) if (amz_sessions_data or {}).get("sessions", 0) > 0 else
                    {"source": "Amazon Internal (Organic)", "clicks": 0, "spend": 0, "sales": 0,
                     "purchases": 0, "cpc": 0, "roas": 0, "pct": 0, "note": "n.m."},
                ],
                "attribution": attr_summary,
            },
            "meta_detail": {
                "cvr_clicks": meta_cvr["clicks"],
                "traffic_clicks": meta_traffic["clicks"],
                "total": meta_cvr["clicks"] + meta_traffic["clicks"],
            },
            "google_clicks": gads["clicks"],
        }

    channel_traffic = {
        "1d": _build_channel_traffic(ga4, amazon_ads, meta_ads, google_ads, 1, {}, amazon_sales),
        "7d": _build_channel_traffic(ga4, amazon_ads, meta_ads, google_ads, 7, amz_sessions_7d, amazon_sales),
        "30d": _build_channel_traffic(ga4, amazon_ads, meta_ads, google_ads, 30, amz_sessions_30d, amazon_sales),
    }
    onz_30 = channel_traffic["30d"]["onzenna"]["total_sessions"]
    amz_30 = channel_traffic["30d"]["amazon"]["total_clicks"]
    print(f"  Channel traffic (30d): Onzenna {onz_30:,} sessions, Amazon {amz_30:,} clicks")

    # ── 7. Build monthly waterfall data ───────────────────────────────────────
    print("\n[7/7] Building monthly waterfall...")

    def monthly_val(data_dict, month, field="net"):
        """Sum field across all entities for a month."""
        total = 0
        for entity_months in data_dict.values():
            v = entity_months.get(month, {})
            total += v.get(field, 0) if isinstance(v, dict) else 0
        return total

    def proj_array(vals):
        """Add full-month projection for partial last month."""
        if is_partial and months and months[-1] == current_month and len(vals) > 0:
            return [round(v * fm_multiplier) if i == len(vals) - 1 else v for i, v in enumerate(vals)]
        return vals[:]

    def proj_for(vals, month_list):
        """Full-month projection for a specific month list."""
        if is_partial and month_list and month_list[-1] == current_month and len(vals) > 0:
            return [round(v * fm_multiplier) if i == len(vals) - 1 else v for i, v in enumerate(vals)]
        return vals[:]

    # Build output arrays
    brand_rev_out = {}
    for brand in BRAND_ORDER:
        vals = []
        for m in months:
            v = brand_monthly.get(brand, {}).get(m, {})
            amz_v = amz_brand_monthly.get(brand, {}).get(m, {})
            vals.append(round(v.get("net", 0) + amz_v.get("net", 0)))
        if any(v > 0 for v in vals):
            brand_rev_out[brand] = {
                "monthly": vals,
                "monthly_proj": proj_array(vals),
                "color": BRAND_COLORS.get(brand, "#94a3b8"),
            }

    channel_rev_out = {}
    for ch in ["Onzenna D2C", "Amazon MP", "Amazon FBA MCF", "TikTok Shop", "Target+", "B2B"]:
        vals = [round(channel_monthly.get(ch, {}).get(m, {}).get("net", 0)) for m in months]
        if any(v > 0 for v in vals):
            channel_rev_out[ch] = {
                "monthly": vals,
                "monthly_proj": proj_array(vals),
                "color": CHANNEL_COLORS.get(ch, "#94a3b8"),
            }

    ad_out = {}
    for platform in ["Amazon Ads", "Meta CVR", "Meta Traffic", "Google Ads"]:
        spend_vals = [round(ad_monthly.get(platform, {}).get(m, {}).get("spend", 0)) for m in months]
        sales_vals = [round(ad_monthly.get(platform, {}).get(m, {}).get("sales", 0)) for m in months]
        impr_vals = [ad_monthly.get(platform, {}).get(m, {}).get("impressions", 0) for m in months]
        click_vals = [ad_monthly.get(platform, {}).get(m, {}).get("clicks", 0) for m in months]
        if any(v > 0 for v in spend_vals):
            ad_out[platform] = {
                "spend": spend_vals,
                "spend_proj": proj_array(spend_vals),
                "sales": sales_vals,
                "sales_proj": proj_array(sales_vals),
                "impressions": impr_vals,
                "clicks": click_vals,
                "color": AD_COLORS.get(platform, "#94a3b8"),
            }

    # ── Ads Landing Channel TROAS ─────────────────────────────────────────
    # Onzenna channel: Google Ads + Meta CVR -> Shopify D2C revenue
    # Amazon channel: Amazon Ads + Meta Traffic -> Amazon MP revenue
    ads_landing = {}
    onz_spend = [0] * len(months)
    amz_spend = [0] * len(months)
    for i, m in enumerate(months):
        for p in ["Google Ads", "Meta CVR"]:
            onz_spend[i] += ad_monthly.get(p, {}).get(m, {}).get("spend", 0)
        for p in ["Amazon Ads", "Meta Traffic"]:
            amz_spend[i] += ad_monthly.get(p, {}).get(m, {}).get("spend", 0)
    onz_rev = [round(channel_monthly.get("Onzenna D2C", {}).get(m, {}).get("net", 0)) for m in months]
    amz_rev = [round(channel_monthly.get("Amazon MP", {}).get(m, {}).get("net", 0)) for m in months]
    ads_landing["Onzenna"] = {
        "spend": [round(v) for v in onz_spend],
        "spend_proj": proj_array([round(v) for v in onz_spend]),
        "revenue": onz_rev,
        "revenue_proj": proj_array(onz_rev),
        "platforms": "Google Ads + Meta CVR",
        "color": "#6366f1",
    }
    ads_landing["Amazon"] = {
        "spend": [round(v) for v in amz_spend],
        "spend_proj": proj_array([round(v) for v in amz_spend]),
        "revenue": amz_rev,
        "revenue_proj": proj_array(amz_rev),
        "platforms": "Amazon Ads + Meta Traffic",
        "color": "#f59e0b",
    }

    # Brand performance (ad + total sales, from 2026-01+)
    brand_perf_out = {}
    for brand in BRAND_ORDER:
        total_sales_vals = []
        ad_spend_vals = []
        ad_sales_vals = []
        for m in ad_months:
            shopify_net = brand_monthly.get(brand, {}).get(m, {}).get("net", 0)
            amz_net = amz_brand_monthly.get(brand, {}).get(m, {}).get("net", 0)
            total_sales_vals.append(round(shopify_net + amz_net))
            ad_spend_vals.append(round(brand_ad_monthly.get(brand, {}).get(m, {}).get("spend", 0)))
            ad_sales_vals.append(round(brand_ad_monthly.get(brand, {}).get(m, {}).get("sales", 0)))
        organic_vals = [max(0, t - a) for t, a in zip(total_sales_vals, ad_sales_vals)]
        if any(v > 0 for v in total_sales_vals) or any(v > 0 for v in ad_spend_vals):
            entry = {
                "total_sales": total_sales_vals,
                "total_sales_proj": proj_for(total_sales_vals, ad_months),
                "ad_spend": ad_spend_vals,
                "ad_spend_proj": proj_for(ad_spend_vals, ad_months),
                "ad_sales": ad_sales_vals,
                "ad_sales_proj": proj_for(ad_sales_vals, ad_months),
                "organic": organic_vals,
                "organic_proj": proj_for(organic_vals, ad_months),
                "color": BRAND_COLORS.get(brand, "#94a3b8"),
            }
            brand_perf_out[brand] = entry

    # Total monthly revenue & costs (GM = Rev - COGS, CM = GM - MKT)
    total_rev_monthly = []
    total_cogs_monthly = []
    total_gm_monthly = []
    total_ad_spend_monthly = []
    total_disc_monthly = []
    _shop_disc_arr = []   # Shopify discounts breakdown
    _amz_fee_arr = []     # Amazon referral fee (15%)
    _fba_arr = []         # FBA fulfillment (None = n.m. pre-Oct)
    _variable_costs_arr = []  # Ref fee + FBA
    _cm_before_mkt_arr = []   # CM before marketing
    _inf_paid_arr = []    # Influencer PAID (PayPal)
    _inf_nonpaid_arr = [] # Influencer NON-PAID (COGS + shipping)
    total_seeding_monthly = []
    total_mkt_monthly = []
    total_cm_monthly = []

    # Pre-build FBA monthly cache (FBA section builds later, but P&L needs it now)
    _fba_monthly_cache = defaultdict(float)
    _fba_snapshot_path = Path(".tmp/datakeeper/amazon_fba_fees.json")
    _asin_fba = {}
    if _fba_snapshot_path.exists():
        try:
            _fba_list = json.loads(_fba_snapshot_path.read_text(encoding="utf-8"))
            for item in _fba_list:
                a = item.get("asin", "")
                f = float(item.get("fba_fee", 0) or 0)
                if a and f > 0:
                    _asin_fba[a] = f
        except Exception:
            pass
    if _asin_fba:
        for r in amazon_sku:
            d = r.get("date", "")
            if not d or d > through:
                continue
            asin = r.get("asin", "")
            fee = _asin_fba.get(asin, 0)
            if fee > 0:
                _fba_monthly_cache[d[:7]] += fee * int(r.get("units", 0) or 0)

    for m in months:
        # Total revenue = Shopify GROSS + Amazon GROSS
        # Both = customer-facing price (what buyers paid).
        # Shopify disc = promo codes. Amazon gross-net = referral fee (shown as Selling Fee in Channel P&L).
        shopify_gross = sum(brand_monthly.get(b, {}).get(m, {}).get("gross", 0) for b in BRAND_ORDER + ["Other"])
        shopify_disc = sum(abs(brand_monthly.get(b, {}).get(m, {}).get("disc", 0)) for b in BRAND_ORDER + ["Other"])
        amz_gross = sum(amz_brand_monthly.get(b, {}).get(m, {}).get("gross", 0) for b in BRAND_ORDER + ["Other"])
        rev = shopify_gross + amz_gross

        # COGS from SKU-level data (or fallback to brand avg)
        cogs = 0
        for b in BRAND_ORDER + ["Other"]:
            cogs += shop_sku_cogs.get(b, {}).get(m, 0)
            cogs += amz_sku_cogs.get(b, {}).get(m, 0)
        # Fallback: if SKU data missing for this month, use brand avg with units
        if cogs == 0 and rev > 0:
            for b in BRAND_ORDER:
                v = brand_monthly.get(b, {}).get(m, {})
                units = v.get("units", 0)
                if units == 0 and v.get("gross", 0) > 0:
                    units = int(v["gross"] / AVG_PRICE.get(b, 25))
                cogs += units * brand_vw_cogs.get(b, brand_vw_cogs.get(b, 0))
                amz_v = amz_brand_monthly.get(b, {}).get(m, {})
                amz_units = amz_v.get("units", 0)
                if amz_units == 0 and amz_v.get("net", 0) > 0:
                    amz_units = int(amz_v["net"] / AVG_PRICE.get(b, 25))
                cogs += amz_units * brand_vw_cogs.get(b, brand_vw_cogs.get(b, 0))
        elif cogs > 0 and rev > 0:
            # Partial SKU data fix (e.g., Oct 2025 SKU starts mid-month):
            # Scale COGS up by daily_units / sku_units ratio
            sku_u = amz_sku_units_m.get(m, 0) + shop_sku_units_m.get(m, 0)
            daily_u = sum(
                amz_brand_monthly.get(b, {}).get(m, {}).get("units", 0)
                + brand_monthly.get(b, {}).get(m, {}).get("units", 0)
                for b in BRAND_ORDER + ["Other"]
            )
            if sku_u > 0 and daily_u > sku_u * 1.2:
                # SKU data covers less than ~83% of daily units -> scale up
                scale = daily_u / sku_u
                cogs = round(cogs * scale)

        gm = rev - cogs  # GM = Total Revenue (gross) - Total COGS

        # Variable Costs (not marketing — cost of doing business on each platform)
        amz_net = sum(amz_brand_monthly.get(b, {}).get(m, {}).get("net", 0) for b in BRAND_ORDER + ["Other"])
        amz_ref_fee = amz_gross - amz_net  # Amazon Referral Fee (~15%)
        fba_fulfill = _fba_monthly_cache.get(m, 0)  # FBA Fulfillment
        variable_costs = amz_ref_fee + fba_fulfill
        _amz_fee_arr.append(round(amz_ref_fee))
        _fba_arr.append(round(fba_fulfill) if fba_fulfill > 0 or m >= "2025-10" else None)

        cm_before_mkt = gm - variable_costs  # CM before MKT = GM - Variable Costs

        # MKT Cost 1: Ad Spend (with Amazon Ads backfill for pre-Dec 2025)
        amz_ad = ad_monthly.get("Amazon Ads", {}).get(m, {}).get("spend", 0)
        if amz_ad == 0 and m in amz_ads_backfill:
            amz_ad = amz_ads_backfill[m]
        onz_ad = sum(
            ad_monthly.get(p, {}).get(m, {}).get("spend", 0)
            for p in ["Meta CVR", "Meta Traffic", "Google Ads"]
        )
        ad_total = amz_ad + onz_ad

        # MKT Cost 2: Discounts (Shopify promo codes only — ref fee moved to variable)
        disc_total = shopify_disc

        # MKT Cost 3: Influencer / Collab
        inf_paid = inf_paid_m.get(m, 0)
        inf_nonpaid = inf_nonpaid_m.get(m, 0)
        seeding_total = inf_paid + inf_nonpaid

        mkt_total = ad_total + disc_total + seeding_total
        cm = cm_before_mkt - mkt_total  # CM after MKT

        total_rev_monthly.append(round(rev))
        total_cogs_monthly.append(round(cogs))
        total_gm_monthly.append(round(gm))
        _variable_costs_arr.append(round(variable_costs))
        _cm_before_mkt_arr.append(round(cm_before_mkt))
        total_ad_spend_monthly.append(round(ad_total))
        total_disc_monthly.append(round(disc_total))
        total_seeding_monthly.append(round(seeding_total))
        total_mkt_monthly.append(round(mkt_total))
        total_cm_monthly.append(round(cm))
        # Breakdown arrays
        _shop_disc_arr.append(round(shopify_disc))
        _inf_paid_arr.append(round(inf_paid))
        _inf_nonpaid_arr.append(round(inf_nonpaid))

    # Paid vs Organic (monthly)
    paid_monthly = []
    for m in months:
        paid = sum(
            ad_monthly.get(p, {}).get(m, {}).get("sales", 0)
            for p in ["Amazon Ads", "Meta CVR", "Meta Traffic", "Google Ads"]
        )
        paid_monthly.append(round(paid))

    organic_monthly = [max(0, rev - paid) for rev, paid in zip(total_rev_monthly, paid_monthly)]

    # ── Build P&L from DataKeeper ────────────────────────────────────────────
    print("[7.5/8] Building P&L...")
    MONTH_NAMES_SHORT = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    all_pnl_months = sorted(months)  # DataKeeper months only

    # Build month labels
    pnl_month_labels = []
    for m in all_pnl_months:
        mi = int(m[5:7])
        yr = m[2:4]
        pnl_month_labels.append(f"{MONTH_NAMES_SHORT[mi-1]} {yr}")

    fy2025_months = [m for m in all_pnl_months if m.startswith("2025-")]
    fy2025_idx = len(fy2025_months)

    def _pnl_with_annual(monthly_vals, fy_count):
        out = [round(v) for v in monthly_vals[:fy_count]]
        out.append(round(sum(monthly_vals[:fy_count])))
        out.extend([round(v) for v in monthly_vals[fy_count:]])
        return out

    # ── Revenue by brand (Gross = Shopify net + Amazon gross) ──
    brand_rev_pnl = {}
    for brand in BRAND_ORDER:
        vals = []
        for m in all_pnl_months:
            shopify_v = brand_monthly.get(brand, {}).get(m, {}).get("gross", 0)
            amz_v = amz_brand_monthly.get(brand, {}).get(m, {}).get("gross", 0)
            vals.append(round(shopify_v + amz_v))
        if any(v > 0 for v in vals):
            brand_rev_pnl[brand] = {
                "values": _pnl_with_annual(vals, fy2025_idx),
                "color": BRAND_COLORS.get(brand, "#94a3b8"),
            }

    # ── Ad spend from DataKeeper (with Amazon Ads backfill) ──
    onz_spend_arr, amz_spend_arr, google_spend_arr, total_ad_arr = [], [], [], []
    for m in all_pnl_months:
        onz = sum(ad_monthly.get(p, {}).get(m, {}).get("spend", 0)
                  for p in ["Meta CVR", "Meta Traffic", "Google Ads"])
        amz = ad_monthly.get("Amazon Ads", {}).get(m, {}).get("spend", 0)
        if amz == 0 and m in amz_ads_backfill:
            amz = amz_ads_backfill[m]
        ggl = ad_monthly.get("Google Ads", {}).get(m, {}).get("spend", 0)
        onz_spend_arr.append(round(onz))
        amz_spend_arr.append(round(amz))
        google_spend_arr.append(round(ggl))
        total_ad_arr.append(round(onz + amz))

    # ── Sales from ads (DataKeeper) ──
    onz_paid_arr, amz_paid_arr, total_paid_arr = [], [], []
    for m in all_pnl_months:
        onz_p = sum(ad_monthly.get(p, {}).get(m, {}).get("sales", 0)
                    for p in ["Meta CVR", "Meta Traffic", "Google Ads"])
        amz_p = ad_monthly.get("Amazon Ads", {}).get(m, {}).get("sales", 0)
        onz_paid_arr.append(round(onz_p))
        amz_paid_arr.append(round(amz_p))
        total_paid_arr.append(round(onz_p + amz_p))

    organic_arr = [max(0, r - p) for r, p in zip(total_rev_monthly, total_paid_arr)]
    cm_after_arr = [g - a for g, a in zip(total_gm_monthly, total_ad_arr)]

    final_labels = list(pnl_month_labels)
    final_labels.insert(fy2025_idx, "FY2025")

    # Mark last month as partial if applicable
    if partial_month_info["is_partial"] and final_labels:
        last_label = final_labels[-1]
        thru_day = int(through.split("-")[2])
        thru_m = int(through.split("-")[1])
        final_labels[-1] = f"{last_label}\n(thru {thru_m}/{thru_day})"

    pnl_polar = {
        "months": final_labels,
        "fy2025_idx": fy2025_idx,
        "partial_month": partial_month_info,
        "brand_sales": brand_rev_pnl,
        "total_revenue": _pnl_with_annual(total_rev_monthly, fy2025_idx),
        "cogs": _pnl_with_annual(total_cogs_monthly, fy2025_idx),
        "gross_margin": _pnl_with_annual(total_gm_monthly, fy2025_idx),
        "variable_costs": _pnl_with_annual(_variable_costs_arr, fy2025_idx),
        "variable_detail": {
            "amz_ref_fee": _pnl_with_annual(_amz_fee_arr, fy2025_idx),
            "fba_fulfillment": _pnl_with_annual([v if v is not None else 0 for v in _fba_arr], fy2025_idx),
            "fba_nm_months": [i for i, v in enumerate(_fba_arr) if v is None],
        },
        "cm_before_mkt": _pnl_with_annual(_cm_before_mkt_arr, fy2025_idx),
        "ad_spend": {
            "onzenna": _pnl_with_annual(onz_spend_arr, fy2025_idx),
            "amazon": _pnl_with_annual(amz_spend_arr, fy2025_idx),
            "total": _pnl_with_annual(total_ad_arr, fy2025_idx),
        },
        "ad_spend_detail": {
            "google": _pnl_with_annual(google_spend_arr, fy2025_idx),
            "amz_grosmimi": _pnl_with_annual([0]*len(all_pnl_months), fy2025_idx),
            "amz_chaenmom": _pnl_with_annual([0]*len(all_pnl_months), fy2025_idx),
            "amz_naeiae": _pnl_with_annual([0]*len(all_pnl_months), fy2025_idx),
        },
        "sales_from_ads": {
            "onzenna": _pnl_with_annual(onz_paid_arr, fy2025_idx),
            "amazon": _pnl_with_annual(amz_paid_arr, fy2025_idx),
            "total": _pnl_with_annual(total_paid_arr, fy2025_idx),
        },
        "organic": {
            "onzenna": _pnl_with_annual([0]*len(all_pnl_months), fy2025_idx),
            "amazon": _pnl_with_annual([0]*len(all_pnl_months), fy2025_idx),
            "total": _pnl_with_annual(organic_arr, fy2025_idx),
        },
        "discounts": _pnl_with_annual(total_disc_monthly, fy2025_idx),
        "discounts_detail": {
            "shopify_disc": _pnl_with_annual(_shop_disc_arr, fy2025_idx),
            "amz_ref_fee": _pnl_with_annual(_amz_fee_arr, fy2025_idx),
        },
        "influencer_spend": _pnl_with_annual(total_seeding_monthly, fy2025_idx),
        "influencer_detail": {
            "paid": _pnl_with_annual(_inf_paid_arr, fy2025_idx),
            "nonpaid": _pnl_with_annual(_inf_nonpaid_arr, fy2025_idx),
        },
        "cm_after_ads": _pnl_with_annual(cm_after_arr, fy2025_idx),
        "cm_final": _pnl_with_annual(
            [g - a - s for g, a, s in zip(total_gm_monthly, total_ad_arr, total_seeding_monthly)],
            fy2025_idx
        ),
    }
    print(f"  P&L built: {len(all_pnl_months)} months, {len(brand_rev_pnl)} brands")

    # ── Channel P&L ──────────────────────────────────────────────────────────
    # Amazon: revenue from amazon_sales_daily, COGS from SKU-level NAS map
    # Shopify D2C (excl Target+): Shopify orders, COGS from SKU-level NAS map
    # Target+: Shopify target+ channel, COGS, fulfillment = N/A
    print("  Building Channel P&L...")

    # Build monthly Amazon fees from amazon_sales
    amz_fees_monthly = defaultdict(float)
    amz_fees_weekly = defaultdict(float)
    for r in amazon_sales:
        d = r.get("date", "")
        if not d or d > through:
            continue
        month = d[:7]
        amz_fees_monthly[month] += float(r.get("fees") or 0)
        wk = _week_key(d)
        amz_fees_weekly[wk] += float(r.get("fees") or 0)

    # Build monthly FBA fulfillment costs
    # Strategy: load FBA fee snapshot (ASIN -> per-unit fee), apply to SKU daily units
    amz_fulfill_monthly = defaultdict(float)
    amz_fulfill_weekly = defaultdict(float)
    has_fba_fees = False

    # Load FBA fee snapshot from local cache (written by data_keeper.py)
    fba_snapshot_path = Path(".tmp/datakeeper/amazon_fba_fees.json")
    asin_fba_fee = {}  # asin -> per-unit FBA fee
    if fba_snapshot_path.exists():
        try:
            import json as _json
            fba_list = _json.loads(fba_snapshot_path.read_text(encoding="utf-8"))
            for item in fba_list:
                asin = item.get("asin", "")
                fee = float(item.get("fba_fee", 0) or 0)
                if asin and fee > 0:
                    asin_fba_fee[asin] = fee
        except Exception as e:
            print(f"  [WARN] FBA snapshot load failed: {e}")

    if asin_fba_fee:
        # Apply FBA fee to each SKU row by ASIN matching
        for r in amazon_sku:
            d = r.get("date", "")
            if not d or d > through:
                continue
            asin = r.get("asin", "")
            fee = asin_fba_fee.get(asin, 0)
            if fee > 0:
                units = int(r.get("units", 0) or 0)
                fba_total = fee * units
                has_fba_fees = True
                amz_fulfill_monthly[d[:7]] += fba_total
                amz_fulfill_weekly[_week_key(d)] += fba_total

    if has_fba_fees:
        print(f"  FBA fulfillment: ${sum(amz_fulfill_monthly.values()):,.0f} total across {len(amz_fulfill_monthly)} months ({len(asin_fba_fee)} ASINs matched)")
    else:
        print("  FBA fulfillment: no data (run: python tools/data_keeper.py --channel amazon_sales)")

    # Channel P&L structure: revenue/cogs/selling_fees/fulfillment/ad_spend/gm/cm per month
    # Selling fees: Amazon 15%, Target+ 15%, Shopify 0%, B2B 0%
    # Fulfillment: Amazon FBA from estimated fees report, others n.m.
    CHANNEL_PNL_ORDER = ["Amazon MP", "Onzenna D2C", "Target+", "B2B"]

    # Build monthly Target+ fees from amazon_sales (channel=Target+)
    tp_fees_monthly = defaultdict(float)
    for r in amazon_sales:
        d = r.get("date", "")
        ch = r.get("channel", "")
        if not d or d > through or ch != "Target+":
            continue
        tp_fees_monthly[d[:7]] += float(r.get("fees") or 0)

    def _channel_pnl_monthly(ch, months_list):
        rev_arr, cogs_arr, sell_arr, fulfill_arr, ad_arr = [], [], [], [], []
        for m in months_list:
            rev = 0
            cogs = 0
            sell_fee = 0
            fulfill = 0
            ad_sp = 0

            if ch == "Amazon MP":
                for b in BRAND_ORDER + ["Other"]:
                    rev += amz_brand_monthly.get(b, {}).get(m, {}).get("gross", 0)
                # SKU-level COGS for Amazon
                for b in BRAND_ORDER + ["Other"]:
                    cogs += amz_sku_cogs.get(b, {}).get(m, 0)
                # Partial month: scale up if SKU units < daily units
                amz_daily_u = sum(amz_brand_monthly.get(b, {}).get(m, {}).get("units", 0) for b in BRAND_ORDER + ["Other"])
                amz_sku_u = amz_sku_units_m.get(m, 0)
                if cogs > 0 and amz_sku_u > 0 and amz_daily_u > amz_sku_u * 1.2:
                    cogs = round(cogs * amz_daily_u / amz_sku_u)
                # Fallback if SKU COGS is 0 but revenue exists
                elif cogs == 0 and rev > 0:
                    for b in BRAND_ORDER + ["Other"]:
                        u = amz_brand_monthly.get(b, {}).get(m, {}).get("units", 0)
                        cogs += u * brand_vw_cogs.get(b, brand_vw_cogs.get(b, 0))
                # Amazon selling fees (referral) + discounts shown in waterfall
                sell_fee = amz_fees_monthly.get(m, 0)
                fulfill = amz_fulfill_monthly.get(m, 0)
                ad_sp = (ad_monthly.get("Amazon Ads", {}).get(m, {}).get("spend", 0)
                         + ad_monthly.get("Meta Traffic", {}).get(m, {}).get("spend", 0))
            elif ch == "Onzenna D2C":
                rev = channel_monthly.get("Onzenna D2C", {}).get(m, {}).get("net", 0)
                # SKU-level COGS — D2C channel only (fixes double-counting bug)
                for b in BRAND_ORDER + ["Other"]:
                    cogs += shop_cogs_by_channel.get("Onzenna D2C", {}).get(b, {}).get(m, 0)
                # Fallback if SKU COGS is 0 but revenue exists
                if cogs == 0 and rev > 0:
                    # Use ratio from total shop_sku_cogs as estimate
                    total_shop = sum(shop_sku_cogs.get(b, {}).get(m, 0) for b in BRAND_ORDER + ["Other"])
                    total_shop_rev = sum(brand_monthly.get(b, {}).get(m, {}).get("net", 0) for b in BRAND_ORDER + ["Other"])
                    cogs_rate = total_shop / total_shop_rev if total_shop_rev else 0.30
                    cogs = round(rev * cogs_rate)
                sell_fee = 0
                fulfill = 0  # n.m.
                ad_sp = (ad_monthly.get("Meta CVR", {}).get(m, {}).get("spend", 0)
                         + ad_monthly.get("Google Ads", {}).get(m, {}).get("spend", 0))
            elif ch == "Target+":
                rev = channel_monthly.get("Target+", {}).get(m, {}).get("net", 0)
                # Target+ COGS from SKU data if available
                tp_sku_cogs = sum(shop_cogs_by_channel.get("Target+", {}).get(b, {}).get(m, 0) for b in BRAND_ORDER + ["Other"])
                if tp_sku_cogs > 0:
                    cogs = tp_sku_cogs
                else:
                    orders = channel_monthly.get("Target+", {}).get(m, {}).get("orders", 0)
                    cogs = orders * 10
                sell_fee = tp_fees_monthly.get(m, 0) or round(rev * 0.15)  # from data or 15%
                fulfill = 0  # n.m.
                ad_sp = 0
            elif ch == "B2B":
                rev = channel_monthly.get("B2B", {}).get(m, {}).get("net", 0)
                # B2B COGS from SKU data if available
                b2b_sku_cogs = sum(shop_cogs_by_channel.get("B2B", {}).get(b, {}).get(m, 0) for b in BRAND_ORDER + ["Other"])
                if b2b_sku_cogs > 0:
                    cogs = b2b_sku_cogs
                else:
                    orders = channel_monthly.get("B2B", {}).get(m, {}).get("orders", 0)
                    cogs = orders * 12
                sell_fee = 0
                fulfill = 0
                ad_sp = 0

            rev_arr.append(round(rev))
            cogs_arr.append(round(cogs))
            sell_arr.append(round(sell_fee))
            # Fulfillment: None = n.m. for months without FBA data
            if ch == "Amazon MP":
                if fulfill > 0:
                    fulfill_arr.append(round(fulfill))
                elif m >= "2025-10" and has_fba_fees:
                    fulfill_arr.append(0)  # genuinely $0 (unlikely but possible)
                else:
                    fulfill_arr.append(None)  # n.m. — no FBA data
            elif ch in ("Onzenna D2C", "Target+"):
                fulfill_arr.append(None)
            else:
                fulfill_arr.append(round(fulfill))
            ad_arr.append(round(ad_sp))

        gm_arr = [r - c - s - (f or 0) for r, c, s, f in zip(rev_arr, cogs_arr, sell_arr, fulfill_arr)]
        cm_arr = [g - a for g, a in zip(gm_arr, ad_arr)]
        fulfill_nm = any(f is None for f in fulfill_arr)
        return {
            "revenue": rev_arr,
            "cogs": cogs_arr,
            "selling_fees": sell_arr,
            "fulfillment": fulfill_arr,
            "fulfillment_nm": fulfill_nm,
            "ad_spend": ad_arr,
            "gross_margin": gm_arr,
            "contribution_margin": cm_arr,
            "color": CHANNEL_PNL_COLORS.get(ch, "#94a3b8"),
        }

    channel_pnl = {}
    for ch in CHANNEL_PNL_ORDER:
        data = _channel_pnl_monthly(ch, months)
        if any(v > 0 for v in data["revenue"]):
            channel_pnl[ch] = data

    pnl_polar["channel_pnl"] = channel_pnl
    print(f"  Channel P&L: {list(channel_pnl.keys())}")

    # ── Klaviyo email marketing ──────────────────────────────────────────────
    # Flow data = cumulative snapshots (same stats repeated daily per flow).
    # Fix: per flow per month, keep only the latest-date row (best snapshot).
    # Campaign data = one row per campaign (no dedup needed).
    print("\n[7/7] Building Klaviyo email metrics...")

    klav_campaigns = [r for r in klaviyo if r.get("source_type") == "campaign"
                      and r.get("date", "") <= through]
    klav_flows_raw = [r for r in klaviyo if r.get("source_type") == "flow"
                      and r.get("date", "") <= through]

    # Dedup flows: latest date per (source_name, month)
    flow_latest = {}  # (name, month) -> row with max date
    for r in klav_flows_raw:
        key = (r.get("source_name", ""), r["date"][:7])
        if key not in flow_latest or r["date"] > flow_latest[key]["date"]:
            flow_latest[key] = r
    klav_flows = list(flow_latest.values())

    def _agg_rows(rows):
        """Aggregate a list of Klaviyo rows into totals."""
        s = sum(int(r.get("sends") or 0) for r in rows)
        o = sum(int(r.get("opens") or 0) for r in rows)
        c = sum(int(r.get("clicks") or 0) for r in rows)
        cv = sum(int(r.get("conversions") or 0) for r in rows)
        rv = sum(float(r.get("revenue") or 0) for r in rows)
        return {"sends": s, "opens": o, "clicks": c, "conversions": cv, "revenue": rv}

    # Monthly arrays aligned to months[]
    def _klav_monthly_arr(rows, months_list):
        out = []
        for m in months_list:
            m_rows = [r for r in rows if r.get("date", "")[:7] == m]
            out.append(_agg_rows(m_rows))
        return out

    camp_arr = _klav_monthly_arr(klav_campaigns, months)
    flow_arr = _klav_monthly_arr(klav_flows, months)

    klaviyo_data = {
        "campaign_monthly": {
            "sends": [v["sends"] for v in camp_arr],
            "opens": [v["opens"] for v in camp_arr],
            "clicks": [v["clicks"] for v in camp_arr],
            "conversions": [v["conversions"] for v in camp_arr],
            "revenue": [round(v["revenue"]) for v in camp_arr],
            "revenue_proj": proj_array([round(v["revenue"]) for v in camp_arr]),
        },
        "flow_monthly": {
            "sends": [v["sends"] for v in flow_arr],
            "opens": [v["opens"] for v in flow_arr],
            "clicks": [v["clicks"] for v in flow_arr],
            "conversions": [v["conversions"] for v in flow_arr],
            "revenue": [round(v["revenue"]) for v in flow_arr],
            "revenue_proj": proj_array([round(v["revenue"]) for v in flow_arr]),
        },
    }

    # Period summaries (7D, 30D) — dedup flows per period too
    def _klav_period_summary(date_from_str):
        p_camps = [r for r in klav_campaigns if r.get("date", "") >= date_from_str]
        p_flows_raw = [r for r in klav_flows_raw if r.get("date", "") >= date_from_str]
        # Dedup flows: latest date per source_name in this period
        fl = {}
        for r in p_flows_raw:
            name = r.get("source_name", "")
            if name not in fl or r["date"] > fl[name]["date"]:
                fl[name] = r
        p_flows = list(fl.values())
        all_rows = p_camps + p_flows
        agg = _agg_rows(all_rows)
        s = agg["sends"]
        camp_r = sum(float(r.get("revenue") or 0) for r in p_camps)
        flow_r = sum(float(r.get("revenue") or 0) for r in p_flows)
        return {
            **agg,
            "revenue": round(agg["revenue"], 2),
            "open_rate": round(agg["opens"] / s * 100, 1) if s else 0,
            "click_rate": round(agg["clicks"] / s * 100, 2) if s else 0,
            "cvr": round(agg["conversions"] / s * 100, 2) if s else 0,
            "rev_per_send": round(agg["revenue"] / s, 3) if s else 0,
            "campaign_revenue": round(camp_r, 2),
            "flow_revenue": round(flow_r, 2),
        }

    def _klav_top_items(date_from_str, source_type, limit=15):
        if source_type == "campaign":
            p_rows = [r for r in klav_campaigns if r.get("date", "") >= date_from_str]
        else:
            p_raw = [r for r in klav_flows_raw if r.get("date", "") >= date_from_str]
            fl = {}
            for r in p_raw:
                name = r.get("source_name", "")
                if name not in fl or r["date"] > fl[name]["date"]:
                    fl[name] = r
            p_rows = list(fl.values())
        # Group by name
        grouped = defaultdict(lambda: {"sends": 0, "opens": 0, "clicks": 0, "conversions": 0, "revenue": 0.0, "date": ""})
        for r in p_rows:
            name = r.get("source_name", "Unknown")
            grouped[name]["sends"] += int(r.get("sends") or 0)
            grouped[name]["opens"] += int(r.get("opens") or 0)
            grouped[name]["clicks"] += int(r.get("clicks") or 0)
            grouped[name]["conversions"] += int(r.get("conversions") or 0)
            grouped[name]["revenue"] += float(r.get("revenue") or 0)
            grouped[name]["date"] = max(grouped[name]["date"], r.get("date", ""))
        top = sorted(grouped.items(), key=lambda x: x[1]["revenue"], reverse=True)[:limit]
        out = []
        for name, v in top:
            s = v["sends"]
            entry = {
                "name": name, "sends": s,
                "open_rate": round(v["opens"] / s * 100, 1) if s else 0,
                "click_rate": round(v["clicks"] / s * 100, 2) if s else 0,
                "cvr": round(v["conversions"] / s * 100, 2) if s else 0,
                "revenue": round(v["revenue"], 2),
            }
            if source_type == "campaign":
                entry["date"] = v["date"]
            out.append(entry)
        return out

    klaviyo_data["summary_7d"] = _klav_period_summary(d7)
    klaviyo_data["summary_30d"] = _klav_period_summary(d30)
    klaviyo_data["top_campaigns_7d"] = _klav_top_items(d7, "campaign")
    klaviyo_data["top_campaigns_30d"] = _klav_top_items(d30, "campaign")
    klaviyo_data["top_flows_7d"] = _klav_top_items(d7, "flow")
    klaviyo_data["top_flows_30d"] = _klav_top_items(d30, "flow")

    s30 = klaviyo_data["summary_30d"]
    print(f"  Klaviyo: {len(klaviyo)} raw rows -> {len(klav_campaigns)} campaigns + {len(klav_flows)} flows (deduped)")
    print(f"  30D: sends={s30['sends']:,} rev=${s30['revenue']:,.0f} | 7D: sends={klaviyo_data['summary_7d']['sends']:,} rev=${klaviyo_data['summary_7d']['revenue']:,.0f}")

    # ── Weekly data for all tabs ─────────────────────────────────────────────
    print("\n[7.8/8] Building weekly aggregations...")

    # Build week list (last 12 weeks)
    all_week_keys = set()
    for entity in list(brand_weekly.values()) + list(amz_brand_weekly.values()) + list(channel_weekly.values()):
        all_week_keys |= set(entity.keys())
    for platform in ad_weekly.values():
        all_week_keys |= set(platform.keys())

    # Last 12 weeks
    _wk_list_12 = []
    for w_offset in range(11, -1, -1):
        dt_w = today - timedelta(weeks=w_offset)
        iy, iw, _ = dt_w.isocalendar()
        wk = f"{iy}-W{iw:02d}"
        if wk not in _wk_list_12:
            _wk_list_12.append(wk)
    weeks = _wk_list_12

    from datetime import date as _date_cls
    def _wk_label(wk_str):
        parts = wk_str.split("-W")
        try:
            d = _date_cls.fromisocalendar(int(parts[0]), int(parts[1]), 1)
            return f"W{int(parts[1])} ({d.strftime('%b %d')})"
        except Exception:
            return wk_str

    week_labels_all = [_wk_label(w) for w in weeks]

    # Weekly brand revenue
    wk_brand_rev = {}
    for brand in BRAND_ORDER:
        vals = [round(brand_weekly.get(brand, {}).get(w, {}).get("net", 0)
                      + amz_brand_weekly.get(brand, {}).get(w, {}).get("net", 0)) for w in weeks]
        if any(v > 0 for v in vals):
            wk_brand_rev[brand] = {
                "weekly": vals,
                "color": BRAND_COLORS.get(brand, "#94a3b8"),
            }

    # Weekly channel revenue
    wk_channel_rev = {}
    for ch in ["Onzenna D2C", "Amazon MP", "Amazon FBA MCF", "TikTok Shop", "Target+", "B2B"]:
        vals = [round(channel_weekly.get(ch, {}).get(w, {}).get("net", 0)) for w in weeks]
        if any(v > 0 for v in vals):
            wk_channel_rev[ch] = {
                "weekly": vals,
                "color": CHANNEL_COLORS.get(ch, "#94a3b8"),
            }

    # Weekly ad performance
    wk_ad_perf = {}
    for platform in ["Amazon Ads", "Meta CVR", "Meta Traffic", "Google Ads"]:
        spend_vals = [round(ad_weekly.get(platform, {}).get(w, {}).get("spend", 0)) for w in weeks]
        sales_vals = [round(ad_weekly.get(platform, {}).get(w, {}).get("sales", 0)) for w in weeks]
        impr_vals = [ad_weekly.get(platform, {}).get(w, {}).get("impressions", 0) for w in weeks]
        click_vals = [ad_weekly.get(platform, {}).get(w, {}).get("clicks", 0) for w in weeks]
        if any(v > 0 for v in spend_vals):
            wk_ad_perf[platform] = {
                "spend": spend_vals,
                "sales": sales_vals,
                "impressions": impr_vals,
                "clicks": click_vals,
                "color": AD_COLORS.get(platform, "#94a3b8"),
            }

    # Weekly ads landing TROAS
    wk_onz_spend = [0] * len(weeks)
    wk_amz_spend = [0] * len(weeks)
    for i, w in enumerate(weeks):
        for p in ["Google Ads", "Meta CVR"]:
            wk_onz_spend[i] += ad_weekly.get(p, {}).get(w, {}).get("spend", 0)
        for p in ["Amazon Ads", "Meta Traffic"]:
            wk_amz_spend[i] += ad_weekly.get(p, {}).get(w, {}).get("spend", 0)
    wk_onz_rev = [round(channel_weekly.get("Onzenna D2C", {}).get(w, {}).get("net", 0)) for w in weeks]
    wk_amz_rev = [round(channel_weekly.get("Amazon MP", {}).get(w, {}).get("net", 0)) for w in weeks]
    wk_ads_landing = {
        "Onzenna": {
            "spend": [round(v) for v in wk_onz_spend],
            "revenue": wk_onz_rev,
            "platforms": "Google Ads + Meta CVR",
            "color": "#6366f1",
        },
        "Amazon": {
            "spend": [round(v) for v in wk_amz_spend],
            "revenue": wk_amz_rev,
            "platforms": "Amazon Ads + Meta Traffic",
            "color": "#f59e0b",
        },
    }

    # Weekly brand performance (ad + total sales)
    wk_brand_perf = {}
    for brand in BRAND_ORDER:
        total_vals = [round(brand_weekly.get(brand, {}).get(w, {}).get("net", 0)
                           + amz_brand_weekly.get(brand, {}).get(w, {}).get("net", 0)) for w in weeks]
        ad_spend_vals = [round(brand_ad_weekly.get(brand, {}).get(w, {}).get("spend", 0)) for w in weeks]
        ad_sales_vals = [round(brand_ad_weekly.get(brand, {}).get(w, {}).get("sales", 0)) for w in weeks]
        organic_vals = [max(0, t - a) for t, a in zip(total_vals, ad_sales_vals)]
        if any(v > 0 for v in total_vals) or any(v > 0 for v in ad_spend_vals):
            wk_brand_perf[brand] = {
                "total_sales": total_vals,
                "ad_spend": ad_spend_vals,
                "ad_sales": ad_sales_vals,
                "organic": organic_vals,
                "color": BRAND_COLORS.get(brand, "#94a3b8"),
            }

    # Weekly waterfall (GM, CM)
    wk_rev_arr = []
    wk_cogs_arr = []
    wk_gm_arr = []
    wk_ad_spend_arr = []
    wk_disc_arr = []
    wk_mkt_arr = []
    wk_cm_arr = []
    for w in weeks:
        shopify_net = sum(brand_weekly.get(b, {}).get(w, {}).get("net", 0) for b in BRAND_ORDER + ["Other"])
        amz_net = sum(amz_brand_weekly.get(b, {}).get(w, {}).get("net", 0) for b in BRAND_ORDER + ["Other"])
        rev = shopify_net + amz_net
        cogs = 0
        for b in BRAND_ORDER:
            units = brand_weekly.get(b, {}).get(w, {}).get("units", 0)
            if units == 0 and brand_weekly.get(b, {}).get(w, {}).get("gross", 0) > 0:
                units = int(brand_weekly[b][w]["gross"] / AVG_PRICE.get(b, 25))
            cogs += units * brand_vw_cogs.get(b, 0)
            amz_units = amz_brand_weekly.get(b, {}).get(w, {}).get("units", 0)
            if amz_units == 0 and amz_brand_weekly.get(b, {}).get(w, {}).get("net", 0) > 0:
                amz_units = int(amz_brand_weekly[b][w]["net"] / AVG_PRICE.get(b, 25))
            cogs += amz_units * brand_vw_cogs.get(b, 0)
        gm = rev - cogs
        ad_sp = sum(ad_weekly.get(p, {}).get(w, {}).get("spend", 0) for p in ["Amazon Ads", "Meta CVR", "Meta Traffic", "Google Ads"])
        disc = sum(brand_weekly.get(b, {}).get(w, {}).get("disc", 0) for b in BRAND_ORDER + ["Other"])
        mkt = ad_sp + abs(disc)
        cm = gm - mkt
        wk_rev_arr.append(round(rev))
        wk_cogs_arr.append(round(cogs))
        wk_gm_arr.append(round(gm))
        wk_ad_spend_arr.append(round(ad_sp))
        wk_disc_arr.append(round(abs(disc)))
        wk_mkt_arr.append(round(mkt))
        wk_cm_arr.append(round(cm))

    weekly_data = {
        "weeks": weeks,
        "week_labels": week_labels_all,
        "brand_revenue": wk_brand_rev,
        "channel_revenue": wk_channel_rev,
        "ad_performance": wk_ad_perf,
        "ads_landing": wk_ads_landing,
        "brand_performance": wk_brand_perf,
        "waterfall": {
            "revenue": wk_rev_arr,
            "cogs": wk_cogs_arr,
            "gross_margin": wk_gm_arr,
            "ad_spend": wk_ad_spend_arr,
            "discounts": wk_disc_arr,
            "mkt_total": wk_mkt_arr,
            "contribution_margin": wk_cm_arr,
        },
        "channel_pnl": {},
    }

    # Build weekly channel P&L
    def _ch_pnl_wk(ch, wlist):
        rev_a, cogs_a, fees_a, ad_a = [], [], [], []
        for w in wlist:
            rev = cogs = fees = ad_sp = 0
            if ch == "Amazon MP":
                for b in BRAND_ORDER + ["Other"]:
                    net = amz_brand_weekly.get(b, {}).get(w, {}).get("net", 0)
                    units = amz_brand_weekly.get(b, {}).get(w, {}).get("units", 0)
                    if units == 0 and net > 0:
                        units = int(net / AVG_PRICE.get(b, 25))
                    rev += net
                    cogs += units * brand_vw_cogs.get(b, 0)
                fees = amz_fees_weekly.get(w, 0)
                ad_sp = (ad_weekly.get("Amazon Ads", {}).get(w, {}).get("spend", 0)
                         + ad_weekly.get("Meta Traffic", {}).get(w, {}).get("spend", 0))
            elif ch == "Onzenna D2C":
                rev = channel_weekly.get("Onzenna D2C", {}).get(w, {}).get("net", 0)
                for b in BRAND_ORDER + ["Other"]:
                    bw = brand_weekly.get(b, {}).get(w, {})
                    b_units = bw.get("units", 0)
                    if b_units == 0 and bw.get("net", 0) > 0:
                        b_units = int(bw["net"] / AVG_PRICE.get(b, 25))
                    cogs += b_units * brand_vw_cogs.get(b, 0)
                ad_sp = (ad_weekly.get("Meta CVR", {}).get(w, {}).get("spend", 0)
                         + ad_weekly.get("Google Ads", {}).get(w, {}).get("spend", 0))
            elif ch == "Target+":
                rev = channel_weekly.get("Target+", {}).get(w, {}).get("net", 0)
                cogs = channel_weekly.get("Target+", {}).get(w, {}).get("orders", 0) * 10
            elif ch == "B2B":
                rev = channel_weekly.get("B2B", {}).get(w, {}).get("net", 0)
                cogs = channel_weekly.get("B2B", {}).get(w, {}).get("orders", 0) * 12
            rev_a.append(round(rev)); cogs_a.append(round(cogs)); fees_a.append(round(fees)); ad_a.append(round(ad_sp))
        gm_a = [r-c-f for r,c,f in zip(rev_a, cogs_a, fees_a)]
        cm_a = [g-a for g,a in zip(gm_a, ad_a)]
        return {"revenue":rev_a,"cogs":cogs_a,"fees":fees_a,"ad_spend":ad_a,"gross_margin":gm_a,"contribution_margin":cm_a,"color":CHANNEL_PNL_COLORS.get(ch,"#94a3b8")}

    for ch in ["Amazon MP", "Onzenna D2C", "Target+", "B2B"]:
        d = _ch_pnl_wk(ch, weeks)
        if any(v > 0 for v in d["revenue"]):
            weekly_data["channel_pnl"][ch] = d

    print(f"  Weekly: {len(weeks)} weeks ({weeks[0]} -> {weeks[-1]})")

    # ── Campaign-level detail (weekly + monthly) ────────────────────────────
    print("\n[8/8] Building campaign-level detail...")

    from datetime import date as _date_type

    # Helper: ISO week label "W12 (Mar 17)"
    def _week_label(iso_yr, iso_wk):
        try:
            d = _date_type.fromisocalendar(iso_yr, iso_wk, 1)
            return f"W{iso_wk} ({d.strftime('%b %d')})"
        except Exception:
            return f"W{iso_wk}"

    # Collect campaign-level daily rows (all platforms)
    campaign_daily = []  # list of dicts with normalized fields

    for r in amazon_ads:
        d = r.get("date", "")
        if not d or d > through:
            continue
        campaign_daily.append({
            "date": d,
            "platform": "Amazon Ads",
            "campaign_id": r.get("campaign_id", ""),
            "campaign_name": r.get("campaign_name", ""),
            "brand": r.get("brand") or "Other",
            "spend": float(r.get("spend") or 0),
            "sales": float(r.get("sales") or 0),
            "impressions": int(r.get("impressions") or 0),
            "clicks": int(r.get("clicks") or 0),
        })

    for r in meta_ads:
        d = r.get("date", "")
        if not d or d > through:
            continue
        cname = (r.get("campaign_name") or "").lower()
        landing = (r.get("landing_url") or "").lower()
        is_amz = "amazon" in cname or "amz" in cname or "amazon" in landing
        platform = "Meta Traffic" if is_amz else "Meta CVR"
        campaign_daily.append({
            "date": d,
            "platform": platform,
            "campaign_id": r.get("campaign_id", ""),
            "campaign_name": r.get("campaign_name", ""),
            "brand": r.get("brand") or "Other",
            "spend": float(r.get("spend") or 0),
            "sales": float(r.get("purchase_value") or 0),
            "impressions": int(r.get("impressions") or 0),
            "clicks": int(r.get("clicks") or 0),
        })

    for r in google_ads:
        d = r.get("date", "")
        if not d or d > through:
            continue
        campaign_daily.append({
            "date": d,
            "platform": "Google Ads",
            "campaign_id": r.get("campaign_id", ""),
            "campaign_name": r.get("campaign_name", ""),
            "brand": "Grosmimi",
            "spend": float(r.get("spend") or 0),
            "sales": float(r.get("conversion_value") or 0),
            "impressions": int(r.get("impressions") or 0),
            "clicks": int(r.get("clicks") or 0),
        })

    # Aggregate by campaign x week and campaign x month
    camp_weekly = defaultdict(lambda: defaultdict(lambda: {
        "spend": 0, "sales": 0, "impressions": 0, "clicks": 0
    }))
    camp_monthly_agg = defaultdict(lambda: defaultdict(lambda: {
        "spend": 0, "sales": 0, "impressions": 0, "clicks": 0
    }))
    camp_meta = {}  # campaign_id -> {name, platform, brand}

    for row in campaign_daily:
        cid = row["campaign_id"]
        d = row["date"]
        month = d[:7]
        dt = datetime.strptime(d, "%Y-%m-%d").date()
        iso_yr, iso_wk, _ = dt.isocalendar()
        wk_key = f"{iso_yr}-W{iso_wk:02d}"

        camp_weekly[cid][wk_key]["spend"] += row["spend"]
        camp_weekly[cid][wk_key]["sales"] += row["sales"]
        camp_weekly[cid][wk_key]["impressions"] += row["impressions"]
        camp_weekly[cid][wk_key]["clicks"] += row["clicks"]

        camp_monthly_agg[cid][month]["spend"] += row["spend"]
        camp_monthly_agg[cid][month]["sales"] += row["sales"]
        camp_monthly_agg[cid][month]["impressions"] += row["impressions"]
        camp_monthly_agg[cid][month]["clicks"] += row["clicks"]

        if cid not in camp_meta:
            camp_meta[cid] = {
                "name": row["campaign_name"],
                "platform": row["platform"],
                "brand": row["brand"],
                "campaign_id": cid,
            }

    # Build week labels (last 12 weeks)
    today_iso = today.isocalendar()
    week_keys = []
    for w_offset in range(11, -1, -1):
        dt_w = today - timedelta(weeks=w_offset)
        iy, iw, _ = dt_w.isocalendar()
        wk = f"{iy}-W{iw:02d}"
        if wk not in week_keys:
            week_keys.append(wk)

    week_labels = []
    for wk in week_keys:
        parts = wk.split("-W")
        week_labels.append(_week_label(int(parts[0]), int(parts[1])))

    # Attribution campaign name -> data mapping (precise matching)
    def _normalize(s):
        """Normalize campaign name for matching: lowercase, strip separators."""
        return s.lower().replace("|", " ").replace("_", " ").replace("-", " ").replace("\t", " ").strip()

    # Pre-build attribution lookup: normalized name -> data
    _attr_norm = {}
    for attr_name, attr_data in attribution.items():
        _attr_norm[_normalize(attr_name)] = (attr_name, attr_data)

    def _match_attribution(meta_name):
        """Match Meta Traffic campaign name to Attribution campaign.
        Uses token overlap with stop-word filtering. Score >= 3 required."""
        if not attribution:
            return None
        mn = _normalize(meta_name)

        # Exact match first
        if mn in _attr_norm:
            return _attr_norm[mn][1]

        STOP = {"amz", "traffic", "wl", "meta", "target", "the", ""}
        meta_tokens = set(mn.split()) - STOP

        best = None
        best_score = 0
        for an, (_, attr_data) in _attr_norm.items():
            attr_tokens = set(an.split()) - STOP
            overlap = meta_tokens & attr_tokens
            score = len(overlap)
            if any(t for t in overlap if t[:4].isdigit() and len(t) >= 6):
                score += 3
            if any(t for t in overlap if t in ("dental", "dentalmom", "stainless", "strawcup",
                                                 "spring", "sale", "bfcm", "naeiae", "grosmimi",
                                                 "chamom", "alpremio", "livfuselli", "knotted")):
                score += 2
            if score > best_score:
                best_score = score
                best = attr_data

        return best if best_score >= 3 else None

    # Build campaign detail entries
    def _camp_entry(cid, agg_dict, period_keys):
        m = camp_meta.get(cid, {})
        spend_arr = [round(agg_dict[cid].get(pk, {}).get("spend", 0), 2) for pk in period_keys]
        sales_arr = [round(agg_dict[cid].get(pk, {}).get("sales", 0), 2) for pk in period_keys]
        impr_arr = [agg_dict[cid].get(pk, {}).get("impressions", 0) for pk in period_keys]
        click_arr = [agg_dict[cid].get(pk, {}).get("clicks", 0) for pk in period_keys]
        total_spend = sum(spend_arr)
        total_sales = sum(sales_arr)
        total_clicks = sum(click_arr)
        total_impr = sum(impr_arr)
        cpc = round(total_spend / total_clicks, 2) if total_clicks > 0 else 0
        ctr = round(total_clicks / total_impr * 100, 2) if total_impr > 0 else 0
        is_traffic = "traffic" in m.get("name", "").lower()

        # Attribution data for traffic campaigns
        attr_sales = 0
        attr_purchases = 0
        attr_brb = 0
        attr_roas = 0
        attr_roas_adj = 0
        if is_traffic and attribution:
            attr_match = _match_attribution(m.get("name", ""))
            if attr_match:
                attr_sales = round(attr_match["sales"], 2)
                attr_purchases = attr_match["purchases"]
                attr_brb = round(attr_match["brb"], 2)
                attr_roas = round(attr_sales / total_spend, 2) if total_spend > 0 else 0
                attr_roas_adj = round((attr_sales + attr_brb) / total_spend, 2) if total_spend > 0 else 0

        entry = {
            "id": cid,
            "campaign_id": m.get("campaign_id", cid),
            "name": m.get("name", ""),
            "platform": m.get("platform", ""),
            "brand": m.get("brand", "Other"),
            "spend": spend_arr,
            "sales": sales_arr,
            "impressions": impr_arr,
            "clicks": click_arr,
            "total_spend": round(total_spend, 2),
            "total_sales": round(total_sales, 2),
            "total_clicks": total_clicks,
            "total_impressions": total_impr,
            "roas": round(total_sales / total_spend, 2) if total_spend > 0 else 0,
            "cpc": cpc,
            "ctr": ctr,
            "is_traffic": is_traffic,
        }
        if is_traffic:
            entry["attr_sales"] = attr_sales
            entry["attr_purchases"] = attr_purchases
            entry["attr_brb"] = attr_brb
            entry["attr_roas"] = attr_roas
            entry["attr_roas_adj"] = attr_roas_adj
        return entry

    # Weekly entries (last 12 weeks, only campaigns with spend)
    weekly_entries = []
    for cid in camp_weekly:
        e = _camp_entry(cid, camp_weekly, week_keys)
        if e["total_spend"] > 0:
            weekly_entries.append(e)
    weekly_entries.sort(key=lambda x: x["total_spend"], reverse=True)

    # Monthly entries (ad_months only)
    monthly_entries = []
    for cid in camp_monthly_agg:
        e = _camp_entry(cid, camp_monthly_agg, ad_months)
        if e["total_spend"] > 0:
            monthly_entries.append(e)
    monthly_entries.sort(key=lambda x: x["total_spend"], reverse=True)

    # Top / Bottom performers
    # CVR campaigns: sort by ROAS. Traffic campaigns (ROAS=0): sort by CPC (lower=better).
    def _top_bottom(entries, min_spend=50):
        qualified = [e for e in entries if e["total_spend"] >= min_spend]
        cvr = [e for e in qualified if not e.get("is_traffic")]
        traffic = [e for e in qualified if e.get("is_traffic") and e["cpc"] > 0]
        # CVR: top = high ROAS, bottom = low ROAS
        top_cvr = sorted(cvr, key=lambda x: x["roas"], reverse=True)[:8]
        bottom_cvr = sorted(cvr, key=lambda x: x["roas"])[:8]
        # Traffic: top = high Attribution ROAS, bottom = low Attribution ROAS
        # Fallback to CPC if no attribution data
        top_traffic = sorted(traffic, key=lambda x: x.get("attr_roas_adj", 0), reverse=True)[:5]
        bottom_traffic = sorted(traffic, key=lambda x: x.get("attr_roas_adj", 999) if x.get("attr_roas_adj", 0) > 0 else 999)[:5]
        return top_cvr, bottom_cvr, top_traffic, bottom_traffic

    weekly_top_cvr, weekly_bottom_cvr, weekly_top_traffic, weekly_bottom_traffic = _top_bottom(weekly_entries, min_spend=20)
    monthly_top_cvr, monthly_bottom_cvr, monthly_top_traffic, monthly_bottom_traffic = _top_bottom(monthly_entries, min_spend=50)

    # New campaigns: appeared this month but not in previous months
    current_month = ad_months[-1] if ad_months else ""
    prev_months = set(ad_months[:-1]) if len(ad_months) > 1 else set()
    new_campaigns = []
    for cid in camp_monthly_agg:
        has_current = camp_monthly_agg[cid].get(current_month, {}).get("spend", 0) > 0
        has_prev = any(camp_monthly_agg[cid].get(pm, {}).get("spend", 0) > 0 for pm in prev_months)
        if has_current and not has_prev:
            e = _camp_entry(cid, camp_monthly_agg, [current_month])
            if e["total_spend"] > 0:
                new_campaigns.append(e)
    new_campaigns.sort(key=lambda x: x["total_spend"], reverse=True)

    campaign_detail = {
        "week_keys": week_keys,
        "week_labels": week_labels,
        "month_keys": ad_months,
        "weekly_all": weekly_entries,
        "weekly_top": weekly_top_cvr,
        "weekly_bottom": weekly_bottom_cvr,
        "weekly_top_traffic": weekly_top_traffic,
        "weekly_bottom_traffic": weekly_bottom_traffic,
        "monthly_all": monthly_entries,
        "monthly_top": monthly_top_cvr,
        "monthly_bottom": monthly_bottom_cvr,
        "monthly_top_traffic": monthly_top_traffic,
        "monthly_bottom_traffic": monthly_bottom_traffic,
        "new_campaigns": new_campaigns,
        "current_month": current_month,
    }

    print(f"  Campaigns: {len(camp_meta)} total, weekly={len(weekly_entries)}, monthly={len(monthly_entries)}")
    print(f"  New this month ({current_month}): {len(new_campaigns)}")
    print(f"  Top/Bottom weekly: {len(weekly_top_cvr)}+{len(weekly_top_traffic)}/{len(weekly_bottom_cvr)}+{len(weekly_bottom_traffic)}, monthly: {len(monthly_top_cvr)}+{len(monthly_top_traffic)}/{len(monthly_bottom_cvr)}+{len(monthly_bottom_traffic)}")

    # ── Serialize brand_ad_by_platform ───────────────────────────────────────
    def _bap_serialize(raw_dict, period_keys):
        out = {}
        for platform in ["Amazon Ads", "Meta CVR", "Meta Traffic", "Google Ads"]:
            brands = dict(raw_dict.get(platform, {}))
            plat_out = {}
            for brand in sorted(brands.keys()):
                spend = [round(brands[brand].get(pk, {}).get("spend", 0), 2) for pk in period_keys]
                if not any(s > 0 for s in spend):
                    continue
                sales = [round(brands[brand].get(pk, {}).get("sales", 0), 2) for pk in period_keys]
                impr = [brands[brand].get(pk, {}).get("impressions", 0) for pk in period_keys]
                clicks = [brands[brand].get(pk, {}).get("clicks", 0) for pk in period_keys]
                plat_out[brand] = {
                    "spend": spend,
                    "sales": sales,
                    "impressions": impr,
                    "clicks": clicks,
                    "color": BRAND_COLORS.get(brand, "#94a3b8"),
                }
            out[platform] = plat_out
        return out

    bap_monthly_out = _bap_serialize(brand_ad_by_platform, ad_months)
    bap_weekly_out = _bap_serialize(brand_ad_by_platform_wk, weeks)
    weekly_data["brand_ad_by_platform"] = bap_weekly_out

    # ── Serialize brand_ad_daily (for Hero Products ad overlay) ───────────────
    all_ad_dates = set()
    for brand_data in brand_ad_daily_raw.values():
        for metric_data in brand_data.values():
            all_ad_dates |= set(metric_data.keys())
    ad_dates_sorted = sorted(all_ad_dates)

    DAILY_METRICS = ["amz_spend", "amz_clicks", "amz_sales",
                     "mt_spend", "mt_clicks", "mt_sales",
                     "gads_spend", "gads_clicks", "gads_sales"]
    brand_ad_daily_out = {"dates": ad_dates_sorted}
    for brand in BRAND_ORDER:
        if brand not in brand_ad_daily_raw:
            continue
        bd = brand_ad_daily_raw[brand]
        brand_out = {}
        for metric in DAILY_METRICS:
            vals = [round(bd[metric].get(d, 0), 2) for d in ad_dates_sorted]
            if any(v > 0 for v in vals):
                brand_out[metric] = vals
        if brand_out:
            brand_ad_daily_out[brand] = brand_out

    print(f"  Brand ad daily: {len(ad_dates_sorted)} dates, {len(brand_ad_daily_out)-1} brands")

    # ── Hero Products ─────────────────────────────────────────────────────────
    print("\n[HERO] Building Hero Products tab...")

    def _build_hero_products():
        """Build hero products: product-category level sales + keywords + content lift."""
        from datetime import date as _date
        cutoff_7d = (today - timedelta(days=7)).isoformat()
        cutoff_30d = (today - timedelta(days=30)).isoformat()

        EXCLUDE_NAMES = {
            "grosmimi eco friendly 304 stainless steel toddler kid feeding divided plate",
            "grosmimi spout",
        }

        def _classify_asin(name, brand):
            t = (name or "").lower()
            if any(e in t for e in EXCLUDE_NAMES):
                return "Other"
            if "alpremio" in t or brand == "Alpremio":
                return "Alpremio"
            if brand == "Grosmimi" or "grosmimi" in t:
                if "stainless" in t and "tumbler" in t:     return "Stainless Tumbler"
                if "stainless" in t:                        return "Stainless Straw Cup"
                if "tumbler" in t or "slow flow" in t:      return "PPSU Tumbler"
                if "baby bottle" in t or "feeding bottle" in t or ("bottle" in t and "straw" not in t and "nipple" not in t):
                    return "PPSU Baby Bottle"
                if "replacement" in t or "accessory" in t or "nipple" in t or "weighted" in t or "strap" in t:
                    return "Accessories"
                return "PPSU Straw Cup"
            if "cha&mom" in t or "cha & mom" in t or "phyto" in t or brand == "CHA&MOM":
                if "wash" in t:             return "Body Wash"
                if "cream" in t:            return "Baby Cream"
                return "Moisturizer"
            if "naeiae" in t or brand == "Naeiae":
                if "pop rice" in t or "rice snack" in t or "rice puff" in t:
                    return "Rice Puff"
                return "Other"
            return "Other"

        EXCLUDE_CATS = {"Accessories", "Other"}
        GROSMIMI_CATEGORIES = ["PPSU Straw Cup", "Stainless Straw Cup", "PPSU Tumbler", "Stainless Tumbler", "PPSU Baby Bottle"]
        CHAMOM_CATEGORIES = ["Moisturizer", "Body Wash", "Baby Cream"]
        CAT_BRAND = {cat: "Grosmimi" for cat in GROSMIMI_CATEGORIES}
        CAT_BRAND.update({cat: "CHA&MOM" for cat in CHAMOM_CATEGORIES})
        CAT_BRAND.update({"Rice Puff": "Naeiae", "Alpremio": "Alpremio"})
        cutoff_90d = (today - timedelta(days=90)).isoformat()

        cat_data = defaultdict(lambda: {
            "brand": "", "sales_7d": 0.0, "units_7d": 0, "sales_30d": 0.0, "units_30d": 0,
            "asins": set(), "asin_names": {},
            "daily": defaultdict(lambda: {"sales": 0.0, "units": 0}),
            "weekly": defaultdict(lambda: {"sales": 0.0, "units": 0}),
        })
        asin_to_cat = {}
        for r in amazon_sku:
            asin = r.get("asin", "")
            if not asin:
                continue
            d = r.get("date", "")
            name = r.get("product_name", "")
            brand = r.get("brand", "")
            if asin not in asin_to_cat:
                asin_to_cat[asin] = _classify_asin(name, brand)
            cat = asin_to_cat[asin]
            sales = float(r.get("ordered_product_sales") or r.get("net_sales") or 0)
            units = int(r.get("units") or 0)
            cd = cat_data[cat]
            if not cd["brand"]:
                cd["brand"] = brand
            cd["asins"].add(asin)
            if asin not in cd["asin_names"]:
                cd["asin_names"][asin] = (name or "")[:80]
            if d >= cutoff_7d:
                cd["sales_7d"] += sales; cd["units_7d"] += units
            if d >= cutoff_30d:
                cd["sales_30d"] += sales; cd["units_30d"] += units
            if d >= cutoff_90d:
                cd["daily"][d]["sales"] += sales; cd["daily"][d]["units"] += units
            try:
                dt = _date.fromisoformat(d)
                wk_key = f"{dt.isocalendar()[0]}-W{dt.isocalendar()[1]:02d}"
                cd["weekly"][wk_key]["sales"] += sales; cd["weekly"][wk_key]["units"] += units
            except Exception:
                pass

        sorted_cats = sorted(
            [(cat, cd) for cat, cd in cat_data.items() if cat not in EXCLUDE_CATS and cd["sales_30d"] > 0],
            key=lambda x: -x[1]["sales_30d"]
        )
        print(f"  Categories: {len(sorted_cats)} ({', '.join(c for c,_ in sorted_cats)})")

        # BA keyword mapping
        cat_keywords = defaultdict(list)
        for r in brand_analytics:
            asin = r.get("asin", "")
            if not r.get("is_ours") or asin not in asin_to_cat:
                continue
            cat = asin_to_cat[asin]
            term = (r.get("search_term") or "").strip().lower()
            if not term:
                continue
            cat_keywords[cat].append({
                "keyword": term,
                "search_freq_rank": int(r.get("search_frequency_rank") or 0),
                "click_share": round(float(r.get("click_share") or 0) * 100, 2),
                "conv_share": round(float(r.get("conversion_share") or 0) * 100, 2),
                "week": r.get("date", ""),
            })

        cat_top_kw = {}
        for cat, entries in cat_keywords.items():
            by_kw = defaultdict(list)
            for e in entries:
                by_kw[e["keyword"]].append(e)
            best = []
            for kw, weeks_data in by_kw.items():
                latest = max(weeks_data, key=lambda x: x["week"])
                rank_history = sorted(weeks_data, key=lambda x: x["week"])
                best.append({
                    "keyword": kw, "search_freq_rank": latest["search_freq_rank"],
                    "rank_weekly": [w["search_freq_rank"] for w in rank_history[-12:]],
                    "rank_week_labels": [w["week"] for w in rank_history[-12:]],
                    "click_share": latest["click_share"], "conv_share": latest["conv_share"],
                })
            cat_top_kw[cat] = sorted(best, key=lambda x: -x["click_share"])[:8]

        # SFR branded keywords
        _BRAND_VARIANTS_PY = {
            'grosmimi': ['grosmimi', 'grosmini', 'grossini', 'gros mimi', 'grossmimi'],
            'naeiae': ['naeiae', 'nae iae'],
            'chaenmom': ['cha&mom', 'chaenmom', 'cha and mom', 'commemoi'],
            'alpremio': ['alpremio'],
        }
        import re as _re2
        def _bslug(b): return _re2.sub(r'[^a-z0-9]', '', (b or '').lower())
        cat_brand_map = {cat: CAT_BRAND.get(cat, cd.get("brand", "")) for cat, cd in sorted_cats}

        cat_sfr_branded = {}
        for cat, entries in cat_keywords.items():
            slug = _bslug(cat_brand_map.get(cat, ''))
            bvars = _BRAND_VARIANTS_PY.get(slug, [slug] if slug else [])
            by_kw = defaultdict(list)
            for e in entries:
                by_kw[e["keyword"]].append(e)
            branded = []
            for kw, weeks_data in by_kw.items():
                kl = kw.lower()
                if bvars and not any(v in kl for v in bvars):
                    continue
                rank_history = sorted(weeks_data, key=lambda x: x["week"])
                if not rank_history:
                    continue
                latest = max(weeks_data, key=lambda x: x["week"])
                branded.append({
                    "keyword": kw, "search_freq_rank": latest["search_freq_rank"],
                    "rank_weekly": [w["search_freq_rank"] for w in rank_history[-12:]],
                    "rank_week_labels": [w["week"] for w in rank_history[-12:]],
                    "click_share": latest["click_share"], "conv_share": latest["conv_share"],
                })
            cat_sfr_branded[cat] = sorted(branded, key=lambda x: -x["click_share"])

        # SQP branded keywords
        sqp_by_brand = defaultdict(lambda: defaultdict(dict))
        for r in sqp_brand:
            b = _bslug(r.get("brand", ""))
            q = (r.get("search_query") or "").strip().lower()
            w = str(r.get("week_end", ""))
            v = int(r.get("search_query_volume") or 0)
            if b and q and w and v:
                sqp_by_brand[b][q][w] = v

        brand_top10 = {}
        for slug, kw_weeks in sqp_by_brand.items():
            all_wks = sorted(set(w for wv in kw_weeks.values() for w in wv))
            kw_totals = {q: sum(wv.values()) for q, wv in kw_weeks.items()}
            top10_kws = sorted(kw_totals, key=lambda q: -kw_totals[q])[:10]
            kw_list = []
            for q in top10_kws:
                wv = kw_weeks[q]
                kw_list.append({
                    "keyword": q, "search_freq_rank": 0, "rank_weekly": [], "rank_week_labels": [],
                    "volume_weekly": [wv.get(w, 0) for w in all_wks], "volume_week_labels": all_wks,
                    "click_share": 0, "conv_share": 0,
                })
            brand_top10[slug] = kw_list

        for cat in set(cat_sfr_branded.keys()) | set(cat_brand_map.keys()):
            slug = _bslug(cat_brand_map.get(cat, ''))
            if slug in brand_top10:
                cat_sfr_branded[cat] = brand_top10[slug]

        # Enrich keywords with ads spend
        st_spend = defaultdict(lambda: {"spend": 0.0, "clicks": 0, "sales": 0.0, "impressions": 0})
        for r in search_terms:
            term = (r.get("search_term") or "").strip().lower()
            if term:
                st_spend[term]["spend"] += float(r.get("spend") or 0)
                st_spend[term]["clicks"] += int(r.get("clicks") or 0)
                st_spend[term]["sales"] += float(r.get("sales") or 0)

        gads_lookup = defaultdict(lambda: {"impressions": 0, "clicks": 0, "spend": 0.0})
        for r in google_search_terms:
            term = (r.get("search_term") or "").strip().lower()
            if term:
                gads_lookup[term]["impressions"] += int(r.get("impressions") or 0)
                gads_lookup[term]["clicks"] += int(r.get("clicks") or 0)
                gads_lookup[term]["spend"] += float(r.get("cost") or r.get("spend") or 0)

        gsc_kw_lookup = defaultdict(lambda: {"impressions": 0, "clicks": 0})
        for r in gsc:
            q = (r.get("query") or "").strip().lower()
            if q:
                gsc_kw_lookup[q]["impressions"] += int(r.get("impressions") or 0)
                gsc_kw_lookup[q]["clicks"] += int(r.get("clicks") or 0)

        for cat, kws in cat_top_kw.items():
            for kw in kws:
                st = st_spend.get(kw["keyword"], {})
                kw["ads_spend_30d"] = round(st.get("spend", 0))
                kw["ads_clicks_30d"] = st.get("clicks", 0)
                kw["ads_sales_30d"] = round(st.get("sales", 0))
                kw["ads_acos"] = round(st.get("spend", 0) / st.get("sales", 1), 2) if st.get("sales", 0) > 0 else 0
                gsc_d = gsc_kw_lookup.get(kw["keyword"].lower(), {})
                gads_d = gads_lookup.get(kw["keyword"].lower(), {})
                kw["google_volume"] = (gsc_d.get("impressions", 0) or 0) + (gads_d.get("impressions", 0) or 0)
                kw["google_ads_impressions"] = gads_d.get("impressions", 0)
                kw["google_ads_clicks"] = gads_d.get("clicks", 0)

        # Build category output
        all_weeks = set()
        for _, cd in sorted_cats:
            all_weeks.update(cd["weekly"].keys())
        week_keys = sorted(all_weeks)[-12:]
        daily_dates = sorted(set(d for _, cd in sorted_cats for d in cd["daily"].keys()))[-90:]

        categories = []
        for cat, cd in sorted_cats:
            categories.append({
                "category": cat, "brand": CAT_BRAND.get(cat, cd["brand"]),
                "asin_count": len(cd["asins"]),
                "sales_7d": round(cd["sales_7d"]), "units_7d": cd["units_7d"],
                "sales_30d": round(cd["sales_30d"]), "units_30d": cd["units_30d"],
                "sales_weekly": [round(cd["weekly"].get(wk, {}).get("sales", 0)) for wk in week_keys],
                "units_weekly": [cd["weekly"].get(wk, {}).get("units", 0) for wk in week_keys],
                "daily_sales": [round(cd["daily"].get(d, {}).get("sales", 0)) for d in daily_dates],
                "daily_units": [cd["daily"].get(d, {}).get("units", 0) for d in daily_dates],
                "top_keywords": cat_top_kw.get(cat, []),
                "sfr_branded": cat_sfr_branded.get(cat, []),
                "asins": sorted([{"asin": a, "name": cd["asin_names"].get(a, "")} for a in cd["asins"]], key=lambda x: x["name"]),
            })

        # Content lift
        post_info = {}
        for p in content_posts:
            pid = p.get("post_id") or p.get("url", "")
            brand = p.get("brand", "")
            pt_raw = p.get("product_types", "")
            ptypes = [t.strip() for t in pt_raw.split(",") if t.strip()] if pt_raw else []
            uname = (p.get("username") or "").strip()
            platform = p.get("platform", "")
            region = (p.get("region") or "us").lower()
            if pid and region == "us":
                post_info[pid] = {"brand": brand, "product_types": ptypes, "username": uname, "platform": platform, "post_date": p.get("post_date", "")}

        ptype_views = defaultdict(lambda: defaultdict(int))
        for m in content_metrics:
            pid = m.get("post_id", "")
            d = m.get("date", "")
            views = int(m.get("views") or 0)
            info = post_info.get(pid)
            if not info or d < cutoff_90d:
                continue
            ptypes = info["product_types"]
            brand = info["brand"]
            if ptypes:
                for pt in ptypes:
                    ptype_views[pt][d] += views
            else:
                if brand == "Grosmimi":
                    for cat in GROSMIMI_CATEGORIES:
                        ptype_views[cat][d] += views
                elif brand == "CHA&MOM":
                    for cat in CHAMOM_CATEGORIES:
                        ptype_views[cat][d] += views
                elif brand == "Naeiae":
                    ptype_views["Rice Puff"][d] += views

        BRAND_KW_MAP = {"Grosmimi": ["grosmimi"], "CHA&MOM": ["cha and mom", "chamom", "cha&mom"],
                        "Naeiae": ["naeiae", "pop rice"], "Onzenna": ["onzenna"], "Alpremio": ["alpremio"]}
        brand_gsc_daily = defaultdict(lambda: defaultdict(int))
        for r in gsc:
            q = (r.get("query") or "").strip().lower()
            d = r.get("date", "")
            impr = int(r.get("impressions") or 0)
            if q and d and impr > 0:
                for bn, kws in BRAND_KW_MAP.items():
                    if any(k in q for k in kws):
                        brand_gsc_daily[bn][d] += impr; break

        asin_days = amz_sessions_raw.get("asin_days", {}) if amz_sessions_raw else {}
        cat_sessions_daily = defaultdict(lambda: defaultdict(int))
        cat_pageviews_daily = defaultdict(lambda: defaultdict(int))
        for cat, cd in sorted_cats:
            for asin in cd.get("asins", set()):
                for d, sv in asin_days.get(asin, {}).items():
                    cat_sessions_daily[cat][d] += sv.get("sessions", 0)
                    cat_pageviews_daily[cat][d] += sv.get("pageViews", 0)

        content_lift = {}
        for cat, cd in sorted_cats:
            cat_brand = cd["brand"]
            content_lift[cat] = {
                "dates": daily_dates,
                "views": [ptype_views.get(cat, {}).get(d, 0) for d in daily_dates],
                "sales": [round(cd["daily"].get(d, {}).get("sales", 0)) for d in daily_dates],
                "units": [cd["daily"].get(d, {}).get("units", 0) for d in daily_dates],
                "gsc_daily": [brand_gsc_daily.get(cat_brand, {}).get(d, 0) for d in daily_dates],
                "sessions_daily": [cat_sessions_daily[cat].get(d, 0) for d in daily_dates],
                "pageviews_daily": [cat_pageviews_daily[cat].get(d, 0) for d in daily_dates],
            }

        # Content creators per category
        username_brand = {}
        for pi in post_info.values():
            u, b = pi.get("username", ""), pi.get("brand", "")
            if u and b and u not in username_brand:
                username_brand[u] = b
        for pi in post_info.values():
            if not pi.get("brand") and pi.get("username"):
                pi["brand"] = username_brand.get(pi["username"], "")

        post_snapshots = defaultdict(dict)
        post_max_views = defaultdict(int)
        post_first_views = defaultdict(int)
        for m in content_metrics:
            pid, d = m.get("post_id", ""), m.get("date", "")
            views = int(m.get("views") or 0)
            if pid and d:
                if views > post_snapshots[pid].get(d, 0):
                    post_snapshots[pid][d] = views
                if views > post_max_views[pid]:
                    post_max_views[pid] = views

        post_daily_delta = defaultdict(dict)
        for pid, snaps in post_snapshots.items():
            sd = sorted(snaps.keys())
            if len(sd) < 2:
                continue
            prev_v = snaps[sd[0]]
            post_first_views[pid] = prev_v
            for d in sd[1:]:
                delta = max(0, snaps[d] - prev_v)
                if delta > 0:
                    post_daily_delta[pid][d] = delta
                prev_v = snaps[d]

        cat_creator_daily = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
        cat_creator_meta = {}
        for pid, deltas in post_daily_delta.items():
            info = post_info.get(pid)
            if not info or not info.get("username"):
                continue
            uname, brand = info["username"], info["brand"]
            ptypes = info["product_types"]
            cats = ptypes if ptypes else (
                GROSMIMI_CATEGORIES if brand == "Grosmimi"
                else CHAMOM_CATEGORIES if brand == "CHA&MOM"
                else ["Rice Puff"] if brand == "Naeiae" else [])
            for cat in cats:
                for d, dv in deltas.items():
                    if d >= cutoff_90d:
                        cat_creator_daily[cat][uname][d] += dv
                key = (cat, uname)
                if key not in cat_creator_meta:
                    cat_creator_meta[key] = {"brand": brand, "platform": info.get("platform", ""), "upload_date": info.get("post_date", ""), "total_views": 0}
                pd_val = info.get("post_date", "")
                if pd_val and (not cat_creator_meta[key]["upload_date"] or pd_val < cat_creator_meta[key]["upload_date"]):
                    cat_creator_meta[key]["upload_date"] = pd_val

        cat_user_posts = defaultdict(lambda: defaultdict(set))
        for pid, info in post_info.items():
            if not info.get("username"):
                continue
            uname, brand = info["username"], info["brand"]
            cats = info["product_types"] if info["product_types"] else (
                GROSMIMI_CATEGORIES if brand == "Grosmimi"
                else CHAMOM_CATEGORIES if brand == "CHA&MOM"
                else ["Rice Puff"] if brand == "Naeiae" else [])
            for cat in cats:
                cat_user_posts[cat][uname].add(pid)
        for (cat, uname), meta in cat_creator_meta.items():
            pids = cat_user_posts.get(cat, {}).get(uname, set())
            meta["total_views"] = sum(post_max_views.get(pid, 0) for pid in pids)
            meta["base_views"] = sum(post_first_views.get(pid, 0) for pid in pids)
            obs = [min(post_snapshots[pid].keys()) for pid in pids if pid in post_snapshots and post_snapshots[pid]]
            meta["first_observed"] = min(obs) if obs else ""

        EXCLUDED_CREATORS = {"rosalieflorentin"}
        content_creators_by_cat = {}
        for cat in [c[0] for c in sorted_cats]:
            totals = []
            for uname, dv in cat_creator_daily.get(cat, {}).items():
                if uname.lower() in EXCLUDED_CREATORS:
                    continue
                growth = sum(dv.values())
                if growth > 0:
                    meta = cat_creator_meta.get((cat, uname), {})
                    ud = meta.get("upload_date", "")
                    if ud and ud < cutoff_90d:
                        continue
                    base = meta.get("base_views", 0)
                    totals.append({
                        "username": uname, "brand": meta.get("brand", ""), "platform": meta.get("platform", ""),
                        "total_views": base + growth, "base_views": base,
                        "first_observed": meta.get("first_observed", ""), "upload_date": ud,
                        "daily_views": [dv.get(d, 0) for d in daily_dates],
                    })
            totals.sort(key=lambda x: -x["total_views"])
            content_creators_by_cat[cat] = totals[:10]

        # Daily spend overlay
        spend_dates = sorted(set(r.get("date", "") for r in amazon_ads if r.get("date", "") >= cutoff_90d))[-90:]
        _sd_set = set(spend_dates)
        _asd, _acd, _msd, _mcd, _msld, _gsd, _atsd, _g4sd = [defaultdict(float) for _ in range(4)] + [defaultdict(int) for _ in range(1)] + [defaultdict(float) for _ in range(1)] + [defaultdict(float) for _ in range(1)] + [defaultdict(int) for _ in range(1)]
        for r in amazon_ads:
            d = r.get("date", "")
            if d in _sd_set:
                _asd[d] += float(r.get("spend") or 0); _acd[d] += int(r.get("clicks") or 0)
        for r in meta_ads:
            d = r.get("date", "")
            if d in _sd_set:
                _msd[d] += float(r.get("spend") or 0); _mcd[d] += int(r.get("clicks") or 0); _msld[d] += float(r.get("purchase_value") or 0)
        for r in google_ads:
            d = r.get("date", "")
            if d in _sd_set:
                _gsd[d] += float(r.get("spend") or 0)
        for r in amazon_sales:
            d = r.get("date", "")
            if d in _sd_set:
                _atsd[d] += float(r.get("ordered_product_sales") or r.get("gross_sales") or 0)
        for r in ga4:
            d = r.get("date", "")
            if d in _sd_set:
                _g4sd[d] += int(r.get("sessions") or 0)

        spend_daily = {
            "dates": spend_dates,
            "amazon": [round(_asd.get(d, 0)) for d in spend_dates],
            "meta": [round(_msd.get(d, 0)) for d in spend_dates],
            "google": [round(_gsd.get(d, 0)) for d in spend_dates],
            "amz_clicks": [_acd.get(d, 0) for d in spend_dates],
            "amz_total_sales": [round(_atsd.get(d, 0)) for d in spend_dates],
            "meta_clicks": [_mcd.get(d, 0) for d in spend_dates],
            "meta_sales": [round(_msld.get(d, 0)) for d in spend_dates],
            "ga4_sessions": [_g4sd.get(d, 0) for d in spend_dates],
        }

        print(f"  Hero categories: {len(categories)}, keywords: {len([k for c in categories for k in c.get('top_keywords',[])])}")
        return {
            "week_keys": week_keys, "daily_dates": daily_dates,
            "categories": categories, "keyword_table": [],
            "content_lift": content_lift, "content_creators_by_cat": content_creators_by_cat,
            "spend_daily": spend_daily,
        }

    hero_data = _build_hero_products()

    # ── Add per-category attribution campaign mapping to hero_products ──
    _attr_30d = channel_traffic["30d"]["amazon"].get("attribution", {})
    _attr_camps = _attr_30d.get("campaigns", [])
    if hero_data and _attr_camps:
        # Ordered list: check most specific keywords first, avoid cross-matching
        # Each campaign is assigned to AT MOST ONE category (first match wins)
        CAT_KEYWORDS_ORDERED = [
            ("PPSU Straw Cup", ["ppsu straw", "ppsu strawcup"]),
            ("Stainless Straw Cup", ["stainless straw", "stainless strawcup", "steel straw"]),
            ("PPSU Tumbler", ["ppsu tumbler"]),
            ("Stainless Tumbler", ["stainless tumbler", "steel tumbler"]),
            ("PPSU Baby Bottle", ["ppsu bottle", "baby bottle", "ppsu baby"]),
            ("Rice Puff", ["rice puff", "rice snack", "naeiae rice"]),
            ("Moisturizer", ["moisturizer", "lotion"]),
            ("Body Wash", ["body wash"]),
            ("Alpremio", ["alpremio"]),
            ("Baby Cream", ["baby cream"]),
        ]

        # Phase 1: assign each campaign to specific category or "unmatched"
        camp_assignments = {}  # campaign_name -> category or None
        for ac in _attr_camps:
            cn_lower = ac["name"].lower().replace("_", " ").replace("|", " ")
            assigned = None
            for cat_name, kws in CAT_KEYWORDS_ORDERED:
                if any(kw in cn_lower for kw in kws):
                    assigned = cat_name
                    break
            camp_assignments[ac["name"]] = assigned

        # Phase 2: build per-category lists (specific matches only)
        cat_attr = {cat_data["category"]: [] for cat_data in hero_data.get("categories", [])}
        unmatched = []  # brand-level generic campaigns
        for ac in _attr_camps:
            assigned = camp_assignments.get(ac["name"])
            if assigned and assigned in cat_attr:
                cat_attr[assigned].append(ac)
            else:
                unmatched.append(ac)

        n_matched = sum(len(v) for v in cat_attr.values())
        n_unmatched = len(unmatched)
        print(f"  Attribution mapping: {n_matched} specific + {n_unmatched} generic (brand-level)")

        # Phase 3: cat_attribution — only specifically matched campaigns
        hero_data["cat_attribution"] = {}
        for cat, camps in cat_attr.items():
            total_sales = sum(c.get("sales", 0) for c in camps)
            total_spend = sum(c.get("spend", 0) for c in camps)
            total_brb = sum(c.get("brb", 0) for c in camps)
            total_orders = sum(c.get("purchases", 0) for c in camps)
            hero_data["cat_attribution"][cat] = {
                "sales": round(total_sales),
                "spend": round(total_spend),
                "brb": round(total_brb, 2),
                "orders": total_orders,
                "roas": round(total_sales / total_spend, 2) if total_spend > 0 else 0,
                "roas_brb": round((total_sales + total_brb) / total_spend, 2) if total_spend > 0 else 0,
                "campaigns": [{
                    "name": c["name"].strip()[:60],
                    "spend": c.get("spend", 0),
                    "sales": c.get("sales", 0),
                    "brb": round(c.get("brb", 0), 2),
                    "clicks": c.get("clicks", 0),
                    "purchases": c.get("purchases", 0),
                    "roas": c.get("roas", 0),
                    "roas_adj": c.get("roas_adj", 0),
                } for c in camps if c.get("spend", 0) > 0 or c.get("sales", 0) > 0],
            }

        # Phase 4: brand_attribution — ALL campaigns (for Brand view toggle)
        brand_attr_map = {}  # brand -> list of all campaigns
        brand_cats = {}  # brand -> set of categories
        for cat_data in hero_data.get("categories", []):
            brand = cat_data["brand"]
            brand_cats.setdefault(brand, set()).add(cat_data["category"])
        # All attribution campaigns are Grosmimi (Meta→AMZ traffic)
        # Group by brand based on campaign name heuristics
        for ac in _attr_camps:
            cn_lower = ac["name"].lower()
            # Default to Grosmimi (most Meta→AMZ campaigns are for Grosmimi)
            brand = "Grosmimi"
            if "naeiae" in cn_lower:
                brand = "Naeiae"
            elif "chamom" in cn_lower or "cha&mom" in cn_lower:
                brand = "CHA&MOM"
            elif "alpremio" in cn_lower:
                brand = "Alpremio"
            brand_attr_map.setdefault(brand, []).append(ac)

        hero_data["brand_attribution"] = {}
        for brand, camps in brand_attr_map.items():
            total_sales = sum(c.get("sales", 0) for c in camps)
            total_spend = sum(c.get("spend", 0) for c in camps)
            total_brb = sum(c.get("brb", 0) for c in camps)
            total_orders = sum(c.get("purchases", 0) for c in camps)
            hero_data["brand_attribution"][brand] = {
                "sales": round(total_sales),
                "spend": round(total_spend),
                "brb": round(total_brb, 2),
                "orders": total_orders,
                "roas": round(total_sales / total_spend, 2) if total_spend > 0 else 0,
                "roas_brb": round((total_sales + total_brb) / total_spend, 2) if total_spend > 0 else 0,
                "campaigns": [{
                    "name": c["name"].strip()[:60],
                    "spend": c.get("spend", 0),
                    "sales": c.get("sales", 0),
                    "brb": round(c.get("brb", 0), 2),
                    "clicks": c.get("clicks", 0),
                    "purchases": c.get("purchases", 0),
                    "roas": c.get("roas", 0),
                    "roas_adj": c.get("roas_adj", 0),
                } for c in camps if c.get("spend", 0) > 0 or c.get("sales", 0) > 0],
            }

        print(f"  Hero category attribution: {sum(1 for v in hero_data['cat_attribution'].values() if v['sales'] > 0)} cats with sales")
        _ba_parts = [f"{b} ${v['sales']:,}" for b, v in hero_data['brand_attribution'].items()]
        print(f"  Hero brand attribution: {', '.join(_ba_parts)}")

    # ── Assemble final data ───────────────────────────────────────────────────
    fin_data = {
        "generated_pst": now_pst.strftime("%Y-%m-%d %H:%M PST"),
        "through_date": through,
        "months": months,
        "data_sources": data_sources,
        "partial_month": partial_month_info,
        "summary": summary,
        "brand_revenue": brand_rev_out,
        "channel_revenue": channel_rev_out,
        "ad_performance": ad_out,
        "ads_landing": ads_landing,
        "ad_start_idx": ad_start_idx,
        "brand_performance": brand_perf_out,
        "paid_organic": {
            "paid": paid_monthly,
            "paid_proj": proj_array(paid_monthly),
            "organic": organic_monthly,
            "organic_proj": proj_array(organic_monthly),
        },
        "waterfall": {
            "revenue": total_rev_monthly,
            "revenue_proj": proj_array(total_rev_monthly),
            "cogs": total_cogs_monthly,
            "cogs_proj": proj_array(total_cogs_monthly),
            "gross_margin": total_gm_monthly,
            "gross_margin_proj": proj_array(total_gm_monthly),
            "variable_costs": _variable_costs_arr,
            "variable_detail": {
                "amz_ref_fee": _amz_fee_arr,
                "fba_fulfillment": [v if v is not None else 0 for v in _fba_arr],
            },
            "cm_before_mkt": _cm_before_mkt_arr,
            "cm_before_mkt_proj": proj_array(_cm_before_mkt_arr),
            "ad_spend": total_ad_spend_monthly,
            "ad_spend_proj": proj_array(total_ad_spend_monthly),
            "discounts": total_disc_monthly,
            "discounts_proj": proj_array(total_disc_monthly),
            "discounts_detail": {
                "shopify_disc": _shop_disc_arr,
            },
            "seeding": total_seeding_monthly,
            "seeding_proj": proj_array(total_seeding_monthly),
            "influencer_detail": {
                "paid": _inf_paid_arr,
                "nonpaid": _inf_nonpaid_arr,
            },
            "mkt_total": total_mkt_monthly,
            "mkt_total_proj": proj_array(total_mkt_monthly),
            "contribution_margin": total_cm_monthly,
            "contribution_margin_proj": proj_array(total_cm_monthly),
        },
        "search_queries": search_data,
        "search_by_brand": search_by_brand,
        "keyword_performance": keyword_performance,
        "gsc_queries": gsc_data,
        "keyword_rankings": keyword_rankings,
        "kw_positions_summary": kw_positions_summary,
        "keyword_volumes": keyword_volumes,
        "brand_analytics": ba_data,
        "brand_analytics_category": ba_category_top,
        "traffic_sources": traffic_data,
        "channel_traffic": channel_traffic,
        "ad_creative_breakdown": ad_creative_data,
        "amz_sessions": {
            "7d": {"sessions": amz_sessions_7d.get("sessions", 0), "pageViews": amz_sessions_7d.get("pageViews", 0)},
            "30d": {"sessions": amz_sessions_30d.get("sessions", 0), "pageViews": amz_sessions_30d.get("pageViews", 0)},
        },
        "pnl_polar": pnl_polar,
        "klaviyo": klaviyo_data,
        "campaign_detail": campaign_detail,
        "brand_ad_by_platform": bap_monthly_out,
        "brand_ad_daily": brand_ad_daily_out,
        "weekly": weekly_data,
        "hero_products": hero_data,
        "search_ranking": _build_autocomplete_data(),
    }

    # ── Write output ──────────────────────────────────────────────────────────
    os.makedirs(OUTPUT.parent, exist_ok=True)
    js_content = "const FIN_DATA = " + json.dumps(fin_data, ensure_ascii=False, indent=1) + ";"
    OUTPUT.write_text(js_content, encoding="utf-8")
    print(f"\n  fin_data.js written: {len(js_content):,} chars")
    print(f"  Summary (30d): Rev ${summary['30d']['total_revenue']:,} | "
          f"Ad ${summary['30d']['total_ad_spend']:,} | "
          f"GM {summary['30d']['gm_pct']}% | "
          f"MER {summary['30d']['mer']}")

    # ── L7: Sanity Check (from 검증이 M-066~M-069) ────────────────────────
    print("\n[SANITY] P&L output validation...")
    warnings = []
    pnl = fin_data["pnl_polar"]
    fy_idx = pnl["fy2025_idx"]
    for i, m in enumerate(pnl["months"]):
        if m == "FY2025":
            continue
        r = pnl["total_revenue"][i]
        c = pnl["cogs"][i]
        ad = pnl["ad_spend"]["total"][i]
        disc = pnl.get("discounts", [0]*len(pnl["months"]))[i]
        inf = pnl.get("influencer_spend", [0]*len(pnl["months"]))[i]
        if r > 0:
            cogs_pct = c / r * 100
            if cogs_pct < 25:
                warnings.append(f"  [WARN] {m}: COGS% {cogs_pct:.0f}% < 25% (partial month data?)")
            elif cogs_pct > 45:
                warnings.append(f"  [WARN] {m}: COGS% {cogs_pct:.0f}% > 45% (revenue using net instead of gross?)")
            # Disc is Shopify only (typically 1-5% of total rev since Amazon is ~80% of rev)
            disc_pct = disc / r * 100
        if ad == 0 and m not in ("Jan 25", "Feb 25", "Mar 25", "Apr 25", "May 25"):
            warnings.append(f"  [WARN] {m}: Ad Spend $0 (backfill missing?)")
        if inf == 0:
            warnings.append(f"  [WARN] {m}: Influencer $0 (PR orders not reflected?)")
    # FY month count check
    fy_month_count = fy_idx
    if fy_month_count < 12 and fy_month_count > 0:
        warnings.append(f"  [WARN] FY2025 has only {fy_month_count} months (expected 12, Jan-May missing?)")
    if warnings:
        for w in warnings:
            print(w)
        print(f"  {len(warnings)} warnings found")
    else:
        print("  All checks PASS")

    if args.push:
        import subprocess
        os.chdir(str(ROOT))
        subprocess.run(["git", "add", str(OUTPUT)], check=True)
        subprocess.run(["git", "commit", "-m", "auto: update financial KPI data [skip ci]"], check=False)
        subprocess.run(["git", "push"], check=True)
        print("  Pushed to GitHub")


def _build_autocomplete_data() -> dict:
    """Load Amazon autocomplete rank data — PG first, fallback to local cache."""
    rows = []

    # Try PG (has all historical data)
    try:
        from data_keeper_client import DataKeeper
        dk = DataKeeper()
        rows = dk.get("amazon_autocomplete_daily", days=120) or []
        if rows:
            print(f"  [autocomplete] {len(rows)} rows from PG")
    except Exception as e:
        print(f"  [autocomplete] PG failed ({e}), trying cache")

    # Fallback to local cache
    if not rows:
        cache_path = os.path.join(os.path.dirname(__file__), "..", ".tmp", "datakeeper", "amazon_autocomplete_daily.json")
        if os.path.exists(cache_path):
            with open(cache_path, "r", encoding="utf-8") as f:
                rows = json.load(f)

    if not rows:
        print("  [autocomplete] No data found")
        return {}

    # Build two structures:
    # 1. "latest" — most recent snapshot per brand/market (for summary cards)
    # 2. "trends" — daily time series per brand/market/keyword (for trend chart)
    latest_date = max(r.get("date", "") for r in rows)

    latest = {}
    trends = {}  # brand → market → keyword → [{date, score, position}, ...]

    for r in rows:
        brand = r.get("brand", "")
        market = r.get("market", "US")
        kw = r.get("keyword", "")
        score = r.get("rank_score", 0)
        pos = r.get("position", -1)
        date = r.get("date", "")

        # Latest snapshot
        if date == latest_date:
            if brand not in latest:
                latest[brand] = {"US": [], "JP": []}
            latest[brand][market].append({"keyword": kw, "score": score, "position": pos, "date": date})

        # Trends
        if brand not in trends:
            trends[brand] = {"US": {}, "JP": {}}
        if kw not in trends[brand][market]:
            trends[brand][market][kw] = []
        trends[brand][market][kw].append({"date": date, "score": score, "position": pos})

    # Sort trend entries by date
    for brand in trends:
        for mkt in trends[brand]:
            for kw in trends[brand][mkt]:
                trends[brand][mkt][kw].sort(key=lambda x: x["date"])

    all_dates = sorted(set(r.get("date", "") for r in rows))
    print(f"  [autocomplete] {len(rows)} rows, {len(all_dates)} dates, {len(latest)} brands")
    return {"latest": latest, "trends": trends, "dates": all_dates}


if __name__ == "__main__":
    generate()
