"""
Excel V8 Dashboard Reader

목표: V8 Excel 파일에서 데이터를 읽어서 
      Streamlit 대시보드에서 쓸 수 있는 Python 형태로 변환

구조:
1. read_v8_file()          - 전체 Excel 열기
2. get_rakuten_data()      - Rakuten 탭에서 월별 데이터 추출
3. get_amazon_data()       - Amazon 탭에서 월별 데이터 추출
4. get_meta_data()         - Meta 탭에서 캠페인별 데이터 추출
5. get_overview_data()     - Overview 탭에서 요약 데이터 추출
"""

from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
from typing import Dict, List, Tuple


# ===================================================================
# 1. Excel 파일 경로 설정
# ===================================================================

PROJECT_ROOT = Path(__file__).parent.parent
EXCEL_PATH = PROJECT_ROOT / "Japan_Marketing Plan_Monthly_V8.xlsx"

print(f"[Dashboard Reader] Excel path: {EXCEL_PATH}")
print(f"[Dashboard Reader] File exists: {EXCEL_PATH.exists()}")


# ===================================================================
# 2. V8 구조 정의 (update_v8.py와 동일)
# ===================================================================

# 분석할 월들
MONTHS = [
    "2025-11", "2025-12", "2026-01", "2026-02",
    "2026-03", "2026-04", "2026-05", "2026-06",
]

# Rakuten/Amazon 행 매핑
def _rak_amz_rows(month_key):
    """
    각 월별 데이터가 몇 행에 있는지 반환
    
    예시:
      2025-11: Row 3-7   (블록 시작=3, TOTAL=7)
      2025-12: Row 8-12  (블록 시작=8, TOTAL=12)
      ...
    """
    idx = MONTHS.index(month_key)
    start = 3 + idx * 5
    total = start + 4
    return start, total


# Meta 행 매핑 (캠페인 개수가 월마다 다름)
META_BLOCKS = {
    "2025-11": {"start": 3, "total": 7, "campaign_start": 3, "campaign_slots": 4},
    "2025-12": {"start": 8, "total": 14, "campaign_start": 8, "campaign_slots": 5},
    "2026-01": {"start": 15, "total": 23, "campaign_start": 15, "campaign_slots": 7},
    "2026-02": {"start": 25, "total": 31, "campaign_start": 25, "campaign_slots": 4},
    "2026-03": {"start": 32, "total": 38, "campaign_start": 32, "campaign_slots": 4},
    "2026-04": {"start": 39, "total": 45, "campaign_start": 39, "campaign_slots": 4},
    "2026-05": {"start": 46, "total": 52, "campaign_start": 46, "campaign_slots": 4},
    "2026-06": {"start": 53, "total": 59, "campaign_start": 53, "campaign_slots": 4},
}


# ===================================================================
# 3. Excel 읽기 함수들 (각 탭별로 분리)
# ===================================================================

def read_v8_file():
    """
    V8 Excel 파일을 열기
    
    Returns:
        openpyxl Workbook 객체 (또는 None if 파일 없음)
    """
    if not EXCEL_PATH.exists():
        print(f"[ERROR] V8 파일을 찾을 수 없습니다: {EXCEL_PATH}")
        return None
    
    try:
        wb = load_workbook(str(EXCEL_PATH), data_only=True)
        print(f"[OK] V8 파일 로드됨: {EXCEL_PATH.name}")
        print(f"     탭 목록: {wb.sheetnames}")
        return wb
    except Exception as e:
        print(f"[ERROR] Excel 열기 실패: {e}")
        return None


def get_rakuten_data(wb, month_key: str) -> Dict:
    """
    RAKUTEN 탭에서 특정 월의 데이터 추출
    
    Args:
        wb: Workbook 객체
        month_key: "2026-01" 형식
    
    Returns:
        {
            "month": "2026-01",
            "budget": 100000,           # D열 (Budget)
            "actual_spend": 95000,      # F열 (Actual Ad Spend)
            "actual_sales": 500000,     # G열 (Actual Sales)
            "notes": "...",             # H열
            "roas": 5.26                # 자동 계산
        }
    """
    
    if month_key not in MONTHS:
        return None
    
    ws = wb["RAKUTEN"]
    _, total_row = _rak_amz_rows(month_key)
    
    # 각 열의 의미 (Excel column = Python column number)
    # B=2, C=3, D=4, E=5, F=6, G=7, H=8
    
    budget = ws.cell(row=total_row, column=4).value or 0  # D = Budget
    actual_spend = ws.cell(row=total_row, column=6).value or 0  # F = Actual Ad Spend
    actual_sales = ws.cell(row=total_row, column=7).value or 0  # G = Actual Sales
    notes = ws.cell(row=total_row, column=8).value or ""  # H = Notes
    
    # ROAS 자동 계산 (Sales / Ad Spend)
    roas = round(actual_sales / actual_spend, 2) if actual_spend > 0 else 0
    
    return {
        "month": month_key,
        "budget": float(budget),
        "actual_spend": float(actual_spend),
        "actual_sales": float(actual_sales),
        "notes": str(notes) if notes else "",
        "roas": roas,
    }


def get_amazon_data(wb, month_key: str) -> Dict:
    """
    AMAZON 탭에서 특정 월의 데이터 추출 (RAKUTEN과 동일 구조)
    """
    
    if month_key not in MONTHS:
        return None
    
    ws = wb["AMAZON"]
    _, total_row = _rak_amz_rows(month_key)
    
    budget = ws.cell(row=total_row, column=4).value or 0
    actual_spend = ws.cell(row=total_row, column=6).value or 0
    actual_sales = ws.cell(row=total_row, column=7).value or 0
    notes = ws.cell(row=total_row, column=8).value or ""
    
    roas = round(actual_sales / actual_spend, 2) if actual_spend > 0 else 0
    
    return {
        "month": month_key,
        "budget": float(budget),
        "actual_spend": float(actual_spend),
        "actual_sales": float(actual_sales),
        "notes": str(notes) if notes else "",
        "roas": roas,
    }


def get_meta_data(wb, month_key: str) -> Dict:
    """
    META 탭에서 특정 월의 데이터 추출
    
    Meta는 캠페인별로 여러 행이므로 조금 복잡합니다.
    
    Returns:
        {
            "month": "2026-01",
            "total_spend": 150000,      # 모든 캠페인의 합
            "campaigns": [
                {
                    "name": "Campaign A",
                    "spend": 50000,
                    "impressions": 100000,
                    "clicks": 5000,
                },
                ...
            ]
        }
    """
    
    if month_key not in META_BLOCKS:
        return None
    
    ws = wb["META"]
    block = META_BLOCKS[month_key]
    
    campaigns = []
    total_spend = 0
    
    # 각 캠페인 행을 순회
    for i in range(block["campaign_slots"]):
        r = block["campaign_start"] + i
        
        campaign_name = ws.cell(row=r, column=4).value  # D = Campaign name
        spend = ws.cell(row=r, column=7).value or 0  # G = Actual Ad Spend
        notes = ws.cell(row=r, column=9).value or ""  # I = Notes (impressions/clicks)
        
        if campaign_name:  # 캠페인명이 있으면 유효한 행
            # Notes에서 impressions, clicks 추출
            # 예: "Impr: 100,000 / Clicks: 5,000"
            impressions = 0
            clicks = 0
            
            if notes:
                # 간단한 파싱 (더 나은 방법도 있지만 이걸로 충분)
                parts = str(notes).split(" / ")
                for part in parts:
                    if "Impr" in part:
                        try:
                            impressions = int(part.split(":")[1].strip().replace(",", ""))
                        except:
                            pass
                    elif "Clicks" in part:
                        try:
                            clicks = int(part.split(":")[1].strip().replace(",", ""))
                        except:
                            pass
            
            campaigns.append({
                "name": str(campaign_name),
                "spend": float(spend),
                "impressions": impressions,
                "clicks": clicks,
            })
            
            total_spend += float(spend)
    
    return {
        "month": month_key,
        "total_spend": total_spend,
        "campaign_count": len(campaigns),
        "campaigns": campaigns,
    }


def get_overview_data(wb) -> Dict:
    """
    Overview(Monthly) 탭에서 모든 월의 요약 데이터 추출
    
    이 탭에는 모든 채널 (Rakuten, Amazon, Meta)의 판매액과 지출이 정리되어 있습니다.
    """
    
    ws = wb["Overview(Monthly)"]
    
    summary = {}
    
    # 각 월마다 Overview 섹션이 있음
    # 예: Nov = Row 4-10, Dec = Row 13-19, Jan = Row 22-28, ...
    
    for month_key in MONTHS:
        # 간단한 방법: 각 월의 TOTAL 행을 찾기
        # 실제로는 더 정확한 매핑이 필요하지만, 우선은 생략
        pass
    
    return summary


# ===================================================================
# 4. 편의 함수: 모든 월의 데이터 한 번에 가져오기
# ===================================================================

def get_all_data() -> Dict:
    """
    V8의 모든 데이터를 한 번에 가져오기
    
    Returns:
        {
            "rakuten": [
                {"month": "2025-11", "actual_sales": ..., "roas": ...},
                ...
            ],
            "amazon": [...],
            "meta": [...],
            "last_updated": "2026-03-03 14:30",
        }
    """
    
    wb = read_v8_file()
    if wb is None:
        return None
    
    data = {
        "rakuten": [],
        "amazon": [],
        "meta": [],
        "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    
    for month_key in MONTHS:
        # Rakuten
        rak = get_rakuten_data(wb, month_key)
        if rak:
            data["rakuten"].append(rak)
        
        # Amazon
        amz = get_amazon_data(wb, month_key)
        if amz:
            data["amazon"].append(amz)
        
        # Meta
        meta = get_meta_data(wb, month_key)
        if meta:
            data["meta"].append(meta)
    
    return data


# ===================================================================
# 5. 테스트 (이 파일을 직접 실행했을 때)
# ===================================================================

if __name__ == "__main__":
    print("\n" + "=" * 60)
    print("Dashboard Reader - 테스트 모드")
    print("=" * 60)
    
    data = get_all_data()
    
    if data:
        print("\n📊 RAKUTEN 데이터:")
        for r in data["rakuten"]:
            print(f"  {r['month']}: Sales ¥{r['actual_sales']:,.0f}, "
                  f"Spend ¥{r['actual_spend']:,.0f}, ROAS {r['roas']:.2f}x")
        
        print("\n📊 AMAZON 데이터:")
        for a in data["amazon"]:
            print(f"  {a['month']}: Sales ¥{a['actual_sales']:,.0f}, "
                  f"Spend ¥{a['actual_spend']:,.0f}, ROAS {a['roas']:.2f}x")
        
        print("\n📊 META 데이터:")
        for m in data["meta"]:
            print(f"  {m['month']}: Total Spend ¥{m['total_spend']:,.0f}, "
                  f"{m['campaign_count']} campaigns")
            for c in m["campaigns"]:
                print(f"    - {c['name']}: ¥{c['spend']:,.0f}")
        
        print(f"\n마지막 업데이트: {data['last_updated']}")
    else:
        print("[ERROR] 데이터를 불러올 수 없습니다.")
