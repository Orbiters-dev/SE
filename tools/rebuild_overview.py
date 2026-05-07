"""
Rebuild Overview(Monthly) tab - V4 clean structure

Per month:
  Direct: Rakuten, Amazon → Subtotal
  Indirect: Meta, Influencer → Subtotal
  TOTAL → Blended ROAS (only here)

Columns: Budget | Plan | Actual (direct metrics) | Total row gets extra Sales + Blended ROAS
Months: Nov 2025 - Jun 2026
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.formatting.rule import CellIsRule

SRC = "Japan_Marketing Plan_Monthly_V2.xlsx"
DST = "Japan_Marketing Plan_Monthly_V5.xlsx"

# ── Colors ──────────────────────────────────────────────────────────────
NAVY       = "1B2A4A"
DARK_BLUE  = "2C3E6B"
MID_BLUE   = "3A5BA0"
LIGHT_BLUE = "D6E4F0"
LIGHT_GRAY = "F2F2F2"
WHITE      = "FFFFFF"
RED_TEXT    = "CC0000"
GREEN_BG   = "E2EFDA"
RED_BG     = "FCE4EC"
YELLOW_BG  = "FFF9C4"
BORDER_CLR = "B0B0B0"
TOTAL_BG   = "E8EAF0"
SUB_DIR_BG = "EBF1F8"
SUB_IND_BG = "F5F0E8"
INDIRECT_BG= "FAFAF5"

# ── Styles ──────────────────────────────────────────────────────────────
thin = Border(
    left=Side("thin", color=BORDER_CLR), right=Side("thin", color=BORDER_CLR),
    top=Side("thin", color=BORDER_CLR), bottom=Side("thin", color=BORDER_CLR))
bot_med = Border(
    left=Side("thin", color=BORDER_CLR), right=Side("thin", color=BORDER_CLR),
    top=Side("thin", color=BORDER_CLR), bottom=Side("medium", color="404040"))
bot_dbl = Border(
    left=Side("thin", color=BORDER_CLR), right=Side("thin", color=BORDER_CLR),
    top=Side("thin", color=BORDER_CLR), bottom=Side("double", color=NAVY))
hdr_bot = Border(
    left=Side("thin", color=BORDER_CLR), right=Side("thin", color=BORDER_CLR),
    top=Side("thin", color=BORDER_CLR), bottom=Side("medium", color=NAVY))

f_title     = Font(name="Aptos", size=16, bold=True, color=NAVY)
f_grp       = Font(name="Aptos", size=11, bold=True, color=WHITE)
f_hdr       = Font(name="Aptos", size=10, bold=True, color=NAVY)
f_month     = Font(name="Aptos", size=11, bold=True, color=NAVY)
f_sec       = Font(name="Aptos", size=9,  bold=True, color="666666")
f_ch        = Font(name="Aptos", size=10, bold=True, color="333333")
f_data      = Font(name="Aptos", size=10, color="333333")
f_sub       = Font(name="Aptos", size=10, bold=True, color="2C3E6B")
f_total     = Font(name="Aptos", size=10, bold=True, color=RED_TEXT)
f_total_d   = Font(name="Aptos", size=10, bold=True, color="1a1a1a")
f_total_roas= Font(name="Aptos", size=11, bold=True, color=RED_TEXT)
f_note      = Font(name="Aptos", size=9, italic=True, color="666666")
f_roas      = Font(name="Aptos", size=10, bold=True, color="333333")

fl_navy     = PatternFill("solid", fgColor=NAVY)
fl_dblue    = PatternFill("solid", fgColor=DARK_BLUE)
fl_mblue    = PatternFill("solid", fgColor=MID_BLUE)
fl_lblue    = PatternFill("solid", fgColor=LIGHT_BLUE)
fl_lgray    = PatternFill("solid", fgColor=LIGHT_GRAY)
fl_white    = PatternFill("solid", fgColor=WHITE)
fl_total    = PatternFill("solid", fgColor=TOTAL_BG)
fl_sub_d    = PatternFill("solid", fgColor=SUB_DIR_BG)
fl_sub_i    = PatternFill("solid", fgColor=SUB_IND_BG)
fl_indir    = PatternFill("solid", fgColor=INDIRECT_BG)

ac = Alignment(horizontal="center", vertical="center", wrap_text=True)
ar = Alignment(horizontal="right", vertical="center")
al = Alignment(horizontal="left", vertical="center")

# Number formats - proper yen
YEN  = '[$\u00a5-411]#,##0'
YENN = '[$\u00a5-411]#,##0;[Red]([$\u00a5-411]#,##0)'
PCT  = '0.0%'
PCTW = '0%'
ROAS = '0.00"x"'

# CM% references from contribution margin sheet
CM = {
    "Rakuten":    "'채널별 공헌이익'!K20",
    "Amazon":     "'채널별 공헌이익'!K21",
    "Meta":       "'채널별 공헌이익'!K22",
    "Influencer": "'채널별 공헌이익'!K23",
}

# ── Column layout ───────────────────────────────────────────────────────
# B: Month  C: Channel
# BUDGET:  D: Budget(Monthly)  E: Budget(Daily)  F: Alloc%
# PLAN:    G: Target ROAS  H: Target Sales  I: CM%  J: CM  K: CM-Budget
# ACTUAL:  L: Ad Spend  M: Ad Sales(Direct)  N: Direct ROAS  O: Budget Exec%  P: Spend Variance
# TOTAL-ONLY: Q: Total Sales  R: Blended ROAS

MONTHS = [
    {
        "label": "2025.11\n(Nov)",
        "budgets":      {"Rakuten": "=RAKUTEN!D7"},
        "target_roas":  {},
        "target_sales": {"Rakuten": 100000},
        "actual_spend": {"Rakuten": 15280},
        "ad_sales":     {"Rakuten": 0},
        "total_sales":  {"Rakuten": 36800},
    },
    {
        "label": "2025.12\n(Dec)",
        "budgets":      {"Rakuten": 200000, "Amazon": 100000, "Meta": 100000, "Influencer": 200000},
        "target_roas":  {},
        "target_sales": {},
        "actual_spend": {"Rakuten": "=RAKUTEN!F12", "Amazon": "=AMAZON!F12", "Meta": "=META!G14", "Influencer": 56520},
        "ad_sales":     {"Rakuten": 5800, "Amazon": 26000},
        "total_sales":  {"Rakuten": 14230, "Amazon": 30200},
    },
    {
        "label": "2026.01\n(Jan)",
        "budgets":      {"Rakuten": 70000, "Amazon": 70000, "Meta": 200000, "Influencer": 200000},
        "target_roas":  {"Rakuten": 1, "Amazon": 1, "Meta": 1, "Influencer": 1},
        "target_sales": {},
        "actual_spend": {"Rakuten": "=RAKUTEN!F14", "Amazon": "=AMAZON!F17", "Meta": "=META!G23", "Influencer": 46000},
        "ad_sales":     {"Rakuten": 28590, "Amazon": 77718},
        "total_sales":  {"Rakuten": "=RAKUTEN!G14", "Amazon": "=AMAZON!G17"},
    },
    {
        "label": "2026.02\n(Feb)",
        "budgets":      {"Rakuten": 50000, "Amazon": 100000, "Meta": 300000, "Influencer": 200000},
        "target_roas":  {},  # feb refs set later
        "target_sales": {"Rakuten": 50000, "Amazon": 100000, "Meta": 300000},
        "actual_spend": {},
        "ad_sales":     {},
        "total_sales":  {},
        "_feb": True,
    },
    {"label": "2026.03\n(Mar)", "budgets": {}, "target_roas": {}, "target_sales": {}, "actual_spend": {}, "ad_sales": {}, "total_sales": {}},
    {"label": "2026.04\n(Apr)", "budgets": {}, "target_roas": {}, "target_sales": {}, "actual_spend": {}, "ad_sales": {}, "total_sales": {}},
    {"label": "2026.05\n(May)", "budgets": {}, "target_roas": {}, "target_sales": {}, "actual_spend": {}, "ad_sales": {}, "total_sales": {}},
    {"label": "2026.06\n(Jun)", "budgets": {}, "target_roas": {}, "target_sales": {}, "actual_spend": {}, "ad_sales": {}, "total_sales": {}},
]


def sc(ws, r, c, val=None, font=None, fill=None, fmt=None, align=None, brd=None):
    """Set cell helper."""
    cell = ws.cell(row=r, column=c)
    if val is not None:
        cell.value = val
    if font: cell.font = font
    if fill: cell.fill = fill
    if fmt:  cell.number_format = fmt
    if align: cell.alignment = align
    cell.border = brd or thin
    return cell


def write_ch_row(ws, r, ch, month, total_ref, indirect=False):
    """Write one channel row (B-P only, no Q/R)."""
    fill = fl_indir if indirect else fl_white
    is_direct = ch in ("Rakuten", "Amazon")

    sc(ws, r, 3, ch, f_ch, fill, align=al)

    bval = month["budgets"].get(ch)
    sc(ws, r, 4, bval, f_data, fill, YEN, ar)
    sc(ws, r, 5, f"=D{r}/30" if bval is not None else None, f_data, fill, YEN, ar)
    sc(ws, r, 6, f"=IFERROR(D{r}/D{total_ref},0)", f_data, fill, PCTW, ar)

    gv = month["target_roas"].get(ch)
    sc(ws, r, 7, gv, f_data, fill, ROAS, ar)

    hv = month["target_sales"].get(ch)
    if hv is None and gv is not None and bval is not None:
        hv = f"=G{r}*D{r}"
    sc(ws, r, 8, hv, f_data, fill, YEN, ar)

    iv = f"={CM[ch]}" if ch in CM else None
    sc(ws, r, 9, iv, f_data, fill, PCT, ar)
    sc(ws, r, 10, f"=I{r}*H{r}", f_data, fill, YENN, ar)
    sc(ws, r, 11, f"=J{r}-D{r}", f_data, fill, YENN, ar)

    lv = month["actual_spend"].get(ch)
    sc(ws, r, 12, lv, f_data, fill, YEN, ar)

    # M: Ad Sales - only for direct channels
    mv = month["ad_sales"].get(ch) if is_direct else None
    sc(ws, r, 13, mv, f_data, fill, YEN, ar)

    # N: Direct ROAS - only for direct channels
    nv = f'=IFERROR(M{r}/L{r},"-")' if is_direct else None
    sc(ws, r, 14, nv, f_roas if is_direct else f_data, fill, ROAS, ar)

    sc(ws, r, 15, f'=IFERROR(L{r}/D{r},"-")', f_data, fill, PCT, ar)
    sc(ws, r, 16, f"=D{r}-L{r}", f_data, fill, YENN, ar)

    # Q, R: empty for individual channels
    sc(ws, r, 17, None, f_data, fill, YEN, ar)
    sc(ws, r, 18, None, f_data, fill, ROAS, ar)


def write_sub(ws, r, label, fr, lr, fill, brd):
    """Subtotal row."""
    sc(ws, r, 3, label, f_sub, fill, align=al, brd=brd)
    for col, fmt in [(4,YEN),(8,YEN),(10,YENN),(11,YENN),(12,YEN),(13,YEN)]:
        cl = get_column_letter(col)
        sc(ws, r, col, f"=SUM({cl}{fr}:{cl}{lr})", f_sub, fill, fmt, ar, brd)
    sc(ws, r, 5, f"=D{r}/30", f_sub, fill, YEN, ar, brd)
    sc(ws, r, 6, None, f_sub, fill, PCTW, ar, brd)
    sc(ws, r, 14, f'=IFERROR(M{r}/L{r},"-")', f_sub, fill, ROAS, ar, brd)
    sc(ws, r, 15, f'=IFERROR(L{r}/D{r},"-")', f_sub, fill, PCT, ar, brd)
    sc(ws, r, 16, f"=D{r}-L{r}", f_sub, fill, YENN, ar, brd)
    for col in [7, 9, 17, 18]:
        sc(ws, r, col, None, f_sub, fill, brd=brd)


def build():
    wb = load_workbook(SRC)
    if "Overview(Monthly)" in wb.sheetnames:
        del wb["Overview(Monthly)"]

    ws = wb.create_sheet("Overview(Monthly)", 0)
    ws.sheet_properties.tabColor = NAVY
    ws.freeze_panes = "D4"

    # Column widths
    for c, w in {"A":2,"B":12,"C":15,"D":16,"E":13,"F":10,"G":12,"H":15,
                 "I":9,"J":14,"K":14,"L":15,"M":15,"N":12,"O":12,"P":14,
                 "Q":16,"R":14}.items():
        ws.column_dimensions[c].width = w

    # Row 1: Title
    ws.row_dimensions[1].height = 36
    ws.merge_cells("B1:R1")
    c = ws["B1"]; c.value = "GROSMIMI JAPAN - MONTHLY MARKETING OVERVIEW"
    c.font = f_title; c.alignment = Alignment(horizontal="left", vertical="center")

    # Row 2: Group headers
    ws.row_dimensions[2].height = 28
    grps = [
        ("B2:C2", "", None),
        ("D2:F2", "BUDGET", fl_navy),
        ("G2:K2", "PLAN", fl_dblue),
        ("L2:P2", "ACTUAL", fl_mblue),
        ("Q2:R2", "BLENDED *", fl_navy),
    ]
    for rng, label, fill in grps:
        mn, mr1, mx, mr2 = range_boundaries(rng)
        for col in range(mn, mx + 1):
            cl = ws.cell(row=2, column=col)
            if col == mn: cl.value = label
            cl.font = f_grp; cl.alignment = ac; cl.border = thin
            if fill: cl.fill = fill
        ws.merge_cells(rng)

    # Row 3: Column headers
    ws.row_dimensions[3].height = 40
    hdrs = [
        ("B","Month"),("C","Channel"),("D","Budget\n(Monthly)"),("E","Budget\n(Daily)"),
        ("F","Alloc %"),("G","Target\nROAS"),("H","Target\nSales"),("I","CM %"),
        ("J","CM"),("K","CM\n- Budget"),("L","Ad Spend"),("M","Ad Sales\n(Direct)"),
        ("N","Direct\nROAS"),("O","Budget\nExec %"),("P","Spend\nVariance"),
        ("Q","Total\nSales *"),("R","Blended\nROAS *"),
    ]
    for cl, label in hdrs:
        c = ws[f"{cl}3"]; c.value = label; c.font = f_hdr
        c.fill = fl_lblue; c.alignment = ac; c.border = hdr_bot

    # ── Monthly blocks ──────────────────────────────────────────────────
    jan_rows = {}
    cur = 4

    for mi, month in enumerate(MONTHS):
        block_start = cur

        # Section: Direct
        sc(ws, cur, 3, "Direct Sales Channels", f_sec, fl_lblue, align=al)
        for col in range(4, 19): sc(ws, cur, col, fill=fl_lblue)
        cur += 1

        # Track row positions for each channel
        ch_row = {}

        dir_first = cur
        for ch in ["Rakuten", "Amazon"]:
            write_ch_row(ws, cur, ch, month, "PH", indirect=False)
            ch_row[ch] = cur
            if month["label"].startswith("2026.01"):
                jan_rows[ch] = cur
            cur += 1
        dir_last = cur - 1

        sub_dir = cur
        write_sub(ws, cur, "Subtotal (Direct)", dir_first, dir_last, fl_sub_d, bot_med)
        cur += 1

        # Section: Indirect
        sc(ws, cur, 3, "Indirect / Traffic Channels  (\u2192 Rakuten)", f_sec, fl_lblue, align=al)
        for col in range(4, 19): sc(ws, cur, col, fill=fl_lblue)
        cur += 1

        ind_first = cur
        for ch in ["Meta", "Influencer"]:
            write_ch_row(ws, cur, ch, month, "PH", indirect=True)
            ch_row[ch] = cur
            if month["label"].startswith("2026.01"):
                jan_rows[ch] = cur
            cur += 1
        ind_last = cur - 1

        sub_ind = cur
        write_sub(ws, cur, "Subtotal (Indirect)", ind_first, ind_last, fl_sub_i, bot_med)
        cur += 1

        # ── Rakuten: Q = Total Sales, R = Blended ROAS (includes Meta+Influencer cost)
        rak_r = ch_row["Rakuten"]
        amz_r = ch_row["Amazon"]
        meta_r = ch_row.get("Meta")
        inf_r = ch_row.get("Influencer")

        ts = month.get("total_sales", {})
        rak_ts = ts.get("Rakuten")
        amz_ts = ts.get("Amazon")

        # Rakuten Q: Total Sales
        sc(ws, rak_r, 17, rak_ts, f_data, fl_white, YEN, ar)
        # Rakuten R: Blended ROAS = Total Sales / (Rakuten spend + Meta spend + Influencer spend)
        sc(ws, rak_r, 18,
           f'=IFERROR(Q{rak_r}/(L{rak_r}+L{meta_r}+L{inf_r}),"-")',
           f_roas, fl_white, ROAS, ar)

        # Amazon Q: Total Sales (reference only)
        sc(ws, amz_r, 17, amz_ts, f_data, fl_white, YEN, ar)
        # Amazon R: no Blended ROAS (would be misleading due to indirect brand effects)
        sc(ws, amz_r, 18, "-", f_data, fl_white, align=ar)

        # TOTAL row
        total_r = cur
        sc(ws, total_r, 3, "TOTAL", f_total, fl_total, align=al, brd=bot_dbl)

        # Sum both subtotals
        for col, fmt in [(4,YEN),(8,YEN),(10,YENN),(11,YENN),(12,YEN),(13,YEN)]:
            cl = get_column_letter(col)
            sc(ws, total_r, col, f"={cl}{sub_dir}+{cl}{sub_ind}", f_total_d, fl_total, fmt, ar, bot_dbl)

        sc(ws, total_r, 5, f"=D{total_r}/30", f_total_d, fl_total, YEN, ar, bot_dbl)
        sc(ws, total_r, 6, 1, f_total_d, fl_total, PCTW, ar, bot_dbl)
        sc(ws, total_r, 14, f'=IFERROR(M{total_r}/L{total_r},"-")', f_total_d, fl_total, ROAS, ar, bot_dbl)
        sc(ws, total_r, 15, f'=IFERROR(L{total_r}/D{total_r},"-")', f_total_d, fl_total, PCT, ar, bot_dbl)
        sc(ws, total_r, 16, f"=D{total_r}-L{total_r}", f_total_d, fl_total, YENN, ar, bot_dbl)

        # Q total: Rakuten Total Sales + Amazon Total Sales
        sc(ws, total_r, 17, f"=Q{rak_r}+Q{amz_r}", f_total_d, fl_total, YEN, ar, bot_dbl)

        # R total: Blended ROAS = Total Sales / Total Ad Spend (all channels)
        sc(ws, total_r, 18, f'=IFERROR(Q{total_r}/L{total_r},"-")',
           f_total_roas, fl_total, ROAS, ar, bot_dbl)

        # Fill empty cells in total
        for col in [7, 9]:
            sc(ws, total_r, col, None, f_total_d, fl_total, brd=bot_dbl)

        # Fix Allocation % references
        for ci, ch in enumerate(["Rakuten", "Amazon"]):
            ws.cell(row=dir_first + ci, column=6).value = f"=IFERROR(D{dir_first+ci}/D{total_r},0)"
        for ci, ch in enumerate(["Meta", "Influencer"]):
            ws.cell(row=ind_first + ci, column=6).value = f"=IFERROR(D{ind_first+ci}/D{total_r},0)"

        # Month label merge
        for r in range(block_start, total_r + 1):
            ws.cell(row=r, column=2).border = thin
        ws.merge_cells(start_row=block_start, start_column=2, end_row=total_r, end_column=2)
        mc = ws.cell(row=block_start, column=2)
        mc.value = month["label"]; mc.font = f_month; mc.alignment = ac; mc.fill = fl_white

        cur = total_r + 2

    # ── Feb Target ROAS refs ────────────────────────────────────────────
    for r in range(4, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v and "2026.02" in str(v):
            for sr in range(r, r + 15):
                cv = ws.cell(row=sr, column=3).value
                if cv == "Rakuten" and jan_rows.get("Rakuten"):
                    ws.cell(row=sr, column=7).value = f"=G{jan_rows['Rakuten']}"
                elif cv == "Amazon" and jan_rows.get("Amazon"):
                    ws.cell(row=sr, column=7).value = f"=G{jan_rows['Amazon']}"
                elif cv == "Meta":
                    ws.cell(row=sr, column=7).value = 1
                elif cv == "TOTAL":
                    break
            break

    # ── Conditional formatting ──────────────────────────────────────────
    last_r = cur - 2
    gf = PatternFill("solid", fgColor=GREEN_BG)
    gn = Font(name="Aptos", size=10, bold=True, color="2E7D32")
    rf = PatternFill("solid", fgColor=RED_BG)
    rn = Font(name="Aptos", size=10, bold=True, color="C62828")
    yf = PatternFill("solid", fgColor=YELLOW_BG)
    yn = Font(name="Aptos", size=10, bold=True, color="F57F17")

    for cl in ["N", "R"]:
        rng = f"{cl}4:{cl}{last_r}"
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=gf, font=gn))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="lessThan", formula=["0.5"], fill=rf, font=rn))
        ws.conditional_formatting.add(rng,
            CellIsRule(operator="between", formula=["0.5","0.999"], fill=yf, font=yn))

    # ── Footnote ────────────────────────────────────────────────────────
    nr = cur
    ws.merge_cells(f"B{nr}:R{nr}")
    nc = ws.cell(row=nr, column=2)
    nc.value = (
        "* Rakuten Blended ROAS = Rakuten Total Sales / (Rakuten + Meta + Influencer Ad Spend). All Meta WL & Influencer land on Rakuten. "
        "Amazon shows Direct ROAS only (Total Sales shown for reference, no Blended ROAS to avoid brand-effect distortion). "
        "TOTAL Blended ROAS = (Rakuten + Amazon Total Sales) / All Ad Spend."
    )
    nc.font = f_note
    nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[nr].height = 36

    ws.sheet_view.showGridLines = False
    wb.save(DST)
    print(f"[OK] Saved: {DST}")


if __name__ == "__main__":
    build()
