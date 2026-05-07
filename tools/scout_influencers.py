#!/usr/bin/env python3
"""
인플루언서 스카우터 — 경쟁사 팔로워 + 해시태그 기반 인플루언서 발굴

사용법:
  python tools/scout_influencers.py --competitors                    # 경쟁사 팔로워 스캔
  python tools/scout_influencers.py --hashtags "#ベビー用品レビュー"  # 해시태그 기반
  python tools/scout_influencers.py --filter                         # 기존 결과 필터링
  python tools/scout_influencers.py --export                         # 엑셀 내보내기

필터 조건 (기본값):
  --max-followers 10000    팔로워 상한
  --min-views 5000         릴스 평균 뷰수 하한
  --max-views 10000        릴스 평균 뷰수 상한
  --min-ratio 0.5          뷰/팔로워 비율 하한
"""

import argparse
import json
import os
import sys
import io
import time
import re
from datetime import datetime, timezone, timedelta
from pathlib import Path

# Windows cp949 인코딩 문제 방지
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import requests
except ImportError:
    print("requests 패키지 필요: pip install requests")
    sys.exit(1)

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

JST = timezone(timedelta(hours=9))
BASE_DIR = Path(__file__).resolve().parent.parent
SCOUT_DIR = BASE_DIR / ".tmp" / "influencer_scout"
RAG_DIR = BASE_DIR / ".tmp" / "influencer_rag" / "profiles"

# 경쟁사 IG 핸들
COMPETITORS = {
    "richell": "richell_baby_official",
    "combi": "combi_babylifedesign",
    "bbox": "bboxforkidsjapan",
    "pigeon": "pigeon_official_jp",
}

# 기본 해시태그
DEFAULT_HASHTAGS = [
    "ベビー用品レビュー",
    "ストローマグ",
    "離乳食グッズ",
    "赤ちゃんのいる生活",
    "育児グッズ",
    "ベビー用品購入品",
]

APIFY_BASE = "https://api.apify.com/v2"


def load_env():
    env_path = BASE_DIR / ".env"
    if load_dotenv:
        load_dotenv(env_path)
    else:
        # manual parse
        if env_path.exists():
            for line in env_path.read_text(encoding="utf-8").splitlines():
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ.setdefault(k.strip(), v.strip())


def get_token():
    load_env()
    token = os.environ.get("APIFY_API_TOKEN", "")
    if not token:
        print("ERROR: APIFY_API_TOKEN not found in .env")
        sys.exit(1)
    return token


def ensure_dirs():
    SCOUT_DIR.mkdir(parents=True, exist_ok=True)


def now_str():
    return datetime.now(JST).strftime("%Y%m%d_%H%M")


# ── Apify Actor 실행 ──────────────────────────────────────

def run_actor(actor_id: str, input_data: dict, token: str, timeout_secs: int = 300) -> dict:
    """Apify actor 실행 후 결과 반환"""
    url = f"{APIFY_BASE}/acts/{actor_id}/runs"
    params = {"token": token}
    headers = {"Content-Type": "application/json"}

    print(f"  Apify actor 시작: {actor_id}...")
    resp = requests.post(url, params=params, json=input_data, headers=headers, timeout=60)
    resp.raise_for_status()
    run_data = resp.json()["data"]
    run_id = run_data["id"]
    print(f"  Run ID: {run_id}")

    # 폴링
    poll_url = f"{APIFY_BASE}/actor-runs/{run_id}"
    start = time.time()
    while time.time() - start < timeout_secs:
        time.sleep(5)
        r = requests.get(poll_url, params={"token": token}, timeout=30)
        status = r.json()["data"]["status"]
        elapsed = int(time.time() - start)
        print(f"  [{elapsed}s] 상태: {status}")
        if status in ("SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"):
            break

    if status != "SUCCEEDED":
        print(f"  ERROR: Actor 실행 실패 ({status})")
        return {"items": []}

    # 결과 가져오기
    dataset_id = r.json()["data"]["defaultDatasetId"]
    items_url = f"{APIFY_BASE}/datasets/{dataset_id}/items"
    items_resp = requests.get(items_url, params={"token": token, "format": "json"}, timeout=60)
    items = items_resp.json()
    print(f"  결과: {len(items)}건")
    return {"items": items}


# ── Step 1: 경쟁사 팔로워 수집 ──────────────────────────────

def scrape_followers(competitors: list[str], token: str, limit: int = 200) -> list[dict]:
    """경쟁사 계정의 팔로워 목록 수집"""
    all_followers = []

    for comp_name in competitors:
        handle = COMPETITORS.get(comp_name, comp_name)
        print(f"\n📡 {comp_name} ({handle}) 팔로워 수집 중...")

        input_data = {
            "usernames": [handle],
            "resultsLimit": limit,
        }

        result = run_actor("monumental_world~instagram-followers-scraper---no-login", input_data, token, timeout_secs=180)
        followers = result.get("items", [])

        for f in followers:
            f["_source_competitor"] = comp_name
            f["_source_handle"] = handle

        all_followers.extend(followers)
        print(f"  {comp_name}: {len(followers)}명 수집")

    # 중복 제거 (username 기준)
    seen = set()
    unique = []
    for f in all_followers:
        username = f.get("username", "")
        if username and username not in seen:
            seen.add(username)
            unique.append(f)

    print(f"\n총 {len(unique)}명 (중복 제거 후)")
    return unique


# ── Step 2: 프로필 상세 수집 ──────────────────────────────

def scrape_profiles(usernames: list[str], token: str) -> list[dict]:
    """프로필 상세 정보 수집 (팔로워 수, 포스트 수 등)"""
    if not usernames:
        return []

    # 배치로 나눠서 처리 (한 번에 최대 50명)
    batch_size = 50
    all_profiles = []

    for i in range(0, len(usernames), batch_size):
        batch = usernames[i:i + batch_size]
        print(f"\n📊 프로필 수집 중... ({i + 1}~{i + len(batch)}/{len(usernames)})")

        input_data = {
            "usernames": batch,
        }

        result = run_actor("apify~instagram-profile-scraper", input_data, token, timeout_secs=300)
        profiles = result.get("items", [])
        all_profiles.extend(profiles)
        print(f"  {len(profiles)}명 프로필 수집")

        if i + batch_size < len(usernames):
            time.sleep(3)  # rate limit 대비

    return all_profiles


# ── Step 3: 릴스 뷰수 수집 ──────────────────────────────

def scrape_reels_views(usernames: list[str], token: str, reels_per_user: int = 6) -> dict:
    """릴스 뷰수 수집 → {username: avg_views}"""
    if not usernames:
        return {}

    views_map = {}
    batch_size = 10  # reel scraper는 배치 작게

    for i in range(0, len(usernames), batch_size):
        batch = usernames[i:i + batch_size]
        print(f"\n🎬 릴스 뷰수 수집 중... ({i + 1}~{i + len(batch)}/{len(usernames)})")

        input_data = {
            "username": batch,
            "resultsLimit": reels_per_user,
        }

        result = run_actor("apify~instagram-reel-scraper", input_data, token, timeout_secs=300)
        posts = result.get("items", [])

        # username별 뷰수 집계
        user_views = {}
        for post in posts:
            if post.get("error"):
                continue
            username = post.get("ownerUsername", "")
            video_views = post.get("videoPlayCount", 0) or post.get("videoViewCount", 0)
            if username and video_views:
                if username not in user_views:
                    user_views[username] = []
                user_views[username].append(video_views)

        for username, views_list in user_views.items():
            if views_list:
                views_map[username] = {
                    "avg_views": int(sum(views_list) / len(views_list)),
                    "max_views": max(views_list),
                    "min_views": min(views_list),
                    "count": len(views_list),
                }

        print(f"    뷰수 데이터: {len(user_views)}명")

        if i + batch_size < len(usernames):
            time.sleep(3)

    return views_map


# ── Step 4: 해시태그 기반 수집 ──────────────────────────────

def scrape_hashtag_posts(hashtags: list[str], token: str, posts_per_tag: int = 100) -> list[str]:
    """해시태그에서 포스트 → 계정 추출"""
    all_usernames = set()

    for tag in hashtags:
        clean_tag = tag.lstrip("#")
        print(f"\n#️⃣ #{clean_tag} 포스트 수집 중...")

        input_data = {
            "hashtags": [clean_tag],
            "resultsLimit": posts_per_tag,
        }

        result = run_actor("scrapesmith~instagram-hashtag-scraper", input_data, token, timeout_secs=180)
        posts = result.get("items", [])

        # error 항목 제외
        valid_posts = [p for p in posts if not p.get("error")]
        usernames = {p.get("ownerUsername", "") for p in valid_posts if p.get("ownerUsername")}
        all_usernames.update(usernames)
        print(f"  #{clean_tag}: {len(valid_posts)}건 포스트 → {len(usernames)}명 발견")

        time.sleep(2)

    print(f"\n총 {len(all_usernames)}명 (해시태그 기반)")
    return list(all_usernames)


# ── 필터링 ──────────────────────────────────────────────

def load_contacted_handles() -> set:
    """RAG에서 이미 연락한 인플루언서 핸들 로드"""
    handles = set()
    if RAG_DIR.exists():
        for f in RAG_DIR.glob("*.json"):
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                handle = data.get("handle", "").lstrip("@").lower()
                if handle:
                    handles.add(handle)
                # 이름도 추가 (핸들 없는 경우 대비)
                name = data.get("name", "").lower()
                if name:
                    handles.add(name)
            except Exception:
                pass
    return handles


def filter_candidates(
    profiles: list[dict],
    views_map: dict,
    max_followers: int = 10000,
    min_views: int = 5000,
    max_views: int = 10000,
    min_ratio: float = 0.5,
) -> list[dict]:
    """조건에 맞는 인플루언서 필터링"""
    contacted = load_contacted_handles()
    candidates = []

    for p in profiles:
        username = p.get("username", "").lower()
        followers = p.get("followersCount", 0) or p.get("followers", 0)
        is_private = p.get("isPrivate", False)
        is_business = p.get("isBusinessAccount", False) or p.get("isBusiness", False)

        # 기본 필터
        if is_private:
            continue
        if followers > max_followers:
            continue
        if followers < 500:  # 너무 작은 계정 제외
            continue

        # 이미 연락한 사람 제외
        if username in contacted:
            continue
        full_name = (p.get("fullName", "") or "").lower()
        if full_name and full_name in contacted:
            continue

        # 뷰수 필터
        views_info = views_map.get(username, {})
        avg_views = views_info.get("avg_views", 0)

        # 뷰수 데이터 없으면 일단 포함 (나중에 확인)
        if avg_views > 0:
            if avg_views < min_views or avg_views > max_views:
                continue
            ratio = avg_views / max(followers, 1)
            if ratio < min_ratio:
                continue
        else:
            ratio = 0

        candidates.append({
            "username": p.get("username", ""),
            "full_name": p.get("fullName", "") or "",
            "followers": followers,
            "following": p.get("followingCount", 0) or p.get("following", 0),
            "posts_count": p.get("postsCount", 0) or p.get("posts", 0),
            "avg_reels_views": avg_views,
            "max_reels_views": views_info.get("max_views", 0),
            "view_follower_ratio": round(ratio, 2),
            "bio": (p.get("biography", "") or p.get("bio", ""))[:100],
            "is_business": is_business,
            "profile_url": f"https://www.instagram.com/{p.get('username', '')}/",
            "source": p.get("_source_competitor", "hashtag"),
        })

    # 뷰/팔로워 비율 높은 순 정렬
    candidates.sort(key=lambda x: x["view_follower_ratio"], reverse=True)

    return candidates


# ── 엑셀 내보내기 ──────────────────────────────────────────

def export_excel(candidates: list[dict], filename: str = None):
    """결과를 엑셀로 내보내기"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl 필요: pip install openpyxl")
        return None

    if not filename:
        filename = f"influencer_scout_{now_str()}.xlsx"

    filepath = SCOUT_DIR / filename
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "인플루언서 후보"

    # 헤더
    headers = [
        "순위", "핸들", "이름", "팔로워", "평균 릴스 뷰",
        "최대 릴스 뷰", "뷰/팔 비율", "포스트 수", "비즈니스",
        "소스", "바이오", "프로필 URL"
    ]

    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 데이터
    for i, c in enumerate(candidates, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=f"@{c['username']}").border = thin_border
        ws.cell(row=row, column=3, value=c["full_name"]).border = thin_border
        ws.cell(row=row, column=4, value=c["followers"]).border = thin_border
        ws.cell(row=row, column=5, value=c["avg_reels_views"]).border = thin_border
        ws.cell(row=row, column=6, value=c["max_reels_views"]).border = thin_border
        ws.cell(row=row, column=7, value=c["view_follower_ratio"]).border = thin_border
        ws.cell(row=row, column=8, value=c["posts_count"]).border = thin_border
        ws.cell(row=row, column=9, value="Y" if c["is_business"] else "N").border = thin_border
        ws.cell(row=row, column=10, value=c["source"]).border = thin_border
        ws.cell(row=row, column=11, value=c["bio"]).border = thin_border
        ws.cell(row=row, column=12, value=c["profile_url"]).border = thin_border

        # 뷰/팔 비율 높은 행 하이라이트
        if c["view_follower_ratio"] >= 1.0:
            highlight = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            for col in range(1, 13):
                ws.cell(row=row, column=col).fill = highlight

    # 열 너비
    widths = [5, 20, 15, 10, 12, 12, 10, 8, 8, 10, 40, 35]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb.save(filepath)
    print(f"\n📊 엑셀 저장: {filepath}")
    return filepath


# ── CLI ──────────────────────────────────────────────────

def main():
    ensure_dirs()

    parser = argparse.ArgumentParser(description="인플루언서 스카우터")
    parser.add_argument("--competitors", nargs="*", default=None,
                        help="경쟁사 팔로워 스캔 (기본: richell combi bbox pigeon)")
    parser.add_argument("--hashtags", nargs="*", default=None,
                        help="해시태그 기반 검색")
    parser.add_argument("--filter", action="store_true",
                        help="기존 결과 다시 필터링")
    parser.add_argument("--export", action="store_true",
                        help="엑셀 내보내기")
    parser.add_argument("--max-followers", type=int, default=10000)
    parser.add_argument("--min-views", type=int, default=5000)
    parser.add_argument("--max-views", type=int, default=10000)
    parser.add_argument("--min-ratio", type=float, default=0.5)
    parser.add_argument("--follower-limit", type=int, default=200,
                        help="경쟁사당 팔로워 수집 수")
    parser.add_argument("--skip-views", action="store_true",
                        help="릴스 뷰수 수집 건너뛰기 (빠른 1차 스캔)")
    parser.add_argument("--cached", action="store_true",
                        help="캐시된 데이터 사용")
    parser.add_argument("--status", action="store_true",
                        help="현재 스카우팅 상태")

    args = parser.parse_args()
    token = get_token()

    cache_file = SCOUT_DIR / "scout_cache.json"

    if args.status:
        if cache_file.exists():
            data = json.loads(cache_file.read_text(encoding="utf-8"))
            print(f"\n📊 스카우팅 현황")
            print(f"  마지막 실행: {data.get('timestamp', '-')}")
            print(f"  수집된 프로필: {len(data.get('profiles', []))}명")
            print(f"  뷰수 데이터: {len(data.get('views_map', {}))}명")
            print(f"  최종 후보: {len(data.get('candidates', []))}명")
            contacted = load_contacted_handles()
            print(f"  제외 목록 (RAG): {len(contacted)}명")
        else:
            print("  아직 스카우팅 데이터 없음")
        return

    # ── 데이터 수집 또는 캐시 로드 ──
    if args.cached and cache_file.exists():
        print("📂 캐시 데이터 로드 중...")
        data = json.loads(cache_file.read_text(encoding="utf-8"))
        profiles = data.get("profiles", [])
        views_map = data.get("views_map", {})
    else:
        all_usernames = []
        profiles = []
        source_map = {}

        # 경쟁사 팔로워 모드
        if args.competitors is not None:
            comp_list = args.competitors if args.competitors else list(COMPETITORS.keys())
            followers = scrape_followers(comp_list, token, limit=args.follower_limit)

            # 공개 계정만 필터
            public_usernames = [
                f.get("username", "")
                for f in followers
                if f.get("username") and not f.get("is_private", False)
            ]
            all_usernames.extend(public_usernames)
            print(f"\n  공개 계정: {len(public_usernames)}명 / 전체 {len(followers)}명")

            # source 정보 보존
            source_map = {f.get("username", ""): f.get("_source_competitor", "") for f in followers}

        # 해시태그 모드
        if args.hashtags is not None:
            tags = args.hashtags if args.hashtags else DEFAULT_HASHTAGS
            ht_usernames = scrape_hashtag_posts(tags, token)
            all_usernames.extend(ht_usernames)
            for u in ht_usernames:
                source_map[u] = "hashtag"

        # 모든 공개 계정의 프로필 상세 수집
        unique_usernames = list(set(u for u in all_usernames if u))
        if unique_usernames:
            print(f"\n📊 프로필 상세 수집: {len(unique_usernames)}명")
            profiles = scrape_profiles(unique_usernames, token)
            # source 정보 추가
            for p in profiles:
                username = p.get("username", "")
                if username in source_map:
                    p["_source_competitor"] = source_map[username]

        # 뷰수 수집
        views_map = {}
        if not args.skip_views:
            # 팔로워 1만 이하만 뷰수 수집 (비용 절약)
            eligible = [
                p.get("username", "")
                for p in profiles
                if (p.get("followersCount", 0) or p.get("followers", 0)) <= args.max_followers
                and (p.get("followersCount", 0) or p.get("followers", 0)) >= 500
                and not p.get("isPrivate", False)
            ]
            if eligible:
                print(f"\n🎬 뷰수 수집 대상: {len(eligible)}명")
                views_map = scrape_reels_views(eligible, token)

        # 캐시 저장
        cache_data = {
            "timestamp": now_str(),
            "profiles": profiles,
            "views_map": views_map,
        }

    # ── 필터링 ──
    candidates = filter_candidates(
        profiles, views_map,
        max_followers=args.max_followers,
        min_views=args.min_views,
        max_views=args.max_views,
        min_ratio=args.min_ratio,
    )

    # 캐시에 후보 포함
    if not (args.cached and cache_file.exists()):
        cache_data["candidates"] = candidates
        cache_file.write_text(
            json.dumps(cache_data, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        print(f"\n💾 캐시 저장: {cache_file}")

    # ── 결과 출력 ──
    print(f"\n{'='*60}")
    print(f"🎯 인플루언서 후보: {len(candidates)}명")
    print(f"{'='*60}")

    for i, c in enumerate(candidates[:20], 1):
        ratio_str = f"{c['view_follower_ratio']:.1f}x" if c['view_follower_ratio'] > 0 else "?"
        print(
            f"  {i:2d}. @{c['username']:20s} | "
            f"팔로워 {c['followers']:>6,} | "
            f"평균뷰 {c['avg_reels_views']:>6,} | "
            f"비율 {ratio_str:>5s} | "
            f"{c['source']}"
        )

    if len(candidates) > 20:
        print(f"  ... 외 {len(candidates) - 20}명 (엑셀에서 확인)")

    # ── 엑셀 내보내기 ──
    if args.export or candidates:
        export_excel(candidates)

    return candidates


if __name__ == "__main__":
    main()
