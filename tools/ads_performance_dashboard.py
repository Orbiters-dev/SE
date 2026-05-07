"""
Ads Performance Dashboard Builder
===================================
Input:  ../Shared/datakeeper/latest/{meta_ads_daily, amazon_ads_daily, amazon_sales_daily}.json
        .tmp/rakuten_sales_weekly.json (optional)

Output: dashboard/Ads_Dashboard_YYYYMMDD_HHMM.xlsx

Tabs:
  1. Overview_Monthly  - Monthly totals (all channels) + RPP manual input (yellow)
  2. 채널별 공헌이익   - Monthly CM by channel (매출 - 광고비 - 수수료)
  3. META              - Weekly Meta ad performance
  4. AMAZON            - Weekly Amazon ads + sales
  5. RAKUTEN           - Weekly sales + RPP manual input (yellow)

Yellow cells = manual input required (Rakuten RPP/sales data).
공헌이익 = 매출 - 광고비 - 플랫폼 수수료
"""

import json
from datetime import datetime, timedelta
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR       = Path(__file__).parent.parent
DK_DIR         = BASE_DIR.parent / "Shared" / "datakeeper" / "latest"
TMP_DIR        = BASE_DIR / ".tmp"
DASH_DIR       = BASE_DIR / "dashboard"
TMP_DIR.mkdir(parents=True, exist_ok=True)
DASH_DIR.mkdir(parents=True, exist_ok=True)
RAKUTEN_TMP    = TMP_DIR / "rakuten_sales_weekly.json"
OUTFILE        = DASH_DIR / f"Ads_Dashboard_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
START_DATE     = "2025-11-01"   # 이 날짜 이후 데이터만 포함

# ── Colors ─────────────────────────────────────────────────────────────────────
C = {
    "meta":      "4267B2",   # Meta blue
    "amazon":    "E47911",   # Amazon orange
    "rakuten":   "BF0000",   # Rakuten red
    "overview":  "2D3561",   # Dark navy
    "white":     "FFFFFF",
    "alt":       "F0F4FF",   # Light blue-gray alternating row
    "input":     "FFF2CC",   # Yellow = manual input
    "input_txt": "7B3800",   # Brown text for yellow headers
    "total":     "D6ECD2",   # Light green = total row
}

# ── Style helpers ──────────────────────────────────────────────────────────────
def _fill(hex_):
    return PatternFill("solid", fgColor=hex_)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, r, c, val, bg=None, txt=C["white"], size=9):
    bg = bg or C["overview"]
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(bold=True, size=size, color=txt)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _border()
    return cell

def _dat(ws, r, c, val, bg=C["white"], bold=False, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(bold=bold, size=9)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(
        horizontal="right" if isinstance(val, (int, float)) else "center",
        vertical="center"
    )
    cell.border = _border()
    if fmt:
        cell.number_format = fmt
    return cell

def _inp(ws, r, c):
    """Yellow manual input cell."""
    cell = ws.cell(row=r, column=c, value=None)
    cell.fill      = _fill(C["input"])
    cell.border    = _border()
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.font      = Font(size=9)
    return cell

def _formula(ws, r, c, formula, fmt=None):
    cell = ws.cell(row=r, column=c, value=formula)
    cell.fill      = _fill(C["alt"])
    cell.border    = _border()
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.font      = Font(size=9, italic=True)
    if fmt:
        cell.number_format = fmt
    return cell

def _col_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _title(ws, merge_range, val, bg, txt=C["white"], size=12, height=28):
    ws.merge_cells(merge_range)
    ref = merge_range.split(":")[0]
    cell = ws[ref]
    cell.value     = val
    cell.font      = Font(bold=True, size=size, color=txt)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    r = int("".join(filter(str.isdigit, ref)))
    ws.row_dimensions[r].height = height
    return cell

# ── Data helpers ───────────────────────────────────────────────────────────────
def _load(path):
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"  [WARN] {path.name}: {e}")
        return []

def _filter(rows):
    """START_DATE 이후 데이터만 반환."""
    return [r for r in rows if r.get("date", "") >= START_DATE]

def _week(date_str):
    """Returns (week_start: date, label: str)."""
    d = datetime.strptime(date_str, "%Y-%m-%d").date()
    start = d - timedelta(days=d.weekday())
    end   = start + timedelta(days=6)
    return start, f"{start.month}/{start.day}~{end.month}/{end.day}"

def _month(date_str):
    d = datetime.strptime(date_str, "%Y-%m-%d")
    return f"{d.year}/{d.month:02d}"

# ── Aggregation ────────────────────────────────────────────────────────────────
def _agg_meta_weekly(rows):
    buckets, labels = defaultdict(lambda: dict(
        spend=0.0, impressions=0, clicks=0, reach=0, purchases=0, purchase_value=0.0
    )), {}
    for r in rows:
        s, lbl = _week(r["date"])
        labels[s] = lbl
        b = buckets[s]
        b["spend"]          += r.get("spend", 0) or 0
        b["impressions"]    += r.get("impressions", 0) or 0
        b["clicks"]         += r.get("clicks", 0) or 0
        b["reach"]          += r.get("reach", 0) or 0
        b["purchases"]      += r.get("purchases", 0) or 0
        b["purchase_value"] += r.get("purchase_value", 0) or 0
    out = []
    for s in sorted(buckets):
        m = buckets[s]
        m["ctr"]  = m["clicks"] / m["impressions"] * 100 if m["impressions"] else 0
        m["roas"] = m["purchase_value"] / m["spend"] if m["spend"] else 0
        out.append((s, labels[s], m))
    return out

def _agg_amz_ads_weekly(rows):
    buckets, labels = defaultdict(lambda: dict(
        spend=0.0, impressions=0, clicks=0, purchases=0, sales=0.0
    )), {}
    for r in rows:
        s, lbl = _week(r["date"])
        labels[s] = lbl
        b = buckets[s]
        b["spend"]       += r.get("spend", 0) or 0
        b["impressions"] += r.get("impressions", 0) or 0
        b["clicks"]      += r.get("clicks", 0) or 0
        b["purchases"]   += r.get("purchases", 0) or 0
        b["sales"]       += r.get("sales", 0) or 0
    out = []
    for s in sorted(buckets):
        m = buckets[s]
        m["acos"] = m["spend"] / m["sales"] * 100 if m["sales"] else 0
        m["roas"] = m["sales"] / m["spend"] if m["spend"] else 0
        out.append((s, labels[s], m))
    return out

def _agg_amz_sales_weekly(rows):
    buckets, labels = defaultdict(lambda: dict(
        gross_sales=0.0, net_sales=0.0, orders=0, units=0, fees=0.0
    )), {}
    for r in rows:
        s, lbl = _week(r["date"])
        labels[s] = lbl
        b = buckets[s]
        b["gross_sales"] += r.get("gross_sales", 0) or 0
        b["net_sales"]   += r.get("net_sales", 0) or 0
        b["orders"]      += r.get("orders", 0) or 0
        b["units"]       += r.get("units", 0) or 0
        b["fees"]        += r.get("fees", 0) or 0
    return [(s, labels[s], buckets[s]) for s in sorted(buckets)]

def _agg_meta_monthly(rows):
    buckets = defaultdict(lambda: dict(
        spend=0.0, impressions=0, clicks=0, purchases=0, purchase_value=0.0
    ))
    for r in rows:
        b = buckets[_month(r["date"])]
        b["spend"]          += r.get("spend", 0) or 0
        b["impressions"]    += r.get("impressions", 0) or 0
        b["clicks"]         += r.get("clicks", 0) or 0
        b["purchases"]      += r.get("purchases", 0) or 0
        b["purchase_value"] += r.get("purchase_value", 0) or 0
    out = []
    for mk in sorted(buckets):
        m = buckets[mk]
        m["roas"] = m["purchase_value"] / m["spend"] if m["spend"] else 0
        out.append((mk, m))
    return out

def _agg_amz_ads_monthly(rows):
    buckets = defaultdict(lambda: dict(spend=0.0, purchases=0, sales=0.0))
    for r in rows:
        b = buckets[_month(r["date"])]
        b["spend"]    += r.get("spend", 0) or 0
        b["purchases"]+= r.get("purchases", 0) or 0
        b["sales"]    += r.get("sales", 0) or 0
    out = []
    for mk in sorted(buckets):
        m = buckets[mk]
        m["acos"] = m["spend"] / m["sales"] * 100 if m["sales"] else 0
        m["roas"] = m["sales"] / m["spend"] if m["spend"] else 0
        out.append((mk, m))
    return out

def _agg_amz_sales_monthly(rows):
    buckets = defaultdict(lambda: dict(gross_sales=0.0, net_sales=0.0, orders=0, units=0))
    for r in rows:
        b = buckets[_month(r["date"])]
        b["gross_sales"] += r.get("gross_sales", 0) or 0
        b["net_sales"]   += r.get("net_sales", 0) or 0
        b["orders"]      += r.get("orders", 0) or 0
        b["units"]       += r.get("units", 0) or 0
    return [(mk, buckets[mk]) for mk in sorted(buckets)]

# ── Tab: META ─────────────────────────────────────────────────────────────────
def build_meta(ws, weekly):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"
    ws.sheet_properties.tabColor = C["meta"]

    _title(ws, "A1:I1", "META 광고 성과 (주별)", C["meta"], height=30)
    ws.row_dimensions[2].height = 34

    hdrs = ["주차", "광고비 ($)", "노출", "클릭", "CTR (%)", "도달", "구매수", "구매금액 ($)", "ROAS"]
    for i, h in enumerate(hdrs, 1):
        _hdr(ws, 2, i, h, C["meta"])

    total = dict(spend=0.0, impressions=0, clicks=0, reach=0, purchases=0, purchase_value=0.0)
    for idx, (_, lbl, m) in enumerate(weekly):
        r  = 3 + idx
        bg = C["white"] if idx % 2 == 0 else C["alt"]
        ws.row_dimensions[r].height = 18
        _dat(ws, r, 1, lbl,                              bg=bg)
        _dat(ws, r, 2, round(m["spend"], 2),             bg=bg, fmt='#,##0.00')
        _dat(ws, r, 3, m["impressions"],                 bg=bg, fmt='#,##0')
        _dat(ws, r, 4, m["clicks"],                      bg=bg, fmt='#,##0')
        _dat(ws, r, 5, round(m["ctr"], 2),               bg=bg, fmt='0.00"%"')
        _dat(ws, r, 6, m["reach"],                       bg=bg, fmt='#,##0')
        _dat(ws, r, 7, m["purchases"],                   bg=bg, fmt='#,##0')
        _dat(ws, r, 8, round(m["purchase_value"], 2),    bg=bg, fmt='#,##0.00')
        _dat(ws, r, 9, round(m["roas"], 2),              bg=bg, fmt='0.00"x"')
        for k in ("spend", "impressions", "clicks", "reach", "purchases", "purchase_value"):
            total[k] += m[k]

    tr = 3 + len(weekly)
    ws.row_dimensions[tr].height = 20
    roas_t = total["purchase_value"] / total["spend"] if total["spend"] else 0
    ctr_t  = total["clicks"] / total["impressions"] * 100 if total["impressions"] else 0
    _dat(ws, tr, 1, "TOTAL",                                   bg=C["total"], bold=True)
    _dat(ws, tr, 2, round(total["spend"], 2),                  bg=C["total"], bold=True, fmt='#,##0.00')
    _dat(ws, tr, 3, total["impressions"],                      bg=C["total"], bold=True, fmt='#,##0')
    _dat(ws, tr, 4, total["clicks"],                           bg=C["total"], bold=True, fmt='#,##0')
    _dat(ws, tr, 5, round(ctr_t, 2),                           bg=C["total"], bold=True, fmt='0.00"%"')
    _dat(ws, tr, 6, total["reach"],                            bg=C["total"], bold=True, fmt='#,##0')
    _dat(ws, tr, 7, total["purchases"],                        bg=C["total"], bold=True, fmt='#,##0')
    _dat(ws, tr, 8, round(total["purchase_value"], 2),         bg=C["total"], bold=True, fmt='#,##0.00')
    _dat(ws, tr, 9, round(roas_t, 2),                          bg=C["total"], bold=True, fmt='0.00"x"')

    _col_w(ws, [15, 13, 12, 10, 10, 12, 10, 14, 10])

# ── Tab: AMAZON ───────────────────────────────────────────────────────────────
def build_amazon(ws, ads_weekly, sales_weekly):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"
    ws.sheet_properties.tabColor = C["amazon"]

    # ── Section 1: 광고 성과 ──
    _title(ws, "A1:H1", "AMAZON 광고 성과 (주별)", C["amazon"], height=28)
    ws.row_dimensions[2].height = 34

    for i, h in enumerate(["주차", "광고비 ($)", "노출", "클릭", "구매수", "귀속매출 ($)", "ACOS (%)", "ROAS"], 1):
        _hdr(ws, 2, i, h, C["amazon"])

    ads_total = dict(spend=0.0, impressions=0, clicks=0, purchases=0, sales=0.0)
    for idx, (_, lbl, m) in enumerate(ads_weekly):
        r  = 3 + idx
        bg = C["white"] if idx % 2 == 0 else C["alt"]
        ws.row_dimensions[r].height = 18
        _dat(ws, r, 1, lbl,                  bg=bg)
        _dat(ws, r, 2, round(m["spend"], 2), bg=bg, fmt='#,##0.00')
        _dat(ws, r, 3, m["impressions"],     bg=bg, fmt='#,##0')
        _dat(ws, r, 4, m["clicks"],          bg=bg, fmt='#,##0')
        _dat(ws, r, 5, m["purchases"],       bg=bg, fmt='#,##0')
        _dat(ws, r, 6, round(m["sales"], 2), bg=bg, fmt='#,##0.00')
        _dat(ws, r, 7, round(m["acos"], 1),  bg=bg, fmt='0.0"%"')
        _dat(ws, r, 8, round(m["roas"], 2),  bg=bg, fmt='0.00"x"')
        for k in ("spend", "impressions", "clicks", "purchases", "sales"):
            ads_total[k] += m[k]

    atr = 3 + len(ads_weekly)
    ws.row_dimensions[atr].height = 20
    acos_t = ads_total["spend"] / ads_total["sales"] * 100 if ads_total["sales"] else 0
    roas_t = ads_total["sales"] / ads_total["spend"] if ads_total["spend"] else 0
    _dat(ws, atr, 1, "TOTAL",                              bg=C["total"], bold=True)
    _dat(ws, atr, 2, round(ads_total["spend"], 2),         bg=C["total"], bold=True, fmt='#,##0.00')
    _dat(ws, atr, 3, ads_total["impressions"],             bg=C["total"], bold=True, fmt='#,##0')
    _dat(ws, atr, 4, ads_total["clicks"],                  bg=C["total"], bold=True, fmt='#,##0')
    _dat(ws, atr, 5, ads_total["purchases"],               bg=C["total"], bold=True, fmt='#,##0')
    _dat(ws, atr, 6, round(ads_total["sales"], 2),         bg=C["total"], bold=True, fmt='#,##0.00')
    _dat(ws, atr, 7, round(acos_t, 1),                    bg=C["total"], bold=True, fmt='0.0"%"')
    _dat(ws, atr, 8, round(roas_t, 2),                    bg=C["total"], bold=True, fmt='0.00"x"')

    # ── Gap ──
    gap = atr + 1
    ws.row_dimensions[gap].height = 12

    # ── Section 2: 판매 성과 ──
    s2 = gap + 1
    _title(ws, f"A{s2}:F{s2}", "AMAZON 판매 성과 (주별)", C["amazon"], size=11, height=26)
    ws.row_dimensions[s2 + 1].height = 34

    for i, h in enumerate(["주차", "총매출 ($)", "순매출 ($)", "주문수", "판매수량", "수수료 ($)"], 1):
        _hdr(ws, s2 + 1, i, h, C["amazon"])

    sales_total = dict(gross_sales=0.0, net_sales=0.0, orders=0, units=0, fees=0.0)
    for idx, (_, lbl, m) in enumerate(sales_weekly):
        r  = s2 + 2 + idx
        bg = C["white"] if idx % 2 == 0 else C["alt"]
        ws.row_dimensions[r].height = 18
        _dat(ws, r, 1, lbl,                            bg=bg)
        _dat(ws, r, 2, round(m["gross_sales"], 2),     bg=bg, fmt='#,##0.00')
        _dat(ws, r, 3, round(m["net_sales"], 2),       bg=bg, fmt='#,##0.00')
        _dat(ws, r, 4, m["orders"],                    bg=bg, fmt='#,##0')
        _dat(ws, r, 5, m["units"],                     bg=bg, fmt='#,##0')
        _dat(ws, r, 6, round(m["fees"], 2),            bg=bg, fmt='#,##0.00')
        for k in ("gross_sales", "net_sales", "orders", "units", "fees"):
            sales_total[k] += m[k]

    str_ = s2 + 2 + len(sales_weekly)
    ws.row_dimensions[str_].height = 20
    _dat(ws, str_, 1, "TOTAL",                                  bg=C["total"], bold=True)
    _dat(ws, str_, 2, round(sales_total["gross_sales"], 2),     bg=C["total"], bold=True, fmt='#,##0.00')
    _dat(ws, str_, 3, round(sales_total["net_sales"], 2),       bg=C["total"], bold=True, fmt='#,##0.00')
    _dat(ws, str_, 4, sales_total["orders"],                    bg=C["total"], bold=True, fmt='#,##0')
    _dat(ws, str_, 5, sales_total["units"],                     bg=C["total"], bold=True, fmt='#,##0')
    _dat(ws, str_, 6, round(sales_total["fees"], 2),            bg=C["total"], bold=True, fmt='#,##0.00')

    _col_w(ws, [15, 13, 13, 12, 10, 12, 12, 10])

# ── Tab: RAKUTEN ──────────────────────────────────────────────────────────────
def build_rakuten(ws, week_rows):
    """
    week_rows: list of (label: str, auto_data: dict or None)
    Columns 1-4: auto data (판매 성과)
    Columns 5-8: yellow manual input (RPP)
    Column 9: ROAS formula
    """
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"
    ws.sheet_properties.tabColor = C["rakuten"]

    _title(ws, "A1:I1", "RAKUTEN 성과 (주별)", C["rakuten"], height=30)
    ws.row_dimensions[2].height = 38

    # Auto headers (1-4)
    for i, h in enumerate(["주차", "총매출", "주문수", "판매수량"], 1):
        _hdr(ws, 2, i, h, C["rakuten"])

    # Manual input headers (5-8) - yellow
    for i, h in enumerate(["RPP 광고비", "RPP 노출", "RPP 클릭", "RPP 매출"], 5):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font      = Font(bold=True, size=9, color=C["input_txt"])
        cell.fill      = _fill("FFD966")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()

    # ROAS header (9)
    _hdr(ws, 2, 9, "RPP ROAS", C["rakuten"])

    # Note row (inserted as merged note)
    note_r = 3
    ws.merge_cells(f"E{note_r}:I{note_r}")
    note_cell = ws.cell(row=note_r, column=5,
                        value="★ 노란색 셀에 라쿠텐 RPP 데이터를 직접 입력하세요")
    note_cell.font      = Font(italic=True, size=8, color=C["input_txt"])
    note_cell.fill      = _fill("FFFACD")
    note_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[note_r].height = 16

    data_start = note_r + 1
    for idx, (lbl, auto) in enumerate(week_rows):
        r  = data_start + idx
        bg = C["white"] if idx % 2 == 0 else C["alt"]
        ws.row_dimensions[r].height = 20

        _dat(ws, r, 1, lbl, bg=bg)
        if auto:
            _dat(ws, r, 2, round(auto.get("gross_sales", 0), 0), bg=bg, fmt='#,##0')
            _dat(ws, r, 3, auto.get("orders", 0),                bg=bg, fmt='#,##0')
            _dat(ws, r, 4, auto.get("units", 0),                 bg=bg, fmt='#,##0')
        else:
            for c in (2, 3, 4):
                _dat(ws, r, c, None, bg=bg)

        for c in (5, 6, 7, 8):
            _inp(ws, r, c)

        # ROAS = RPP매출(H) / RPP광고비(E)
        E = get_column_letter(5)
        H = get_column_letter(8)
        _formula(ws, r, 9, f'=IF({E}{r}=0,"",{H}{r}/{E}{r})', fmt='0.00"x"')

    _col_w(ws, [15, 13, 10, 10, 13, 11, 10, 13, 11])

# ── Tab: 채널별 공헌이익 ──────────────────────────────────────────────────────
def build_cm(ws, meta_m, amz_ads_m, amz_sales_m):
    """
    공헌이익 = 매출 - 광고비 - 플랫폼 수수료

    Columns:
      A     : 월
      B-E   : META   (구매금액, 광고비, 공헌이익, CM%)       [auto]
      F-J   : AMAZON (총매출, 광고비, 수수료, 공헌이익, CM%) [auto]
      K-O   : RAKUTEN(매출, RPP광고비, 수수료, 공헌이익, CM%)[manual K~M, formula N~O]
      P-S   : TOTAL  (전체매출, 전체광고비, 공헌이익, CM%)   [formula]
    """
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B4"
    ws.sheet_properties.tabColor = "1D7044"  # Green

    _title(ws, "A1:S1", "채널별 공헌이익 (월별)", "1D7044", height=32)

    # Row 2: section labels
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 36

    ws.merge_cells("A2:A3")
    a2 = ws["A2"]
    a2.value     = "월"
    a2.font      = Font(bold=True, size=9, color=C["white"])
    a2.fill      = _fill("1D7044")
    a2.alignment = Alignment(horizontal="center", vertical="center")
    a2.border    = _border()

    def _sec(cs, ce, label, bg, txt=C["white"]):
        ws.merge_cells(f"{get_column_letter(cs)}2:{get_column_letter(ce)}2")
        cell = ws.cell(row=2, column=cs, value=label)
        cell.font      = Font(bold=True, size=9, color=txt)
        cell.fill      = _fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _border()

    _sec(2,  5,  "META",              C["meta"])
    _sec(6,  10, "AMAZON",            C["amazon"])
    _sec(11, 15, "RAKUTEN",           C["rakuten"])
    _sec(16, 19, "TOTAL",             "1D7044")

    # Row 3: column headers
    col_defs = [
        # META (2-5)
        (2,  C["meta"],    C["white"],     "구매금액 ($)"),
        (3,  C["meta"],    C["white"],     "광고비 ($)"),
        (4,  C["meta"],    C["white"],     "공헌이익 ($)"),
        (5,  C["meta"],    C["white"],     "CM%"),
        # AMAZON (6-10)
        (6,  C["amazon"],  C["white"],     "총매출 ($)"),
        (7,  C["amazon"],  C["white"],     "광고비 ($)"),
        (8,  C["amazon"],  C["white"],     "수수료 ($)"),
        (9,  C["amazon"],  C["white"],     "공헌이익 ($)"),
        (10, C["amazon"],  C["white"],     "CM%"),
        # RAKUTEN (11-15)
        (11, "FFD966",     C["input_txt"], "매출 (수동)"),
        (12, "FFD966",     C["input_txt"], "RPP광고비 (수동)"),
        (13, "FFD966",     C["input_txt"], "수수료 (수동)"),
        (14, C["rakuten"], C["white"],     "공헌이익"),
        (15, C["rakuten"], C["white"],     "CM%"),
        # TOTAL (16-19)
        (16, "1D7044",     C["white"],     "전체 매출"),
        (17, "1D7044",     C["white"],     "전체 광고비"),
        (18, "1D7044",     C["white"],     "공헌이익"),
        (19, "1D7044",     C["white"],     "CM%"),
    ]
    for col, bg, txt, label in col_defs:
        cell = ws.cell(row=3, column=col, value=label)
        cell.font      = Font(bold=True, size=8, color=txt)
        cell.fill      = _fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()

    # Data
    all_months = sorted(set(
        [m for m, _ in meta_m] +
        [m for m, _ in amz_ads_m] +
        [m for m, _ in amz_sales_m]
    ))
    meta_d     = {m: d for m, d in meta_m}
    amz_ads_d  = {m: d for m, d in amz_ads_m}
    amz_sales_d= {m: d for m, d in amz_sales_m}

    for idx, month in enumerate(all_months):
        r  = 4 + idx
        bg = C["white"] if idx % 2 == 0 else C["alt"]
        ws.row_dimensions[r].height = 20

        _dat(ws, r, 1, month, bg=bg)

        # META
        md   = meta_d.get(month, {})
        rev  = round(md.get("purchase_value", 0), 2)
        spend= round(md.get("spend", 0), 2)
        _dat(ws, r, 2, rev,   bg=bg, fmt='#,##0.00')
        _dat(ws, r, 3, spend, bg=bg, fmt='#,##0.00')
        B = get_column_letter(2); CC = get_column_letter(3)
        _formula(ws, r, 4, f"={B}{r}-{CC}{r}", fmt='#,##0.00')
        _formula(ws, r, 5, f"=IF({B}{r}=0,\"\",({B}{r}-{CC}{r})/{B}{r})", fmt='0.0%')

        # AMAZON
        ad  = amz_ads_d.get(month, {})
        sd  = amz_sales_d.get(month, {})
        g_sales = round(sd.get("gross_sales", 0), 2)
        ad_spend= round(ad.get("spend", 0), 2)
        fees    = round(sd.get("fees", 0) if sd else 0, 2)
        _dat(ws, r, 6, g_sales,  bg=bg, fmt='#,##0.00')
        _dat(ws, r, 7, ad_spend, bg=bg, fmt='#,##0.00')
        _dat(ws, r, 8, fees,     bg=bg, fmt='#,##0.00')
        F = get_column_letter(6); G = get_column_letter(7); H = get_column_letter(8)
        _formula(ws, r, 9,  f"={F}{r}-{G}{r}-{H}{r}", fmt='#,##0.00')
        _formula(ws, r, 10, f"=IF({F}{r}=0,\"\",({F}{r}-{G}{r}-{H}{r})/{F}{r})", fmt='0.0%')

        # RAKUTEN — manual input
        for c in (11, 12, 13):
            _inp(ws, r, c)
        K = get_column_letter(11); L = get_column_letter(12); M = get_column_letter(13)
        _formula(ws, r, 14, f"=IF({K}{r}=0,\"\",{K}{r}-{L}{r}-{M}{r})", fmt='#,##0.00')
        _formula(ws, r, 15, f"=IF({K}{r}=0,\"\",({K}{r}-{L}{r}-{M}{r})/{K}{r})", fmt='0.0%')

        # TOTAL
        D_ = get_column_letter(4); I_ = get_column_letter(9); N_ = get_column_letter(14)
        _formula(ws, r, 16, f"={B}{r}+{F}{r}+IF({K}{r}=\"\",0,{K}{r})", fmt='#,##0.00')
        _formula(ws, r, 17, f"={CC}{r}+{G}{r}+IF({L}{r}=\"\",0,{L}{r})", fmt='#,##0.00')
        P = get_column_letter(16); Q = get_column_letter(17)
        _formula(ws, r, 18, f"={D_}{r}+{I_}{r}+IF({N_}{r}=\"\",0,{N_}{r})", fmt='#,##0.00')
        R_ = get_column_letter(18)
        _formula(ws, r, 19, f"=IF({P}{r}=0,\"\",{R_}{r}/{P}{r})", fmt='0.0%')

    # Total row
    tr = 4 + len(all_months)
    ws.row_dimensions[tr].height = 22
    _dat(ws, tr, 1, "TOTAL", bg=C["total"], bold=True)
    sum_cols = {2: '#,##0.00', 3: '#,##0.00', 6: '#,##0.00', 7: '#,##0.00', 8: '#,##0.00',
                16: '#,##0.00', 17: '#,##0.00'}
    for col, fmt in sum_cols.items():
        cl = get_column_letter(col)
        cell = ws.cell(row=tr, column=col, value=f"=SUM({cl}4:{cl}{tr-1})")
        cell.font = Font(bold=True, size=9); cell.fill = _fill(C["total"])
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _border(); cell.number_format = fmt

    _col_w(ws, [10, 13, 12, 13, 9, 13, 12, 12, 13, 9, 12, 13, 12, 13, 9, 13, 13, 13, 9])


# ── Tab: Overview_Monthly ─────────────────────────────────────────────────────
def build_overview(ws, meta_m, amz_ads_m, amz_sales_m):
    """
    Columns:
      A     : 월
      B-F   : META (광고비, 노출, 클릭, 구매금액, ROAS)         [5]
      G-J   : AMAZON 광고 (광고비, 귀속매출, ACOS, ROAS)        [4]
      K-N   : AMAZON 판매 (총매출, 순매출, 주문수, 판매수량)    [4]
      O-P   : RAKUTEN 판매 (총매출, 주문수)                     [2]
      Q-T   : RPP 수동입력 (광고비, 노출, 클릭, 매출)           [4, yellow]
      U     : RPP ROAS (formula)                                [1]
    """
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B4"
    ws.sheet_properties.tabColor = C["overview"]

    # Row 1: main title
    _title(ws, "A1:U1", "월별 채널 통합 성과 개요", C["overview"], size=13, height=34)

    # Row 2: section labels (merged)
    ws.row_dimensions[2].height = 22

    def _sec(col_s, col_e, label, bg, txt=C["white"]):
        s_ltr = get_column_letter(col_s)
        e_ltr = get_column_letter(col_e)
        ws.merge_cells(f"{s_ltr}2:{e_ltr}2")
        c = ws.cell(row=2, column=col_s, value=label)
        c.font      = Font(bold=True, size=9, color=txt)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border()

    # A2:A3 merged for "월"
    ws.merge_cells("A2:A3")
    m_cell = ws["A2"]
    m_cell.value     = "월"
    m_cell.font      = Font(bold=True, size=9, color=C["white"])
    m_cell.fill      = _fill(C["overview"])
    m_cell.alignment = Alignment(horizontal="center", vertical="center")
    m_cell.border    = _border()

    _sec(2,  6,  "META",              C["meta"])
    _sec(7,  10, "AMAZON 광고",       C["amazon"])
    _sec(11, 14, "AMAZON 판매",       C["amazon"])
    _sec(15, 16, "RAKUTEN 판매",      C["rakuten"])
    _sec(17, 20, "RAKUTEN RPP (수동입력)", "FFD966", C["input_txt"])
    _sec(21, 21, "RPP",               C["rakuten"])

    # Row 3: column headers
    ws.row_dimensions[3].height = 38

    col_hdrs = [
        # META (2-6)
        (2,  C["meta"],    C["white"],     "광고비 ($)"),
        (3,  C["meta"],    C["white"],     "노출"),
        (4,  C["meta"],    C["white"],     "클릭"),
        (5,  C["meta"],    C["white"],     "구매금액 ($)"),
        (6,  C["meta"],    C["white"],     "ROAS"),
        # AMAZON ADS (7-10)
        (7,  C["amazon"],  C["white"],     "광고비 ($)"),
        (8,  C["amazon"],  C["white"],     "귀속매출 ($)"),
        (9,  C["amazon"],  C["white"],     "ACOS (%)"),
        (10, C["amazon"],  C["white"],     "ROAS"),
        # AMAZON SALES (11-14)
        (11, C["amazon"],  C["white"],     "총매출 ($)"),
        (12, C["amazon"],  C["white"],     "순매출 ($)"),
        (13, C["amazon"],  C["white"],     "주문수"),
        (14, C["amazon"],  C["white"],     "판매수량"),
        # RAKUTEN (15-16)
        (15, C["rakuten"], C["white"],     "총매출"),
        (16, C["rakuten"], C["white"],     "주문수"),
        # RPP manual (17-20)
        (17, "FFD966",     C["input_txt"], "RPP 광고비"),
        (18, "FFD966",     C["input_txt"], "RPP 노출"),
        (19, "FFD966",     C["input_txt"], "RPP 클릭"),
        (20, "FFD966",     C["input_txt"], "RPP 매출"),
        # RPP ROAS (21)
        (21, C["rakuten"], C["white"],     "RPP ROAS"),
    ]
    for col, bg, txt, label in col_hdrs:
        cell = ws.cell(row=3, column=col, value=label)
        cell.font      = Font(bold=True, size=8, color=txt)
        cell.fill      = _fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()

    # ── Data rows ──
    all_months = sorted(set(
        [m for m, _ in meta_m] +
        [m for m, _ in amz_ads_m] +
        [m for m, _ in amz_sales_m]
    ))
    meta_d     = {m: d for m, d in meta_m}
    amz_ads_d  = {m: d for m, d in amz_ads_m}
    amz_sales_d= {m: d for m, d in amz_sales_m}

    for idx, month in enumerate(all_months):
        r  = 4 + idx
        bg = C["white"] if idx % 2 == 0 else C["alt"]
        ws.row_dimensions[r].height = 20

        _dat(ws, r, 1, month, bg=bg)

        md = meta_d.get(month, {})
        _dat(ws, r, 2,  round(md.get("spend", 0), 2),         bg=bg, fmt='#,##0.00')
        _dat(ws, r, 3,  md.get("impressions", 0),              bg=bg, fmt='#,##0')
        _dat(ws, r, 4,  md.get("clicks", 0),                   bg=bg, fmt='#,##0')
        _dat(ws, r, 5,  round(md.get("purchase_value", 0), 2), bg=bg, fmt='#,##0.00')
        _dat(ws, r, 6,  round(md.get("roas", 0), 2),           bg=bg, fmt='0.00"x"')

        ad = amz_ads_d.get(month, {})
        _dat(ws, r, 7,  round(ad.get("spend", 0), 2),          bg=bg, fmt='#,##0.00')
        _dat(ws, r, 8,  round(ad.get("sales", 0), 2),          bg=bg, fmt='#,##0.00')
        _dat(ws, r, 9,  round(ad.get("acos", 0), 1),           bg=bg, fmt='0.0"%"')
        _dat(ws, r, 10, round(ad.get("roas", 0), 2),           bg=bg, fmt='0.00"x"')

        sd = amz_sales_d.get(month, {})
        _dat(ws, r, 11, round(sd.get("gross_sales", 0), 2),    bg=bg, fmt='#,##0.00')
        _dat(ws, r, 12, round(sd.get("net_sales", 0), 2),      bg=bg, fmt='#,##0.00')
        _dat(ws, r, 13, sd.get("orders", 0),                   bg=bg, fmt='#,##0')
        _dat(ws, r, 14, sd.get("units", 0),                    bg=bg, fmt='#,##0')

        # Rakuten: auto data not available from Data Keeper
        _dat(ws, r, 15, None, bg=bg)
        _dat(ws, r, 16, None, bg=bg)

        # RPP manual input
        for c in (17, 18, 19, 20):
            _inp(ws, r, c)

        Q = get_column_letter(17)  # RPP광고비
        T = get_column_letter(20)  # RPP매출
        _formula(ws, r, 21, f'=IF({Q}{r}=0,"",{T}{r}/{Q}{r})', fmt='0.00"x"')

    # ── Total row ──
    tr = 4 + len(all_months)
    ws.row_dimensions[tr].height = 22
    _dat(ws, tr, 1, "TOTAL", bg=C["total"], bold=True)

    sum_cols = {
        2: '#,##0.00', 3: '#,##0', 4: '#,##0', 5: '#,##0.00',
        7: '#,##0.00', 8: '#,##0.00',
        11: '#,##0.00', 12: '#,##0.00', 13: '#,##0', 14: '#,##0',
    }
    for col, fmt in sum_cols.items():
        cl = get_column_letter(col)
        cell = ws.cell(row=tr, column=col,
                       value=f"=SUM({cl}4:{cl}{tr - 1})")
        cell.font      = Font(bold=True, size=9)
        cell.fill      = _fill(C["total"])
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border    = _border()
        cell.number_format = fmt

    _col_w(ws, [10, 12, 11, 10, 13, 9, 12, 13, 9, 9, 13, 12, 9, 9, 11, 9, 12, 10, 9, 12, 10])

# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print("=== Ads Performance Dashboard Builder ===\n")

    # ── Load ──
    print(f"[1/4] Loading Data Keeper (기간: {START_DATE} 이후)...")
    meta_rows      = _filter(_load(DK_DIR / "meta_ads_daily.json"))
    amz_ads_rows   = _filter(_load(DK_DIR / "amazon_ads_daily.json"))
    amz_sales_rows = _filter(_load(DK_DIR / "amazon_sales_daily.json"))
    print(f"  Meta ads:     {len(meta_rows):,} rows")
    print(f"  Amazon ads:   {len(amz_ads_rows):,} rows")
    print(f"  Amazon sales: {len(amz_sales_rows):,} rows")

    rakuten_tmp = None
    if RAKUTEN_TMP.exists():
        raw = _load(RAKUTEN_TMP)
        if raw and isinstance(raw, dict):
            rakuten_tmp = raw.get("summary", {})
            print(f"  Rakuten:      loaded from .tmp (summary)")
    else:
        print(f"  Rakuten:      no data (run fetch_rakuten_sales.py first)")

    # ── Aggregate ──
    print("\n[2/4] Aggregating...")
    meta_weekly      = _agg_meta_weekly(meta_rows)
    amz_ads_weekly   = _agg_amz_ads_weekly(amz_ads_rows)
    amz_sales_weekly = _agg_amz_sales_weekly(amz_sales_rows)
    meta_monthly     = _agg_meta_monthly(meta_rows)
    amz_ads_monthly  = _agg_amz_ads_monthly(amz_ads_rows)
    amz_sales_monthly= _agg_amz_sales_monthly(amz_sales_rows)

    # Week labels for Rakuten tab (union of META + Amazon weeks)
    all_week_starts = sorted(set(
        [s for s, _, _ in meta_weekly] +
        [s for s, _, _ in amz_ads_weekly]
    ))
    labels_map = {}
    for s, lbl, _ in meta_weekly:
        labels_map[s] = lbl
    for s, lbl, _ in amz_ads_weekly:
        labels_map.setdefault(s, lbl)
    rakuten_rows = [(labels_map[s], None) for s in all_week_starts]

    print(f"  Meta:   {len(meta_weekly)} weeks, {len(meta_monthly)} months")
    print(f"  Amazon: {len(amz_ads_weekly)} weeks, {len(amz_ads_monthly)} months")

    # ── Build ──
    print("\n[3/4] Building workbook...")
    wb = Workbook()

    ws_overview = wb.active
    ws_overview.title = "Overview_Monthly"
    ws_cm      = wb.create_sheet("채널별 공헌이익")
    ws_meta    = wb.create_sheet("META")
    ws_amazon  = wb.create_sheet("AMAZON")
    ws_rakuten = wb.create_sheet("RAKUTEN")

    print("  Overview_Monthly...")
    build_overview(ws_overview, meta_monthly, amz_ads_monthly, amz_sales_monthly)
    print("  채널별 공헌이익...")
    build_cm(ws_cm, meta_monthly, amz_ads_monthly, amz_sales_monthly)
    print("  META...")
    build_meta(ws_meta, meta_weekly)
    print("  AMAZON...")
    build_amazon(ws_amazon, amz_ads_weekly, amz_sales_weekly)
    print("  RAKUTEN...")
    build_rakuten(ws_rakuten, rakuten_rows)

    # ── Save ──
    print(f"\n[4/4] Saving...")
    wb.save(OUTFILE)
    print(f"\n[OK] {OUTFILE.name}")
    print(f"     {OUTFILE}")
    print("\nNotes:")
    print("  - 노란색 셀 = 수동 입력 (Rakuten RPP 데이터)")
    print("  - RPP ROAS는 광고비 + 매출 입력 시 자동 계산")
    print("  - Rakuten 판매 데이터: fetch_rakuten_sales.py 실행 후 .tmp 확인")


if __name__ == "__main__":
    main()
