import json
import uuid
from datetime import datetime, timedelta
from decimal import Decimal

from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

# CORS origins allowed for GH Pages dashboards
_CORS_ORIGINS = (
    'https://orbiters-dev.github.io',
)


def _cors_headers(request, response):
    """Add CORS headers if origin matches allowed list."""
    origin = request.META.get('HTTP_ORIGIN', '')
    if origin in _CORS_ORIGINS:
        response['Access-Control-Allow-Origin'] = origin
        response['Access-Control-Allow-Methods'] = 'GET, POST, PUT, OPTIONS'
        response['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
        response['Access-Control-Max-Age'] = '86400'
    return response

from .models import (
    OnzUser,
    OnzOnboarding,
    OnzEngagementEvent,
    OnzRecommendationCache,
    OnzLoyaltySurvey,
    OnzCreatorProfile,
    OnzGiftingApplication,
    OnzInfluencerOutreach,
    GmailContact,
    PipelineConfig,
    PipelineCreator,
    PipelineExecutionLog,
    PipelineStatusChange,
    EmailReplyConfig,
    FAQEntry,
    EmailReplyLog,
    PipelineConversation,
    CreatorPipeline,
    CreatorContent,
)


def _json_body(request):
    """Parse JSON body from request."""
    return json.loads(request.body.decode("utf-8"))


def _serialize(obj):
    """Convert model instance to dict."""
    data = {}
    for field in obj._meta.fields:
        val = getattr(obj, field.name)
        if isinstance(val, datetime):
            val = val.isoformat()
        elif isinstance(val, uuid.UUID):
            val = str(val)
        elif isinstance(val, Decimal):
            val = float(val)
        data[field.name] = val
    # Deserialize JSON text fields
    for key in ("interests", "concerns", "purchase_factors", "discovery_channels",
                "content_preferences", "support_network", "shopping_categories",
                "other_channels", "content_types", "product_handles", "post_slugs"):
        if key in data and isinstance(data[key], str):
            try:
                data[key] = json.loads(data[key])
            except (json.JSONDecodeError, TypeError):
                pass
    return data


def _parse_date(val):
    """Parse date string to date object."""
    if not val:
        return None
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _json_field(val):
    """Convert list/dict to JSON string for TextField storage."""
    if isinstance(val, (list, dict)):
        return json.dumps(val)
    if isinstance(val, str):
        return val
    return "[]"


# --- Users ---


@csrf_exempt
@require_http_methods(["POST"])
def create_user(request):
    """Create or update user profile (upsert by id or email)."""
    body = _json_body(request)
    user_id = body.get("id")
    email = body.get("email")

    if not email:
        return JsonResponse({"error": "email is required"}, status=400)

    defaults = {
        "full_name": body.get("full_name", ""),
        "pregnancy_stage": body.get("pregnancy_stage", ""),
        "auth_provider": body.get("auth_provider", "email"),
        "interests": _json_field(body.get("interests", [])),
    }
    if body.get("baby_dob"):
        defaults["baby_dob"] = _parse_date(body["baby_dob"])
    if body.get("shopify_customer_id"):
        defaults["shopify_customer_id"] = body["shopify_customer_id"]

    if user_id:
        user, created = OnzUser.objects.update_or_create(
            id=user_id, defaults={**defaults, "email": email}
        )
    else:
        user, created = OnzUser.objects.update_or_create(
            email=email, defaults=defaults
        )

    return JsonResponse(_serialize(user), status=201 if created else 200)


@csrf_exempt
@require_http_methods(["GET", "PUT"])
def get_or_update_user(request, user_id):
    """Get or update user profile by ID."""
    try:
        user = OnzUser.objects.get(id=user_id)
    except OnzUser.DoesNotExist:
        return JsonResponse({"error": "User not found"}, status=404)

    if request.method == "PUT":
        body = _json_body(request)
        for field in ("full_name", "pregnancy_stage", "auth_provider", "shopify_customer_id"):
            if field in body:
                setattr(user, field, body[field])
        if "baby_dob" in body:
            user.baby_dob = _parse_date(body["baby_dob"])
        if "interests" in body:
            user.interests = _json_field(body["interests"])
        if "klaviyo_synced" in body:
            user.klaviyo_synced = body["klaviyo_synced"]
        user.save()

    return JsonResponse(_serialize(user))


# --- Onboarding ---


@csrf_exempt
@require_http_methods(["POST"])
def save_onboarding(request):
    """Save or update onboarding data."""
    body = _json_body(request)
    user_id = body.get("user_id")
    if not user_id:
        return JsonResponse({"error": "user_id is required"}, status=400)

    try:
        user = OnzUser.objects.get(id=user_id)
    except OnzUser.DoesNotExist:
        return JsonResponse({"error": "User not found"}, status=404)

    defaults = {
        "journey_stage": body.get("journey_stage", ""),
        "has_other_children": body.get("has_other_children", False),
        "other_children_count": body.get("other_children_count", ""),
        "concerns": _json_field(body.get("concerns", [])),
        "purchase_frequency": body.get("purchase_frequency", ""),
    }
    if body.get("baby_birthday"):
        defaults["baby_birthday"] = _parse_date(body["baby_birthday"])

    ob, created = OnzOnboarding.objects.update_or_create(user=user, defaults=defaults)
    return JsonResponse(_serialize(ob), status=201 if created else 200)


@csrf_exempt
@require_http_methods(["GET"])
def get_onboarding(request, user_id):
    """Get onboarding data for a user."""
    try:
        ob = OnzOnboarding.objects.get(user_id=user_id)
    except OnzOnboarding.DoesNotExist:
        return JsonResponse({"error": "Not found"}, status=404)
    return JsonResponse(_serialize(ob))


# --- Engagement Events ---


@csrf_exempt
@require_http_methods(["POST"])
def log_engagement(request):
    """Log a user engagement event."""
    body = _json_body(request)
    user_id = body.get("user_id")
    action = body.get("action")
    resource_type = body.get("resource_type")
    resource_id = body.get("resource_id")

    if not all([user_id, action, resource_type, resource_id]):
        return JsonResponse({"error": "user_id, action, resource_type, resource_id required"}, status=400)

    event = OnzEngagementEvent.objects.create(
        user_id=user_id,
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
    )
    return JsonResponse(_serialize(event), status=201)


@csrf_exempt
@require_http_methods(["GET"])
def get_engagement(request, user_id):
    """Get recent engagement events for a user."""
    limit = int(request.GET.get("limit", 10))
    events = OnzEngagementEvent.objects.filter(user_id=user_id)[:limit]
    return JsonResponse([_serialize(e) for e in events], safe=False)


# --- Recommendation Cache ---


@csrf_exempt
@require_http_methods(["GET", "PUT"])
def get_or_update_recommendations(request, user_id):
    """Get or upsert recommendation cache."""
    if request.method == "PUT":
        body = _json_body(request)
        rec, created = OnzRecommendationCache.objects.update_or_create(
            user_id=user_id,
            defaults={
                "product_handles": _json_field(body.get("product_handles", [])),
                "post_slugs": _json_field(body.get("post_slugs", [])),
            },
        )
        return JsonResponse(_serialize(rec), status=201 if created else 200)

    # GET
    try:
        rec = OnzRecommendationCache.objects.get(user_id=user_id)
    except OnzRecommendationCache.DoesNotExist:
        return JsonResponse({"error": "Not found"}, status=404)
    return JsonResponse(_serialize(rec))


# --- Loyalty Survey ---


@csrf_exempt
@require_http_methods(["POST"])
def save_loyalty_survey(request):
    """Save loyalty survey responses."""
    body = _json_body(request)
    user_id = body.get("user_id")
    if not user_id:
        return JsonResponse({"error": "user_id is required"}, status=400)

    defaults = {
        "purchase_factors": _json_field(body.get("purchase_factors", [])),
        "discovery_channels": _json_field(body.get("discovery_channels", [])),
        "content_preferences": _json_field(body.get("content_preferences", [])),
        "sms_opt_in": body.get("sms_opt_in", False),
        "routine_type": body.get("routine_type", ""),
        "feeding_method": body.get("feeding_method", ""),
        "support_network": _json_field(body.get("support_network", [])),
        "shopping_categories": _json_field(body.get("shopping_categories", [])),
        "discount_code": body.get("discount_code", ""),
    }

    survey, created = OnzLoyaltySurvey.objects.update_or_create(
        user_id=user_id, defaults=defaults
    )
    return JsonResponse(_serialize(survey), status=201 if created else 200)


# --- Creator Survey ---


@csrf_exempt
@require_http_methods(["POST"])
def save_creator_survey(request):
    """Save creator/influencer profile data."""
    body = _json_body(request)
    user_id = body.get("user_id")
    if not user_id:
        return JsonResponse({"error": "user_id is required"}, status=400)

    defaults = {
        "creator_level": body.get("creator_level", ""),
        "primary_platform": body.get("primary_platform", ""),
        "primary_handle": body.get("primary_handle", ""),
        "other_channels": _json_field(body.get("other_channels", [])),
        "following_size": body.get("following_size", ""),
        "content_types": _json_field(body.get("content_types", [])),
        "has_brand_deals": body.get("has_brand_deals"),
    }

    profile, created = OnzCreatorProfile.objects.update_or_create(
        user_id=user_id, defaults=defaults
    )
    return JsonResponse(_serialize(profile), status=201 if created else 200)


# --- Status ---


@csrf_exempt
@require_http_methods(["GET"])
def get_status(request, user_id):
    """Get completion status for all onboarding stages."""
    try:
        user = OnzUser.objects.get(id=user_id)
    except OnzUser.DoesNotExist:
        return JsonResponse({"error": "User not found"}, status=404)

    return JsonResponse({
        "user_id": str(user.id),
        "has_profile": True,
        "has_onboarding": OnzOnboarding.objects.filter(user=user).exists(),
        "has_loyalty_survey": OnzLoyaltySurvey.objects.filter(user=user).exists(),
        "has_creator_profile": OnzCreatorProfile.objects.filter(user=user).exists(),
        "engagement_count": OnzEngagementEvent.objects.filter(user=user).count(),
        "has_recommendations": OnzRecommendationCache.objects.filter(user=user).exists(),
    })


# --- Gifting Applications ---


@csrf_exempt
@require_http_methods(["POST"])
def save_gifting(request):
    """Save or update a gifting application (upsert by email + draft_order)."""
    body = _json_body(request)
    email = body.get("email", "").strip().lower()
    if not email:
        return JsonResponse({"error": "email is required"}, status=400)

    # Parse nested structures from n8n payload (supports both nested and flat formats)
    personal = body.get("personal_info", {})
    baby = body.get("baby_info", {})
    addr = body.get("shipping_address", {})
    child_1 = baby.get("child_1") or {}
    child_2 = baby.get("child_2") or {}

    defaults = {
        "full_name": body.get("full_name") or personal.get("full_name", ""),
        "phone": body.get("phone") or personal.get("phone", ""),
        "instagram": body.get("instagram") or personal.get("instagram", ""),
        "tiktok": body.get("tiktok") or personal.get("tiktok", ""),
        "child_1_birthday": _parse_date(child_1.get("birthday") or body.get("child_1_birthday")),
        "child_1_age_months": child_1.get("age_months") or body.get("child_1_age_months"),
        "child_2_birthday": _parse_date(child_2.get("birthday") or body.get("child_2_birthday")),
        "child_2_age_months": child_2.get("age_months") or body.get("child_2_age_months"),
        "selected_products": _json_field(body.get("selected_products", [])),
        "address_street": addr.get("street", "") or body.get("street", ""),
        "address_apt": addr.get("apt", "") or body.get("apt", ""),
        "address_city": addr.get("city", "") or body.get("city", ""),
        "address_state": addr.get("state", "") or body.get("state", ""),
        "address_zip": addr.get("zip", "") or body.get("zip", ""),
        "address_country": addr.get("country", "US") or body.get("country", "US"),
        "shopify_customer_id": str(body.get("shopify_customer_id") or ""),
    }

    # Optional fields that may come from n8n after draft order creation
    if body.get("shopify_draft_order_id"):
        defaults["shopify_draft_order_id"] = str(body["shopify_draft_order_id"])
    if body.get("shopify_draft_order_name"):
        defaults["shopify_draft_order_name"] = body["shopify_draft_order_name"]
    if body.get("airtable_record_id"):
        defaults["airtable_record_id"] = body["airtable_record_id"]
    if body.get("status"):
        defaults["status"] = body["status"]
    if body.get("submitted_at"):
        try:
            defaults["submitted_at"] = datetime.fromisoformat(
                body["submitted_at"].replace("Z", "+00:00")
            )
        except (ValueError, AttributeError):
            pass

    # Upsert: match by email. If draft_order_id given, use that for more precise match.
    draft_id = defaults.get("shopify_draft_order_id", "")
    if draft_id:
        app, created = OnzGiftingApplication.objects.update_or_create(
            email=email, shopify_draft_order_id=draft_id,
            defaults=defaults,
        )
    else:
        app, created = OnzGiftingApplication.objects.update_or_create(
            email=email, status="submitted",
            defaults=defaults,
        )

    # Auto-update PipelineCreator to "Need Review" when gifting form submitted
    try:
        creator = PipelineCreator.objects.filter(email__iexact=email).first()
        if creator and creator.pipeline_status not in (
            "Need Review", "Send Contract", "Contract Signed",
            "Sample Sent", "Sample Shipped", "Sample Delivered", "Posted",
        ):
            old_status = creator.pipeline_status
            creator.pipeline_status = "Need Review"
            creator.save(update_fields=["pipeline_status"])
            PipelineStatusChange.objects.create(
                creator_email=email,
                from_status=old_status,
                to_status="Need Review",
                changed_by="gifting-form",
            )
    except Exception:
        pass  # Don't fail gifting save if pipeline update fails

    return JsonResponse(_serialize(app), status=201 if created else 200)


@csrf_exempt
@require_http_methods(["POST"])
def update_gifting(request):
    """Update gifting application fields (from Airtable sync, status changes, etc.)."""
    body = _json_body(request)
    app_id = body.get("id")
    email = body.get("email", "").strip().lower()

    if not app_id and not email:
        return JsonResponse({"error": "id or email required"}, status=400)

    try:
        if app_id:
            app = OnzGiftingApplication.objects.get(id=app_id)
        else:
            app = OnzGiftingApplication.objects.filter(email=email).order_by("-created_at").first()
            if not app:
                return JsonResponse({"error": "Not found"}, status=404)
    except OnzGiftingApplication.DoesNotExist:
        return JsonResponse({"error": "Not found"}, status=404)

    # Update allowed fields
    for field in ("status", "airtable_record_id", "shopify_draft_order_id",
                  "shopify_draft_order_name", "shopify_customer_id",
                  "instagram", "tiktok"):
        if field in body:
            setattr(app, field, body[field])
    app.save()

    return JsonResponse(_serialize(app))


@csrf_exempt
@require_http_methods(["GET"])
def list_gifting(request):
    """List gifting applications with optional filters."""
    qs = OnzGiftingApplication.objects.all()

    # Filters
    email = request.GET.get("email")
    status = request.GET.get("status")
    if email:
        qs = qs.filter(email=email.lower())
    if status:
        qs = qs.filter(status=status)

    limit = int(request.GET.get("limit", 50))
    apps = qs[:limit]

    return JsonResponse([_serialize(a) for a in apps], safe=False)


# --- Influencer Outreach (Outbound / Pathlight pipeline) ---


@csrf_exempt
@require_http_methods(["POST"])
def save_outreach(request):
    """Upsert an outbound influencer outreach record (Pathlight pipeline).
    Separate from inbound gifting applications.
    Upsert key: airtable_record_id (if given), else email.
    """
    body = _json_body(request)
    email = body.get("email", "").strip().lower()
    airtable_record_id = body.get("airtable_record_id", "").strip()

    if not email and not airtable_record_id:
        return JsonResponse({"error": "email or airtable_record_id required"}, status=400)

    defaults = {
        "email": email,
        "ig_handle": body.get("ig_handle", body.get("instagram", "")),
        "tiktok_handle": body.get("tiktok_handle", body.get("tiktok", "")),
        "platform": body.get("platform", ""),
        "full_name": body.get("full_name", ""),
        "outreach_type": body.get("outreach_type", ""),
        "outreach_status": body.get("outreach_status", "Not Started"),
        "airtable_base_id": body.get("airtable_base_id", ""),
        "airtable_conversation_id": body.get("airtable_conversation_id", ""),
        "source": body.get("source", "manual"),
        "environment": body.get("environment", "wj_test"),
    }
    for field in ("shopify_customer_id", "shopify_draft_order_id", "shopify_draft_order_name"):
        if body.get(field):
            defaults[field] = str(body[field])

    if airtable_record_id:
        obj, created = OnzInfluencerOutreach.objects.update_or_create(
            airtable_record_id=airtable_record_id, defaults=defaults
        )
    else:
        obj, created = OnzInfluencerOutreach.objects.update_or_create(
            email=email, defaults=defaults
        )

    return JsonResponse(_serialize(obj), status=201 if created else 200)


@csrf_exempt
@require_http_methods(["POST"])
def update_outreach(request):
    """Update outreach record fields (status, shopify IDs, etc.)."""
    body = _json_body(request)
    airtable_record_id = body.get("airtable_record_id", "").strip()
    email = body.get("email", "").strip().lower()

    if not airtable_record_id and not email:
        return JsonResponse({"error": "airtable_record_id or email required"}, status=400)

    try:
        if airtable_record_id:
            obj = OnzInfluencerOutreach.objects.get(airtable_record_id=airtable_record_id)
        else:
            obj = OnzInfluencerOutreach.objects.filter(email=email).order_by("-created_at").first()
            if not obj:
                return JsonResponse({"error": "Not found"}, status=404)
    except OnzInfluencerOutreach.DoesNotExist:
        return JsonResponse({"error": "Not found"}, status=404)

    for field in ("outreach_status", "outreach_type", "shopify_customer_id",
                  "shopify_draft_order_id", "shopify_draft_order_name",
                  "airtable_conversation_id", "ig_handle", "tiktok_handle"):
        if field in body:
            setattr(obj, field, body[field])
    obj.save()

    return JsonResponse(_serialize(obj))


@csrf_exempt
@require_http_methods(["GET"])
def list_outreach(request):
    """List outreach records with optional filters."""
    qs = OnzInfluencerOutreach.objects.all()

    email = request.GET.get("email")
    status = request.GET.get("status")
    environment = request.GET.get("environment")
    airtable_record_id = request.GET.get("airtable_record_id")

    if email:
        qs = qs.filter(email=email.lower())
    if status:
        qs = qs.filter(outreach_status=status)
    if environment:
        qs = qs.filter(environment=environment)
    if airtable_record_id:
        qs = qs.filter(airtable_record_id=airtable_record_id)

    limit = int(request.GET.get("limit", 50))
    return JsonResponse([_serialize(o) for o in qs[:limit]], safe=False)


# --- Pipeline Config ---


def _serialize_config(obj):
    """Serialize PipelineConfig to dict with proper date handling."""
    data = {}
    for field in obj._meta.fields:
        val = getattr(obj, field.name)
        if isinstance(val, datetime):
            val = val.isoformat()
        elif hasattr(val, 'isoformat'):  # date objects
            val = val.isoformat() if val else None
        data[field.name] = val
    return data


@csrf_exempt
def get_pipeline_config_today(request):
    """Get today's pipeline config. Creates default if none exists."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))
    from datetime import date as date_cls
    today = date_cls.today()
    config, created = PipelineConfig.objects.get_or_create(date=today)
    return _cors_headers(request, JsonResponse(_serialize_config(config)))


@csrf_exempt
def get_or_save_pipeline_config(request, config_date):
    """Get or update pipeline config for a specific date."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))
    from datetime import date as date_cls
    try:
        d = datetime.strptime(config_date, "%Y-%m-%d").date()
    except ValueError:
        return _cors_headers(request, JsonResponse({"error": "Invalid date format. Use YYYY-MM-DD"}, status=400))

    if request.method == "GET":
        try:
            config = PipelineConfig.objects.get(date=d)
        except PipelineConfig.DoesNotExist:
            return _cors_headers(request, JsonResponse({"error": "Config not found for this date"}, status=404))
        return _cors_headers(request, JsonResponse(_serialize_config(config)))

    # POST — upsert
    body = _json_body(request)
    defaults = {}
    # Batch control
    if "update_date" in body:
        defaults["update_date"] = _parse_date(body["update_date"])
    if "start_from_beginning" in body:
        defaults["start_from_beginning"] = bool(body["start_from_beginning"])
    if "creators_contacted" in body:
        defaults["creators_contacted"] = int(body["creators_contacted"])
    if "ht_threshold" in body:
        defaults["ht_threshold"] = int(body["ht_threshold"])
    # Feature toggles
    for field in ("rag_email_dedup", "apify_autofill"):
        if field in body:
            defaults[field] = bool(body[field])
    if "human_in_loop" in body:
        defaults["human_in_loop"] = body["human_in_loop"]
    if "sender_email" in body:
        defaults["sender_email"] = body["sender_email"]
    # Templates & forms
    for field in ("outreach_template_id", "grosmimi_form_url", "chaenmom_form_url",
                  "naeiae_form_url", "ht_form_url"):
        if field in body:
            defaults[field] = body[field]
    # Brand allocation
    for field in ("alloc_grosmimi", "alloc_chaenmom", "alloc_naeiae"):
        if field in body:
            defaults[field] = int(body[field])
    # Computed fields (from preview tool)
    for field in ("eligible_total", "eligible_grosmimi", "eligible_chaenmom",
                  "eligible_naeiae", "eligible_unknown", "ht_count", "lt_count"):
        if field in body:
            defaults[field] = int(body[field])
    # Meta
    if "updated_by" in body:
        defaults["updated_by"] = body["updated_by"]

    config, created = PipelineConfig.objects.update_or_create(
        date=d, defaults=defaults
    )
    return _cors_headers(request, JsonResponse(_serialize_config(config), status=201 if created else 200))


@csrf_exempt
def pipeline_config_history(request):
    """Get recent pipeline config history (last 30 days by default)."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))
    limit = int(request.GET.get("limit", 30))
    configs = PipelineConfig.objects.all()[:limit]
    return _cors_headers(request, JsonResponse([_serialize_config(c) for c in configs], safe=False))


# --- Pipeline Conversations (draft storage for n8n) ---


@csrf_exempt
def pipeline_conversations(request):
    """GET: list conversations. POST: create new. DELETE: remove conversations."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    if request.method == 'GET':
        qs = PipelineConversation.objects.all()
        email = request.GET.get("creator_email")
        if email:
            qs = qs.filter(creator_email=email)
        status = request.GET.get("status")
        if status:
            qs = qs.filter(status=status)
        limit = min(int(request.GET.get("limit", 100)), 500)
        items = list(qs[:limit].values())
        for item in items:
            for k in ("id",):
                if k in item:
                    item[k] = str(item[k])
            for k in ("created_at", "updated_at", "email_date"):
                if k in item and item[k]:
                    item[k] = item[k].isoformat()
        return _cors_headers(request, JsonResponse({"results": items, "total": qs.count()}))

    if request.method == 'POST':
        body = json.loads(request.body or "{}")
        creator_email = body.get("creator_email", "")
        status_val = body.get("status", "Draft Ready")
        # Parse email_date if provided
        email_date_val = None
        if body.get("email_date"):
            from django.utils.dateparse import parse_datetime
            email_date_val = parse_datetime(body["email_date"])
        defaults = {
            "creator_handle": body.get("creator_handle", ""),
            "direction": body.get("direction", "Outbound"),
            "channel": body.get("channel", "Email"),
            "subject": body.get("subject", ""),
            "message_content": body.get("message_content", ""),
            "brand": body.get("brand", ""),
            "outreach_type": body.get("outreach_type", "LT"),
            "gmail_thread_id": body.get("gmail_thread_id", ""),
            "gmail_message_id": body.get("gmail_message_id", ""),
            "is_auto_sent": body.get("is_auto_sent", False),
            "email_date": email_date_val,
        }
        # Upsert: one Draft Ready conversation per creator_email+brand
        if creator_email and status_val == "Draft Ready":
            lookup = {"creator_email": creator_email, "status": "Draft Ready"}
            brand_val = body.get("brand", "")
            if brand_val:
                lookup["brand"] = brand_val
            conv, created = PipelineConversation.objects.update_or_create(
                defaults=defaults, **lookup
            )
        else:
            # Dedup by gmail_message_id if provided
            gmail_mid = body.get("gmail_message_id", "")
            if gmail_mid:
                existing = PipelineConversation.objects.filter(gmail_message_id=gmail_mid).first()
                if existing:
                    return _cors_headers(request, JsonResponse({"id": str(existing.id), "status": existing.status, "duplicate": True}, status=200))
            conv = PipelineConversation.objects.create(
                creator_email=creator_email, status=status_val, **defaults
            )
            created = True
        return _cors_headers(request, JsonResponse({"id": str(conv.id), "status": conv.status}, status=201 if created else 200))

    if request.method == 'DELETE':
        email = request.GET.get("creator_email")
        if email:
            deleted, _ = PipelineConversation.objects.filter(creator_email=email).delete()
        else:
            # Delete all non-Draft-Ready conversations (keep drafts)
            keep_drafts = request.GET.get("keep_drafts", "true") == "true"
            if keep_drafts:
                deleted, _ = PipelineConversation.objects.exclude(status="Draft Ready").delete()
            else:
                deleted, _ = PipelineConversation.objects.all().delete()
        return _cors_headers(request, JsonResponse({"deleted": deleted}))

    return _cors_headers(request, JsonResponse({"error": "Method not allowed"}, status=405))


# --- Tables (for monitoring) ---


@csrf_exempt
@require_http_methods(["GET"])
def list_tables(request):
    """List all onzenna tables with row counts."""
    tables = {
        "onz_users": OnzUser.objects.count(),
        "onz_onboarding": OnzOnboarding.objects.count(),
        "onz_engagement_events": OnzEngagementEvent.objects.count(),
        "onz_recommendation_cache": OnzRecommendationCache.objects.count(),
        "onz_loyalty_survey": OnzLoyaltySurvey.objects.count(),
        "onz_creator_profile": OnzCreatorProfile.objects.count(),
        "onz_gifting_applications": OnzGiftingApplication.objects.count(),
        "onz_influencer_outreach": OnzInfluencerOutreach.objects.count(),
        "gk_gmail_contacts": GmailContact.objects.count(),
        "onz_pipeline_config": PipelineConfig.objects.count(),
        "onz_pipeline_creators": PipelineCreator.objects.count(),
        "onz_pipeline_execution_log": PipelineExecutionLog.objects.count(),
        "onz_pipeline_status_changes": PipelineStatusChange.objects.count(),
    }
    return JsonResponse({"tables": tables})


# --- Gmail RAG Contact Lookup ---


@csrf_exempt
@require_http_methods(["GET"])
def check_gmail_contact(request):
    """Check if an email exists in the Gmail RAG contact index."""
    email = request.GET.get("email", "").strip().lower()
    if not email:
        return JsonResponse({"error": "email parameter required"}, status=400)
    try:
        c = GmailContact.objects.get(email=email)
        return JsonResponse({
            "found": True,
            "email": c.email,
            "name": c.name,
            "account": c.account,
            "total_sent": c.total_sent,
            "total_received": c.total_received,
            "last_contact_date": c.last_contact_date.isoformat() if c.last_contact_date else None,
            "last_subject": c.last_subject,
        })
    except GmailContact.DoesNotExist:
        return JsonResponse({"found": False, "email": email})


@csrf_exempt
def bulk_check_gmail_contacts(request):
    """Bulk check multiple emails against Gmail RAG contact index."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return _cors_headers(request, JsonResponse({"error": "Invalid JSON"}, status=400))

    emails = body.get("emails", [])
    if not emails or not isinstance(emails, list):
        return _cors_headers(request, JsonResponse({"error": "emails array required"}, status=400))

    emails_lower = [e.strip().lower() for e in emails if isinstance(e, str)]

    # 1. Gmail RAG check (affiliates@onzenna + hello@zezebaebae)
    found = GmailContact.objects.filter(email__in=emails_lower)
    found_map = {}
    for c in found:
        found_map[c.email] = {
            "name": c.name,
            "account": c.account,
            "total_sent": c.total_sent,
            "last_contact_date": c.last_contact_date.isoformat() if c.last_contact_date else None,
        }

    # 2. ManyChat cross-check — is_manychat_contact=True in pipeline_creators
    manychat_emails = set(
        PipelineCreator.objects.filter(
            is_manychat_contact=True,
            email__in=emails_lower,
        ).values_list("email", flat=True)
    )
    for e in manychat_emails:
        if e not in found_map:
            found_map[e] = {
                "name": "",
                "account": "manychat",
                "total_sent": 0,
                "last_contact_date": None,
            }

    results = {}
    for e in emails_lower:
        results[e] = found_map.get(e, None)

    return _cors_headers(request, JsonResponse({"results": results, "total_checked": len(emails_lower), "total_found": len(found_map)}))


@csrf_exempt
@require_http_methods(["POST"])
def sync_gmail_contacts(request):
    """Batch upsert Gmail contacts from SQLite sync script."""
    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    contacts = body.get("contacts", [])
    if not contacts or not isinstance(contacts, list):
        return JsonResponse({"error": "contacts array required"}, status=400)

    created = 0
    updated = 0
    for c in contacts:
        email = c.get("email", "").strip().lower()
        if not email:
            continue
        obj, was_created = GmailContact.objects.update_or_create(
            email=email,
            defaults={
                "name": c.get("name", ""),
                "domain": c.get("domain", ""),
                "account": c.get("account", "zezebaebae"),
                "first_contact_date": c.get("first_contact_date"),
                "last_contact_date": c.get("last_contact_date"),
                "last_subject": c.get("last_subject", ""),
                "total_sent": c.get("total_sent", 0),
                "total_received": c.get("total_received", 0),
            },
        )
        if was_created:
            created += 1
        else:
            updated += 1

    return JsonResponse({"created": created, "updated": updated, "total": len(contacts)})


# --- Pipeline Creators (CRM Dashboard) ---


def _serialize_creator(obj):
    """Serialize PipelineCreator / PipelineExecutionLog to dict."""
    data = {}
    for field in obj._meta.fields:
        val = getattr(obj, field.name)
        if isinstance(val, datetime):
            val = val.isoformat()
        elif isinstance(val, uuid.UUID):
            val = str(val)
        elif hasattr(val, 'isoformat'):
            val = val.isoformat() if val else None
        data[field.name] = val
    # JSONFields are already native list/dict from Django — no extra parsing needed
    return data


@csrf_exempt
def pipeline_creators_list(request):
    """GET: list creators with filters. POST: upsert creator."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    if request.method == 'POST':
        body = _json_body(request)
        email = body.get("email", "").strip().lower()
        if not email:
            return _cors_headers(request, JsonResponse({"error": "email is required"}, status=400))

        defaults = {
            "ig_handle": body.get("ig_handle", ""),
            "tiktok_handle": body.get("tiktok_handle", ""),
            "full_name": body.get("full_name", ""),
            "platform": body.get("platform", ""),
            "pipeline_status": body.get("pipeline_status", "Not Started"),
            "brand": body.get("brand", ""),
            "outreach_type": body.get("outreach_type", ""),
            "source": body.get("source", "manual"),
            "notes": body.get("notes", ""),
        }
        if body.get("followers") is not None:
            defaults["followers"] = int(body["followers"]) if body["followers"] else None
        if body.get("avg_views") is not None:
            defaults["avg_views"] = int(body["avg_views"]) if body["avg_views"] else None
        if body.get("initial_discovery_date"):
            defaults["initial_discovery_date"] = _parse_date(body["initial_discovery_date"])
        # Content fields (from Syncly)
        for f in ("top_post_url", "top_post_transcript", "top_post_caption"):
            if body.get(f) is not None:
                defaults[f] = str(body[f])
        for f in ("top_post_views", "views_30d", "likes_30d"):
            if body.get(f) is not None:
                defaults[f] = int(body[f]) if body[f] else None
        if body.get("top_post_date"):
            defaults["top_post_date"] = _parse_date(body["top_post_date"])
        for f in ("shopify_customer_id", "shopify_draft_order_id",
                   "shopify_draft_order_name", "airtable_record_id"):
            if body.get(f):
                defaults[f] = str(body[f])

        # New multi-source fields
        for f in ("sources", "gmail_accounts", "pr_products",
                   "apify_posted_brands", "apify_posts"):
            if f in body and isinstance(body[f], list):
                defaults[f] = body[f]
        for f in ("phone", "child_1_birthday", "child_2_birthday", "collaboration_status"):
            if f in body:
                defaults[f] = str(body[f])
        for f in ("is_shopify_pr", "is_apify_tagged", "is_manychat_contact", "is_business_account"):
            if f in body:
                defaults[f] = bool(body[f])
        if "business_category" in body:
            defaults["business_category"] = str(body["business_category"])
        for f in ("contact_count", "gmail_total_sent", "gmail_total_received", "apify_post_count"):
            if body.get(f) is not None:
                defaults[f] = int(body[f])
        if body.get("gmail_first_contact"):
            defaults["gmail_first_contact"] = _parse_date(body["gmail_first_contact"])
        if body.get("gmail_last_contact"):
            defaults["gmail_last_contact"] = _parse_date(body["gmail_last_contact"])
        if body.get("apify_last_post_date"):
            defaults["apify_last_post_date"] = _parse_date(body["apify_last_post_date"])
        if body.get("first_contacted_at"):
            try:
                defaults["first_contacted_at"] = datetime.fromisoformat(
                    body["first_contacted_at"].replace("Z", "+00:00"))
            except (ValueError, AttributeError):
                pass
        if body.get("last_contacted_at"):
            try:
                defaults["last_contacted_at"] = datetime.fromisoformat(
                    body["last_contacted_at"].replace("Z", "+00:00"))
            except (ValueError, AttributeError):
                pass
        if body.get("apify_last_crawled_at"):
            try:
                defaults["apify_last_crawled_at"] = datetime.fromisoformat(
                    body["apify_last_crawled_at"].replace("Z", "+00:00"))
            except (ValueError, AttributeError):
                pass

        creator, created = PipelineCreator.objects.update_or_create(
            email=email, defaults=defaults
        )

        # Record status change if status explicitly set and not default
        if body.get("pipeline_status") and body["pipeline_status"] != "Not Started":
            PipelineStatusChange.objects.create(
                creator_email=email,
                from_status="Not Started",
                to_status=body["pipeline_status"],
                changed_by=body.get("changed_by", "api"),
            )

        return _cors_headers(request, JsonResponse(_serialize_creator(creator), status=201 if created else 200))

    # GET — list with filters
    qs = PipelineCreator.objects.all()

    search = request.GET.get("search", "").strip()
    if search:
        from django.db.models import Q
        qs = qs.filter(
            Q(email__icontains=search) |
            Q(ig_handle__icontains=search) |
            Q(tiktok_handle__icontains=search) |
            Q(full_name__icontains=search)
        )

    status = request.GET.get("status")
    if status:
        qs = qs.filter(pipeline_status=status)

    region = request.GET.get("region")
    if region:
        qs = qs.filter(region__iexact=region)

    brand = request.GET.get("brand")
    if brand:
        qs = qs.filter(brand=brand)

    source = request.GET.get("source")
    if source:
        qs = qs.filter(source=source)

    outreach_type = request.GET.get("type")
    if outreach_type:
        qs = qs.filter(outreach_type=outreach_type)

    # assigned_to filter: maps owner name → brand(s)
    _OWNER_BRANDS = {
        "Jeehoo": ["Grosmimi"],
        "Laeeka": ["CHA&MOM"],
        "Soyeon": ["Naeiae"],
    }
    assigned_to = request.GET.get("assigned_to")
    if assigned_to is not None:
        if assigned_to == "":
            # Unassigned = brands not in any owner mapping
            all_owned = [b for bs in _OWNER_BRANDS.values() for b in bs]
            qs = qs.exclude(brand__in=all_owned)
        elif assigned_to in _OWNER_BRANDS:
            qs = qs.filter(brand__in=_OWNER_BRANDS[assigned_to])

    # Discovery date filter (for weekly batch selection)
    discovery_date = request.GET.get("discovery_date")
    if discovery_date:
        qs = qs.filter(initial_discovery_date=discovery_date)

    # Cross-check filters
    email_suffix = request.GET.get("email_suffix")
    if email_suffix:
        qs = qs.filter(email__endswith=email_suffix)

    if request.GET.get("is_shopify_pr") == "true":
        qs = qs.filter(is_shopify_pr=True)

    if request.GET.get("is_apify_tagged") == "true":
        qs = qs.filter(is_apify_tagged=True)

    if request.GET.get("is_manychat_contact") == "true":
        qs = qs.filter(is_manychat_contact=True)

    # Content type filter: video / image (based on gk_content_posts URL pattern)
    content_type = request.GET.get("content_type")
    if content_type in ("video", "image"):
        from django.db import connection
        if content_type == "video":
            ct_sql = """
                SELECT DISTINCT LOWER(username) FROM gk_content_posts
                WHERE (url LIKE '%%/reel/%%' OR url LIKE '%%/reels/%%' OR url LIKE '%%tiktok%%'
                       OR (transcript IS NOT NULL AND LENGTH(transcript) > 10))
                AND username IS NOT NULL AND username != ''
            """
        else:
            ct_sql = """
                SELECT DISTINCT LOWER(username) FROM gk_content_posts
                WHERE url NOT LIKE '%%/reel/%%' AND url NOT LIKE '%%/reels/%%'
                AND url NOT LIKE '%%tiktok%%'
                AND (transcript IS NULL OR LENGTH(transcript) <= 10)
                AND username IS NOT NULL AND username != ''
            """
        with connection.cursor() as cursor:
            cursor.execute(ct_sql)
            ct_handles = {r[0] for r in cursor.fetchall()}
        if ct_handles:
            from django.db.models import Q as _Q
            qs = qs.filter(_Q(ig_handle__in=[h for h in ct_handles]) |
                           _Q(tiktok_handle__in=[h for h in ct_handles]))
        else:
            qs = qs.none()

    # Ordering (whitelist to prevent field enumeration)
    ALLOWED_ORDERS = {
        "-created_at", "created_at", "-followers", "followers",
        "-avg_views", "avg_views", "email", "-email",
        "-updated_at", "updated_at", "pipeline_status", "-pipeline_status",
        "-initial_discovery_date", "initial_discovery_date",
    }
    order = request.GET.get("order", "-created_at")
    if order not in ALLOWED_ORDERS:
        order = "-created_at"
    qs = qs.order_by(order)

    # Pagination (safe parsing)
    try:
        page = max(1, int(request.GET.get("page", 1)))
    except (ValueError, TypeError):
        page = 1
    try:
        limit = min(500, max(1, int(request.GET.get("limit", 50))))
    except (ValueError, TypeError):
        limit = 50
    total = qs.count()
    offset = (page - 1) * limit
    creators = list(qs[offset:offset + limit])

    # ── Enrich with latest post data from gk_content_posts ──
    handles = set()
    for c in creators:
        if c.ig_handle:
            handles.add(c.ig_handle.lower())
        if c.tiktok_handle:
            handles.add(c.tiktok_handle.lower())

    latest_posts = {}  # handle -> {views, likes, post_date, followers}
    if handles:
        from django.db import connection
        placeholders = ",".join(["%s"] * len(handles))
        sql = f"""
            SELECT DISTINCT ON (LOWER(username))
                LOWER(username), views_30d, likes_30d, post_date, followers
            FROM gk_content_posts
            WHERE LOWER(username) IN ({placeholders})
              AND username IS NOT NULL AND username != ''
            ORDER BY LOWER(username), post_date DESC NULLS LAST
        """
        with connection.cursor() as cursor:
            cursor.execute(sql, list(handles))
            for row in cursor.fetchall():
                latest_posts[row[0]] = {
                    "latest_views": row[1],
                    "latest_likes": row[2],
                    "latest_post_date": row[3].isoformat() if row[3] else None,
                    "cp_followers": row[4],
                }

    results = []
    for c in creators:
        data = _serialize_creator(c)
        # Merge latest post data
        lp = latest_posts.get((c.ig_handle or "").lower()) or latest_posts.get((c.tiktok_handle or "").lower()) or {}
        data["latest_views"] = lp.get("latest_views")
        data["latest_likes"] = lp.get("latest_likes")
        data["latest_post_date"] = lp.get("latest_post_date")
        # Fallback followers from content_posts if creator has 0/null
        if not data.get("followers") and lp.get("cp_followers"):
            data["followers"] = lp["cp_followers"]
        results.append(data)

    return _cors_headers(request, JsonResponse({
        "results": results,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit,
    }))


@csrf_exempt
def pipeline_creator_detail(request, creator_id):
    """GET: creator detail. PUT: update creator (with status audit)."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    try:
        creator = PipelineCreator.objects.get(id=creator_id)
    except PipelineCreator.DoesNotExist:
        return _cors_headers(request, JsonResponse({"error": "Creator not found"}, status=404))

    if request.method == 'PUT':
        body = _json_body(request)
        old_status = creator.pipeline_status

        # Brand assignment lock: Not Started → brand change blocked
        if "brand" in body and body["brand"] and creator.pipeline_status == "Not Started":
            return _cors_headers(request, JsonResponse(
                {"error": "Cannot assign brand while status is 'Not Started'. Set status to 'Draft Ready' or higher first."},
                status=400))

        for field in ("ig_handle", "tiktok_handle", "full_name", "platform",
                      "pipeline_status", "brand", "outreach_type", "source", "notes",
                      "shopify_customer_id", "shopify_draft_order_id",
                      "shopify_draft_order_name", "airtable_record_id",
                      "phone", "child_1_birthday", "child_2_birthday", "collaboration_status"):
            if field in body:
                setattr(creator, field, body[field])

        if "followers" in body:
            creator.followers = int(body["followers"]) if body["followers"] else None
        if "avg_views" in body:
            creator.avg_views = int(body["avg_views"]) if body["avg_views"] else None
        if "initial_discovery_date" in body:
            creator.initial_discovery_date = _parse_date(body["initial_discovery_date"])

        # JSONField updates
        for f in ("sources", "gmail_accounts", "pr_products",
                   "apify_posted_brands", "apify_posts"):
            if f in body and isinstance(body[f], list):
                setattr(creator, f, body[f])
        # Boolean flags
        for f in ("is_shopify_pr", "is_apify_tagged", "is_manychat_contact"):
            if f in body:
                setattr(creator, f, bool(body[f]))
        # Integer fields
        for f in ("contact_count", "gmail_total_sent", "gmail_total_received", "apify_post_count"):
            if f in body and body[f] is not None:
                setattr(creator, f, int(body[f]))
        # Date fields
        for f in ("gmail_first_contact", "gmail_last_contact", "apify_last_post_date"):
            if f in body:
                setattr(creator, f, _parse_date(body[f]))
        # Datetime fields
        for f in ("first_contacted_at", "last_contacted_at", "apify_last_crawled_at"):
            if f in body and body[f]:
                try:
                    setattr(creator, f, datetime.fromisoformat(
                        body[f].replace("Z", "+00:00")))
                except (ValueError, AttributeError):
                    pass

        creator.save()

        # Audit status change
        new_status = creator.pipeline_status
        if new_status != old_status:
            PipelineStatusChange.objects.create(
                creator_email=creator.email,
                from_status=old_status,
                to_status=new_status,
                changed_by=body.get("changed_by", "dashboard"),
            )

        return _cors_headers(request, JsonResponse(_serialize_creator(creator)))

    # GET
    return _cors_headers(request, JsonResponse(_serialize_creator(creator)))


@csrf_exempt
@require_http_methods(["PUT", "OPTIONS"])
def pipeline_creator_by_email(request, email):
    """PUT: update creator by email address (for n8n webhook callbacks)."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))
    try:
        creator = PipelineCreator.objects.get(email=email)
    except PipelineCreator.DoesNotExist:
        return _cors_headers(request, JsonResponse({"error": "Creator not found"}, status=404))
    body = _json_body(request)
    old_status = creator.pipeline_status
    if "pipeline_status" in body:
        creator.pipeline_status = body["pipeline_status"]
    if "brand" in body:
        creator.brand = body["brand"]
    creator.save()
    if creator.pipeline_status != old_status:
        PipelineStatusChange.objects.create(
            creator_email=creator.email,
            from_status=old_status,
            to_status=creator.pipeline_status,
            changed_by=body.get("changed_by", "n8n"),
        )
    return _cors_headers(request, JsonResponse(_serialize_creator(creator)))


@csrf_exempt
def pipeline_creators_stats(request):
    """Aggregate stats: by status, by brand, recent activity."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    from django.db.models import Count, Q
    from datetime import timedelta
    from datetime import date as date_cls

    total = PipelineCreator.objects.count()

    # Status breakdown
    status_counts = dict(
        PipelineCreator.objects.values_list('pipeline_status')
        .annotate(c=Count('id'))
        .values_list('pipeline_status', 'c')
    )

    # Brand breakdown
    brand_counts = dict(
        PipelineCreator.objects.values_list('brand')
        .annotate(c=Count('id'))
        .values_list('brand', 'c')
    )

    # This week's new creators
    week_ago = date_cls.today() - timedelta(days=7)
    new_this_week = PipelineCreator.objects.filter(created_at__date__gte=week_ago).count()

    # Type breakdown
    type_counts = dict(
        PipelineCreator.objects.values_list('outreach_type')
        .annotate(c=Count('id'))
        .values_list('outreach_type', 'c')
    )

    # LT/HT status breakdown (for Y-fork funnel)
    lt_status = dict(
        PipelineCreator.objects.filter(outreach_type='LT')
        .values_list('pipeline_status')
        .annotate(c=Count('id'))
        .values_list('pipeline_status', 'c')
    )
    ht_status = dict(
        PipelineCreator.objects.filter(outreach_type='HT')
        .values_list('pipeline_status')
        .annotate(c=Count('id'))
        .values_list('pipeline_status', 'c')
    )

    # @discovered.syncly email count (unconfirmed)
    discovered_email_count = PipelineCreator.objects.filter(
        email__endswith="@discovered.syncly"
    ).count()

    # Discovery date breakdown (Not Started only, for batch dropdown)
    from django.db.models.functions import TruncDate
    discovery_date_qs = (
        PipelineCreator.objects
        .filter(pipeline_status='Not Started')
        .exclude(initial_discovery_date__isnull=True)
        .values('initial_discovery_date')
        .annotate(c=Count('id'))
        .order_by('-initial_discovery_date')
    )
    by_discovery_date = {}
    for row in discovery_date_qs:
        d = row['initial_discovery_date']
        by_discovery_date[str(d)] = row['c']

    # Cross-check summary
    shopify_pr_count = PipelineCreator.objects.filter(is_shopify_pr=True).count()
    apify_tagged_count = PipelineCreator.objects.filter(is_apify_tagged=True).count()
    manychat_count = PipelineCreator.objects.filter(is_manychat_contact=True).count()

    # Data source freshness (from gk_* tables)
    from django.db import connection
    data_sources = {}
    try:
        with connection.cursor() as cur:
            # Syncly (source='syncly' in content_posts)
            cur.execute("SELECT COUNT(*), MAX(collected_at) FROM gk_content_posts WHERE source='syncly'")
            row = cur.fetchone()
            data_sources["syncly"] = {"count": row[0] or 0, "last_collected": str(row[1]) if row[1] else None}

            # Apify (source='apify' or all posts)
            cur.execute("SELECT COUNT(*), MAX(collected_at) FROM gk_content_posts")
            row = cur.fetchone()
            data_sources["apify"] = {"count": row[0] or 0, "last_collected": str(row[1]) if row[1] else None}

            # Shopify PR orders
            cur.execute("SELECT COUNT(*), MAX(collected_at) FROM gk_influencer_orders")
            row = cur.fetchone()
            data_sources["shopify_pr"] = {"count": row[0] or 0, "last_collected": str(row[1]) if row[1] else None}

            # Gmail contacts
            cur.execute("SELECT COUNT(*), MAX(synced_at) FROM gk_gmail_contacts")
            row = cur.fetchone()
            data_sources["gmail"] = {"count": row[0] or 0, "last_collected": str(row[1]) if row[1] else None}
    except Exception:
        pass  # gk_ tables may not exist in all envs

    return _cors_headers(request, JsonResponse({
        "total": total,
        "by_status": status_counts,
        "by_status_lt": lt_status,
        "by_status_ht": ht_status,
        "by_brand": brand_counts,
        "by_type": type_counts,
        "new_this_week": new_this_week,
        "discovered_email_count": discovered_email_count,
        "shopify_pr_count": shopify_pr_count,
        "apify_tagged_count": apify_tagged_count,
        "manychat_count": manychat_count,
        "data_sources": data_sources,
        "by_discovery_date": by_discovery_date,
    }))


@csrf_exempt
def datapool_stats(request):
    """Aggregate stats for Creator Datapool Dashboard visualization."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    from django.db.models import Count, Q, Avg, Max, Min
    from django.db.models.functions import TruncDate
    from datetime import date as date_cls

    qs = PipelineCreator.objects.all()
    total = qs.count()

    # --- Region distribution ---
    region_counts = dict(
        qs.values_list('region')
        .annotate(c=Count('id'))
        .values_list('region', 'c')
    )

    # --- Pipeline status distribution ---
    status_counts = dict(
        qs.values_list('pipeline_status')
        .annotate(c=Count('id'))
        .values_list('pipeline_status', 'c')
    )

    # --- Brand distribution ---
    brand_counts = dict(
        qs.values_list('brand')
        .annotate(c=Count('id'))
        .values_list('brand', 'c')
    )

    # --- Outreach type (LT/HT) ---
    type_counts = dict(
        qs.values_list('outreach_type')
        .annotate(c=Count('id'))
        .values_list('outreach_type', 'c')
    )

    # --- Email source (normal vs @discovered.* vs @noemail.placeholder) ---
    discovered_count = qs.filter(email__contains='@discovered.').count()
    noemail_count = qs.filter(email__contains='@noemail.').count()
    no_email = qs.filter(Q(email='') | Q(email__isnull=True)).count()
    valid_email = total - discovered_count - noemail_count - no_email

    # --- Followers distribution (buckets) ---
    followers_buckets = {
        '0-1K': qs.filter(followers__gte=0, followers__lt=1000).count(),
        '1K-5K': qs.filter(followers__gte=1000, followers__lt=5000).count(),
        '5K-10K': qs.filter(followers__gte=5000, followers__lt=10000).count(),
        '10K-50K': qs.filter(followers__gte=10000, followers__lt=50000).count(),
        '50K-100K': qs.filter(followers__gte=50000, followers__lt=100000).count(),
        '100K+': qs.filter(followers__gte=100000).count(),
        'Unknown': qs.filter(followers__isnull=True).count(),
    }

    # --- Daily intake (last 30 days) ---
    thirty_days_ago = date_cls.today() - timedelta(days=30)
    daily_intake = list(
        qs.filter(created_at__date__gte=thirty_days_ago)
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(c=Count('id'))
        .order_by('day')
        .values_list('day', 'c')
    )
    daily_intake_dict = {str(d): c for d, c in daily_intake}

    # --- Platform distribution ---
    platform_counts = dict(
        qs.values_list('platform')
        .annotate(c=Count('id'))
        .values_list('platform', 'c')
    )

    # --- Business account ratio ---
    biz_true = qs.filter(is_business_account=True).count()
    biz_false = total - biz_true

    # --- Source tags distribution ---
    from django.db import connection
    source_tag_counts = {}
    try:
        with connection.cursor() as cur:
            cur.execute("""
                SELECT s.val, COUNT(*)
                FROM onz_pipeline_creators c,
                     LATERAL jsonb_array_elements_text(c.sources) AS s(val)
                GROUP BY s.val ORDER BY COUNT(*) DESC
            """)
            for row in cur.fetchall():
                source_tag_counts[row[0]] = row[1]
    except Exception:
        pass

    # --- Follower stats ---
    fstats = qs.exclude(followers__isnull=True).aggregate(
        avg=Avg('followers'), mx=Max('followers'), mn=Min('followers')
    )

    # --- Content Pool stats (CreatorContent table) ---
    from django.db.models import Sum
    content_qs = CreatorContent.objects.all()
    content_total = content_qs.count()
    content_non_partnered = content_qs.filter(content_type="non_partnered").count()
    content_partnered = content_qs.filter(content_type="partnered").count()
    content_with_transcript = content_qs.exclude(transcript="").count()
    content_with_views = content_qs.exclude(views__isnull=True).exclude(views=0)
    content_views_stats = content_with_views.aggregate(
        avg=Avg('views'), total=Sum('views')
    )
    content_by_platform = dict(
        content_qs.values_list('platform')
        .annotate(c=Count('id'))
        .values_list('platform', 'c')
    )

    # Last 30 days content stats
    thirty_days_ago_content = date_cls.today() - timedelta(days=30)
    recent_content = content_qs.filter(post_date__gte=thirty_days_ago_content)
    content_30d_posts = recent_content.count()
    content_30d_views = recent_content.aggregate(total=Sum('views'))['total'] or 0

    # Creators with content (linked)
    creators_with_content = content_qs.values('creator_id').distinct().count()

    # Top creators by content count
    top_creators = list(
        content_qs.values('creator__ig_handle', 'creator_id')
        .annotate(post_count=Count('id'))
        .order_by('-post_count')[:10]
    )

    # --- Pipeline campaigns stats (CreatorPipeline table) ---
    pipeline_qs = CreatorPipeline.objects.all()
    pipeline_total = pipeline_qs.count()
    pipeline_by_brand = dict(
        pipeline_qs.values_list('brand')
        .annotate(c=Count('id'))
        .values_list('brand', 'c')
    )
    pipeline_by_status = dict(
        pipeline_qs.values_list('pipeline_status')
        .annotate(c=Count('id'))
        .values_list('pipeline_status', 'c')
    )

    return _cors_headers(request, JsonResponse({
        "total": total,
        "by_region": region_counts,
        "by_status": status_counts,
        "by_brand": brand_counts,
        "by_outreach_type": type_counts,
        "email_source": {
            "valid": valid_email,
            "discovered": discovered_count,
            "noemail_placeholder": noemail_count,
            "no_email": no_email,
        },
        "followers_buckets": followers_buckets,
        "followers_stats": {
            "avg": round(fstats['avg'] or 0),
            "max": fstats['mx'] or 0,
            "min": fstats['mn'] or 0,
        },
        "daily_intake": daily_intake_dict,
        "by_platform": platform_counts,
        "business_account": {"yes": biz_true, "no": biz_false},
        "by_source_tag": source_tag_counts,
        "content_pool": {
            "total": content_total,
            "non_partnered": content_non_partnered,
            "partnered": content_partnered,
            "with_transcript": content_with_transcript,
            "creators_linked": creators_with_content,
            "avg_views": round(content_views_stats['avg'] or 0),
            "total_views": content_views_stats['total'] or 0,
            "posts_30d": content_30d_posts,
            "views_30d": content_30d_views,
            "by_platform": content_by_platform,
            "top_creators": [
                {"handle": t['creator__ig_handle'] or '(unknown)', "count": t['post_count']}
                for t in top_creators
            ],
        },
        "pipeline_campaigns": {
            "total": pipeline_total,
            "by_brand": pipeline_by_brand,
            "by_status": pipeline_by_status,
        },
    }))


@csrf_exempt
def creator_content_list(request):
    """GET: list CreatorContent with pagination + filters."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    from django.db.models import Q

    qs = CreatorContent.objects.select_related('creator').all()

    # Filters
    search = request.GET.get("search", "").strip()
    if search:
        qs = qs.filter(
            Q(creator__ig_handle__icontains=search) |
            Q(creator__tiktok_handle__icontains=search) |
            Q(caption__icontains=search) |
            Q(transcript__icontains=search)
        )
    platform = request.GET.get("platform", "").strip()
    if platform:
        qs = qs.filter(platform__iexact=platform)
    content_type = request.GET.get("content_type", "").strip()
    if content_type:
        qs = qs.filter(content_type=content_type)
    region = request.GET.get("region", "").strip()
    if region:
        qs = qs.filter(creator__region__iexact=region)
    has_transcript = request.GET.get("has_transcript", "").strip()
    if has_transcript == "1":
        qs = qs.exclude(transcript="")
    elif has_transcript == "0":
        qs = qs.filter(Q(transcript="") | Q(transcript__isnull=True))
    has_caption = request.GET.get("has_caption", "").strip()
    if has_caption == "1":
        qs = qs.exclude(caption="").exclude(caption__isnull=True)
    elif has_caption == "0":
        qs = qs.filter(Q(caption="") | Q(caption__isnull=True))
    views_filter = request.GET.get("views_min", "").strip()
    if views_filter == "1m":
        qs = qs.filter(views__gte=1000000)
    elif views_filter == "100k":
        qs = qs.filter(views__gte=100000)
    elif views_filter == "10k":
        qs = qs.filter(views__gte=10000)
    elif views_filter == "1k":
        qs = qs.filter(views__gte=1000)
    elif views_filter == "0":
        qs = qs.filter(Q(views=0) | Q(views__isnull=True))
    likes_filter = request.GET.get("likes_min", "").strip()
    if likes_filter == "100k":
        qs = qs.filter(likes__gte=100000)
    elif likes_filter == "10k":
        qs = qs.filter(likes__gte=10000)
    elif likes_filter == "1k":
        qs = qs.filter(likes__gte=1000)
    elif likes_filter == "100":
        qs = qs.filter(likes__gte=100)
    elif likes_filter == "0":
        qs = qs.filter(Q(likes=0) | Q(likes__isnull=True))
    has_comments = request.GET.get("has_comments", "").strip()
    if has_comments == "1":
        qs = qs.filter(comments__gt=0)
    elif has_comments == "0":
        qs = qs.filter(Q(comments=0) | Q(comments__isnull=True))
    post_date_range = request.GET.get("post_date_range", "").strip()
    if post_date_range:
        from datetime import timedelta
        from django.utils import timezone
        days_map = {"7d": 7, "30d": 30, "60d": 60, "90d": 90}
        days = days_map.get(post_date_range)
        if days:
            cutoff = (timezone.now() - timedelta(days=days)).date()
            qs = qs.filter(post_date__gte=cutoff)

    # Ordering
    order = request.GET.get("order", "-views")
    ALLOWED = {"-views", "views", "-created_at", "created_at", "-post_date", "post_date", "-likes", "likes", "-comments", "comments"}
    if order not in ALLOWED:
        order = "-views"
    from django.db.models import F
    # Push NULLs to the end for descending sorts
    if order.startswith("-"):
        field = order[1:]
        qs = qs.order_by(F(field).desc(nulls_last=True))
    else:
        qs = qs.order_by(F(order).asc(nulls_last=True))

    # Pagination
    try:
        page = max(1, int(request.GET.get("page", 1)))
    except (ValueError, TypeError):
        page = 1
    try:
        limit = min(500, max(1, int(request.GET.get("limit", 100))))
    except (ValueError, TypeError):
        limit = 100
    total = qs.count()
    offset = (page - 1) * limit
    items = qs[offset:offset + limit]

    results = []
    for c in items:
        results.append({
            "id": str(c.id),
            "creator_handle": c.creator.ig_handle or c.creator.tiktok_handle or "",
            "creator_region": c.creator.region or "",
            "post_url": c.post_url,
            "platform": c.platform,
            "post_date": str(c.post_date) if c.post_date else "",
            "content_type": c.content_type,
            "partner_brand": c.partner_brand,
            "views": c.views,
            "likes": c.likes,
            "comments": c.comments,
            "transcript": c.transcript or "",
            "caption": c.caption or "",
            "quality_score": c.quality_score,
            "fit_score": c.fit_score,
            "scene_fit": c.scene_fit,
            "created_at": c.created_at.isoformat() if c.created_at else "",
        })

    return _cors_headers(request, JsonResponse({
        "results": results,
        "total": total,
        "page": page,
        "limit": limit,
        "pages": (total + limit - 1) // limit,
    }))


@csrf_exempt
def pipeline_filter_stats(request):
    """GET: Server-side filter pipeline statistics for Feature Toggle dashboard.

    Avoids loading 10K+ records into the browser by computing filter stats on the server.
    Accepts ?discovery_date=YYYY-MM-DD to scope to a single batch week.
    """
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    from django.db.models import Q

    discovery_date = request.GET.get("discovery_date")
    region = request.GET.get("region")

    qs = PipelineCreator.objects.filter(pipeline_status='Not Started')
    if region:
        qs = qs.filter(region__iexact=region)
    if discovery_date:
        # Support comma-separated dates for week groups (e.g. 2026-02-04,2026-02-05,...)
        dates = [d.strip() for d in discovery_date.split(",") if d.strip()]
        if len(dates) == 1:
            qs = qs.filter(initial_discovery_date=dates[0])
        else:
            qs = qs.filter(initial_discovery_date__in=dates)

    total = qs.count()
    with_email = qs.exclude(
        Q(email__contains='@discovered.') | Q(email='') | Q(email__isnull=True)
    ).count()
    no_email = total - with_email
    business = qs.filter(is_business_account=True).count()
    ht = qs.filter(outreach_type='HT').count()
    lt = total - ht

    # HT/LT with real email breakdown (for LT-only filter pipeline)
    email_filter = Q(email__endswith='@discovered.syncly') | Q(email='')
    ht_with_email = qs.filter(outreach_type='HT').exclude(email_filter).count()
    lt_with_email = with_email - ht_with_email

    # Apify fill rate: actual measured from DB (LT only — HT doesn't need Apify)
    lt_qs = qs.exclude(outreach_type='HT')
    apify_crawled = lt_qs.filter(is_apify_tagged=True).count()
    apify_with_email = lt_qs.filter(is_apify_tagged=True).exclude(email_filter).count()
    apify_fill_rate = round(apify_with_email / apify_crawled, 4) if apify_crawled > 0 else 0.09

    return _cors_headers(request, JsonResponse({
        "total": total,
        "with_email": with_email,
        "no_email": no_email,
        "business_accounts": business,
        "ht": ht,
        "lt": lt,
        "ht_with_email": ht_with_email,
        "lt_with_email": lt_with_email,
        "apify_crawled": apify_crawled,
        "apify_with_email": apify_with_email,
        "apify_fill_rate": apify_fill_rate,
    }))


@csrf_exempt
def pipeline_creators_bulk_status(request):
    """Bulk update status for multiple creators."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    body = _json_body(request)
    creator_ids = body.get("ids", [])
    new_status = body.get("status", "")
    changed_by = body.get("changed_by", "dashboard")

    if not creator_ids or not new_status:
        return _cors_headers(request, JsonResponse({"error": "ids and status required"}, status=400))

    updated = 0
    for cid in creator_ids:
        try:
            creator = PipelineCreator.objects.get(id=cid)
            old_status = creator.pipeline_status
            if old_status != new_status:
                creator.pipeline_status = new_status
                creator.save()
                PipelineStatusChange.objects.create(
                    creator_email=creator.email,
                    from_status=old_status,
                    to_status=new_status,
                    changed_by=changed_by,
                )
                updated += 1
        except PipelineCreator.DoesNotExist:
            continue

    return _cors_headers(request, JsonResponse({"updated": updated, "total": len(creator_ids)}))


# --- Pipeline Execution Log ---


@csrf_exempt
def pipeline_execution_log(request):
    """GET: list execution logs. POST: add new log entry."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    if request.method == 'POST':
        body = _json_body(request)
        action_type = body.get("action_type", "")
        if not action_type:
            return _cors_headers(request, JsonResponse({"error": "action_type required"}, status=400))

        log = PipelineExecutionLog.objects.create(
            action_type=action_type,
            triggered_by=body.get("triggered_by", ""),
            target_count=int(body.get("target_count", 0)),
            status=body.get("status", "pending"),
            details=json.dumps(body.get("details", {})) if isinstance(body.get("details"), dict) else body.get("details", "{}"),
        )
        if body.get("completed_at"):
            try:
                log.completed_at = datetime.fromisoformat(body["completed_at"].replace("Z", "+00:00"))
                log.save()
            except (ValueError, AttributeError):
                pass

        return _cors_headers(request, JsonResponse(_serialize_creator(log), status=201))

    # GET — list
    limit = int(request.GET.get("limit", 50))
    action_type = request.GET.get("action_type")
    qs = PipelineExecutionLog.objects.all()
    if action_type:
        qs = qs.filter(action_type=action_type)

    logs = qs[:limit]
    results = []
    for log in logs:
        d = {}
        for field in log._meta.fields:
            val = getattr(log, field.name)
            if isinstance(val, datetime):
                val = val.isoformat()
            elif isinstance(val, uuid.UUID):
                val = str(val)
            d[field.name] = val
        # Parse details JSON
        if isinstance(d.get("details"), str):
            try:
                d["details"] = json.loads(d["details"])
            except (json.JSONDecodeError, TypeError):
                pass
        results.append(d)

    return _cors_headers(request, JsonResponse({"results": results, "total": qs.count()}, safe=False))


# --- Syncly Discovery Import ---


@csrf_exempt
def import_syncly_discovery(request):
    """POST: Import creators from gk_content_posts into pipeline_creators.

    Finds creators in content_posts that don't exist in pipeline_creators yet.
    Creates them with status='Not Started', source='syncly'.

    Body (optional):
      brand: filter by brand (default: all)
      limit: max creators to import (default: 50)
      days: look back N days (default: 30)
    """
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    body = _json_body(request)
    brand_filter = body.get("brand", "")
    region = body.get("region", "")
    source_label = body.get("source", "syncly")
    limit = int(body.get("limit", 50))
    days = int(body.get("days", 90))

    from django.db import connection
    from datetime import timedelta
    from datetime import date as date_cls

    cutoff = date_cls.today() - timedelta(days=days)

    # Query gk_content_posts for unique creators not already in pipeline
    sql = """
        SELECT DISTINCT ON (cp.username)
            cp.username, cp.nickname, cp.followers, cp.platform,
            cp.brand, cp.caption, cp.url, cp.post_date
        FROM gk_content_posts cp
        LEFT JOIN onz_pipeline_creators pc
            ON LOWER(cp.username) = LOWER(pc.ig_handle)
            OR LOWER(cp.username) = LOWER(pc.tiktok_handle)
        WHERE pc.id IS NULL
          AND cp.post_date >= %s
          AND cp.username IS NOT NULL
          AND cp.username != ''
    """
    params = [cutoff]

    if region:
        sql += " AND LOWER(cp.region) = LOWER(%s)"
        params.append(region)

    if brand_filter:
        sql += " AND LOWER(cp.brand) = LOWER(%s)"
        params.append(brand_filter)

    sql += " ORDER BY cp.username, cp.followers DESC NULLS LAST LIMIT %s"
    params.append(limit)

    created = 0
    skipped = 0
    imported = []

    try:
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            rows = cursor.fetchall()

        for row in rows:
            username, nickname, followers, platform, brand, caption, url, post_date = row

            # Clean handle (remove @)
            handle = (username or "").lstrip("@").strip()
            if not handle:
                skipped += 1
                continue

            # Determine platform
            plat = (platform or "").lower()
            ig_handle = handle if "tiktok" not in plat else ""
            tiktok_handle = handle if "tiktok" in plat else ""

            # Generate placeholder email (normalize: lowercase, strip dots/dashes/underscores)
            normalized_handle = handle.lower().replace('.', '_').replace('-', '_')
            email = f"{normalized_handle}@discovered.syncly"

            # Check if already exists by handle (case-insensitive) or normalized email
            handle_lower = handle.lower()
            exists = PipelineCreator.objects.filter(
                ig_handle__iexact=handle_lower
            ).exists() or PipelineCreator.objects.filter(
                tiktok_handle__iexact=handle_lower
            ).exists() or PipelineCreator.objects.filter(
                email__iexact=email
            ).exists()

            if exists:
                skipped += 1
                continue

            creator = PipelineCreator.objects.create(
                email=email,
                ig_handle=ig_handle,
                tiktok_handle=tiktok_handle,
                full_name=nickname or handle,
                platform="TikTok" if "tiktok" in plat else "Instagram",
                pipeline_status="Not Started",
                brand=brand or "",
                outreach_type="LT" if (followers or 0) < 100000 else "HT",
                source=source_label,
                region=region or "us",
                followers=followers,
                avg_views=int((followers or 0) * (0.15 if "tiktok" in plat else 0.08)),
                initial_discovery_date=post_date or date_cls.today(),
                notes=url or "",
            )
            created += 1
            imported.append({
                "id": str(creator.id),
                "handle": handle,
                "brand": brand,
                "followers": followers,
            })

    except Exception as e:
        return _cors_headers(request, JsonResponse({
            "error": str(e),
            "created": created,
            "skipped": skipped,
        }, status=500))

    # Auto-sync transcripts after import (fixes 85% missing transcripts)
    transcript_result = {"updated": 0, "checked": 0, "matched": 0}
    if created > 0 and region:
        try:
            from django.db import connection as _conn
            _sync_sql = """
                SELECT
                    cp.username, cp.transcript, cp.url, cp.views_30d,
                    cp.post_date, pc.id AS creator_id
                FROM gk_content_posts cp
                INNER JOIN onz_pipeline_creators pc
                    ON LOWER(cp.username) = LOWER(pc.ig_handle)
                WHERE cp.transcript IS NOT NULL
                  AND LENGTH(cp.transcript) >= 20
                  AND cp.username IS NOT NULL AND cp.username != ''
                  AND LOWER(cp.region) = LOWER(%s)
                ORDER BY cp.views_30d DESC NULLS LAST
                LIMIT 500
            """
            with _conn.cursor() as cur:
                cur.execute(_sync_sql, [region])
                t_rows = cur.fetchall()

            best_by_creator = {}
            for uname, transcript, t_url, views, pdate, cid in t_rows:
                cid_str = str(cid)
                if cid_str not in best_by_creator or (views or 0) > (best_by_creator[cid_str].get("views") or 0):
                    best_by_creator[cid_str] = {
                        "transcript": transcript, "url": t_url or "",
                        "views": views, "post_date": pdate,
                    }

            t_updated = 0
            for cid_str, data in best_by_creator.items():
                try:
                    cr = PipelineCreator.objects.get(id=cid_str)
                    changed = False
                    if data["transcript"] and data["transcript"] != cr.top_post_transcript:
                        cr.top_post_transcript = data["transcript"]
                        changed = True
                    if data["url"] and data["url"] != cr.top_post_url:
                        cr.top_post_url = data["url"]
                        changed = True
                    if data["views"] and data["views"] != cr.top_post_views:
                        cr.top_post_views = data["views"]
                        changed = True
                    if data["post_date"] and data["post_date"] != cr.top_post_date:
                        cr.top_post_date = data["post_date"]
                        changed = True
                    if changed:
                        cr.save()
                        t_updated += 1
                except PipelineCreator.DoesNotExist:
                    continue

            transcript_result = {"updated": t_updated, "checked": len(t_rows), "matched": len(best_by_creator)}
        except Exception as te:
            transcript_result["error"] = str(te)

    return _cors_headers(request, JsonResponse({
        "created": created,
        "skipped": skipped,
        "imported": imported,
        "transcript_sync": transcript_result,
    }, status=201))


# --- Syncly Sheet → Excel → DB (2-step pipeline) ---

_SYNCLY_SHEET_ID = "1dIAhP8wCEdFulSAai3K-RoZTvLBIaWxAK7hzInBsF0o"

# Output_updated column indices (0-based)
_OUT_COL = {
    "username": 6, "level": 8, "post_url": 15, "text": 18,
    "transcript": 19, "caption": 20, "post_date": 21,
    "followers": 32, "avg_view": 33, "views_30d": 39, "likes_30d": 40,
}


def _syncly_excel_path():
    """Return path to the clean Excel file on the server."""
    import os as _os
    proj = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
    return _os.path.join(proj, "Data Storage", "syncly_creators_clean.xlsx")


def _safe_int(val):
    if not val:
        return None
    try:
        return int(float(str(val).replace(",", "").strip()))
    except (ValueError, TypeError):
        return None


# ── Syncly Pipeline: Upload Excel → DB (single step, no Google Sheets) ──

@csrf_exempt
def syncly_upload_excel(request):
    """POST: Upload Excel → save on server → upsert to pipeline_creators.

    One-step pipeline. Accepts multipart/form-data with field 'file'.
    Saves file, reads it, upserts creators + content fields to DB.
    Returns stats (created, updated, skipped, etc.)
    """
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))
    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    import os as _os, re as _re
    from datetime import date as date_cls

    f = request.FILES.get('file')
    if not f:
        return _cors_headers(request, JsonResponse({"error": "No file uploaded. Use field name 'file'."}, status=400))

    if not f.name.endswith(('.xlsx', '.xls')):
        return _cors_headers(request, JsonResponse({"error": "Only .xlsx/.xls files accepted"}, status=400))

    try:
        import openpyxl

        xlsx_path = _syncly_excel_path()
        _os.makedirs(_os.path.dirname(xlsx_path), exist_ok=True)

        # 1. Save uploaded file
        with open(xlsx_path, 'wb') as dest:
            for chunk in f.chunks():
                dest.write(chunk)

        # 2. Read Excel
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return _cors_headers(request, JsonResponse({"error": "Excel has no data rows"}, status=400))

        hdr = [str(h or "").strip() for h in rows[0]]
        data_rows = rows[1:]

        def col_idx(name):
            nl = name.lower()
            for i, h in enumerate(hdr):
                if h.lower() == nl:
                    return i
            for i, h in enumerate(hdr):
                if nl in h.lower():
                    return i
            return -1

        def _parse_date(raw):
            if raw is None:
                return None
            if hasattr(raw, 'date') and callable(getattr(raw, 'date', None)):
                return raw.date()
            raw = str(raw).strip()
            if not raw:
                return None
            if " " in raw:
                raw = raw.split(" ")[0]
            try:
                if "-" in raw:
                    parts = raw.split("-")
                    if len(parts) == 3:
                        y = int(parts[0])
                        if y < 100:
                            y += 2000
                        return date_cls(y, int(parts[1]), int(parts[2]))
            except Exception:
                pass
            return None

        # Column indices
        idx_username = col_idx("Username")
        idx_email = col_idx("Email")
        idx_platform = col_idx("Platform")
        idx_collab = col_idx("제휴 상태")
        if idx_collab < 0:
            idx_collab = col_idx("\uc81c\ud734 \uc0c1\ud0dc")
        idx_discovery = col_idx("최초 발견")
        if idx_discovery < 0:
            idx_discovery = col_idx("발견")
        idx_followers_cr = col_idx("Followers")

        # Content columns
        idx_top_post_url = col_idx("top_post_url")
        idx_top_post_transcript = col_idx("top_post_transcript")
        if idx_top_post_transcript < 0:
            idx_top_post_transcript = col_idx("transcript")
        idx_top_post_caption = col_idx("top_post_caption")
        if idx_top_post_caption < 0:
            idx_top_post_caption = col_idx("caption")
        idx_top_post_views = col_idx("top_post_views")
        idx_views_30d = col_idx("views_30d")
        idx_likes_30d = col_idx("likes_30d")
        idx_avg_view = col_idx("avg_view")
        idx_followers_output = col_idx("followers_output")
        idx_post_date = col_idx("top_post_date")
        if idx_post_date < 0:
            idx_post_date = col_idx("post_date")

        # 3. Process rows
        stats = {"total_rows": len(data_rows), "skip_no_email": 0,
                 "skip_collab": 0, "skip_invalid": 0, "eligible": 0,
                 "with_content": 0, "created": 0, "updated": 0, "dupes": 0}

        ht_threshold = 100000
        try:
            cfg = PipelineConfig.objects.order_by('-date').first()
            if cfg and cfg.ht_threshold:
                ht_threshold = cfg.ht_threshold
        except Exception:
            pass

        for row in data_rows:
            def cell(idx, _row=row):
                return str(_row[idx] or "").strip() if 0 <= idx < len(_row) else ""

            username = cell(idx_username).lstrip("@").strip()
            if not username or not _re.match(r'^[a-zA-Z0-9._]+$', username):
                stats["skip_invalid"] += 1
                continue

            email_val = cell(idx_email)
            if not email_val or "@" not in email_val or "@discovered." in email_val.lower():
                stats["skip_no_email"] += 1
                continue

            collab = cell(idx_collab)
            if collab:
                stats["skip_collab"] += 1
                continue

            stats["eligible"] += 1

            raw_d = row[idx_discovery] if 0 <= idx_discovery < len(row) else None
            disc_date = _parse_date(str(raw_d) if raw_d else "")

            handle = username.lower()
            plat = "TikTok" if "tiktok" in cell(idx_platform).lower() else "Instagram"
            ig_handle = handle if plat == "Instagram" else ""
            tiktok_handle = handle if plat == "TikTok" else ""

            avg_v = _safe_int(cell(idx_avg_view)) if idx_avg_view >= 0 else None
            views_30d = _safe_int(cell(idx_views_30d)) if idx_views_30d >= 0 else None
            avg_v = avg_v or views_30d or 0
            outreach_type = "HT" if avg_v >= ht_threshold else "LT"
            followers = (_safe_int(cell(idx_followers_output)) if idx_followers_output >= 0 else None) \
                        or (_safe_int(cell(idx_followers_cr)) if idx_followers_cr >= 0 else None) or 0

            content_fields = {}
            post_url = cell(idx_top_post_url) if idx_top_post_url >= 0 else ""
            if post_url:
                stats["with_content"] += 1
                content_fields["top_post_url"] = post_url
                content_fields["top_post_transcript"] = cell(idx_top_post_transcript) if idx_top_post_transcript >= 0 else ""
                content_fields["top_post_caption"] = cell(idx_top_post_caption) if idx_top_post_caption >= 0 else ""
                content_fields["top_post_views"] = _safe_int(cell(idx_top_post_views)) if idx_top_post_views >= 0 else None
                content_fields["views_30d"] = views_30d
                content_fields["likes_30d"] = _safe_int(cell(idx_likes_30d)) if idx_likes_30d >= 0 else None
                pd = _parse_date(cell(idx_post_date)) if idx_post_date >= 0 else None
                if pd:
                    content_fields["top_post_date"] = pd

            # Upsert
            existing = PipelineCreator.objects.filter(ig_handle__iexact=handle).first() or \
                       PipelineCreator.objects.filter(tiktok_handle__iexact=handle).first()

            if existing:
                changed = False
                if "@discovered." in (existing.email or ""):
                    if not PipelineCreator.objects.filter(email=email_val).exclude(id=existing.id).exists():
                        existing.email = email_val
                        changed = True
                    else:
                        stats["dupes"] += 1
                        continue

                for field, val in content_fields.items():
                    if val is not None and val != "":
                        setattr(existing, field, val)
                        changed = True

                if avg_v and (not existing.avg_views or avg_v > existing.avg_views):
                    existing.avg_views = avg_v
                    changed = True
                if followers and (not existing.followers or followers > existing.followers):
                    existing.followers = followers
                    changed = True
                if not existing.initial_discovery_date and disc_date:
                    existing.initial_discovery_date = disc_date
                    changed = True

                if changed:
                    existing.save()
                    stats["updated"] += 1
                continue

            if PipelineCreator.objects.filter(email=email_val).exists():
                stats["dupes"] += 1
                continue

            try:
                PipelineCreator.objects.create(
                    email=email_val,
                    ig_handle=ig_handle,
                    tiktok_handle=tiktok_handle,
                    full_name=handle,
                    platform=plat,
                    pipeline_status="Not Started",
                    outreach_type=outreach_type,
                    source="syncly",
                    followers=followers,
                    avg_views=avg_v,
                    initial_discovery_date=disc_date or date_cls.today(),
                    notes=f"Excel upload. Post: {post_url or 'N/A'}",
                    **content_fields,
                )
                stats["created"] += 1
            except Exception:
                stats["dupes"] += 1

        stats["filename"] = f.name
        stats["uploaded_at"] = datetime.now().isoformat()
        return _cors_headers(request, JsonResponse(stats, status=200))

    except Exception as e:
        import traceback
        return _cors_headers(request, JsonResponse({
            "error": str(e), "traceback": traceback.format_exc()
        }, status=500))


@csrf_exempt
def syncly_import_excel(request):
    """POST: Read cleaned Excel → import/update pipeline_creators.

    Step 2 of the 2-step pipeline.

    Body (optional):
      week: filter to specific week (YYYY-MM-DD). If omitted, imports ALL rows.
      clean_placeholders: true to delete @discovered.* placeholders first (default true)
    """
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))
    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    import os as _os, re as _re
    from datetime import date as date_cls

    try:
        import openpyxl
    except ImportError:
        return _cors_headers(request, JsonResponse({"error": "openpyxl not installed"}, status=500))

    xlsx_path = _syncly_excel_path()
    if not _os.path.exists(xlsx_path):
        return _cors_headers(request, JsonResponse({
            "error": "Excel file not found. Run Step 1 (Export) first.",
            "path": xlsx_path,
        }, status=404))

    try:
        body = _json_body(request)
    except Exception:
        body = {}
    week_str = body.get("week", "")
    clean_ph = body.get("clean_placeholders", True)

    def _parse_date(raw):
        if raw is None:
            return None
        # openpyxl returns datetime objects directly
        if hasattr(raw, 'date') and callable(getattr(raw, 'date', None)):
            return raw.date()
        raw = str(raw).strip()
        if not raw:
            return None
        # Strip time component from stringified datetime ("2026-03-15 00:00:00")
        if " " in raw:
            raw = raw.split(" ")[0]
        try:
            if "-" in raw:
                parts = raw.split("-")
                if len(parts) == 3:
                    y = int(parts[0])
                    if y < 100:
                        y += 2000
                    return date_cls(y, int(parts[1]), int(parts[2]))
            elif len(raw) == 6 and raw.isdigit():
                return date_cls(2000 + int(raw[:2]), int(raw[2:4]), int(raw[4:6]))
        except Exception:
            pass
        return None

    try:
        # Read Excel
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return _cors_headers(request, JsonResponse({
                "error": "Excel file is empty", "total_rows": 0
            }, status=400))

        hdr = [str(h or "") for h in rows[0]]
        data_rows = rows[1:]

        def col_idx(name):
            # Exact match first, then substring fallback
            nl = name.lower()
            for i, h in enumerate(hdr):
                if h.lower() == nl:
                    return i
            for i, h in enumerate(hdr):
                if nl in h.lower():
                    return i
            return -1

        idx_username = col_idx("Username")
        idx_email = col_idx("Email")
        idx_platform = col_idx("Platform")
        idx_collab = col_idx("제휴 상태")
        if idx_collab < 0:
            idx_collab = col_idx("\uc81c\ud734 \uc0c1\ud0dc")
        idx_discovery = col_idx("최초 발견")
        if idx_discovery < 0:
            idx_discovery = col_idx("발견")
        idx_followers_cr = col_idx("Followers")

        # Content columns
        idx_top_post_url = col_idx("top_post_url")
        idx_top_post_transcript = col_idx("top_post_transcript")
        idx_top_post_caption = col_idx("top_post_caption")
        idx_top_post_views = col_idx("top_post_views")
        idx_top_post_date = col_idx("top_post_date")
        idx_views_30d = col_idx("views_30d")
        idx_likes_30d = col_idx("likes_30d")
        idx_avg_view = col_idx("avg_view")
        idx_followers_output = col_idx("followers_output")

        # Week filter
        target_dates = None
        if week_str:
            wd = _parse_date(week_str)
            if wd:
                target_dates = set([wd + timedelta(days=i) for i in range(7)])

        # Filter rows
        eligible = []
        stats = {"total_rows": len(data_rows), "matched_week": 0,
                 "skip_no_email": 0, "skip_collab": 0, "skip_invalid": 0}

        for row in data_rows:
            def cell(idx, _row=row):
                return str(_row[idx] or "").strip() if 0 <= idx < len(_row) else ""

            # Week filter
            if target_dates:
                raw_d = row[idx_discovery] if 0 <= idx_discovery < len(row) else None
                disc_date = _parse_date(str(raw_d) if raw_d else "")
                if not disc_date or disc_date not in target_dates:
                    continue
                stats["matched_week"] += 1

            username = cell(idx_username).lstrip("@").strip()
            if not username or not _re.match(r'^[a-zA-Z0-9._]+$', username):
                stats["skip_invalid"] += 1
                continue

            email_val = cell(idx_email)
            if not email_val or "@" not in email_val or "@discovered." in email_val.lower():
                stats["skip_no_email"] += 1
                continue

            collab = cell(idx_collab)
            if collab:
                stats["skip_collab"] += 1
                continue

            raw_d = row[idx_discovery] if 0 <= idx_discovery < len(row) else None
            disc_date = _parse_date(str(raw_d) if raw_d else "")

            entry = {
                "username": username.lower(),
                "email": email_val,
                "platform": cell(idx_platform),
                "discovery_date": disc_date,
                "followers_cr": _safe_int(cell(idx_followers_cr)) if idx_followers_cr >= 0 else None,
                "top_post_url": cell(idx_top_post_url) if idx_top_post_url >= 0 else "",
                "top_post_transcript": cell(idx_top_post_transcript) if idx_top_post_transcript >= 0 else "",
                "top_post_caption": cell(idx_top_post_caption) if idx_top_post_caption >= 0 else "",
                "top_post_views": _safe_int(cell(idx_top_post_views)) if idx_top_post_views >= 0 else None,
                "top_post_date": _parse_date(cell(idx_top_post_date)) if idx_top_post_date >= 0 else None,
                "views_30d": _safe_int(cell(idx_views_30d)) if idx_views_30d >= 0 else None,
                "likes_30d": _safe_int(cell(idx_likes_30d)) if idx_likes_30d >= 0 else None,
                "avg_views": _safe_int(cell(idx_avg_view)) if idx_avg_view >= 0 else None,
                "followers": _safe_int(cell(idx_followers_output)) if idx_followers_output >= 0 else None,
            }
            eligible.append(entry)

        stats["eligible"] = len(eligible)
        stats["with_content"] = sum(1 for e in eligible if e.get("top_post_url"))

        if not eligible:
            return _cors_headers(request, JsonResponse(stats, status=200))

        # Clean @discovered.* placeholders if requested
        cleaned_ph = 0
        if clean_ph:
            ph_qs = PipelineCreator.objects.filter(
                email__icontains="@discovered.",
                pipeline_status="Not Started",
            )
            if target_dates:
                ph_qs = ph_qs.filter(initial_discovery_date__in=list(target_dates))
            cleaned_ph = ph_qs.delete()[0]

        # HT threshold from config
        ht_threshold = 100000
        try:
            cfg = PipelineConfig.objects.order_by('-date').first()
            if cfg and cfg.ht_threshold:
                ht_threshold = cfg.ht_threshold
        except Exception:
            pass

        created = 0
        updated = 0
        dupes = 0

        for e in eligible:
            handle = e["username"]
            email_val = e["email"]
            plat = "TikTok" if "tiktok" in (e["platform"] or "").lower() else "Instagram"
            ig_handle = handle if plat == "Instagram" else ""
            tiktok_handle = handle if plat == "TikTok" else ""
            disc_date = e["discovery_date"] or date_cls.today()

            avg_v = e.get("avg_views") or e.get("views_30d") or 0
            outreach_type = "HT" if avg_v >= ht_threshold else "LT"
            followers = e.get("followers") or e.get("followers_cr") or 0

            content_fields = {
                "top_post_url": e.get("top_post_url") or "",
                "top_post_transcript": e.get("top_post_transcript") or "",
                "top_post_caption": e.get("top_post_caption") or "",
                "top_post_views": e.get("top_post_views"),
                "top_post_date": e.get("top_post_date"),
                "views_30d": e.get("views_30d"),
                "likes_30d": e.get("likes_30d"),
            }

            existing = PipelineCreator.objects.filter(ig_handle__iexact=handle).first() or \
                       PipelineCreator.objects.filter(tiktok_handle__iexact=handle).first()

            if existing:
                changed = False
                if "@discovered." in (existing.email or ""):
                    if not PipelineCreator.objects.filter(email=email_val).exclude(id=existing.id).exists():
                        existing.email = email_val
                        changed = True
                    else:
                        dupes += 1
                        continue

                if e.get("top_post_url"):
                    for field, val in content_fields.items():
                        if val is not None and val != "":
                            setattr(existing, field, val)
                            changed = True

                if e.get("avg_views") and (not existing.avg_views or e["avg_views"] > existing.avg_views):
                    existing.avg_views = e["avg_views"]
                    changed = True
                if followers and (not existing.followers or followers > existing.followers):
                    existing.followers = followers
                    changed = True
                if not existing.initial_discovery_date:
                    existing.initial_discovery_date = disc_date
                    changed = True

                if changed:
                    existing.save()
                    updated += 1
                continue

            if PipelineCreator.objects.filter(email=email_val).exists():
                dupes += 1
                continue

            try:
                PipelineCreator.objects.create(
                    email=email_val,
                    ig_handle=ig_handle,
                    tiktok_handle=tiktok_handle,
                    full_name=handle,
                    platform=plat,
                    pipeline_status="Not Started",
                    outreach_type=outreach_type,
                    source="syncly",
                    followers=followers,
                    avg_views=avg_v,
                    initial_discovery_date=disc_date,
                    notes=f"Excel import. Post: {e.get('top_post_url') or 'N/A'}",
                    **content_fields,
                )
                created += 1
            except Exception:
                dupes += 1

        stats["cleaned_placeholders"] = cleaned_ph
        stats["created"] = created
        stats["updated"] = updated
        stats["dupes"] = dupes
        stats["import_date"] = datetime.now().isoformat()
        return _cors_headers(request, JsonResponse(stats, status=200))

    except Exception as e:
        import traceback
        return _cors_headers(request, JsonResponse({
            "error": str(e), "traceback": traceback.format_exc()
        }, status=500))


@csrf_exempt
def syncly_excel_status(request):
    """GET: Return status of the Syncly Excel pipeline (file info + last counts)."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    import os as _os

    xlsx_path = _syncly_excel_path()
    result = {"excel_exists": False, "excel_path": xlsx_path}

    if _os.path.exists(xlsx_path):
        stat = _os.stat(xlsx_path)
        result["excel_exists"] = True
        result["excel_date"] = datetime.fromtimestamp(stat.st_mtime).isoformat()
        result["excel_size_kb"] = round(stat.st_size / 1024, 1)

        # Count rows
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            result["excel_rows"] = ws.max_row - 1 if ws.max_row else 0  # minus header
            wb.close()
        except Exception:
            result["excel_rows"] = None

    # DB stats: how many pipeline_creators from source=syncly
    try:
        from django.db.models import Q
        total_syncly = PipelineCreator.objects.filter(source="syncly").count()
        placeholder = PipelineCreator.objects.filter(
            source="syncly", email__icontains="@discovered.").count()
        with_email = total_syncly - placeholder
        result["db_syncly_total"] = total_syncly
        result["db_syncly_with_email"] = with_email
        result["db_syncly_placeholder"] = placeholder
    except Exception:
        pass

    return _cors_headers(request, JsonResponse(result))


# ---------------------------------------------------------------------------
# Syncly Autofill Emails — Apify IG Profile Scraper + Firecrawl
# ---------------------------------------------------------------------------

import re as _re_mod
import os as _os

_EMAIL_RE = _re_mod.compile(
    r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}'
)
_IGNORE_DOMAINS = {
    'discovered.syncly', 'onzenna.com', 'orbiters.co.kr',
    'zezebaebae.com', 'example.com', 'email.com', 'gmail.con',
    'instagram.com', 'tiktok.com', 'facebook.com', 'twitter.com',
    'sentry.io', 'sentry-next.wixpress.com',
}


def _extract_emails(text):
    """Extract valid emails from text, excluding internal/platform domains."""
    if not text:
        return []
    found = _EMAIL_RE.findall(text)
    return [
        e.lower() for e in found
        if e.split('@')[1].lower() not in _IGNORE_DOMAINS
    ]


def _apify_fetch_profiles(handles, token):
    """Call Apify Instagram Profile Scraper for a batch of handles.
    Returns dict: {handle: {bio, external_url, ...}}
    """
    import urllib.request
    import urllib.parse

    actor_id = "apify~instagram-profile-scraper"
    url = f"https://api.apify.com/v2/acts/{actor_id}/runs?token={token}"
    payload = json.dumps({
        "usernames": handles,
    }).encode()
    req = urllib.request.Request(
        url, data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            run_data = json.loads(resp.read())
    except Exception as e:
        return {"_error": f"Apify run start failed: {e}"}

    run_id = run_data.get("data", {}).get("id")
    if not run_id:
        return {"_error": "No run ID returned"}

    # Poll for completion (max 3 min)
    import time
    status_url = f"https://api.apify.com/v2/actor-runs/{run_id}?token={token}"
    for _ in range(36):  # 36 * 5s = 180s
        time.sleep(5)
        try:
            with urllib.request.urlopen(status_url, timeout=15) as resp:
                status_data = json.loads(resp.read())
            status = status_data.get("data", {}).get("status")
            if status in ("SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"):
                break
        except Exception:
            continue

    if status != "SUCCEEDED":
        return {"_error": f"Apify run status: {status}"}

    # Fetch dataset
    dataset_id = run_data.get("data", {}).get("defaultDatasetId")
    if not dataset_id:
        return {"_error": "No dataset ID"}

    ds_url = f"https://api.apify.com/v2/datasets/{dataset_id}/items?token={token}&limit=200"
    try:
        with urllib.request.urlopen(ds_url, timeout=30) as resp:
            items = json.loads(resp.read())
    except Exception as e:
        return {"_error": f"Dataset fetch failed: {e}"}

    result = {}
    for item in items:
        uname = (item.get("username") or "").lower()
        if uname:
            result[uname] = {
                "bio": item.get("biography") or item.get("bio") or "",
                "external_url": item.get("externalUrl") or item.get("external_url") or "",
                "full_name": item.get("fullName") or item.get("full_name") or "",
            }
    return result


def _firecrawl_extract_email(url, api_key):
    """Use Firecrawl to scrape a URL and extract email addresses."""
    import urllib.request

    fc_url = "https://api.firecrawl.dev/v1/scrape"
    payload = json.dumps({
        "url": url,
        "formats": ["markdown"],
        "onlyMainContent": True,
    }).encode()
    req = urllib.request.Request(
        fc_url, data=payload,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read())
        md = data.get("data", {}).get("markdown", "")
        return _extract_emails(md)
    except Exception:
        return []


@csrf_exempt
def syncly_autofill_emails(request):
    """POST: Find emails for placeholder creators via Apify + Firecrawl.

    Query params:
      ?limit=20  — max creators to process (default 20, max 50)
      ?source=last_upload — only process creators from last Excel upload
    """
    if request.method == "OPTIONS":
        resp = HttpResponse(status=204)
        return _cors_headers(request, resp)
    if request.method != "POST":
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    # Get API keys from env
    apify_token = _os.environ.get("APIFY_API_TOKEN", "")
    firecrawl_key = _os.environ.get("FIRECRAWL_API_KEY", "")

    if not apify_token:
        return _cors_headers(request, JsonResponse({
            "error": "APIFY_API_TOKEN not configured on server"
        }, status=500))

    # Parse params
    limit = min(int(request.GET.get("limit", "20")), 50)

    # Find placeholder creators (respect region filter from dashboard)
    region = request.GET.get("region", "").strip().lower()
    from django.db.models import Q
    qs = PipelineCreator.objects.filter(
        Q(email__icontains="@discovered.") | Q(email__icontains="@noemail.")
    ).exclude(
        ig_handle=""
    ).exclude(
        ig_handle__isnull=True
    )
    if region:
        qs = qs.filter(region__iexact=region)
    qs = qs.order_by("-created_at")[:limit]

    creators = list(qs)
    if not creators:
        return _cors_headers(request, JsonResponse({
            "total_checked": 0, "found": 0, "not_found": 0,
            "message": "No placeholder creators with IG handles to process"
        }))

    # Batch Apify call
    handles = [c.ig_handle.lstrip("@").strip() for c in creators if c.ig_handle]
    profiles = _apify_fetch_profiles(handles, apify_token)

    if "_error" in profiles:
        return _cors_headers(request, JsonResponse({
            "error": profiles["_error"],
            "total_checked": len(handles),
        }, status=502))

    # Process results
    found_list = []
    not_found_list = []

    for c in creators:
        handle = (c.ig_handle or "").lstrip("@").strip().lower()
        profile = profiles.get(handle, {})
        bio = profile.get("bio", "")
        ext_url = profile.get("external_url", "")

        # 1st: Try bio email
        emails = _extract_emails(bio)

        # 2nd: Try linked website via Firecrawl
        if not emails and ext_url and firecrawl_key:
            emails = _firecrawl_extract_email(ext_url, firecrawl_key)

        if emails:
            best_email = emails[0]
            # Check uniqueness before saving
            if PipelineCreator.objects.filter(email=best_email).exclude(pk=c.pk).exists():
                not_found_list.append(handle)
                continue
            c.email = best_email
            try:
                c.save(update_fields=["email"])
            except Exception:
                not_found_list.append(handle)
                continue
            found_list.append({
                "handle": handle,
                "email": best_email,
                "source": "bio" if _extract_emails(bio) else "website",
            })
        else:
            not_found_list.append(handle)

    stats = {
        "total_checked": len(creators),
        "found": len(found_list),
        "not_found": len(not_found_list),
        "updated": found_list,
        "not_found_handles": not_found_list[:20],
    }
    return _cors_headers(request, JsonResponse(stats))


@csrf_exempt
def syncly_content_import(request):
    """POST: Enrich DB creators with content data from uploaded Excel.

    Reads the uploaded Excel file (Step 1), finds content columns
    (transcript, post_url, views, etc.), matches by username,
    and updates existing PipelineCreator records.

    No Google Sheets dependency — reads from local Excel only.
    """
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    import os as _os
    from datetime import date as date_cls

    xlsx_path = _syncly_excel_path()
    if not _os.path.exists(xlsx_path):
        return _cors_headers(request, JsonResponse({
            "error": "No Excel file found. Upload one first (Step 1).",
            "path": xlsx_path,
        }, status=404))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return _cors_headers(request, JsonResponse({"error": "Excel has no data rows"}, status=400))

        hdr = [str(h or "").strip() for h in rows[0]]
        data_rows = rows[1:]

        def col_idx(name):
            nl = name.lower()
            for i, h in enumerate(hdr):
                if h.lower() == nl:
                    return i
            for i, h in enumerate(hdr):
                if nl in h.lower():
                    return i
            return -1

        # Find content columns
        i_username = col_idx("Username")
        i_transcript = col_idx("top_post_transcript")
        if i_transcript < 0:
            i_transcript = col_idx("transcript")
        i_post_url = col_idx("top_post_url")
        if i_post_url < 0:
            i_post_url = col_idx("post_url")
        i_caption = col_idx("top_post_caption")
        if i_caption < 0:
            i_caption = col_idx("caption")
        i_views = col_idx("top_post_views")
        if i_views < 0:
            i_views = col_idx("views_30d")
        i_likes = col_idx("likes_30d")
        i_avg_view = col_idx("avg_view")
        i_followers = col_idx("followers_output")
        if i_followers < 0:
            i_followers = col_idx("Followers")
        i_post_date = col_idx("top_post_date")
        if i_post_date < 0:
            i_post_date = col_idx("post_date")

        if i_username < 0:
            return _cors_headers(request, JsonResponse({
                "error": "No 'Username' column found in Excel",
                "columns": hdr[:20],
            }, status=400))

        def _cell(row, idx):
            if idx < 0 or idx >= len(row):
                return ""
            v = row[idx]
            if v is None:
                return ""
            return str(v).strip()

        updated = 0
        skipped = 0
        not_found = 0

        for row in data_rows:
            username = _cell(row, i_username).lstrip("@").lower()
            if not username:
                skipped += 1
                continue

            # Find existing creator by handle
            existing = PipelineCreator.objects.filter(ig_handle__iexact=username).first() or \
                       PipelineCreator.objects.filter(tiktok_handle__iexact=username).first()

            if not existing:
                not_found += 1
                continue

            changed = False

            # Update content fields
            transcript = _cell(row, i_transcript)
            if transcript and (not existing.top_post_transcript or len(transcript) > len(existing.top_post_transcript or "")):
                existing.top_post_transcript = transcript
                changed = True

            post_url = _cell(row, i_post_url)
            if post_url and not existing.top_post_url:
                existing.top_post_url = post_url
                changed = True

            caption = _cell(row, i_caption)
            if caption and not existing.top_post_caption:
                existing.top_post_caption = caption
                changed = True

            views = _safe_int(_cell(row, i_views))
            if views and (not existing.top_post_views or views > existing.top_post_views):
                existing.top_post_views = views
                changed = True

            views_30d = _safe_int(_cell(row, i_views))
            if views_30d and (not existing.views_30d or views_30d > existing.views_30d):
                existing.views_30d = views_30d
                changed = True

            likes = _safe_int(_cell(row, i_likes))
            if likes and (not existing.likes_30d or likes > existing.likes_30d):
                existing.likes_30d = likes
                changed = True

            avg_v = _safe_int(_cell(row, i_avg_view))
            if avg_v and (not existing.avg_views or avg_v > existing.avg_views):
                existing.avg_views = avg_v
                changed = True

            followers = _safe_int(_cell(row, i_followers))
            if followers and (not existing.followers or followers > existing.followers):
                existing.followers = followers
                changed = True

            # Post date
            pd_raw = _cell(row, i_post_date)
            if pd_raw and "-" in pd_raw:
                try:
                    if " " in pd_raw:
                        pd_raw = pd_raw.split(" ")[0]
                    pp = pd_raw.split("-")
                    existing.top_post_date = date_cls(int(pp[0]), int(pp[1]), int(pp[2]))
                    changed = True
                except Exception:
                    pass

            if changed:
                existing.save()
                updated += 1
            else:
                skipped += 1

        return _cors_headers(request, JsonResponse({
            "updated": updated,
            "skipped": skipped,
            "not_found": not_found,
            "total_excel_rows": len(data_rows),
            "created": 0,
        }, status=200))

    except Exception as e:
        import traceback
        return _cors_headers(request, JsonResponse({
            "error": str(e),
            "traceback": traceback.format_exc(),
        }, status=500))


# ========== EMAIL REPLY CONFIG ==========

def _serialize_email_config(cfg, include_faq=False):
    """Serialize EmailReplyConfig to dict, optionally including FAQ entries."""
    data = {
        "brand": cfg.brand,
        "is_active": cfg.is_active,
        "version": cfg.version,
        "classification": {
            "prompt": cfg.classification_prompt,
            "model": cfg.classification_model,
        },
        "auto_send": {
            "lt": cfg.lt_auto_send,
            "ht": cfg.ht_auto_send,
        },
        "templates": {
            "accept": cfg.accept_template,
            "faq_gap": cfg.faq_gap_template,
            "normal": cfg.normal_template,
            "decline": cfg.decline_template,
        },
        "sender_name": cfg.sender_name,
        "sign_off": cfg.sign_off,
        "outreach_prompts": {
            "lt": cfg.outreach_lt_prompt,
            "ht": cfg.outreach_ht_prompt,
        },
        "content_guidelines": {
            "hashtags": json.loads(cfg.hashtags) if cfg.hashtags else [],
            "product_mentions": json.loads(cfg.product_mentions) if cfg.product_mentions else [],
            "deadline_days": cfg.deadline_days,
        },
        "gifting_form_url": cfg.gifting_form_url,
        "updated_at": cfg.updated_at.isoformat() if cfg.updated_at else None,
        "updated_by": cfg.updated_by,
    }
    if include_faq:
        faqs = FAQEntry.objects.filter(
            brand__in=[cfg.brand, "all"], is_active=True
        ).order_by("-priority", "category")
        data["faq_entries"] = [
            {
                "id": str(f.id),
                "brand": f.brand,
                "question": f.question,
                "answer": f.answer,
                "keywords": json.loads(f.keywords) if f.keywords else [],
                "category": f.category,
                "priority": f.priority,
            }
            for f in faqs
        ]
    return data


@csrf_exempt
def email_config_list(request):
    """GET: list all brand configs. POST not used here."""
    if request.method == "OPTIONS":
        return _cors_headers(request, HttpResponse(status=204))
    configs = EmailReplyConfig.objects.all()
    data = [_serialize_email_config(c) for c in configs]
    return _cors_headers(request, JsonResponse(data, safe=False))


@csrf_exempt
def email_config_detail(request, brand):
    """GET: n8n fetches full config + FAQ for a brand. POST: create/update."""
    if request.method == "OPTIONS":
        return _cors_headers(request, HttpResponse(status=204))

    if request.method == "GET":
        try:
            cfg = EmailReplyConfig.objects.get(brand=brand)
        except EmailReplyConfig.DoesNotExist:
            return _cors_headers(request, JsonResponse({"error": f"No config for brand '{brand}'"}, status=404))
        return _cors_headers(request, JsonResponse(_serialize_email_config(cfg, include_faq=True)))

    if request.method == "POST":
        body = _json_body(request)
        cfg, created = EmailReplyConfig.objects.update_or_create(
            brand=brand,
            defaults={
                "is_active": body.get("is_active", True),
                "classification_prompt": body.get("classification_prompt", ""),
                "classification_model": body.get("classification_model", "claude-sonnet-4-20250514"),
                "lt_auto_send": body.get("lt_auto_send", True),
                "ht_auto_send": body.get("ht_auto_send", False),
                "sender_name": body.get("sender_name", ""),
                "sign_off": body.get("sign_off", ""),
                "accept_template": body.get("accept_template", ""),
                "faq_gap_template": body.get("faq_gap_template", ""),
                "normal_template": body.get("normal_template", ""),
                "decline_template": body.get("decline_template", ""),
                "outreach_lt_prompt": body.get("outreach_lt_prompt", ""),
                "outreach_ht_prompt": body.get("outreach_ht_prompt", ""),
                "hashtags": json.dumps(body.get("hashtags", [])),
                "product_mentions": json.dumps(body.get("product_mentions", [])),
                "deadline_days": body.get("deadline_days", 30),
                "gifting_form_url": body.get("gifting_form_url", ""),
                "updated_by": body.get("updated_by", ""),
            },
        )
        if not created:
            cfg.version += 1
            cfg.save(update_fields=["version"])
        return _cors_headers(request, JsonResponse(
            _serialize_email_config(cfg), status=201 if created else 200
        ))

    return _cors_headers(request, JsonResponse({"error": "Method not allowed"}, status=405))


@csrf_exempt
def faq_list(request):
    """GET: list FAQ entries (optional ?brand=X). POST: create new."""
    if request.method == "OPTIONS":
        return _cors_headers(request, HttpResponse(status=204))

    if request.method == "GET":
        qs = FAQEntry.objects.filter(is_active=True)
        brand = request.GET.get("brand")
        if brand:
            qs = qs.filter(brand__in=[brand, "all"])
        data = [
            {
                "id": str(f.id),
                "brand": f.brand,
                "question": f.question,
                "answer": f.answer,
                "keywords": json.loads(f.keywords) if f.keywords else [],
                "category": f.category,
                "priority": f.priority,
                "created_at": f.created_at.isoformat() if f.created_at else None,
                "updated_at": f.updated_at.isoformat() if f.updated_at else None,
            }
            for f in qs
        ]
        return _cors_headers(request, JsonResponse(data, safe=False))

    if request.method == "POST":
        body = _json_body(request)
        f = FAQEntry.objects.create(
            brand=body.get("brand", "all"),
            question=body.get("question", ""),
            answer=body.get("answer", ""),
            keywords=json.dumps(body.get("keywords", [])),
            category=body.get("category", ""),
            priority=body.get("priority", 0),
        )
        return _cors_headers(request, JsonResponse({"id": str(f.id), "created": True}, status=201))

    return _cors_headers(request, JsonResponse({"error": "Method not allowed"}, status=405))


@csrf_exempt
def faq_detail(request, faq_id):
    """PUT: update. DELETE: soft-delete."""
    if request.method == "OPTIONS":
        return _cors_headers(request, HttpResponse(status=204))

    try:
        f = FAQEntry.objects.get(id=faq_id)
    except FAQEntry.DoesNotExist:
        return _cors_headers(request, JsonResponse({"error": "FAQ not found"}, status=404))

    if request.method == "PUT":
        body = _json_body(request)
        for field in ("brand", "question", "answer", "category"):
            if field in body:
                setattr(f, field, body[field])
        if "keywords" in body:
            f.keywords = json.dumps(body["keywords"])
        if "priority" in body:
            f.priority = body["priority"]
        if "is_active" in body:
            f.is_active = body["is_active"]
        f.save()
        return _cors_headers(request, JsonResponse({"updated": True}))

    if request.method == "DELETE":
        f.is_active = False
        f.save(update_fields=["is_active"])
        return _cors_headers(request, JsonResponse({"deleted": True}))

    return _cors_headers(request, JsonResponse({"error": "Method not allowed"}, status=405))


@csrf_exempt
def reply_log_create(request):
    """POST: n8n writes audit log. GET: dashboard reads recent logs."""
    if request.method == "OPTIONS":
        return _cors_headers(request, HttpResponse(status=204))

    if request.method == "GET":
        days = int(request.GET.get("days", 7))
        cutoff = datetime.now() - timedelta(days=days)
        qs = EmailReplyLog.objects.filter(processed_at__gte=cutoff)
        brand = request.GET.get("brand")
        if brand:
            qs = qs.filter(brand=brand)
        data = [
            {
                "id": str(r.id),
                "creator_email": r.creator_email,
                "brand": r.brand,
                "outreach_type": r.outreach_type,
                "intent": r.intent,
                "confidence": r.confidence,
                "auto_sent": r.auto_sent,
                "template_used": r.template_used,
                "incoming_subject": r.incoming_subject,
                "config_version": r.config_version,
                "processed_at": r.processed_at.isoformat() if r.processed_at else None,
            }
            for r in qs[:100]
        ]
        return _cors_headers(request, JsonResponse(data, safe=False))

    if request.method == "POST":
        body = _json_body(request)
        r = EmailReplyLog.objects.create(
            creator_email=body.get("creator_email", ""),
            brand=body.get("brand", ""),
            outreach_type=body.get("outreach_type", ""),
            intent=body.get("intent", "Unknown"),
            confidence=body.get("confidence"),
            auto_sent=body.get("auto_sent", False),
            template_used=body.get("template_used", ""),
            faq_entry_id=body.get("faq_entry_id"),
            incoming_subject=body.get("incoming_subject", ""),
            incoming_snippet=body.get("incoming_snippet", ""),
            outgoing_body=body.get("outgoing_body", ""),
            config_version=body.get("config_version", 1),
        )
        return _cors_headers(request, JsonResponse({"id": str(r.id), "created": True}, status=201))

    return _cors_headers(request, JsonResponse({"error": "Method not allowed"}, status=405))


# --- Cross-Check: Gmail + Apify + Shopify PR batch match ---


@csrf_exempt
def pipeline_creators_cross_check(request):
    """POST: Batch cross-check PipelineCreators against Gmail, Apify, Shopify PR.

    Query params:
        sources=gmail,apify,shopify_pr  (comma-separated, default: all)

    Gmail: match email → enrich gmail_* fields
    Apify: match ig_handle/tiktok_handle against gk_content_posts → enrich apify_* fields
    Shopify PR: match ig_handle/email against gk_influencer_orders → enrich pr_* fields
    """
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    from django.db import connection

    requested = request.GET.get("sources", "gmail,apify,shopify_pr")
    run_sources = [s.strip() for s in requested.split(",")]

    total_creators = PipelineCreator.objects.count()
    result = {"total_creators": total_creators}

    # --- Gmail cross-check ---
    if "gmail" in run_sources:
        gmail_matched = 0
        gmail_updated = 0
        all_gmail = {c.email: c for c in GmailContact.objects.all()}

        batch_size = 500
        off = 0
        while off < total_creators:
            creators = list(PipelineCreator.objects.all()[off:off + batch_size])
            for cr in creators:
                gc = all_gmail.get(cr.email)
                if not gc:
                    continue
                gmail_matched += 1
                changed = False
                if gc.first_contact_date and not cr.gmail_first_contact:
                    cr.gmail_first_contact = gc.first_contact_date.date() if hasattr(gc.first_contact_date, 'date') else gc.first_contact_date
                    changed = True
                if gc.last_contact_date:
                    gc_date = gc.last_contact_date.date() if hasattr(gc.last_contact_date, 'date') else gc.last_contact_date
                    if not cr.gmail_last_contact or gc_date > cr.gmail_last_contact:
                        cr.gmail_last_contact = gc_date
                        changed = True
                if gc.total_sent and gc.total_sent > cr.gmail_total_sent:
                    cr.gmail_total_sent = gc.total_sent
                    changed = True
                if gc.total_received and gc.total_received > cr.gmail_total_received:
                    cr.gmail_total_received = gc.total_received
                    changed = True
                acct = gc.account or "zezebaebae"
                current_accounts = cr.gmail_accounts or []
                if acct not in current_accounts:
                    cr.gmail_accounts = current_accounts + [acct]
                    changed = True
                current_sources = cr.sources or []
                if "gmail_inbound" not in current_sources:
                    cr.sources = current_sources + ["gmail_inbound"]
                    changed = True
                if changed:
                    cr.save()
                    gmail_updated += 1
            off += batch_size

        result["gmail_matched"] = gmail_matched
        result["gmail_updated"] = gmail_updated
        result["total_gmail_contacts"] = len(all_gmail)

    # --- Apify cross-check (gk_content_posts → PipelineCreator) ---
    if "apify" in run_sources:
        apify_matched = 0
        apify_updated = 0

        # Build handle→posts map from gk_content_posts
        with connection.cursor() as cur:
            cur.execute("""
                SELECT username, platform, brand, post_date, url,
                       views_30d, likes_30d, comments_30d, post_id
                FROM gk_content_posts
                WHERE post_date >= CURRENT_DATE - INTERVAL '180 days'
                ORDER BY post_date DESC
            """)
            cols = [c[0] for c in cur.description]
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]

        # Group by lowercase handle
        handle_posts = {}
        for r in rows:
            h = (r["username"] or "").lower().lstrip("@")
            if not h:
                continue
            handle_posts.setdefault(h, []).append(r)

        # Match against PipelineCreator
        batch_size = 500
        off = 0
        while off < total_creators:
            creators = list(PipelineCreator.objects.all()[off:off + batch_size])
            for cr in creators:
                ig = (cr.ig_handle or "").lower().lstrip("@")
                tt = (cr.tiktok_handle or "").lower().lstrip("@")
                posts = []
                if ig and ig in handle_posts:
                    posts.extend(handle_posts[ig])
                if tt and tt in handle_posts:
                    posts.extend(handle_posts[tt])
                if not posts:
                    continue

                apify_matched += 1
                changed = False

                # Deduplicate by post_id
                seen = set()
                unique_posts = []
                for p in posts:
                    if p["post_id"] not in seen:
                        seen.add(p["post_id"])
                        unique_posts.append(p)

                new_count = len(unique_posts)
                brands = list(set(p["brand"] for p in unique_posts if p.get("brand")))
                last_date = max((p["post_date"] for p in unique_posts if p.get("post_date")), default=None)

                # Build apify_posts JSON (top 50 by date)
                top_posts = sorted(unique_posts, key=lambda x: x.get("post_date") or "", reverse=True)[:50]
                apify_json = [{
                    "post_id": p["post_id"],
                    "url": p.get("url", ""),
                    "platform": p.get("platform", ""),
                    "post_date": str(p["post_date"]) if p.get("post_date") else "",
                    "views": p.get("views_30d", 0),
                    "likes": p.get("likes_30d", 0),
                    "comments": p.get("comments_30d", 0),
                } for p in top_posts]

                if not cr.is_apify_tagged:
                    cr.is_apify_tagged = True
                    changed = True
                if new_count != cr.apify_post_count:
                    cr.apify_post_count = new_count
                    changed = True
                if sorted(brands) != sorted(cr.apify_posted_brands or []):
                    cr.apify_posted_brands = brands
                    changed = True
                if last_date and (not cr.apify_last_post_date or last_date > cr.apify_last_post_date):
                    cr.apify_last_post_date = last_date
                    changed = True

                cr.apify_posts = apify_json
                cr.apify_last_crawled_at = datetime.now()
                changed = True

                src = cr.sources or []
                if "apify" not in src:
                    cr.sources = src + ["apify"]
                    changed = True

                if not cr.collaboration_status and new_count > 0:
                    cr.collaboration_status = "active"
                    changed = True

                if changed:
                    cr.save()
                    apify_updated += 1
            off += batch_size

        result["apify_matched"] = apify_matched
        result["apify_updated"] = apify_updated
        result["apify_total_posts"] = len(rows)

    # --- Shopify PR cross-check (gk_influencer_orders → PipelineCreator) ---
    if "shopify_pr" in run_sources:
        pr_matched = 0
        pr_updated = 0

        with connection.cursor() as cur:
            cur.execute("""
                SELECT account_handle, customer_email, brand,
                       product_names, product_types, shipping_date, order_name
                FROM gk_influencer_orders
                ORDER BY shipping_date DESC NULLS LAST
            """)
            cols = [c[0] for c in cur.description]
            pr_rows = [dict(zip(cols, r)) for r in cur.fetchall()]

        # Build handle→orders and email→orders maps
        handle_orders = {}
        email_orders = {}
        for r in pr_rows:
            h = (r.get("account_handle") or "").lower().lstrip("@")
            e = (r.get("customer_email") or "").lower()
            if h:
                handle_orders.setdefault(h, []).append(r)
            if e:
                email_orders.setdefault(e, []).append(r)

        batch_size = 500
        off = 0
        while off < total_creators:
            creators = list(PipelineCreator.objects.all()[off:off + batch_size])
            for cr in creators:
                ig = (cr.ig_handle or "").lower().lstrip("@")
                email_lower = (cr.email or "").lower()
                orders = []
                if ig and ig in handle_orders:
                    orders.extend(handle_orders[ig])
                if email_lower and email_lower in email_orders:
                    orders.extend(email_orders[email_lower])
                if not orders:
                    continue

                pr_matched += 1
                changed = False

                # Deduplicate by order_name
                seen = set()
                unique_orders = []
                for o in orders:
                    key = o.get("order_name", "")
                    if key and key not in seen:
                        seen.add(key)
                        unique_orders.append(o)

                # Build pr_products JSON
                pr_products = []
                for o in unique_orders:
                    products = (o.get("product_names") or "").split(",")
                    for prod in products:
                        prod = prod.strip()
                        if prod:
                            pr_products.append({
                                "brand": o.get("brand", ""),
                                "product": prod,
                                "order_date": str(o["shipping_date"]) if o.get("shipping_date") else "",
                            })

                if not cr.is_shopify_pr:
                    cr.is_shopify_pr = True
                    changed = True
                if pr_products and pr_products != (cr.pr_products or []):
                    cr.pr_products = pr_products
                    changed = True

                src = cr.sources or []
                if "shopify_pr" not in src:
                    cr.sources = src + ["shopify_pr"]
                    changed = True

                if changed:
                    cr.save()
                    pr_updated += 1
            off += batch_size

        result["shopify_pr_matched"] = pr_matched
        result["shopify_pr_updated"] = pr_updated
        result["shopify_pr_total_orders"] = len(pr_rows)

    # Always include current flag counts
    result["shopify_pr_count"] = PipelineCreator.objects.filter(is_shopify_pr=True).count()
    result["apify_tagged_count"] = PipelineCreator.objects.filter(is_apify_tagged=True).count()
    result["manychat_count"] = PipelineCreator.objects.filter(is_manychat_contact=True).count()

    return _cors_headers(request, JsonResponse(result, status=200))


# --- Sync Transcripts from content_posts → pipeline_creators ---

@csrf_exempt
def sync_transcripts(request):
    """POST: Sync transcripts from gk_content_posts to pipeline_creators.

    Finds content_posts with transcripts and updates matching pipeline_creators.

    Body:
      region: "jp" or "us" (required)
      limit: max posts to check (default: 300)
      min_length: minimum transcript length (default: 20)
    """
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    body = _json_body(request)
    region = body.get("region", "")
    limit = int(body.get("limit", 300))
    min_length = int(body.get("min_length", 20))

    from django.db import connection

    # Find content_posts with transcripts, joined to pipeline_creators
    sql = """
        SELECT
            cp.username,
            cp.transcript,
            cp.url,
            cp.views_30d,
            cp.post_date,
            pc.id AS creator_id
        FROM gk_content_posts cp
        INNER JOIN onz_pipeline_creators pc
            ON LOWER(cp.username) = LOWER(pc.ig_handle)
        WHERE cp.transcript IS NOT NULL
          AND LENGTH(cp.transcript) >= %s
          AND cp.username IS NOT NULL
          AND cp.username != ''
    """
    params = [min_length]

    if region:
        sql += " AND LOWER(cp.region) = LOWER(%s)"
        params.append(region)

    sql += " ORDER BY cp.views_30d DESC NULLS LAST LIMIT %s"
    params.append(limit)

    updated = 0
    checked = 0
    matched = 0

    try:
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            rows = cursor.fetchall()

        checked = len(rows)

        # Group by creator_id, pick best (highest views)
        best_by_creator = {}
        for username, transcript, url, views, post_date, creator_id in rows:
            matched += 1
            cid = str(creator_id)
            if cid not in best_by_creator or (views or 0) > (best_by_creator[cid].get("views") or 0):
                best_by_creator[cid] = {
                    "transcript": transcript,
                    "url": url or "",
                    "views": views,
                    "post_date": post_date,
                }

        # Update pipeline_creators
        for cid, data in best_by_creator.items():
            try:
                cr = PipelineCreator.objects.get(id=cid)
                changed = False
                if data["transcript"] and data["transcript"] != cr.top_post_transcript:
                    cr.top_post_transcript = data["transcript"]
                    changed = True
                if data["url"] and data["url"] != cr.top_post_url:
                    cr.top_post_url = data["url"]
                    changed = True
                if data["views"] and data["views"] != cr.top_post_views:
                    cr.top_post_views = data["views"]
                    changed = True
                if data["post_date"] and data["post_date"] != cr.top_post_date:
                    cr.top_post_date = data["post_date"]
                    changed = True
                if changed:
                    cr.save()
                    updated += 1
            except PipelineCreator.DoesNotExist:
                continue

    except Exception as e:
        return _cors_headers(request, JsonResponse({
            "error": str(e),
            "updated": updated,
            "checked": checked,
            "matched": matched,
        }, status=500))

    return _cors_headers(request, JsonResponse({
        "updated": updated,
        "checked": checked,
        "matched": matched,
    }))


# --- GK Transcripts (for brand keyword matching) ---

@csrf_exempt
def gk_transcripts(request):
    """POST: return concatenated transcripts per username from gk_content_posts."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))
    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    body = _json_body(request)
    usernames = body.get("usernames", [])
    if not usernames or not isinstance(usernames, list):
        return _cors_headers(request, JsonResponse({"results": {}, "total": 0}))

    # Limit to 500 usernames
    usernames = usernames[:500]
    placeholders = ",".join(["%s"] * len(usernames))

    from django.db import connection
    with connection.cursor() as cur:
        cur.execute(f"""
            SELECT LOWER(username), STRING_AGG(COALESCE(transcript, '') || ' ' || COALESCE(caption, ''), ' ')
            FROM gk_content_posts
            WHERE LOWER(username) IN ({placeholders})
            GROUP BY LOWER(username)
        """, [u.lower() for u in usernames])
        rows = cur.fetchall()

    results = {}
    for username, text in rows:
        if text and text.strip():
            results[username] = text.strip()[:2000]  # cap at 2000 chars per user

    return _cors_headers(request, JsonResponse({
        "results": results,
        "total": len(results),
    }))


# --- Email Verify (basic syntax + @discovered.* check) ---

@csrf_exempt
def email_verify(request):
    """POST: validate emails — syntax check + @discovered.* filter."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))
    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    try:
        body = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return _cors_headers(request, JsonResponse({"error": "Invalid JSON"}, status=400))

    emails = body.get("emails", [])
    if not emails or not isinstance(emails, list):
        return _cors_headers(request, JsonResponse({"error": "emails array required"}, status=400))

    import re
    email_re = re.compile(r'^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$')

    results = {}
    for e in emails:
        if not isinstance(e, str):
            continue
        e_lower = e.strip().lower()
        if not e_lower:
            continue
        if '@discovered.' in e_lower:
            results[e_lower] = {"valid": False, "reason": "placeholder email (@discovered.*)"}
        elif not email_re.match(e_lower):
            results[e_lower] = {"valid": False, "reason": "invalid format"}
        else:
            results[e_lower] = {"valid": True, "reason": None}

    return _cors_headers(request, JsonResponse({
        "results": results,
        "total_checked": len(results),
        "total_valid": sum(1 for v in results.values() if v["valid"]),
        "total_invalid": sum(1 for v in results.values() if not v["valid"]),
    }))


# --- Transcript Language Check (ASCII ratio for English detection) ---

@csrf_exempt
def transcript_lang_check(request):
    """POST: Check transcript language for given usernames.

    Body:
      usernames: list of ig_handles to check

    Returns per-username: {is_english, ratio, transcript_count}
    Uses gk_content_posts transcripts joined by username.
    ASCII ratio >= 0.8 = English.
    """
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    body = _json_body(request)
    usernames = body.get("usernames", [])

    if not usernames:
        return _cors_headers(request, JsonResponse({"results": {}}))

    from django.db import connection

    # Get all transcripts for these usernames from gk_content_posts
    placeholders = ",".join(["%s"] * len(usernames))
    sql = f"""
        SELECT LOWER(username) as uname, transcript
        FROM gk_content_posts
        WHERE LOWER(username) IN ({placeholders})
          AND transcript IS NOT NULL
          AND LENGTH(transcript) >= 20
    """
    params = [u.lower() for u in usernames]

    results = {}
    try:
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            rows = cursor.fetchall()

        # Group transcripts by username
        by_user = {}
        for uname, transcript in rows:
            by_user.setdefault(uname, []).append(transcript)

        # Also check top_post_transcript from pipeline_creators
        pc_placeholders = ",".join(["%s"] * len(usernames))
        pc_sql = f"""
            SELECT LOWER(ig_handle) as uname, top_post_transcript
            FROM onz_pipeline_creators
            WHERE LOWER(ig_handle) IN ({pc_placeholders})
              AND top_post_transcript IS NOT NULL
              AND LENGTH(top_post_transcript) >= 20
        """
        with connection.cursor() as cursor:
            cursor.execute(pc_sql, params)
            pc_rows = cursor.fetchall()

        for uname, transcript in pc_rows:
            if uname not in by_user:
                by_user.setdefault(uname, []).append(transcript)

        # Calculate ASCII ratio per username
        for uname in [u.lower() for u in usernames]:
            transcripts = by_user.get(uname, [])
            if not transcripts:
                results[uname] = {"is_english": False, "ratio": 0, "transcript_count": 0}
                continue

            # Combine all transcripts for ratio calculation
            combined = " ".join(transcripts)
            if len(combined) == 0:
                results[uname] = {"is_english": False, "ratio": 0, "transcript_count": len(transcripts)}
                continue

            ascii_count = sum(1 for c in combined if ord(c) < 128)
            ratio = ascii_count / len(combined)
            results[uname] = {
                "is_english": ratio >= 0.8,
                "ratio": round(ratio, 3),
                "transcript_count": len(transcripts),
            }

    except Exception as e:
        return _cors_headers(request, JsonResponse({"error": str(e)}, status=500))

    return _cors_headers(request, JsonResponse({"results": results}))


@csrf_exempt
def classify_region_by_transcript(request):
    """POST: Auto-classify creator region based on transcript language ratio.

    Analyzes all transcripts (gk_content_posts + top_post_transcript) per creator.
    If Japanese char ratio >= threshold → region='jp', else → region='us'.

    Body:
      dry_run: true/false (default: true) — preview only, no DB update
      threshold: float (default: 0.15) — min Japanese char ratio to classify as JP
      min_chars: int (default: 30) — min combined transcript length to classify
      only_empty: true/false (default: false) — only classify creators with empty region

    Returns: {classified: [...], summary: {to_jp, to_us, unchanged, skipped}}
    """
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    body = _json_body(request)
    dry_run = body.get("dry_run", True)
    threshold = float(body.get("threshold", 0.15))
    min_chars = int(body.get("min_chars", 30))
    only_empty = body.get("only_empty", False)

    from django.db import connection

    try:
        with connection.cursor() as cursor:
            # Get all creators with their combined transcript text
            sql = """
                WITH content_transcripts AS (
                    SELECT
                        LOWER(cp.username) AS uname,
                        STRING_AGG(cp.transcript, ' ') AS combined
                    FROM gk_content_posts cp
                    WHERE cp.transcript IS NOT NULL
                      AND LENGTH(cp.transcript) >= 20
                      AND cp.username IS NOT NULL AND cp.username != ''
                    GROUP BY LOWER(cp.username)
                ),
                pipeline_transcripts AS (
                    SELECT
                        LOWER(pc.ig_handle) AS uname,
                        pc.top_post_transcript AS combined
                    FROM onz_pipeline_creators pc
                    WHERE pc.top_post_transcript IS NOT NULL
                      AND LENGTH(pc.top_post_transcript) >= 20
                ),
                merged AS (
                    SELECT uname, combined FROM content_transcripts
                    UNION ALL
                    SELECT uname, combined FROM pipeline_transcripts
                ),
                per_creator AS (
                    SELECT
                        uname,
                        STRING_AGG(combined, ' ') AS all_text
                    FROM merged
                    GROUP BY uname
                )
                SELECT
                    pc.id, pc.ig_handle, pc.region, pc.country,
                    pc.followers, pc.brand, pc.pipeline_status,
                    LENGTH(m.all_text) AS total_chars,
                    (LENGTH(m.all_text) - LENGTH(
                        REGEXP_REPLACE(m.all_text, E'[\\u3040-\\u309F\\u30A0-\\u30FF]', '', 'g')
                    )) AS jp_chars
                FROM onz_pipeline_creators pc
                INNER JOIN per_creator m ON LOWER(pc.ig_handle) = m.uname
                WHERE LENGTH(m.all_text) >= %s
                ORDER BY pc.ig_handle
            """
            cursor.execute(sql, [min_chars])
            cols = [c[0] for c in cursor.description]
            rows = [dict(zip(cols, r)) for r in cursor.fetchall()]

        classified = []
        to_jp = 0
        to_us = 0
        unchanged = 0
        skipped = 0

        for row in rows:
            total = row['total_chars'] or 1
            jp_pct = (row['jp_chars'] or 0) / total
            new_region = 'jp' if jp_pct >= threshold else 'us'
            current_region = (row['region'] or '').strip().lower()

            if only_empty and current_region:
                skipped += 1
                continue

            changed = (new_region != current_region)
            if changed:
                if new_region == 'jp':
                    to_jp += 1
                else:
                    to_us += 1
            else:
                unchanged += 1

            if changed:
                classified.append({
                    "ig_handle": row['ig_handle'],
                    "current_region": row['region'] or '',
                    "new_region": new_region,
                    "jp_ratio": round(jp_pct, 3),
                    "total_chars": total,
                    "followers": row['followers'],
                    "brand": row['brand'] or '',
                    "status": row['pipeline_status'],
                })

        # Apply updates if not dry_run
        updated = 0
        if not dry_run and classified:
            with connection.cursor() as cursor:
                for item in classified:
                    cursor.execute(
                        "UPDATE onz_pipeline_creators SET region = %s, updated_at = NOW() WHERE ig_handle = %s",
                        [item['new_region'], item['ig_handle']]
                    )
                    updated += 1

        resp = {
            "dry_run": dry_run,
            "threshold": threshold,
            "min_chars": min_chars,
            "summary": {
                "to_jp": to_jp,
                "to_us": to_us,
                "unchanged": unchanged,
                "skipped": skipped,
                "updated": updated,
            },
            "classified": classified,
        }
        return _cors_headers(request, JsonResponse(resp))

    except Exception as e:
        import traceback
        return _cors_headers(request, JsonResponse({"error": str(e), "trace": traceback.format_exc()}, status=500))


# --- Run CI Pipeline (Whisper transcript trigger) ---

@csrf_exempt
def run_ci_pipeline(request):
    """POST: Identify content_posts needing Whisper transcripts.

    Returns list of posts that need transcription (no transcript, video, views >= min_views).
    Does NOT run Whisper itself — that's handled by a separate background process.

    Body:
      region: "jp" or "us" (required)
      max: max posts to process (default: 20)
      min_views: minimum views threshold (default: 3000)
    """
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST required"}, status=405))

    body = _json_body(request)
    region = body.get("region", "")
    max_posts = int(body.get("max", 20))
    min_views = int(body.get("min_views", 3000))

    from django.db import connection

    sql = """
        SELECT cp.post_id, cp.username, cp.url, cp.views_30d, cp.post_date
        FROM gk_content_posts cp
        WHERE (cp.transcript IS NULL OR cp.transcript = '')
          AND cp.url IS NOT NULL
          AND cp.url != ''
          AND (cp.views_30d IS NOT NULL AND cp.views_30d >= %s)
    """
    params = [min_views]

    if region:
        sql += " AND LOWER(cp.region) = LOWER(%s)"
        params.append(region)

    sql += " ORDER BY cp.views_30d DESC NULLS LAST LIMIT %s"
    params.append(max_posts)

    try:
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            rows = cursor.fetchall()

        pending = []
        for post_id, username, url, views, post_date in rows:
            pending.append({
                "post_id": post_id,
                "username": username,
                "url": url,
                "views": views,
                "post_date": str(post_date) if post_date else None,
            })

        # Launch analyze_video_content.py as background process if posts found
        pid = None
        if pending and not body.get("dry_run"):
            import subprocess, os as _os
            script = _os.path.join(
                _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
                "tools", "analyze_video_content.py"
            )
            if _os.path.exists(script):
                try:
                    cmd = [
                        "python3", script,
                        "--region", region or "jp",
                        "--max", str(max_posts),
                        "--min-views", str(min_views),
                    ]
                    # Pass OPENAI_API_KEY from .env to subprocess
                    env = _os.environ.copy()
                    env_file = _os.path.join(
                        _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))),
                        ".env"
                    )
                    if _os.path.exists(env_file):
                        with open(env_file) as f:
                            for line in f:
                                line = line.strip()
                                if line and not line.startswith("#") and "=" in line:
                                    k, v = line.split("=", 1)
                                    env[k.strip()] = v.strip()
                    proc = subprocess.Popen(
                        cmd,
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        start_new_session=True,
                        env=env,
                    )
                    pid = proc.pid
                except Exception as e:
                    pid = f"launch_error: {e}"

        return _cors_headers(request, JsonResponse({
            "status": "ok",
            "region": region,
            "max": max_posts,
            "min_views": min_views,
            "pending_count": len(pending),
            "pid": pid,
            "pending": pending,
        }))

    except Exception as e:
        return _cors_headers(request, JsonResponse({
            "error": str(e),
        }, status=500))


# --- Discovery Posts API (gk_content_posts for JP/US CRM dashboards) ---

@csrf_exempt
@require_http_methods(["GET"])
def discovery_posts(request):
    """GET: List content posts from gk_content_posts for CRM discovery tab.

    Params:
      region: "jp" or "us" (optional)
      limit: max posts (default 2000)
      content_type: "video" or "image" (optional)
    """
    region = request.GET.get("region", "")
    limit = min(int(request.GET.get("limit", 2000)), 5000)
    content_type = request.GET.get("content_type", "")

    from django.db import connection

    sql = """
        SELECT cp.username, cp.nickname, cp.followers, cp.platform,
               cp.url, cp.post_date, cp.brand, cp.caption, cp.hashtags,
               cp.views_30d, cp.likes_30d, cp.comments_30d,
               cp.transcript, cp.region, cp.source,
               cp.brand_fit_score, cp.scene_fit, cp.subject_age,
               cp.scene_tags, cp.content_quality_score
        FROM gk_content_posts cp
        WHERE cp.username IS NOT NULL AND cp.username != ''
    """
    params = []

    if region:
        sql += " AND LOWER(cp.region) = LOWER(%s)"
        params.append(region)

    # content_type filter: infer from URL pattern (no content_type column)
    if content_type == "video":
        sql += " AND (cp.url LIKE '%%/reel/%%' OR cp.url LIKE '%%tiktok%%' OR (cp.transcript IS NOT NULL AND LENGTH(cp.transcript) > 20))"
    elif content_type == "image":
        sql += " AND cp.url NOT LIKE '%%/reel/%%' AND cp.url NOT LIKE '%%tiktok%%' AND (cp.transcript IS NULL OR LENGTH(cp.transcript) <= 20)"

    sql += " ORDER BY cp.views_30d DESC NULLS LAST LIMIT %s"
    params.append(limit)

    try:
        with connection.cursor() as cursor:
            cursor.execute(sql, params)
            cols = [desc[0] for desc in cursor.description]
            rows = cursor.fetchall()

        results = []
        for row in rows:
            r = dict(zip(cols, row))
            results.append({
                "handle": r.get("username", ""),
                "full_name": r.get("nickname", ""),
                "views": r.get("views_30d") or 0,
                "likes": r.get("likes_30d") or 0,
                "comments_count": r.get("comments_30d") or 0,
                "followers": r.get("followers") or 0,
                "post_date": str(r.get("post_date", "")) if r.get("post_date") else "",
                "url": r.get("url", ""),
                "platform": r.get("platform", ""),
                "source": r.get("source", ""),
                "caption": r.get("caption", ""),
                "hashtags": r.get("hashtags", ""),
                "transcript": r.get("transcript", ""),
                "brand": r.get("brand", ""),
                "brand_fit_score": r.get("brand_fit_score"),
                "scene_fit": r.get("scene_fit", ""),
                "subject_age": r.get("subject_age", ""),
                "scene_tags": r.get("scene_tags", ""),
                "delivery_verbal_score": r.get("content_quality_score"),
                "content_type": "video" if (r.get("url") and ("/reel/" in r["url"] or "tiktok" in r["url"])) or (r.get("transcript") and len(r["transcript"]) > 20) else "image",
                "outreach_email": "",
            })

        return _cors_headers(request, JsonResponse({"results": results}))

    except Exception as e:
        return _cors_headers(request, JsonResponse({"error": str(e)}, status=500))


# ── Creator Duplicate Detection + Merge ──────────────────────────────────────

@csrf_exempt
def creator_find_duplicates(request):
    """GET: Find duplicate creator groups by email, phone, name, or handle overlap."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))

    region = request.GET.get("region")
    qs = PipelineCreator.objects.all()
    if region:
        qs = qs.filter(region__iexact=region)

    creators = list(qs.values("id", "ig_handle", "tiktok_handle", "full_name",
                              "email", "phone", "region", "pipeline_status",
                              "brand", "followers", "platform"))

    import re
    email_map = {}
    phone_map = {}
    name_map = {}
    handle_map = {}

    for c in creators:
        cid = str(c["id"])
        email = (c["email"] or "").strip().lower()
        if email and "@discovered." not in email:
            email_map.setdefault(email, []).append(cid)
        phone = re.sub(r"[^\d]", "", c["phone"] or "")
        if len(phone) >= 7:
            phone_map.setdefault(phone, []).append(cid)
        name = (c["full_name"] or "").strip().lower()
        rgn = (c["region"] or "us").lower()
        if name and len(name) > 2:
            name_map.setdefault((name, rgn), []).append(cid)
        ig = (c["ig_handle"] or "").strip().lower()
        tt = (c["tiktok_handle"] or "").strip().lower()
        if ig:
            handle_map.setdefault(ig, []).append(cid)
        if tt:
            handle_map.setdefault(tt, []).append(cid)

    # Union-find for grouping
    parent = {}

    def find(x):
        while parent.get(x, x) != x:
            parent[x] = parent.get(parent[x], parent[x])
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    reasons = {}

    def add_pair(id1, id2, reason):
        union(id1, id2)
        key = tuple(sorted([id1, id2]))
        reasons.setdefault(key, set()).add(reason)

    for ids in email_map.values():
        if len(ids) > 1:
            for i in range(len(ids)):
                for j in range(i + 1, len(ids)):
                    add_pair(ids[i], ids[j], "email")
    for ids in phone_map.values():
        if len(ids) > 1:
            for i in range(len(ids)):
                for j in range(i + 1, len(ids)):
                    add_pair(ids[i], ids[j], "phone")
    for ids in name_map.values():
        if len(ids) > 1:
            for i in range(len(ids)):
                for j in range(i + 1, len(ids)):
                    add_pair(ids[i], ids[j], "name")
    for ids in handle_map.values():
        if len(ids) > 1:
            for i in range(len(ids)):
                for j in range(i + 1, len(ids)):
                    add_pair(ids[i], ids[j], "handle")

    # Bio cross-link: IG bio mentions TikTok handle (or vice versa)
    try:
        from django.db import connection
        with connection.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT ON (LOWER(username))
                    LOWER(username), bio_text
                FROM gk_content_posts
                WHERE bio_text IS NOT NULL AND LENGTH(bio_text) > 5
                ORDER BY LOWER(username), post_date DESC NULLS LAST
            """)
            bio_map = {r[0]: r[1] for r in cur.fetchall()}

        tt_re = re.compile(r'(?:tik\s*tok|tt)\s*[:\-@/]\s*@?([a-zA-Z0-9_.]{2,30})', re.IGNORECASE)
        tt_url_re = re.compile(r'tiktok\.com/@([a-zA-Z0-9_.]{2,30})', re.IGNORECASE)
        ig_re = re.compile(r'(?:instagram|ig|insta)\s*[:\-@/]\s*@?([a-zA-Z0-9_.]{2,30})', re.IGNORECASE)
        domain_re = re.compile(r'\.(com|co|org|net|io|uk)$')

        # Build reverse maps: ig_handle->cid, tt_handle->cid
        ig_cid = {}
        tt_cid = {}
        for c in creators:
            cid = str(c["id"])
            ig = (c["ig_handle"] or "").strip().lower()
            tt = (c["tiktok_handle"] or "").strip().lower()
            if ig:
                ig_cid[ig] = cid
            if tt:
                tt_cid[tt] = cid

        for handle, bio in bio_map.items():
            source_cid = ig_cid.get(handle) or tt_cid.get(handle)
            if not source_cid:
                continue
            # If source is IG creator, look for TT handle in bio
            if handle in ig_cid:
                for pat in [tt_re, tt_url_re]:
                    m = pat.search(bio)
                    if m:
                        found = m.group(1).strip('.').lower()
                        if found in tt_cid and not domain_re.search(found):
                            target_cid = tt_cid[found]
                            if source_cid != target_cid:
                                add_pair(source_cid, target_cid, "bio_crosslink")
                        break
            # If source is TT creator, look for IG handle in bio
            if handle in tt_cid:
                m = ig_re.search(bio)
                if m:
                    found = m.group(1).strip('.').lower()
                    if found in ig_cid and not domain_re.search(found):
                        target_cid = ig_cid[found]
                        if source_cid != target_cid:
                            add_pair(source_cid, target_cid, "bio_crosslink")
    except Exception as e:
        pass  # bio crosslink is best-effort

    groups = {}
    c_by_id = {str(c["id"]): c for c in creators}
    for cid in c_by_id:
        root = find(cid)
        groups.setdefault(root, []).append(cid)

    result = []
    for gid, (root, members) in enumerate(
            [(r, m) for r, m in groups.items() if len(m) > 1]):
        group_reasons = set()
        for i in range(len(members)):
            for j in range(i + 1, len(members)):
                key = tuple(sorted([members[i], members[j]]))
                group_reasons.update(reasons.get(key, set()))
        result.append({
            "group_id": gid,
            "match_reasons": sorted(group_reasons),
            "confidence": len(group_reasons),
            "creators": [c_by_id[m] for m in members],
        })
    result.sort(key=lambda g: -g["confidence"])

    return _cors_headers(request, JsonResponse({
        "duplicate_groups": result,
        "total_groups": len(result),
    }))


@csrf_exempt
def creator_merge(request):
    """POST: Merge secondary creators into a primary creator."""
    if request.method == 'OPTIONS':
        return _cors_headers(request, HttpResponse(status=204))
    if request.method != 'POST':
        return _cors_headers(request, JsonResponse({"error": "POST only"}, status=405))

    body = _json_body(request)
    primary_id = body.get("primary_id")
    secondary_ids = body.get("secondary_ids", [])

    if not primary_id or not secondary_ids:
        return _cors_headers(request, JsonResponse(
            {"error": "primary_id and secondary_ids required"}, status=400))

    try:
        primary = PipelineCreator.objects.get(id=primary_id)
    except PipelineCreator.DoesNotExist:
        return _cors_headers(request, JsonResponse(
            {"error": f"Primary {primary_id} not found"}, status=404))

    merged_handles = []
    for sid in secondary_ids:
        try:
            sec = PipelineCreator.objects.get(id=sid)
        except PipelineCreator.DoesNotExist:
            continue

        merged_handles.append(sec.ig_handle or sec.tiktok_handle or sec.email)

        # Fill empty fields on primary from secondary
        for f in ["ig_handle", "tiktok_handle", "full_name", "phone",
                   "business_category", "child_1_birthday", "child_2_birthday"]:
            if not getattr(primary, f) and getattr(sec, f):
                setattr(primary, f, getattr(sec, f))

        # Prefer non-discovered email
        if "@discovered." in (primary.email or ""):
            if sec.email and "@discovered." not in sec.email:
                primary.email = sec.email

        # Numeric: take max
        for f in ["followers", "avg_views"]:
            if (getattr(sec, f) or 0) > (getattr(primary, f) or 0):
                setattr(primary, f, getattr(sec, f))

        # Merge source tags
        if sec.sources:
            existing = set(primary.sources or [])
            for s in sec.sources:
                existing.add(s)
            primary.sources = list(existing)

        # Sum contact counts
        primary.contact_count = (primary.contact_count or 0) + (sec.contact_count or 0)
        primary.gmail_total_sent = (primary.gmail_total_sent or 0) + (sec.gmail_total_sent or 0)
        primary.gmail_total_received = (primary.gmail_total_received or 0) + (sec.gmail_total_received or 0)

        # Boolean flags: OR
        for f in ["is_shopify_pr", "is_apify_tagged", "is_manychat_contact", "is_business_account"]:
            if getattr(sec, f):
                setattr(primary, f, True)

        # Platform: Multi if different
        if sec.platform and primary.platform and sec.platform != primary.platform:
            primary.platform = "Multi"

        sec.delete()

    merge_note = f"Merged from: {', '.join(merged_handles)}"
    primary.notes = f"{primary.notes}\n{merge_note}" if primary.notes else merge_note
    primary.save()

    return _cors_headers(request, JsonResponse({
        "merged": True,
        "primary": _serialize_creator(primary),
        "merged_count": len(merged_handles),
        "merged_handles": merged_handles,
    }))
